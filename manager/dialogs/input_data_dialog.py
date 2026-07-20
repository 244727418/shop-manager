# -*- coding: utf-8 -*-
"""录入数据对话框"""
import csv
import ctypes
import io
import os
import re

from PyQt5.QtWidgets import (
    QApplication, QDialog, QVBoxLayout, QFormLayout, QLabel, QLineEdit,
    QFrame, QMessageBox, QPushButton, QShortcut, QWidget
)
from PyQt5.QtGui import QCursor, QKeySequence, QRegularExpressionValidator
from PyQt5.QtCore import QEvent, QRegularExpression, QTimer, Qt, pyqtSignal


def _normalize_number_text(value):
    return str(value or "").replace(",", "").replace("，", "").strip()


def _parse_number(value, integer=False):
    text = _normalize_number_text(value)
    if not text:
        return 0
    return int(text) if integer else float(text)


def _refund_header_columns(headers):
    normalized = [
        re.sub(r"[\s（）()_\-]+", "", str(value or "").strip().lower())
        for value in headers
    ]
    refund_col = next(
        (index for index, text in enumerate(normalized) if "退款金额" in text),
        None,
    )
    order_col = next(
        (
            index for index, text in enumerate(normalized)
            if any(keyword in text for keyword in ("订单号", "订单编号", "主订单号", "订单id", "orderid"))
        ),
        None,
    )
    return refund_col, order_col


def _refund_amount(value):
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value).replace(",", "").replace("，", ""))
    return float(match.group(0)) if match else None


def _refund_order_id(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _summarize_refund_rows(rows):
    iterator = iter(rows)
    refund_col = order_col = None
    for row in iterator:
        detected_refund, detected_order = _refund_header_columns(row)
        if refund_col is None and detected_refund is not None:
            refund_col = detected_refund
        if order_col is None and detected_order is not None:
            order_col = detected_order
        if refund_col is not None and order_col is not None:
            break
    else:
        missing = []
        if refund_col is None:
            missing.append("退款金额")
        if order_col is None:
            missing.append("订单编号")
        raise ValueError("没有识别到表头：" + "、".join(missing))

    amount = 0.0
    order_ids = set()
    for row in iterator:
        order_id = _refund_order_id(row[order_col]) if order_col < len(row) else ""
        value = _refund_amount(row[refund_col]) if refund_col < len(row) else None
        if not order_id or value is None:
            continue
        amount += value
        order_ids.add(order_id)
    return round(amount, 2), len(order_ids)


def _summarize_refund_workbook(file_path):
    if os.path.splitext(file_path)[1].lower() != ".xlsx":
        raise ValueError("退款表识别目前只支持 .xlsx 文件")
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, data_only=True, read_only=True)
    try:
        for sheet in workbook.worksheets:
            sheet.reset_dimensions()
            try:
                result = _summarize_refund_rows(sheet.iter_rows(values_only=True))
            except ValueError:
                continue
            return result, sheet.title
    finally:
        workbook.close()
    raise ValueError("所有工作表中都没有识别到“退款金额”和“订单编号”表头")


class RefundImportPanel(QFrame):
    recognized = pyqtSignal(object, object)

    def __init__(self, parent=None):
        super().__init__(
            parent,
            Qt.Tool | Qt.FramelessWindowHint | Qt.WindowDoesNotAcceptFocus,
        )
        self.setAttribute(Qt.WA_ShowWithoutActivating, True)
        self.setFrameShape(QFrame.StyledPanel)
        self.setStyleSheet(
            "RefundImportPanel { background: white; border: 1px solid #8aa4b8; }"
        )
        self.resize(290, 190)
        self.setAcceptDrops(True)
        self.source_path = ""
        self.pasted_rows = None

        layout = QVBoxLayout(self)
        self.drop_label = QLabel("拖入 .xlsx 文件\n或按 Ctrl+V 粘贴文件/Excel单元格")
        self.drop_label.setAlignment(Qt.AlignCenter)
        self.drop_label.setWordWrap(True)
        self.drop_label.setStyleSheet(
            "QLabel { border: 1px dashed #8aa4b8; background: #f7fafc; padding: 12px; }"
        )
        layout.addWidget(self.drop_label, 1)

        self.result_label = QLabel("等待导入退款表格")
        self.result_label.setWordWrap(True)
        layout.addWidget(self.result_label)

        btn_recognize = QPushButton("自动识别并填写")
        btn_recognize.setFocusPolicy(Qt.NoFocus)
        btn_recognize.clicked.connect(self.recognize)
        layout.addWidget(btn_recognize)

    def _set_file(self, file_path):
        if os.path.splitext(file_path)[1].lower() != ".xlsx":
            QMessageBox.warning(self, "文件格式", "请选择 .xlsx 格式的退款表格")
            return
        self.source_path = file_path
        self.pasted_rows = None
        self.drop_label.setText(os.path.basename(file_path))
        self.result_label.setText("文件已接收，点击“自动识别并填写”")

    def paste_clipboard(self):
        mime = QApplication.clipboard().mimeData()
        local_files = [url.toLocalFile() for url in mime.urls() if url.isLocalFile()]
        if local_files:
            self._set_file(local_files[0])
            return
        text = mime.text().strip()
        if not text:
            return
        if os.path.isfile(text.strip('"')):
            self._set_file(text.strip('"'))
            return
        delimiter = "\t" if "\t" in text else ","
        self.source_path = ""
        self.pasted_rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
        self.drop_label.setText(f"已粘贴 {len(self.pasted_rows)} 行单元格数据")
        self.result_label.setText("数据已接收，点击“自动识别并填写”")

    def dragEnterEvent(self, event):
        if any(
            url.isLocalFile() and url.toLocalFile().lower().endswith(".xlsx")
            for url in event.mimeData().urls()
        ):
            event.acceptProposedAction()

    def dropEvent(self, event):
        files = [url.toLocalFile() for url in event.mimeData().urls() if url.isLocalFile()]
        if files:
            self._set_file(files[0])
            event.acceptProposedAction()

    def recognize(self):
        try:
            if self.source_path:
                (amount, orders), sheet_name = _summarize_refund_workbook(self.source_path)
                source_text = f"Sheet：{sheet_name}"
            elif self.pasted_rows is not None:
                amount, orders = _summarize_refund_rows(self.pasted_rows)
                source_text = "剪贴板数据"
            else:
                QMessageBox.information(self, "提示", "请先拖入或粘贴退款表格")
                return
            parts = [source_text]
            parts.append(f"退款金额：{amount:.2f}" if amount is not None else "未识别退款金额列")
            parts.append(f"退款订单：{orders}" if orders is not None else "未识别订单号列")
            self.result_label.setText("｜".join(parts))
            self.recognized.emit(amount, orders)
        except Exception as exc:
            QMessageBox.warning(self, "识别失败", str(exc))


class InputDataDialog(QDialog):
    """手动录入数据对话框"""
    def __init__(self, parent=None, initial_data=None, tutorial_mode=False):
        super().__init__(parent)
        self.setWindowTitle("📝 录入数据")
        self.resize(900 if tutorial_mode else 250, 680)
        if not tutorial_mode:
            self.setFixedWidth(250)
        self.calculated_values = {}
        self.initial_data = initial_data or {}
        self.refund_import_panel = None
        self.refund_panel_anchor = None
        self._refund_paste_key_down = False
        self.init_ui()
        self.refund_paste_shortcut = QShortcut(QKeySequence.Paste, self)
        self.refund_paste_shortcut.setContext(Qt.ApplicationShortcut)
        self.refund_paste_shortcut.setEnabled(False)
        self.refund_paste_shortcut.activated.connect(self._paste_refund_clipboard)
        self.refund_paste_timer = QTimer(self)
        self.refund_paste_timer.setInterval(50)
        self.refund_paste_timer.timeout.connect(self._poll_refund_global_paste)
        self.refund_paste_timer.start()
        if self.initial_data:
            self.apply_initial_data()

    def init_ui(self):
        layout = QVBoxLayout(self)

        title = QLabel("📝 请填写以下数据（手动输入项）")
        title.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50; margin-bottom: 10px;")
        layout.addWidget(title)

        self.input_form_widget = QWidget()
        self.input_form_widget.setMaximumWidth(420)
        form_layout = QFormLayout(self.input_form_widget)
        form_layout.setSpacing(10)

        self.input_fields = {}
        input_fields_config = [
            ("实发订单", "实际发货的订单数量"),
            ("实发金额", "实际发货的总金额（元）"),
            ("毛利润", "毛利金额（元）"),
            ("退款金额", "退款的总金额（元）"),
            ("退款订单", "退款的订单数量"),
            ("推广费", "推广费用（元）"),
            ("扣款", "扣款金额（元）"),
            ("其他服务", "其他服务费用（元）"),
            ("其他", "其他费用（可以为负值）"),
        ]

        for field_name, tooltip in input_fields_config:
            le = QLineEdit()
            le.setPlaceholderText(f"请输入{field_name}")
            le.setToolTip(tooltip)
            le.setStyleSheet("padding: 6px; border: 1px solid #ccc; border-radius: 3px; font-size: 13px;")
            if field_name in ["实发金额", "毛利润", "退款金额", "推广费", "扣款", "其他服务", "其他"]:
                le.setValidator(QRegularExpressionValidator(QRegularExpression(r"\s*[+-]?[\d,，]*(?:\.\d*)?\s*")))
            elif field_name in ["实发订单", "退款订单"]:
                le.setValidator(QRegularExpressionValidator(QRegularExpression(r"\s*[\d,，]*\s*")))
            le.textChanged.connect(lambda text, editor=le: self._normalize_number_input(editor, text))
            le.installEventFilter(self)
            self.input_fields[field_name] = le
            label = QLabel(f"{field_name}:")
            label.setStyleSheet("color: #555; font-weight: bold;")
            form_layout.addRow(label, le)

        layout.addWidget(self.input_form_widget, 0, Qt.AlignLeft)

        self.btn_calculate = QPushButton("🧮 计算并预览")
        self.btn_calculate.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 5px;
                font-size: 13px;
            }
            QPushButton:hover { background-color: #2980b9; }
        """)
        self.btn_calculate.clicked.connect(self.calculate)
        layout.addWidget(self.btn_calculate)

        self.preview_label = QLabel()
        self.preview_label.setStyleSheet("""
            background-color: #f8f9fa;
            border: 1px solid #dee2e6;
            border-radius: 5px;
            padding: 15px;
            font-size: 12px;
        """)
        self.preview_label.setWordWrap(True)
        layout.addWidget(self.preview_label)

        btn_layout = QVBoxLayout()
        self.btn_confirm = QPushButton("✅ 确认保存")
        self.btn_confirm.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 5px;
                font-size: 13px;
            }
            QPushButton:hover { background-color: #219a52; }
        """)
        self.btn_confirm.clicked.connect(self.accept)
        btn_cancel = QPushButton("取消")
        btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
                font-size: 13px;
            }
            QPushButton:hover { background-color: #7f8c8d; }
        """)
        btn_cancel.clicked.connect(self.reject)
        btn_layout.addWidget(self.btn_confirm)
        btn_layout.addWidget(btn_cancel)
        layout.addLayout(btn_layout)

    def eventFilter(self, watched, event):
        if event.type() in (QEvent.FocusIn, QEvent.MouseButtonPress) and watched in self.input_fields.values():
            if watched in (
                self.input_fields.get("退款金额"),
                self.input_fields.get("退款订单"),
            ):
                self._show_refund_import_panel(watched)
            elif self.refund_import_panel:
                self.refund_import_panel.hide()
        return super().eventFilter(watched, event)

    def _show_refund_import_panel(self, anchor=None):
        if self.refund_import_panel is None:
            self.refund_import_panel = RefundImportPanel(self)
            self.refund_import_panel.recognized.connect(self._apply_refund_import)
        self.refund_panel_anchor = anchor or self.input_fields.get("退款金额")
        self._position_refund_import_panel()
        self.refund_import_panel.show()
        self.refund_import_panel.raise_()

    def _position_refund_import_panel(self):
        if self.refund_import_panel is None:
            return
        anchor = self.refund_panel_anchor or self.input_fields.get("退款金额")
        anchor_point = anchor.mapToGlobal(anchor.rect().topRight())
        x = anchor_point.x() + 8
        y = anchor_point.y()
        screen = QApplication.screenAt(anchor_point)
        if screen and x + self.refund_import_panel.width() > screen.availableGeometry().right():
            x = anchor.mapToGlobal(anchor.rect().topLeft()).x() - self.refund_import_panel.width() - 8
        self.refund_import_panel.move(x, y)

    def _paste_refund_clipboard(self):
        if self.refund_import_panel and self.refund_import_panel.isVisible():
            self.refund_import_panel.paste_clipboard()

    def _poll_refund_global_paste(self):
        panel = self.refund_import_panel
        if os.name != "nt" or panel is None or not panel.isVisible():
            self.refund_paste_shortcut.setEnabled(False)
            self._refund_paste_key_down = False
            return
        hovered = panel.frameGeometry().contains(QCursor.pos())
        self.refund_paste_shortcut.setEnabled(False)
        user32 = ctypes.windll.user32
        key_down = bool(user32.GetAsyncKeyState(0x11) & 0x8000) and bool(
            user32.GetAsyncKeyState(ord("V")) & 0x8000
        )
        if hovered and key_down and not self._refund_paste_key_down:
            panel.paste_clipboard()
        self._refund_paste_key_down = key_down

    def _apply_refund_import(self, amount, orders):
        if amount is not None:
            self.input_fields["退款金额"].setText(f"{amount:.2f}".rstrip("0").rstrip("."))
        if orders is not None:
            self.input_fields["退款订单"].setText(str(orders))

    def moveEvent(self, event):
        super().moveEvent(event)
        if self.refund_import_panel and self.refund_import_panel.isVisible():
            self._position_refund_import_panel()

    def closeEvent(self, event):
        self.refund_paste_shortcut.setEnabled(False)
        if self.refund_import_panel:
            self.refund_import_panel.close()
        super().closeEvent(event)

    @staticmethod
    def _normalize_number_input(editor, text):
        normalized = _normalize_number_text(text)
        if normalized != text:
            editor.setText(normalized)

    def apply_initial_data(self):
        field_map = {
            "actual_orders": "实发订单",
            "actual_amount": "实发金额",
            "gross_profit": "毛利润",
            "refund_amount": "退款金额",
            "refund_orders": "退款订单",
            "promotion_fee": "推广费",
            "deduction": "扣款",
            "other_service": "其他服务",
            "other": "其他",
        }
        for key, field_name in field_map.items():
            if key not in self.initial_data:
                continue
            value = self.initial_data.get(key)
            if value is None or value == "":
                continue
            if isinstance(value, float):
                text = f"{value:.2f}".rstrip("0").rstrip(".")
            else:
                text = str(value)
            self.input_fields[field_name].setText(text)
        self.calculate()

    def calculate(self):
        try:
            actual_orders = _parse_number(self.input_fields["实发订单"].text(), integer=True)
            actual_amount = _parse_number(self.input_fields["实发金额"].text())
            gross_profit = _parse_number(self.input_fields["毛利润"].text())
            refund_amount = _parse_number(self.input_fields["退款金额"].text())
            refund_orders = _parse_number(self.input_fields["退款订单"].text(), integer=True)
            promotion_fee = _parse_number(self.input_fields["推广费"].text())
            deduction = _parse_number(self.input_fields["扣款"].text())
            other_service = _parse_number(self.input_fields["其他服务"].text())
            other = _parse_number(self.input_fields["其他"].text())

            if actual_orders == 0 or actual_amount == 0:
                self.preview_label.setText("⚠️ 实发订单和实发金额不能为0")
                return

            gross_margin_rate = (gross_profit / actual_amount * 100) if actual_amount > 0 else 0
            refund_rate_by_amount = (refund_amount / actual_amount * 100) if actual_amount > 0 else 0
            refund_rate_by_orders = (refund_orders / actual_orders * 100) if actual_orders > 0 else 0
            unit_price = actual_amount / actual_orders
            promotion_ratio = (promotion_fee / actual_amount * 100) if actual_amount > 0 else 0
            tech_fee = actual_amount * 0.006
            net_profit = gross_profit - refund_amount - promotion_fee - deduction - other_service + other - tech_fee
            net_margin_rate = (net_profit / actual_amount * 100) if actual_amount > 0 else 0
            profit_per_order = net_profit / actual_orders if actual_orders > 0 else 0

            self.calculated_values = {
                "gross_margin_rate": gross_margin_rate,
                "refund_rate_by_amount": refund_rate_by_amount,
                "refund_rate_by_orders": refund_rate_by_orders,
                "unit_price": unit_price,
                "promotion_ratio": promotion_ratio,
                "tech_fee": tech_fee,
                "net_profit": net_profit,
                "net_margin_rate": net_margin_rate,
                "profit_per_order": profit_per_order,
            }

            preview_text = f"""
<b>自动计算结果预览：</b><br>
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━<br>
📊 毛利率: <span style="color:#27ae60">{gross_margin_rate:.2f}%</span><br>
📉 金额退款率: <span style="color:#e74c3c">{refund_rate_by_amount:.2f}%</span><br>
📉 订单退款率: <span style="color:#e74c3c">{refund_rate_by_orders:.2f}%</span><br>
💰 件单价: <span style="color:#3498db">¥{unit_price:.2f}</span><br>
📢 推广占比: <span style="color:#9b59b6">{promotion_ratio:.2f}%</span><br>
🔧 技术服务费: <span style="color:#f39c12">¥{tech_fee:.2f}</span><br>
💵 净利润: <span style="color:#27ae60;font-weight:bold">¥{net_profit:.2f}</span><br>
📈 净利率: <span style="color:#27ae60">{net_margin_rate:.2f}%</span><br>
📊 单笔利润: <span style="color:#3498db">¥{profit_per_order:.2f}</span>
            """
            self.preview_label.setText(preview_text)

        except ValueError:
            self.preview_label.setText("⚠️ 请检查输入数值是否正确")

    def get_data(self):
        return {
            "actual_orders": _parse_number(self.input_fields["实发订单"].text(), integer=True),
            "actual_amount": _parse_number(self.input_fields["实发金额"].text()),
            "gross_profit": _parse_number(self.input_fields["毛利润"].text()),
            "refund_amount": _parse_number(self.input_fields["退款金额"].text()),
            "refund_orders": _parse_number(self.input_fields["退款订单"].text(), integer=True),
            "promotion_fee": _parse_number(self.input_fields["推广费"].text()),
            "deduction": _parse_number(self.input_fields["扣款"].text()),
            "other_service": _parse_number(self.input_fields["其他服务"].text()),
            "other": _parse_number(self.input_fields["其他"].text()),
            **self.calculated_values
        }
