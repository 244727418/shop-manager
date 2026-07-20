# -*- coding: utf-8 -*-
"""店铺毛利管理对话框"""
import os
import json
import time
import re
import hashlib
import threading
import traceback
from io import BytesIO
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QGridLayout, QLabel, QTableWidget, QTableWidgetItem,
    QWidget, QLineEdit, QPushButton, QMessageBox, QMenu, QAction,
    QAbstractItemView, QFileDialog, QComboBox, QScrollArea, QHeaderView,
    QApplication, QPlainTextEdit, QProgressDialog, QCheckBox, QToolTip,
    QListWidget, QListWidgetItem, QDialogButtonBox, QInputDialog
)
from PyQt5.QtCore import Qt, QEvent, QPropertyAnimation, QEasingCurve, QRect, QTimer, pyqtSignal, QByteArray, QBuffer, QIODevice, QObject, QThread, QSize, QEventLoop
from PyQt5.QtGui import QColor, QCursor, QPixmap, QDoubleValidator, QFont
from PyQt5.QtWidgets import QGraphicsOpacityEffect
from PyQt5.QtGui import QClipboard

try:
    from manager.crash_report import append_event
except ImportError:
    try:
        from crash_report import append_event
    except ImportError:
        append_event = None

try:
    from manager.window_icons import apply_window_icon
except ImportError:
    from window_icons import apply_window_icon

try:
    from manager.file_dialog_memory import remembered_existing_directory, remembered_open_file, remembered_save_file
except ImportError:
    from file_dialog_memory import remembered_existing_directory, remembered_open_file, remembered_save_file

try:
    from manager.pdd_browser_monitor import BrowserMonitorError, PddBrowserMonitor
except ImportError:
    from pdd_browser_monitor import BrowserMonitorError, PddBrowserMonitor

try:
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _price_match_summary_html(
    item_count,
    matched_count,
    unmatched_count,
    spec_count,
    price_count,
    marketing_count,
    product_count,
):
    if item_count and not unmatched_count:
        return '<span style="color:#27ae60;">完全匹配</span>'
    if not item_count:
        return '<span style="color:#c0392b;">未抓取到链接</span>'

    parts = []
    if matched_count:
        parts.append(f'<span style="color:#27ae60;">匹配 {matched_count}</span>')
    parts.append(f'<span style="color:#c0392b;">未匹配 {unmatched_count}</span>')
    reasons = [
        ("规格", spec_count),
        ("价格", price_count),
        ("活动/营销", marketing_count),
        ("商品ID", product_count),
    ]
    parts.extend(
        f'<span style="color:#c0392b;">{label} {count}</span>'
        for label, count in reasons
        if count
    )
    return "　".join(parts)


class StoreAiReportWorker(QObject):
    finished = pyqtSignal(str, object)
    failed = pyqtSignal(str)

    def __init__(self, api_url, headers, data):
        super().__init__()
        self.api_url = api_url
        self.headers = headers
        self.data = data
        self.cancelled = False

    def cancel(self):
        self.cancelled = True

    def run(self):
        try:
            import requests
        except ImportError:
            self.failed.emit("缺少 requests 依赖，无法调用 AI。")
            return

        response = None
        try:
            for attempt in range(3):
                if self.cancelled:
                    self.failed.emit("已取消生成报告。")
                    return
                response = requests.post(self.api_url, headers=self.headers, json=self.data, timeout=90)
                if response.status_code not in (500, 503):
                    break
                if attempt < 2:
                    time.sleep(2 * (attempt + 1))

            if self.cancelled:
                self.failed.emit("已取消生成报告。")
                return
            if response is not None and response.status_code == 200:
                result = response.json()
                usage = result.get("usage") or {}
                self.finished.emit(result["choices"][0]["message"]["content"].strip(), usage)
                return

            status = response.status_code if response is not None else "无响应"
            detail = response.text.strip()[:500] if response is not None and response.text else ""
            if status == 503:
                message = "AI调用失败：503\nDeepSeek服务器当前过载，请稍后重试。"
            else:
                message = f"AI调用失败：{status}"
            if detail:
                message += f"\n\n返回内容：{detail}"
            self.failed.emit(message)
        except requests.exceptions.Timeout:
            if not self.cancelled:
                self.failed.emit("AI调用超时，请稍后重试。")
        except requests.exceptions.ConnectionError as e:
            if not self.cancelled:
                self.failed.emit(f"AI连接失败：{str(e)}")
        except Exception as e:
            if not self.cancelled:
                self.failed.emit(f"生成报告失败：{str(e)}")


class PddUnmatchedTaskWindow(QDialog):
    """独立显示价格管理未匹配规格链接，用户可手动标记已处理。"""

    open_code_requested = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(None)
        self.records = {}
        self.setWindowTitle("未匹配规格链接")
        self.setWindowFlags(
            Qt.Window
            | Qt.WindowMinimizeButtonHint
            | Qt.WindowMaximizeButtonHint
            | Qt.WindowCloseButtonHint
        )
        self.setWindowModality(Qt.NonModal)
        self.resize(740, 360)

        layout = QVBoxLayout(self)
        self.lbl_summary = QLabel("未匹配规格链接 0 个")
        self.lbl_summary.setStyleSheet("font-size: 12px; color: #555; padding: 4px 0;")
        layout.addWidget(self.lbl_summary)

        self.table = QTableWidget()
        self.table.setFocusPolicy(Qt.NoFocus)
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["完成", "商品ID", "未匹配原因", "商品标题", "抓取添加编码"])
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setStyleSheet("QTableWidget::item:focus { outline: none; }")
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.table.setColumnWidth(0, 64)
        self.table.setColumnWidth(1, 150)
        self.table.setColumnWidth(2, 180)
        self.table.setColumnWidth(4, 100)
        self.table.cellClicked.connect(self._handle_cell_clicked)
        layout.addWidget(self.table)

    def upsert_records(self, records):
        for record in records or []:
            product_id = str(record.get("product_id") or "").strip()
            if not product_id:
                continue
            existing = self.records.get(product_id, {})
            done = bool(existing.get("done"))
            self.records[product_id] = {
                "product_id": product_id,
                "title": record.get("title") or existing.get("title") or "",
                "reasons": record.get("reasons") or existing.get("reasons") or [],
                "done": done,
            }
        self.refresh()

    def remove_product_ids(self, product_ids):
        for product_id in product_ids or []:
            self.records.pop(str(product_id), None)
        self.refresh()

    def refresh(self):
        records = list(self.records.values())
        records.sort(key=lambda row: (bool(row.get("done")), row.get("product_id", "")))
        self.table.setRowCount(len(records))
        done_count = 0
        for row_index, record in enumerate(records):
            product_id = record.get("product_id", "")
            done = bool(record.get("done"))
            if done:
                done_count += 1
            checkbox = QCheckBox()
            checkbox.setChecked(done)
            checkbox.stateChanged.connect(lambda _state, pid=product_id: self.mark_done(pid))
            checkbox_widget = QWidget()
            checkbox_layout = QHBoxLayout(checkbox_widget)
            checkbox_layout.setContentsMargins(0, 0, 0, 0)
            checkbox_layout.setAlignment(Qt.AlignCenter)
            checkbox_layout.addWidget(checkbox)
            self.table.setCellWidget(row_index, 0, checkbox_widget)

            values = [
                product_id,
                "、".join(record.get("reasons") or []),
                record.get("title", ""),
            ]
            for col_offset, value in enumerate(values, start=1):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignCenter if col_offset != 3 else Qt.AlignVCenter | Qt.AlignLeft)
                if col_offset == 1:
                    item.setToolTip("单击复制商品ID")
                if done:
                    font = item.font()
                    font.setStrikeOut(True)
                    item.setFont(font)
                    item.setForeground(QColor("#95a5a6"))
                    item.setBackground(QColor("#f2f3f4"))
                self.table.setItem(row_index, col_offset, item)
            match_button = QPushButton("抓取添加编码")
            match_button.setStyleSheet("QPushButton { padding: 1px; }")
            match_button.setEnabled(not done and any("规格" in reason for reason in record.get("reasons") or []))
            match_button.clicked.connect(lambda _checked=False, pid=product_id: self.open_code_requested.emit(pid))
            self.table.setCellWidget(row_index, 4, match_button)
        self.lbl_summary.setText(f"未匹配规格链接 {len(records)} 个，已手动完成 {done_count} 个")

    def mark_done(self, product_id):
        product_id = str(product_id or "").strip()
        if product_id in self.records:
            self.records[product_id]["done"] = True
            self.refresh()

    def _handle_cell_clicked(self, row, col):
        if col != 1:
            return
        item = self.table.item(row, col)
        product_id = item.text().strip() if item else ""
        if product_id:
            QApplication.clipboard().setText(product_id)
            self.lbl_summary.setText(f"商品ID {product_id} 已复制。")


class PddProductMatchDialog(QDialog):
    """拼多多商品列表 ID 抓取与本地店铺商品匹配测试窗口。"""

    link_saved = pyqtSignal(str)
    auto_code_search_finished = pyqtSignal(dict)

    def __init__(
        self,
        db,
        monitor,
        default_store_id=None,
        parent=None,
        mode="combined",
        store_id_provider=None,
        owner=None,
        initial_product_id="",
        auto_search=False,
        unmatched_records_provider=None,
    ):
        super().__init__(parent)
        self.db = db
        self.monitor = monitor
        self.default_store_id = default_store_id
        self.mode = mode
        self.store_id_provider = store_id_provider
        self.owner = owner or parent
        self.initial_product_id = str(initial_product_id or "").strip()
        self.auto_search = bool(auto_search)
        self.unmatched_records_provider = unmatched_records_provider
        self._auto_code_search_running = False
        self.missing_ids = []
        self.last_debug_info = {}
        self.last_product_id = ""
        self.last_title = ""
        self.last_product_images = []
        self.last_product_image_data = None
        self.last_specs = []
        self.last_image_download_error = ""
        self.last_price_management_info = {}
        self.price_sync_rows = {}
        self.price_result_cards = {}
        self.price_result_items = {}
        self.price_unmatched_spec_product_ids = []
        self.price_unmatched_records = {}
        self.price_current_page_matched_product_ids = []
        self.unmatched_task_window = None
        self.unmatched_code_dialog = None
        self.setWindowTitle("拼多多链接抓取")
        if self.mode in ("code", "price"):
            self.setWindowFlags(
                Qt.Window
                | Qt.WindowMinimizeButtonHint
                | Qt.WindowMaximizeButtonHint
                | Qt.WindowCloseButtonHint
            )
            self.setWindowModality(Qt.NonModal)
        self.resize(1280 if self.mode == "price" else 1180, 680)
        self.init_ui()
        self.load_stores()
        self.apply_mode_layout()
        self.auto_code_search_finished.connect(self._finish_initial_code_search)
        if self.mode == "price" and self.initial_product_id and self.auto_search:
            QTimer.singleShot(300, self._auto_search_initial_price_product)

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)

        self.top_widget = QWidget()
        top_layout = QHBoxLayout(self.top_widget)
        top_layout.setContentsMargins(0, 0, 0, 0)

        self.chk_store_scope = QCheckBox("按所选店铺匹配")
        self.chk_store_scope.setChecked(True)
        self.chk_store_scope.setEnabled(False)
        top_layout.addWidget(self.chk_store_scope)

        self.combo_store = QComboBox()
        self.combo_store.setMinimumWidth(220)
        top_layout.addWidget(self.combo_store)

        self.btn_open_browser = QPushButton("打开商家端")
        self.btn_open_browser.clicked.connect(self.open_browser)
        top_layout.addWidget(self.btn_open_browser)

        self.btn_scan = QPushButton("抓取添加编码界面")
        self.btn_scan.clicked.connect(self.scan_current_page)
        self.btn_scan.setStyleSheet(
            "QPushButton { background-color: #f39c12; color: white; border: none; padding: 6px 12px; border-radius: 4px; font-weight: bold; }"
            "QPushButton:hover { background-color: #d68910; }"
            "QPushButton:disabled { background-color: #bdc3c7; }"
        )
        top_layout.addWidget(self.btn_scan)

        self.btn_scan_price = QPushButton("抓取价格管理")
        self.btn_scan_price.clicked.connect(self.scan_price_management)
        self.btn_scan_price.setStyleSheet(
            "QPushButton { background-color: #8e44ad; color: white; border: none; padding: 6px 12px; border-radius: 4px; font-weight: bold; }"
            "QPushButton:hover { background-color: #7d3c98; }"
            "QPushButton:disabled { background-color: #bdc3c7; }"
        )
        top_layout.addWidget(self.btn_scan_price)

        self.btn_refresh_price_match = QPushButton("刷新匹配状态")
        self.btn_refresh_price_match.clicked.connect(self.refresh_all_price_management_matches)
        self.btn_refresh_price_match.setVisible(False)
        self.btn_refresh_price_match.setEnabled(False)
        self.btn_refresh_price_match.setStyleSheet(
            "QPushButton { background-color: #16a085; color: white; border: none; padding: 6px 12px; border-radius: 4px; font-weight: bold; }"
            "QPushButton:hover { background-color: #138d75; }"
            "QPushButton:disabled { background-color: #bdc3c7; }"
        )
        self.btn_sync_all_price = QPushButton("一键同步所有未匹配价格和促销")
        self.btn_sync_all_price.clicked.connect(self.sync_all_price_management_products)
        self.btn_sync_all_price.setVisible(False)
        self.btn_sync_all_price.setEnabled(False)
        self.btn_sync_all_price.setStyleSheet(
            "QPushButton { background-color: #27ae60; color: white; border: none; padding: 6px 12px; border-radius: 4px; font-weight: bold; }"
            "QPushButton:hover { background-color: #229954; }"
            "QPushButton:disabled { background-color: #bdc3c7; }"
        )
        top_layout.addStretch()
        layout.addWidget(self.top_widget)

        self.lbl_summary = QLabel("先手动打开拼多多“添加/编辑商品编码”窗口，或进入价格管理页面，再点击对应抓取按钮。")
        self.lbl_summary.setWordWrap(True)
        self.lbl_summary.setStyleSheet("color: #555; font-size: 12px; padding: 6px 0;")
        layout.addWidget(self.lbl_summary)

        self.unmatched_specs_widget = QWidget()
        unmatched_specs_layout = QHBoxLayout(self.unmatched_specs_widget)
        unmatched_specs_layout.setContentsMargins(0, 0, 0, 0)
        self.lbl_unmatched_specs = QLabel("未匹配规格链接ID：无")
        self.lbl_unmatched_specs.setWordWrap(True)
        self.lbl_unmatched_specs.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.lbl_unmatched_specs.setStyleSheet("color: #8e44ad; font-size: 12px; padding: 2px 0;")
        unmatched_specs_layout.addWidget(self.lbl_unmatched_specs, 1)
        self.unmatched_specs_widget.setVisible(False)
        layout.addWidget(self.unmatched_specs_widget)

        self.start_widget = QWidget()
        start_layout = QHBoxLayout(self.start_widget)
        start_layout.setContentsMargins(0, 0, 0, 0)
        self.btn_start_scan = QPushButton("开始抓取")
        self.btn_start_scan.clicked.connect(self.start_current_mode_scan)
        self.btn_start_scan.setStyleSheet(
            "QPushButton { background-color: #2e86de; color: white; border: none; padding: 7px 14px; border-radius: 4px; font-weight: bold; }"
            "QPushButton:hover { background-color: #1b4f72; }"
            "QPushButton:disabled { background-color: #bdc3c7; }"
        )
        start_layout.addWidget(self.btn_start_scan)
        start_layout.addWidget(self.btn_refresh_price_match)
        start_layout.addWidget(self.btn_sync_all_price)
        start_layout.addStretch()
        self.start_widget.setVisible(False)
        layout.addWidget(self.start_widget)

        self.header_widget = QWidget()
        header_layout = QGridLayout(self.header_widget)
        header_layout.setContentsMargins(0, 0, 0, 0)
        self.lbl_pdd_product_title = QLabel("标题: --")
        self.lbl_software_product_title = QLabel("软件标题: --")
        self.lbl_pdd_product_images = QLabel("主图")
        self.product_images_widget = QWidget()
        self.product_images_layout = QHBoxLayout(self.product_images_widget)
        self.product_images_layout.setContentsMargins(4, 4, 4, 4)
        self.product_images_layout.setSpacing(0)
        self.product_images_layout.setAlignment(Qt.AlignCenter)
        self.product_images_layout.addWidget(self.lbl_pdd_product_images)
        self.lbl_pdd_product_title.setWordWrap(True)
        self.lbl_software_product_title.setWordWrap(True)
        self.lbl_pdd_product_images.setWordWrap(True)
        for label in (self.lbl_pdd_product_title, self.lbl_software_product_title):
            label.setAlignment(Qt.AlignCenter)
            label.setMinimumHeight(46)
            label.setStyleSheet("background-color: #f8f9fa; border: 1px solid #dee2e6; border-radius: 4px; padding: 6px;")
        self.product_images_widget.setFixedSize(96, 96)
        self.product_images_widget.setStyleSheet("background-color: #f8f9fa; border: 1px solid #dee2e6; border-radius: 4px;")
        header_layout.addWidget(self.product_images_widget, 0, 0, 2, 1, Qt.AlignCenter)
        header_layout.addWidget(self.lbl_pdd_product_title, 0, 1)
        header_layout.addWidget(self.lbl_software_product_title, 1, 1)
        header_layout.setColumnStretch(1, 1)
        layout.addWidget(self.header_widget)

        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["规格图", "规格信息", "规格编码", "价格", "商品匹配", "软件规格", "原始文本"])
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setFocusPolicy(Qt.NoFocus)
        self.table.setStyleSheet(
            "QTableWidget::item { color: #000; font-weight: normal; }"
            "QTableWidget::item:selected { color: #000; font-weight: normal; background-color: #dbeafe; outline: none; }"
            "QTableWidget::item:focus { outline: none; border: none; }"
        )
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_code_table_context_menu)
        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(58)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setColumnWidth(0, 58)
        self.table.setColumnWidth(1, 210)
        self.table.setColumnWidth(2, 150)
        self.table.setColumnWidth(3, 70)
        self.table.setColumnWidth(4, 72)
        self.table.setColumnWidth(5, 150)
        layout.addWidget(self.table)

        self.price_scroll_area = QScrollArea()
        self.price_scroll_area.setWidgetResizable(True)
        self.price_scroll_area.setVisible(False)
        self.price_scroll_content = QWidget()
        self.price_scroll_layout = QVBoxLayout(self.price_scroll_content)
        self.price_scroll_layout.setContentsMargins(0, 0, 0, 0)
        self.price_scroll_layout.setSpacing(10)
        self.price_scroll_layout.addStretch()
        self.price_scroll_area.setWidget(self.price_scroll_content)
        layout.addWidget(self.price_scroll_area)

        bottom = QWidget()
        bottom_layout = QHBoxLayout(bottom)
        bottom_layout.setContentsMargins(0, 0, 0, 0)
        self.btn_save_to_store = QPushButton("创建/覆盖到本店铺")
        self.btn_save_to_store.clicked.connect(self.save_current_link_to_store)
        self.btn_save_to_store.setStyleSheet(
            "QPushButton { background-color: #27ae60; color: white; border: none; padding: 6px 12px; border-radius: 4px; font-weight: bold; }"
            "QPushButton:hover { background-color: #229954; }"
        )
        self.chk_include_product_image = QCheckBox("包含主图")
        self.chk_include_product_image.setChecked(True)
        self.chk_include_product_image.setToolTip("取消勾选后，创建/覆盖时不写入商品主图，保留软件里已有或手动上传的主图。")
        bottom_layout.addWidget(self.chk_include_product_image)
        bottom_layout.addWidget(self.btn_save_to_store)
        self.btn_overwrite_without_price = QPushButton("覆盖除价格之外的信息到软件")
        self.btn_overwrite_without_price.clicked.connect(self.overwrite_current_link_without_price)
        self.btn_overwrite_without_price.setStyleSheet(
            "QPushButton { background-color: #16a085; color: white; border: none; padding: 6px 12px; border-radius: 4px; font-weight: bold; }"
            "QPushButton:hover { background-color: #138d75; }"
        )
        bottom_layout.addWidget(self.btn_overwrite_without_price)
        self.btn_open_current_spec = QPushButton("打开当前链接规格与毛利管理")
        self.btn_open_current_spec.clicked.connect(self.open_current_code_product_spec_dialog)
        self.btn_open_current_spec.setEnabled(False)
        bottom_layout.addWidget(self.btn_open_current_spec)
        self.btn_copy_debug = QPushButton("复制调试JSON")
        self.btn_copy_debug.clicked.connect(self.copy_debug_json)
        bottom_layout.addWidget(self.btn_copy_debug)
        self.btn_copy_unmatched_specs = QPushButton("未匹配规格链接")
        self.btn_copy_unmatched_specs.clicked.connect(self.show_unmatched_task_window)
        self.btn_copy_unmatched_specs.setVisible(False)
        self.btn_copy_unmatched_specs.setEnabled(False)
        bottom_layout.addWidget(self.btn_copy_unmatched_specs)
        bottom_layout.addStretch()
        btn_close = QPushButton("关闭")
        btn_close.clicked.connect(self.accept)
        bottom_layout.addWidget(btn_close)
        layout.addWidget(bottom)

    def load_stores(self):
        self.combo_store.clear()
        fixed_store_id = self.default_store_id
        if self.store_id_provider:
            try:
                provider_store_id = self.store_id_provider()
                if provider_store_id:
                    fixed_store_id = provider_store_id
            except Exception:
                pass

        if fixed_store_id is not None:
            rows = self.db.safe_fetchall("SELECT id, name FROM stores WHERE id=?", (fixed_store_id,))
            if rows:
                store_id, store_name = rows[0]
                self.combo_store.addItem(str(store_name or f"店铺{store_id}"), int(store_id))
            else:
                self.combo_store.addItem(f"店铺{fixed_store_id}", int(fixed_store_id))
            self.combo_store.setEnabled(False)
            self.combo_store.setToolTip("已固定为当前打开窗口所属店铺，不能在此切换店铺。")
            return

        rows = self.db.safe_fetchall("SELECT id, name FROM stores ORDER BY sort_order, id")
        for store_id, store_name in rows:
            self.combo_store.addItem(str(store_name or f"店铺{store_id}"), int(store_id))
        self.combo_store.setEnabled(True)

    def current_store_id(self):
        if self.store_id_provider:
            try:
                store_id = self.store_id_provider()
                if store_id:
                    return store_id
            except Exception:
                pass
        return self.combo_store.currentData()

    def _activate_browser_store_context(self):
        store_id = self.current_store_id()
        if store_id and hasattr(self.monitor, "set_store_context"):
            self.monitor.set_store_context(store_id)
        return store_id

    def apply_mode_layout(self):
        if self.mode == "code":
            self.setWindowTitle("抓取添加编码界面")
            self.resize(1080, 640)
            self.top_widget.hide()
            self.start_widget.show()
            self.btn_scan_price.hide()
            self.btn_refresh_price_match.hide()
            self.btn_sync_all_price.hide()
            self.btn_copy_unmatched_specs.hide()
            self.btn_save_to_store.show()
            self.btn_overwrite_without_price.show()
            self.btn_open_current_spec.show()
            if hasattr(self, "chk_include_product_image"):
                self.chk_include_product_image.show()
            self._set_code_table_mode()
            self._refresh_unmatched_spec_ids_panel()
            self.lbl_summary.setText("已打开抓取添加编码界面窗口。请确认浏览器停在添加/编辑编码界面后，点击“开始抓取”。")
        elif self.mode == "price":
            self.setWindowTitle("抓取价格管理")
            self.resize(1380, 720)
            self.top_widget.hide()
            self.start_widget.show()
            self.btn_save_to_store.hide()
            self.btn_overwrite_without_price.hide()
            self.btn_open_current_spec.hide()
            self.btn_refresh_price_match.show()
            self.btn_refresh_price_match.setEnabled(False)
            self.btn_sync_all_price.show()
            self.btn_sync_all_price.setEnabled(False)
            if hasattr(self, "chk_include_product_image"):
                self.chk_include_product_image.hide()
            self._set_price_table_mode()
            if hasattr(self, "unmatched_specs_widget"):
                self.unmatched_specs_widget.hide()
            if self.initial_product_id and self.auto_search:
                self.lbl_summary.setText(
                    f"已打开抓取价格管理窗口。将尝试搜索商品ID {self.initial_product_id}，请确认页面结果后点击“开始抓取”。"
                )
            elif self.initial_product_id:
                self.lbl_summary.setText(
                    f"已打开抓取价格管理窗口。请确认浏览器已进入价格管理页面，之后再点击“开始抓取”。当前商品ID：{self.initial_product_id}"
                )
            else:
                self.lbl_summary.setText("已打开抓取价格管理窗口。请确认浏览器停在价格管理页面后，点击“开始抓取”。")

    def _auto_search_initial_price_product(self):
        product_id = str(self.initial_product_id or "").strip()
        if not product_id:
            return
        try:
            result = self.monitor.search_price_management_product(product_id)
        except Exception as e:
            result = {"ok": False, "status": f"自动搜索失败：{e}"}
        status = result.get("status") if isinstance(result, dict) else ""
        if not status:
            status = f"已尝试搜索商品ID {product_id}"
        self.lbl_summary.setText(f"{status}。请确认价格管理页面结果后，点击“开始抓取”。")

    def start_initial_code_search(self, product_id):
        product_id = str(product_id or "").strip()
        if not product_id:
            return
        QApplication.clipboard().setText(product_id)
        self.initial_product_id = product_id
        if self._auto_code_search_running:
            return
        self._auto_code_search_running = True
        store_id = self.current_store_id()
        store_name = self.combo_store.currentText().strip()
        self.lbl_summary.setText(f"商品ID {product_id} 已复制，正在打开商品列表并自动搜索...")
        threading.Thread(
            target=self._run_initial_code_search,
            args=(product_id, store_id, store_name),
            daemon=True,
        ).start()

    def _run_initial_code_search(self, product_id, store_id, store_name):
        try:
            result = self.monitor.open_goods_list_and_search_product(
                product_id,
                expected_store_name=store_name,
                store_id=store_id,
            )
        except Exception as e:
            result = {"ok": False, "status": f"自动搜索失败：{e}"}
        try:
            self.auto_code_search_finished.emit(result if isinstance(result, dict) else {})
        except RuntimeError:
            pass

    def _finish_initial_code_search(self, result):
        self._auto_code_search_running = False
        status = str((result or {}).get("status") or "未能自动搜索商品ID").strip()
        self.lbl_summary.setText(status)

    def start_current_mode_scan(self):
        if self.mode == "price":
            self.scan_price_management()
        else:
            self.scan_current_page()

    def _set_scan_controls_enabled(self, enabled):
        for button_name in ("btn_scan", "btn_scan_price", "btn_start_scan"):
            button = getattr(self, button_name, None)
            if button is not None:
                button.setEnabled(enabled)

    def open_browser(self):
        try:
            store_id = self._activate_browser_store_context()
            if hasattr(self.monitor, "activate_store_browser"):
                self.monitor.activate_store_browser(store_id, open_url=True, open_new_tab=False)
            else:
                self.monitor.open_merchant_page(store_id)
            self.lbl_summary.setText("已打开商家端。请手动进入“添加/编辑商品编码”窗口或价格管理页面，再点击对应抓取按钮。")
        except Exception as e:
            QMessageBox.warning(self, "拼多多链接抓取", f"打开商家端失败：{e}")

    def _local_products_for_store(self, store_id):
        rows = self.db.safe_fetchall(
            "SELECT id, name, title, image_data FROM products WHERE store_id=? AND COALESCE(is_archived, 0)=0",
            (store_id,),
        )
        result = {}
        for sys_id, product_id, title, image_data in rows:
            product_id = str(product_id or "").strip()
            if product_id:
                result[product_id] = {"id": sys_id, "title": str(title or ""), "image_data": image_data}
        return result

    def _local_specs_for_product(self, product_sys_id):
        if not product_sys_id:
            return {}
        rows = self.db.safe_fetchall(
            "SELECT spec_code, spec_name FROM product_specs WHERE product_id=?",
            (product_sys_id,),
        )
        result = {}
        for spec_code, spec_name in rows:
            raw_code = str(spec_code or "").strip()
            normalized_code = self._normalize_spec_code(raw_code)
            if raw_code:
                result[raw_code] = str(spec_name or "").strip()
            if normalized_code:
                result[normalized_code] = str(spec_name or "").strip()
        return result

    def _normalize_spec_code(self, value):
        return "".join(ch for ch in str(value or "").strip() if ch.isalnum())

    def _local_specs_by_name_for_product(self, product_sys_id):
        if not product_sys_id:
            return {}
        rows = self.db.safe_fetchall(
            "SELECT id, spec_name, sale_price FROM product_specs WHERE product_id=?",
            (product_sys_id,),
        )
        result = {}
        for spec_id, spec_name, sale_price in rows:
            name = str(spec_name or "").strip()
            if name:
                spec_payload = {
                    "id": spec_id,
                    "spec_name": name,
                    "sale_price": float(sale_price or 0),
                }
                result[name] = spec_payload
                normalized_name = self._normalize_spec_name_for_match(name)
                if normalized_name:
                    result.setdefault(normalized_name, spec_payload)
                if self.db.get_setting(f"product_spec_two_level_{product_sys_id}", "0") == "1" and " " in name:
                    first, second = name.rsplit(" ", 1)
                    combined_name = f"{first.strip()}-{second.strip()}"
                    result.setdefault(combined_name, spec_payload)
                    result.setdefault(self._normalize_spec_name_for_match(combined_name), spec_payload)
        return result

    def _normalize_spec_name_for_match(self, value):
        text = str(value or "").strip()
        text = re.sub(r"\s*-\s*", " ", text)
        text = text.replace(" ", "").replace("\u3000", "")
        text = text.replace("丨", "|")
        return text

    def _dedupe_specs(self, specs):
        best_by_key = {}
        order = []
        draft_warning_pattern = re.compile(r"(草稿箱|正在编辑|确认修改当前商品编码|自动删除草稿|去查看草稿)")
        def normalize_name(value):
            text = str(value or "").strip()
            for prefix in ("规格信息 ", "规格编码 ", "规格 ", "编码 "):
                while text.startswith(prefix):
                    text = text[len(prefix):].strip()
            text = re.sub(r"\s*(?:当前价|价格|售价)[:：]?\s*[￥¥]?\s*\d+(?:\.\d{1,2})?.*$", "", text).strip()
            return text

        def score(item):
            fields = [
                normalize_name(item.get("spec_info", "")),
                str(item.get("spec_code", "") or "").strip(),
                str(item.get("price", "") or "").strip(),
                str(item.get("image", "") or "").strip(),
            ]
            raw_text = str(item.get("raw_text", "") or "")
            aggregate_penalty = 500 if raw_text.count("¥") > 1 or len(raw_text) > 160 else 0
            concise_bonus = 80 if len(raw_text) <= 80 else 40 if len(raw_text) <= 120 else 0
            return sum(1 for value in fields if value) * 1000 + concise_bonus - aggregate_penalty

        for spec in specs:
            spec_code = str(spec.get("spec_code", "") or "").strip()
            spec_name = normalize_name(spec.get("spec_info", ""))
            compact_name = spec_name.replace(" ", "").replace("\u3000", "")
            compact_raw = str(spec.get("raw_text", "") or "").replace(" ", "").replace("\u3000", "")
            compact_raw = compact_raw.replace("请输入", "").replace("已输入", "")
            if draft_warning_pattern.search(spec_name) or draft_warning_pattern.search(str(spec.get("raw_text", "") or "")):
                continue
            if compact_name.startswith("商品编码") and spec_code and not str(spec.get("price", "") or "").strip():
                continue
            if re.fullmatch(r"商品编码[A-Za-z0-9_-]{3,}", compact_raw or ""):
                continue
            if spec_code and not spec_name and not str(spec.get("price", "") or "").strip() and not str(spec.get("image", "") or "").strip():
                continue
            if not spec_code and not spec_name:
                continue
            key = spec_name or spec_code
            cleaned = dict(spec)
            cleaned["spec_info"] = spec_name
            cleaned["raw_text"] = str(cleaned.get("raw_text", "") or "").replace("请输入", "").replace("已输入", "").strip()
            if key not in best_by_key:
                best_by_key[key] = cleaned
                order.append(key)
            elif score(cleaned) > score(best_by_key[key]):
                best_by_key[key] = cleaned
        return [best_by_key[key] for key in order]

    def _set_product_image_previews(self, image_urls):
        self.last_product_image_data = None
        while self.product_images_layout.count():
            item = self.product_images_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        image_urls = [str(url or "").strip() for url in image_urls if str(url or "").strip()]
        if not image_urls:
            empty = QLabel("主图: --")
            empty.setAlignment(Qt.AlignCenter)
            self.product_images_layout.addWidget(empty, 0, Qt.AlignCenter)
            return

        for image_url in image_urls[:1]:
            label = QLabel()
            label.setAlignment(Qt.AlignCenter)
            label.setFixedSize(88, 88)
            label.setToolTip(image_url)
            pixmap = QPixmap()
            image_data = self._download_image_bytes(self._preview_image_url(image_url), timeout=3)
            if image_data:
                self.last_product_image_data = image_data
                pixmap.loadFromData(image_data)
            if not pixmap.isNull():
                label.setPixmap(pixmap.scaled(88, 88, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            else:
                label.setText("加载失败")
            self.product_images_layout.addWidget(label, 0, Qt.AlignCenter)

    def _set_image_preview_cell(self, row, col, image_url):
        image_url = str(image_url or "").strip()
        if not image_url:
            item = QTableWidgetItem("")
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, col, item)
            return

        label = QLabel()
        label.setAlignment(Qt.AlignCenter)
        label.setFixedSize(54, 54)
        label.setToolTip(image_url)
        pixmap = QPixmap()
        image_data = self._download_image_bytes(self._preview_image_url(image_url), timeout=3)
        if image_data:
            pixmap.loadFromData(image_data)

        if not pixmap.isNull():
            label.setPixmap(pixmap.scaled(50, 50, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            self.table.setCellWidget(row, col, label)
            self.table.setRowHeight(row, 58)
        else:
            item = QTableWidgetItem("加载失败")
            item.setTextAlignment(Qt.AlignCenter)
            item.setToolTip(
                f"{image_url}\n\n{self.last_image_download_error}"
                if self.last_image_download_error else image_url
            )
            item.setData(Qt.UserRole, image_url)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, col, item)

    def _code_table_image_url(self, row):
        widget = self.table.cellWidget(row, 0)
        if widget is not None:
            return str(widget.toolTip() or "").strip()
        item = self.table.item(row, 0)
        if not item:
            return ""
        return str(item.data(Qt.UserRole) or item.toolTip() or item.text() or "").strip()

    def _set_code_table_item(self, row, col, value, editable=True, color=None):
        item = QTableWidgetItem(str(value or ""))
        item.setTextAlignment(Qt.AlignCenter)
        item.setToolTip(str(value or ""))
        if not editable:
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        if color:
            item.setForeground(QColor(color))
        self.table.setItem(row, col, item)

    def _code_table_specs_from_view(self):
        specs = []
        for row in range(self.table.rowCount()):
            def cell_text(col):
                item = self.table.item(row, col)
                return str(item.text() if item else "").strip()

            spec_info = cell_text(1)
            spec_code = cell_text(2)
            price = cell_text(3)
            raw_text = cell_text(6)
            image = self._code_table_image_url(row)
            if not spec_info and not spec_code and not price and not image:
                continue
            if not spec_info and not image:
                continue
            specs.append({
                "spec_info": spec_info,
                "spec_code": spec_code,
                "price": price,
                "image": image,
                "raw_text": raw_text,
            })
        return self._dedupe_specs(specs)

    def _sync_code_table_specs_to_state(self):
        if self.mode == "code" and self.table.isVisible():
            self.last_specs = self._code_table_specs_from_view()

    def show_code_table_context_menu(self, pos):
        if self.mode != "code" or not self.table.isVisible():
            return
        row = self.table.rowAt(pos.y())
        if row < 0:
            return
        self.table.selectRow(row)
        menu = QMenu(self)
        action_delete = QAction("删除当前规格行", self)
        action_delete.triggered.connect(lambda _checked=False, r=row: self.delete_code_table_row(r))
        menu.addAction(action_delete)
        menu.exec_(self.table.viewport().mapToGlobal(pos))

    def delete_code_table_row(self, row):
        if row < 0 or row >= self.table.rowCount():
            return
        self.table.removeRow(row)
        self._sync_code_table_specs_to_state()
        self.lbl_summary.setText(f"已删除 1 行规格，当前界面剩余 {self.table.rowCount()} 行。保存/覆盖将使用当前界面内容。")

    def _preview_image_url(self, image_url):
        return re.sub(r"imageView2/2/w/\d+(?:/q/\d+)?", "imageView2/2/w/200/q/50", str(image_url or "").strip())

    def _download_image_bytes(self, image_url, timeout=5):
        image_url = str(image_url or "").strip()
        self.last_image_download_error = ""
        if not image_url:
            return None
        try:
            import requests
            response = requests.get(
                image_url,
                timeout=timeout,
                headers={
                    "User-Agent": "Mozilla/5.0",
                    "Referer": "https://mms.pinduoduo.com/",
                },
            )
            response.raise_for_status()
            if response.content:
                return response.content
            self.last_image_download_error = "empty response body"
        except Exception as e:
            self.last_image_download_error = f"{type(e).__name__}: {e}"
            print(f"image download failed {image_url}: {self.last_image_download_error}")
        return None

    def _parse_price_value(self, value):
        text = str(value or "").strip().replace("￥", "").replace("¥", "").replace(",", "")
        try:
            return float(text) if text else None
        except ValueError:
            return None

    def _money_equal(self, left, right):
        if left is None or right is None:
            return False
        return abs(float(left) - float(right)) <= 0.01

    def _fmt_money(self, value):
        try:
            return f"{float(value):.2f}"
        except Exception:
            return "--"

    def _price_tag_flags(self, tag_type):
        text = str(tag_type or "").strip()
        return {
            "is_limited_time": 1 if text == "限时限量购" else 0,
            "is_marketing": 1 if text == "营销活动" else 0,
        }

    def _set_code_table_mode(self):
        self.price_sync_rows = {}
        self.price_unmatched_spec_product_ids = []
        self.price_unmatched_records = {}
        self.price_current_page_matched_product_ids = []
        self.btn_save_to_store.setEnabled(True)
        if hasattr(self, "btn_open_current_spec"):
            self.btn_open_current_spec.setVisible(self.mode != "price")
            self.btn_open_current_spec.setEnabled(False)
        if hasattr(self, "header_widget"):
            self.header_widget.show()
        self.lbl_pdd_product_title.show()
        self.lbl_software_product_title.show()
        self.product_images_widget.show()
        if hasattr(self, "btn_copy_unmatched_specs"):
            self.btn_copy_unmatched_specs.setVisible(False)
            self.btn_copy_unmatched_specs.setEnabled(False)
        if hasattr(self, "btn_sync_all_price"):
            self.btn_sync_all_price.setVisible(False)
            self.btn_sync_all_price.setEnabled(False)
        if hasattr(self, "chk_include_product_image"):
            self.chk_include_product_image.setVisible(True)
        self.price_scroll_area.setVisible(False)
        self.table.setVisible(True)
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        self._clear_price_result_cards()
        self.table.clear()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["规格图", "规格信息", "规格编码", "价格", "商品匹配", "软件规格", "原始文本"])
        self.table.horizontalHeader().setStretchLastSection(True)
        widths = [58, 210, 150, 70, 72, 150, 220]
        for col, width in enumerate(widths):
            self.table.setColumnWidth(col, width)

    def _set_price_table_mode(self):
        self.btn_save_to_store.setEnabled(False)
        if hasattr(self, "btn_open_current_spec"):
            self.btn_open_current_spec.setVisible(False)
        self.price_unmatched_spec_product_ids = []
        self.price_unmatched_records = {}
        self.price_current_page_matched_product_ids = []
        if hasattr(self, "header_widget"):
            self.header_widget.hide()
        if hasattr(self, "btn_copy_unmatched_specs"):
            self.btn_copy_unmatched_specs.setVisible(True)
            self.btn_copy_unmatched_specs.setEnabled(False)
        if hasattr(self, "btn_sync_all_price"):
            self.btn_sync_all_price.setVisible(True)
            self.btn_sync_all_price.setEnabled(False)
        if hasattr(self, "chk_include_product_image"):
            self.chk_include_product_image.setVisible(False)
        self.lbl_pdd_product_title.setText("价格管理: --")
        self.lbl_software_product_title.setText("匹配结果: --")
        self._set_product_image_previews([])
        self.table.setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.price_scroll_area.setVisible(True)
        self._clear_price_result_cards()

    def _clear_price_result_cards(self):
        if not hasattr(self, "price_scroll_layout"):
            return
        self.price_result_cards = {}
        self.price_result_items = {}
        self.price_unmatched_spec_product_ids = []
        self.price_unmatched_records = {}
        self.price_current_page_matched_product_ids = []
        if hasattr(self, "btn_copy_unmatched_specs"):
            self.btn_copy_unmatched_specs.setEnabled(False)
        if hasattr(self, "btn_refresh_price_match"):
            self.btn_refresh_price_match.setEnabled(False)
        if hasattr(self, "btn_sync_all_price"):
            self.btn_sync_all_price.setEnabled(False)
        while self.price_scroll_layout.count():
            item = self.price_scroll_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
        self.price_scroll_layout.addStretch()

    def _product_status_text(self, compare):
        target_tag = compare.get("target_tag") or {}
        if int(target_tag.get("is_limited_time") or 0):
            return "限时限量购"
        if int(target_tag.get("is_marketing") or 0):
            return "营销活动"
        return "裸价"

    def _product_discount_text(self, compare):
        target_discount = compare.get("target_discount") or {}
        coupon = float(target_discount.get("coupon_amount") or 0)
        new_customer = float(target_discount.get("new_customer_discount") or 0)
        parts = []
        if coupon > 0:
            parts.append(f"优惠券 {coupon:.2f} 元")
        if new_customer > 0:
            parts.append(f"新客立减 {new_customer:.2f} 元")
        return "；".join(parts) if parts else "无优惠券"

    def _set_price_card_image(self, label, image_url, image_data):
        pixmap = QPixmap()
        image_url = str(image_url or "").strip()
        if image_url:
            downloaded = self._download_image_bytes(image_url, timeout=3)
            if downloaded:
                pixmap.loadFromData(downloaded)
        if pixmap.isNull() and image_data:
            try:
                pixmap.loadFromData(image_data)
            except Exception:
                pixmap = QPixmap()
        if pixmap.isNull():
            label.setText("无图")
            label.setAlignment(Qt.AlignCenter)
            return
        target_size = max(24, min(label.width() or 64, label.height() or 64) - 6)
        label.setPixmap(pixmap.scaled(target_size, target_size, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        label.setAlignment(Qt.AlignCenter)

    def _copy_price_product_id(self, product_id):
        product_id = str(product_id or "").strip()
        QApplication.clipboard().setText(product_id)
        self.lbl_summary.setText(f"商品ID {product_id} 已复制。")

    def _add_price_result_card(self, item, local_product, compare, insert_at=None):
        product_id = str(item.get("product_id") or "")
        card = QWidget()
        card.setStyleSheet(
            "QWidget#priceCard { background-color: #ffffff; border: 1px solid #dfe6e9; border-radius: 6px; }"
        )
        card.setObjectName("priceCard")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(10, 10, 10, 10)
        card_layout.setSpacing(6)

        top = QWidget()
        top_layout = QHBoxLayout(top)
        top_layout.setContentsMargins(0, 0, 0, 0)
        top_layout.setSpacing(8)

        image_label = QLabel()
        image_label.setFixedSize(58, 58)
        image_label.setStyleSheet("background-color: #f8f9fa; border: 1px solid #dfe6e9; border-radius: 4px; color: #95a5a6;")
        self._set_price_card_image(image_label, item.get("image"), local_product.get("image_data"))
        top_layout.addWidget(image_label, 0, Qt.AlignVCenter)

        id_btn = QPushButton(f"商品ID: {product_id}")
        id_btn.setToolTip("单击复制商品ID")
        id_btn.setStyleSheet(
            "QPushButton { background-color: #ecf5ff; color: #21618c; border: 1px solid #aed6f1; padding: 4px 8px; border-radius: 4px; text-align: left; font-weight: bold; }"
            "QPushButton:hover { background-color: #d6eaf8; }"
        )
        id_btn.clicked.connect(lambda _checked=False, pid=product_id: self._copy_price_product_id(pid))
        top_layout.addWidget(id_btn, 0)

        status_label = QLabel(self._product_status_text(compare))
        status_label.setAlignment(Qt.AlignCenter)
        status_color = "#c0392b" if self._product_status_text(compare) == "限时限量购" else "#7d3c98" if self._product_status_text(compare) == "营销活动" else "#1e8449"
        status_label.setStyleSheet(f"color: {status_color}; background-color: #f8f9fa; border: 1px solid #dfe6e9; border-radius: 4px; padding: 4px 8px; font-weight: bold;")
        top_layout.addWidget(status_label, 0)

        discount_label = QLabel(self._product_discount_text(compare))
        discount_label.setStyleSheet("color: #7d6608; background-color: #fcf3cf; border: 1px solid #f7dc6f; border-radius: 4px; padding: 4px 8px;")
        top_layout.addWidget(discount_label, 0)

        top_layout.addStretch()

        action_widget = QWidget()
        action_layout = QVBoxLayout(action_widget)
        action_layout.setContentsMargins(0, 0, 0, 0)
        action_layout.setSpacing(5)

        sync_btn = QPushButton("同步价格和营销")
        sync_btn.setEnabled(bool(compare.get("can_sync")))
        sync_btn.setToolTip("规格名称一致且存在差异时可同步；已匹配或规格不一致时不可同步。")
        sync_btn.setStyleSheet(
            "QPushButton { background-color: #27ae60; color: white; border: none; padding: 5px 10px; border-radius: 4px; font-weight: bold; }"
            "QPushButton:hover { background-color: #229954; }"
            "QPushButton:disabled { background-color: #bdc3c7; color: #666; }"
        )
        sync_btn.clicked.connect(lambda _checked=False, pid=product_id: self.sync_price_management_product(pid))
        action_layout.addWidget(sync_btn)

        edit_btn = QPushButton("手动编辑")
        edit_btn.setToolTip("打开软件里的规格与毛利管理窗口，关闭后刷新当前匹配状态。")
        edit_btn.setStyleSheet(
            "QPushButton { background-color: #3498db; color: white; border: none; padding: 5px 10px; border-radius: 4px; font-weight: bold; }"
            "QPushButton:hover { background-color: #2e86c1; }"
        )
        edit_btn.clicked.connect(lambda _checked=False, pid=product_id: self.open_price_product_spec_dialog(pid))
        action_layout.addWidget(edit_btn)
        code_btn = QPushButton("抓取添加编码")
        code_btn.setEnabled(bool(self._price_compare_categories(compare).get("spec")))
        code_btn.setToolTip("仅打开抓取添加编码窗口，不自动搜索商品列表。")
        code_btn.setStyleSheet(
            "QPushButton { background-color: #f39c12; color: white; border: none; padding: 1px; border-radius: 4px; font-weight: bold; }"
            "QPushButton:hover { background-color: #d68910; }"
            "QPushButton:disabled { background-color: #bdc3c7; color: #666; }"
        )
        code_btn.clicked.connect(lambda _checked=False, pid=product_id: self.open_unmatched_code_match(pid))
        action_layout.addWidget(code_btn)
        top_layout.addWidget(action_widget, 0)
        card_layout.addWidget(top)

        title_label = QLabel(item.get("title") or local_product.get("title") or "--")
        title_label.setWordWrap(True)
        title_label.setStyleSheet("font-size: 13px; color: #2c3e50; padding: 0;")
        card_layout.addWidget(title_label)

        spec_table = QTableWidget()
        spec_table.setColumnCount(10)
        spec_table.setHorizontalHeaderLabels(["规格名称", "软件价格", "抓取价格", "价格标签", "优惠券", "新客立减", "店铺满减", "软件券后价", "抓取实收价格", "匹配状态"])
        spec_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        spec_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        spec_table.setWordWrap(True)
        spec_table.verticalHeader().setVisible(False)
        spec_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        spec_table.horizontalHeader().setStretchLastSection(True)
        widths = [280, 76, 76, 92, 76, 84, 92, 88, 104, 210]
        for col, width in enumerate(widths):
            spec_table.setColumnWidth(col, width)
        rows = compare.get("rows") or []
        spec_table.setRowCount(len(rows))
        spec_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        spec_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        for row_index, row_data in enumerate(rows):
            values = [
                row_data.get("spec_name", ""),
                self._fmt_money(row_data.get("local_sale_price")),
                self._fmt_money(row_data.get("before_price")),
                row_data.get("price_tag") or row_data.get("price_tag_type") or "--",
                self._fmt_money(row_data.get("coupon_amount")),
                self._fmt_money(row_data.get("new_customer_discount")),
                row_data.get("store_full_reduction_text") or "--",
                self._fmt_money(row_data.get("local_final_price")),
                self._fmt_money(row_data.get("final_receipt")),
                row_data.get("status", ""),
            ]
            for col, value in enumerate(values):
                table_item = QTableWidgetItem(str(value))
                table_item.setTextAlignment(Qt.AlignVCenter | (Qt.AlignLeft if col in (0, 9) else Qt.AlignCenter))
                if col in (0, 9):
                    table_item.setFlags(table_item.flags() | Qt.ItemIsEnabled)
                    table_item.setToolTip(str(value))
                if col == 9:
                    table_item.setForeground(QColor("#27ae60" if value == "匹配" else "#e74c3c"))
                spec_table.setItem(row_index, col, table_item)
        spec_table.resizeRowsToContents()
        header_height = spec_table.horizontalHeader().sizeHint().height()
        rows_height = sum(spec_table.rowHeight(i) for i in range(spec_table.rowCount()))
        spec_table.setFixedHeight(header_height + rows_height + 8)
        card_layout.addWidget(spec_table)
        if insert_at is None:
            insert_at = max(0, self.price_scroll_layout.count() - 1)
        self.price_scroll_layout.insertWidget(insert_at, card)
        self.price_result_cards[product_id] = card
        self.price_result_items[product_id] = {"item": item, "local_product": local_product}

    def _product_discount_target(self, item):
        coupon_values = []
        new_customer_values = []
        store_full_reduction_values = []
        for spec in item.get("specs") or []:
            coupon = self._parse_price_value(spec.get("coupon_amount"))
            new_customer = self._parse_price_value(spec.get("new_customer_discount"))
            store_full_reduction = self._parse_price_value(spec.get("store_full_reduction_amount"))
            if coupon is not None:
                coupon_values.append(coupon)
            if new_customer is not None:
                new_customer_values.append(new_customer)
            if store_full_reduction is not None:
                store_full_reduction_values.append(store_full_reduction)
        return {
            "coupon_amount": max(coupon_values) if coupon_values else 0.0,
            "new_customer_discount": max(new_customer_values) if new_customer_values else 0.0,
            "store_full_reduction_amount": max(store_full_reduction_values) if store_full_reduction_values else 0.0,
        }

    def _product_tag_target(self, item):
        has_limited_time = False
        has_marketing = False
        for spec in item.get("specs") or []:
            flags = self._price_tag_flags(spec.get("price_tag_type"))
            has_limited_time = has_limited_time or bool(flags["is_limited_time"])
            has_marketing = has_marketing or bool(flags["is_marketing"])
        return {
            "is_limited_time": 1 if has_limited_time else 0,
            "is_marketing": 1 if has_marketing else 0,
        }

    def _build_price_compare_rows(self, item, local_product):
        product_db_id = local_product.get("id")
        product_rows = self.db.safe_fetchall(
            "SELECT title, coupon_amount, new_customer_discount, is_limited_time, is_marketing FROM products WHERE id=?",
            (product_db_id,),
        )
        product_state = product_rows[0] if product_rows else ("", 0, 0, 0, 0)
        local_title, local_coupon, local_new_customer, local_limited, local_marketing = product_state
        local_specs = self._local_specs_by_name_for_product(product_db_id)
        target_discount = self._product_discount_target(item)
        target_tag = self._product_tag_target(item)
        rows = []
        can_sync = True
        all_matched = True
        issues = []

        for spec in item.get("specs") or []:
            spec_name = str(spec.get("spec_name") or "").strip()
            local_spec = local_specs.get(spec_name) or local_specs.get(self._normalize_spec_name_for_match(spec_name))
            before_price = self._parse_price_value(spec.get("before_price"))
            final_receipt = self._parse_price_value(spec.get("final_receipt"))
            store_full_reduction_amount = self._parse_price_value(spec.get("store_full_reduction_amount")) or 0.0
            store_full_reduction_threshold = self._parse_price_value(spec.get("store_full_reduction_threshold"))
            effective_store_full_reduction = store_full_reduction_amount
            if store_full_reduction_threshold is not None and before_price is not None and before_price < store_full_reduction_threshold:
                effective_store_full_reduction = 0.0
            status_parts = []
            row_can_sync = True
            row_matched = True
            local_sale_price = local_spec.get("sale_price") if local_spec else None
            local_final_price = None
            if local_sale_price is not None:
                local_final_price = local_sale_price - max(
                    float(local_coupon or 0),
                    float(local_new_customer or 0),
                    float(effective_store_full_reduction or 0),
                )

            if not local_spec:
                status_parts.append("规格不匹配")
                row_can_sync = False
                row_matched = False
            if before_price is None:
                status_parts.append("缺少券前价")
                row_can_sync = False
                row_matched = False
            elif local_sale_price is not None and not self._money_equal(local_sale_price, before_price):
                status_parts.append("券前价不匹配")
                row_matched = False
            if final_receipt is None:
                status_parts.append("缺少实收")
                row_can_sync = False
                row_matched = False
            elif local_final_price is not None and not self._money_equal(local_final_price, final_receipt):
                status_parts.append("券后价不匹配")
                row_matched = False

            if not self._money_equal(local_coupon or 0, target_discount["coupon_amount"]):
                status_parts.append("优惠券不匹配")
                row_matched = False
            if not self._money_equal(local_new_customer or 0, target_discount["new_customer_discount"]):
                status_parts.append("新客立减不匹配")
                row_matched = False
            if int(local_limited or 0) != target_tag["is_limited_time"]:
                status_parts.append("限时限量购不匹配")
                row_matched = False
            if int(local_marketing or 0) != target_tag["is_marketing"]:
                status_parts.append("营销活动不匹配")
                row_matched = False

            if not row_can_sync:
                can_sync = False
            if not row_matched:
                all_matched = False
                issues.extend(status_parts)

            rows.append({
                "product_db_id": product_db_id,
                "product_id": item.get("product_id", ""),
                "product_title": item.get("title") or local_title or local_product.get("title", ""),
                "spec_name": spec_name,
                "local_spec_id": local_spec.get("id") if local_spec else None,
                "local_sale_price": local_sale_price,
                "before_price": before_price,
                "price_tag": spec.get("price_tag") or ("拼单价" if spec.get("price_tag_type") == "裸价" else ""),
                "price_tag_type": spec.get("price_tag_type", ""),
                "coupon_amount": target_discount["coupon_amount"],
                "new_customer_discount": target_discount["new_customer_discount"],
                "store_full_reduction_amount": store_full_reduction_amount,
                "store_full_reduction_threshold": store_full_reduction_threshold,
                "store_full_reduction_text": spec.get("store_full_reduction_text") or "",
                "local_final_price": local_final_price,
                "final_receipt": final_receipt,
                "status": "匹配" if row_matched else "不匹配：" + "、".join(status_parts),
                "row_can_sync": row_can_sync,
                "raw_item": item,
            })

        return {
            "rows": rows,
            "can_sync": can_sync and not all_matched,
            "all_matched": all_matched,
            "issues": issues,
            "target_discount": target_discount,
            "target_tag": target_tag,
        }

    def _price_compare_categories(self, compare):
        rows = compare.get("rows") or []
        issues = []
        for row in rows:
            status = str(row.get("status") or "")
            if status and status != "匹配":
                issues.append(status)
        issue_text = "、".join(issues + [str(item) for item in (compare.get("issues") or [])])
        return {
            "spec": (not rows) or ("规格不匹配" in issue_text),
            "price": any(keyword in issue_text for keyword in ("券前价不匹配", "券后价不匹配", "缺少券前价", "缺少实收")),
            "marketing": any(keyword in issue_text for keyword in ("优惠券不匹配", "新客立减不匹配", "限时限量购不匹配", "营销活动不匹配")),
        }

    def _price_unmatched_reason_labels(self, categories, product_unmatched=False):
        reasons = []
        if product_unmatched:
            reasons.append("商品ID未匹配")
        if categories.get("spec"):
            reasons.append("规格未匹配")
        if categories.get("price"):
            reasons.append("价格未匹配")
        if categories.get("marketing"):
            reasons.append("活动/营销工具未匹配")
        return reasons or ["未匹配"]

    def _is_spec_unmatched_record(self, record):
        return any("规格" in str(reason) for reason in (record or {}).get("reasons") or [])

    def _spec_unmatched_records(self):
        return [
            record for record in (self.price_unmatched_records or {}).values()
            if self._is_spec_unmatched_record(record)
        ]

    def _provided_spec_unmatched_records(self):
        provider = getattr(self, "unmatched_records_provider", None)
        records = provider() if callable(provider) else []
        if isinstance(records, dict):
            records = records.values()
        return [record for record in (records or []) if self._is_spec_unmatched_record(record)]

    def _refresh_unmatched_spec_ids_panel(self):
        if not hasattr(self, "unmatched_specs_widget"):
            return
        records = self._provided_spec_unmatched_records()
        ids = []
        for record in records:
            product_id = str((record or {}).get("product_id") or "").strip()
            if product_id and product_id not in ids:
                ids.append(product_id)
        self._code_unmatched_spec_ids = ids
        visible = self.mode == "code" and bool(ids)
        self.unmatched_specs_widget.setVisible(visible)
        if visible:
            self.lbl_unmatched_specs.setText("已记录未匹配规格ID：" + "、".join(ids))

    def _sort_records_by_time(self, records):
        def key(record):
            text = str(record.get("time", "") if isinstance(record, dict) else "")
            try:
                hour, minute = text.split(":", 1)
                return (int(hour), int(minute))
            except Exception:
                return (99, 99)
        return sorted(records or [], key=key)

    def _record_product_operation(self, product_db_id, text, metric, old="", new="", change_type="pdd_sync"):
        main_app = getattr(self.owner, "main_app", None) or self.owner
        if main_app is not None and hasattr(main_app, "record_product_operation"):
            main_app.record_product_operation(product_db_id, text, metric, old, new, change_type)
            return
        now = datetime.now()
        time_str = now.strftime("%H:%M")
        rows = self.db.safe_fetchall(
            "SELECT records_json FROM records WHERE product_id=? AND year=? AND month=? AND day=?",
            (product_db_id, now.year, now.month, now.day),
        )
        records = []
        if rows and rows[0][0]:
            try:
                records = json.loads(rows[0][0])
            except Exception:
                records = []
        if not isinstance(records, list):
            records = []
        records.append({
            "time": time_str,
            "text": text,
            "changes": [{
                "time": time_str,
                "metric": metric,
                "old": "" if old is None else str(old),
                "new": "" if new is None else str(new),
                "text": text,
                "type": change_type,
            }],
        })
        records = self._sort_records_by_time(records)
        self.db.safe_execute(
            "INSERT OR REPLACE INTO records (product_id, year, month, day, records_json) VALUES (?, ?, ?, ?, ?)",
            (product_db_id, now.year, now.month, now.day, json.dumps(records, ensure_ascii=False)),
        )

    def _execute_required(self, query, params=()):
        cursor = self.db.safe_execute(query, params)
        if cursor is None:
            raise RuntimeError(f"数据库写入失败：{query[:80]}")
        return cursor

    def _autosave_archive_after_link_write(self):
        try:
            self.db.conn.commit()
        except Exception as e:
            print(f"link write commit failed: {e}")

        candidates = [self.owner, self.parent(), getattr(self.owner, "main_app", None)]
        for candidate in candidates:
            if candidate is not None and hasattr(candidate, "autosave_current_archive"):
                ok, result = candidate.autosave_current_archive()
                if not ok:
                    print(f"link write archive autosave failed: {result}")
                return ok
        return True

    def _refresh_store_margin_after_link_write(self, product_db_id=None, created=False):
        targets = []

        def add_target(obj):
            if obj is not None and obj not in targets:
                targets.append(obj)

        owner = self.owner or self.parent()
        add_target(owner)
        add_target(self.parent())
        add_target(getattr(owner, "main_app", None))

        store_id = self.current_store_id()
        for target in list(targets):
            store_dialogs = getattr(target, "store_margin_dialogs", None)
            if isinstance(store_dialogs, dict) and store_id in store_dialogs:
                add_target(store_dialogs.get(store_id))

        for target in targets:
            schedule = getattr(target, "schedule_pdd_link_refresh", None)
            if callable(schedule):
                schedule(store_id, product_db_id, created=created)
                continue
            for method_name in ("load_specs", "load_products", "force_refresh_frozen_table"):
                method = getattr(target, method_name, None)
                if callable(method):
                    method()
                    break

    def _refresh_after_price_sync(self, product_id=None, product_db_id=None):
        if product_db_id is None and product_id is not None:
            rows = (self.price_sync_rows.get(str(product_id)) or {}).get("rows") or []
            product_db_id = rows[0].get("product_db_id") if rows else None
        if callable(append_event):
            append_event(f"pdd_price_sync:refresh product_id={product_id} product_db_id={product_db_id}")
        if isinstance(product_db_id, (list, tuple, set)):
            for db_id in product_db_id:
                self._refresh_store_margin_after_link_write(db_id)
        else:
            self._refresh_store_margin_after_link_write(product_db_id)
        if product_id is None:
            self.refresh_all_price_management_matches()
        else:
            self.refresh_price_management_product(product_id)

    def _refresh_code_table_match_status(self):
        if self.mode != "code" or not self.table.isVisible():
            return
        store_id = self.current_store_id()
        product_id = str(self.last_product_id or "").strip()
        specs = self._code_table_specs_from_view()
        self.last_specs = specs

        local_products = self._local_products_for_store(store_id) if store_id else {}
        product_match = local_products.get(product_id) if product_id else None
        matched = bool(product_match)
        self.missing_ids = [] if matched or not product_id else [product_id]
        software_title = product_match.get("title", "") if product_match else ""
        local_specs = self._local_specs_for_product(product_match.get("id")) if product_match else {}

        self.lbl_software_product_title.setText(f"软件标题: {software_title or '--'}")
        matched_specs = 0
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 2)
            spec_code = str(item.text() if item else "").strip()
            normalized_spec_code = self._normalize_spec_code(spec_code)
            spec_matched = bool(
                spec_code and (
                    spec_code in local_specs or normalized_spec_code in local_specs
                )
            )
            if spec_matched:
                matched_specs += 1
            local_spec_name = ""
            if spec_code:
                local_spec_name = local_specs.get(spec_code, "") or local_specs.get(normalized_spec_code, "")
            local_spec_display = local_spec_name or (
                "已匹配规格" if spec_matched else ("未匹配规格" if matched and spec_code else "")
            )
            self._set_code_table_item(
                row,
                4,
                "已匹配" if matched else "未匹配",
                editable=False,
                color="#27ae60" if matched else "#e74c3c",
            )
            self._set_code_table_item(
                row,
                5,
                local_spec_display,
                editable=False,
                color="#27ae60" if spec_matched else "#e67e22",
            )

        if hasattr(self, "btn_open_current_spec"):
            self.btn_open_current_spec.setEnabled(bool(product_match))
        self.lbl_summary.setText(
            f"当前编码窗口：商品ID {product_id or '未识别'}，规格 {len(specs)} 条，"
            f"图片 {len(self.last_product_images or [])} 张，"
            f"{'已匹配本店商品' if matched else '未匹配本店商品'}，"
            f"规格匹配 {matched_specs}/{len(specs)}。"
        )

    def save_current_link_to_store(self):
        store_id = self.current_store_id()
        if not store_id:
            QMessageBox.warning(self, "提示", "请先选择要写入的软件店铺。")
            return
        product_id = str(self.last_product_id or "").strip()
        title = str(self.last_title or "").strip()
        self.table.clearFocus()
        self._sync_code_table_specs_to_state()
        specs = self._dedupe_specs(self.last_specs or [])
        if not product_id or not title:
            QMessageBox.warning(self, "提示", "当前抓取结果缺少商品ID或标题，请先重新抓取添加编码界面。")
            return
        if not specs:
            QMessageBox.warning(self, "提示", "当前抓取结果没有规格信息，请先重新抓取添加编码界面。")
            return

        try:
            rows = self.db.safe_fetchall(
                "SELECT id, title FROM products WHERE store_id=? AND name=? AND COALESCE(is_archived, 0)=0",
                (store_id, product_id),
            )
            existing = rows[0] if rows else None
            is_update = bool(existing)
            product_db_id = existing[0] if existing else None
            old_title = existing[1] if existing else ""
            include_product_image = self.chk_include_product_image.isChecked() if hasattr(self, "chk_include_product_image") else True
            product_image_data = (
                self.last_product_image_data
                or self._download_image_bytes((self.last_product_images or [""])[0])
            ) if include_product_image else None
            if include_product_image and not product_image_data:
                raise RuntimeError(f"主图下载失败：{self.last_image_download_error or '未抓取到主图'}")

            if is_update:
                if include_product_image and product_image_data:
                    self._execute_required(
                        "UPDATE products SET title=?, image_data=? WHERE id=?",
                        (title, product_image_data, product_db_id),
                    )
                else:
                    self._execute_required(
                        "UPDATE products SET title=? WHERE id=?",
                        (title, product_db_id),
                    )
            else:
                result = self.db.safe_fetchall("SELECT MAX(sort_order) FROM products WHERE store_id=?", (store_id,))
                max_order = result[0][0] if result and result[0][0] is not None else 0
                self._execute_required(
                    "INSERT INTO products (store_id, name, title, image_data, sort_order) VALUES (?, ?, ?, ?, ?)",
                    (store_id, product_id, title, product_image_data, max_order + 1),
                )
                product_db_id = self.db.safe_fetchall("SELECT last_insert_rowid()")[0][0]

            valid_specs = [
                spec for spec in specs
                if str(spec.get("spec_info", "") or "").strip() or str(spec.get("spec_code", "") or "").strip()
            ]
            default_weight = (100.0 / len(valid_specs)) if valid_specs else 0.0
            if is_update:
                old_rows = self.db.safe_fetchall(
                    "SELECT id, spec_code, spec_name, sale_price, weight_percent FROM product_specs WHERE product_id=?",
                    (product_db_id,),
                )
                old_by_code = {}
                old_by_name = {}
                old_weight_total = 0.0
                for row in old_rows:
                    spec_id, spec_code, spec_name, sale_price, weight_percent = row
                    payload = {
                        "id": spec_id,
                        "spec_code": str(spec_code or "").strip(),
                        "spec_name": str(spec_name or "").strip(),
                        "sale_price": float(sale_price or 0),
                        "weight_percent": float(weight_percent or 0),
                    }
                    old_weight_total += payload["weight_percent"]
                    code_key = self._normalize_spec_code(payload["spec_code"])
                    name_key = self._normalize_spec_name_for_match(payload["spec_name"])
                    if code_key:
                        old_by_code.setdefault(code_key, payload)
                    if name_key:
                        old_by_name.setdefault(name_key, payload)

                kept_ids = []
                use_default_weight = old_weight_total <= 0 and bool(valid_specs)
                for spec in valid_specs:
                    spec_name = str(spec.get("spec_info", "") or "").strip()
                    spec_code = str(spec.get("spec_code", "") or "").strip()
                    sale_price = self._parse_price_value(spec.get("price", ""))
                    old_spec = old_by_code.get(self._normalize_spec_code(spec_code)) or old_by_name.get(self._normalize_spec_name_for_match(spec_name))
                    spec_image_data = self._download_image_bytes(spec.get("image", ""))
                    if old_spec:
                        kept_ids.append(old_spec["id"])
                        weight_percent = default_weight if use_default_weight else old_spec["weight_percent"]
                        if spec_image_data:
                            self._execute_required(
                                "UPDATE product_specs SET spec_name=?, spec_code=?, sale_price=?, weight_percent=?, spec_image_data=? WHERE id=?",
                                (spec_name or spec_code, spec_code, sale_price, weight_percent, spec_image_data, old_spec["id"]),
                            )
                        else:
                            self._execute_required(
                                "UPDATE product_specs SET spec_name=?, spec_code=?, sale_price=?, weight_percent=? WHERE id=?",
                                (spec_name or spec_code, spec_code, sale_price, weight_percent, old_spec["id"]),
                            )
                    else:
                        self._execute_required(
                            "INSERT INTO product_specs (product_id, spec_name, spec_code, sale_price, weight_percent, spec_image_data) VALUES (?, ?, ?, ?, ?, ?)",
                            (product_db_id, spec_name or spec_code, spec_code, sale_price, 0, spec_image_data),
                        )
                        kept_ids.append(self.db.safe_fetchall("SELECT last_insert_rowid()")[0][0])
                for spec_id, *_rest in old_rows:
                    if spec_id not in kept_ids:
                        self._execute_required("DELETE FROM product_specs WHERE id=?", (spec_id,))
            else:
                for spec in valid_specs:
                    spec_name = str(spec.get("spec_info", "") or "").strip()
                    spec_code = str(spec.get("spec_code", "") or "").strip()
                    sale_price = self._parse_price_value(spec.get("price", ""))
                    spec_image_data = self._download_image_bytes(spec.get("image", ""))
                    self._execute_required(
                        "INSERT INTO product_specs (product_id, spec_name, spec_code, sale_price, weight_percent, spec_image_data) VALUES (?, ?, ?, ?, ?, ?)",
                    (product_db_id, spec_name or spec_code, spec_code, sale_price, default_weight, spec_image_data),
                    )

            action_text = "覆盖链接" if is_update else "新建链接"
            image_text = "包含主图" if include_product_image else "不包含主图"
            self._record_product_operation(
                product_db_id,
                f"拼多多抓取{action_text}：商品ID {product_id}，标题：{title}，规格 {len(specs)} 个，{image_text}",
                metric=action_text,
                old=old_title if is_update else "",
                new=title,
                change_type="pdd_link_sync",
            )
            self._autosave_archive_after_link_write()

            if hasattr(self, "btn_open_current_spec"):
                self.btn_open_current_spec.setEnabled(True)
            self.lbl_summary.setText(
                f"已按当前界面内容{'覆盖' if is_update else '创建'}本店铺链接：{product_id}，规格 {len(specs)} 个，{image_text}。"
            )
            self._refresh_code_table_match_status()
            QMessageBox.information(self, "写入完成", f"已{'覆盖' if is_update else '创建'}本店铺链接：{product_id}\n规格：{len(specs)} 个")
            self.link_saved.emit(product_id)
            QTimer.singleShot(0, lambda pid=product_db_id, created=not is_update: self._refresh_store_margin_after_link_write(pid, created))
        except Exception as e:
            QMessageBox.warning(self, "写入失败", f"创建/覆盖链接失败：{e}")

    def overwrite_current_link_without_price(self):
        store_id = self.current_store_id()
        if not store_id:
            QMessageBox.warning(self, "提示", "请先选择要写入的软件店铺。")
            return
        product_id = str(self.last_product_id or "").strip()
        title = str(self.last_title or "").strip()
        self.table.clearFocus()
        self._sync_code_table_specs_to_state()
        specs = self._dedupe_specs(self.last_specs or [])
        if not product_id or not title:
            QMessageBox.warning(self, "提示", "当前抓取结果缺少商品ID或标题，请先重新抓取添加编码界面。")
            return
        if not specs:
            QMessageBox.warning(self, "提示", "当前抓取结果没有规格信息，请先重新抓取添加编码界面。")
            return

        rows = self.db.safe_fetchall(
            "SELECT id, title FROM products WHERE store_id=? AND name=? AND COALESCE(is_archived, 0)=0",
            (store_id, product_id),
        )
        if not rows:
            QMessageBox.information(self, "未匹配本店铺", "软件里还没有这个商品，请先使用“创建/覆盖到本店铺”。")
            return

        product_db_id, old_title = rows[0]
        try:
            include_product_image = self.chk_include_product_image.isChecked() if hasattr(self, "chk_include_product_image") else True
            product_image_data = (
                self.last_product_image_data
                or self._download_image_bytes((self.last_product_images or [""])[0])
            ) if include_product_image else None
            if include_product_image and not product_image_data:
                raise RuntimeError(f"主图下载失败：{self.last_image_download_error or '未抓取到主图'}")
            if include_product_image and product_image_data:
                self.db.safe_execute(
                    "UPDATE products SET title=?, image_data=? WHERE id=?",
                    (title, product_image_data, product_db_id),
                )
            else:
                self.db.safe_execute("UPDATE products SET title=? WHERE id=?", (title, product_db_id))

            old_rows = self.db.safe_fetchall(
                "SELECT id, spec_code, spec_name, sale_price, weight_percent FROM product_specs WHERE product_id=?",
                (product_db_id,),
            )
            old_by_code = {}
            old_by_name = {}
            for row in old_rows:
                spec_id, spec_code, spec_name, sale_price, weight_percent = row
                payload = {
                    "id": spec_id,
                    "spec_code": str(spec_code or "").strip(),
                    "spec_name": str(spec_name or "").strip(),
                    "sale_price": float(sale_price or 0),
                    "weight_percent": float(weight_percent or 0),
                }
                code_key = self._normalize_spec_code(payload["spec_code"])
                name_key = self._normalize_spec_name_for_match(payload["spec_name"])
                if code_key:
                    old_by_code.setdefault(code_key, payload)
                if name_key:
                    old_by_name.setdefault(name_key, payload)

            kept_ids = []
            for spec in specs:
                spec_name = str(spec.get("spec_info", "") or "").strip()
                spec_code = str(spec.get("spec_code", "") or "").strip()
                if not spec_name and not spec_code:
                    continue
                old_spec = old_by_code.get(self._normalize_spec_code(spec_code)) or old_by_name.get(self._normalize_spec_name_for_match(spec_name))
                spec_image_data = self._download_image_bytes(spec.get("image", ""))
                if old_spec:
                    kept_ids.append(old_spec["id"])
                    if spec_image_data:
                        self.db.safe_execute(
                            "UPDATE product_specs SET spec_name=?, spec_code=?, spec_image_data=? WHERE id=?",
                            (spec_name or spec_code, spec_code, spec_image_data, old_spec["id"]),
                        )
                    else:
                        self.db.safe_execute(
                            "UPDATE product_specs SET spec_name=?, spec_code=? WHERE id=?",
                            (spec_name or spec_code, spec_code, old_spec["id"]),
                        )
                else:
                    self.db.safe_execute(
                        "INSERT INTO product_specs (product_id, spec_name, spec_code, sale_price, weight_percent, spec_image_data) VALUES (?, ?, ?, ?, ?, ?)",
                        (product_db_id, spec_name or spec_code, spec_code, 0, 0, spec_image_data),
                    )
                    kept_ids.append(self.db.safe_fetchall("SELECT last_insert_rowid()")[0][0])

            old_ids = [row[0] for row in old_rows]
            for spec_id in old_ids:
                if spec_id not in kept_ids:
                    self.db.safe_execute("DELETE FROM product_specs WHERE id=?", (spec_id,))

            image_text = "包含主图" if include_product_image else "不包含主图"
            self._record_product_operation(
                product_db_id,
                f"拼多多覆盖非价格信息：商品ID {product_id}，标题：{title}，规格 {len(kept_ids)} 个，价格保持不变，{image_text}",
                metric="覆盖非价格信息",
                old=old_title,
                new=title,
                change_type="pdd_link_non_price_sync",
            )
            self.lbl_summary.setText(
                f"已按当前界面内容覆盖商品 {product_id} 的非价格信息，规格 {len(kept_ids)} 个，{image_text}。"
            )
            self._refresh_code_table_match_status()
            QMessageBox.information(self, "覆盖完成", f"已覆盖商品 {product_id} 的非价格信息。\n规格：{len(kept_ids)} 个\n{image_text}\n已有规格价格保持不变；新增规格价格为 0。")
            self.link_saved.emit(product_id)
            QTimer.singleShot(0, lambda pid=product_db_id: self._refresh_store_margin_after_link_write(pid))
        except Exception as e:
            QMessageBox.warning(self, "覆盖失败", f"覆盖除价格之外的信息失败：{e}")

    def open_current_code_product_spec_dialog(self):
        store_id = self.current_store_id()
        product_id = str(self.last_product_id or "").strip()
        if not store_id or not product_id:
            QMessageBox.information(self, "提示", "请先抓取并匹配当前链接。")
            return
        local_product = self._local_products_for_store(store_id).get(product_id)
        if not local_product:
            QMessageBox.information(self, "提示", "当前链接还未匹配到本店铺商品，请先创建/覆盖到本店铺。")
            return
        parent = self.owner or self.parent()
        try:
            if parent and hasattr(parent, "open_spec_dialog"):
                parent.open_spec_dialog(local_product.get("id"), product_id, local_product.get("title", "") or self.last_title)
            elif parent and hasattr(parent, "main_app") and hasattr(parent.main_app, "open_product_spec_dialog"):
                parent.main_app.open_product_spec_dialog(self.db, local_product.get("id"), product_id, local_product.get("title", "") or self.last_title, parent)
        except Exception as e:
            QMessageBox.warning(self, "打开失败", f"打开规格与毛利管理失败：{e}")

    def scan_current_page(self):
        store_id = self._activate_browser_store_context()
        if not store_id:
            QMessageBox.warning(self, "提示", "请先选择要匹配的软件店铺。")
            return

        self._set_code_table_mode()
        self._refresh_unmatched_spec_ids_panel()
        self._set_scan_controls_enabled(False)
        QApplication.processEvents()
        try:
            info = self.monitor.inspect()
        except Exception as e:
            QMessageBox.warning(self, "拼多多链接抓取", f"读取浏览器页面失败：{e}")
            return
        finally:
            self._set_scan_controls_enabled(True)

        self.last_debug_info = info
        detail = info.get("current_code_detail") or {}
        product_id = str(detail.get("product_id") or "").strip()
        title = str(detail.get("title") or "").strip()
        product_images = [str(x).strip() for x in (detail.get("product_images") or []) if str(x).strip()]
        specs = self._dedupe_specs(detail.get("specs") or [])
        self.last_product_id = product_id
        self.last_title = title
        self.last_product_images = product_images
        self.last_specs = specs

        local_products = self._local_products_for_store(store_id)
        product_match = local_products.get(product_id) if product_id else None
        matched = bool(product_match)
        self.missing_ids = [] if matched or not product_id else [product_id]
        software_title = product_match.get("title", "") if product_match else ""
        local_specs = self._local_specs_for_product(product_match.get("id")) if product_match else {}

        self.lbl_pdd_product_title.setText(f"标题: {title or '--'}")
        self.lbl_software_product_title.setText(f"软件标题: {software_title or '--'}")
        self._set_product_image_previews(product_images)
        self.product_images_widget.setToolTip("\n".join(product_images))

        self.table.setRowCount(len(specs))
        for row, spec in enumerate(specs):
            spec_code = str(spec.get("spec_code", "") or "").strip()
            normalized_spec_code = self._normalize_spec_code(spec_code)
            local_spec_name = local_specs.get(spec_code, "") or local_specs.get(normalized_spec_code, "") if spec_code else ""
            spec_matched = bool(spec_code and (spec_code in local_specs or normalized_spec_code in local_specs))
            local_spec_display = local_spec_name or ("已匹配规格" if spec_matched else ("未匹配规格" if matched and spec_code else ""))
            values = [
                spec.get("image", ""),
                spec.get("spec_info", ""),
                spec_code,
                spec.get("price", ""),
                "已匹配" if matched else "未匹配",
                local_spec_display,
                spec.get("raw_text", ""),
            ]
            for col, value in enumerate(values):
                if col == 0:
                    self._set_image_preview_cell(row, col, value)
                    continue
                editable = col in (1, 2, 3, 6)
                color = None
                if col == 4:
                    color = "#27ae60" if matched else "#e74c3c"
                if col == 5:
                    color = "#27ae60" if spec_matched else "#e67e22"
                self._set_code_table_item(row, col, value, editable=editable, color=color)

        matched_specs = sum(1 for spec in specs if str(spec.get("spec_code", "") or "").strip() in local_specs)
        self.lbl_summary.setText(
            f"当前编码窗口：商品ID {product_id or '未识别'}，规格 {len(specs)} 条，"
            f"图片 {len(product_images)} 张，{'已匹配本店商品' if matched else '未匹配本店商品'}，"
            f"规格匹配 {matched_specs}/{len(specs)}。"
        )
        self.lbl_summary.setToolTip(json.dumps(info, ensure_ascii=False, indent=2))
        self._refresh_unmatched_spec_ids_panel()
        if hasattr(self, "btn_open_current_spec"):
            self.btn_open_current_spec.setEnabled(bool(product_match))

        if not product_id and not specs:
            QMessageBox.information(
                self,
                "未识别到添加编码界面信息",
                "当前页面没有读取到商品ID或规格信息。请确认已经手动打开“添加/编辑商品编码”窗口；如果仍为空，请提供该窗口里标题、商品ID、规格行外层容器的 class/id 名称。",
            )

    def scan_price_management(self):
        store_id = self._activate_browser_store_context()
        if not store_id:
            QMessageBox.warning(self, "提示", "请先选择要匹配的软件店铺。")
            return

        self._set_price_table_mode()
        self._set_scan_controls_enabled(False)
        self.lbl_summary.setText("正在滚动并抓取当前价格管理页面，请稍候...")
        QApplication.processEvents()
        try:
            info = self.monitor.inspect_price_management()
        except Exception as e:
            QMessageBox.warning(self, "拼多多链接抓取", f"读取价格管理页面失败：{e}")
            return
        finally:
            self._set_scan_controls_enabled(True)

        self.last_debug_info = info
        self.last_price_management_info = info
        self.price_sync_rows = {}
        self.price_unmatched_spec_product_ids = []
        self.price_unmatched_records = {}
        self.price_current_page_matched_product_ids = []
        if not info.get("is_price_management"):
            self.lbl_summary.setText(info.get("status") or "当前页面不是价格管理界面。")
            self.lbl_summary.setToolTip(json.dumps(info, ensure_ascii=False, indent=2))
            QMessageBox.information(self, "不是价格管理界面", "当前页面未识别到价格管理、券前价、商家出资优惠或单件预估实收。请先进入拼多多价格管理页面。")
            return

        local_products = self._local_products_for_store(store_id)
        items = info.get("items") or []
        matched_items = []
        product_unmatched_count = 0
        full_matched_count = 0
        total_unmatched_count = 0
        spec_unmatched_count = 0
        price_unmatched_count = 0
        marketing_unmatched_count = 0
        total_specs = 0
        matched_specs = 0
        blocked_specs = 0
        for item in items:
            product_id = str(item.get("product_id") or "").strip()
            local_product = local_products.get(product_id)
            if not local_product:
                product_unmatched_count += 1
                total_unmatched_count += 1
                if product_id:
                    self.price_unmatched_records[product_id] = {
                        "product_id": product_id,
                        "title": item.get("title", ""),
                        "reasons": self._price_unmatched_reason_labels({}, product_unmatched=True),
                    }
                continue
            compare = self._build_price_compare_rows(item, local_product)
            if not compare["rows"]:
                self.price_unmatched_spec_product_ids.append(product_id)
                total_unmatched_count += 1
                spec_unmatched_count += 1
                self.price_unmatched_records[product_id] = {
                    "product_id": product_id,
                    "title": item.get("title") or local_product.get("title", ""),
                    "reasons": self._price_unmatched_reason_labels({"spec": True}),
                }
                continue
            if any(not row.get("local_spec_id") for row in compare["rows"]):
                self.price_unmatched_spec_product_ids.append(product_id)
            categories = self._price_compare_categories(compare)
            if compare.get("all_matched"):
                full_matched_count += 1
                self.price_current_page_matched_product_ids.append(product_id)
            else:
                total_unmatched_count += 1
                if categories["spec"]:
                    spec_unmatched_count += 1
                if categories["price"]:
                    price_unmatched_count += 1
                if categories["marketing"]:
                    marketing_unmatched_count += 1
                self.price_unmatched_records[product_id] = {
                    "product_id": product_id,
                    "title": item.get("title") or local_product.get("title", ""),
                    "reasons": self._price_unmatched_reason_labels(categories),
                }
            matched_items.append((item, local_product, compare))
            total_specs += len(item.get("specs") or [])
            matched_specs += len(compare["rows"])
            blocked_specs += sum(1 for row in compare["rows"] if not row.get("row_can_sync"))

        display_rows = []
        for item, local_product, compare in matched_items:
            product_id = str(item.get("product_id") or "")
            self.price_sync_rows[product_id] = compare
            self.price_result_items[product_id] = {"item": item, "local_product": local_product}
            for row in compare["rows"]:
                display_rows.append(row)
            self._add_price_result_card(item, local_product, compare)

        self.price_unmatched_spec_product_ids = sorted(set(self.price_unmatched_spec_product_ids), key=self.price_unmatched_spec_product_ids.index)
        if hasattr(self, "btn_copy_unmatched_specs"):
            self.btn_copy_unmatched_specs.setEnabled(bool(self._spec_unmatched_records()))
        if hasattr(self, "btn_refresh_price_match"):
            self.btn_refresh_price_match.setEnabled(bool(self.price_result_items))
        if hasattr(self, "btn_sync_all_price"):
            self.btn_sync_all_price.setEnabled(any(compare.get("can_sync") for compare in (self.price_sync_rows or {}).values()))
        self.lbl_summary.setStyleSheet("font-size: 20px; font-weight: bold; padding: 6px 0;")
        self.lbl_summary.setText(
            _price_match_summary_html(
                len(items),
                full_matched_count,
                total_unmatched_count,
                spec_unmatched_count,
                price_unmatched_count,
                marketing_unmatched_count,
                product_unmatched_count,
            )
        )
        self.lbl_summary.setToolTip(json.dumps(info, ensure_ascii=False, indent=2))
        if self.unmatched_task_window is not None:
            self.unmatched_task_window.remove_product_ids(self.price_current_page_matched_product_ids)
            self.unmatched_task_window.upsert_records(self._spec_unmatched_records())
        if self.unmatched_code_dialog is not None:
            self.unmatched_code_dialog._refresh_unmatched_spec_ids_panel()

        if not display_rows:
            QMessageBox.information(self, "没有可展示的匹配链接", "已抓到价格管理页面，但没有商品ID能匹配当前软件店铺，或没有识别到规格价格。")

    def sync_price_management_product(self, product_id, show_message=True, refresh_after=True):
        product_id = str(product_id or "").strip()
        compare = self.price_sync_rows.get(product_id)
        if not compare:
            message = "当前链接没有可同步的价格管理抓取结果。"
            if show_message:
                QMessageBox.warning(self, "提示", message)
            return False, message
        if not compare.get("can_sync"):
            message = "该链接已匹配，或存在规格名称不一致/缺少关键价格字段，不能同步。"
            if show_message:
                QMessageBox.information(self, "无需同步", message)
            return False, message

        rows = compare.get("rows") or []
        product_db_id = rows[0].get("product_db_id") if rows else None
        if not product_db_id:
            message = "没有找到软件商品，不能同步。"
            if show_message:
                QMessageBox.warning(self, "提示", message)
            return False, message
        if callable(append_event):
            append_event(f"pdd_price_sync:item:start product_id={product_id} product_db_id={product_db_id}")

        try:
            old_product_rows = self.db.safe_fetchall(
                "SELECT coupon_amount, new_customer_discount, is_limited_time, is_marketing FROM products WHERE id=?",
                (product_db_id,),
            )
            old_coupon, old_new_customer, old_limited, old_marketing = old_product_rows[0] if old_product_rows else (0, 0, 0, 0)
            changes = []
            for row in rows:
                if not row.get("row_can_sync"):
                    continue
                spec_id = row.get("local_spec_id")
                new_price = row.get("before_price")
                if not spec_id or new_price is None:
                    continue
                old_price = row.get("local_sale_price")
                if not self._money_equal(old_price, new_price):
                    changes.append(f"{row.get('spec_name')} 价格 {self._fmt_money(old_price)}→{self._fmt_money(new_price)}")
                self.db.safe_execute(
                    "UPDATE product_specs SET sale_price=? WHERE id=?",
                    (float(new_price), spec_id),
                )

            target_discount = compare.get("target_discount") or {}
            target_tag = compare.get("target_tag") or {}
            coupon_amount = float(target_discount.get("coupon_amount") or 0)
            new_customer_discount = float(target_discount.get("new_customer_discount") or 0)
            is_limited_time = int(target_tag.get("is_limited_time") or 0)
            is_marketing = int(target_tag.get("is_marketing") or 0)
            if not self._money_equal(old_coupon or 0, coupon_amount):
                changes.append(f"优惠券 {self._fmt_money(old_coupon)}→{self._fmt_money(coupon_amount)}")
            if not self._money_equal(old_new_customer or 0, new_customer_discount):
                changes.append(f"新客立减 {self._fmt_money(old_new_customer)}→{self._fmt_money(new_customer_discount)}")
            if int(old_limited or 0) != is_limited_time:
                changes.append(f"限时限量购 {'已报名' if old_limited else '未报名'}→{'已报名' if is_limited_time else '未报名'}")
            if int(old_marketing or 0) != is_marketing:
                changes.append(f"营销活动 {'已报名' if old_marketing else '未报名'}→{'已报名' if is_marketing else '未报名'}")

            self.db.safe_execute(
                "UPDATE products SET coupon_amount=?, new_customer_discount=?, is_limited_time=?, is_marketing=? WHERE id=?",
                (coupon_amount, new_customer_discount, is_limited_time, is_marketing, product_db_id),
            )
            if changes:
                self._record_product_operation(
                    product_db_id,
                    "；".join(changes),
                    metric="改价",
                    old="",
                    new="；".join(changes),
                    change_type="pdd_price_marketing_sync",
                )

            if show_message:
                QMessageBox.information(self, "同步完成", f"已同步商品 {product_id} 的价格和营销信息。")
            if refresh_after:
                QTimer.singleShot(0, lambda pid=product_id, dbid=product_db_id: self._refresh_after_price_sync(pid, dbid))
            if callable(append_event):
                append_event(f"pdd_price_sync:item:done product_id={product_id} product_db_id={product_db_id}")
            return True, f"已同步商品 {product_id}"
        except Exception as e:
            message = f"同步价格和营销失败：{e}"
            if show_message:
                QMessageBox.warning(self, "同步失败", message)
            return False, message

    def sync_all_price_management_products(self):
        syncable_ids = [
            str(product_id)
            for product_id, compare in (self.price_sync_rows or {}).items()
            if compare.get("can_sync")
        ]
        if not syncable_ids:
            self.lbl_summary.setText("当前没有可一键同步的未匹配链接。规格不匹配、商品未匹配、已完全匹配的链接会自动跳过。")
            QMessageBox.information(self, "无需同步", "当前没有可一键同步的未匹配链接。")
            return

        self.btn_sync_all_price.setEnabled(False)
        self._set_scan_controls_enabled(False)
        QApplication.processEvents()
        success = 0
        failed = []
        synced_db_ids = []
        try:
            for product_id in syncable_ids:
                ok, message = self.sync_price_management_product(product_id, show_message=False, refresh_after=False)
                if ok:
                    success += 1
                    rows = (self.price_sync_rows.get(str(product_id)) or {}).get("rows") or []
                    if rows and rows[0].get("product_db_id"):
                        synced_db_ids.append(rows[0].get("product_db_id"))
                else:
                    failed.append(f"{product_id}: {message}")
        finally:
            self._set_scan_controls_enabled(True)
            if hasattr(self, "btn_sync_all_price"):
                self.btn_sync_all_price.setEnabled(any(compare.get("can_sync") for compare in (self.price_sync_rows or {}).values()))

        summary = f"一键同步完成：成功 {success} 个，跳过/失败 {len(failed)} 个。"
        if failed:
            summary += " " + "；".join(failed[:5])
            self.lbl_summary.setToolTip("\n".join(failed))
        self.lbl_summary.setText(summary)
        QMessageBox.information(self, "一键同步完成", summary)
        QTimer.singleShot(0, lambda ids=synced_db_ids: self._refresh_after_price_sync(None, ids))

    def open_price_product_spec_dialog(self, product_id):
        product_id = str(product_id or "").strip()
        compare = self.price_sync_rows.get(product_id) or {}
        rows = compare.get("rows") or []
        product_db_id = rows[0].get("product_db_id") if rows else None
        product_title = rows[0].get("product_title") if rows else ""
        if not product_db_id:
            QMessageBox.warning(self, "提示", "没有找到软件商品，不能打开规格与毛利管理。")
            return

        parent = self.owner or self.parent()
        dialog = None
        try:
            if parent and hasattr(parent, "open_product_spec_dialog"):
                dialog = parent.open_product_spec_dialog(self.db, product_db_id, product_id, product_title, parent)
            elif parent and hasattr(parent, "open_spec_dialog"):
                dialog = parent.open_spec_dialog(product_db_id, product_id, product_title)
            elif parent and hasattr(parent, "main_app") and hasattr(parent.main_app, "open_product_spec_dialog"):
                dialog = parent.main_app.open_product_spec_dialog(self.db, product_db_id, product_id, product_title, parent)
        except Exception as e:
            QMessageBox.warning(self, "打开失败", f"打开规格与毛利管理失败：{e}")
            return

        if dialog is None:
            QMessageBox.warning(self, "打开失败", "没有找到规格与毛利管理窗口入口。")
            return
        if hasattr(dialog, "finished"):
            dialog.finished.connect(lambda result=0, pid=product_id: self._refresh_price_product_after_edit(pid, result))
        elif dialog is not None and hasattr(dialog, "destroyed"):
            dialog.destroyed.connect(lambda _=None: None)
        self.lbl_summary.setText(f"已打开商品ID {product_id} 的规格与毛利管理窗口。保存后会刷新该链接匹配状态，取消不刷新。")

    def _refresh_price_product_after_edit(self, product_id, result):
        if result != QDialog.Accepted:
            return
        QTimer.singleShot(120, lambda pid=product_id: self.refresh_price_management_product(pid))

    def refresh_price_management_product(self, product_id):
        product_id = str(product_id or "").strip()
        cached = self.price_result_items.get(product_id) or {}
        item = cached.get("item")
        if not item:
            self.lbl_summary.setText(f"商品ID {product_id} 没有缓存的价格管理结果，无法单独刷新。")
            return
        local_product = self._local_products_for_store(self.current_store_id()).get(product_id)
        if not local_product:
            self.lbl_summary.setText(f"商品ID {product_id} 当前不在本店铺，无法刷新匹配状态。")
            return
        old_card = self.price_result_cards.get(product_id)
        insert_at = max(0, self.price_scroll_layout.count() - 1)
        if old_card is not None:
            index = self.price_scroll_layout.indexOf(old_card)
            if index >= 0:
                insert_at = index
                removed = self.price_scroll_layout.takeAt(index)
                widget = removed.widget()
                if widget is not None:
                    widget.deleteLater()
        compare = self._build_price_compare_rows(item, local_product)
        self.price_sync_rows[product_id] = compare
        self.price_unmatched_spec_product_ids = [
            pid for pid in (self.price_unmatched_spec_product_ids or []) if str(pid) != product_id
        ]
        if not compare.get("rows") or any(not row.get("local_spec_id") for row in compare.get("rows") or []):
            self.price_unmatched_spec_product_ids.append(product_id)
        if compare.get("all_matched"):
            self.price_unmatched_records.pop(product_id, None)
            self.price_current_page_matched_product_ids.append(product_id)
        else:
            categories = self._price_compare_categories(compare)
            self.price_unmatched_records[product_id] = {
                "product_id": product_id,
                "title": item.get("title") or local_product.get("title", ""),
                "reasons": self._price_unmatched_reason_labels(categories),
            }
        self.price_unmatched_spec_product_ids = sorted(set(self.price_unmatched_spec_product_ids), key=self.price_unmatched_spec_product_ids.index)
        self.price_current_page_matched_product_ids = sorted(set(self.price_current_page_matched_product_ids), key=self.price_current_page_matched_product_ids.index)
        if hasattr(self, "btn_copy_unmatched_specs"):
            self.btn_copy_unmatched_specs.setEnabled(bool(self._spec_unmatched_records()))
        if hasattr(self, "btn_sync_all_price"):
            self.btn_sync_all_price.setEnabled(any(compare.get("can_sync") for compare in (self.price_sync_rows or {}).values()))
        if self.unmatched_task_window is not None:
            if compare.get("all_matched") or not self._is_spec_unmatched_record(self.price_unmatched_records.get(product_id, {})):
                self.unmatched_task_window.remove_product_ids([product_id])
            else:
                self.unmatched_task_window.upsert_records([self.price_unmatched_records[product_id]])
        if self.unmatched_code_dialog is not None:
            self.unmatched_code_dialog._refresh_unmatched_spec_ids_panel()
        self._add_price_result_card(item, local_product, compare, insert_at=insert_at)
        self.lbl_summary.setText(f"商品ID {product_id} 已根据保存后的软件规格刷新匹配状态。")

    def refresh_all_price_management_matches(self):
        product_ids = list((self.price_result_items or {}).keys())
        if not product_ids:
            self.lbl_summary.setText("当前没有缓存的价格管理抓取结果，请先点击“开始抓取”。")
            return

        self.price_unmatched_spec_product_ids = []
        self.price_unmatched_records = {}
        self.price_current_page_matched_product_ids = []
        refreshed_count = 0
        for product_id in product_ids:
            cached = self.price_result_items.get(product_id) or {}
            item = cached.get("item")
            if not item:
                continue
            local_product = self._local_products_for_store(self.current_store_id()).get(str(product_id))
            if not local_product:
                self.price_unmatched_records[str(product_id)] = {
                    "product_id": str(product_id),
                    "title": item.get("title") or "",
                    "reasons": ["商品ID未匹配"],
                }
                continue
            self.refresh_price_management_product(str(product_id))
            refreshed_count += 1

        self.lbl_summary.setText(f"已基于当前缓存重新刷新 {refreshed_count}/{len(product_ids)} 个价格管理链接的匹配状态。")
        if hasattr(self, "btn_sync_all_price"):
            self.btn_sync_all_price.setEnabled(any(compare.get("can_sync") for compare in (self.price_sync_rows or {}).values()))

    def copy_missing_ids(self):
        QApplication.clipboard().setText("\n".join(self.missing_ids))
        self.lbl_summary.setText(f"已复制 {len(self.missing_ids)} 个未匹配商品ID。")

    def show_unmatched_task_window(self):
        if self.unmatched_task_window is None:
            self.unmatched_task_window = PddUnmatchedTaskWindow()
            self.unmatched_task_window.open_code_requested.connect(self.open_unmatched_code_match)
            self.unmatched_task_window.destroyed.connect(lambda _=None: setattr(self, "unmatched_task_window", None))
        self.unmatched_task_window.remove_product_ids(self.price_current_page_matched_product_ids)
        self.unmatched_task_window.upsert_records(self._spec_unmatched_records())
        if self.unmatched_task_window.isMinimized():
            self.unmatched_task_window.showNormal()
        else:
            self.unmatched_task_window.show()
        self.unmatched_task_window.raise_()
        self.unmatched_task_window.activateWindow()
        self.lbl_summary.setText(
            f"已生成未匹配规格链接窗口：本次新增/更新 {len(self._spec_unmatched_records())} 个规格未匹配链接。"
        )

    def open_unmatched_code_match(self, product_id):
        product_id = str(product_id or "").strip()
        store_id = self.current_store_id()
        if not product_id or not store_id:
            return
        try:
            if self.unmatched_code_dialog is None:
                self.unmatched_code_dialog = PddProductMatchDialog(
                    self.db,
                    self.monitor,
                    default_store_id=store_id,
                    parent=None,
                    mode="code",
                    store_id_provider=self.current_store_id,
                    owner=self.owner or self,
                    unmatched_records_provider=lambda: (
                        self.unmatched_task_window.records
                        if self.unmatched_task_window is not None
                        else self.price_unmatched_records
                    ),
                )
                self.unmatched_code_dialog.link_saved.connect(self._unmatched_code_saved)
                self.unmatched_code_dialog.destroyed.connect(lambda _=None: setattr(self, "unmatched_code_dialog", None))
            else:
                self.unmatched_code_dialog.unmatched_records_provider = lambda: (
                    self.unmatched_task_window.records
                    if self.unmatched_task_window is not None
                    else self.price_unmatched_records
                )
                self.unmatched_code_dialog._refresh_unmatched_spec_ids_panel()
            if self.unmatched_code_dialog.isMinimized():
                self.unmatched_code_dialog.showNormal()
            else:
                self.unmatched_code_dialog.show()
            self.unmatched_code_dialog.raise_()
            self.unmatched_code_dialog.activateWindow()
            self.unmatched_code_dialog.lbl_summary.setText(
                f"已打开抓取添加编码窗口。商品ID {product_id} 可在上方未匹配规格ID中复制；请手动打开对应添加编码界面后点击开始抓取。"
            )
        except Exception as e:
            QMessageBox.warning(self, "抓取添加编码", f"打开商品 {product_id} 的抓取添加编码失败：{e}")

    def _unmatched_code_saved(self, product_id):
        product_id = str(product_id or "").strip()
        if product_id in self.price_result_items:
            QTimer.singleShot(0, lambda pid=product_id: self.refresh_price_management_product(pid))

    def copy_unmatched_spec_product_ids(self):
        ids = [str(product_id).strip() for product_id in (self.price_unmatched_spec_product_ids or []) if str(product_id).strip()]
        ids = sorted(set(ids), key=ids.index)
        if not ids:
            self.lbl_summary.setText("当前没有未匹配规格的链接ID。")
            return
        QApplication.clipboard().setText("\n".join(ids))
        self.lbl_summary.setText(f"已复制 {len(ids)} 个未匹配规格的链接ID。")

    def copy_debug_json(self):
        QApplication.clipboard().setText(json.dumps(self.last_debug_info or {}, ensure_ascii=False, indent=2))
        self.lbl_summary.setText("已复制当前页面调试JSON。")


class PddPromotionStatusDialog(QDialog):
    """展示当前商品推广页面抓取到的直通车推广状态和软件投产设置对比。"""

    def __init__(self, db, monitor, store_id_provider, parent=None, owner=None):
        super().__init__(parent)
        self.db = db
        self.monitor = monitor
        self.store_id_provider = store_id_provider
        self.owner = owner or parent
        self.last_debug_info = {}
        self.captured_items = {}
        self.setWindowTitle("拼多多推广状态抓取")
        self.resize(1180, 720)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)

        top = QHBoxLayout()
        self.lbl_summary = QLabel("请先在浏览器切到商品推广页面，再点击“抓取当前推广页”。")
        self.lbl_summary.setWordWrap(True)
        self.lbl_summary.setStyleSheet("color: #555;")
        top.addWidget(self.lbl_summary, 1)
        self.btn_scan = QPushButton("抓取当前推广页")
        self.btn_scan.clicked.connect(self.scan_current_page)
        self._style_button(self.btn_scan, "#8e44ad", "#7d3c98", "#5b2c6f")
        top.addWidget(self.btn_scan)
        self.btn_sync_all = QPushButton("一键同步修改所有")
        self.btn_sync_all.clicked.connect(self.sync_all_items)
        self._style_button(self.btn_sync_all, "#27ae60", "#1e8449", "#145a32")
        top.addWidget(self.btn_sync_all)
        self.btn_clear = QPushButton("清空结果")
        self.btn_clear.clicked.connect(self.clear_results)
        self._style_button(self.btn_clear, "#566573", "#34495e", "#273746")
        top.addWidget(self.btn_clear)
        self.btn_copy_debug = QPushButton("复制调试JSON")
        self.btn_copy_debug.clicked.connect(self.copy_debug_json)
        self._style_button(self.btn_copy_debug, "#3498db", "#21618c", "#154360")
        top.addWidget(self.btn_copy_debug)
        layout.addLayout(top)

        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "主图", "商品ID", "软件标题", "抓取推广", "抓取类型", "抓取数值",
            "软件当前设置", "匹配结果", "抓取原文", "同步修改"
        ])
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(64)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setColumnWidth(0, 64)
        self.table.setColumnWidth(1, 135)
        self.table.setColumnWidth(2, 240)
        self.table.setColumnWidth(3, 82)
        self.table.setColumnWidth(4, 118)
        self.table.setColumnWidth(5, 88)
        self.table.setColumnWidth(6, 230)
        self.table.setColumnWidth(7, 170)
        self.table.setColumnWidth(9, 96)
        layout.addWidget(self.table)

        bottom = QHBoxLayout()
        bottom.addStretch()
        btn_close = QPushButton("关闭")
        btn_close.clicked.connect(self.close)
        self._style_button(btn_close, "#7f8c8d", "#566573", "#424949")
        bottom.addWidget(btn_close)
        layout.addLayout(bottom)

    def _button_style(self, bg, hover, pressed=None):
        pressed = pressed or hover
        return (
            f"QPushButton {{ background-color: {bg}; color: white; border: 1px solid {hover}; "
            "padding: 7px 14px; border-radius: 5px; font-weight: bold; }}"
            f"QPushButton:hover {{ background-color: {hover}; border: 1px solid {pressed}; }}"
            f"QPushButton:pressed {{ background-color: {pressed}; padding-top: 8px; padding-left: 15px; }}"
            "QPushButton:disabled { background-color: #bdc3c7; border-color: #aeb6bf; color: #f4f6f7; }"
        )

    def _style_button(self, button, bg, hover, pressed=None):
        button.setCursor(Qt.PointingHandCursor)
        button.setStyleSheet(self._button_style(bg, hover, pressed))

    def _show_toast(self, message):
        main_app = self._main_app()
        if main_app is not None and hasattr(main_app, "show_toast"):
            main_app.show_toast(message)
        else:
            self.lbl_summary.setText(message)

    def current_store_id(self):
        try:
            return self.store_id_provider() if self.store_id_provider else None
        except Exception:
            return None

    def _load_local_products(self):
        store_id = self.current_store_id()
        if not store_id:
            return {}
        rows = self.db.safe_fetchall(
            """SELECT id, name, title, image_data, COALESCE(current_roi, 0),
                      COALESCE(transaction_bid, 0), COALESCE(roi_input_mode, 'roi'),
                      COALESCE(is_natural_flow, 0), COALESCE(is_sitewide_managed, 0)
               FROM products
               WHERE store_id=? AND COALESCE(is_archived, 0)=0""",
            (store_id,),
        )
        return {
            str(row[1] or "").strip(): {
                "db_id": row[0],
                "product_id": str(row[1] or "").strip(),
                "title": row[2] or "",
                "image_data": row[3],
                "current_roi": float(row[4] or 0),
                "transaction_bid": float(row[5] or 0),
                "roi_input_mode": row[6] if row[6] in ("roi", "bid") else "roi",
                "is_natural_flow": bool(row[7]),
                "is_sitewide_managed": bool(row[8]),
            }
            for row in rows
        }

    def _numeric_value(self, value):
        match = re.search(r"-?\d+(?:\.\d+)?", str(value or "").replace(",", ""))
        return float(match.group(0)) if match else None

    def _software_setting_text(self, product):
        if not product:
            return "软件未找到该商品ID"
        if product.get("is_natural_flow"):
            return "自然流｜软件当前未开启直通车"
        if product.get("is_sitewide_managed"):
            return "全站托管｜软件当前不是直通车模式"
        if product.get("roi_input_mode") == "bid":
            return f"稳定成本推广｜净成交出价 ¥{product.get('transaction_bid', 0):.2f}｜折算投产 {product.get('current_roi', 0):.2f}"
        return f"稳定成本推广｜净目标投产比 {product.get('current_roi', 0):.2f}"

    def _compare_item(self, item, product):
        if not product:
            return "未匹配：软件没有这个商品ID", False

        ad_enabled = bool(item.get("ad_enabled"))
        software_active = (not product.get("is_natural_flow")) and (not product.get("is_sitewide_managed"))
        problems = []
        if ad_enabled != software_active:
            if ad_enabled:
                problems.append("开关不一致：抓取已开启，软件当前未开启；同步后会改为稳定成本推广")
            else:
                problems.append("开关不一致：抓取已关闭，软件当前为推广模式；同步后会改为自然流")

        if not software_active:
            if not problems:
                return "匹配：软件非直通车推广，页面也未开启", True
            return "；".join(problems), False

        bid_type = str(item.get("bid_type") or "")
        bid_value = self._numeric_value(item.get("bid_value"))
        expected_type = "净成交出价" if product.get("roi_input_mode") == "bid" else "净目标投产比"
        expected_value = product.get("transaction_bid") if product.get("roi_input_mode") == "bid" else product.get("current_roi")

        if expected_type not in bid_type:
            problems.append(f"类型不一致：抓取{bid_type or '未识别'}，软件{expected_type}")
        if bid_value is None:
            problems.append("抓取数值未识别")
        elif abs(float(bid_value) - float(expected_value or 0)) > 0.01:
            problems.append(f"数值不一致：抓取{bid_value:.2f}，软件{float(expected_value or 0):.2f}")

        if problems:
            return "；".join(problems), False
        return "匹配", True

    def _set_item(self, row, col, text, color=None):
        item = QTableWidgetItem(str(text or ""))
        item.setTextAlignment(Qt.AlignCenter)
        item.setToolTip(str(text or ""))
        if color:
            item.setForeground(QColor(color))
        self.table.setItem(row, col, item)
        return item

    def _set_image_cell(self, row, image_data):
        label = QLabel()
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("padding: 1px; background: #fff;")
        if image_data:
            pixmap = QPixmap()
            pixmap.loadFromData(bytes(image_data))
            if not pixmap.isNull():
                label.setPixmap(pixmap.scaled(58, 58, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            else:
                label.setText("无图")
        else:
            label.setText("无图")
        self.table.setCellWidget(row, 0, label)

    def _set_sync_button(self, row, product_id, enabled=True):
        button = QPushButton("同步")
        button.setEnabled(bool(enabled))
        self._style_button(button, "#2e86de", "#1b4f72", "#154360")
        button.clicked.connect(lambda _checked=False, pid=product_id: self.sync_single_item(pid))
        self.table.setCellWidget(row, 9, button)

    def render_items(self):
        items = list(self.captured_items.values())
        items.sort(key=lambda item: str(item.get("product_id") or ""))
        local_products = self._load_local_products()
        self.table.setRowCount(0)
        matched_count = 0
        for item in items:
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setRowHeight(row, 64)
            product_id = str(item.get("product_id") or "").strip()
            product = local_products.get(product_id)
            result_text, matched = self._compare_item(item, product)
            if matched:
                matched_count += 1
            self._set_image_cell(row, product.get("image_data") if product else None)
            self._set_item(row, 1, product_id)
            self._set_item(row, 2, product.get("title", "") if product else "软件未匹配")
            self._set_item(row, 3, item.get("ad_status_text") or ("开启" if item.get("ad_enabled") else "关闭"), "#008000" if item.get("ad_enabled") else "#999999")
            self._set_item(row, 4, item.get("bid_type") or "--")
            self._set_item(row, 5, item.get("bid_value") or "--")
            self._set_item(row, 6, self._software_setting_text(product))
            self._set_item(row, 7, result_text, "#008000" if matched else "#d00000")
            self._set_item(row, 8, item.get("raw_text") or "")
            item["_sync_needed"] = bool(product) and not matched
            self._set_sync_button(row, product_id, enabled=item["_sync_needed"])
        return matched_count, len(items)

    def scan_current_page(self):
        try:
            if hasattr(self.monitor, "set_store_context"):
                self.monitor.set_store_context(self.current_store_id())
            info = self.monitor.inspect_promotion_status_page()
        except Exception as e:
            QMessageBox.warning(self, "抓取失败", f"读取当前推广页面失败：{e}")
            return

        self.last_debug_info = info
        items = info.get("items") or []
        for item in items:
            product_id = str(item.get("product_id") or "").strip()
            key = product_id or str(item.get("data_testid") or len(self.captured_items))
            self.captured_items[key] = item

        status = info.get("status") or f"已识别 {len(items)} 条推广状态"
        matched_count, total_count = self.render_items()
        self.lbl_summary.setText(f"{status}｜本次新增/更新 {len(items)} 条｜累计 {total_count} 条｜软件匹配 {matched_count}/{total_count} 条｜页面：{info.get('title') or '--'}")
        self.lbl_summary.setToolTip(json.dumps(info, ensure_ascii=False, indent=2))
        if not items:
            detail = info.get("status") or "当前页面没有识别到商品推广开关。请确认已经切到商品推广列表，并且表格行已加载出来。"
            if info.get("ok") is False:
                QMessageBox.warning(self, "抓取失败", detail)
            else:
                QMessageBox.information(self, "未识别到数据", detail)

    def clear_results(self):
        self.captured_items = {}
        self.table.setRowCount(0)
        self.lbl_summary.setText("已清空推广状态抓取结果。")

    def _main_app(self):
        owner = self.owner
        if owner is not None and hasattr(owner, "main_app"):
            return owner.main_app
        return owner

    def _record_sync_operation(self, product_db_id, old_text, new_text):
        main_app = self._main_app()
        text = f"拼多多推广状态同步：{old_text} → {new_text}"
        if main_app is not None and hasattr(main_app, "record_product_operation"):
            main_app.record_product_operation(
                product_db_id,
                text,
                metric="推广状态同步",
                old=old_text,
                new=new_text,
                change_type="pdd_promotion_status_sync",
            )

    def _refresh_after_sync(self):
        if append_event:
            append_event("pdd_promotion_sync:refresh:start")
        main_app = self._main_app()
        owner = self.owner
        if owner is not None and hasattr(owner, "load_data"):
            owner.load_data()
        if owner is not None and hasattr(owner, "update_summary"):
            owner.update_summary()
        if main_app is not None and hasattr(main_app, "refresh_store_cards"):
            store_id = self.current_store_id()
            if store_id:
                main_app.refresh_store_cards(store_id)
        if append_event:
            append_event("pdd_promotion_sync:refresh:done")

    def _target_setting_from_item(self, item, product):
        ad_enabled = bool(item.get("ad_enabled"))
        bid_type = str(item.get("bid_type") or "")
        bid_value = self._numeric_value(item.get("bid_value"))
        if not product:
            raise ValueError("软件没有这个商品ID")
        if not ad_enabled:
            return {
                "is_natural_flow": 1,
                "is_sitewide_managed": 0,
                "roi_input_mode": product.get("roi_input_mode") or "roi",
                "current_roi": product.get("current_roi") or 0,
                "transaction_bid": product.get("transaction_bid") or 0,
                "text": "自然流",
            }
        if bid_value is None:
            raise ValueError("推广开启但抓取数值为空，未同步")
        if "净成交出价" in bid_type:
            return {
                "is_natural_flow": 0,
                "is_sitewide_managed": 0,
                "roi_input_mode": "bid",
                "current_roi": product.get("current_roi") or 0,
                "transaction_bid": bid_value,
                "text": f"稳定成本推广｜净成交出价 ¥{bid_value:.2f}",
            }
        return {
            "is_natural_flow": 0,
            "is_sitewide_managed": 0,
            "roi_input_mode": "roi",
            "current_roi": bid_value,
            "transaction_bid": product.get("transaction_bid") or 0,
            "text": f"稳定成本推广｜净目标投产比 {bid_value:.2f}",
        }

    def _sync_item(self, product_id, item=None, refresh=True):
        product_id = str(product_id or "").strip()
        if append_event:
            append_event(f"pdd_promotion_sync:item:start product_id={product_id} refresh={refresh}")
        if not item:
            item = self.captured_items.get(product_id)
        if not item:
            return False, f"{product_id or '空ID'}：没有抓取数据"
        product = self._load_local_products().get(product_id)
        if not product:
            return False, f"{product_id}：软件没有这个商品ID"
        result_text, matched = self._compare_item(item, product)
        if matched:
            if append_event:
                append_event(f"pdd_promotion_sync:item:skip_matched product_id={product_id}")
            return True, f"{product_id}：已匹配，无需同步"
        old_text = self._software_setting_text(product)
        try:
            target = self._target_setting_from_item(item, product)
        except Exception as e:
            return False, f"{product_id}：{e}"
        unchanged = (
            int(product.get("is_natural_flow") or 0) == target["is_natural_flow"]
            and int(product.get("is_sitewide_managed") or 0) == target["is_sitewide_managed"]
            and (
                target["is_natural_flow"]
                or (
                    str(product.get("roi_input_mode") or "roi") == target["roi_input_mode"]
                    and abs(float(product.get(
                        "transaction_bid" if target["roi_input_mode"] == "bid" else "current_roi"
                    ) or 0) - float(target[
                        "transaction_bid" if target["roi_input_mode"] == "bid" else "current_roi"
                    ] or 0)) <= 0.01
                )
            )
        )
        if unchanged:
            if append_event:
                append_event(f"pdd_promotion_sync:item:skip_unchanged product_id={product_id}")
            return True, f"{product_id}：推广数据未变化"
        if append_event:
            append_event(f"pdd_promotion_sync:item:update_db product_id={product_id}")
        self.db.safe_execute(
            """UPDATE products
               SET is_natural_flow=?, is_sitewide_managed=?, roi_input_mode=?,
                   current_roi=?, transaction_bid=?
               WHERE id=?""",
            (
                target["is_natural_flow"],
                target["is_sitewide_managed"],
                target["roi_input_mode"],
                target["current_roi"],
                target["transaction_bid"],
                product["db_id"],
            ),
        )
        if append_event:
            append_event(f"pdd_promotion_sync:item:record product_id={product_id}")
        self._record_sync_operation(product["db_id"], old_text, target["text"])
        if refresh:
            self._refresh_after_sync()
            self.render_items()
        if append_event:
            append_event(f"pdd_promotion_sync:item:done product_id={product_id}")
        return True, f"{product_id}：已同步为{target['text']}"

    def sync_single_item(self, product_id):
        ok, message = self._sync_item(product_id, refresh=True)
        if ok:
            self.lbl_summary.setText(message)
        else:
            QMessageBox.warning(self, "同步失败", message)

    def sync_all_items(self):
        if append_event:
            append_event("pdd_promotion_sync:all:start")
        if not self.captured_items:
            QMessageBox.information(self, "提示", "请先抓取推广状态。")
            return
        local_products = self._load_local_products()
        sync_items = {}
        skipped = 0
        for key, item in (self.captured_items or {}).items():
            product_id = str(item.get("product_id") or key or "").strip()
            product = local_products.get(product_id)
            _result_text, matched = self._compare_item(item, product)
            if product and not matched:
                sync_items[key] = item
            else:
                skipped += 1
        if not sync_items:
            if append_event:
                append_event("pdd_promotion_sync:all:no_changes")
            self.lbl_summary.setText("当前抓取结果全部匹配，无需同步。")
            QMessageBox.information(self, "无需同步", "当前抓取结果全部匹配，没有需要修改的推广设置。")
            return
        reply = QMessageBox.question(
            self,
            "确认同步",
            f"确定同步 {len(sync_items)} 条不一致的推广设置吗？\n已匹配的 {skipped} 条不会修改、不会记录操作记录。",
        )
        if reply != QMessageBox.Yes:
            return
        success = 0
        failed = []
        for key, item in list(sync_items.items()):
            product_id = str(item.get("product_id") or key or "").strip()
            ok, message = self._sync_item(product_id, item=item, refresh=False)
            if ok:
                success += 1
            else:
                failed.append(message)
        self._refresh_after_sync()
        self.render_items()
        if append_event:
            append_event(f"pdd_promotion_sync:all:done success={success} failed={len(failed)} skipped={skipped}")
        summary = f"一键同步完成：成功 {success} 条，跳过/失败 {len(failed)} 条。"
        if failed:
            summary += " " + "；".join(failed[:5])
        self.lbl_summary.setText(summary)
        if failed:
            self.lbl_summary.setToolTip("\n".join(failed))

    def sync_all_items(self):
        if append_event:
            append_event("pdd_promotion_sync:all:start")
        if not self.captured_items:
            self._show_toast("请先抓取推广状态")
            return
        local_products = self._load_local_products()
        sync_items = {}
        skipped = 0
        for key, item in (self.captured_items or {}).items():
            product_id = str(item.get("product_id") or key or "").strip()
            product = local_products.get(product_id)
            _result_text, matched = self._compare_item(item, product)
            if product and not matched:
                sync_items[key] = item
            else:
                skipped += 1
        if not sync_items:
            if append_event:
                append_event("pdd_promotion_sync:all:no_changes")
            self.lbl_summary.setText("当前抓取结果全部匹配，无需同步。")
            self._show_toast("无需修改")
            return
        success = 0
        failed = []
        for key, item in list(sync_items.items()):
            product_id = str(item.get("product_id") or key or "").strip()
            ok, message = self._sync_item(product_id, item=item, refresh=False)
            if ok:
                success += 1
            else:
                failed.append(message)
        self._refresh_after_sync()
        self.render_items()
        if append_event:
            append_event(f"pdd_promotion_sync:all:done success={success} failed={len(failed)} skipped={skipped}")
        summary = f"一键同步完成：已修改 {success} 条，跳过/失败 {len(failed)} 条。"
        if failed:
            summary += " " + "；".join(failed[:5])
        self.lbl_summary.setText(summary)
        self.lbl_summary.setToolTip("\n".join(failed) if failed else "")
        self._show_toast(f"已修改 {success} 条")

    def copy_debug_json(self):
        QApplication.clipboard().setText(json.dumps(self.last_debug_info or {}, ensure_ascii=False, indent=2))
        self.lbl_summary.setText("已复制推广状态抓取调试JSON。")


class PddLinkControlDialog(QDialog):
    """拼多多链接抓取主控小窗口。"""

    def __init__(self, db, monitor, default_store_id=None, parent=None):
        super().__init__(parent)
        self.db = db
        self.monitor = monitor
        self.default_store_id = default_store_id
        self.code_dialog = None
        self.price_dialog = None
        self.promotion_status_dialog = None
        self.setWindowTitle("拼多多链接抓取")
        self.resize(620, 260)
        self.init_ui()
        self.load_stores()
        QTimer.singleShot(250, self.refresh_browser_display)

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)

        store_row = QHBoxLayout()
        store_row.addWidget(QLabel("店铺"))
        self.combo_store = QComboBox()
        self.combo_store.setMinimumWidth(260)
        store_row.addWidget(self.combo_store, 1)
        layout.addLayout(store_row)

        self.lbl_browser_display = QLabel("浏览器状态读取中...")
        self.lbl_browser_display.setWordWrap(True)
        self.lbl_browser_display.setMinimumHeight(86)
        self.lbl_browser_display.setStyleSheet(
            "background-color: #111827; color: #d1fae5; border-radius: 6px; padding: 10px; font-size: 12px;"
        )
        layout.addWidget(self.lbl_browser_display)

        button_row = QHBoxLayout()
        self.btn_open_browser = QPushButton("打开商家端")
        self.btn_open_browser.clicked.connect(self.open_browser)
        button_row.addWidget(self.btn_open_browser)

        self.btn_open_code = QPushButton("抓取添加编码界面")
        self.btn_open_code.clicked.connect(self.open_code_dialog)
        button_row.addWidget(self.btn_open_code)

        self.btn_open_price = QPushButton("抓取价格管理")
        self.btn_open_price.clicked.connect(self.open_price_dialog)
        button_row.addWidget(self.btn_open_price)

        self.btn_open_promotion_status = QPushButton("抓取推广状态")
        self.btn_open_promotion_status.clicked.connect(self.open_promotion_status_dialog)
        button_row.addWidget(self.btn_open_promotion_status)
        layout.addLayout(button_row)

        close_row = QHBoxLayout()
        close_row.addStretch()
        btn_close = QPushButton("关闭")
        btn_close.clicked.connect(self.close)
        close_row.addWidget(btn_close)
        layout.addLayout(close_row)

    def load_stores(self):
        self.combo_store.clear()
        if self.default_store_id is not None:
            rows = self.db.safe_fetchall("SELECT id, name FROM stores WHERE id=?", (self.default_store_id,))
            if rows:
                store_id, store_name = rows[0]
                self.combo_store.addItem(str(store_name or f"店铺{store_id}"), int(store_id))
            else:
                self.combo_store.addItem(f"店铺{self.default_store_id}", int(self.default_store_id))
            self.combo_store.setEnabled(False)
            self.combo_store.setToolTip("已固定为当前打开窗口所属店铺，不能在此切换店铺。")
            return

        rows = self.db.safe_fetchall("SELECT id, name FROM stores ORDER BY sort_order, id")
        for store_id, store_name in rows:
            self.combo_store.addItem(str(store_name or f"店铺{store_id}"), int(store_id))
        self.combo_store.setEnabled(True)

    def current_store_id(self):
        return self.combo_store.currentData()

    def _activate_browser_store_context(self):
        store_id = self.current_store_id()
        if store_id and hasattr(self.monitor, "set_store_context"):
            self.monitor.set_store_context(store_id)
        return store_id

    def open_browser(self):
        try:
            store_id = self._activate_browser_store_context()
            if hasattr(self.monitor, "activate_store_browser"):
                self.monitor.activate_store_browser(store_id, open_url=True, open_new_tab=False)
            else:
                self.monitor.open_merchant_page(store_id)
            self.refresh_browser_display()
        except Exception as e:
            QMessageBox.warning(self, "拼多多链接抓取", f"打开商家端失败：{e}")

    def _window_parent(self):
        return self.parent() or self

    def _show_result_dialog(self, dialog):
        if dialog.isMinimized():
            dialog.showNormal()
        else:
            dialog.show()
        dialog.setWindowState(dialog.windowState() & ~Qt.WindowMinimized | Qt.WindowActive)
        dialog.raise_()
        dialog.activateWindow()

    def open_code_dialog(self):
        self._activate_browser_store_context()
        self.refresh_browser_display()
        if self.code_dialog is None:
            self.code_dialog = PddProductMatchDialog(
                self.db,
                self.monitor,
                default_store_id=self.current_store_id(),
                parent=None,
                mode="code",
                store_id_provider=self.current_store_id,
                owner=self._window_parent(),
                unmatched_records_provider=lambda: (
                    self.price_dialog.unmatched_task_window.records
                    if self.price_dialog is not None and self.price_dialog.unmatched_task_window is not None
                    else self.price_dialog.price_unmatched_records
                    if self.price_dialog is not None
                    else {}
                ),
            )
            self.code_dialog.destroyed.connect(lambda _=None: setattr(self, "code_dialog", None))
        else:
            self.code_dialog.unmatched_records_provider = lambda: (
                self.price_dialog.unmatched_task_window.records
                if self.price_dialog is not None and self.price_dialog.unmatched_task_window is not None
                else self.price_dialog.price_unmatched_records
                if self.price_dialog is not None
                else {}
            )
            self.code_dialog._refresh_unmatched_spec_ids_panel()
        self._show_result_dialog(self.code_dialog)

    def open_price_dialog(self):
        self._activate_browser_store_context()
        self.refresh_browser_display()
        if self.price_dialog is None:
            self.price_dialog = PddProductMatchDialog(
                self.db,
                self.monitor,
                default_store_id=self.current_store_id(),
                parent=None,
                mode="price",
                store_id_provider=self.current_store_id,
                owner=self._window_parent(),
            )
            self.price_dialog.destroyed.connect(lambda _=None: setattr(self, "price_dialog", None))
        self._show_result_dialog(self.price_dialog)

    def open_promotion_status_dialog(self):
        self._activate_browser_store_context()
        self.refresh_browser_display()
        if self.promotion_status_dialog is None:
            self.promotion_status_dialog = PddPromotionStatusDialog(
                self.db,
                self.monitor,
                self.current_store_id,
                parent=None,
                owner=self._window_parent(),
            )
            self.promotion_status_dialog.destroyed.connect(lambda _=None: setattr(self, "promotion_status_dialog", None))
        self._show_result_dialog(self.promotion_status_dialog)

    def refresh_browser_display(self):
        try:
            self._activate_browser_store_context()
            info = self.monitor.inspect()
        except Exception as e:
            self.lbl_browser_display.setText(f"浏览器状态：检测失败\n{e}")
            return
        title = info.get("title") or "--"
        url = info.get("url") or "--"
        status = info.get("status") or "未知状态"
        detail = info.get("current_code_detail") or {}
        page_type = (
            "价格管理"
            if info.get("is_price_management") or "价格管理" in title or "goods-price-management" in url
            else "商品推广" if info.get("is_promotion_page") or any(marker in f"{title} {url}" for marker in ("商品推广", "商品营销", "推广平台", "推广", "promotion", "advert", "ad_manage", "marketing", "traffic"))
            else "添加编码界面" if detail.get("product_id") or detail.get("specs")
            else "商家端"
        )
        code_text = f"添加编码：{detail.get('product_id') or '未识别'} / 规格 {len(detail.get('specs') or [])} 条"
        self.lbl_browser_display.setText(
            f"浏览器状态：{status}\n页面类型：{page_type}\n标题：{title}\n{code_text}\nURL：{url[:90]}"
        )
        self.lbl_browser_display.setToolTip(json.dumps(info, ensure_ascii=False, indent=2))


class ScalableTableWidget(QTableWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.scale_factor = 1.0

    def wheelEvent(self, event):
        if event.modifiers() & Qt.ControlModifier:
            delta = event.angleDelta().y()
            if delta > 0:
                self.scale_factor = min(2.0, self.scale_factor + 0.1)
            else:
                self.scale_factor = max(0.5, self.scale_factor - 0.1)
            self.apply_scale()
            event.accept()
            return
        super().wheelEvent(event)

    def apply_scale(self):
        base_row_height = 60
        base_header_height = 50

        row_height = int(base_row_height * self.scale_factor)
        header_height = int(base_header_height * self.scale_factor)

        for row in range(self.rowCount()):
            if row == 0 or row % 2 != 0:
                self.setRowHeight(row, row_height)
            else:
                self.setRowHeight(row, int(row_height * 0.4))

        header = self.horizontalHeader()
        header.setFixedHeight(header_height)

        font_size = int(16 * self.scale_factor)
        header_font = QFont()
        header_font.setPointSize(max(10, int(16 * self.scale_factor)))
        header.setFont(header_font)

        table_font = QFont()
        table_font.setPointSize(font_size)
        self.setFont(table_font)

        for row in range(self.rowCount()):
            for col in range(self.columnCount()):
                item = self.item(row, col)
                if item:
                    font = item.font()
                    font.setPointSize(font_size)
                    item.setFont(font)


class ImageCell(QWidget):
    cell_hovered = pyqtSignal(int)
    image_view_requested = pyqtSignal(int)
    paste_requested = pyqtSignal(int)

    def __init__(self, index, parent=None):
        super().__init__(parent)
        self.index = index
        self.current_pixmap = None
        self.setStyleSheet("border: 1px solid #c7ccd4; background-color: #ffffff; border-radius: 0px;")
        layout = QVBoxLayout(self)
        layout.setContentsMargins(1, 1, 1, 1)
        layout.setSpacing(0)
        layout.setAlignment(Qt.AlignCenter)
        self.image_label = QLabel("Ctrl+V\n粘贴图片")
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setStyleSheet("border: none; color: #999999; font-size: 11px; background: transparent;")
        layout.addWidget(self.image_label)
        self.setMouseTracking(True)
        self.setFocusPolicy(Qt.ClickFocus)
        self.setCursor(Qt.PointingHandCursor)

    def enterEvent(self, event):
        self.cell_hovered.emit(self.index)
        self.setFocus()
        super().enterEvent(event)

    def leaveEvent(self, event):
        self.clearFocus()
        super().leaveEvent(event)

    def keyPressEvent(self, event):
        if event.modifiers() & Qt.ControlModifier and event.key() == Qt.Key_V:
            self.paste_requested.emit(self.index)
            return
        super().keyPressEvent(event)

    def mouseDoubleClickEvent(self, event):
        if self.current_pixmap:
            self.image_view_requested.emit(self.index)
        super().mouseDoubleClickEvent(event)

    def set_image(self, pixmap):
        self.current_pixmap = pixmap
        self.image_label.setText("")
        self._refresh_thumbnail()

    def clear_image(self):
        self.current_pixmap = None
        self.image_label.clear()
        self.image_label.setText("Ctrl+V\n粘贴图片")

    def has_image(self):
        return self.current_pixmap is not None

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._refresh_thumbnail()

    def _refresh_thumbnail(self):
        if not self.current_pixmap or self.current_pixmap.isNull():
            return
        target_width = max(1, self.width() - 2)
        target_height = max(1, self.height() - 2)
        scaled = self.current_pixmap.scaled(
            target_width,
            target_height,
            Qt.KeepAspectRatioByExpanding,
            Qt.SmoothTransformation,
        )
        x = max(0, (scaled.width() - target_width) // 2)
        y = max(0, (scaled.height() - target_height) // 2)
        self.image_label.setPixmap(scaled.copy(x, y, target_width, target_height))


def _next_store_image_slot(rows):
    return max((int(row[0]) for row in rows), default=-1) + 1


def _should_clear_weekly_images(latest_end_date, new_end_date, is_overwrite):
    return bool(not is_overwrite and latest_end_date and new_end_date and new_end_date > latest_end_date)


class LargeMarginDataDialog(QDialog):
    """放大版毛利数据表格窗口"""
    FORMULAS = {
        "日期": None,
        "实发订单": None,
        "实发金额": None,
        "毛利润": None,
        "毛利率": "毛利润 / 实发金额 × 100%",
        "退款金额": None,
        "金额退款率": "退款金额 / 实发金额 × 100%",
        "退款订单": None,
        "订单退款率": "退款订单 / 实发订单 × 100%",
        "件单价": "实发金额 / 实发订单",
        "推广费": None,
        "推广占比": "推广费 / 实发金额 × 100%",
        "技术服务费": "实发金额 × 0.6%",
        "扣款": None,
        "其他服务": None,
        "其他": None,
        "净利润": "毛利润 - 退款金额 - 推广费 - 扣款 - 其他服务 + 其他 - 技术服务费",
        "净利率": "净利润 / 实发金额 × 100%",
        "单笔利润": "净利润 / 实发订单",
        "日盈亏": "净利润 / 天数",
    }

    def load_all_data(self):
        """从数据库加载所有历史数据"""
        try:
            records = self.db.safe_fetchall("""
                SELECT start_date, end_date, actual_orders, actual_amount, gross_profit,
                       refund_amount, refund_orders, promotion_fee, deduction, other_service, other,
                       gross_margin_rate, refund_rate_by_amount, refund_rate_by_orders,
                       unit_price, promotion_ratio, tech_fee,
                       net_profit, net_margin_rate, profit_per_order
                FROM manual_margin_data WHERE store_id=? ORDER BY start_date ASC, end_date ASC
            """, (self.store_id,))
            return records
        except Exception as e:
            print(f"加载历史数据失败: {e}")
            return []

    def _add_week_comparison_row(self, row, current, previous, GREEN, RED, GRAY):
        """添加较上周对比数据行"""
        current_net_profit = current[17] if current[17] else 0
        previous_net_profit = previous[17] if previous[17] else 0
        current_net_margin = current[18] if current[18] else 0
        previous_net_margin = previous[18] if previous[18] else 0

        current_daily = 0
        if current[0] and current[1]:
            try:
                from datetime import datetime
                start_dt = datetime.strptime(current[0], "%Y-%m-%d")
                end_dt = datetime.strptime(current[1], "%Y-%m-%d")
                days = max(1, (end_dt - start_dt).days + 1)
                current_daily = current_net_profit / days if days > 0 else 0
            except:
                pass

        previous_daily = 0
        if previous[0] and previous[1]:
            try:
                from datetime import datetime
                start_dt = datetime.strptime(previous[0], "%Y-%m-%d")
                end_dt = datetime.strptime(previous[1], "%Y-%m-%d")
                days = max(1, (end_dt - start_dt).days + 1)
                previous_daily = previous_net_profit / days if days > 0 else 0
            except:
                pass

        for col in range(20):
            item = QTableWidgetItem()
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable & ~Qt.ItemIsSelectable)

            if col == 0:
                item.setText("较上周")
                item.setBackground(QColor("#e8e8e8"))
            elif col == 1:
                if previous[2] and previous[2] != 0:
                    change = ((current[2] or 0) - (previous[2] or 0)) / abs(previous[2]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 2:
                if (previous[3] or 0) != 0:
                    change = ((current[3] or 0) - (previous[3] or 0)) / abs(previous[3]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 3:
                if (previous[4] or 0) != 0:
                    change = ((current[4] or 0) - (previous[4] or 0)) / abs(previous[4]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 4:
                change = (current[11] or 0) - (previous[11] or 0)
                icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                item.setText(f"{icon} {abs(change):.1f}%")
                item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
            elif col == 5:
                if (previous[5] or 0) != 0:
                    change = ((current[5] or 0) - (previous[5] or 0)) / abs(previous[5]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(RED if change > 0 else GREEN if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 6:
                change = (current[12] or 0) - (previous[12] or 0)
                icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                item.setText(f"{icon} {abs(change):.1f}%")
                item.setForeground(RED if change > 0 else GREEN if change < 0 else GRAY)
            elif col == 7:
                if previous[6] and previous[6] != 0:
                    change = ((current[6] or 0) - (previous[6] or 0)) / abs(previous[6]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(RED if change > 0 else GREEN if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 8:
                change = (current[13] or 0) - (previous[13] or 0)
                icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                item.setText(f"{icon} {abs(change):.1f}%")
                item.setForeground(RED if change > 0 else GREEN if change < 0 else GRAY)
            elif col == 9:
                if (previous[14] or 0) != 0:
                    change = ((current[14] or 0) - (previous[14] or 0)) / abs(previous[14]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 10:
                if (previous[7] or 0) != 0:
                    change = ((current[7] or 0) - (previous[7] or 0)) / abs(previous[7]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 11:
                change = (current[15] or 0) - (previous[15] or 0)
                icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                item.setText(f"{icon} {abs(change):.1f}%")
                item.setForeground(RED if change > 0 else GREEN if change < 0 else GRAY)
            elif col == 12:
                if (previous[16] or 0) != 0:
                    change = ((current[16] or 0) - (previous[16] or 0)) / abs(previous[16]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(RED if change > 0 else GREEN if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 13:
                if (previous[8] or 0) != 0:
                    change = ((current[8] or 0) - (previous[8] or 0)) / abs(previous[8]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(RED if change > 0 else GREEN if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 14:
                if (previous[9] or 0) != 0:
                    change = ((current[9] or 0) - (previous[9] or 0)) / abs(previous[9]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(RED if change > 0 else GREEN if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 15:
                if (previous[10] or 0) != 0:
                    change = ((current[10] or 0) - (previous[10] or 0)) / abs(previous[10]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 16:
                if previous_net_profit != 0:
                    change = (current_net_profit - previous_net_profit) / abs(previous_net_profit) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 17:
                change = current_net_margin - previous_net_margin
                icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                item.setText(f"{icon} {abs(change):.1f}%")
                item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
            elif col == 18:
                if (previous[19] or 0) != 0:
                    change = ((current[19] or 0) - (previous[19] or 0)) / abs(previous[19]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 19:
                if previous_daily != 0:
                    change = (current_daily - previous_daily) / abs(previous_daily) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)

            item.setFont(QFont("", -1, QFont.Bold))
            self.table.setItem(row, col, item)
        self.table.setRowHeight(row, 22)

    def __init__(self, store_name, store_id, db, parent=None):
        super().__init__(parent)
        self.store_id = store_id
        self.db = db
        self.parent_dialog = parent
        self.setWindowTitle(f"📊 {store_name} - 毛利数据明细（放大版）")
        self.setStyleSheet("background-color: #f5f5f5;")

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)

        header_label = QLabel("📈 毛利数据明细 - 放大查看模式（点击右上角关闭按钮退出）")
        header_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50; padding: 10px;")
        main_layout.addWidget(header_label)

        self.table = ScalableTableWidget()
        self.table.setColumnCount(21)
        self.table.setHorizontalHeaderLabels([
            "日期", "实发订单", "实发金额", "毛利润", "毛利率", "退款金额", "金额退款率",
            "退款订单", "订单退款率", "件单价", "推广费", "推广占比",
            "技术服务费", "扣款", "其他服务", "其他", "净利润",
            "净利率", "单笔利润", "日盈亏", "操作"
        ])

        records = self.load_all_data()
        total_rows = 2 * len(records) - 1
        self.table.setRowCount(total_rows)

        GREEN = QColor("#27ae60")
        RED = QColor("#e74c3c")
        GRAY = QColor("#999999")

        current_table_row = 0
        for i, record in enumerate(records):
            table_row = current_table_row
            start_date = record[0] if record[0] else ""
            end_date = record[1] if record[1] else ""
            start_display = start_date[5:10] if start_date and len(start_date) >= 10 else start_date
            end_display = end_date[5:10] if end_date and len(end_date) >= 10 else end_date
            date_str = f"{start_display}\n{end_display}"

            days = 1
            if start_date and end_date:
                try:
                    from datetime import datetime
                    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
                    days = max(1, (end_dt - start_dt).days + 1)
                except:
                    pass

            net_profit = record[17] if record[17] else 0
            daily_profit = net_profit / days if days > 0 else 0

            values = [
                date_str,
                str(int(record[2])) if record[2] else "0",
                f"¥{record[3]:.2f}" if record[3] else "¥0.00",
                f"¥{record[4]:.2f}" if record[4] else "¥0.00",
                f"{record[11]:.2f}%" if record[11] else "0.00%",
                f"¥{record[5]:.2f}" if record[5] else "¥0.00",
                f"{record[12]:.2f}%" if record[12] else "0.00%",
                str(int(record[6])) if record[6] else "0",
                f"{record[13]:.2f}%" if record[13] else "0.00%",
                f"¥{record[14]:.2f}" if record[14] else "¥0.00",
                f"¥{record[7]:.2f}" if record[7] else "¥0.00",
                f"{record[15]:.2f}%" if record[15] else "0.00%",
                f"¥{record[16]:.2f}" if record[16] else "¥0.00",
                f"¥{record[8]:.2f}" if record[8] else "¥0.00",
                f"¥{record[9]:.2f}" if record[9] else "¥0.00",
                f"¥{record[10]:.2f}" if record[10] else "¥0.00",
                f"¥{record[17]:.2f}" if record[17] else "¥0.00",
                f"{record[18]:.2f}%" if record[18] else "0.00%",
                f"¥{record[19]:.2f}" if record[19] else "¥0.00",
                f"¥{daily_profit:.2f}",
            ]

            for j, value in enumerate(values):
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                if j == 0:
                    item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                if j == 0:
                    pass
                elif j in [1, 2, 3, 5, 7, 10, 13, 14, 15]:
                    item.setBackground(QColor("#c8e6c9"))
                    item.setForeground(QColor("#1b5e20"))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                else:
                    item.setBackground(QColor("#bbdefb"))
                    item.setForeground(QColor("#0d47a1"))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                self.table.setItem(table_row, j, item)

            delete_btn = QPushButton("🗑️")
            delete_btn.setStyleSheet("""
                QPushButton {
                    background-color: #e74c3c;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    font-size: 14px;
                    font-weight: bold;
                    padding: 5px 10px;
                }
                QPushButton:hover {
                    background-color: #c0392b;
                }
            """)
            delete_btn.clicked.connect(lambda checked, r=table_row, idx=i: self.delete_data_row_with_comparison(r, idx))
            self.table.setCellWidget(table_row, 20, delete_btn)
            self.table.setRowHeight(table_row, 60)

            if i > 0:
                self._add_week_comparison_row(table_row + 1, record, records[i - 1], GREEN, RED, GRAY)
                current_table_row += 2
            else:
                current_table_row += 1

        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(True)
        self.table.setGridStyle(Qt.SolidLine)
        self.table.setAlternatingRowColors(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)

        table_font = QFont()
        table_font.setPointSize(16)
        self.table.setFont(table_font)

        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #cccccc;
                font-size: 16px;
                border: 2px solid #cccccc;
                border-radius: 6px;
                background-color: white;
            }
            QTableWidget::item {
                padding: 1px;
                text-align: center;
                border: 1px solid #cccccc;
                font-size: 16px;
            }
            QTableWidget::item:selected {
                background-color: #e6f3ff;
                color: black;
                outline: none;
            }
            QHeaderView {
                border: none;
            }
            QHeaderView::section {
                background-color: #e0e0e0;
                padding: 1px;
                margin: 0px;
                border: none;
                border-left: 1px solid #cccccc;
                border-bottom: 1px solid #cccccc;
                border-right: 1px solid #cccccc;
                font-size: 16px;
                font-weight: bold;
                min-height: 50px;
            }
        """)

        self.header = self.table.horizontalHeader()
        self.header.setSectionResizeMode(QHeaderView.Interactive)
        self.header.setMinimumSectionSize(80)
        self.header.setStretchLastSection(True)
        self.header.setMouseTracking(True)
        self.header.viewport().setMouseTracking(True)
        self.header.viewport().installEventFilter(self)

        main_layout.addWidget(self.table)

        # 底部按钮行
        bottom_btn_widget = QWidget()
        bottom_btn_layout = QHBoxLayout(bottom_btn_widget)
        bottom_btn_layout.setContentsMargins(0, 0, 0, 0)
        
        self.btn_calculate_total = QPushButton("🧮 计算总和")
        self.btn_calculate_total.setFixedHeight(45)
        self.btn_calculate_total.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 14px;
                font-weight: bold;
                padding: 8px 20px;
            }
            QPushButton:hover {
                background-color: #219a52;
            }
            QPushButton:pressed {
                background-color: #1e8449;
            }
        """)
        self.btn_calculate_total.clicked.connect(self.calculate_total)
        bottom_btn_layout.addWidget(self.btn_calculate_total)
        
        bottom_btn_layout.addStretch()
        
        close_btn = QPushButton("关闭")
        close_btn.setFixedHeight(45)
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 14px;
                font-weight: bold;
                padding: 8px 20px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
            QPushButton:pressed {
                background-color: #a93226;
            }
        """)
        close_btn.clicked.connect(self.accept)
        bottom_btn_layout.addWidget(close_btn)
        
        main_layout.addWidget(bottom_btn_widget)

        self.records = records

        for col in range(self.table.columnCount()):
            self.header.setSectionResizeMode(col, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(True)
        QApplication.processEvents()
        total_width = self.table.horizontalHeader().length() + self.table.verticalHeader().width() + 50
        screen = QApplication.desktop().screenGeometry()
        window_width = min(max(total_width, 1200), screen.width() - 100)
        data_rows = (self.table.rowCount() + 1) // 2
        comparison_rows = self.table.rowCount() - data_rows
        window_height = min(max(200 + data_rows * 60 + comparison_rows * 12, 600), screen.height() - 100)
        self.resize(window_width, window_height)

    def showEvent(self, event):
        super().showEvent(event)
        self.reload_data()

    def reload_data(self):
        records = self.load_all_data()
        self.table.setRowCount(0)
        self.table.clearContents()

        if not records:
            self.records = []
            return

        total_rows = 2 * len(records) - 1
        self.table.setRowCount(total_rows)

        GREEN = QColor("#27ae60")
        RED = QColor("#e74c3c")
        GRAY = QColor("#999999")

        current_table_row = 0
        for i, record in enumerate(records):
            table_row = current_table_row
            start_date = record[0] if record[0] else ""
            end_date = record[1] if record[1] else ""
            start_display = start_date[5:10] if start_date and len(start_date) >= 10 else start_date
            end_display = end_date[5:10] if end_date and len(end_date) >= 10 else end_date
            date_str = f"{start_display}\n{end_display}"

            days = 1
            if start_date and end_date:
                try:
                    from datetime import datetime
                    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
                    days = max(1, (end_dt - start_dt).days + 1)
                except:
                    pass

            net_profit = record[17] if record[17] else 0
            daily_profit = net_profit / days if days > 0 else 0

            values = [
                date_str,
                str(int(record[2])) if record[2] else "0",
                f"¥{record[3]:.2f}" if record[3] else "¥0.00",
                f"¥{record[4]:.2f}" if record[4] else "¥0.00",
                f"{record[11]:.2f}%" if record[11] else "0.00%",
                f"¥{record[5]:.2f}" if record[5] else "¥0.00",
                f"{record[12]:.2f}%" if record[12] else "0.00%",
                str(int(record[6])) if record[6] else "0",
                f"{record[13]:.2f}%" if record[13] else "0.00%",
                f"¥{record[14]:.2f}" if record[14] else "¥0.00",
                f"¥{record[7]:.2f}" if record[7] else "¥0.00",
                f"{record[15]:.2f}%" if record[15] else "0.00%",
                f"¥{record[16]:.2f}" if record[16] else "¥0.00",
                f"¥{record[8]:.2f}" if record[8] else "¥0.00",
                f"¥{record[9]:.2f}" if record[9] else "¥0.00",
                f"¥{record[10]:.2f}" if record[10] else "¥0.00",
                f"¥{record[17]:.2f}" if record[17] else "¥0.00",
                f"{record[18]:.2f}%" if record[18] else "0.00%",
                f"¥{record[19]:.2f}" if record[19] else "¥0.00",
                f"¥{daily_profit:.2f}",
            ]

            for j, value in enumerate(values):
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                if j == 0:
                    item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                if j == 0:
                    pass
                elif j in [1, 2, 3, 5, 7, 10, 13, 14, 15]:
                    item.setBackground(QColor("#c8e6c9"))
                    item.setForeground(QColor("#1b5e20"))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                else:
                    item.setBackground(QColor("#bbdefb"))
                    item.setForeground(QColor("#0d47a1"))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                self.table.setItem(table_row, j, item)

            delete_btn = QPushButton("🗑️")
            delete_btn.setStyleSheet("""
                QPushButton {
                    background-color: #e74c3c;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    font-size: 14px;
                    font-weight: bold;
                    padding: 5px 10px;
                }
                QPushButton:hover {
                    background-color: #c0392b;
                }
            """)
            delete_btn.clicked.connect(lambda checked, r=table_row, idx=i: self.delete_data_row_with_comparison(r, idx))
            self.table.setCellWidget(table_row, 20, delete_btn)
            self.table.setRowHeight(table_row, 60)

            if i > 0:
                self._add_week_comparison_row(table_row + 1, record, records[i - 1], GREEN, RED, GRAY)
                current_table_row += 2
            else:
                current_table_row += 1

        self.records = records

    def show_context_menu(self, pos):
        menu = QMenu()
        hint_action = QAction("ℹ️ 对比行不可操作", self.table)
        hint_action.setEnabled(False)
        menu.addAction(hint_action)
        menu.exec_(self.table.viewport().mapToGlobal(pos))

    def delete_data_row_with_comparison(self, row, record_index):
        date_item = self.table.item(row, 0)
        date_text = date_item.text().split('\n')[0] if date_item else ""

        reply = QMessageBox.question(self, "确认删除", f"确定要删除 {date_text} 这行数据及其对比行吗？",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.No:
            return

        if record_index < len(self.records):
            record = self.records[record_index]
            self.db.safe_execute("""
                DELETE FROM manual_margin_data WHERE store_id=? AND start_date=? AND end_date=?
            """, (self.store_id, record[0], record[1]))

        if self.parent_dialog and hasattr(self.parent_dialog, 'update_current_history_label'):
            self.parent_dialog.update_current_history_label()
        if self.parent_dialog and hasattr(self.parent_dialog, 'refresh_manual_data_display'):
            self.parent_dialog.refresh_manual_data_display()

        self.reload_data()
        QApplication.processEvents()

        self.show_toast("✅ 已删除")

    def show_toast(self, message):
        """显示气泡提示（淡入淡出0.5秒）"""
        if not hasattr(self, 'toast_label'):
            self.toast_label = QLabel(self)
            self.toast_label.setStyleSheet("""
                background-color: rgba(0, 0, 0, 180);
                color: white;
                padding: 8px 16px;
                border-radius: 4px;
                font-size: 14px;
            """)
            self.toast_label.setAttribute(Qt.WA_TranslucentBackground)
            self.toast_label.setWindowFlags(Qt.FramelessWindowHint)
            self.toast_opacity = QGraphicsOpacityEffect()
            self.toast_label.setGraphicsEffect(self.toast_opacity)
            self.toast_opacity.setOpacity(0)

        self.toast_label.setText(message)
        self.toast_label.adjustSize()
        parent_pos = self.mapToGlobal(self.rect().bottomLeft())
        x = parent_pos.x() + (self.width() - self.toast_label.width()) // 2
        y = parent_pos.y() - self.toast_label.height() - 10
        self.toast_label.move(x, y)
        self.toast_label.show()
        self.toast_label.repaint()

        self.toast_opacity.setOpacity(1)
        QApplication.processEvents()

        QTimer.singleShot(500, self.fade_out_toast)

    def fade_out_toast(self):
        if hasattr(self, 'toast_opacity'):
            self.toast_opacity.setOpacity(0)
            QTimer.singleShot(500, self.toast_label.hide if hasattr(self, 'toast_label') else lambda: None)

    def calculate_total(self):
        """计算选中行或所有行的总和并弹出窗口显示"""
        records = self.load_all_data()
        if not records:
            QMessageBox.information(self, "提示", "无数据可计算")
            return

        def is_data_row(row):
            return row == 0 or row % 2 == 1

        selected_rows = self.table.selectionModel().selectedRows()

        if selected_rows:
            rows_to_calculate = [row.row() for row in selected_rows if is_data_row(row.row())]
        else:
            rows_to_calculate = [row for row in range(self.table.rowCount()) if is_data_row(row)]

        if not rows_to_calculate:
            QMessageBox.information(self, "提示", "没有可计算的数据行")
            return

        # 初始化总和
        total_orders = 0  # 实发订单
        total_amount = 0  # 实发金额
        total_gross_profit = 0  # 毛利润
        total_refund_amount = 0  # 退款金额
        total_refund_orders = 0  # 退款订单
        total_promotion_fee = 0  # 推广费
        total_deduction = 0  # 扣款
        total_other_service = 0  # 其他服务
        total_other = 0  # 其他
        total_days = 0  # 总天数

        for row in rows_to_calculate:
            if row >= self.table.rowCount():
                continue
            
            # 获取日期单元格并计算天数
            date_item = self.table.item(row, 0)
            if date_item:
                date_text = date_item.text()
                if '\n' in date_text:
                    parts = date_text.split('\n')
                    if len(parts) >= 2:
                        start_date_str = parts[0].strip()
                        end_date_str = parts[1].strip()
                        try:
                            from datetime import datetime
                            # 假设年份为当前年份
                            current_year = datetime.now().year
                            start_dt = datetime.strptime(f"{current_year}-{start_date_str}", "%Y-%m-%d")
                            end_dt = datetime.strptime(f"{current_year}-{end_date_str}", "%Y-%m-%d")
                            days = max(1, (end_dt - start_dt).days + 1)
                            total_days += days
                        except:
                            total_days += 1
                    else:
                        total_days += 1
                else:
                    total_days += 1
            else:
                total_days += 1
            
            # 获取单元格数据
            orders_item = self.table.item(row, 1)
            amount_item = self.table.item(row, 2)
            gross_profit_item = self.table.item(row, 3)
            refund_amount_item = self.table.item(row, 5)
            refund_orders_item = self.table.item(row, 7)
            promotion_fee_item = self.table.item(row, 10)
            deduction_item = self.table.item(row, 13)
            other_service_item = self.table.item(row, 14)
            other_item = self.table.item(row, 15)

            # 累加
            if orders_item:
                try:
                    total_orders += int(orders_item.text().replace('¥', '').replace(',', ''))
                except:
                    pass
            if amount_item:
                try:
                    total_amount += float(amount_item.text().replace('¥', '').replace(',', ''))
                except:
                    pass
            if gross_profit_item:
                try:
                    total_gross_profit += float(gross_profit_item.text().replace('¥', '').replace(',', ''))
                except:
                    pass
            if refund_amount_item:
                try:
                    total_refund_amount += float(refund_amount_item.text().replace('¥', '').replace(',', ''))
                except:
                    pass
            if refund_orders_item:
                try:
                    total_refund_orders += int(refund_orders_item.text().replace('¥', '').replace(',', ''))
                except:
                    pass
            if promotion_fee_item:
                try:
                    total_promotion_fee += float(promotion_fee_item.text().replace('¥', '').replace(',', ''))
                except:
                    pass
            if deduction_item:
                try:
                    total_deduction += float(deduction_item.text().replace('¥', '').replace(',', ''))
                except:
                    pass
            if other_service_item:
                try:
                    total_other_service += float(other_service_item.text().replace('¥', '').replace(',', ''))
                except:
                    pass
            if other_item:
                try:
                    total_other += float(other_item.text().replace('¥', '').replace(',', ''))
                except:
                    pass

        # 计算派生值
        tech_fee = total_amount * 0.006  # 技术服务费
        net_profit = total_gross_profit - total_refund_amount - total_promotion_fee - total_deduction - total_other_service + total_other - tech_fee
        
        # 计算百分比
        gross_margin_rate = (total_gross_profit / total_amount * 100) if total_amount > 0 else 0
        refund_rate_by_amount = (total_refund_amount / total_amount * 100) if total_amount > 0 else 0
        refund_rate_by_orders = (total_refund_orders / total_orders * 100) if total_orders > 0 else 0
        unit_price = (total_amount / total_orders) if total_orders > 0 else 0
        promotion_ratio = (total_promotion_fee / total_amount * 100) if total_amount > 0 else 0
        net_margin_rate = (net_profit / total_amount * 100) if total_amount > 0 else 0
        profit_per_order = (net_profit / total_orders) if total_orders > 0 else 0
        daily_profit = (net_profit / total_days) if total_days > 0 else 0

        # 弹出总和窗口
        total_dialog = QDialog(self)
        total_dialog.setWindowTitle("📊 数据总和")
        total_dialog.setStyleSheet("background-color: #f5f5f5;")
        total_layout = QVBoxLayout(total_dialog)
        total_layout.setContentsMargins(10, 10, 10, 10)
        
        title_label = QLabel(f"📈 数据总和（共{len(rows_to_calculate)}行，总计{total_days}天）")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50; padding: 10px;")
        total_layout.addWidget(title_label)
        
        # 创建表格显示总和
        total_table = QTableWidget()
        total_table.setColumnCount(20)
        total_table.setHorizontalHeaderLabels([
            "日期", "实发订单", "实发金额", "毛利润", "毛利率", "退款金额", "金额退款率",
            "退款订单", "订单退款率", "件单价", "推广费", "推广占比",
            "技术服务费", "扣款", "其他服务", "其他", "净利润",
            "净利率", "单笔利润", "日盈亏"
        ])
        
        total_table.setRowCount(1)
        
        # 填充数据
        values = [
            f"总计\n({len(rows_to_calculate)}行\n{total_days}天)",
            str(int(total_orders)),
            f"¥{total_amount:.2f}",
            f"¥{total_gross_profit:.2f}",
            f"{gross_margin_rate:.2f}%",
            f"¥{total_refund_amount:.2f}",
            f"{refund_rate_by_amount:.2f}%",
            str(int(total_refund_orders)),
            f"{refund_rate_by_orders:.2f}%",
            f"¥{unit_price:.2f}",
            f"¥{total_promotion_fee:.2f}",
            f"{promotion_ratio:.2f}%",
            f"¥{tech_fee:.2f}",
            f"¥{total_deduction:.2f}",
            f"¥{total_other_service:.2f}",
            f"¥{total_other:.2f}",
            f"¥{net_profit:.2f}",
            f"{net_margin_rate:.2f}%",
            f"¥{profit_per_order:.2f}",
            f"¥{daily_profit:.2f}"
        ]
        
        for j, value in enumerate(values):
            item = QTableWidgetItem(value)
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            if j == 0:
                item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                item.setBackground(QColor("#e8e8e8"))
            elif j in [1, 2, 3, 5, 7, 10, 13, 14, 15]:
                item.setBackground(QColor("#c8e6c9"))
                item.setForeground(QColor("#1b5e20"))
                font = item.font()
                font.setBold(True)
                item.setFont(font)
            else:
                item.setBackground(QColor("#bbdefb"))
                item.setForeground(QColor("#0d47a1"))
                font = item.font()
                font.setBold(True)
                item.setFont(font)
            total_table.setItem(0, j, item)
        
        total_table.setRowHeight(0, 60)
        total_table.verticalHeader().setVisible(False)
        total_table.setShowGrid(True)
        total_table.setGridStyle(Qt.SolidLine)
        total_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        
        table_font = QFont()
        table_font.setPointSize(16)
        total_table.setFont(table_font)
        
        total_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #cccccc;
                font-size: 16px;
                border: 2px solid #cccccc;
                border-radius: 6px;
                background-color: white;
            }
            QTableWidget::item {
                padding: 1px;
                text-align: center;
                border: 1px solid #cccccc;
                font-size: 16px;
            }
            QTableWidget::item:selected {
                background-color: #0078d7;
                color: white;
            }
            QHeaderView {
                border: none;
            }
            QHeaderView::section {
                background-color: #e0e0e0;
                padding: 1px;
                margin: 0px;
                border: none;
                border-left: 1px solid #cccccc;
                border-bottom: 1px solid #cccccc;
                border-right: 1px solid #cccccc;
                font-size: 16px;
                font-weight: bold;
                min-height: 50px;
            }
        """)
        
        header = total_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)
        header.setMinimumSectionSize(80)
        header.setStretchLastSection(True)
        
        total_layout.addWidget(total_table)
        
        # 调整列宽以适应内容
        for col in range(total_table.columnCount()):
            header.setSectionResizeMode(col, QHeaderView.ResizeToContents)
        total_table.horizontalHeader().setStretchLastSection(True)
        QApplication.processEvents()
        
        # 设置窗口大小
        total_width = header.length() + total_table.verticalHeader().width() + 50
        screen = QApplication.desktop().screenGeometry()
        window_width = min(max(total_width, 1200), screen.width() - 100)
        window_height = min(300, screen.height() - 100)
        total_dialog.resize(window_width, window_height)
        
        # 关闭按钮
        close_btn = QPushButton("关闭")
        close_btn.setFixedHeight(45)
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 14px;
                font-weight: bold;
                padding: 8px 20px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
            QPushButton:pressed {
                background-color: #a93226;
            }
        """)
        close_btn.clicked.connect(total_dialog.accept)
        total_layout.addWidget(close_btn)
        
        total_dialog.exec_()

    def eventFilter(self, obj, event):
        if obj == self.header.viewport() and event.type() == QEvent.MouseMove:
            col = self.header.logicalIndexAt(event.pos())
            if 0 <= col < len(self.FORMULAS):
                col_names = list(self.FORMULAS.keys())
                formula = self.FORMULAS[col_names[col]]
                if formula:
                    self.header.setToolTip(formula)
                else:
                    self.header.setToolTip("")
            return super().eventFilter(obj, event)
        return super().eventFilter(obj, event)


class StoreMarginDialog(QDialog):
    """店铺毛利综合管理对话框"""
    def __init__(self, store_id, store_name, main_app, parent=None, save_callback=None):
        super().__init__(parent)
        self.setAttribute(Qt.WA_DeleteOnClose, True)
        apply_window_icon(self, "store")
        self.store_id = store_id
        self.store_name = store_name
        self.main_app = main_app
        self.db = main_app.db
        self.product_weights = {}
        self.is_balancing = False
        self.save_callback = save_callback
        self._disposing = False
        self.is_reading_mode = False
        self.large_dialog = None
        self.ai_report_dialog = None
        self.ai_report_debug_dialog = None
        self.ai_report_thread = None
        self.ai_report_worker = None
        self.pdd_browser_monitor = None
        self.pdd_product_match_dialog = None
        self.pdd_link_control_dialog = None
        self._button_tooltip_widget = None
        self._button_tooltip_text = ""
        self._button_tooltip_timer = QTimer(self)
        self._button_tooltip_timer.setSingleShot(True)
        self._button_tooltip_timer.timeout.connect(self._show_delayed_button_tooltip)

        self.setWindowTitle(f"🏪 店铺毛利管理 - {store_name}")
        self.resize(1700, 800)

        self.toast_label = QLabel(self)
        self.toast_label.setWindowFlags(Qt.Tool | Qt.WindowStaysOnTopHint | Qt.FramelessWindowHint)
        self.toast_label.setAttribute(Qt.WA_TranslucentBackground)
        self.toast_label.setStyleSheet("""
            background-color: rgba(128, 128, 128, 0.5);
            color: black;
            padding: 10px 20px;
            border-radius: 8px;
            font-size: 14px;
        """)
        self.toast_label.setAlignment(Qt.AlignCenter)
        self.toast_label.hide()
        self.toast_label.setGraphicsEffect(QGraphicsOpacityEffect(opacity=0.5))
        self.toast_opacity_effect = QGraphicsOpacityEffect(opacity=0.5)
        self.toast_label.setGraphicsEffect(self.toast_opacity_effect)

        self.toast_fade_out_animation = QPropertyAnimation(self.toast_opacity_effect, b"opacity")
        self.toast_fade_out_animation.setDuration(500)
        self.toast_fade_out_animation.setStartValue(0.5)
        self.toast_fade_out_animation.setEndValue(0.0)
        self.toast_fade_out_animation.setEasingCurve(QEasingCurve.OutCubic)
        self.toast_fade_out_animation.finished.connect(self.toast_label.hide)

        self.init_ui()
        self._restore_latest_import_history_if_needed()
        self._normalize_imported_order_store_ids()
        self.load_products()
        self._apply_store_margin_button_tooltips()
        self.refresh_manual_data_display()

    def _build_weekly_images_panel(self):
        self.weekly_images_panel = QWidget(self)
        self.weekly_images_panel.setObjectName("WeeklyImagesPanel")
        self.weekly_images_panel.setAttribute(Qt.WA_StyledBackground, True)
        self.weekly_images_panel.setStyleSheet("""
            QWidget#WeeklyImagesPanel {
                background: white;
                border: 1px solid #9aa7b2;
                border-radius: 4px;
            }
        """)
        self.weekly_images_panel.resize(760, 360)
        panel_layout = QVBoxLayout(self.weekly_images_panel)
        panel_layout.setContentsMargins(10, 8, 10, 10)
        panel_layout.setSpacing(6)

        header = QHBoxLayout()
        title = QLabel("本周附带图片")
        title.setStyleSheet("font-size: 14px; font-weight: bold; color: #263746; border: none;")
        self.lbl_weekly_image_count = QLabel("0 张")
        self.lbl_weekly_image_count.setStyleSheet("color: #667788; border: none;")
        self.btn_clear_weekly_images = QPushButton("清空图片")
        self.btn_clear_weekly_images.setFixedSize(76, 28)
        self.btn_clear_weekly_images.setStyleSheet("QPushButton { padding: 1px; }")
        self.btn_clear_weekly_images.clicked.connect(self.clear_all_weekly_images)
        header.addWidget(title)
        header.addWidget(self.lbl_weekly_image_count)
        header.addStretch()
        header.addWidget(self.btn_clear_weekly_images)
        panel_layout.addLayout(header)

        self.weekly_images_scroll = QScrollArea()
        self.weekly_images_scroll.setWidgetResizable(True)
        self.weekly_images_scroll.setFrameShape(QScrollArea.NoFrame)
        self.weekly_images_content = QWidget()
        self.weekly_images_grid = QGridLayout(self.weekly_images_content)
        self.weekly_images_grid.setContentsMargins(1, 1, 1, 1)
        self.weekly_images_grid.setSpacing(1)
        self.weekly_images_grid.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.weekly_images_scroll.setWidget(self.weekly_images_content)
        panel_layout.addWidget(self.weekly_images_scroll)
        self.weekly_image_cells = {}
        self.weekly_images_panel.hide()

        app = QApplication.instance()
        if app is not None:
            app.installEventFilter(self)
            self._weekly_app_filter_installed = True

    def _position_weekly_images_panel(self):
        panel = getattr(self, "weekly_images_panel", None)
        button = getattr(self, "btn_weekly_images", None)
        if panel is None or button is None:
            return
        width = min(760, max(420, self.width() - 32))
        image_count = len(getattr(self, "weekly_image_cells", {})) + 1
        columns = max(1, (width - 22) // 87)
        image_rows = (image_count + columns - 1) // columns
        height = min(340, max(143, 56 + image_rows * 87))
        height = min(height, max(126, self.height() - 32))
        panel.resize(width, height)
        anchor = button.mapTo(self, button.rect().bottomRight())
        x = max(8, min(anchor.x() - width, self.width() - width - 8))
        y = anchor.y() + 4
        if y + height > self.height() - 8:
            y = max(8, button.mapTo(self, button.rect().topLeft()).y() - height - 4)
        panel.move(x, y)

    def toggle_weekly_images_panel(self):
        if self.weekly_images_panel.isVisible():
            self.weekly_images_panel.hide()
            return
        self.load_weekly_images()
        self._position_weekly_images_panel()
        self.weekly_images_panel.show()
        self.weekly_images_panel.raise_()

    def load_weekly_images(self):
        rows = self.db.safe_fetchall(
            "SELECT slot_index, image_data FROM store_temp_images WHERE store_id=? ORDER BY slot_index",
            (self.store_id,),
        )
        while self.weekly_images_grid.count():
            item = self.weekly_images_grid.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
        self.weekly_image_cells = {}
        columns = max(1, (self.weekly_images_panel.width() - 22) // 87)
        for position, (slot_index, image_data) in enumerate(rows):
            slot_index = int(slot_index)
            cell = ImageCell(slot_index, self.weekly_images_content)
            cell.setFixedSize(86, 86)
            cell.setToolTip("双击查看大图；查看窗口中按 Ctrl+滚轮缩放")
            pixmap = QPixmap()
            pixmap.loadFromData(bytes(image_data or b""))
            if not pixmap.isNull():
                cell.set_image(pixmap)
            cell.paste_requested.connect(lambda _index: self._append_weekly_image_from_clipboard())
            cell.image_view_requested.connect(self.show_weekly_image_viewer)
            self.weekly_images_grid.addWidget(cell, position // columns, position % columns)
            self.weekly_image_cells[slot_index] = cell

        add_slot = _next_store_image_slot(rows)
        add_cell = ImageCell(add_slot, self.weekly_images_content)
        add_cell.setFixedSize(86, 86)
        add_cell.paste_requested.connect(lambda _index: self._append_weekly_image_from_clipboard())
        position = len(rows)
        self.weekly_images_grid.addWidget(add_cell, position // columns, position % columns)
        self.lbl_weekly_image_count.setText(f"{len(rows)} 张")

    def _append_weekly_image_from_clipboard(self):
        clipboard = QApplication.clipboard()
        if not clipboard.mimeData().hasImage():
            self.show_toast("剪贴板中没有图片")
            return False
        pixmap = QPixmap.fromImage(clipboard.image())
        if pixmap.isNull():
            self.show_toast("无法读取剪贴板图片")
            return False
        rows = self.db.safe_fetchall(
            "SELECT slot_index FROM store_temp_images WHERE store_id=?",
            (self.store_id,),
        )
        slot_index = _next_store_image_slot(rows)
        byte_array = QByteArray()
        buffer = QBuffer(byte_array)
        buffer.open(QIODevice.WriteOnly)
        if not pixmap.save(buffer, "PNG"):
            self.show_toast("图片转换失败")
            return False
        self.db.safe_execute(
            """INSERT INTO store_temp_images (store_id, slot_index, image_data, created_time)
               VALUES (?, ?, ?, ?)""",
            (self.store_id, slot_index, bytes(byte_array), datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        )
        self.load_weekly_images()
        self.show_toast("已添加附带图片")
        return True

    def clear_all_weekly_images(self):
        rows = self.db.safe_fetchall(
            "SELECT 1 FROM store_temp_images WHERE store_id=? LIMIT 1",
            (self.store_id,),
        )
        if not rows:
            self.show_toast("当前没有附带图片")
            return
        reply = QMessageBox.question(
            self,
            "确认清空",
            "确定清空当前店铺的全部附带图片吗？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        self.db.safe_execute("DELETE FROM store_temp_images WHERE store_id=?", (self.store_id,))
        self.load_weekly_images()

    def show_weekly_image_viewer(self, slot_index):
        slots = [slot for slot, cell in sorted(self.weekly_image_cells.items()) if cell.has_image()]
        if int(slot_index) not in slots:
            return
        try:
            from manager.dialogs.product_spec import SpecImageViewerDialog
        except ImportError:
            from dialogs.product_spec import SpecImageViewerDialog

        class WeeklyImageViewerDialog(SpecImageViewerDialog):
            def __init__(viewer_self, pixmaps, current_index, parent):
                viewer_self.pixmaps = pixmaps
                viewer_self.current_index = current_index
                super().__init__(pixmaps[current_index], parent)
                nav = QHBoxLayout()
                viewer_self.btn_previous = QPushButton("← 上一张")
                viewer_self.btn_next = QPushButton("下一张 →")
                viewer_self.page_label = QLabel()
                viewer_self.btn_previous.clicked.connect(lambda: viewer_self.switch_image(-1))
                viewer_self.btn_next.clicked.connect(lambda: viewer_self.switch_image(1))
                nav.addWidget(viewer_self.btn_previous)
                nav.addStretch()
                nav.addWidget(viewer_self.page_label)
                nav.addStretch()
                nav.addWidget(viewer_self.btn_next)
                viewer_self.layout().insertLayout(0, nav)
                viewer_self.refresh_page_label()

            def refresh_page_label(viewer_self):
                total = len(viewer_self.pixmaps)
                viewer_self.page_label.setText(f"{viewer_self.current_index + 1} / {total}")
                viewer_self.btn_previous.setEnabled(total > 1)
                viewer_self.btn_next.setEnabled(total > 1)
                viewer_self.setWindowTitle(f"附带图片 {viewer_self.current_index + 1}/{total}")

            def switch_image(viewer_self, direction):
                if len(viewer_self.pixmaps) <= 1:
                    return
                viewer_self.current_index = (viewer_self.current_index + direction) % len(viewer_self.pixmaps)
                viewer_self.original_pixmap = viewer_self.pixmaps[viewer_self.current_index]
                viewer_self.scale_factor = 1.0
                viewer_self._refresh_image()
                QTimer.singleShot(0, viewer_self._center_image_view)
                viewer_self.refresh_page_label()

            def keyPressEvent(viewer_self, event):
                if event.key() == Qt.Key_Left:
                    viewer_self.switch_image(-1)
                    return
                if event.key() == Qt.Key_Right:
                    viewer_self.switch_image(1)
                    return
                super().keyPressEvent(event)

        pixmaps = [self.weekly_image_cells[slot].current_pixmap for slot in slots]
        dialog = WeeklyImageViewerDialog(pixmaps, slots.index(int(slot_index)), self)
        dialog.exec_()

    def toggle_reading_mode(self):
        """切换阅览模式 - 弹出放大版数据表格窗口"""
        self.is_reading_mode = not self.is_reading_mode
        if self.is_reading_mode:
            self.btn_reading_mode.setText("📖 退出阅览")
            self.btn_reading_mode.setStyleSheet("font-size: 11px; padding: 3px 5px; background-color: #e74c3c; color: white; border-radius: 3px;")
            self.show_toast("已打开放大版数据窗口")
            if self.large_dialog is None:
                self.large_dialog = LargeMarginDataDialog(self.store_name, self.store_id, self.db, self)
                self.large_dialog.setAttribute(Qt.WA_QuitOnClose, False)
                self.large_dialog.show()
            else:
                self.large_dialog.reload_data()
                self.large_dialog.show()
                self.large_dialog.activateWindow()
            self.is_reading_mode = False
            self.btn_reading_mode.setText("🔍 阅览模式")
            self.btn_reading_mode.setStyleSheet("font-size: 11px; padding: 3px 5px; background-color: #3498db; color: white; border-radius: 3px;")
        else:
            if self.large_dialog and self.large_dialog.isVisible():
                self.large_dialog.close()
            self.btn_reading_mode.setText("🔍 阅览模式")
            self.btn_reading_mode.setStyleSheet("font-size: 11px; padding: 3px 5px; background-color: #3498db; color: white; border-radius: 3px;")
            self.show_toast("已退出阅览模式")
            self.is_reading_mode = False

    def show_toast(self, message):
        """显示气泡提示（淡入淡出0.5秒，不透明度50%）"""
        self.toast_fade_out_animation.stop()
        self.toast_opacity_effect.setOpacity(0.5)
        self.toast_label.setText(message)
        self.toast_label.adjustSize()
        parent_pos = self.mapToGlobal(self.rect().bottomLeft())
        x = parent_pos.x() + (self.width() - self.toast_label.width()) // 2
        y = parent_pos.y() - 80
        self.toast_label.move(x, y)
        self.toast_label.show()
        QTimer.singleShot(500, self.fade_out_toast)

    def fade_out_toast(self):
        """淡出气泡提示"""
        self.toast_fade_out_animation.start()

    def get_pdd_browser_monitor(self):
        if self.pdd_browser_monitor is None:
            main_app = getattr(self, "main_app", None)
            if main_app is not None and hasattr(main_app, "_get_pdd_browser_monitor"):
                self.pdd_browser_monitor = main_app._get_pdd_browser_monitor()
                return self.pdd_browser_monitor
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            self.pdd_browser_monitor = PddBrowserMonitor(base_dir)
        return self.pdd_browser_monitor

    def open_pdd_merchant_test(self):
        """打开拼多多链接抓取主控窗口。"""
        try:
            monitor = self.get_pdd_browser_monitor()
            if self.pdd_link_control_dialog is None:
                self.pdd_link_control_dialog = PddLinkControlDialog(
                    self.db,
                    monitor,
                    default_store_id=self.store_id,
                    parent=self,
                )
                self.pdd_link_control_dialog.destroyed.connect(lambda _=None: setattr(self, "pdd_link_control_dialog", None))
            self.pdd_link_control_dialog.show()
            self.pdd_link_control_dialog.raise_()
            self.pdd_link_control_dialog.activateWindow()
        except Exception as e:
            QMessageBox.warning(self, "拼多多链接抓取", f"打开抓取窗口失败：{e}")

    def get_sys_id_by_user_id(self, user_id):
        """根据用户ID获取系统ID"""
        for sys_id, uid in self.sys_id_to_user_id.items():
            if uid == user_id:
                return sys_id
        return None

    def _get_spec_display_mode(self, setting_key):
        mode = self.db.get_setting(setting_key, "code")
        return mode if mode in ("code", "name") else "code"

    def _get_spec_display_text(self, user_product_id, spec_code, mode):
        if not spec_code:
            return "无"
        if mode != "name":
            return str(spec_code)
        sys_id = self.get_sys_id_by_user_id(user_product_id)
        if not sys_id:
            return f"未匹配规格\n{spec_code}"
        res = self.db.safe_fetchall(
            "SELECT spec_name FROM product_specs WHERE product_id=? AND spec_code=?",
            (sys_id, spec_code)
        )
        spec_name = res[0][0] if res and res[0][0] else ""
        return str(spec_name).strip() if str(spec_name).strip() else f"未匹配规格\n{spec_code}"

    def _get_valid_spec_codes(self, user_product_id):
        sys_id = self.get_sys_id_by_user_id(user_product_id)
        if not sys_id:
            return set()
        rows = self.db.safe_fetchall(
            "SELECT spec_code FROM product_specs WHERE product_id=? AND spec_code IS NOT NULL AND spec_code<>''",
            (sys_id,)
        )
        return {str(row[0]).strip() for row in rows if row and str(row[0]).strip()}

    def _set_spec_label_text(self, label, text, mode, alert=False):
        label.setWordWrap(True)
        label.setText(str(text or "无"))
        alert = alert or str(text or "").startswith("未匹配规格")
        color = "#e74c3c" if alert else ("black" if text and text != "无" else "#95a5a6")
        font_size = 13 if mode == "name" else 19
        weight = "font-weight: bold;" if alert else ""
        label.setStyleSheet(f"color: {color}; font-size: {font_size}px; {weight}")

    def _parse_amount_value(self, value):
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip()
        if not text or text.lower() in ("none", "nan"):
            return 0.0
        for token in ("￥", "¥", "元", ",", " "):
            text = text.replace(token, "")
        try:
            return float(text)
        except (TypeError, ValueError):
            return 0.0

    def get_main_spec(self, prod_id):
        """获取商品的主卖规格"""
        spec_counts = self.db.safe_fetchall(
            "SELECT spec_code, order_count FROM imported_orders WHERE product_id=?",
            (prod_id,)
        )
        if not spec_counts:
            return None, 0
        valid_spec_codes = self._get_valid_spec_codes(prod_id)
        if valid_spec_codes:
            spec_counts = [sc for sc in spec_counts if sc[0] and str(sc[0]).strip() in valid_spec_codes]
        total_orders = sum(sc[1] for sc in spec_counts if sc[1])
        if total_orders == 0:
            return None, 0
        max_spec = max(spec_counts, key=lambda x: x[1] if x[1] else 0)
        return max_spec[0] if max_spec[0] else None, max_spec[1] if max_spec[1] else 0

    def _snapshot_has_refunds(self, snapshot):
        """检查历史快照里是否包含退款数据。"""
        for data in snapshot.get("orders", {}).values():
            if isinstance(data, dict) and (data.get("refund_count") or 0) > 0:
                return True
        return False

    def _restore_orders_from_snapshot(self, snapshot, import_time=None):
        """按历史快照恢复当前店铺 imported_orders，不触发界面副作用。"""
        orders_data = snapshot.get("orders", {})
        if not orders_data:
            return False

        restore_time = import_time or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.db.safe_execute("DELETE FROM imported_orders WHERE store_id=?", (self.store_id,))

        for key, data in orders_data.items():
            if not isinstance(data, dict):
                continue
            parts = key.split("_", 1)
            if len(parts) < 2:
                continue

            user_product_id = parts[0]
            spec_code = parts[1]
            order_count = data.get("count", 0)
            refund_count = data.get("refund_count", 0)
            actual_amount = data.get("actual_amount", 0)
            dates = data.get("dates", [])
            earliest_date = min(dates) if dates else None
            latest_date = max(dates) if dates else None
            date_range = f"{earliest_date}~{latest_date}" if earliest_date and latest_date else None

            self.db.safe_execute("""
                INSERT OR REPLACE INTO imported_orders
                (store_id, product_id, spec_code, order_count, import_time, order_date, actual_amount, refund_count)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (self.store_id, user_product_id, spec_code, order_count,
                  restore_time, date_range, actual_amount, refund_count))

        return True

    def _restore_latest_import_history_if_needed(self):
        """当前订单缺少退款数时，自动从最新历史快照恢复。"""
        current = self.db.safe_fetchall("""
            SELECT COUNT(*), COALESCE(SUM(COALESCE(refund_count, 0)), 0)
            FROM imported_orders
            WHERE store_id=?
        """, (self.store_id,))
        current_count = current[0][0] if current else 0
        current_refunds = current[0][1] if current else 0
        if current_count > 0 and current_refunds > 0:
            return False

        history_records = self.db.safe_fetchall("""
            SELECT import_time, snapshot_data
            FROM import_history
            WHERE store_id=? AND snapshot_data IS NOT NULL AND snapshot_data != ''
            ORDER BY import_time DESC, id DESC
            LIMIT 1
        """, (self.store_id,))
        if not history_records:
            return False

        import_time, snapshot_data = history_records[0]
        try:
            snapshot = json.loads(snapshot_data)
        except Exception:
            return False

        if not self._snapshot_has_refunds(snapshot):
            return False

        return self._restore_orders_from_snapshot(snapshot, import_time)

    def _current_store_product_codes(self):
        rows = self.db.safe_fetchall(
            "SELECT name FROM products WHERE store_id=? AND COALESCE(is_archived, 0)=0",
            (self.store_id,)
        )
        return {str(row[0]).strip() for row in rows if row and str(row[0]).strip()}

    def _normalize_imported_order_store_ids(self):
        """把属于当前店铺商品ID的旧订单记录归到当前店铺，兼容跨电脑本地存档的 store_id 差异。"""
        product_codes = self._current_store_product_codes()
        if not product_codes:
            return 0

        safe_product_codes = []
        for product_code in product_codes:
            count_rows = self.db.safe_fetchall(
                "SELECT COUNT(*) FROM products WHERE name=? AND COALESCE(is_archived, 0)=0",
                (product_code,)
            )
            if count_rows and int(count_rows[0][0] or 0) == 1:
                safe_product_codes.append(product_code)
        if not safe_product_codes:
            return 0

        fixed_count = 0
        for product_code in safe_product_codes:
            rows = self.db.safe_fetchall(
                """SELECT id, product_id, spec_code, order_count, import_time, order_date, actual_amount, refund_count
                   FROM imported_orders
                   WHERE product_id=? AND store_id<>?""",
                (product_code, self.store_id)
            )
            for row in rows:
                order_id, prod_id, spec_code, order_count, import_time, order_date, actual_amount, refund_count = row
                existing = self.db.safe_fetchall(
                    """SELECT id, order_count, actual_amount, refund_count, order_date
                       FROM imported_orders
                       WHERE store_id=? AND product_id=? AND spec_code=?""",
                    (self.store_id, prod_id, spec_code)
                )
                if existing:
                    existing_id, old_count, old_amount, old_refund, old_date = existing[0]
                    merged_dates = []
                    for value in (old_date, order_date):
                        if value:
                            for part in str(value).split("~"):
                                part = part.strip()
                                if part and part not in merged_dates:
                                    merged_dates.append(part)
                    merged_date = "~".join(merged_dates) if merged_dates else None
                    self.db.safe_execute(
                        """UPDATE imported_orders
                           SET order_count=?, actual_amount=?, refund_count=?, order_date=?, import_time=?
                           WHERE id=?""",
                        (
                            int(old_count or 0) + int(order_count or 0),
                            float(old_amount or 0) + float(actual_amount or 0),
                            int(old_refund or 0) + int(refund_count or 0),
                            merged_date,
                            import_time,
                            existing_id
                        )
                    )
                    self.db.safe_execute("DELETE FROM imported_orders WHERE id=?", (order_id,))
                else:
                    self.db.safe_execute(
                        "UPDATE imported_orders SET store_id=? WHERE id=?",
                        (self.store_id, order_id)
                    )
                fixed_count += 1
        if fixed_count:
            try:
                self.db.conn.commit()
            except Exception:
                pass
            print(f"[STORE_MARGIN] 已修正 {fixed_count} 条订单记录的店铺归属")
        return fixed_count

    def get_user_id_by_sys_id(self, sys_id):
        """根据系统ID获取用户ID"""
        return self.sys_id_to_user_id.get(sys_id)

    def _button_tooltip_for(self, button):
        button_tips = {
            getattr(self, "btn_last_week", None): "将过往数据分析的日期范围快速设置为最近七天。",
            getattr(self, "btn_input_data", None): "打开录入数据窗口，手动填写当前日期范围的实发订单、实发金额、毛利润、退款和推广等店铺毛利数据。",
            getattr(self, "btn_import_data", None): "导入过往店铺算账表格，写入顶部数据周期板块。ERP模式按原表导入；表格模式会读取订单明细并自动预填可计算的数据。",
            getattr(self, "btn_export_excel", None): "导出顶部过往算账数据，并附带店铺商品权重和商品规格售卖情况。",
            getattr(self, "btn_reading_mode", None): "打开放大阅览窗口，便于单独查看过往数据分析表。",
            getattr(self, "btn_weekly_images", None): "查看或追加当前店铺的本周附带图片；展开后将鼠标移入图片区域，可按 Ctrl+V 直接粘贴图片。",
            getattr(self, "btn_clear_weekly_images", None): "清空当前店铺已经保存的全部本周附带图片。",
            getattr(self, "btn_profit_calc", None): "重新计算当前店铺商品权重、综合毛利和相关利润指标。",
            getattr(self, "btn_import_orders", None): "导入订单规格数据，用于计算店铺商品权重、主卖规格、单量、销售额和退款结构。",
            getattr(self, "btn_history", None): "查看当前店铺的订单导入历史，可恢复或删除历史导入记录。",
            getattr(self, "btn_ai_report", None): "打开AI店铺周报窗口，基于近两周财务、售卖结构和操作记录生成文字报告。",
            getattr(self, "btn_promotion_data", None): "打开推广数据分析窗口，导入和查看当前店铺各链接的每日真实推广数据。",
            getattr(self, "btn_display_settings", None): "设置主卖规格和退款占比最多规格的显示方式，支持规格编码或规格名称并全局生效。",
            getattr(self, "btn_save_store_settings", None): "保存当前店铺的综合设置，包括全站托管投产比和店铺通用满减梯度。",
            getattr(self, "btn_save", None): "保存当前店铺商品权重和综合毛利相关改动，并记录必要的店铺操作记录。",
            getattr(self, "btn_close", None): "关闭店铺毛利管理窗口。",
        }
        tip = button_tips.get(button)
        if tip:
            return tip
        text = re.sub(r"\s+", "", button.text() or "")
        if "查看商品" in text:
            return "打开该链接的规格与毛利管理窗口，查看和编辑规格、成本、价格、活动和投产信息。"
        if text == "关闭":
            return "关闭当前窗口。"
        if text == "保存":
            return "保存当前窗口中的设置或数据。"
        if text == "取消":
            return "取消本次操作并关闭窗口。"
        if text == "确定" or text == "确认":
            return "确认当前选择并继续执行。"
        if "复制" in text:
            return "复制当前窗口展示的内容到剪贴板。"
        return ""

    def _apply_store_margin_button_tooltips(self):
        for button in self.findChildren(QPushButton):
            tip = self._button_tooltip_for(button)
            if not tip:
                continue
            button.setToolTip("")
            button.setProperty("store_margin_tooltip_text", tip)
            if not button.property("store_margin_tooltip_filter_installed"):
                button.installEventFilter(self)
                button.setProperty("store_margin_tooltip_filter_installed", True)

    def _start_button_tooltip_timer(self, button, text):
        self._button_tooltip_timer.stop()
        QToolTip.hideText()
        self._button_tooltip_widget = button
        self._button_tooltip_text = text
        self._button_tooltip_timer.start(1000)

    def _stop_button_tooltip_timer(self, button=None):
        if button is None or button == self._button_tooltip_widget:
            self._button_tooltip_timer.stop()
            QToolTip.hideText()
            self._button_tooltip_widget = None
            self._button_tooltip_text = ""

    def _show_delayed_button_tooltip(self):
        button = self._button_tooltip_widget
        if not button or not button.isVisible() or not button.underMouse():
            return
        text = self._button_tooltip_text or button.property("store_margin_tooltip_text") or ""
        if not text:
            return
        pos = button.mapToGlobal(button.rect().bottomLeft())
        QToolTip.showText(pos, text, button)

    def eventFilter(self, obj, event):
        if getattr(self, "_disposing", False):
            return False
        panel = getattr(self, "weekly_images_panel", None)
        button = getattr(self, "btn_weekly_images", None)
        if panel is not None and panel.isVisible():
            if event.type() == QEvent.KeyPress and event.modifiers() & Qt.ControlModifier and event.key() == Qt.Key_V:
                if panel.rect().contains(panel.mapFromGlobal(QCursor.pos())):
                    self._append_weekly_image_from_clipboard()
                    return True
            if event.type() == QEvent.MouseButtonPress and hasattr(event, "globalPos"):
                global_pos = event.globalPos()
                inside_panel = panel.rect().contains(panel.mapFromGlobal(global_pos))
                inside_button = button and button.rect().contains(button.mapFromGlobal(global_pos))
                if not inside_panel and not inside_button:
                    panel.hide()
        if isinstance(obj, QPushButton):
            tooltip_text = obj.property("store_margin_tooltip_text")
            if tooltip_text:
                if event.type() == QEvent.Enter:
                    self._start_button_tooltip_timer(obj, tooltip_text)
                elif event.type() in (QEvent.Leave, QEvent.MouseButtonPress, QEvent.Hide):
                    self._stop_button_tooltip_timer(obj)
        if event.type() == QEvent.MouseButtonDblClick:
            if isinstance(obj, QLineEdit):
                row = obj.property("row")
                prod_id = obj.property("prod_id")
                if row is not None and prod_id is not None:
                    return False
        return False

    def on_weight_changed(self, user_id, text):
        if user_id not in self.product_weights:
            return
        sender = self.sender()
        try:
            new_weight = float(text) if text else 0
        except ValueError:
            new_weight = 0
        new_weight = max(0, min(100, new_weight))
        total_locked = sum(
            data.get("weight", 0) for uid, data in self.product_weights.items()
            if uid != user_id and data.get("locked", 0)
        )
        max_allowed = 100 - total_locked
        if new_weight > max_allowed:
            new_weight = max_allowed
        self.product_weights[user_id]["weight"] = new_weight
        self.db.safe_execute("UPDATE stores SET weight_synced=0 WHERE id=?", (self.store_id,))
        self.main_app.refresh_store_weight_sync_flag(self.store_id)
        self.rebalance_unlocked_weights(user_id)
        self.update_weight_inputs()
        self.calculate_total_margin()

    def on_weight_editing_finished(self, user_id):
        if user_id not in self.product_weights:
            return
        sys_id = self.get_sys_id_by_user_id(user_id)
        if not sys_id:
            return
        new_weight = self.product_weights[user_id].get("weight", 0)
        self.db.safe_execute("UPDATE products SET store_weight=? WHERE id=?", (new_weight, sys_id))
        self.update_weight_inputs()

    def rebalance_unlocked_weights(self, changed_user_id):
        total_locked = sum(
            data.get("weight", 0) for data in self.product_weights.values() if data.get("locked", 0)
        )
        changed_weight = self.product_weights[changed_user_id]["weight"]
        remaining = max(0, 100 - total_locked - changed_weight)
        unlocked_prods = [
            uid for uid, data in self.product_weights.items()
            if uid != changed_user_id and not data.get("locked", 0)
        ]
        if not unlocked_prods:
            return
        avg_weight = remaining / len(unlocked_prods)
        for uid in unlocked_prods:
            self.product_weights[uid]["weight"] = avg_weight
            sys_id = self.get_sys_id_by_user_id(uid)
            if sys_id:
                self.db.safe_execute("UPDATE products SET store_weight=? WHERE id=?", (avg_weight, sys_id))

    def update_weight_inputs(self):
        for row in range(self.table.rowCount()):
            prod_id = self.table.item(row, 1).data(Qt.UserRole)
            if not prod_id or prod_id not in self.product_weights:
                continue
            cell_widget = self.table.cellWidget(row, 6)
            if not cell_widget:
                continue
            weight_input = cell_widget.findChild(QLineEdit)
            if weight_input:
                weight = self.product_weights[prod_id]["weight"]
                weight_str = str(int(weight)) if weight == int(weight) else f"{weight:.1f}"
                weight_input.blockSignals(True)
                weight_input.setText(weight_str)
                weight_input.blockSignals(False)

    def _format_roi_value(self, value):
        try:
            value = float(value or 0)
        except (TypeError, ValueError):
            value = 0.0
        return f"{value:.2f}" if value > 0 else "未设置"

    def _record_sitewide_roi_change(self, old_roi, new_roi):
        if abs(float(old_roi or 0) - float(new_roi or 0)) <= 0.0001:
            return
        try:
            now = datetime.now()
            time_str = now.strftime("%H:%M")
            log_text = f"【全站托管投产】{self._format_roi_value(old_roi)} → {self._format_roi_value(new_roi)}"
            records = self.db.get_store_record(self.store_id, now.year, now.month, now.day)
            records.append({"time": time_str, "text": log_text})
            self.db.save_store_record(self.store_id, now.year, now.month, now.day, records)
            if hasattr(self.main_app, "refresh_store_cards"):
                QTimer.singleShot(0, lambda: self.main_app.refresh_store_cards(self.store_id))
        except Exception as e:
            print(f"保存全站托管投产记录失败: {e}")

    def _format_store_discount_rules_for_input(self):
        try:
            rules = self.db.get_store_discount_rules(self.store_id)
        except Exception:
            rules = []
        parts = []
        for rule in rules:
            threshold = float(rule.get("threshold") or 0)
            discount = float(rule.get("discount") or 0)
            threshold_text = f"{threshold:.2f}".rstrip("0").rstrip(".")
            discount_text = f"{discount:.2f}".rstrip("0").rstrip(".")
            parts.append(f"满{threshold_text}减{discount_text}")
        return "；".join(parts)

    def _parse_store_discount_rules_input(self):
        text = self.store_discount_rules_input.text().strip() if hasattr(self, "store_discount_rules_input") else ""
        if not text:
            return []
        matches = re.findall(r"满\s*([0-9]+(?:\.[0-9]+)?)\s*减\s*([0-9]+(?:\.[0-9]+)?)", text)
        if not matches:
            raise ValueError("店铺满减格式不正确，请按“满100减10；满200减30”填写。")
        rules = []
        for threshold_text, discount_text in matches:
            threshold = float(threshold_text)
            discount = float(discount_text)
            if threshold <= 0 or discount <= 0:
                raise ValueError("满减门槛和减免金额必须大于0。")
            if discount >= threshold:
                raise ValueError("满减金额不能大于或等于门槛金额。")
            rules.append({"threshold": threshold, "discount": discount})
        return rules

    def _record_store_discount_rules_change(self, old_text, new_text):
        if old_text == new_text:
            return
        try:
            now = datetime.now()
            time_str = now.strftime("%H:%M")
            log_text = f"【店铺满减】{old_text or '未设置'} → {new_text or '未设置'}"
            records = self.db.get_store_record(self.store_id, now.year, now.month, now.day)
            records.append({"time": time_str, "text": log_text})
            self.db.save_store_record(self.store_id, now.year, now.month, now.day, records)
            if hasattr(self.main_app, "refresh_store_cards"):
                QTimer.singleShot(0, lambda: self.main_app.refresh_store_cards(self.store_id))
        except Exception as e:
            print(f"保存店铺满减记录失败: {e}")

    def save_sitewide_roi(self, show_toast=True):
        try:
            sitewide_roi = float(self.sitewide_roi_input.text().strip()) if self.sitewide_roi_input.text().strip() else 0.0
            if sitewide_roi < 0:
                sitewide_roi = 0.0
            old_rows = self.db.safe_fetchall("SELECT sitewide_roi FROM stores WHERE id=?", (self.store_id,))
            old_roi = float(old_rows[0][0] or 0) if old_rows else 0.0
            self.db.safe_execute("UPDATE stores SET sitewide_roi=? WHERE id=?", (sitewide_roi, self.store_id))
            self._record_sitewide_roi_change(old_roi, sitewide_roi)
            if show_toast and hasattr(self.main_app, "show_toast"):
                self.main_app.show_toast("全站托管投产比已保存")
            return True
        except ValueError:
            if show_toast and hasattr(self.main_app, "show_toast"):
                self.main_app.show_toast("全站托管投产比格式不正确")
            return False

    def save_store_discount_rules(self, show_toast=True):
        try:
            old_text = self.db.format_store_discount_rules(self.store_id)
            rules = self._parse_store_discount_rules_input()
            clean_rules = self.db.save_store_discount_rules(self.store_id, rules)
            self.store_discount_rules_input.setText(self._format_store_discount_rules_for_input())
            new_text = self.db.format_store_discount_rules(self.store_id)
            self._record_store_discount_rules_change(old_text, new_text)
            if show_toast and hasattr(self.main_app, "show_toast"):
                self.main_app.show_toast("店铺满减梯度已保存")
            return True
        except ValueError as e:
            QMessageBox.warning(self, "店铺满减格式错误", str(e))
            return False
        except Exception as e:
            QMessageBox.warning(self, "保存失败", f"保存店铺满减失败：{e}")
            return False

    def save_store_comprehensive_settings(self, show_toast=True):
        if not self.save_sitewide_roi(show_toast=False):
            if show_toast and hasattr(self.main_app, "show_toast"):
                self.main_app.show_toast("全站托管投产比格式不正确")
            return False
        if not self.save_store_discount_rules(show_toast=False):
            return False
        if show_toast and hasattr(self.main_app, "show_toast"):
            self.main_app.show_toast("店铺综合设置已保存")
        self.calculate_total_margin()
        return True

    def save_weights(self):
        if append_event:
            append_event(f"store_margin:save_weights:start store_id={self.store_id}")
        old_margin = self.calculate_total_margin()
        if not self.save_store_comprehensive_settings(show_toast=False):
            return
        updates = []
        for row in range(self.table.rowCount()):
            prod_id = self.table.item(row, 1).data(Qt.UserRole)
            if not prod_id or prod_id not in self.product_weights:
                continue
            sys_id = self.get_sys_id_by_user_id(prod_id)
            if not sys_id:
                continue
            weight = self.product_weights[prod_id]["weight"]
            is_locked = self.product_weights[prod_id]["locked"]
            updates.append((weight, is_locked, sys_id))
        self.db.save_product_weights(updates)
        if append_event:
            append_event(f"store_margin:save_weights:db_done store_id={self.store_id} count={len(updates)}")
        saved_count = len(updates)
        new_margin = self.calculate_total_margin()
        if old_margin is not None and new_margin is not None and abs(old_margin - new_margin) > 0.01:
            self.save_margin_log(old_margin, new_margin)
        self.main_app.show_toast(f"✅ 已保存 {saved_count} 项权重数据")
        callback = self.save_callback
        store_id = self.store_id
        if callback:
            callback(store_id, new_margin)
        self.close()

    def save_margin_log(self, old_margin, new_margin):
        try:
            time_str = datetime.now().strftime("%H:%M")
            change = new_margin - old_margin
            change_str = f"+{change:.1f}%" if change > 0 else f"{change:.1f}%"
            log_text = f"【权重保存】综合毛利: {old_margin:.1f}% → {new_margin:.1f}% ({change_str})"
            year = datetime.now().year
            month = datetime.now().month
            day = datetime.now().day
            records = self.db.get_store_record(self.store_id, year, month, day)
            records.append({"time": time_str, "text": log_text})
            self.db.save_store_record(self.store_id, year, month, day, records)
        except Exception as e:
            print(f"保存毛利日志失败: {e}")

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        # ====== 店铺综合设置板块 ======
        settings_widget = QWidget()
        self.settings_widget = settings_widget
        settings_widget.setStyleSheet("""
            QWidget {
                background-color: #f8fbff;
                border: 1px solid #d6e4f0;
                border-radius: 6px;
            }
            QLabel {
                border: none;
                background: transparent;
            }
            QLineEdit {
                background: white;
            }
        """)
        settings_layout = QHBoxLayout(settings_widget)
        settings_layout.setContentsMargins(10, 8, 10, 8)
        settings_layout.setSpacing(8)

        settings_title = QLabel("店铺综合设置")
        settings_title.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50;")
        settings_layout.addWidget(settings_title)

        settings_layout.addWidget(QLabel("全站托管投产:"))
        self.sitewide_roi_input = QLineEdit()
        self.sitewide_roi_input.setFixedWidth(80)
        self.sitewide_roi_input.setPlaceholderText("投产比")
        self.sitewide_roi_input.setStyleSheet("padding: 5px; border: 1px solid #8e44ad; border-radius: 4px; font-weight: bold;")
        try:
            sitewide_rows = self.db.safe_fetchall("SELECT sitewide_roi FROM stores WHERE id=?", (self.store_id,))
            sitewide_roi = float(sitewide_rows[0][0] or 0) if sitewide_rows else 0.0
            self.sitewide_roi_input.setText(f"{sitewide_roi:.2f}" if sitewide_roi > 0 else "")
        except Exception:
            pass
        self.sitewide_roi_input.returnPressed.connect(self.save_store_comprehensive_settings)
        settings_layout.addWidget(self.sitewide_roi_input)

        settings_layout.addWidget(QLabel("店铺满减梯度:"))
        self.store_discount_rules_input = QLineEdit()
        self.store_discount_rules_input.setMinimumWidth(260)
        self.store_discount_rules_input.setPlaceholderText("例如：满100减10；满200减30")
        self.store_discount_rules_input.setStyleSheet("padding: 5px; border: 1px solid #27ae60; border-radius: 4px;")
        self.store_discount_rules_input.setText(self._format_store_discount_rules_for_input())
        self.store_discount_rules_input.returnPressed.connect(self.save_store_comprehensive_settings)
        settings_layout.addWidget(self.store_discount_rules_input, 1)

        self.btn_save_store_settings = QPushButton("保存设置")
        self.btn_save_store_settings.setStyleSheet("""
            QPushButton {
                background-color: #2c7be5;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #1a68d1; }
        """)
        self.btn_save_store_settings.clicked.connect(self.save_store_comprehensive_settings)
        settings_layout.addWidget(self.btn_save_store_settings)
        layout.addWidget(settings_widget)

        # ====== 板块1: 过往数据分析板块 ======
        historical_widget = QWidget()
        historical_widget.setStyleSheet("border: 1px solid #dee2e6; border-radius: 8px;")
        historical_layout = QVBoxLayout(historical_widget)
        historical_layout.setContentsMargins(0, 0, 0, 0)

        # 日期选择行
        date_row = QWidget()
        date_layout = QHBoxLayout(date_row)
        date_layout.setContentsMargins(0, 0, 0, 0)

        date_label = QLabel("数据周期:")
        date_label.setStyleSheet("font-size: 12px; color: #666; padding: 0 5px;")

        # 使用日期选择器
        from PyQt5.QtWidgets import QDateEdit
        from PyQt5.QtCore import QDate

        self.date_start_input = QDateEdit()
        self.date_start_input.setCalendarPopup(True)
        self.date_start_input.setDate(QDate.currentDate().addDays(-7))
        self.date_start_input.setDisplayFormat("yyyy-MM-dd")
        self.date_start_input.setFixedWidth(100)
        self.date_start_input.setStyleSheet("font-size: 11px; padding: 2px;")

        self.date_separator = QLabel("~")
        self.date_separator.setStyleSheet("font-size: 12px; color: #666; padding: 0 5px;")

        self.date_end_input = QDateEdit()
        self.date_end_input.setCalendarPopup(True)
        self.date_end_input.setDate(QDate.currentDate().addDays(-1))
        self.date_end_input.setDisplayFormat("yyyy-MM-dd")
        self.date_end_input.setFixedWidth(100)
        self.date_end_input.setStyleSheet("font-size: 11px; padding: 2px;")

        # 快捷按钮
        self.btn_last_week = QPushButton("📅 近七天")
        self.btn_last_week.setFixedWidth(80)
        self.btn_last_week.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                padding: 6px 10px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
            QPushButton:pressed {
                background-color: #6c7a7d;
            }
        """)
        self.btn_last_week.clicked.connect(self.set_last_week)

        self.btn_input_data = QPushButton("📝 录入数据")
        self.btn_input_data.setFixedWidth(90)
        self.btn_input_data.setStyleSheet("""
            QPushButton {
                background-color: #4f86ad;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                padding: 6px 10px;
            }
            QPushButton:hover {
                background-color: #40769b;
            }
            QPushButton:pressed {
                background-color: #356687;
            }
        """)
        self.btn_input_data.clicked.connect(self.open_input_data_dialog)

        self.btn_import_data = QPushButton("📂 导入数据")
        self.btn_import_data.setFixedWidth(90)
        self.btn_import_data.setStyleSheet("""
            QPushButton {
                background-color: #3f739e;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                padding: 6px 10px;
            }
            QPushButton:hover {
                background-color: #356388;
            }
            QPushButton:pressed {
                background-color: #2c5474;
            }
        """)
        self.btn_import_data.clicked.connect(self.import_data)

        self.btn_export_excel = QPushButton("导出Excel")
        self.btn_export_excel.setFixedWidth(90)
        self.btn_export_excel.setStyleSheet("""
            QPushButton {
                background-color: #5b91b8;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                padding: 6px 10px;
            }
            QPushButton:hover {
                background-color: #4c80a6;
            }
            QPushButton:pressed {
                background-color: #406f92;
            }
        """)
        self.btn_export_excel.clicked.connect(self._export_margin_excel_from_button)

        self.combo_margin_data_mode = QComboBox()
        self.combo_margin_data_mode.addItem("ERP模式", "erp")
        self.combo_margin_data_mode.addItem("表格模式", "table")
        self.combo_margin_data_mode.setFixedWidth(92)
        self.combo_margin_data_mode.setStyleSheet("font-size: 12px; padding: 2px;")
        current_mode = self.db.get_setting("store_margin_data_mode", "erp")
        self.combo_margin_data_mode.setCurrentIndex(1 if current_mode == "table" else 0)
        self.combo_margin_data_mode.currentIndexChanged.connect(self.on_margin_data_mode_changed)

        self.btn_reading_mode = QPushButton("🔍 阅览模式")
        self.btn_reading_mode.setFixedWidth(90)
        self.btn_reading_mode.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                padding: 6px 10px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #2471a3;
            }
        """)
        self.btn_reading_mode.clicked.connect(self.toggle_reading_mode)

        self.btn_weekly_images = QPushButton("本周附带图片")
        self.btn_weekly_images.setFixedWidth(112)
        self.btn_weekly_images.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                padding: 6px 10px;
            }
            QPushButton:hover { background-color: #138496; }
        """)
        self.btn_weekly_images.clicked.connect(self.toggle_weekly_images_panel)

        self.lbl_current_history = QLabel("📍 当前: 暂无数据")
        self.lbl_current_history.setStyleSheet("""
            QLabel {
                color: #3498db;
                font-size: 12px;
                font-weight: bold;
                padding: 6px 12px;
                background-color: #e8f4fc;
                border-radius: 4px;
            }
        """)

        date_layout.addWidget(date_label)
        date_layout.addWidget(self.date_start_input)
        date_layout.addWidget(self.date_separator)
        date_layout.addWidget(self.date_end_input)
        date_layout.addWidget(self.btn_last_week)
        date_layout.addWidget(self.btn_input_data)
        date_layout.addWidget(self.btn_import_data)
        date_layout.addWidget(self.btn_export_excel)
        date_layout.addWidget(self.combo_margin_data_mode)
        date_layout.addWidget(self.btn_reading_mode)
        date_layout.addWidget(self.lbl_current_history)
        date_layout.addStretch()
        date_layout.addWidget(self.btn_weekly_images)

        historical_layout.addWidget(date_row)
        self._build_weekly_images_panel()

        # 子板块C: 手动录入数据表格
        self.margin_data_table = QTableWidget()
        self.margin_data_table.setColumnCount(20)  # 日期 + 19个指标
        # 设置表头
        self.margin_data_table.setHorizontalHeaderLabels([
            "日期", "实发订单", "实发金额", "毛利润", "毛利率", "退款金额", "金额退款率",
            "退款订单", "订单退款率", "件单价", "推广费", "推广占比",
            "技术服务费", "扣款", "其他服务", "其他", "净利润",
            "净利率", "单笔利润", "日盈亏"
        ])
        self.margin_data_table.verticalHeader().setVisible(False)
        self.margin_data_table.setShowGrid(True)
        self.margin_data_table.setGridStyle(Qt.SolidLine)
        self.margin_data_table.setAlternatingRowColors(False)
        # Excel风格标准表格
        self.margin_data_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #cccccc;
                font-size: 14px;
                border: 1px solid #cccccc;
                border-radius: 0px;
                margin: 0px;
                background-color: white;
            }
            QTableWidget::item {
                padding: 0px;
                text-align: center;
                border: 1px solid #cccccc;
                font-size: 14px;
            }
            QTableWidget::item:selected {
                background-color: #e6f3ff;
                color: black;
                outline: none;
            }
            QHeaderView {
                border: none;
                margin: 0px;
            }
            QHeaderView::section {
                background-color: #e0e0e0;
                padding: 0px;
                margin: 0px;
                border: none;
                border-left: 1px solid #cccccc;
                border-bottom: 1px solid #cccccc;
                border-right: 1px solid #cccccc;
                font-size: 14px;
                font-weight: bold;
                min-height: 45px;
            }
            QTableCornerButton::section {
                border-radius: 0px;
            }
        """)
        # 设置表格字体大小
        from PyQt5.QtGui import QFont
        table_font = QFont()
        table_font.setPointSize(14)
        self.margin_data_table.setFont(table_font)

        header = self.margin_data_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setMinimumSectionSize(50)
        self.margin_data_table._initial_width = None
        self.margin_data_table.setMinimumHeight(60)
        self.margin_data_table.setMaximumHeight(100)
        self.margin_data_table.verticalHeader().setVisible(False)
        self.margin_data_table.setShowGrid(True)
        self.margin_data_table.setGridStyle(Qt.SolidLine)
        
        historical_layout.addWidget(self.margin_data_table)

        # 周环比对比表格
        self.week_table = QTableWidget()
        self.week_table.setColumnCount(20)
        self.week_table.setRowCount(1)
        self.week_table.verticalHeader().setVisible(False)
        self.week_table.horizontalHeader().setVisible(False)
        self.week_table.setShowGrid(True)
        self.week_table.setGridStyle(Qt.SolidLine)
        self.week_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #cccccc;
                font-size: 14px;
                border: 1px solid #cccccc;
                border-radius: 0px;
                background-color: white;
            }
            QTableWidget::item {
                padding: 0px;
                text-align: center;
                border: 1px solid #cccccc;
                font-size: 14px;
            }
            QHeaderView::section {
                background-color: #d0d0d0;
                padding: 0px;
                border: 1px solid #cccccc;
                font-size: 14px;
                font-weight: bold;
                min-height: 30px;
            }
        """)
        week_header = self.week_table.horizontalHeader()
        week_header.setSectionResizeMode(QHeaderView.Stretch)
        self.week_table.setMaximumHeight(40)
        
        historical_layout.addWidget(self.week_table)

        # 板块标题：商品规格毛利权重
        section_title_2 = QLabel("📦 订单规格毛利权重")
        section_title_2.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50; padding: 5px 10px; border-radius: 4px;")
        historical_layout.addWidget(section_title_2)

        # 毛利明细表格
        self.table = QTableWidget()
        self.table.setColumnCount(14)
        self.table.setHorizontalHeaderLabels(["图片", "商品 ID", "商品标题", "综合成本", "客单价", "毛利", "权重 (%)", "权重对比\n(较上周)", "单量", "单量对比\n(较上周)", "销售额", "主卖规格", "退款率", "退款占比\n最多规格"])
        self.table.setAlternatingRowColors(False)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setFocusPolicy(Qt.NoFocus)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        self.table.cellChanged.connect(self.on_cell_changed)
        self.table.cellClicked.connect(self.on_cell_clicked)
        self.table.cellDoubleClicked.connect(self.on_cell_double_clicked)
        
        # 设置列宽自适应填充
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)
        header.setMinimumSectionSize(50)
        header.setSortIndicatorShown(True)
        header.setSortIndicator(-1, Qt.AscendingOrder)
        self._order_sort_column = -1
        self._order_sort_order = Qt.AscendingOrder
        header.sectionClicked.connect(self.sort_order_table_by_column)
        header.setSectionResizeMode(0, QHeaderView.Fixed)
        self.table.setColumnWidth(0, 72)
        # 商品标题列固定200像素
        header.setSectionResizeMode(2, QHeaderView.Fixed)
        self.table.setColumnWidth(2, 200)
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #dfe3e6;
                border: 1px solid #cfd4d8;
                border-radius: 0px;
                background-color: white;
            }
            QTableWidget::item {
                padding: 0px;
                text-align: center;
                border: 0px;
                background-color: white;
            }
            QTableWidget::item:focus {
                border: 0px;
                outline: none;
            }
            QTableWidget QLabel,
            QTableWidget QLineEdit {
                background: transparent;
                border: none;
                border-radius: 0px;
                padding: 0px;
                margin: 0px;
            }
            QWidget#plainTableCell {
                background: transparent;
                border: none;
                border-radius: 0px;
            }
            QTableWidget::item:selected {
                background-color: #e6f3ff;
                color: black;
                outline: none;
            }
            QTableWidget::item:hover {
                background-color: #d4edda;
            }
            QHeaderView {
                border: none;
                background-color: white;
            }
            QHeaderView::section {
                background-color: white;
                padding: 0px;
                margin: 0px;
                border: 0px;
                border-right: 1px solid #dfe3e6;
                border-bottom: 1px solid #cfd4d8;
                font-weight: bold;
                min-height: 35px;
                text-align: center;
            }
        """)
        historical_layout.addWidget(self.table)

        btn_widget = QWidget()
        btn_layout = QHBoxLayout(btn_widget)
        self.btn_profit_calc = QPushButton("🧮 计算利润")
        self.btn_profit_calc.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
            QPushButton:pressed {
                background-color: #7d3c98;
            }
        """)
        self.btn_profit_calc.clicked.connect(self.open_profit_calculator)
        self.btn_import_orders = QPushButton("📥 导入订单")
        self.btn_import_orders.setStyleSheet("""
            QPushButton {
                background-color: #5b63a9;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #4d5597;
            }
            QPushButton:pressed {
                background-color: #424986;
            }
        """)
        self.btn_import_orders.clicked.connect(self.import_orders)
        
        # 历史记录按钮
        self.btn_history = QPushButton("📜 全部记录")
        self.btn_history.setStyleSheet("""
            QPushButton {
                background-color: #6b73b9;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #5c64a7;
            }
            QPushButton:pressed {
                background-color: #4e5695;
            }
        """)
        self.btn_history.clicked.connect(self.show_import_history)

        self.btn_ai_report = QPushButton("AI生成报告")
        self.btn_ai_report.setStyleSheet("""
            QPushButton {
                background-color: #16a085;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #138d75;
            }
            QPushButton:pressed {
                background-color: #117a65;
            }
        """)
        self.btn_ai_report.clicked.connect(self.open_ai_report_dialog)

        self.btn_promotion_data = QPushButton("推广数据分析")
        self.btn_promotion_data.setStyleSheet("""
            QPushButton {
                background-color: #d35400;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #ba4a00;
            }
        """)
        self.btn_promotion_data.clicked.connect(self.open_promotion_data_dialog)

        self.btn_display_settings = QPushButton("设置")
        self.btn_display_settings.setStyleSheet("""
            QPushButton {
                background-color: #7f8c8d;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                padding: 6px 10px;
            }
            QPushButton:hover {
                background-color: #707b7c;
            }
        """)
        self.btn_display_settings.clicked.connect(self.open_store_margin_display_settings)

        self.btn_pdd_merchant_test = QPushButton("抓取链接")
        self.btn_pdd_merchant_test.setToolTip("打开拼多多链接抓取窗口，可抓取添加编码界面、价格管理和推广状态")
        self.btn_pdd_merchant_test.setStyleSheet("""
            QPushButton {
                background-color: #f39c12;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                padding: 6px 10px;
            }
            QPushButton:hover {
                background-color: #d68910;
            }
        """)
        self.btn_pdd_merchant_test.clicked.connect(self.open_pdd_merchant_test)

        self.lbl_total_margin = QLabel("综合毛利: 0.00%")
        self.lbl_total_margin.setStyleSheet(
            "font-size: 14px; font-weight: bold; color: #e74c3c; background-color: #fdeaa8; padding: 6px 12px; border-radius: 6px;"
        )

        self.lbl_total_orders = QLabel("总单量: 0")
        self.lbl_total_orders.setStyleSheet(
            "font-size: 14px; font-weight: bold; color: #3498db; background-color: #e8f4fc; padding: 6px 12px; border-radius: 6px;"
        )

        self.lbl_order_range = QLabel("当前订单时间范围: --")
        self.lbl_order_range.setStyleSheet(
            "font-size: 14px; font-weight: bold; color: #8e44ad; background-color: #f5eef8; padding: 6px 12px; border-radius: 6px;"
        )

        self.btn_save = QPushButton("💾 保存")
        self.btn_save.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
            QPushButton:pressed {
                background-color: #004085;
            }
        """)
        self.btn_save.clicked.connect(self.save_weights)
        self.btn_close = QPushButton("关闭")
        self.btn_close.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
            QPushButton:pressed {
                background-color: #a93226;
            }
        """)
        self.btn_close.clicked.connect(self.accept)
        btn_layout.addWidget(self.btn_profit_calc)
        btn_layout.addWidget(self.btn_import_orders)
        btn_layout.addWidget(self.btn_history)
        btn_layout.addWidget(self.btn_ai_report)
        btn_layout.addWidget(self.btn_promotion_data)
        btn_layout.addWidget(self.btn_pdd_merchant_test)
        btn_layout.addWidget(self.btn_display_settings)
        btn_layout.addSpacing(10)
        btn_layout.addWidget(self.lbl_total_margin)
        btn_layout.addWidget(self.lbl_total_orders)
        btn_layout.addWidget(self.lbl_order_range)
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_save)
        btn_layout.addWidget(self.btn_close)
        historical_layout.addWidget(btn_widget)

        layout.addWidget(historical_widget)

    def resizeEvent(self, event):
        """窗口大小改变时同步两表列宽"""
        super().resizeEvent(event)
        self.sync_table_widths()
        if getattr(self, "weekly_images_panel", None) is not None and self.weekly_images_panel.isVisible():
            self._position_weekly_images_panel()

    def closeEvent(self, event):
        self._prepare_for_account_switch()
        super().closeEvent(event)

    def _prepare_for_account_switch(self):
        self._disposing = True
        app = QApplication.instance()
        if app is not None and getattr(self, "_weekly_app_filter_installed", False):
            app.removeEventFilter(self)
            self._weekly_app_filter_installed = False
        for timer_name in ("_button_tooltip_timer",):
            timer = getattr(self, timer_name, None)
            if timer is not None:
                timer.stop()
        anim = getattr(self, "toast_fade_out_animation", None)
        if anim is not None:
            anim.stop()
        for widget in [self, *self.findChildren(QWidget)]:
            try:
                widget.removeEventFilter(self)
                QApplication.removePostedEvents(widget)
            except Exception:
                pass
        for child_name in (
            "large_dialog", "promotion_data_dialog", "pdd_product_match_dialog",
            "pdd_link_control_dialog", "pdd_code_dialog", "pdd_price_dialog",
            "promotion_status_dialog",
        ):
            child = getattr(self, child_name, None)
            if child is not None:
                try:
                    child.close()
                    child.deleteLater()
                except Exception:
                    pass
                setattr(self, child_name, None)

    def open_promotion_data_dialog(self):
        existing = getattr(self, "promotion_data_dialog", None)
        if existing is not None:
            existing.showNormal() if existing.isMinimized() else existing.show()
            existing.raise_()
            existing.activateWindow()
            return
        try:
            from manager.dialogs.promotion_data import PromotionDataDialog
        except ImportError:
            from dialogs.promotion_data import PromotionDataDialog
        dialog = PromotionDataDialog(self.store_id, self.store_name, self.db, self.main_app, None)
        dialog.setWindowModality(Qt.NonModal)
        dialog.setAttribute(Qt.WA_DeleteOnClose, True)
        dialog.destroyed.connect(
            lambda _=None, d=dialog: setattr(self, "promotion_data_dialog", None)
            if getattr(self, "promotion_data_dialog", None) is d else None
        )
        self.promotion_data_dialog = dialog
        dialog.show()
        dialog.raise_()
        dialog.activateWindow()

    def open_store_margin_display_settings(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("订单规格显示设置")
        dialog.resize(360, 190)
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(14, 14, 14, 14)

        main_row = QHBoxLayout()
        main_row.addWidget(QLabel("主卖规格显示："))
        combo_main = QComboBox()
        combo_main.addItem("规格编码", "code")
        combo_main.addItem("规格名称", "name")
        main_mode = self._get_spec_display_mode("store_margin_main_spec_display")
        combo_main.setCurrentIndex(1 if main_mode == "name" else 0)
        main_row.addWidget(combo_main)
        layout.addLayout(main_row)

        refund_row = QHBoxLayout()
        refund_row.addWidget(QLabel("退款最多规格显示："))
        combo_refund = QComboBox()
        combo_refund.addItem("规格编码", "code")
        combo_refund.addItem("规格名称", "name")
        refund_mode = self._get_spec_display_mode("store_margin_refund_spec_display")
        combo_refund.setCurrentIndex(1 if refund_mode == "name" else 0)
        refund_row.addWidget(combo_refund)
        layout.addLayout(refund_row)

        btn_row = QHBoxLayout()
        btn_save = QPushButton("保存")
        btn_cancel = QPushButton("取消")
        btn_save.setStyleSheet("QPushButton { background-color: #27ae60; color: white; padding: 6px 16px; border-radius: 4px; }")
        btn_row.addStretch()
        btn_row.addWidget(btn_save)
        btn_row.addWidget(btn_cancel)
        layout.addStretch()
        layout.addLayout(btn_row)

        def save_settings():
            self.db.set_setting("store_margin_main_spec_display", combo_main.currentData())
            self.db.set_setting("store_margin_refund_spec_display", combo_refund.currentData())
            dialog.accept()
            self.load_products()
            self.show_toast("规格显示设置已保存")

        btn_save.clicked.connect(save_settings)
        btn_cancel.clicked.connect(dialog.reject)
        dialog.exec_()

    def load_products(self):
        try:
            self.table.cellChanged.disconnect()
        except (TypeError, RuntimeError):
            pass
        products_raw = self.db.safe_fetchall(
            """SELECT id, name, title, image_data, sort_order, product_category_label,
                      store_weight, store_weight_locked
               FROM products WHERE store_id=? AND COALESCE(is_archived, 0)=0
                 AND COALESCE(is_violation, 0)=0""",
            (self.store_id,),
        )
        if hasattr(self.main_app, "_sort_products_for_display"):
            sorted_products = self.main_app._sort_products_for_display([
                (prod[0], prod[1], prod[2], prod[3], prod[4], prod[5])
                for prod in products_raw
            ])
            extra_by_id = {prod[0]: (prod[6], prod[7]) for prod in products_raw}
            products = [
                (prod[0], prod[1], prod[2], prod[3], *extra_by_id.get(prod[0], (0, 0)))
                for prod in sorted_products
            ]
        else:
            products = [
                (prod[0], prod[1], prod[2], prod[3], prod[6], prod[7])
                for prod in sorted(products_raw, key=lambda p: (p[4] if p[4] is not None else p[0], p[0]))
            ]
        self.sys_id_to_user_id = {}
        self.product_weights = {}
        self.refund_widgets = {}
        for prod in products:
            sys_id, prod_id, prod_title, image_data, store_weight, store_locked = prod
            self.sys_id_to_user_id[sys_id] = prod_id
            self.product_weights[prod_id] = {"sys_id": sys_id, "weight": store_weight or 0, "locked": 0}
        self.calculate_weights_from_orders()
        self.table.setRowCount(len(products))
        for row, prod in enumerate(products):
            sys_id, prod_id, prod_title, image_data, store_weight, store_locked = prod
            store_weight = self.product_weights.get(prod_id, {}).get("weight", store_weight or 0)
            if prod_id in self.product_weights:
                self.product_weights[prod_id]["locked"] = store_locked or 0
            img_widget = QWidget()
            img_layout = QVBoxLayout(img_widget)
            img_layout.setContentsMargins(1, 1, 1, 1)
            img_layout.setSpacing(0)
            img_layout.setAlignment(Qt.AlignCenter)
            img_label = QLabel()
            img_label.setFixedSize(70, 70)
            img_label.setScaledContents(False)
            img_label.setAlignment(Qt.AlignCenter)
            if image_data:
                pixmap = QPixmap()
                try:
                    pixmap.loadFromData(bytes(image_data))
                except Exception as e:
                    print(f"店铺毛利商品图片加载失败 product={prod_id}: {e}")
                    pixmap = QPixmap()
                if not pixmap.isNull():
                    scaled = pixmap.scaled(70, 70, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
                    x = max(0, (scaled.width() - 70) // 2)
                    y = max(0, (scaled.height() - 70) // 2)
                    img_label.setPixmap(scaled.copy(x, y, 70, 70))
                else:
                    img_label.setText("❌")
                    img_label.setStyleSheet("color: #999; border: 1px solid #ddd;")
            else:
                img_label.setText("📷")
                img_label.setStyleSheet("color: #999; border: 1px solid #ddd;")
            img_layout.addWidget(img_label)
            self.table.setCellWidget(row, 0, img_widget)
            self.table.setRowHeight(row, 72)
            item_id = QTableWidgetItem(str(prod_id))
            item_id.setFlags(item_id.flags() & ~Qt.ItemIsEditable)
            item_id.setFont(QFont("Microsoft YaHei", 9))
            self.table.setItem(row, 1, item_id)
            item_title = QTableWidgetItem(prod_title or "")
            item_title.setFlags(item_title.flags() & ~Qt.ItemIsEditable)
            title_font = QFont("Microsoft YaHei", 9)
            title_font.setUnderline(True)
            item_title.setFont(title_font)
            item_title.setForeground(QColor("#1769aa"))
            item_title.setToolTip("单击打开规格与毛利管理")
            self.table.setItem(row, 2, item_title)
            cost, price, margin = self.get_product_margin(sys_id)
            cost_item = QTableWidgetItem(f"¥{cost:.2f}" if cost else "¥0.00")
            cost_item.setFlags(cost_item.flags() & ~Qt.ItemIsEditable)
            cost_item.setTextAlignment(Qt.AlignCenter)
            cost_item.setFont(QFont("Microsoft YaHei", 19))
            self.table.setItem(row, 3, cost_item)
            item_price = QTableWidgetItem(f"¥{price:.2f}" if price else "¥0.00")
            item_price.setFlags(item_price.flags() & ~Qt.ItemIsEditable)
            item_price.setTextAlignment(Qt.AlignCenter)
            item_price.setFont(QFont("Microsoft YaHei", 19))
            self.table.setItem(row, 4, item_price)
            margin_text = f"{margin:.2f}%" if margin else "0.00%"
            item_margin = QTableWidgetItem(margin_text)
            item_margin.setFlags(item_margin.flags() & ~Qt.ItemIsEditable)
            item_margin.setTextAlignment(Qt.AlignCenter)
            item_margin.setFont(QFont("Microsoft YaHei", 19))
            if margin and margin < 10:
                item_margin.setBackground(QColor("#ffcccc"))
            elif margin and margin > 30:
                item_margin.setBackground(QColor("#ccffcc"))
            self.table.setItem(row, 5, item_margin)
            weight = store_weight or 0
            is_locked = store_locked or 0
            weight_widget = QWidget()
            weight_widget.setObjectName("plainTableCell")
            weight_layout = QHBoxLayout(weight_widget)
            weight_layout.setContentsMargins(0, 0, 0, 0)
            weight_layout.setSpacing(0)
            left_widget = QWidget()
            left_widget.setObjectName("plainTableCell")
            left_layout = QVBoxLayout(left_widget)
            left_layout.setContentsMargins(0, 0, 0, 0)
            weight_str = str(int(weight)) if weight == int(weight) else f"{weight:.1f}"
            weight_input = QLineEdit(weight_str)
            weight_input.setAlignment(Qt.AlignCenter)
            weight_input.setFixedHeight(25)
            weight_input.setReadOnly(True)
            weight_input.setStyleSheet(
                "QLineEdit { background: transparent; border: none; border-radius: 0px; padding: 0px; font-weight: bold; color: #2e7d32; }"
            )
            weight_input.installEventFilter(self)
            weight_input.setProperty("row", row)
            weight_input.setProperty("prod_id", prod_id)
            left_layout.addWidget(weight_input)
            weight_layout.addWidget(left_widget, 1)
            self.table.setCellWidget(row, 6, weight_widget)
            
            # 新增：权重对比列（第 8 列）
            weight_compare_widget = QWidget()
            weight_compare_widget.setObjectName("plainTableCell")
            weight_compare_layout = QHBoxLayout(weight_compare_widget)
            weight_compare_layout.setContentsMargins(0, 0, 0, 0)
            weight_compare_label = QLabel("-")
            weight_compare_label.setAlignment(Qt.AlignCenter)
            weight_compare_label.setStyleSheet("color: black; font-size: 19px;")
            weight_compare_layout.addWidget(weight_compare_label)
            self.table.setCellWidget(row, 7, weight_compare_widget)
            
            order_label_widget = QWidget()
            order_label_widget.setObjectName("plainTableCell")
            order_label_layout = QHBoxLayout(order_label_widget)
            order_label_layout.setContentsMargins(0, 0, 0, 0)
            order_label = QLabel("")
            order_label.setAlignment(Qt.AlignCenter)
            order_label.setStyleSheet("color: black; font-size: 19px;")
            order_label_layout.addWidget(order_label)
            self.table.setCellWidget(row, 8, order_label_widget)
            
            # 新增：单量对比列（第 10 列）
            order_compare_widget = QWidget()
            order_compare_widget.setObjectName("plainTableCell")
            order_compare_layout = QHBoxLayout(order_compare_widget)
            order_compare_layout.setContentsMargins(0, 0, 0, 0)
            order_compare_label = QLabel("-")
            order_compare_label.setAlignment(Qt.AlignCenter)
            order_compare_label.setStyleSheet("color: black; font-size: 19px;")
            order_compare_layout.addWidget(order_compare_label)
            self.table.setCellWidget(row, 9, order_compare_widget)
            
            main_spec_widget = QWidget()
            main_spec_widget.setObjectName("plainTableCell")
            main_spec_layout = QHBoxLayout(main_spec_widget)
            main_spec_layout.setContentsMargins(0, 0, 0, 0)
            main_spec_label = QLabel("-")
            main_spec_label.setAlignment(Qt.AlignCenter)
            main_spec_label.setWordWrap(True)
            main_spec_label.setStyleSheet("color: black; font-size: 19px;")
            main_spec_layout.addWidget(main_spec_label)
            self.table.setCellWidget(row, 11, main_spec_widget)
            if prod_id in self.product_weights:
                self.product_weights[prod_id]["main_spec"] = main_spec_label
                main_spec_code, spec_orders = self.get_main_spec(prod_id)
                main_spec_mode = self._get_spec_display_mode("store_margin_main_spec_display")
                if spec_orders > 0 and main_spec_code:
                    main_spec_text = self._get_spec_display_text(prod_id, main_spec_code, main_spec_mode)
                    self._set_spec_label_text(main_spec_label, main_spec_text, main_spec_mode)
                elif spec_orders == 0:
                    self._set_spec_label_text(main_spec_label, "无", main_spec_mode)
            refund_orders_label = QLabel("无")
            refund_orders_label.setAlignment(Qt.AlignCenter)
            refund_orders_label.setStyleSheet("color: #95a5a6; font-size: 19px;")
            self.table.setCellWidget(row, 12, refund_orders_label)
            self.refund_widgets[row] = {'orders': refund_orders_label}
            refund_ratio_label = QLabel("无")
            refund_ratio_label.setAlignment(Qt.AlignCenter)
            refund_ratio_label.setWordWrap(True)
            refund_ratio_label.setStyleSheet("color: #95a5a6; font-size: 19px;")
            self.table.setCellWidget(row, 13, refund_ratio_label)
            self.refund_widgets[row]['ratio'] = refund_ratio_label
            self.table.setItem(row, 10, QTableWidgetItem("-"))
            self.table.item(row, 10).setFont(QFont("Microsoft YaHei", 14))
            self.table.item(row, 10).setTextAlignment(Qt.AlignCenter)
            self.table.item(row, 1).setData(Qt.UserRole, prod_id)
            order_label.setProperty("prod_id", prod_id)
            self._update_order_label_for_row(row, weight_input, order_label, prod_id)
        self.table.cellChanged.connect(self.on_cell_changed)
        self.update_weight_inputs()
        self.calculate_total_margin()
        self.update_current_history_label()
        self.update_orders_display()
        self.update_compare_columns()
        self.update_product_avg_price()
        self._apply_store_margin_button_tooltips()

    def update_product_avg_price(self):
        """更新所有商品的客单价和销售额列"""
        self._normalize_imported_order_store_ids()
        for row in range(self.table.rowCount()):
            prod_id_item = self.table.item(row, 1)
            if not prod_id_item:
                continue
            user_product_id = prod_id_item.data(Qt.UserRole)
            if not user_product_id:
                continue
            sys_id = self.get_sys_id_by_user_id(user_product_id)
            if not sys_id:
                self.table.item(row, 4).setText("-")
                self.table.item(row, 10).setText("-")
                continue
            imported_totals = self.db.safe_fetchall(
                """SELECT COALESCE(SUM(order_count), 0), COALESCE(SUM(actual_amount), 0)
                   FROM imported_orders
                   WHERE store_id=? AND product_id=?""",
                (self.store_id, user_product_id)
            )
            imported_orders = imported_totals[0][0] if imported_totals else 0
            imported_amount = imported_totals[0][1] if imported_totals else 0
            if imported_orders and imported_amount and imported_amount > 0:
                avg_price = imported_amount / imported_orders
                self.table.item(row, 4).setText(f"¥{avg_price:.2f}")
                self.table.item(row, 10).setText(f"¥{imported_amount:.2f}")
                continue
            spec_sales = self.db.safe_fetchall(
                "SELECT ps.sale_price, io.order_count FROM product_specs ps "
                "LEFT JOIN imported_orders io ON io.product_id = ? AND io.spec_code = ps.spec_code "
                "WHERE ps.product_id = ? AND COALESCE(ps.is_temporarily_off_shelf, 0)=0",
                (user_product_id, sys_id)
            )
            total_amount = 0.0
            total_orders = 0
            for sale_price, order_count in spec_sales:
                if sale_price and order_count:
                    total_amount += sale_price * order_count
                    total_orders += order_count
            if total_orders > 0:
                avg_price = total_amount / total_orders
                self.table.item(row, 4).setText(f"¥{avg_price:.2f}")
                self.table.item(row, 10).setText(f"¥{total_amount:.2f}")
            else:
                self.table.item(row, 4).setText("-")
                self.table.item(row, 10).setText("-")

    def refresh_manual_data_display(self):
        """刷新手动录入数据展示（只显示最近一次数据）"""
        try:
            records = self.load_manual_data()
            
            # 只取最后一条记录（最近一次）
            if not records:
                self.margin_data_table.setRowCount(0)
                return
                
            record = records[-1]  # 取最后一条

            self.margin_data_table.setRowCount(1)
            
            # 显示日期范围：开始日期 和 结束日期 分两行显示
            start_date = record[0] if record[0] else ""
            end_date = record[1] if record[1] else ""
            start_display = start_date[5:10] if start_date and len(start_date) >= 10 else start_date
            end_display = end_date[5:10] if end_date and len(end_date) >= 10 else end_date
            date_str = f"{start_display}\n{end_display}"

            # 计算天数
            days = 1
            if start_date and end_date:
                try:
                    from datetime import datetime
                    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
                    days = max(1, (end_dt - start_dt).days + 1)
                except:
                    pass
            
            # 计算日盈亏
            net_profit = record[17] if record[17] else 0
            daily_profit = net_profit / days if days > 0 else 0
            
            values = [
                date_str,  # 0: 日期
                str(int(record[2])),  # 1: 实发订单
                f"¥{record[3]:.2f}",  # 2: 实发金额
                f"¥{record[4]:.2f}",  # 3: 毛利润
                f"{record[11]:.2f}%",  # 4: 毛利率
                f"¥{record[5]:.2f}",  # 5: 退款金额
                f"{record[12]:.2f}%",  # 6: 金额退款率
                str(int(record[6])),  # 7: 退款订单
                f"{record[13]:.2f}%",  # 8: 订单退款率
                f"¥{record[14]:.2f}",  # 9: 件单价
                f"¥{record[7]:.2f}",  # 10: 推广费
                f"{record[15]:.2f}%",  # 11: 推广占比
                f"¥{record[16]:.2f}",  # 12: 技术服务费
                f"¥{record[8]:.2f}",  # 13: 扣款
                f"¥{record[9]:.2f}",  # 14: 其他服务
                f"¥{record[10]:.2f}",  # 15: 其他
                f"¥{record[17]:.2f}",  # 16: 净利润
                f"{record[18]:.2f}%",  # 17: 净利率
                f"¥{record[19]:.2f}",  # 18: 单笔利润
                f"¥{daily_profit:.2f}",  # 19: 日盈亏
            ]

            for j, value in enumerate(values):
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)

                if j == 0:
                    item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)

                if j == 0:
                    pass
                elif j in [1, 2, 3, 5, 7, 10, 13, 14, 15]:
                    item.setBackground(QColor("#c8e6c9"))
                    item.setForeground(QColor("#1b5e20"))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                else:
                    item.setBackground(QColor("#bbdefb"))
                    item.setForeground(QColor("#0d47a1"))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)

                self.margin_data_table.setItem(0, j, item)

            self.margin_data_table.setRowHeight(0, 60)

            # 计算周环比变化
            self.calculate_week_comparison(records)
            
            # 同步两表列宽
            self.sync_table_widths()

        except Exception as e:
            print(f"刷新手动数据展示失败: {e}")
            import traceback
            traceback.print_exc()

    def sync_table_widths(self):
        """同步两个表格的列宽，确保完全对齐"""
        try:
            for col in range(20):
                width = self.margin_data_table.columnWidth(col)
                self.week_table.setColumnWidth(col, width)
        except Exception as e:
            print(f"同步列宽失败: {e}")

    def _excel_headers(self, table):
        headers = []
        for col in range(table.columnCount()):
            item = table.horizontalHeaderItem(col)
            headers.append(item.text().replace("\n", " ") if item else "")
        return headers

    def _table_cell_text(self, table, row, col):
        widget = table.cellWidget(row, col)
        if widget:
            if isinstance(widget, QLabel):
                return widget.text()
            line_edit = widget.findChild(QLineEdit)
            if line_edit:
                return line_edit.text()
            labels = widget.findChildren(QLabel)
            label_texts = [label.text() for label in labels if label.text() and label.text() not in ("❌", "-")]
            if label_texts:
                return "\n".join(label_texts)
        item = table.item(row, col)
        return item.text() if item else ""

    def _format_manual_record_for_export(self, record):
        start_date = record[0] if record[0] else ""
        end_date = record[1] if record[1] else ""
        date_text = self._export_period_day_text(start_date, end_date)
        days = 1
        if start_date and end_date:
            try:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                end_dt = datetime.strptime(end_date, "%Y-%m-%d")
                days = max(1, (end_dt - start_dt).days + 1)
            except Exception:
                days = 1
        net_profit = float(record[17] or 0)
        daily_profit = net_profit / days if days else 0
        return [
            date_text,
            str(int(record[2] or 0)),
            f"¥{float(record[3] or 0):.2f}",
            f"¥{float(record[4] or 0):.2f}",
            f"{float(record[11] or 0):.2f}%",
            f"¥{float(record[5] or 0):.2f}",
            f"{float(record[12] or 0):.2f}%",
            str(int(record[6] or 0)),
            f"{float(record[13] or 0):.2f}%",
            f"¥{float(record[14] or 0):.2f}",
            f"¥{float(record[7] or 0):.2f}",
            f"{float(record[15] or 0):.2f}%",
            f"¥{float(record[16] or 0):.2f}",
            f"¥{float(record[8] or 0):.2f}",
            f"¥{float(record[9] or 0):.2f}",
            f"¥{float(record[10] or 0):.2f}",
            f"¥{float(record[17] or 0):.2f}",
            f"{float(record[18] or 0):.2f}%",
            f"¥{float(record[19] or 0):.2f}",
            f"¥{daily_profit:.2f}",
        ]

    def _export_period_day_text(self, start_date, end_date):
        days = []
        for value in (start_date, end_date):
            day_text = self._export_day_text(value)
            if day_text and day_text not in days:
                days.append(day_text)
        return "~".join(days)

    def _export_day_text(self, value):
        text = str(value or "").strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
            try:
                dt = datetime.strptime(text, fmt)
                return f"{dt.month}月{dt.day}日"
            except Exception:
                pass
        match = re.search(r"(\d{1,2})月(\d{1,2})日", text)
        if match:
            return f"{int(match.group(1))}月{int(match.group(2))}日"
        match = re.search(r"\d{4}[-/](\d{1,2})[-/](\d{1,2})", text)
        if match:
            return f"{int(match.group(1))}月{int(match.group(2))}日"
        return text

    def _manual_record_numeric_values_for_export(self, record):
        days = 1
        if record[0] and record[1]:
            try:
                start_dt = datetime.strptime(record[0], "%Y-%m-%d")
                end_dt = datetime.strptime(record[1], "%Y-%m-%d")
                days = max(1, (end_dt - start_dt).days + 1)
            except Exception:
                days = 1
        net_profit = float(record[17] or 0)
        return [
            None,
            float(record[2] or 0),
            float(record[3] or 0),
            float(record[4] or 0),
            float(record[11] or 0),
            float(record[5] or 0),
            float(record[12] or 0),
            float(record[6] or 0),
            float(record[13] or 0),
            float(record[14] or 0),
            float(record[7] or 0),
            float(record[15] or 0),
            float(record[16] or 0),
            float(record[8] or 0),
            float(record[9] or 0),
            float(record[10] or 0),
            net_profit,
            float(record[18] or 0),
            float(record[19] or 0),
            net_profit / days if days else 0,
        ]

    def _manual_compare_export_row(self, current, previous):
        current_values = self._manual_record_numeric_values_for_export(current)
        previous_values = self._manual_record_numeric_values_for_export(previous)
        rate_cols = {4, 6, 8, 11, 17}
        values = ["较上期"]
        directions = ["flat"]
        for col in range(1, 20):
            curr = current_values[col] or 0
            prev = previous_values[col] or 0
            diff = curr - prev
            if abs(diff) < 0.000001:
                values.append("→ 0.0%")
                directions.append("flat")
                continue
            icon = "↑" if diff > 0 else "↓"
            directions.append("up" if diff > 0 else "down")
            if col in rate_cols:
                values.append(f"{icon} {abs(diff):.1f}%")
            elif abs(prev) > 0.000001:
                values.append(f"{icon} {abs(diff / abs(prev) * 100):.1f}%")
            else:
                values.append(f"{icon} 新增")
        return values, directions

    def _style_excel_sheet(self, ws, widths=None):
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(name="SimHei", bold=True, color="1F2933", size=13)
        data_font = Font(name="SimHei", color="000000", size=12)
        thin = Side(style="thin", color="B7C4D1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                else:
                    cell.font = data_font
        if widths:
            for col_idx, width in widths.items():
                ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _style_historical_export_sheet(self, ws, row_types, compare_directions, widths=None):
        header_fill = PatternFill("solid", fgColor="EAF4FF")
        header_font = Font(name="SimHei", bold=True, color="1F2933", size=12)
        manual_font = Font(name="SimHei", bold=True, color="1B8F3A", size=12)
        calculated_font = Font(name="SimHei", bold=True, color="1D4ED8", size=12)
        neutral_font = Font(name="SimHei", bold=True, color="374151", size=12)
        compare_up_font = Font(name="SimHei", bold=True, color="16A34A", size=12)
        compare_down_font = Font(name="SimHei", bold=True, color="DC2626", size=12)
        compare_flat_font = Font(name="SimHei", bold=True, color="6B7280", size=12)
        white_fill = PatternFill("solid", fgColor="FFFFFF")
        compare_fill = PatternFill("solid", fgColor="F7F8FA")
        thin = Side(style="thin", color="CAD5E2")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        manual_cols = {2, 3, 4, 6, 8, 11, 14, 15, 16}

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                else:
                    row_type = row_types.get(cell.row, "data")
                    if row_type == "compare":
                        cell.fill = compare_fill
                        direction = compare_directions.get((cell.row, cell.column), "flat")
                        if cell.column == 1:
                            cell.font = neutral_font
                        elif direction == "up":
                            cell.font = compare_up_font
                        elif direction == "down":
                            cell.font = compare_down_font
                        else:
                            cell.font = compare_flat_font
                    else:
                        cell.fill = white_fill
                        if cell.column == 1:
                            cell.font = neutral_font
                        elif cell.column in manual_cols:
                            cell.font = manual_font
                        else:
                            cell.font = calculated_font
        if widths:
            for col_idx, width in widths.items():
                ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _write_historical_export_sheet(self, wb):
        ws = wb.active
        ws.title = "过往数据分析"
        ws.append(self._excel_headers(self.margin_data_table))
        row_types = {}
        compare_directions = {}
        records = self.load_manual_data()
        for index, record in enumerate(records):
            ws.append(self._format_manual_record_for_export(record))
            data_row = ws.max_row
            row_types[data_row] = "data"
            ws.row_dimensions[data_row].height = 34
            if index > 0:
                compare_values, directions = self._manual_compare_export_row(record, records[index - 1])
                ws.append(compare_values)
                compare_row = ws.max_row
                row_types[compare_row] = "compare"
                ws.row_dimensions[compare_row].height = 24
                for col_idx, direction in enumerate(directions, start=1):
                    compare_directions[(compare_row, col_idx)] = direction
        self._style_historical_export_sheet(
            ws,
            row_types,
            compare_directions,
            {
                1: 20, 2: 12, 3: 14, 4: 14, 5: 12,
                6: 14, 7: 12, 8: 12, 9: 12, 10: 12,
                11: 12, 12: 12, 13: 14, 14: 12, 15: 14,
                16: 12, 17: 14, 18: 12, 19: 14, 20: 14,
            }
        )
        ws.freeze_panes = "A2"
        self._append_reading_mode_images_to_historical_sheet(ws)

    def _export_month_text(self, value):
        text = str(value or "").strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
            try:
                dt = datetime.strptime(text, fmt)
                return f"{dt.month}月"
            except Exception:
                pass
        match = re.search(r"(\d{1,2})月", text)
        if match:
            return f"{int(match.group(1))}月"
        match = re.search(r"\d{4}[-/](\d{1,2})", text)
        if match:
            return f"{int(match.group(1))}月"
        return ""

    def _reading_mode_images_for_export(self):
        try:
            return self.db.safe_fetchall(
                """SELECT slot_index, image_data, created_time
                   FROM store_temp_images
                   WHERE store_id=?
                   ORDER BY slot_index""",
                (self.store_id,),
            )
        except Exception as e:
            print(f"读取本周附带图片失败: {e}")
            return []

    def _add_reading_export_image(self, ws, image_data, cell, image_refs, label):
        if not image_data:
            return None
        try:
            image_stream, display_width, display_height = self._high_res_export_image_stream(
                image_data,
                max_embed_size=None,
                max_display_width=96,
                max_display_height=96,
            )
            if image_stream is None:
                return None
            image = ExcelImage(image_stream)
            image.width = display_width
            image.height = display_height
            ws.add_image(image, cell)
            image_refs.append(image_stream)
            return image
        except Exception as e:
            print(f"导出{label}失败: {e}")
            return None

    def _high_res_export_image_stream(self, image_data, max_embed_size=2400, max_display_width=360, max_display_height=260):
        from PIL import Image as PilImage

        source = PilImage.open(BytesIO(image_data))
        if source.mode not in ("RGB", "RGBA"):
            source = source.convert("RGBA")
        width, height = source.size
        if width <= 0 or height <= 0:
            return None, 0, 0

        embed = source.copy()
        if max_embed_size and max(width, height) > max_embed_size:
            embed.thumbnail((int(max_embed_size), int(max_embed_size)), PilImage.LANCZOS)
            width, height = embed.size

        display_scale = min(float(max_display_width) / width, float(max_display_height) / height, 1.0)
        display_width = max(1, int(width * display_scale))
        display_height = max(1, int(height * display_scale))

        stream = BytesIO()
        embed.save(stream, format="PNG", optimize=True)
        stream.seek(0)
        return stream, display_width, display_height

    def _thumbnail_export_image_stream(self, image_data, max_width, max_height):
        from PIL import Image as PilImage

        source = PilImage.open(BytesIO(image_data))
        if source.mode not in ("RGB", "RGBA"):
            source = source.convert("RGBA")
        width, height = source.size
        if width <= 0 or height <= 0:
            return None
        source.thumbnail((int(max_width), int(max_height)), PilImage.LANCZOS)
        stream = BytesIO()
        source.save(stream, format="PNG", optimize=True)
        stream.seek(0)
        return stream

    def _append_reading_mode_images_to_historical_sheet(self, ws):
        rows = self._reading_mode_images_for_export()
        if not rows:
            return
        image_refs = getattr(ws, "_image_stream_refs", [])
        start_row = ws.max_row + 2
        max_col = max(20, ws.max_column)
        title_fill = PatternFill("solid", fgColor="EAF4FF")
        label_fill = PatternFill("solid", fgColor="F8FAFC")
        header_font = Font(name="SimHei", bold=True, color="1F2933", size=13)
        data_font = Font(name="SimHei", color="000000", size=12)
        thin = Side(style="thin", color="CAD5E2")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_col)
        title_cell = ws.cell(start_row, 1, "本周附带图片")
        title_cell.fill = title_fill
        title_cell.font = header_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        title_cell.border = border
        current_row = start_row + 1
        ws.cell(current_row, 1, "月份")
        ws.cell(current_row, 2, "位置")
        ws.cell(current_row, 3, "图片")
        for col in range(1, 4):
            cell = ws.cell(current_row, col)
            cell.fill = label_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=6)
        current_row += 1

        for slot_index, image_data, created_time in rows:
            ws.cell(current_row, 1, self._export_month_text(created_time))
            ws.cell(current_row, 2, f"图片{int(slot_index or 0) + 1}")
            for col in range(1, 4):
                cell = ws.cell(current_row, col)
                cell.font = data_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
                if col < 3:
                    cell.fill = PatternFill("solid", fgColor="FFFFFF")
            ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=6)
            image = self._add_reading_export_image(ws, image_data, f"C{current_row}", image_refs, "本周附带图片")
            if image:
                ws.row_dimensions[current_row].height = self._excel_row_height_for_pixels(100)
            current_row += 1
        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 12)
        ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 12)
        ws._image_stream_refs = image_refs

    def _product_image_map_for_export(self):
        rows = self.db.safe_fetchall(
            "SELECT name, image_data FROM products WHERE store_id=? AND COALESCE(is_archived, 0)=0",
            (self.store_id,)
        )
        return {str(name): image_data for name, image_data in rows if name is not None and image_data}

    def _compare_export_text_and_direction(self, table, row, col):
        text = self._table_cell_text(table, row, col).strip()
        widget = table.cellWidget(row, col)
        style = ""
        if widget:
            if isinstance(widget, QLabel):
                style = widget.styleSheet() or ""
            else:
                labels = widget.findChildren(QLabel)
                if labels:
                    style = labels[0].styleSheet() or ""
        clean = re.sub(r"[\U0001F7E2\U0001F534\U000026AA\U000026AB\U00002B06\U00002B07\U000027A1\ufe0f]", "", text)
        clean = clean.replace("⬆", "").replace("⬇", "").replace("➡", "").strip()
        if "#27ae60" in style.lower() or "green" in style.lower():
            return f"↑ {clean}", "up"
        if "#c0392b" in style.lower() or "red" in style.lower():
            return f"↓ {clean}", "down"
        if clean in ("", "-", "无"):
            return clean or "-", "flat"
        return f"→ {clean}", "flat"

    def _write_orders_export_sheet(self, wb):
        ws = wb.create_sheet("店铺商品权重")
        headers = self._excel_headers(self.table)
        if headers and headers[-1] == "操作":
            headers = headers[:-1]
        ws.append(headers)
        image_map = self._product_image_map_for_export()
        image_refs = []
        compare_styles = []
        image_size = self._export_product_image_size()
        for row in range(self.table.rowCount()):
            product_id = self._table_cell_text(self.table, row, 1)
            values = []
            for col in range(len(headers)):
                if col == 0:
                    values.append("")
                elif col in (7, 9):
                    compare_text, direction = self._compare_export_text_and_direction(self.table, row, col)
                    values.append(compare_text)
                    compare_styles.append((row + 2, col + 1, direction))
                else:
                    values.append(self._table_cell_text(self.table, row, col))
            ws.append(values)
            excel_row = row + 2
            self._set_square_image_cell(ws, excel_row, 1, image_size)
            image_data = image_map.get(product_id)
            if image_data:
                self._add_export_image(ws, image_data, f"A{excel_row}", image_size, image_refs, "商品")
        ws._image_stream_refs = image_refs
        self._style_excel_sheet(ws, {1: self._excel_column_width_for_pixels(image_size), 2: 14, 3: 24, 4: 11, 5: 11, 6: 10, 7: 10, 8: 12, 9: 10, 10: 12, 11: 12, 12: 16, 13: 10, 14: 18})
        weight_col = None
        for col_idx, header in enumerate(headers, start=1):
            if str(header or "").strip() == "权重":
                weight_col = col_idx
                break
        if weight_col:
            ws.cell(1, weight_col).font = Font(name="SimHei", bold=True, color="1F2933", size=15)
            for excel_row in range(2, ws.max_row + 1):
                cell = ws.cell(excel_row, weight_col)
                text = str(cell.value or "").strip()
                if text and text not in ("-", "无") and "%" not in text:
                    number = self._first_number_for_export(text)
                    cell.value = f"{number:.2f}%" if number is not None else f"{text}%"
                cell.font = Font(name="SimHei", color="000000", size=14)
        for excel_row, excel_col, direction in compare_styles:
            color = "16A34A" if direction == "up" else "DC2626" if direction == "down" else "6B7280"
            ws.cell(excel_row, excel_col).font = Font(name="SimHei", color=color, size=12, bold=direction in ("up", "down"))
        ws.freeze_panes = "A2"

    def _safe_export_float(self, value, default=0.0):
        try:
            if value is None or value == "":
                return default
            return float(value)
        except (TypeError, ValueError):
            return default

    def _fmt_export_money(self, value):
        return f"¥{self._safe_export_float(value):.2f}"

    def _fmt_export_pct(self, value):
        return f"{self._safe_export_float(value):.2f}%"

    def _fmt_export_number(self, value, digits=2, suffix=""):
        return f"{self._safe_export_float(value):.{digits}f}{suffix}"

    def _split_spec_name_for_export(self, product_sys_id, spec_name):
        text = str(spec_name or "").strip()
        return text

    def _promotion_summary_for_export(self, ctx):
        parts = []
        if ctx["coupon"] > 0:
            parts.append(f"有{ctx['coupon']:.0f}元优惠券")
        if ctx["new_customer"] > 0:
            parts.append(f"有{ctx['new_customer']:.0f}元新客立减")
        store_discount_text = ctx.get("store_discount_text") or ""
        if store_discount_text and store_discount_text != "未设置":
            parts.append(f"店铺满减：{store_discount_text}")
        if ctx["is_limited_time"]:
            parts.append("限时限量购")
        if ctx["is_marketing"]:
            parts.append("活动")
        return "；".join(parts) if parts else "这个链接是裸价"

    def _roi_summary_for_export(self, ctx):
        parts = [f"推广方式：{ctx['promotion_mode']}"]
        if ctx["current_roi"] > 0:
            parts.append(f"当前投产：{ctx['current_roi']:.2f}")
        if ctx["transaction_bid"] > 0:
            parts.append(f"成交出价：{ctx['transaction_bid']:.2f}")
        if ctx["margin_pct"] > 0:
            parts.append(f"综合毛利率：{ctx['margin_pct']:.2f}%")
        if ctx["gross_break_even"] > 0:
            parts.append(f"毛保本投产：{ctx['gross_break_even']:.2f}")
        if ctx["net_break_even"] > 0:
            parts.append(f"净保本投产：{ctx['net_break_even']:.2f}")
        if ctx["roi_multiple"] > 0:
            parts.append(f"投产倍数：{ctx['roi_multiple']:.2f}倍")
        if ctx["scale_roi"] > 0:
            parts.append(f"放量投产：{ctx['scale_roi']:.2f}")
        if ctx["promotion_ratio"] > 0:
            parts.append(f"推广占比：{ctx['promotion_ratio']:.2f}%")
        if ctx["net_profit_rate"]:
            parts.append(f"净利率：{ctx['net_profit_rate']:.2f}%")
        return "；".join(parts)

    def _previous_week_range_for_export(self):
        today = datetime.now().date()
        this_week_monday = today - timedelta(days=today.weekday())
        start_date = this_week_monday - timedelta(days=7)
        end_date = this_week_monday - timedelta(days=1)
        return start_date, end_date

    def _weekday_cn_for_export(self, record_date):
        return ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][record_date.weekday()]

    def _operation_record_color_for_export(self, record_date):
        colors = ["1D4ED8", "16A34A", "9333EA", "EA580C", "0891B2", "DB2777", "4B5563"]
        return colors[record_date.weekday() % len(colors)]

    def _style_operation_record_export_row(self, ws, row, max_col, color, bold=False, start_col=1):
        font = Font(name="SimHei", color=color, size=12, bold=bold)
        for col in range(start_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _first_number_for_export(self, value):
        match = re.search(r"-?\d+(?:\.\d+)?", str(value or ""))
        if not match:
            return None
        try:
            return float(match.group(0))
        except ValueError:
            return None

    def _brief_change_text_for_export(self, change, fallback_text=""):
        if not isinstance(change, dict):
            return str(change or fallback_text or "").strip()
        metric = str(change.get("metric", "") or "").strip()
        text = str(change.get("text", "") or "").strip()
        old = str(change.get("old", "") or "").strip()
        new = str(change.get("new", "") or "").strip()
        combined = f"{metric} {text} {old} {new}"

        def new_text():
            return new or text or fallback_text

        if "推广模式" in combined or "自然流" in combined or "全站" in combined or "稳定成本" in combined:
            return f"变到{new_text()}" if new_text() else "调整推广方式"
        if "成交出价" in combined or "出价" in combined:
            return f"成交出价变到{new_text()}" if new_text() else "调整成交出价"
        if "投产" in combined or "ROI" in combined or "roi" in combined:
            old_num = self._first_number_for_export(old)
            new_num = self._first_number_for_export(new)
            if old_num is not None and new_num is not None:
                action = "提投产到" if new_num > old_num else "降投产到" if new_num < old_num else "投产变到"
                return f"{action}{new}"
            return f"调整投产：{new_text()}" if new_text() else "调整投产"
        if "售价" in combined or "价格" in combined or "涨价" in combined or "降价" in combined:
            old_num = self._first_number_for_export(old)
            new_num = self._first_number_for_export(new)
            if old_num is not None and new_num is not None:
                action = "涨价格到" if new_num > old_num else "降价格到" if new_num < old_num else "价格变到"
                return f"{action}{new}"
            return text or (f"改价格到{new}" if new else "调整价格")
        if "优惠券" in combined:
            return f"优惠券变到{new_text()}" if new_text() else "调整优惠券"
        if "新客立减" in combined:
            return f"新客立减变到{new_text()}" if new_text() else "调整新客立减"
        if "限时" in combined or "营销" in combined or "活动" in combined:
            return text or (f"活动变到{new}" if new else "调整活动")
        if "规格新增" in combined or "新增规格" in combined:
            return text or "新增规格"
        if "规格删除" in combined or "删除规格" in combined:
            return text or "删除规格"
        if "规格名称" in combined:
            return text or (f"规格名称变到{new}" if new else "修改规格名称")
        return text or fallback_text or (f"{metric}：{old}→{new}" if metric or old or new else "")

    def _record_briefs_for_export(self, product_sys_id):
        start_date, end_date = self._previous_week_range_for_export()
        rows = self.db.safe_fetchall(
            """SELECT year, month, day, records_json
               FROM records
               WHERE product_id=?""",
            (product_sys_id,),
        )
        briefs = []
        for year, month, day, records_json in rows:
            try:
                record_date = datetime(int(year), int(month), int(day)).date()
            except Exception:
                continue
            if record_date < start_date or record_date > end_date:
                continue
            try:
                records = json.loads(records_json) if records_json else []
            except Exception:
                records = []
            for record in records:
                if not isinstance(record, dict):
                    continue
                time_text = str(record.get("time", "") or "").strip()
                fallback_text = str(record.get("text", "") or "").strip()
                changes = record.get("changes") or []
                if isinstance(changes, list) and changes:
                    texts = [self._brief_change_text_for_export(change, fallback_text) for change in changes]
                    content = "；".join([text for text in texts if text])
                else:
                    content = fallback_text
                content = re.sub(r"\s+", " ", str(content or "")).strip()
                if not content:
                    continue
                prefix = f"{record_date.month}/{record_date.day} {self._weekday_cn_for_export(record_date)}"
                if time_text:
                    prefix += f" {time_text}"
                briefs.append({
                    "date": record_date,
                    "time": time_text,
                    "text": f"{prefix} {content}",
                    "has_spec": "规格" in content,
                })
        briefs.sort(key=lambda item: (item["date"], item["time"]))
        range_text = f"{start_date.month}/{start_date.day}-{end_date.month}/{end_date.day}"
        return range_text, briefs

    def _products_for_specs_export(self):
        rows = self.db.safe_fetchall(
            """SELECT id, name, title, image_data, sort_order, product_category_label
               FROM products
               WHERE store_id=? AND COALESCE(is_archived, 0)=0""",
            (self.store_id,),
        )
        detail_ids = getattr(self, "export_detail_product_ids", None)
        if detail_ids is not None:
            detail_ids = {str(value) for value in detail_ids}
            rows = [row for row in rows if str(row[1]) in detail_ids]
        if hasattr(self.main_app, "_sort_products_for_display"):
            sorted_rows = self.main_app._sort_products_for_display(rows)
        else:
            sorted_rows = sorted(rows, key=lambda p: (p[4] if p[4] is not None else p[0], p[0]))
        return sorted_rows

    def _product_order_map_for_export(self, product_id):
        rows = self.db.safe_fetchall(
            """SELECT spec_code, COALESCE(SUM(order_count), 0), COALESCE(SUM(refund_count), 0)
               FROM imported_orders
               WHERE store_id=? AND product_id=?
               GROUP BY spec_code""",
            (self.store_id, product_id),
        )
        return {
            str(spec_code or ""): {
                "orders": int(order_count or 0),
                "refunds": int(refund_count or 0),
            }
            for spec_code, order_count, refund_count in rows
        }

    def _product_export_context(self, product):
        sys_id, product_id, title, image_data, _sort_order, category_label = product
        product_rows = self.db.safe_fetchall(
            """SELECT coupon_amount, new_customer_discount, current_roi, COALESCE(transaction_bid, 0),
                      return_rate, is_limited_time, is_marketing, is_natural_flow,
                      is_sitewide_managed, COALESCE(roi_input_mode, 'roi'), product_memo
               FROM products WHERE id=?""",
            (sys_id,),
        )
        prod = product_rows[0] if product_rows else (0, 0, 0, 0, 0, 0, 0, 0, 0, "roi", "")
        coupon = self._safe_export_float(prod[0])
        new_customer = self._safe_export_float(prod[1])
        product_discount = max(coupon, new_customer)
        saved_roi = self._safe_export_float(prod[2])
        transaction_bid = self._safe_export_float(prod[3])
        return_rate = self._safe_export_float(prod[4])
        is_limited_time = bool(prod[5])
        is_marketing = bool(prod[6])
        is_natural_flow = bool(prod[7])
        is_sitewide_managed = bool(prod[8]) and not is_natural_flow
        roi_input_mode = prod[9] if prod[9] in ("roi", "bid") else "roi"
        product_memo = str(prod[10] or "")
        sitewide_rows = self.db.safe_fetchall("SELECT sitewide_roi FROM stores WHERE id=?", (self.store_id,))
        sitewide_roi = self._safe_export_float(sitewide_rows[0][0]) if sitewide_rows else 0.0
        store_discount_text = self.db.format_store_discount_rules(self.store_id)

        spec_rows = self.db.safe_fetchall(
            """SELECT spec_name, spec_code, sale_price, weight_percent, is_locked, spec_image_data
               FROM product_specs WHERE product_id=?
                 AND COALESCE(is_temporarily_off_shelf, 0)=0""",
            (sys_id,),
        )
        cost_map = {}
        spec_codes = [str(row[1] or "") for row in spec_rows if row[1]]
        if spec_codes:
            for i in range(0, len(spec_codes), 800):
                chunk = spec_codes[i:i + 800]
                placeholders = ",".join(["?"] * len(chunk))
                cost_rows = self.db.safe_fetchall(
                    f"SELECT spec_code, cost_price FROM cost_library WHERE spec_code IN ({placeholders})",
                    tuple(chunk),
                )
                cost_map.update({str(code): self._safe_export_float(cost) for code, cost in cost_rows})

        order_map = self._product_order_map_for_export(product_id)
        valid_spec_codes = {str(row[1] or "") for row in spec_rows if row[1]}
        recognized_orders = sum(
            data["orders"] for code, data in order_map.items()
            if code in valid_spec_codes
        )

        specs = []
        total_weight = 0.0
        weighted_margin = 0.0
        weighted_final_price = 0.0
        weighted_profit = 0.0
        fallback_prices = []
        fallback_profits = []
        for row in spec_rows:
            spec_name = str(row[0] or "")
            spec_code = str(row[1] or "")
            sale_price = self._safe_export_float(row[2])
            weight_percent = self._safe_export_float(row[3])
            is_locked = bool(row[4])
            spec_image_data = row[5] if len(row) > 5 else None
            cost = cost_map.get(spec_code, 0.0)
            store_discount, _rule = self.db.calculate_store_discount(self.store_id, sale_price)
            effective_discount = max(product_discount, store_discount)
            final_price = sale_price - effective_discount
            margin_pct = ((final_price - cost) / final_price * 100) if final_price > 0 and cost > 0 else 0.0
            gross_profit = final_price - cost
            order_data = order_map.get(spec_code, {"orders": 0, "refunds": 0})
            order_count = order_data["orders"]
            refund_count = order_data["refunds"]
            if recognized_orders > 0:
                display_weight = order_count / recognized_orders * 100
                weight_source = "导入订单"
            else:
                display_weight = weight_percent
                weight_source = "手动权重"
            refund_ratio = refund_count / order_count * 100 if order_count > 0 and refund_count > 0 else 0.0
            display_spec_name = self._split_spec_name_for_export(sys_id, spec_name)
            specs.append({
                "spec_image_data": spec_image_data,
                "spec_name": display_spec_name,
                "spec_code": spec_code,
                "cost": cost,
                "sale_price": sale_price,
                "final_price": final_price,
                "effective_discount": effective_discount,
                "margin_pct": margin_pct,
                "gross_profit": gross_profit,
                "weight": display_weight,
                "weight_source": weight_source,
                "is_locked": is_locked,
                "order_count": order_count,
                "refund_count": refund_count,
                "refund_ratio": refund_ratio,
            })
            if final_price > 0:
                fallback_prices.append(final_price)
                fallback_profits.append(gross_profit)
            if final_price > 0 and display_weight > 0:
                weighted_final_price += final_price * display_weight
                weighted_profit += gross_profit * display_weight
                total_weight += display_weight
                weighted_margin += (margin_pct / 100.0) * display_weight

        specs.sort(key=lambda item: item["final_price"])
        margin_decimal = weighted_margin / total_weight if total_weight > 0 else 0.0
        margin_pct = margin_decimal * 100
        avg_price = weighted_final_price / total_weight if total_weight > 0 else (
            sum(fallback_prices) / len(fallback_prices) if fallback_prices else 0.0
        )
        avg_profit = weighted_profit / total_weight if total_weight > 0 else (
            sum(fallback_profits) / len(fallback_profits) if fallback_profits else 0.0
        )
        gross_break_even = 1 / margin_decimal if margin_decimal > 0 else 0.0
        net_margin_formula = margin_decimal * (1 - return_rate / 100) - 0.006
        net_break_even = 1 / net_margin_formula if net_margin_formula > 0 else 0.0

        promotion_mode = "稳定成本推广"
        current_roi = saved_roi
        if is_natural_flow:
            promotion_mode = "自然流"
            current_roi = 0.0
        elif is_sitewide_managed:
            promotion_mode = "全站托管"
            current_roi = sitewide_roi
        elif roi_input_mode == "bid" and transaction_bid > 0 and avg_price > 0:
            promotion_mode = "稳定成本推广（成交出价）"
            current_roi = avg_price / transaction_bid

        if is_natural_flow:
            net_profit_rate = net_margin_formula * 100
            promotion_ratio = 0.0
            roi_multiple = 0.0
        elif roi_input_mode == "bid" and not is_sitewide_managed and transaction_bid > 0 and avg_price > 0:
            net_profit_amount = avg_profit * (1 - return_rate / 100) - (avg_price * 0.006) - transaction_bid
            net_profit_rate = net_profit_amount / avg_price * 100
            promotion_ratio = transaction_bid / avg_price * 100
            roi_multiple = current_roi / net_break_even if net_break_even > 0 else 0.0
        elif current_roi > 0:
            net_profit_rate = (margin_decimal * (1 - return_rate / 100) - 0.006 - (1 / current_roi)) * 100
            promotion_ratio = 1 / current_roi * 100
            roi_multiple = current_roi / net_break_even if net_break_even > 0 else 0.0
        else:
            net_profit_rate = 0.0
            promotion_ratio = 0.0
            roi_multiple = 0.0

        return {
            "sys_id": sys_id,
            "product_id": str(product_id or ""),
            "title": str(title or ""),
            "image_data": image_data,
            "category_label": str(category_label or ""),
            "product_memo": product_memo,
            "coupon": coupon,
            "new_customer": new_customer,
            "store_discount_text": store_discount_text,
            "max_discount": max([product_discount] + [spec.get("effective_discount", 0.0) for spec in specs]) if specs else product_discount,
            "is_limited_time": is_limited_time,
            "is_marketing": is_marketing,
            "is_natural_flow": is_natural_flow,
            "is_sitewide_managed": is_sitewide_managed,
            "promotion_mode": promotion_mode,
            "roi_input_mode": "成交出价" if roi_input_mode == "bid" else "投产比",
            "current_roi": current_roi,
            "transaction_bid": transaction_bid,
            "return_rate": return_rate,
            "margin_pct": margin_pct,
            "gross_break_even": gross_break_even,
            "net_break_even": net_break_even,
            "roi_multiple": roi_multiple,
            "scale_roi": net_break_even * 0.8 if net_break_even > 0 else 0.0,
            "promotion_ratio": promotion_ratio,
            "net_profit_rate": net_profit_rate,
            "specs": specs,
        }

    def _set_range_fill(self, ws, row, start_col, end_col, fill):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).fill = fill

    def _style_product_specs_block(self, ws, start_row, end_row, max_col):
        thin = Side(style="thin", color="C9D4E2")
        thick = Side(style="thick", color="000000")
        for row in range(start_row, end_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if row == start_row:
                    cell.font = Font(name="SimHei", color="000000", size=13, bold=True)
                else:
                    cell.font = Font(name="SimHei", color="000000", size=12, bold=cell.font.bold)
                left = thick if col == 1 else thin
                right = thick if col == max_col else thin
                top = thick if row == start_row else thin
                bottom = thick if row == end_row else thin
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    def _style_product_specs_export_range(self, ws, start_row, end_row, start_col, end_col):
        thin = Side(style="thin", color="C9D4E2")
        thick = Side(style="thick", color="000000")
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if row in (start_row, start_row + 2, start_row + 3):
                    cell.font = Font(name="SimHei", color="000000", size=13, bold=True)
                else:
                    cell.font = Font(name="SimHei", color="000000", size=12, bold=cell.font.bold)
                left = thick if col == start_col else thin
                right = thick if col == end_col else thin
                top = thick if row == start_row else thin
                bottom = thick if row == end_row else thin
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    def _excel_column_width_for_pixels(self, pixels):
        return max(1.0, (float(pixels) - 5.0) / 7.0)

    def _excel_row_height_for_pixels(self, pixels):
        return float(pixels) * 0.75

    def _export_product_image_size(self):
        return 96

    def _export_image_quality_config(self):
        mode = getattr(self, "export_image_quality", "clear")
        configs = {
            "clear": {"label": "清晰版", "square": 1600, "reading": 2400},
            "balanced": {"label": "均衡版", "square": 1000, "reading": 1400},
            "light": {"label": "轻量版", "square": 640, "reading": 900},
        }
        return configs.get(mode, configs["clear"])

    def _export_image_embed_size(self):
        return int(self._export_image_quality_config().get("square", 1600))

    def _set_square_image_cell(self, ws, row, col, size):
        ws.row_dimensions[row].height = self._excel_row_height_for_pixels(size)
        ws.column_dimensions[get_column_letter(col)].width = self._excel_column_width_for_pixels(size)

    def _square_image_stream(self, image_data, display_size, max_embed_size=None):
        from PIL import Image as PilImage

        max_embed_size = max_embed_size or self._export_image_embed_size()
        cache = getattr(self, "_excel_export_image_cache", None)
        cache_key = None
        if cache is not None:
            cache_key = (hashlib.md5(bytes(image_data)).hexdigest(), display_size, max_embed_size)
            cached = cache.get(cache_key)
            if cached:
                return BytesIO(cached)

        source = PilImage.open(BytesIO(image_data))
        if source.mode not in ("RGB", "RGBA"):
            source = source.convert("RGBA")
        width, height = source.size
        if width <= 0 or height <= 0:
            return None
        side = min(width, height)
        left = max(0, (width - side) // 2)
        top = max(0, (height - side) // 2)
        square = source.crop((left, top, left + side, top + side))
        target_size = side
        if target_size > max_embed_size:
            target_size = max_embed_size
        elif target_size < display_size:
            target_size = display_size
        if target_size != side:
            square = square.resize((target_size, target_size), PilImage.LANCZOS)
        stream = BytesIO()
        square.save(stream, format="PNG", optimize=True)
        stream.seek(0)
        if cache is not None and cache_key:
            cache[cache_key] = stream.getvalue()
            stream.seek(0)
        return stream

    def _add_export_image(self, ws, image_data, cell, size, image_refs, label):
        if not image_data:
            return
        try:
            image_stream = self._square_image_stream(image_data, size)
            if image_stream is None:
                return
            image = ExcelImage(image_stream)
            image.width = size
            image.height = size
            ws.add_image(image, cell)
            image_refs.append(image_stream)
        except Exception as e:
            print(f"导出{label}图片失败: {e}")

    def _operation_record_export_cells(self, record_info):
        record_date = record_info["date"]
        time_text = str(record_info.get("time") or "").strip()
        date_text = f"{record_date.month}/{record_date.day} {self._weekday_cn_for_export(record_date)}"
        if time_text:
            date_text += f" {time_text}"
        content = str(record_info.get("text") or "").strip()
        if content.startswith(date_text):
            content = content[len(date_text):].strip()
        return date_text, content or "-"

    def _write_product_specs_export_sheet(self, wb):
        ws = wb.create_sheet("商品规格售卖情况")
        image_size = self._export_product_image_size()

        basic_fill = PatternFill("solid", fgColor="EAF4FF")
        promo_fill = PatternFill("solid", fgColor="E8F7EF")
        roi_fill = PatternFill("solid", fgColor="FFF3D6")
        spec_fill = PatternFill("solid", fgColor="F3E8FF")
        operation_fill = PatternFill("solid", fgColor="EEF2FF")
        white_fill = PatternFill("solid", fgColor="FFFFFF")
        header_font = Font(name="SimHei", color="000000", bold=True, size=13)
        image_refs = []

        products = self._products_for_specs_export()
        if not products:
            ws.append(["暂无商品数据"])
            ws["A1"].font = header_font
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
            self._style_product_specs_block(ws, 1, 1, 6)
            return

        spec_fields = [
            ("规格图", "_image"),
            ("规格名称", "spec_name"),
            ("规格编码", "spec_code"),
            ("成本", "cost"),
            ("实际价格", "final_price"),
            ("毛利率", "margin_pct"),
            ("毛利润", "gross_profit"),
            ("权重", "weight"),
            ("单量", "order_count"),
            ("退款订单", "refund_count"),
            ("退款占比", "refund_ratio"),
        ]
        block_width = len(spec_fields)
        block_gap = 1
        info_row = 1
        category_row = 2
        spec_title_row = 3
        spec_header_row = 4
        spec_first_row = 5
        max_written_row = spec_header_row

        for product_index, product in enumerate(products):
            ctx = self._product_export_context(product)
            record_range, record_lines = self._record_briefs_for_export(ctx["sys_id"])
            specs = ctx.get("specs", [])
            start_col = 1 + product_index * (block_width + block_gap)
            end_col = start_col + block_width - 1
            block_start_row = info_row

            self._set_square_image_cell(ws, info_row, start_col, image_size)
            for col_idx in range(start_col + 1, end_col + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 16
            ws.column_dimensions[get_column_letter(start_col)].width = self._excel_column_width_for_pixels(image_size)

            ws.cell(info_row, start_col, "")
            if start_col + 1 <= end_col:
                ws.merge_cells(start_row=info_row, start_column=start_col + 1, end_row=info_row, end_column=end_col)
            ws.cell(info_row, start_col + 1, f"商品ID：{ctx['product_id']}\n标题：{ctx['title']}")
            self._set_range_fill(ws, info_row, start_col, end_col, basic_fill)
            self._add_export_image(ws, ctx["image_data"], f"{get_column_letter(start_col)}{info_row}", image_size, image_refs, "商品")

            ws.merge_cells(start_row=category_row, start_column=start_col, end_row=category_row, end_column=end_col)
            ws.cell(category_row, start_col, f"商品分类：{ctx['category_label'] or '-'}")
            self._set_range_fill(ws, category_row, start_col, end_col, basic_fill)

            ws.merge_cells(start_row=spec_title_row, start_column=start_col, end_row=spec_title_row, end_column=end_col)
            ws.cell(spec_title_row, start_col, "规格数据")
            self._set_range_fill(ws, spec_title_row, start_col, end_col, spec_fill)
            ws.cell(spec_title_row, start_col).font = header_font

            for offset, (label, _key) in enumerate(spec_fields):
                cell = ws.cell(spec_header_row, start_col + offset, label)
                cell.fill = spec_fill
                cell.font = header_font

            current_row = spec_first_row
            if specs:
                for spec in specs:
                    for offset, (_label, key) in enumerate(spec_fields):
                        col_idx = start_col + offset
                        cell = ws.cell(current_row, col_idx)
                        if key == "_image":
                            self._set_square_image_cell(ws, current_row, col_idx, image_size)
                            self._add_export_image(ws, spec["spec_image_data"], f"{get_column_letter(col_idx)}{current_row}", image_size, image_refs, "规格")
                        elif key in ("cost", "final_price", "gross_profit"):
                            cell.value = self._fmt_export_money(spec[key])
                        elif key == "margin_pct":
                            cell.value = self._fmt_export_pct(spec[key])
                        elif key == "weight":
                            cell.value = f"{self._safe_export_float(spec[key]):.2f}%"
                        elif key == "order_count":
                            cell.value = f"{int(spec[key])}单"
                        elif key == "refund_count":
                            cell.value = f"{int(spec[key])}单" if spec[key] else "无"
                        elif key == "refund_ratio":
                            cell.value = self._fmt_export_pct(spec[key]) if spec[key] else "无"
                        else:
                            cell.value = spec.get(key, "")
                        cell.fill = white_fill
                    current_row += 1
            else:
                ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=end_col)
                ws.cell(current_row, start_col, "暂无规格数据")
                self._set_range_fill(ws, current_row, start_col, end_col, white_fill)
                current_row += 1

            ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=end_col)
            ws.cell(current_row, start_col, "促销活动优惠情况")
            self._set_range_fill(ws, current_row, start_col, end_col, promo_fill)
            ws.cell(current_row, start_col).font = header_font
            current_row += 1

            promotion_text = self._promotion_summary_for_export(ctx)
            ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=end_col)
            ws.cell(current_row, start_col, promotion_text)
            self._set_range_fill(ws, current_row, start_col, end_col, white_fill)
            ws.row_dimensions[current_row].height = max(ws.row_dimensions[current_row].height or 0, 28)
            current_row += 1

            ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=end_col)
            ws.cell(current_row, start_col, "投产比分析")
            self._set_range_fill(ws, current_row, start_col, end_col, roi_fill)
            ws.cell(current_row, start_col).font = header_font
            current_row += 1

            roi_text = self._roi_summary_for_export(ctx)
            ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=end_col)
            ws.cell(current_row, start_col, roi_text)
            self._set_range_fill(ws, current_row, start_col, end_col, white_fill)
            roi_lines = max(1, len(roi_text) // 48 + 1, roi_text.count("；") + 1)
            ws.row_dimensions[current_row].height = max(ws.row_dimensions[current_row].height or 0, 24 + roi_lines * 18)
            current_row += 1

            ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=end_col)
            ws.cell(current_row, start_col, f"上一周操作记录（{record_range}）")
            self._set_range_fill(ws, current_row, start_col, end_col, operation_fill)
            ws.cell(current_row, start_col).font = header_font
            current_row += 1

            spec_names_for_bold = [
                str(spec.get("spec_name") or "").strip()
                for spec in specs
                if str(spec.get("spec_name") or "").strip()
            ]
            operation_row_styles = []
            if record_lines:
                for record_info in record_lines:
                    date_text, content = self._operation_record_export_cells(record_info)
                    ws.cell(current_row, start_col, date_text)
                    if start_col + 1 <= end_col:
                        ws.merge_cells(start_row=current_row, start_column=start_col + 1, end_row=current_row, end_column=end_col)
                    ws.cell(current_row, start_col + 1, content)
                    self._set_range_fill(ws, current_row, start_col, end_col, white_fill)
                    ws.row_dimensions[current_row].height = max(ws.row_dimensions[current_row].height or 0, 30)
                    record_text = str(record_info.get("text") or "")
                    contains_spec_name = any(spec_name and spec_name in record_text for spec_name in spec_names_for_bold)
                    operation_row_styles.append((
                        current_row,
                        self._operation_record_color_for_export(record_info["date"]),
                        bool(record_info.get("has_spec") or contains_spec_name),
                    ))
                    current_row += 1
            else:
                ws.cell(current_row, start_col, "-")
                if start_col + 1 <= end_col:
                    ws.merge_cells(start_row=current_row, start_column=start_col + 1, end_row=current_row, end_column=end_col)
                ws.cell(current_row, start_col + 1, "上一周无操作记录")
                self._set_range_fill(ws, current_row, start_col, end_col, white_fill)
                ws.row_dimensions[current_row].height = max(ws.row_dimensions[current_row].height or 0, 30)
                operation_row_styles.append((current_row, "6B7280", False))
                current_row += 1

            block_end_row = current_row - 1
            self._style_product_specs_export_range(ws, block_start_row, block_end_row, start_col, end_col)
            for row_idx, color, bold in operation_row_styles:
                if block_start_row <= row_idx <= block_end_row:
                    self._style_operation_record_export_row(ws, row_idx, end_col, color, bold, start_col)
            max_written_row = max(max_written_row, block_end_row)

        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{get_column_letter(ws.max_column)}{max_written_row}"
        ws._image_stream_refs = image_refs

    def _write_product_specs_export_sheet(self, wb):
        ws = wb.create_sheet("商品规格售卖情况")
        image_size = self._export_product_image_size()
        image_refs = []
        products = self._products_for_specs_export()
        fills = {
            "title": PatternFill("solid", fgColor="DCEBFF"),
            "section": PatternFill("solid", fgColor="EAF4FF"),
            "spec": PatternFill("solid", fgColor="F3E8FF"),
            "roi": PatternFill("solid", fgColor="FFF3D6"),
            "white": PatternFill("solid", fgColor="FFFFFF"),
        }
        header_font = Font(name="SimHei", color="000000", bold=True, size=14)
        normal_font = Font(name="SimHei", color="000000", size=13)
        input_fill = PatternFill("solid", fgColor="E8F7EF")
        formula_fill = PatternFill("solid", fgColor="FFF7ED")
        headers = ["规格图", "规格名称", "规格编码", "成本", "实际价格", "毛利率", "毛利润", "权重", "单量", "退款订单", "退款占比"]
        for col, width in enumerate([14, 18, 18, 12, 12, 12, 12, 11, 10, 10, 12], start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
        if not products:
            ws.append(["本次未选择需要详细展示的链接"])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            ws["A1"].font = header_font
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            self._style_product_specs_export_range(ws, 1, 1, 1, len(headers))
            return
        current_row = 1
        for product in products:
            ctx = self._product_export_context(product)
            specs = ctx.get("specs", [])
            block_start = current_row
            self._set_square_image_cell(ws, current_row, 1, image_size)
            ws.cell(current_row, 2, "商品ID")
            ws.cell(current_row, 3, ctx["product_id"])
            ws.cell(current_row, 4, "商品标题")
            ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=11)
            ws.cell(current_row, 5, ctx["title"])
            self._set_range_fill(ws, current_row, 1, 11, fills["title"])
            self._add_export_image(ws, ctx["image_data"], f"A{current_row}", image_size, image_refs, "商品")
            current_row += 1
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
            ws.cell(current_row, 1, f"链接备注：{ctx['product_memo'] or '-'}")
            self._set_range_fill(ws, current_row, 1, 11, fills["white"])
            current_row += 1
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
            ws.cell(current_row, 1, f"促销活动优惠情况：{self._promotion_summary_for_export(ctx)}")
            self._set_range_fill(ws, current_row, 1, 11, fills["white"])
            current_row += 1
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
            ws.cell(current_row, 1, f"商品分类：{ctx['category_label'] or '-'}")
            self._set_range_fill(ws, current_row, 1, 11, fills["white"])
            current_row += 1
            ws.cell(current_row, 1, "规格售卖情况")
            self._set_range_fill(ws, current_row, 1, 11, fills["spec"])
            ws.cell(current_row, 1).font = header_font
            current_row += 1
            for offset, label in enumerate(headers, start=1):
                cell = ws.cell(current_row, offset, label)
                cell.fill = fills["spec"]
                cell.font = header_font
            current_row += 1
            spec_first_row = current_row
            if specs:
                for spec in specs:
                    self._set_square_image_cell(ws, current_row, 1, image_size)
                    self._add_export_image(ws, spec.get("spec_image_data"), f"A{current_row}", image_size, image_refs, "规格")
                    values = {
                        2: spec.get("spec_name", ""),
                        3: spec.get("spec_code", ""),
                        4: self._safe_export_float(spec.get("cost")),
                        5: self._safe_export_float(spec.get("final_price")),
                        6: f'=IF(E{current_row}>0,(E{current_row}-D{current_row})/E{current_row},0)',
                        7: f"=E{current_row}-D{current_row}",
                        8: self._safe_export_float(spec.get("weight")) / 100,
                        9: int(spec.get("order_count") or 0),
                        10: int(spec.get("refund_count") or 0),
                        11: f'=IF(I{current_row}>0,J{current_row}/I{current_row},0)',
                    }
                    for col, value in values.items():
                        cell = ws.cell(current_row, col, value)
                        cell.fill = input_fill if col in (4, 5, 8, 9, 10) else (formula_fill if col in (6, 7, 11) else fills["white"])
                        cell.font = normal_font
                    for col in (4, 5, 7):
                        ws.cell(current_row, col).number_format = '¥#,##0.00'
                    for col in (6, 8, 11):
                        ws.cell(current_row, col).number_format = '0.00%'
                    current_row += 1
            else:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
                ws.cell(current_row, 1, "暂无规格数据")
                self._set_range_fill(ws, current_row, 1, 11, fills["white"])
                current_row += 1
            spec_last_row = max(spec_first_row, current_row - 1)
            summary_row = current_row
            ws.cell(summary_row, 1, "综合")
            ws.cell(summary_row, 2, "综合毛利率")
            ws.cell(summary_row, 3, f"=IF(SUM(H{spec_first_row}:H{spec_last_row})>0,SUMPRODUCT(F{spec_first_row}:F{spec_last_row},H{spec_first_row}:H{spec_last_row})/SUM(H{spec_first_row}:H{spec_last_row}),0)")
            ws.cell(summary_row, 4, "加权客单价")
            ws.cell(summary_row, 5, f"=IF(SUM(H{spec_first_row}:H{spec_last_row})>0,SUMPRODUCT(E{spec_first_row}:E{spec_last_row},H{spec_first_row}:H{spec_last_row})/SUM(H{spec_first_row}:H{spec_last_row}),0)")
            ws.cell(summary_row, 6, "单笔毛利")
            ws.cell(summary_row, 7, f"=IF(SUM(H{spec_first_row}:H{spec_last_row})>0,SUMPRODUCT(G{spec_first_row}:G{spec_last_row},H{spec_first_row}:H{spec_last_row})/SUM(H{spec_first_row}:H{spec_last_row}),0)")
            self._set_range_fill(ws, summary_row, 1, 11, fills["section"])
            ws.cell(summary_row, 3).number_format = '0.00%'
            ws.cell(summary_row, 5).number_format = '¥#,##0.00'
            ws.cell(summary_row, 7).number_format = '¥#,##0.00'
            current_row += 2
            ws.cell(current_row, 1, "投产比分析")
            self._set_range_fill(ws, current_row, 1, 4, fills["roi"])
            ws.cell(current_row, 1).font = header_font
            current_row += 1
            for col, label in enumerate(["指标", "数值", "说明", "可编辑"], start=1):
                ws.cell(current_row, col, label).fill = fills["roi"]
                ws.cell(current_row, col).font = header_font
            current_row += 1
            def metric(label, value="", note="", editable=False, number_format=None):
                nonlocal current_row
                row = current_row
                ws.cell(row, 1, label)
                ws.cell(row, 2, value)
                ws.cell(row, 3, note)
                ws.cell(row, 4, "是" if editable else "")
                for col in range(1, 5):
                    ws.cell(row, col).fill = input_fill if editable and col == 2 else (formula_fill if str(value).startswith("=") and col == 2 else fills["white"])
                    ws.cell(row, col).font = normal_font
                if number_format:
                    ws.cell(row, 2).number_format = number_format
                current_row += 1
                return row
            metric("推广方式", ctx["promotion_mode"])
            is_bid_mode = "成交出价" in str(ctx.get("roi_input_mode") or "")
            if is_bid_mode:
                bid_row = metric("成交出价", self._safe_export_float(ctx["transaction_bid"]), "可手动修改", True, '¥#,##0.00')
                roi_row = None
            else:
                roi_row = metric("当前投产", self._safe_export_float(ctx["current_roi"]), "可手动修改", True, "0.00")
                bid_row = None
            margin_row = metric("综合毛利率", f"=C{summary_row}", "来自上方规格权重公式", False, "0.00%")
            metric("毛保本投产", f"=IF(B{margin_row}>0,1/B{margin_row},0)", "", False, "0.00")
            return_factor = f"(1-{self._safe_export_float(ctx.get('return_rate'))}/100)"
            net_base = f"(B{margin_row}*{return_factor}-0.006)"
            net_be_row = metric("净保本投产", f"=IF({net_base}>0,1/{net_base},0)", "", False, "0.00")
            effective_roi_formula = f"IF(B{bid_row}>0,E{summary_row}/B{bid_row},0)" if is_bid_mode else f"B{roi_row}"
            metric("投产倍数", f"=IF(B{net_be_row}>0,{effective_roi_formula}/B{net_be_row},0)", "", False, "0.00")
            metric("推广占比", f"=IF({effective_roi_formula}>0,1/{effective_roi_formula},0)", "", False, "0.00%")
            if ctx.get("is_natural_flow"):
                net_rate_formula = f"={net_base}"
            elif is_bid_mode:
                net_rate_formula = f"=IF(E{summary_row}>0,(G{summary_row}*{return_factor}-E{summary_row}*0.006-B{bid_row})/E{summary_row},0)"
            else:
                net_rate_formula = f"={net_base}-IF(B{roi_row}>0,1/B{roi_row},0)"
            net_rate_row = metric("净利率", net_rate_formula, "修改当前出价或上方规格后自动变化", False, "0.00%")
            spend_row = metric("预计花费", 100, "默认100限额，可改成任意花费", True, '¥#,##0.00')
            if is_bid_mode:
                expected_profit_formula = f"=IF(B{bid_row}>0,(B{spend_row}/B{bid_row})*E{summary_row}*{return_factor}*(B{margin_row}-0.006)-B{spend_row},0)"
            else:
                expected_profit_formula = f"=B{spend_row}*{effective_roi_formula}*{return_factor}*(B{margin_row}-0.006)-B{spend_row}"
            metric("预计盈亏", expected_profit_formula, "按主界面预计盈亏口径，引用预计花费", False, '¥#,##0.00')
            current_row += 1
            record_range, record_lines = self._record_briefs_for_export(ctx["sys_id"])
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
            ws.cell(current_row, 1, f"上一周操作记录（{record_range}）")
            self._set_range_fill(ws, current_row, 1, 11, fills["section"])
            ws.cell(current_row, 1).font = header_font
            current_row += 1
            operation_styles = []
            if record_lines:
                spec_names = [str(spec.get("spec_name") or "").strip() for spec in specs]
                for record_info in record_lines:
                    date_text, content = self._operation_record_export_cells(record_info)
                    ws.cell(current_row, 1, date_text)
                    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=11)
                    ws.cell(current_row, 2, content)
                    self._set_range_fill(ws, current_row, 1, 11, fills["white"])
                    contains_spec_name = any(name and name in str(record_info.get("text") or "") for name in spec_names)
                    operation_styles.append((current_row, self._operation_record_color_for_export(record_info["date"]), bool(record_info.get("has_spec") or contains_spec_name)))
                    current_row += 1
            else:
                ws.cell(current_row, 1, "-")
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=11)
                ws.cell(current_row, 2, "上一周无操作记录")
                self._set_range_fill(ws, current_row, 1, 11, fills["white"])
                operation_styles.append((current_row, "6B7280", False))
                current_row += 1
            block_end = current_row - 1
            self._style_product_specs_export_range(ws, block_start, block_end, 1, 11)
            for row_idx, color, bold in operation_styles:
                self._style_operation_record_export_row(ws, row_idx, 11, color, bold, 1)
            for row in range(block_start, block_end + 1):
                ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, 28)
                for col in range(1, 12):
                    cell = ws.cell(row, col)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    if not cell.font or cell.font.name is None:
                        cell.font = normal_font
            current_row += 2
        ws.freeze_panes = "A1"
        ws._image_stream_refs = image_refs

    def _default_margin_excel_path(self, folder, safe_store_name):
        today = datetime.now().strftime("%Y%m%d")
        file_path = os.path.join(folder, f"店铺毛利_{safe_store_name}_{today}.xlsx")
        suffix = 2
        while os.path.exists(file_path):
            file_path = os.path.join(folder, f"店铺毛利_{safe_store_name}_{today}_{suffix}.xlsx")
            suffix += 1
        return file_path

    def _export_margin_excel_from_button(self):
        return self.export_margin_excel()

    def export_margin_excel(self, dialog_parent=None, file_path=None, show_success=True, progress_title="导出Excel"):
        progress = None
        try:
            if isinstance(dialog_parent, bool):
                dialog_parent = None
            parent = dialog_parent if isinstance(dialog_parent, QWidget) else self
            if not OPENPYXL_AVAILABLE:
                QMessageBox.warning(parent, "缺少依赖", "请先安装 openpyxl 库：\npip install openpyxl")
                return False
            detail_map = StoreMarginExcelExporter.select_detail_products(
                parent, self.db, self.main_app, [(self.store_id, self.store_name)]
            )
            if detail_map is None:
                return False
            self.export_detail_product_ids = detail_map.get(int(self.store_id), set())
            safe_store_name = "".join(ch for ch in self.store_name if ch not in r'\/:*?"<>|').strip() or "店铺"
            if not file_path:
                folder = remembered_existing_directory(parent, self.db, "选择导出保存文件夹")
                if not folder:
                    return False
                file_path = self._default_margin_excel_path(folder, safe_store_name)
            if not file_path:
                return False
            if not file_path.lower().endswith(".xlsx"):
                file_path += ".xlsx"
            progress = QProgressDialog("正在准备导出...", "取消", 0, 100, parent)
            progress.setWindowTitle(progress_title)
            progress.setWindowModality(Qt.WindowModal)
            progress.setMinimumDuration(0)
            progress.setAutoClose(False)
            progress.setAutoReset(False)

            def update_progress(value, text):
                progress.setValue(int(value))
                progress.setLabelText(text)
                QApplication.processEvents()
                return not progress.wasCanceled()

            self._last_export_error = ""
            self._export_error_dialog_shown = False
            ok = self.export_margin_excel_to_path(file_path, update_progress)
            was_canceled = progress.wasCanceled()
            progress.setValue(100)
            progress.close()
            progress = None
            if ok:
                if show_success:
                    QMessageBox.information(parent, "导出成功", f"数据已导出到：\n{file_path}")
                return True
            if show_success and not was_canceled and not getattr(self, "_export_error_dialog_shown", False):
                errors = getattr(self, "_batch_export_errors", None)
                detail = "\n".join(errors[-3:]) if errors else getattr(self, "_last_export_error", "")
                if not detail:
                    detail = "导出流程已返回失败，但未提供具体错误。"
                QMessageBox.warning(parent, "导出失败", detail)
            return False
        except Exception as e:
            if progress is not None:
                try:
                    progress.close()
                except Exception:
                    pass
            detail = traceback.format_exc()
            parent = dialog_parent if isinstance(dialog_parent, QWidget) else self
            QMessageBox.critical(parent, "导出失败", f"导出时发生异常：\n{e}\n\n{detail}")
            return False

    def export_margin_excel_to_path(self, file_path, progress_callback=None):
        try:
            self._excel_export_image_cache = {}
            self._last_export_error = ""
            if progress_callback and not progress_callback(5, "正在创建 Excel 工作簿..."):
                return False
            wb = Workbook()
            if progress_callback and not progress_callback(18, "正在写入过往数据分析..."):
                return False
            self._write_historical_export_sheet(wb)
            if progress_callback and not progress_callback(42, "正在写入店铺商品权重..."):
                return False
            self._write_orders_export_sheet(wb)
            if progress_callback and not progress_callback(68, "正在写入商品规格售卖情况和缩略图..."):
                return False
            self._write_product_specs_export_sheet(wb)
            if progress_callback and not progress_callback(92, "正在保存 Excel 文件..."):
                return False
            wb.save(file_path)
            if progress_callback:
                progress_callback(100, "导出完成")
            return True
        except Exception as e:
            self._last_export_error = traceback.format_exc()
            if hasattr(self, "_batch_export_errors"):
                self._batch_export_errors.append(self._last_export_error)
            else:
                self._export_error_dialog_shown = True
                QMessageBox.warning(self, "导出失败", f"导出时发生错误：\n{str(e)}")
            return False
        finally:
            self._excel_export_image_cache = {}

    def calculate_week_comparison(self, records):
        """计算并显示周环比变化"""
        if len(records) < 2:
            for col in range(20):
                item = QTableWidgetItem("暂无上周数据")
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                item.setBackground(QColor("#f5f5f5"))
                self.week_table.setItem(0, col, item)
            return

        current = records[-1]
        previous = records[-2]

        current_net_profit = current[17] if current[17] else 0
        previous_net_profit = previous[17] if previous[17] else 0
        current_net_margin = current[18] if current[18] else 0
        previous_net_margin = previous[18] if previous[18] else 0

        current_daily = 0
        if current[0] and current[1]:
            try:
                from datetime import datetime
                start_dt = datetime.strptime(current[0], "%Y-%m-%d")
                end_dt = datetime.strptime(current[1], "%Y-%m-%d")
                days = max(1, (end_dt - start_dt).days + 1)
                current_daily = current_net_profit / days if days > 0 else 0
            except:
                pass

        previous_daily = 0
        if previous[0] and previous[1]:
            try:
                from datetime import datetime
                start_dt = datetime.strptime(previous[0], "%Y-%m-%d")
                end_dt = datetime.strptime(previous[1], "%Y-%m-%d")
                days = max(1, (end_dt - start_dt).days + 1)
                previous_daily = previous_net_profit / days if days > 0 else 0
            except:
                pass

        GREEN = QColor("#27ae60")
        RED = QColor("#e74c3c")
        GRAY = QColor("#999999")

        for col in range(20):
            item = QTableWidgetItem()
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)

            if col == 0:
                item.setText("较上周")
                item.setBackground(QColor("#e8e8e8"))
            elif col == 1:
                if previous[2] and previous[2] != 0:
                    change = ((current[2] or 0) - (previous[2] or 0)) / abs(previous[2]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 2:
                if (previous[3] or 0) != 0:
                    change = ((current[3] or 0) - (previous[3] or 0)) / abs(previous[3]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 3:
                if (previous[4] or 0) != 0:
                    change = ((current[4] or 0) - (previous[4] or 0)) / abs(previous[4]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 4:
                change = (current[11] or 0) - (previous[11] or 0)
                icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                item.setText(f"{icon} {abs(change):.1f}%")
                item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
            elif col == 5:
                if (previous[5] or 0) != 0:
                    change = ((current[5] or 0) - (previous[5] or 0)) / abs(previous[5]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(RED if change > 0 else GREEN if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 6:
                change = (current[12] or 0) - (previous[12] or 0)
                icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                item.setText(f"{icon} {abs(change):.1f}%")
                item.setForeground(RED if change > 0 else GREEN if change < 0 else GRAY)
            elif col == 7:
                if previous[6] and previous[6] != 0:
                    change = ((current[6] or 0) - (previous[6] or 0)) / abs(previous[6]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(RED if change > 0 else GREEN if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 8:
                change = (current[13] or 0) - (previous[13] or 0)
                icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                item.setText(f"{icon} {abs(change):.1f}%")
                item.setForeground(RED if change > 0 else GREEN if change < 0 else GRAY)
            elif col == 9:
                if (previous[14] or 0) != 0:
                    change = ((current[14] or 0) - (previous[14] or 0)) / abs(previous[14]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 10:
                if (previous[7] or 0) != 0:
                    change = ((current[7] or 0) - (previous[7] or 0)) / abs(previous[7]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 11:
                change = (current[15] or 0) - (previous[15] or 0)
                icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                item.setText(f"{icon} {abs(change):.1f}%")
                item.setForeground(RED if change > 0 else GREEN if change < 0 else GRAY)
            elif col == 12:
                if (previous[16] or 0) != 0:
                    change = ((current[16] or 0) - (previous[16] or 0)) / abs(previous[16]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(RED if change > 0 else GREEN if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 13:
                if (previous[8] or 0) != 0:
                    change = ((current[8] or 0) - (previous[8] or 0)) / abs(previous[8]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(RED if change > 0 else GREEN if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 14:
                if (previous[9] or 0) != 0:
                    change = ((current[9] or 0) - (previous[9] or 0)) / abs(previous[9]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(RED if change > 0 else GREEN if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 15:
                if (previous[10] or 0) != 0:
                    change = ((current[10] or 0) - (previous[10] or 0)) / abs(previous[10]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 16:
                if previous_net_profit != 0:
                    change = (current_net_profit - previous_net_profit) / abs(previous_net_profit) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 17:
                change = current_net_margin - previous_net_margin
                icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                item.setText(f"{icon} {abs(change):.1f}%")
                item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
            elif col == 18:
                if (previous[19] or 0) != 0:
                    change = ((current[19] or 0) - (previous[19] or 0)) / abs(previous[19]) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)
            elif col == 19:
                if previous_daily != 0:
                    change = (current_daily - previous_daily) / abs(previous_daily) * 100
                    icon = "↑" if change > 0 else "↓" if change < 0 else "→"
                    item.setText(f"{icon} {abs(change):.1f}%")
                    item.setForeground(GREEN if change > 0 else RED if change < 0 else GRAY)
                else:
                    item.setText("→ 0.0%")
                    item.setForeground(GRAY)

            self.week_table.setItem(0, col, item)

    def _fmt_money(self, value):
        try:
            return f"{float(value or 0):.2f}"
        except (TypeError, ValueError):
            return "0.00"

    def _fmt_number(self, value):
        try:
            num = float(value or 0)
            return int(num) if num == int(num) else round(num, 2)
        except (TypeError, ValueError):
            return 0

    def _days_between(self, start_date, end_date):
        if not start_date or not end_date:
            return 1
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            return max(1, (end_dt - start_dt).days + 1)
        except Exception:
            return 1

    def _manual_margin_record_to_dict(self, record):
        start_date = record[0] or ""
        end_date = record[1] or ""
        days = self._days_between(start_date, end_date)
        net_profit = float(record[17] or 0)
        return {
            "date_range": f"{start_date}~{end_date}" if start_date or end_date else "",
            "days": days,
            "actual_orders": self._fmt_number(record[2]),
            "actual_amount": round(float(record[3] or 0), 2),
            "gross_profit": round(float(record[4] or 0), 2),
            "gross_margin_rate": round(float(record[11] or 0), 2),
            "refund_amount": round(float(record[5] or 0), 2),
            "refund_rate_by_amount": round(float(record[12] or 0), 2),
            "refund_orders": self._fmt_number(record[6]),
            "refund_rate_by_orders": round(float(record[13] or 0), 2),
            "unit_price": round(float(record[14] or 0), 2),
            "promotion_fee": round(float(record[7] or 0), 2),
            "promotion_ratio": round(float(record[15] or 0), 2),
            "tech_fee": round(float(record[16] or 0), 2),
            "deduction": round(float(record[8] or 0), 2),
            "other_service": round(float(record[9] or 0), 2),
            "other": round(float(record[10] or 0), 2),
            "net_profit": round(net_profit, 2),
            "net_margin_rate": round(float(record[18] or 0), 2),
            "profit_per_order": round(float(record[19] or 0), 2),
            "daily_profit": round(net_profit / days if days > 0 else 0, 2),
        }

    def _change_summary(self, current, previous):
        if not previous:
            return {"available": False, "note": "无上一周对比数据"}
        metrics = [
            ("actual_orders", "实发订单", "pct"),
            ("actual_amount", "实发金额", "pct"),
            ("gross_profit", "毛利润", "pct"),
            ("gross_margin_rate", "毛利率", "point"),
            ("refund_amount", "退款金额", "pct"),
            ("refund_rate_by_amount", "金额退款率", "point"),
            ("refund_orders", "退款订单", "pct"),
            ("refund_rate_by_orders", "订单退款率", "point"),
            ("unit_price", "件单价", "pct"),
            ("promotion_fee", "推广费", "pct"),
            ("promotion_ratio", "推广占比", "point"),
            ("net_profit", "净利润", "pct"),
            ("net_margin_rate", "净利率", "point"),
            ("profit_per_order", "单笔利润", "pct"),
            ("daily_profit", "日盈亏", "pct"),
        ]
        changes = []
        for key, label, mode in metrics:
            curr = float(current.get(key) or 0)
            prev = float(previous.get(key) or 0)
            diff = curr - prev
            item = {"metric": label, "current": round(curr, 2), "previous": round(prev, 2), "diff": round(diff, 2)}
            if mode == "point":
                item["change"] = f"{diff:+.2f}个百分点"
            elif prev != 0:
                item["change"] = f"{(diff / abs(prev) * 100):+.2f}%"
            else:
                item["change"] = "上周为0，无法计算百分比"
            changes.append(item)
        return {"available": True, "items": changes}

    def _build_historical_margin_summary(self, records):
        parsed = [self._manual_margin_record_to_dict(r) for r in records]
        total_amount = sum(float(r["actual_amount"] or 0) for r in parsed)
        total_orders = sum(float(r["actual_orders"] or 0) for r in parsed)
        total_net_profit = sum(float(r["net_profit"] or 0) for r in parsed)
        total_refund_amount = sum(float(r["refund_amount"] or 0) for r in parsed)
        total_promotion_fee = sum(float(r["promotion_fee"] or 0) for r in parsed)
        profitable = [r for r in parsed if float(r["net_profit"] or 0) > 0]
        loss = [r for r in parsed if float(r["net_profit"] or 0) < 0]
        best = max(parsed, key=lambda r: float(r["net_profit"] or 0)) if parsed else None
        worst = min(parsed, key=lambda r: float(r["net_profit"] or 0)) if parsed else None
        return {
            "period_count": len(parsed),
            "total_actual_orders": self._fmt_number(total_orders),
            "total_actual_amount": round(total_amount, 2),
            "total_net_profit": round(total_net_profit, 2),
            "overall_net_margin_rate": round((total_net_profit / total_amount * 100) if total_amount else 0, 2),
            "total_refund_amount": round(total_refund_amount, 2),
            "overall_refund_rate_by_amount": round((total_refund_amount / total_amount * 100) if total_amount else 0, 2),
            "total_promotion_fee": round(total_promotion_fee, 2),
            "overall_promotion_ratio": round((total_promotion_fee / total_amount * 100) if total_amount else 0, 2),
            "profitable_period_count": len(profitable),
            "loss_period_count": len(loss),
            "best_period": best,
            "worst_period": worst,
            "recent_periods": parsed[-6:],
        }

    def _parse_date_safe(self, value):
        try:
            return datetime.strptime(value, "%Y-%m-%d").date() if value else None
        except Exception:
            return None

    def _build_operation_record_range(self, current, previous):
        starts = [self._parse_date_safe(current.get("date_range", "").split("~")[0] if current else "")]
        ends = [self._parse_date_safe(current.get("date_range", "").split("~")[-1] if current else "")]
        if previous:
            starts.append(self._parse_date_safe(previous.get("date_range", "").split("~")[0]))
            ends.append(self._parse_date_safe(previous.get("date_range", "").split("~")[-1]))
        starts = [d for d in starts if d]
        ends = [d for d in ends if d]
        if not starts or not ends:
            return None, None, "无有效日期范围"
        start_date = min(starts)
        end_date = max(ends)
        return start_date, end_date, f"{start_date.strftime('%Y-%m-%d')}~{end_date.strftime('%Y-%m-%d')}"

    def _record_text_for_ai(self, record):
        text = str(record.get("text", "") or "").strip()
        changes = record.get("changes", []) or []
        change_texts = []
        if isinstance(changes, list):
            for change in changes:
                if isinstance(change, dict):
                    change_text = str(change.get("text", "") or "").strip()
                    metric = str(change.get("metric", "") or "").strip()
                    old = str(change.get("old", "") or "").strip()
                    new = str(change.get("new", "") or "").strip()
                    if change_text:
                        change_texts.append(change_text)
                    elif metric or old or new:
                        change_texts.append(f"{metric}: {old} -> {new}".strip())
                elif change:
                    change_texts.append(str(change))
        if text and change_texts:
            return f"{text}；结构化变化：{'；'.join(change_texts)}"
        if change_texts:
            return "；".join(change_texts)
        return text

    def _build_product_operation_records(self, current, previous):
        start_date, end_date, range_text = self._build_operation_record_range(current, previous)
        if not start_date or not end_date:
            return {"available": False, "range": range_text, "items": [], "note": "无有效日期范围，未读取商品操作记录"}

        products = self.db.safe_fetchall(
            "SELECT id, name, title FROM products WHERE store_id=? AND COALESCE(is_archived, 0)=0 ORDER BY sort_order",
            (self.store_id,)
        )
        if not products:
            return {"available": False, "range": range_text, "items": [], "note": "当前店铺没有商品"}

        product_map = {int(row[0]): {"product_id": row[1] or "", "product_title": row[2] or ""} for row in products}
        rows = self.db.safe_fetchall(
            "SELECT product_id, year, month, day, records_json FROM records WHERE product_id IN ({})".format(
                ",".join(["?"] * len(product_map))
            ),
            tuple(product_map.keys())
        )
        items = []
        for product_sys_id, year, month, day, records_json in rows:
            try:
                record_date = datetime(int(year), int(month), int(day)).date()
            except Exception:
                continue
            if record_date < start_date or record_date > end_date:
                continue
            try:
                day_records = json.loads(records_json) if records_json else []
            except Exception:
                day_records = []
            product_info = product_map.get(int(product_sys_id), {})
            for record in day_records:
                if not isinstance(record, dict):
                    continue
                content = self._record_text_for_ai(record)
                if not content:
                    continue
                items.append({
                    "date": record_date.strftime("%Y-%m-%d"),
                    "time": record.get("time", "") or "",
                    "product_id": product_info.get("product_id", ""),
                    "product_title": product_info.get("product_title", ""),
                    "content": content,
                })

        items.sort(key=lambda x: (x["date"], x["time"], x["product_id"]))
        return {
            "available": bool(items),
            "range": range_text,
            "items": items,
            "note": "本周+上周范围内的当前店铺商品操作记录" if items else "本周+上周范围内没有商品操作记录",
        }

    def _format_ai_report_spec(self, spec_code, spec_name, display_mode):
        spec_code = str(spec_code or "").strip()
        spec_name = str(spec_name or "").strip()
        if not spec_code and not spec_name:
            return ""
        if display_mode == "name":
            return f"【{spec_name or spec_code}】"
        return spec_code

    def _build_sales_structure_summary(self):
        spec_display_mode = self._get_spec_display_mode("store_margin_main_spec_display")
        rows = self.db.safe_fetchall("""
            SELECT io.product_id, io.spec_code,
                   SUM(COALESCE(io.order_count, 0)) AS order_count,
                   SUM(COALESCE(io.refund_count, 0)) AS refund_count,
                   p.id, p.title, p.coupon_amount, p.new_customer_discount,
                   ps.spec_name, ps.sale_price, cl.cost_price
            FROM imported_orders io
            LEFT JOIN products p ON p.store_id = io.store_id AND p.name = io.product_id
            LEFT JOIN product_specs ps ON ps.product_id = p.id AND ps.spec_code = io.spec_code
            LEFT JOIN cost_library cl ON cl.spec_code = io.spec_code
            WHERE io.store_id=?
            GROUP BY io.product_id, io.spec_code
        """, (self.store_id,))
        if not rows:
            return {"available": False, "note": "无订单规格毛利权重数据"}

        total_orders = sum(int(row[2] or 0) for row in rows)
        total_refunds = sum(int(row[3] or 0) for row in rows)
        total_amount = 0.0
        product_map = {}
        specs = []

        for row in rows:
            product_id, spec_code, order_count, refund_count, sys_id, title, coupon, new_customer, spec_name, sale_price, cost_price = row
            if sys_id and not spec_name:
                continue
            order_count = int(order_count or 0)
            refund_count = int(refund_count or 0)
            sale_price = float(sale_price or 0)
            cost_price = float(cost_price or 0)
            spec_display = self._format_ai_report_spec(spec_code, spec_name, spec_display_mode)
            store_discount, _rule = self.db.calculate_store_discount(self.store_id, sale_price)
            discount = max(float(coupon or 0), float(new_customer or 0), float(store_discount or 0))
            final_price = max(0, sale_price - discount)
            amount = sale_price * order_count
            total_amount += amount
            margin_rate = None
            gross_profit = None
            if final_price > 0 and cost_price > 0:
                margin_rate = (final_price - cost_price) / final_price * 100
                gross_profit = (final_price - cost_price) * order_count
            refund_rate = (refund_count / order_count * 100) if order_count else 0
            item = {
                "product_id": product_id or "",
                "product_title": title or "",
                "spec": spec_display,
                "orders": order_count,
                "order_share": round((order_count / total_orders * 100) if total_orders else 0, 2),
                "sales_amount": round(amount, 2),
                "sale_price": round(sale_price, 2),
                "cost_price": round(cost_price, 2),
                "margin_rate": round(margin_rate, 2) if margin_rate is not None else None,
                "gross_profit_estimate": round(gross_profit, 2) if gross_profit is not None else None,
                "refund_orders": refund_count,
                "refund_rate": round(refund_rate, 2),
            }
            specs.append(item)

            prod = product_map.setdefault(product_id or "", {
                "product_id": product_id or "",
                "product_title": title or "",
                "orders": 0,
                "sales_amount": 0.0,
                "refund_orders": 0,
                "specs": [],
            })
            prod["orders"] += order_count
            prod["sales_amount"] += amount
            prod["refund_orders"] += refund_count
            prod["specs"].append(item)

        products = []
        for prod in product_map.values():
            prod_specs = sorted(prod["specs"], key=lambda x: x["orders"], reverse=True)
            max_refund_spec = max(prod_specs, key=lambda x: x["refund_rate"]) if prod_specs else None
            products.append({
                "product_id": prod["product_id"],
                "product_title": prod["product_title"],
                "orders": prod["orders"],
                "order_share": round((prod["orders"] / total_orders * 100) if total_orders else 0, 2),
                "sales_amount": round(prod["sales_amount"], 2),
                "refund_orders": prod["refund_orders"],
                "refund_rate": round((prod["refund_orders"] / prod["orders"] * 100) if prod["orders"] else 0, 2),
                "main_spec": prod_specs[0]["spec"] if prod_specs else "",
                "max_refund_spec": max_refund_spec["spec"] if max_refund_spec and max_refund_spec["refund_orders"] > 0 else "",
                "top_specs": prod_specs[:5],
            })

        return {
            "available": True,
            "spec_display_mode": "规格名称" if spec_display_mode == "name" else "规格编码",
            "spec_display_rule": "所有规格相关信息使用【规格名称】展示，不输出规格编码" if spec_display_mode == "name" else "所有规格相关信息使用规格编码展示",
            "total_orders": total_orders,
            "total_sales_amount": round(total_amount, 2),
            "total_refund_orders": total_refunds,
            "overall_refund_rate": round((total_refunds / total_orders * 100) if total_orders else 0, 2),
            "top_products": sorted(products, key=lambda x: x["orders"], reverse=True)[:10],
            "top_specs": sorted(specs, key=lambda x: x["orders"], reverse=True)[:20],
        }

    def _build_store_report_context(self, records):
        current = self._manual_margin_record_to_dict(records[-1])
        previous = self._manual_margin_record_to_dict(records[-2]) if len(records) >= 2 else None
        order_range = self.lbl_order_range.text() if hasattr(self, "lbl_order_range") else ""
        return {
            "store_name": self.store_name,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "data_notes": [
                "最新一周和上一周的财务指标按已发货/已收货维度统计。",
                "订单规格毛利权重表按已拼单订单统计，订单数量和财务表可能存在误差，分析只看大方向。",
                "财务数据的退款订单和退款率按本周申请售后维度统计，可能包含很久之前成交但本周退款的顾客。",
                "订单规格结构主要反映最新一周拼单顾客，不代表历史成交客群。",
            ],
            "current_week": current,
            "previous_week": previous,
            "week_comparison": self._change_summary(current, previous),
            "historical_summary": self._build_historical_margin_summary(records),
            "sales_structure": self._build_sales_structure_summary(),
            "product_operation_records": self._build_product_operation_records(current, previous),
            "current_order_range_label": order_range,
        }

    def _estimate_tokens(self, text):
        if not text:
            return 0
        cjk_chars = 0
        other_chars = 0
        for ch in text:
            if "\u4e00" <= ch <= "\u9fff":
                cjk_chars += 1
            elif not ch.isspace():
                other_chars += 1
        return int(cjk_chars * 1.15 + other_chars / 4) + 1

    def _estimate_messages_tokens(self, messages):
        try:
            text = json.dumps(messages, ensure_ascii=False)
        except Exception:
            text = str(messages)
        return self._estimate_tokens(text)

    def _format_token_status(self, input_estimate, usage=None):
        usage = usage or {}
        prompt_tokens = usage.get("prompt_tokens")
        completion_tokens = usage.get("completion_tokens")
        total_tokens = usage.get("total_tokens")
        if prompt_tokens is not None or completion_tokens is not None or total_tokens is not None:
            prompt_text = str(prompt_tokens) if prompt_tokens is not None else f"约 {input_estimate}"
            completion_text = str(completion_tokens) if completion_tokens is not None else "接口未返回"
            if total_tokens is None and prompt_tokens is not None and completion_tokens is not None:
                total_tokens = prompt_tokens + completion_tokens
            total_text = str(total_tokens) if total_tokens is not None else "接口未返回"
            return f"输入 {prompt_text} tokens | 输出 {completion_text} tokens | 总计 {total_text} tokens"
        return f"输入约 {input_estimate} tokens | 输出 未生成 | 总计 未生成"

    def _show_debug_prompt_dialog(self, messages, input_tokens, usage=None):
        if self.ai_report_debug_dialog and self.ai_report_debug_dialog.isVisible():
            self.ai_report_debug_dialog.close()

        dialog = QDialog(self.ai_report_dialog or self)
        dialog.setWindowTitle("调试提示词")
        dialog.setModal(False)
        dialog.setWindowModality(Qt.NonModal)
        dialog.setAttribute(Qt.WA_DeleteOnClose, False)
        dialog.resize(900, 680)
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(12, 12, 12, 12)

        token_label = QLabel(self._format_token_status(input_tokens, usage))
        token_label.setStyleSheet("font-size: 13px; color: #2c3e50; padding: 6px; background: #f4f6f7; border-radius: 4px;")
        layout.addWidget(token_label)

        prompt_text = QPlainTextEdit()
        prompt_text.setReadOnly(True)
        prompt_text.setPlainText(json.dumps(messages, ensure_ascii=False, indent=2))
        prompt_text.setStyleSheet("font-size: 12px;")
        layout.addWidget(prompt_text, 1)

        btn_row = QHBoxLayout()
        btn_copy = QPushButton("复制提示词")
        btn_close = QPushButton("关闭")
        btn_copy.clicked.connect(lambda: QApplication.clipboard().setText(prompt_text.toPlainText()))
        btn_close.clicked.connect(dialog.close)
        btn_row.addStretch()
        btn_row.addWidget(btn_copy)
        btn_row.addWidget(btn_close)
        layout.addLayout(btn_row)
        dialog.finished.connect(lambda _=0: setattr(self, "ai_report_debug_dialog", None))
        self.ai_report_debug_dialog = dialog
        dialog.show()
        dialog.raise_()
        dialog.activateWindow()

    def open_ai_report_dialog(self):
        if self.ai_report_dialog and self.ai_report_dialog.isVisible():
            self.ai_report_dialog.raise_()
            self.ai_report_dialog.activateWindow()
            return

        api_key = self.db.get_setting("ai_api_key", "")
        if not api_key:
            QMessageBox.warning(self, "提示", "请先在 AI API 配置中填写 API Key。")
            return

        records = self.load_manual_data()
        if not records:
            QMessageBox.warning(self, "提示", "需要先录入至少一条店铺毛利数据，才能生成报告。")
            return

        context = self._build_store_report_context(records)
        current_range = context["current_week"].get("date_range") or "无"
        previous = context.get("previous_week")
        previous_range = previous.get("date_range") if previous else "无上一周数据"
        order_range = context.get("current_order_range_label") or "无订单数据"

        dialog = QDialog(self)
        dialog.setWindowTitle(f"AI店铺周报 - {self.store_name}")
        dialog.setModal(False)
        dialog.setWindowModality(Qt.NonModal)
        dialog.setAttribute(Qt.WA_DeleteOnClose, False)
        dialog.resize(920, 720)
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(12, 12, 12, 12)

        info = QLabel(
            f"店铺：{self.store_name}    最新一周：{current_range}    上一周：{previous_range}\n"
            f"{order_range}"
        )
        info.setWordWrap(True)
        info.setStyleSheet("font-size: 13px; color: #2c3e50; padding: 8px; background: #f4f6f7; border-radius: 4px;")
        layout.addWidget(info)

        note_label = QLabel("补充分析要求（可选）：")
        note_label.setStyleSheet("font-size: 13px; font-weight: bold;")
        layout.addWidget(note_label)

        note_input = QPlainTextEdit()
        note_input.setPlaceholderText("例如：重点看利润下滑原因、退款问题、主卖规格是否健康、下周怎么调整。")
        note_input.setMaximumHeight(90)
        layout.addWidget(note_input)

        result_label = QLabel("AI生成结果：")
        result_label.setStyleSheet("font-size: 13px; font-weight: bold;")
        layout.addWidget(result_label)

        result_text = QPlainTextEdit()
        result_text.setReadOnly(True)
        result_text.setPlaceholderText("点击“生成报告”后，报告会显示在这里。")
        result_text.setStyleSheet("font-size: 13px; line-height: 1.5;")
        layout.addWidget(result_text, 1)

        token_status_label = QLabel("")
        token_status_label.setStyleSheet("font-size: 12px; color: #7f8c8d; padding: 4px 2px;")
        layout.addWidget(token_status_label)

        btn_row = QHBoxLayout()
        btn_generate = QPushButton("生成报告")
        btn_debug = QPushButton("调试提示词")
        btn_copy = QPushButton("复制报告")
        btn_close = QPushButton("关闭")
        btn_copy.setEnabled(False)
        btn_generate.setStyleSheet("QPushButton { background-color: #16a085; color: white; font-weight: bold; padding: 8px 18px; border-radius: 4px; }")
        btn_debug.setStyleSheet("QPushButton { background-color: #7f8c8d; color: white; font-weight: bold; padding: 8px 18px; border-radius: 4px; }")
        btn_copy.setStyleSheet("QPushButton { background-color: #3498db; color: white; font-weight: bold; padding: 8px 18px; border-radius: 4px; } QPushButton:disabled { background-color: #bdc3c7; }")
        btn_close.setStyleSheet("QPushButton { padding: 8px 18px; border-radius: 4px; }")
        btn_row.addStretch()
        btn_row.addWidget(btn_generate)
        btn_row.addWidget(btn_debug)
        btn_row.addWidget(btn_copy)
        btn_row.addWidget(btn_close)
        layout.addLayout(btn_row)

        prompt_state = {"messages": [], "input_tokens": 0, "usage": None}

        def refresh_prompt_state(clear_usage=False):
            if clear_usage:
                prompt_state["usage"] = None
            messages = self._build_store_report_messages(context, note_input.toPlainText().strip())
            input_tokens = self._estimate_messages_tokens(messages)
            prompt_state["messages"] = messages
            prompt_state["input_tokens"] = input_tokens
            token_status_label.setText(self._format_token_status(input_tokens, prompt_state.get("usage")))
            return messages, input_tokens

        refresh_prompt_state()
        note_input.textChanged.connect(lambda: refresh_prompt_state(True))

        def generate_report():
            messages, input_tokens = refresh_prompt_state()
            prompt_state["usage"] = None
            token_status_label.setText(self._format_token_status(input_tokens))
            btn_generate.setEnabled(False)
            btn_debug.setEnabled(False)
            btn_copy.setEnabled(False)
            result_text.setPlainText("正在生成报告...")

            def on_success(report, usage):
                prompt_state["usage"] = usage or {}
                result_text.setPlainText(report)
                btn_copy.setEnabled(True)
                token_status_label.setText(self._format_token_status(input_tokens, prompt_state["usage"]))

            def on_finished():
                btn_generate.setEnabled(True)
                btn_debug.setEnabled(True)

            self._start_ai_store_report(messages, input_tokens, dialog, on_success, on_finished)

        def copy_report():
            text = result_text.toPlainText().strip()
            if text:
                QApplication.clipboard().setText(text)
                self.show_toast("报告已复制")

        btn_generate.clicked.connect(generate_report)
        btn_debug.clicked.connect(lambda: self._show_debug_prompt_dialog(prompt_state["messages"], prompt_state["input_tokens"], prompt_state.get("usage")))
        btn_copy.clicked.connect(copy_report)
        btn_close.clicked.connect(dialog.close)
        dialog.finished.connect(lambda _=0: setattr(self, "ai_report_dialog", None))
        self.ai_report_dialog = dialog
        dialog.show()
        dialog.raise_()
        dialog.activateWindow()

    def _build_store_report_messages(self, context, custom_note):
        system_prompt = """
你是电商店铺经营分析助手。只基于用户提供的数据写周报，不要编造没有给出的数字。
输出要求：
1. 不要使用 Markdown 格式，不要使用 Markdown 标题、表格、代码块、引用块，也不要把输入表格完整复述出来。
2. 报告要像一段可以直接转发给别人看的文字，最多使用换行、空格、简单序号和符号来区分层次，比如“一、本周结论”“①”“-”。
3. 先写本周整体结论，再写核心指标变化，再写售卖结构和主卖规格判断，再结合商品操作记录判断可能影响，最后写下周动作。
4. 操作记录只能作为解释线索：能对应上数据变化就说明，不能对应就不要硬凑原因。
5. 只保留必要数字和关键判断，优先总结大方向；不要长篇解释、不要客套话、不要营销腔。
6. 订单规格数据和财务数据口径不同，出现订单数差异时只提醒一次，不要反复解释。
7. 如果没有商品操作记录，要明确说“本周和上周没有可参考的商品操作记录”，不要臆测操作原因。
8. 规格相关信息必须按输入数据里的 spec_display_rule 执行：如果是规格名称，只能使用【】中的规格名称；如果是规格编码，才使用规格编码。
""".strip()
        user_prompt = {
            "任务": "生成店铺本周经营文字报告",
            "用户补充分析要求": custom_note or "无",
            "数据": context,
        }
        return [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": json.dumps(user_prompt, ensure_ascii=False, indent=2)},
        ]

    def _call_ai_for_store_report(self, context, custom_note, parent_dialog):
        try:
            import requests
        except ImportError:
            QMessageBox.warning(parent_dialog, "错误", "缺少 requests 依赖，无法调用 AI。")
            return ""

        api_key = self.db.get_setting("ai_api_key", "")
        api_url = self.db.get_setting("ai_api_url", "https://api.deepseek.com/chat/completions")
        model = self.db.get_setting("ai_model", "deepseek-v4-flash")
        if not api_key:
            QMessageBox.warning(parent_dialog, "提示", "请先在 AI API 配置中填写 API Key。")
            return ""

        progress = QProgressDialog("正在调用 AI 生成报告...", "取消", 0, 0, parent_dialog)
        progress.setWindowTitle("AI处理中")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()
        QApplication.processEvents()

        headers = {
            "Authorization": f"Bearer {api_key.strip()}",
            "Content-Type": "application/json",
        }
        data = {
            "model": model,
            "messages": self._build_store_report_messages(context, custom_note),
            "max_tokens": 4096,
            "temperature": 0.35,
        }

        response = None
        try:
            for attempt in range(3):
                if progress.wasCanceled():
                    progress.close()
                    return ""
                response = requests.post(api_url, headers=headers, json=data, timeout=90)
                if response.status_code not in (500, 503):
                    break
                if attempt < 2:
                    time.sleep(2 * (attempt + 1))
            progress.close()

            if response is not None and response.status_code == 200:
                result = response.json()
                return result["choices"][0]["message"]["content"].strip()

            status = response.status_code if response is not None else "无响应"
            detail = response.text.strip()[:500] if response is not None and response.text else ""
            if status == 503:
                message = "AI调用失败：503\nDeepSeek服务器当前过载，请稍后重试。"
            else:
                message = f"AI调用失败：{status}"
            if detail:
                message += f"\n\n返回内容：{detail}"
            QMessageBox.warning(parent_dialog, "错误", message)
            return ""
        except requests.exceptions.Timeout:
            progress.close()
            QMessageBox.warning(parent_dialog, "错误", "AI调用超时，请稍后重试。")
            return ""
        except requests.exceptions.ConnectionError as e:
            progress.close()
            QMessageBox.warning(parent_dialog, "错误", f"AI连接失败：{str(e)}")
            return ""
        except Exception as e:
            progress.close()
            QMessageBox.warning(parent_dialog, "错误", f"生成报告失败：{str(e)}")
            return ""

    def _start_ai_store_report(self, messages, input_tokens, parent_dialog, on_success, on_finished):
        api_key = self.db.get_setting("ai_api_key", "")
        api_url = self.db.get_setting("ai_api_url", "https://api.deepseek.com/chat/completions")
        model = self.db.get_setting("ai_model", "deepseek-v4-flash")
        if not api_key:
            QMessageBox.warning(parent_dialog, "提示", "请先在 AI API 配置中填写 API Key。")
            on_finished()
            return

        headers = {
            "Authorization": f"Bearer {api_key.strip()}",
            "Content-Type": "application/json",
        }
        data = {
            "model": model,
            "messages": messages,
            "max_tokens": 4096,
            "temperature": 0.35,
        }

        progress = QProgressDialog("正在调用 AI 生成报告...", "取消", 0, 0, parent_dialog)
        progress.setWindowTitle("AI处理中")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()

        thread = QThread(parent_dialog)
        worker = StoreAiReportWorker(api_url, headers, data)
        worker.moveToThread(thread)
        self.ai_report_thread = thread
        self.ai_report_worker = worker

        def cleanup():
            progress.close()
            on_finished()
            worker.deleteLater()
            thread.quit()
            thread.wait()
            thread.deleteLater()
            if self.ai_report_worker is worker:
                self.ai_report_worker = None
            if self.ai_report_thread is thread:
                self.ai_report_thread = None

        def handle_success(report, usage):
            on_success(report, usage)
            cleanup()

        def handle_failure(message):
            if message != "已取消生成报告。":
                QMessageBox.warning(parent_dialog, "错误", message)
            cleanup()

        progress.canceled.connect(worker.cancel)
        progress.canceled.connect(lambda: progress.setLabelText("正在取消..."))
        thread.started.connect(worker.run)
        worker.finished.connect(handle_success)
        worker.failed.connect(handle_failure)
        thread.start()

    def load_manual_data(self):
        """从数据库加载手动录入数据"""
        try:
            records = self.db.safe_fetchall("""
                SELECT start_date, end_date, actual_orders, actual_amount, gross_profit,
                       refund_amount, refund_orders, promotion_fee, deduction, other_service, other,
                       gross_margin_rate, refund_rate_by_amount, refund_rate_by_orders,
                       unit_price, promotion_ratio, tech_fee, net_profit, net_margin_rate, profit_per_order
                FROM manual_margin_data WHERE store_id=? ORDER BY start_date ASC, end_date ASC
            """, (self.store_id,))
            return records
        except Exception as e:
            print(f"加载手动数据失败: {e}")
            return []

    def _complete_manual_margin_data(self, data):
        completed = dict(data or {})
        actual_orders = int(completed.get("actual_orders", 0) or 0)
        actual_amount = float(completed.get("actual_amount", 0) or 0)
        gross_profit = float(completed.get("gross_profit", 0) or 0)
        refund_amount = float(completed.get("refund_amount", 0) or 0)
        refund_orders = int(completed.get("refund_orders", 0) or 0)
        promotion_fee = float(completed.get("promotion_fee", 0) or 0)
        deduction = float(completed.get("deduction", 0) or 0)
        other_service = float(completed.get("other_service", 0) or 0)
        other = float(completed.get("other", 0) or 0)

        completed["actual_orders"] = actual_orders
        completed["actual_amount"] = actual_amount
        completed["gross_profit"] = gross_profit
        completed["refund_amount"] = refund_amount
        completed["refund_orders"] = refund_orders
        completed["promotion_fee"] = promotion_fee
        completed["deduction"] = deduction
        completed["other_service"] = other_service
        completed["other"] = other

        completed["gross_margin_rate"] = (gross_profit / actual_amount * 100) if actual_amount > 0 else 0
        completed["refund_rate_by_amount"] = (refund_amount / actual_amount * 100) if actual_amount > 0 else 0
        completed["refund_rate_by_orders"] = (refund_orders / actual_orders * 100) if actual_orders > 0 else 0
        completed["unit_price"] = actual_amount / actual_orders if actual_orders > 0 else 0
        completed["promotion_ratio"] = (promotion_fee / actual_amount * 100) if actual_amount > 0 else 0
        completed["tech_fee"] = actual_amount * 0.006
        completed["net_profit"] = gross_profit - refund_amount - promotion_fee - deduction - other_service + other - completed["tech_fee"]
        completed["net_margin_rate"] = (completed["net_profit"] / actual_amount * 100) if actual_amount > 0 else 0
        completed["profit_per_order"] = completed["net_profit"] / actual_orders if actual_orders > 0 else 0
        return completed

    def _flush_imported_data_to_archive(self):
        try:
            self.db.conn.commit()
        except Exception as e:
            print(f"commit imported store margin data failed: {e}")

    def delete_manual_data(self, start_date):
        """删除手动录入数据"""
        reply = QMessageBox.question(self, "确认删除", "确定删除这条数据吗？")
        if reply == QMessageBox.Yes:
            try:
                self.db.safe_execute(
                    "DELETE FROM manual_margin_data WHERE store_id=? AND start_date=?",
                    (self.store_id, start_date)
                )
                # 先清空表格，等待UI更新后再刷新数据
                self.margin_data_table.setRowCount(0)
                QApplication.processEvents()
                self.refresh_manual_data_display()
                self.update_current_history_label()
                self.show_toast("✅ 已删除数据")
            except Exception as e:
                QMessageBox.warning(self, "错误", f"删除失败: {e}")

    def open_input_data_dialog(self):
        """打开录入数据对话框"""
        self.open_input_data_dialog_with_initial()

    def open_input_data_dialog_with_initial(self, initial_data=None):
        """打开录入数据对话框，可选预填部分字段。"""
        from .input_data_dialog import InputDataDialog
        dialog = InputDataDialog(self, initial_data=initial_data)
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            # 添加日期信息
            start_date = self.date_start_input.date().toString("yyyy-MM-dd")
            end_date = self.date_end_input.date().toString("yyyy-MM-dd")
            data["start_date"] = start_date
            data["end_date"] = end_date
            self.save_manual_data(self._complete_manual_margin_data(data))
            self._flush_imported_data_to_archive()
            self.refresh_manual_data_display()
            self.update_current_history_label()

    def on_margin_data_mode_changed(self, *_args):
        mode = self.combo_margin_data_mode.currentData() or "erp"
        self.db.set_setting("store_margin_data_mode", mode)
        text = "表格模式" if mode == "table" else "ERP模式"
        self.show_toast(f"已切换为{text}")

    def is_table_margin_data_mode(self):
        return self.db.get_setting("store_margin_data_mode", "erp") == "table"

    def _read_import_rows(self, file_path):
        import os
        ext = os.path.splitext(file_path)[1].lower()
        if ext in ['.xlsx', '.xls']:
            return self._read_excel_sheet_rows(file_path, data_only=True)
        import csv
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                with open(file_path, 'r', encoding=encoding, newline='') as f:
                    return list(csv.reader(f))
            except UnicodeDecodeError:
                continue
        raise ValueError("CSV 文件编码无法识别，请使用 UTF-8 或 GB18030 编码")

    def _read_excel_sheet_rows(self, file_path, data_only=False):
        if os.path.splitext(file_path)[1].lower() == ".xls":
            import xlrd
            book = xlrd.open_workbook(file_path)
            try:
                sheet_name = self._choose_excel_sheet(book.sheet_names())
                if sheet_name is None:
                    return None
                sheet = book.sheet_by_name(sheet_name)
                rows = []
                for row_index in range(sheet.nrows):
                    row = []
                    for cell in sheet.row(row_index):
                        value = cell.value
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            value = xlrd.xldate_as_datetime(value, book.datemode).strftime("%Y-%m-%d")
                        row.append(value)
                    rows.append(row)
                return rows
            finally:
                book.release_resources()

        import openpyxl
        from datetime import datetime as dt
        wb = openpyxl.load_workbook(file_path, data_only=data_only)
        try:
            sheet_names = [sheet.title for sheet in wb.worksheets]
            sheet_name = self._choose_excel_sheet(sheet_names)
            if sheet_name is None:
                return None
            rows = []
            for row in wb[sheet_name].iter_rows(values_only=True):
                rows.append([
                    cell.strftime("%Y-%m-%d") if isinstance(cell, dt) else cell
                    for cell in row
                ])
            return rows
        finally:
            wb.close()

    def _choose_excel_sheet(self, sheet_names):
        if not sheet_names:
            raise ValueError("Excel 文件中没有可读取的工作表")
        if len(sheet_names) == 1:
            return sheet_names[0]
        sheet_name, ok = QInputDialog.getItem(
            self, "选择工作表", "检测到多个 Sheet，请选择要导入的工作表：",
            sheet_names, 0, False,
        )
        return sheet_name if ok else None

    def _cell_text(self, value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _get_row_value(self, row, col_index):
        if col_index is None or col_index < 0 or col_index >= len(row):
            return None
        return row[col_index]

    def _parse_quantity_value(self, value):
        try:
            return max(1, int(float(str(value).replace(",", "").strip())))
        except (TypeError, ValueError):
            return 1

    def _parse_order_date_value(self, value):
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        text = str(value).strip()
        if not text:
            return None
        try:
            import re
            match = re.search(r'(\d{4})[/-](\d{1,2})[/-](\d{1,2})', text)
            if match:
                return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
            match = re.search(r'(\d{1,2})[/-](\d{1,2})', text)
            if match:
                return f"{datetime.now().year:04d}-{int(match.group(1)):02d}-{int(match.group(2)):02d}"
        except Exception:
            return None
        return None

    def _format_import_order_date(self, value):
        parsed = self._parse_order_date_value(value)
        if not parsed:
            return None
        _year, month, day = parsed.split("-")
        return f"{int(month)}/{int(day)}"

    def _is_effective_shipped_status(self, status_text):
        text = str(status_text or "").strip()
        return "已发货" in text or "已收货" in text

    def _is_refund_status(self, status_text):
        text = str(status_text or "").strip()
        refund_keywords = ["退款成功", "已发货退款", "已收货退款", "售后成功", "退货退款成功"]
        return any(keyword in text for keyword in refund_keywords)

    def _show_margin_import_missing_dialog(self, issues):
        dialog = QDialog(self)
        dialog.setWindowTitle("缺少必要成本或匹配信息")
        dialog.resize(1050, 620)
        layout = QVBoxLayout(dialog)

        tip = QLabel("以下有效订单行缺少必要成本或匹配信息，已阻止预填毛利润。商品ID、标题、规格编码和问题可复制。")
        tip.setWordWrap(True)
        tip.setStyleSheet("font-size: 13px; color: #2c3e50; padding: 6px;")
        layout.addWidget(tip)

        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(["行号", "图片", "商品ID", "商品标题", "规格编码", "问题"])
        table.setRowCount(len(issues))
        table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        table.setSelectionBehavior(QAbstractItemView.SelectItems)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.verticalHeader().setVisible(False)
        table.setColumnWidth(0, 70)
        table.setColumnWidth(1, 78)
        table.setColumnWidth(2, 170)
        table.setColumnWidth(3, 300)
        table.setColumnWidth(4, 190)
        table.horizontalHeader().setSectionResizeMode(5, QHeaderView.Stretch)

        def readonly_text(text):
            line = QLineEdit(str(text or ""))
            line.setReadOnly(True)
            line.setStyleSheet("border: none; background: transparent; padding: 4px;")
            return line

        for row, issue in enumerate(issues):
            row_item = QTableWidgetItem(str(issue.get("row", "")))
            row_item.setTextAlignment(Qt.AlignCenter)
            table.setItem(row, 0, row_item)

            img_label = QLabel()
            img_label.setAlignment(Qt.AlignCenter)
            image_data = issue.get("image_data")
            if image_data:
                pixmap = QPixmap()
                pixmap.loadFromData(image_data)
                if not pixmap.isNull():
                    img_label.setPixmap(pixmap.scaled(56, 56, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                else:
                    img_label.setText("-")
            else:
                img_label.setText("-")
            table.setCellWidget(row, 1, img_label)
            table.setCellWidget(row, 2, readonly_text(issue.get("product_code", "")))
            table.setCellWidget(row, 3, readonly_text(issue.get("product_title", "")))
            table.setCellWidget(row, 4, readonly_text(issue.get("spec_code", "")))
            table.setCellWidget(row, 5, readonly_text(issue.get("message", "")))
            table.setRowHeight(row, 64)

        layout.addWidget(table)

        copy_text = "\n".join(
            "\t".join([
                str(issue.get("row", "")),
                str(issue.get("product_code", "")),
                str(issue.get("product_title", "")),
                str(issue.get("spec_code", "")),
                str(issue.get("message", "")),
            ])
            for issue in issues
        )

        btn_layout = QHBoxLayout()
        btn_copy = QPushButton("复制全部")
        btn_close = QPushButton("关闭")

        def copy_all():
            QApplication.clipboard().setText("行号\t商品ID\t商品标题\t规格编码\t问题\n" + copy_text)
            self.show_toast("已复制缺失信息")

        btn_copy.clicked.connect(copy_all)
        btn_close.clicked.connect(dialog.accept)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_copy)
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)

        dialog.exec_()

    def import_margin_data_from_order_table(self):
        """表格模式：从订单明细计算毛利录入预填数据。"""
        file_path, _ = remembered_open_file(
            self, self.db, "选择订单明细文件",
            "Excel文件 (*.xlsx *.xls);;CSV文件 (*.csv);;所有文件 (*.*)"
        )
        if not file_path:
            return

        try:
            rows = self._read_import_rows(file_path)
            if rows is None:
                return
            if len(rows) < 2:
                QMessageBox.warning(self, "错误", "文件数据少于2行，无法导入")
                return

            headers = [str(h).strip() if h else "" for h in rows[0]]
            col_mapping = self._show_column_mapping_dialog(headers, self._auto_detect_columns(headers))
            if not col_mapping:
                return

            required_fields = {
                "product_id": "商品ID列",
                "spec_code": "规格编码列",
                "quantity": "数量列",
                "order_status": "订单状态列",
                "actual_amount": "实收金额列",
            }
            missing_required = [name for key, name in required_fields.items() if col_mapping.get(key) is None]
            if missing_required:
                QMessageBox.warning(self, "缺少列映射", "表格模式请先选择：" + "、".join(missing_required))
                return

            products = self.db.safe_fetchall(
            "SELECT id, name, title, image_data FROM products WHERE store_id=? AND COALESCE(is_archived, 0)=0",
                (self.store_id,)
            )
            product_code_to_id = {str(name).strip(): prod_id for prod_id, name, _title, _image_data in products if name is not None}
            product_info_by_code = {
                str(name).strip(): {"title": title or "", "image_data": image_data}
                for _prod_id, name, title, image_data in products
                if name is not None
            }
            if not product_code_to_id:
                QMessageBox.warning(self, "提示", "当前店铺没有商品，无法计算表格模式毛利")
                return

            all_specs = {}
            for prod_id in product_code_to_id.values():
                spec_rows = self.db.safe_fetchall(
                    "SELECT spec_code FROM product_specs WHERE product_id=?",
                    (prod_id,)
                )
                all_specs[prod_id] = {str(row[0]).strip() for row in spec_rows if row and row[0] is not None}

            cost_cache = {}
            missing_issues = []
            valid_rows = 0
            total_amount = 0.0
            total_gross_profit = 0.0
            refund_orders = 0
            parsed_dates = []

            for row_idx, row in enumerate(rows[1:], start=2):
                if not row or all(str(cell).strip() == "" for cell in row if cell is not None):
                    continue

                status_text = self._cell_text(self._get_row_value(row, col_mapping["order_status"]))
                if not self._is_effective_shipped_status(status_text):
                    continue

                product_code = self._cell_text(self._get_row_value(row, col_mapping["product_id"]))
                spec_code = self._cell_text(self._get_row_value(row, col_mapping["spec_code"]))
                quantity = self._parse_quantity_value(self._get_row_value(row, col_mapping["quantity"]))
                actual_amount = self._parse_amount_value(self._get_row_value(row, col_mapping["actual_amount"]))
                order_date = self._parse_order_date_value(self._get_row_value(row, col_mapping.get("order_date")))
                if order_date:
                    parsed_dates.append(order_date)

                prod_id = product_code_to_id.get(product_code)
                product_info = product_info_by_code.get(product_code, {})
                if prod_id is None:
                    missing_issues.append({
                        "row": row_idx,
                        "product_code": product_code or "空",
                        "product_title": "",
                        "spec_code": spec_code or "空",
                        "image_data": None,
                        "message": "商品ID不在当前店铺",
                    })
                    continue
                if spec_code not in all_specs.get(prod_id, set()):
                    missing_issues.append({
                        "row": row_idx,
                        "product_code": product_code,
                        "product_title": product_info.get("title", ""),
                        "spec_code": spec_code or "空",
                        "image_data": product_info.get("image_data"),
                        "message": "规格编码不匹配当前商品",
                    })
                    continue

                if spec_code not in cost_cache:
                    cost_rows = self.db.safe_fetchall(
                        "SELECT product_cost, unit_weight FROM cost_library WHERE spec_code=?",
                        (spec_code,)
                    )
                    if not cost_rows:
                        cost_cache[spec_code] = None
                    else:
                        cost_cache[spec_code] = cost_rows[0]
                cost_info = cost_cache.get(spec_code)
                if not cost_info or cost_info[0] is None or cost_info[1] is None or float(cost_info[0] or 0) <= 0 or float(cost_info[1] or 0) <= 0:
                    missing_issues.append({
                        "row": row_idx,
                        "product_code": product_code,
                        "product_title": product_info.get("title", ""),
                        "spec_code": spec_code,
                        "image_data": product_info.get("image_data"),
                        "message": "成本库缺少产品成本或重量",
                    })
                    continue

                product_cost = float(cost_info[0] or 0)
                unit_weight = float(cost_info[1] or 0)
                total_weight = unit_weight * quantity
                shipping_fee = self.db.calculate_cost_shipping_fee(total_weight)
                misc_fee = self.db.get_cost_misc_fee()
                total_cost = product_cost * quantity + shipping_fee + misc_fee

                valid_rows += 1
                total_amount += actual_amount
                total_gross_profit += actual_amount - total_cost
                if self._is_refund_status(status_text):
                    refund_orders += 1

            if missing_issues:
                self._show_margin_import_missing_dialog(missing_issues)
                return

            if valid_rows <= 0:
                QMessageBox.information(self, "提示", "没有识别到状态为已发货或已收货的有效订单行")
                return

            if parsed_dates:
                from PyQt5.QtCore import QDate
                start = min(parsed_dates)
                end = max(parsed_dates)
                self.date_start_input.setDate(QDate.fromString(start, "yyyy-MM-dd"))
                self.date_end_input.setDate(QDate.fromString(end, "yyyy-MM-dd"))

            initial_data = {
                "actual_orders": valid_rows,
                "actual_amount": round(total_amount, 2),
                "gross_profit": round(total_gross_profit, 2),
                "refund_orders": refund_orders,
            }
            start_date = self.date_start_input.date().toString("yyyy-MM-dd")
            end_date = self.date_end_input.date().toString("yyyy-MM-dd")
            data = self._complete_manual_margin_data(initial_data)
            data["start_date"] = start_date
            data["end_date"] = end_date
            result = self.save_manual_data(data)
            if result:
                self._flush_imported_data_to_archive()
                self.refresh_manual_data_display()
                self.update_current_history_label()
                self.show_toast(f"✅ 表格导入已自动保存：{start_date} ~ {end_date}")

        except Exception as e:
            QMessageBox.warning(self, "错误", f"表格模式导入失败: {e}")

    def import_data(self):
        """导入Excel/CSV数据"""
        if self.is_table_margin_data_mode():
            self.import_margin_data_from_order_table()
            return

        file_path, _ = remembered_open_file(
            self, self.db, "选择导入文件",
            "Excel文件 (*.xlsx *.xls);;CSV文件 (*.csv);;所有文件 (*.*)"
        )
        if not file_path:
            return

        try:
            import os
            ext = os.path.splitext(file_path)[1].lower()

            if ext in ['.xlsx', '.xls']:
                rows = self._read_excel_sheet_rows(file_path)
                if rows is None:
                    return
            else:
                import csv
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    rows = list(reader)

            if len(rows) < 2:
                QMessageBox.warning(self, "错误", "文件数据少于2行，无法导入")
                return

            header = [str(h).strip() if h else "" for h in rows[0]]

            # 自动识别列
            col_map = {}
            for i, h in enumerate(header):
                h_str = str(h).strip()
                h_lower = h_str.lower()
                if "日期" in h_str or "date" in h_lower:
                    col_map["start_date"] = i
                elif "实发订单" in h_str:
                    col_map["actual_orders"] = i
                elif "实发金额" in h_str:
                    col_map["actual_amount"] = i
                elif "毛利润" in h_str and "净" not in h_str:
                    col_map["gross_profit"] = i
                elif "退款金额" in h_str:
                    col_map["refund_amount"] = i
                elif "退款订单" in h_str:
                    col_map["refund_orders"] = i
                elif "推广费" in h_str:
                    col_map["promotion_fee"] = i
                elif "扣款" in h_str:
                    col_map["deduction"] = i
                elif "其他服务" in h_str:
                    col_map["other_service"] = i
                elif h_str.strip() == "其他" or (h_str.strip().startswith("其他") and "服务" not in h_str):
                    col_map["other"] = i

            # 总是弹出列映射对话框，自动识别到的帮用户选上
            dialog = QDialog(self)
            dialog.setWindowTitle("选择列映射")
            dialog.resize(500, 400)
            layout = QVBoxLayout(dialog)

            recognized_count = len(col_map)
            layout.addWidget(QLabel(f"已自动识别 {recognized_count} 个字段，其他字段请手动选择："))

            fields = [
                ("start_date", "日期（必填）"),
                ("actual_orders", "实发订单"),
                ("actual_amount", "实发金额"),
                ("gross_profit", "毛利润"),
                ("refund_amount", "退款金额"),
                ("refund_orders", "退款订单"),
                ("promotion_fee", "推广费"),
                ("deduction", "扣款"),
                ("other_service", "其他服务"),
                ("other", "其他"),
            ]

            combos = {}
            for field_key, field_name in fields:
                row_layout = QHBoxLayout()
                row_layout.addWidget(QLabel(field_name))
                combo = QComboBox()
                combo.addItem("（不导入）", -1)
                for idx, h in enumerate(header):
                    combo.addItem(f"{idx}: {h}", idx)
                # 自动设置已识别的列
                if field_key in col_map:
                    combo.setCurrentIndex(col_map[field_key] + 1)  # +1 因为第一个是"（不导入）"
                row_layout.addWidget(combo)
                layout.addLayout(row_layout)
                combos[field_key] = combo

            btn_layout = QHBoxLayout()
            ok_btn = QPushButton("确定")
            cancel_btn = QPushButton("取消")
            btn_layout.addWidget(ok_btn)
            btn_layout.addWidget(cancel_btn)
            layout.addLayout(btn_layout)

            def on_ok():
                new_col_map = {}
                for field_key, combo in combos.items():
                    idx = combo.currentData()
                    if idx != -1:
                        new_col_map[field_key] = idx
                dialog.col_map = new_col_map
                dialog.accept()

            ok_btn.clicked.connect(on_ok)
            cancel_btn.clicked.connect(dialog.reject)

            if dialog.exec_() != QDialog.Accepted:
                return

            col_map = dialog.col_map

            if "start_date" not in col_map:
                QMessageBox.warning(self, "错误", "请至少选择日期列")
                return

            imported_count = 0
            overwritten_count = 0
            for row_idx in range(1, len(rows)):
                row = rows[row_idx]
                if not row or all(str(cell).strip() == "" for cell in row):
                    continue

                # 解析日期
                date_str = str(row[col_map["start_date"]]).strip()
                try:
                    from datetime import datetime
                    import re
                    start_date = None
                    end_date = None

                    dash_match = re.match(r'(\d{1,2})\.(\d{1,2})-(\d{1,2})\.(\d{1,2})', date_str)
                    if dash_match:
                        current_year = datetime.now().year
                        start_date = datetime(current_year, int(dash_match.group(1)), int(dash_match.group(2)))
                        end_date = datetime(current_year, int(dash_match.group(3)), int(dash_match.group(4)))
                    else:
                        date_str_clean = date_str.replace("/", "-").replace("~", "-")
                        parts = date_str_clean.split("-")
                        if len(parts) == 3:
                            if len(parts[0]) == 4:
                                start_date = end_date = datetime(int(parts[0]), int(parts[1]), int(parts[2]))
                            else:
                                current_year = datetime.now().year
                                start_date = end_date = datetime(current_year, int(parts[0]), int(parts[1]))
                        elif len(parts) == 2:
                            current_year = datetime.now().year
                            start_date = end_date = datetime(current_year, int(parts[0]), int(parts[1]))
                        elif len(parts) == 1 and parts[0]:
                            if "." in parts[0]:
                                sub_parts = parts[0].split(".")
                            elif "/" in parts[0]:
                                sub_parts = parts[0].split("/")
                            else:
                                sub_parts = None
                            if sub_parts and len(sub_parts) == 2:
                                current_year = datetime.now().year
                                start_date = end_date = datetime(current_year, int(sub_parts[0]), int(sub_parts[1]))

                    if start_date is None:
                        continue
                    start_date_str = start_date.strftime("%Y-%m-%d")
                    end_date_str = end_date.strftime("%Y-%m-%d")
                except:
                    continue

                # 解析数值
                def get_float(val, default=0.0):
                    try:
                        s = str(val).replace("¥", "").replace(",", "").replace("%", "").strip()
                        return float(s) if s else default
                    except:
                        return default

                def get_int(val, default=0):
                    try:
                        s = str(val).replace(",", "").strip()
                        return int(float(s)) if s else default
                    except:
                        return default

                data = {
                    "start_date": start_date_str,
                    "end_date": end_date_str,
                    "actual_orders": get_int(row[col_map.get("actual_orders", -1)]),
                    "actual_amount": get_float(row[col_map.get("actual_amount", -1)]),
                    "gross_profit": get_float(row[col_map.get("gross_profit", -1)]),
                    "refund_amount": get_float(row[col_map.get("refund_amount", -1)]),
                    "refund_orders": get_int(row[col_map.get("refund_orders", -1)]),
                    "promotion_fee": get_float(row[col_map.get("promotion_fee", -1)]),
                    "deduction": get_float(row[col_map.get("deduction", -1)]),
                    "other_service": get_float(row[col_map.get("other_service", -1)]),
                    "other": get_float(row[col_map.get("other", -1)]),
                }

                # 计算自动指标
                if data["actual_amount"] > 0:
                    data["gross_margin_rate"] = (data["gross_profit"] / data["actual_amount"]) * 100
                    data["refund_rate_by_amount"] = (data["refund_amount"] / data["actual_amount"]) * 100
                    data["promotion_ratio"] = (data["promotion_fee"] / data["actual_amount"]) * 100
                    data["unit_price"] = data["actual_amount"] / data["actual_orders"] if data["actual_orders"] > 0 else 0
                else:
                    data["gross_margin_rate"] = 0
                    data["refund_rate_by_amount"] = 0
                    data["promotion_ratio"] = 0
                    data["unit_price"] = 0

                if data["actual_orders"] > 0:
                    data["refund_rate_by_orders"] = (data["refund_orders"] / data["actual_orders"]) * 100
                else:
                    data["refund_rate_by_orders"] = 0

                data["tech_fee"] = data["actual_amount"] * 0.006

                data["net_profit"] = (
                    data["gross_profit"]
                    - data["refund_amount"]
                    - data["promotion_fee"]
                    - data["deduction"]
                    - data["other_service"]
                    + data["other"]
                    - data["tech_fee"]
                )

                if data["actual_amount"] > 0:
                    data["net_margin_rate"] = (data["net_profit"] / data["actual_amount"]) * 100
                else:
                    data["net_margin_rate"] = 0

                if data["actual_orders"] > 0:
                    data["profit_per_order"] = data["net_profit"] / data["actual_orders"]
                else:
                    data["profit_per_order"] = 0

                result = self.save_manual_data(data)
                if result == "new":
                    imported_count += 1
                elif result == "overwrite":
                    imported_count += 1
                    overwritten_count += 1

            self.refresh_manual_data_display()
            self.update_current_history_label()
            self._flush_imported_data_to_archive()
            if overwritten_count > 0:
                self.show_toast(f"✅ 导入成功：新增 {imported_count - overwritten_count} 条，覆盖 {overwritten_count} 条")
            else:
                self.show_toast(f"✅ 已导入 {imported_count} 条数据")

        except Exception as e:
            QMessageBox.warning(self, "错误", f"导入失败: {e}")

    def save_manual_data(self, data):
        """保存手动录入数据"""
        try:
            from datetime import datetime
            start_date = data.get("start_date", "")
            end_date = data.get("end_date", "")

            # 检查是否有相同日期的记录
            existing = self.db.safe_fetchall("""
                SELECT actual_orders, actual_amount, gross_profit, refund_amount,
                       refund_orders, promotion_fee, deduction, other_service, other
                FROM manual_margin_data
                WHERE store_id=? AND start_date=? AND end_date=?
            """, (self.store_id, start_date, end_date))

            if existing:
                old_record = existing[0]
                new_values = (
                    data.get("actual_orders", 0),
                    data.get("actual_amount", 0),
                    data.get("gross_profit", 0),
                    data.get("refund_amount", 0),
                    data.get("refund_orders", 0),
                    data.get("promotion_fee", 0),
                    data.get("deduction", 0),
                    data.get("other_service", 0),
                    data.get("other", 0),
                )
                if old_record == new_values:
                    return False  # 数据相同，跳过

                reply = QMessageBox.question(
                    self, "确认覆盖",
                    f"该日期范围 ({start_date} ~ {end_date}) 已存在数据，是否覆盖？"
                )
                if reply != QMessageBox.Yes:
                    return False
                is_overwrite = True
            else:
                is_overwrite = False

            latest_rows = self.db.safe_fetchall(
                "SELECT MAX(end_date) FROM manual_margin_data WHERE store_id=?",
                (self.store_id,),
            )
            latest_end_date = str(latest_rows[0][0] or "") if latest_rows else ""
            clear_weekly_images = _should_clear_weekly_images(latest_end_date, end_date, is_overwrite)

            created_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # 使用REPLACE自动覆盖已存在的记录
            self.db.safe_execute("""
                INSERT OR REPLACE INTO manual_margin_data (
                    store_id, start_date, end_date,
                    actual_orders, actual_amount, gross_profit,
                    refund_amount, refund_orders, promotion_fee,
                    deduction, other_service, other,
                    gross_margin_rate, refund_rate_by_amount, refund_rate_by_orders,
                    unit_price, promotion_ratio, tech_fee,
                    net_profit, net_margin_rate, profit_per_order,
                    created_time
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                self.store_id,
                data.get("start_date", ""),
                data.get("end_date", ""),
                data.get("actual_orders", 0),
                data.get("actual_amount", 0),
                data.get("gross_profit", 0),
                data.get("refund_amount", 0),
                data.get("refund_orders", 0),
                data.get("promotion_fee", 0),
                data.get("deduction", 0),
                data.get("other_service", 0),
                data.get("other", 0),
                data.get("gross_margin_rate", 0),
                data.get("refund_rate_by_amount", 0),
                data.get("refund_rate_by_orders", 0),
                data.get("unit_price", 0),
                data.get("promotion_ratio", 0),
                data.get("tech_fee", 0),
                data.get("net_profit", 0),
                data.get("net_margin_rate", 0),
                data.get("profit_per_order", 0),
                created_time
            ))
            if clear_weekly_images:
                try:
                    self.db.safe_execute("DELETE FROM store_temp_images WHERE store_id=?", (self.store_id,))
                    if getattr(self, "weekly_images_panel", None) is not None:
                        self.load_weekly_images()
                except Exception as image_error:
                    QMessageBox.warning(self, "图片清理失败", f"毛利数据已保存，但上一周期附带图片清理失败：{image_error}")
            return "overwrite" if is_overwrite else "new"
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败: {e}")
            return False

    def save_historical_data(self):
        """保存当前导入订单数据到历史记录"""
        try:
            start_date = self.date_start_input.date().toString("yyyy-MM-dd")
            end_date = self.date_end_input.date().toString("yyyy-MM-dd")
            
            if start_date > end_date:
                QMessageBox.warning(self, "提示", "开始日期不能晚于结束日期")
                return
            
            # 计算天数差
            from datetime import datetime
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            days_diff = (end_dt - start_dt).days + 1
            
            if days_diff <= 0:
                QMessageBox.warning(self, "提示", "日期范围无效")
                return
            
            # 获取店铺的所有导入订单数据
            orders = self.db.safe_fetchall(
                """SELECT io.actual_amount, io.order_count FROM imported_orders io
                   WHERE io.store_id=? AND EXISTS (
                       SELECT 1 FROM products p
                       JOIN product_specs ps ON ps.product_id=p.id AND ps.spec_code=io.spec_code
                         AND COALESCE(ps.is_temporarily_off_shelf, 0)=0
                       WHERE p.store_id=io.store_id AND p.name=io.product_id
                         AND COALESCE(p.is_archived, 0)=0 AND COALESCE(p.is_violation, 0)=0
                   )""",
                (self.store_id,)
            )
            
            if not orders:
                QMessageBox.information(self, "提示", "当前没有导入的订单数据")
                return
            
            # 计算总数据
            total_amount = 0.0
            total_orders = 0
            
            for actual_amount, order_count in orders:
                if actual_amount:
                    total_amount += actual_amount
                if order_count:
                    total_orders += order_count
            
            # 计算客单价和日均数据
            avg_price = total_amount / total_orders if total_orders > 0 else 0
            daily_amount = total_amount / days_diff
            daily_orders = total_orders / days_diff
            
            # 保存到数据库
            created_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            self.db.safe_execute(
                "INSERT OR REPLACE INTO historical_data (store_id, start_date, end_date, total_amount, total_orders, avg_price, daily_amount, daily_orders, created_time) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (self.store_id, start_date, end_date, total_amount, total_orders, avg_price, daily_amount, daily_orders, created_time)
            )
            
            # 更新显示
            self.lbl_daily_amount.setText(f"日销售金额: ¥{daily_amount:.2f}")
            self.lbl_daily_orders.setText(f"日单量: {daily_orders:.1f}单")
            
            # 显示保存成功信息
            self.show_toast(f"✅ 已保存 {start_date} ~ {end_date} 的数据")
            
        except Exception as e:
            print(f"保存历史数据失败: {e}")
            QMessageBox.warning(self, "错误", f"保存历史数据失败: {e}")

    def view_historical_data(self):
        """查看历史数据"""
        try:
            # 创建历史数据查看对话框
            dialog = QDialog(self)
            dialog.setWindowTitle(f"📊 {self.store_name} - 历史数据")
            dialog.resize(600, 400)
            
            layout = QVBoxLayout(dialog)
            
            # 标题
            title_label = QLabel("📈 历史数据分析记录")
            title_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50; margin-bottom: 10px;")
            layout.addWidget(title_label)
            
            # 获取历史数据
            historical_data = self.db.safe_fetchall(
                "SELECT id, start_date, end_date, total_amount, total_orders, avg_price, daily_amount, daily_orders, created_time FROM historical_data WHERE store_id=? ORDER BY start_date DESC, end_date DESC",
                (self.store_id,)
            )
            
            if not historical_data:
                no_data_label = QLabel("暂无历史数据记录")
                no_data_label.setStyleSheet("font-size: 14px; color: #999; text-align: center; padding: 20px;")
                layout.addWidget(no_data_label)
            else:
                # 创建滚动区域
                scroll_area = QScrollArea()
                scroll_widget = QWidget()
                scroll_layout = QVBoxLayout(scroll_widget)
                scroll_layout.setSpacing(5)
                
                for data in historical_data:
                    data_id, start_date, end_date, total_amount, total_orders, avg_price, daily_amount, daily_orders, created_time = data
                    
                    # 创建数据卡片
                    card_widget = QWidget()
                    card_widget.setStyleSheet("background-color: #f8f9fa; border: 1px solid #dee2e6; border-radius: 5px; padding: 8px;")
                    card_layout = QVBoxLayout(card_widget)
                    card_layout.setContentsMargins(8, 8, 8, 8)
                    
                    # 日期行
                    date_row = QWidget()
                    date_layout = QHBoxLayout(date_row)
                    date_layout.setContentsMargins(0, 0, 0, 0)
                    
                    date_label = QLabel(f"📅 {start_date} ~ {end_date}")
                    date_label.setStyleSheet("font-size: 12px; font-weight: bold; color: #2c3e50;")
                    
                    delete_btn = QPushButton("🗑️")
                    delete_btn.setFixedSize(20, 20)
                    delete_btn.setStyleSheet("font-size: 10px; padding: 0px; background-color: #dc3545; color: white; border-radius: 2px;")
                    delete_btn.clicked.connect(lambda checked, d_id=data_id: self.delete_historical_data(d_id, dialog))
                    
                    date_layout.addWidget(date_label)
                    date_layout.addStretch()
                    date_layout.addWidget(delete_btn)
                    
                    # 数据行
                    data_row = QWidget()
                    data_layout = QHBoxLayout(data_row)
                    data_layout.setContentsMargins(0, 5, 0, 0)
                    
                    amount_label = QLabel(f"总额: ¥{total_amount:.2f}")
                    amount_label.setStyleSheet("font-size: 11px; color: #e74c3c;")
                    
                    orders_label = QLabel(f"订单: {total_orders}单")
                    orders_label.setStyleSheet("font-size: 11px; color: #3498db;")
                    
                    avg_label = QLabel(f"客单价: ¥{avg_price:.2f}")
                    avg_label.setStyleSheet("font-size: 11px; color: #27ae60;")
                    
                    daily_amount_label = QLabel(f"日销: ¥{daily_amount:.2f}")
                    daily_amount_label.setStyleSheet("font-size: 11px; color: #9b59b6;")
                    
                    daily_orders_label = QLabel(f"日单: {daily_orders:.1f}")
                    daily_orders_label.setStyleSheet("font-size: 11px; color: #f39c12;")
                    
                    data_layout.addWidget(amount_label)
                    data_layout.addWidget(orders_label)
                    data_layout.addWidget(avg_label)
                    data_layout.addWidget(daily_amount_label)
                    data_layout.addWidget(daily_orders_label)
                    
                    card_layout.addWidget(date_row)
                    card_layout.addWidget(data_row)
                    
                    scroll_layout.addWidget(card_widget)
                
                scroll_layout.addStretch()
                scroll_area.setWidget(scroll_widget)
                scroll_area.setWidgetResizable(True)
                layout.addWidget(scroll_area)
            
            # 关闭按钮
            close_btn = QPushButton("关闭")
            close_btn.clicked.connect(dialog.accept)
            layout.addWidget(close_btn)
            
            dialog.exec_()
            
        except Exception as e:
            print(f"查看历史数据失败: {e}")
            QMessageBox.warning(self, "错误", f"查看历史数据失败: {e}")

    def set_last_week(self):
        """设置近七天的日期范围（昨天到过去七天）"""
        from PyQt5.QtCore import QDate
        
        # 结束日期设置为昨天
        yesterday = QDate.currentDate().addDays(-1)
        # 开始日期设置为七天前
        seven_days_ago = yesterday.addDays(-6)
        
        self.date_start_input.setDate(seven_days_ago)
        self.date_end_input.setDate(yesterday)
        
        self.show_toast(f"已设置日期范围: {seven_days_ago.toString('yyyy-MM-dd')} ~ {yesterday.toString('yyyy-MM-dd')}")

    def delete_historical_data(self, data_id, parent_dialog):
        """删除历史数据"""
        reply = QMessageBox.question(self, "确认删除", "确定删除这条历史数据记录吗？")
        if reply == QMessageBox.Yes:
            try:
                self.db.safe_execute("DELETE FROM historical_data WHERE id=?", (data_id,))
                # 刷新对话框
                parent_dialog.accept()
                self.view_historical_data()
                self.show_toast("✅ 已删除历史数据")
            except Exception as e:
                QMessageBox.warning(self, "错误", f"删除失败: {e}")

    def _update_order_label_for_row(self, row, weight_input, order_label, prod_id):
        """更新单量显示标签"""
        spec_counts = self.db.safe_fetchall(
            "SELECT spec_code, order_count, refund_count FROM imported_orders WHERE product_id=?",
            (prod_id,)
        )
        total_prod_orders = sum(sc[1] for sc in spec_counts) if spec_counts else 0
        if total_prod_orders > 0:
            order_label.setText(f"{total_prod_orders}单")
            weight_input.setToolTip(f"订单数: {total_prod_orders}单")
        else:
            order_label.setText("0单")
            weight_input.setToolTip("")
        refund_orders_label = self.table.cellWidget(row, 12)
        refund_ratio_label = self.table.cellWidget(row, 13)
        if spec_counts:
            total_orders = sum(sc[1] or 0 for sc in spec_counts)
            total_refund = sum(sc[2] or 0 for sc in spec_counts)
            if total_orders > 0 and total_refund > 0:
                refund_rate = total_refund / total_orders * 100
                if refund_orders_label and hasattr(refund_orders_label, 'setText'):
                    refund_orders_label.setText(f"{refund_rate:.2f}%")
                    refund_orders_label.setStyleSheet("color: #e74c3c; font-size: 19px; font-weight: bold;")
                max_refund_spec = None
                max_refund_rate = -1
                for spec_code, oc, rc in spec_counts:
                    oc = oc or 0
                    rc = rc or 0
                    if oc > 0 and rc > 0:
                        sr = rc / oc
                        if sr > max_refund_rate:
                            max_refund_rate = sr
                            max_refund_spec = spec_code
                if refund_ratio_label and hasattr(refund_ratio_label, 'setText'):
                    if max_refund_spec:
                        refund_ratio_label.setText(str(max_refund_spec))
                        refund_ratio_label.setStyleSheet("color: #e74c3c; font-size: 19px; font-weight: bold;")
                    else:
                        refund_ratio_label.setText("无")
                        refund_ratio_label.setStyleSheet("color: #95a5a6; font-size: 19px;")
            elif total_orders > 0 and total_refund == 0:
                if refund_orders_label and hasattr(refund_orders_label, 'setText'):
                    refund_orders_label.setText("0.00%")
                    refund_orders_label.setStyleSheet("color: #27ae60; font-size: 19px;")
                if refund_ratio_label and hasattr(refund_ratio_label, 'setText'):
                    refund_ratio_label.setText("无")
                    refund_ratio_label.setStyleSheet("color: #95a5a6; font-size: 19px;")
            else:
                if refund_orders_label and hasattr(refund_orders_label, 'setText'):
                    refund_orders_label.setText("无")
                    refund_orders_label.setStyleSheet("color: #95a5a6; font-size: 19px;")
                if refund_ratio_label and hasattr(refund_ratio_label, 'setText'):
                    refund_ratio_label.setText("无")
                    refund_ratio_label.setStyleSheet("color: #95a5a6; font-size: 19px;")
        else:
            if refund_orders_label and hasattr(refund_orders_label, 'setText'):
                refund_orders_label.setText("无")
                refund_orders_label.setStyleSheet("color: #95a5a6; font-size: 19px;")
            if refund_ratio_label and hasattr(refund_ratio_label, 'setText'):
                refund_ratio_label.setText("无")
                refund_ratio_label.setStyleSheet("color: #95a5a6; font-size: 19px;")

    def get_product_margin(self, product_id):
        metrics = self.db.calculate_product_gross_margin_metrics(product_id)
        price = float(metrics.get("avg_final_price") or 0)
        profit = float(metrics.get("avg_gross_profit") or 0)
        margin = float(metrics.get("gross_margin_pct") or 0)
        return max(0.0, price - profit), price, margin

    def calculate_total_margin(self):
        total_weight = 0
        total_weighted_margin = 0
        for row in range(self.table.rowCount()):
            prod_id = self.table.item(row, 1).data(Qt.UserRole)
            cell_widget = self.table.cellWidget(row, 6)
            if not cell_widget:
                continue
            weight_input = cell_widget.findChild(QLineEdit)
            margin_item = self.table.item(row, 5)
            if not weight_input or not margin_item:
                continue
            try:
                weight = float(weight_input.text()) if weight_input.text() else 0
                margin = float(margin_item.text().replace("%", ""))
            except ValueError:
                continue
            total_weight += weight
            total_weighted_margin += margin * weight
        total_margin = (total_weighted_margin / total_weight) if total_weight > 0 else 0
        self.lbl_total_margin.setText(f"综合毛利: {total_margin:.2f}%")
        if total_weight > 100:
            self.lbl_total_margin.setToolTip(f"⚠️ 权重总和超过100% ({total_weight:.1f}%)，可能导致毛利计算不准")
        else:
            self.lbl_total_margin.setToolTip("")
        if total_margin < 10:
            self.lbl_total_margin.setStyleSheet(
                "font-size: 20px; font-weight: bold; color: #c0392b; background-color: #fdeaa8; padding: 10px 20px; border-radius: 8px;"
            )
        elif total_margin > 30:
            self.lbl_total_margin.setStyleSheet(
                "font-size: 20px; font-weight: bold; color: #27ae60; background-color: #d5f5e3; padding: 10px 20px; border-radius: 8px;"
            )
        else:
            self.lbl_total_margin.setStyleSheet(
                "font-size: 20px; font-weight: bold; color: #e74c3c; background-color: #fdeaa8; padding: 10px 20px; border-radius: 8px;"
            )
        return total_margin

    def open_profit_calculator(self):
        margin_text = self.lbl_total_margin.text()
        try:
            margin_rate = float(margin_text.replace("%", "").replace("综合毛利:", "").strip())
        except ValueError:
            margin_rate = 0.0
        avg_price = self.calculate_weighted_avg_price()
        self.main_app.open_profit_calculator_dialog(
            margin_rate, avg_price, self.store_id, self.store_name, "store", self, self.db
        )

    def calculate_weighted_avg_price(self):
        total_weight = 0.0
        weighted_price = 0.0
        for row in range(self.table.rowCount()):
            prod_id = self.table.item(row, 1).data(Qt.UserRole)
            cell_widget = self.table.cellWidget(row, 6)
            if not cell_widget:
                continue
            weight_input = cell_widget.findChild(QLineEdit)
            price_item = self.table.item(row, 4)
            if not weight_input or not price_item:
                continue
            try:
                weight = float(weight_input.text()) if weight_input.text() else 0
                price = float(price_item.text()) if price_item.text() else 0
                if price > 0 and weight > 0:
                    weighted_price += price * weight
                    total_weight += weight
            except ValueError:
                continue
        return weighted_price / total_weight if total_weight > 0 else 0.0

    def on_cell_changed(self, row, col):
        pass

    def toggle_lock(self, row, user_id):
        sys_id = self.get_sys_id_by_user_id(user_id)
        if not sys_id:
            return
        current = self.product_weights.get(user_id, {})
        is_locked = current.get("locked", 0)
        new_locked = 1 if not is_locked else 0
        self.db.safe_execute("UPDATE products SET store_weight_locked=? WHERE id=?", (new_locked, sys_id))
        if new_locked == 1:
            total_locked = sum(
                data.get("weight", 0) for uid, data in self.product_weights.items()
                if uid != user_id and data.get("locked", 0)
            )
            remaining = 100 - total_locked
            if current.get("weight", 0) > remaining:
                self.db.safe_execute("UPDATE products SET store_weight=? WHERE id=?", (remaining, sys_id))
        self.load_products()

    def calculate_weights_from_orders(self):
        self._normalize_imported_order_store_ids()
        order_data = self.db.safe_fetchall("""
            SELECT io.product_id, SUM(io.order_count) as total_orders
            FROM imported_orders io
            WHERE io.store_id=? AND EXISTS (
                SELECT 1 FROM products p
                JOIN product_specs ps ON ps.product_id=p.id AND ps.spec_code=io.spec_code
                  AND COALESCE(ps.is_temporarily_off_shelf, 0)=0
                WHERE p.store_id=io.store_id AND p.name=io.product_id
                  AND COALESCE(p.is_archived, 0)=0 AND COALESCE(p.is_violation, 0)=0
            )
            GROUP BY io.product_id
        """, (self.store_id,))
        product_orders = {row[0]: row[1] for row in order_data if row[1]}
        total_store_orders = sum(product_orders.values())
        if total_store_orders <= 0:
            return
        for prod_id in self.product_weights:
            orders = product_orders.get(prod_id, 0)
            weight = (orders / total_store_orders) * 100 if orders > 0 else 0
            sys_id = self.product_weights[prod_id].get("sys_id")
            if sys_id:
                self.db.safe_execute("UPDATE products SET store_weight=? WHERE id=?", (weight, sys_id))
            self.product_weights[prod_id]["weight"] = weight

    def auto_balance_weights(self):
        unlocked_rows = []
        for row in range(self.table.rowCount()):
            prod_id = self.table.item(row, 1).data(Qt.UserRole)
            if prod_id and not self.product_weights.get(prod_id, {}).get("locked", 0):
                unlocked_rows.append(prod_id)
        if not unlocked_rows:
            return
        total_locked = sum(
            data.get("weight", 0) for data in self.product_weights.values() if data.get("locked", 0)
        )
        remaining = 100 - total_locked
        avg_weight = (remaining / len(unlocked_rows)) if remaining > 0 else 0
        for user_id in unlocked_rows:
            sys_id = self.get_sys_id_by_user_id(user_id)
            if sys_id:
                self.db.safe_execute("UPDATE products SET store_weight=? WHERE id=?", (avg_weight, sys_id))
        self.load_products()

    def _order_table_sort_value(self, column, text):
        text = str(text or "").strip()
        if column not in {1, 3, 4, 5, 6, 7, 8, 9, 10, 12}:
            return text.casefold()
        match = re.search(r"-?\d+(?:\.\d+)?", text.replace(",", ""))
        if not match:
            return 0
        if column == 1:
            return int(float(match.group()))
        value = float(match.group())
        return -abs(value) if "↓" in text else value

    def sort_order_table_by_column(self, column):
        if column == 0:
            return
        if self._order_sort_column == column:
            self._order_sort_order = (
                Qt.DescendingOrder
                if self._order_sort_order == Qt.AscendingOrder
                else Qt.AscendingOrder
            )
        else:
            self._order_sort_column = column
            self._order_sort_order = Qt.DescendingOrder

        originals = []
        for row in range(self.table.rowCount()):
            text = self._table_cell_text(self.table, row, column)
            item = self.table.item(row, column)
            if item is None:
                item = QTableWidgetItem()
                self.table.setItem(row, column, item)
            originals.append((item, item.data(Qt.DisplayRole)))
            item.setData(Qt.DisplayRole, self._order_table_sort_value(column, text))

        self.table.sortItems(column, self._order_sort_order)
        for item, original in originals:
            item.setData(Qt.DisplayRole, original)
        self.table.horizontalHeader().setSortIndicator(column, self._order_sort_order)

    def on_cell_clicked(self, row, col):
        if col != 2:
            return
        product_item = self.table.item(row, 1)
        if product_item:
            prod_id = product_item.data(Qt.UserRole)
            if prod_id:
                self.open_spec_dialog_by_id(prod_id)

    def on_cell_double_clicked(self, row, col):
        if col == 1:
            prod_id = self.table.item(row, 1).data(Qt.UserRole)
            if prod_id:
                clipboard = QApplication.clipboard()
                clipboard.setText(str(prod_id))
                self.show_toast(f"✅ 商品ID {prod_id} 已复制到剪贴板")
                return
        if col == 6:
            prod_id = self.table.item(row, 1).data(Qt.UserRole)
            if prod_id:
                is_locked = self.product_weights.get(prod_id, {}).get("locked", 0)
                if is_locked:
                    self.toggle_lock(row, prod_id)
                    QMessageBox.information(self, "已解锁", "权重已解锁，可以编辑！")
                else:
                    self.is_editing = True
                    self.table.editItem(self.table.item(row, 6))

    def show_context_menu(self, pos):
        item = self.table.itemAt(pos)
        if not item:
            return
        row, col = item.row(), item.column()
        prod_id = self.table.item(row, 1).data(Qt.UserRole)
        menu = QMenu(self)
        action_edit = QAction("📦 编辑规格", self)
        action_edit.triggered.connect(lambda: self.open_spec_dialog_by_id(prod_id))
        menu.addAction(action_edit)
        if col == 6 and prod_id:
            pass
        menu.exec_(self.table.viewport().mapToGlobal(pos))

    def show_margin_data_context_menu(self, pos):
        """显示财务数据表格的右键菜单"""
        item = self.margin_data_table.itemAt(pos)
        if not item:
            return
        row = item.row()
        
        # 获取该行的日期数据
        date_item = self.margin_data_table.item(row, 0)
        if not date_item:
            return
            
        date_text = date_item.text()
        # 日期格式可能是 "04-01\n04-07" 或 "04-01~04-07"
        if "\n" in date_text:
            parts = date_text.split("\n")
        elif "~" in date_text:
            parts = date_text.split("~")
        else:
            return

        if len(parts) >= 2 and parts[0].strip():
            start_date_short = parts[0].strip()
            current_year = datetime.now().year
            start_date_full = f"{current_year}-{start_date_short}"
        else:
            QMessageBox.warning(self, "错误", "该数据的开始日期为空，无法删除")
            return
            
        menu = QMenu(self)
        action_delete = QAction("🗑️ 删除此行数据", self)
        action_delete.triggered.connect(lambda: self.delete_manual_data(start_date_full))
        menu.addAction(action_delete)
        
        menu.exec_(self.margin_data_table.viewport().mapToGlobal(pos))

    def show_week_comparison(self):
        """显示周环比对比结果"""
        current_index = self.combo_current.currentIndex()
        previous_index = self.combo_previous.currentIndex()
        
        if current_index < 0 or previous_index < 0:
            self.comparison_result.setText("请选择当前周和对比周")
            return
            
        current_data = self.combo_current.itemData(current_index)
        previous_data = self.combo_previous.itemData(previous_index)
        
        if not current_data or not previous_data:
            self.comparison_result.setText("数据加载失败，请重试")
            return
            
        # 解析数据
        current_values = current_data
        previous_values = previous_data
        
        # 计算关键指标变化
        result_text = "📈 周环比对比结果:\n\n"
        
        # 实发订单变化
        current_orders = current_values[2]  # actual_orders
        previous_orders = previous_values[2]
        order_change = current_orders - previous_orders
        order_change_pct = (order_change / previous_orders * 100) if previous_orders > 0 else 0
        order_icon = "📈" if order_change > 0 else "📉" if order_change < 0 else "➡️"
        result_text += f"{order_icon} 实发订单: {current_orders}单 (上周: {previous_orders}单) "
        if order_change != 0:
            result_text += f"变化: {order_change:+d}单 ({order_change_pct:+.1f}%)\n"
        else:
            result_text += "持平\n"
        
        # 实发金额变化
        current_amount = current_values[3]  # actual_amount
        previous_amount = previous_values[3]
        amount_change = current_amount - previous_amount
        amount_change_pct = (amount_change / previous_amount * 100) if previous_amount > 0 else 0
        amount_icon = "📈" if amount_change > 0 else "📉" if amount_change < 0 else "➡️"
        result_text += f"{amount_icon} 实发金额: ¥{current_amount:.2f} (上周: ¥{previous_amount:.2f}) "
        if amount_change != 0:
            result_text += f"变化: ¥{amount_change:+.2f} ({amount_change_pct:+.1f}%)\n"
        else:
            result_text += "持平\n"
        
        # 净利润变化
        current_profit = current_values[17]  # net_profit
        previous_profit = previous_values[17]
        profit_change = current_profit - previous_profit
        profit_change_pct = (profit_change / previous_profit * 100) if previous_profit > 0 else 0
        profit_icon = "📈" if profit_change > 0 else "📉" if profit_change < 0 else "➡️"
        result_text += f"{profit_icon} 净利润: ¥{current_profit:.2f} (上周: ¥{previous_profit:.2f}) "
        if profit_change != 0:
            result_text += f"变化: ¥{profit_change:+.2f} ({profit_change_pct:+.1f}%)\n"
        else:
            result_text += "持平\n"
        
        # 净利率变化
        current_margin = current_values[18]  # net_margin_rate
        previous_margin = previous_values[18]
        margin_change = current_margin - previous_margin
        margin_icon = "📈" if margin_change > 0 else "📉" if margin_change < 0 else "➡️"
        result_text += f"{margin_icon} 净利率: {current_margin:.2f}% (上周: {previous_margin:.2f}%) "
        if margin_change != 0:
            result_text += f"变化: {margin_change:+.2f}%\n"
        else:
            result_text += "持平\n"
        
        # 单笔利润变化
        current_ppo = current_values[19]  # profit_per_order
        previous_ppo = previous_values[19]
        ppo_change = current_ppo - previous_ppo
        ppo_change_pct = (ppo_change / previous_ppo * 100) if previous_ppo > 0 else 0
        ppo_icon = "📈" if ppo_change > 0 else "📉" if ppo_change < 0 else "➡️"
        result_text += f"{ppo_icon} 单笔利润: ¥{current_ppo:.2f} (上周: ¥{previous_ppo:.2f}) "
        if ppo_change != 0:
            result_text += f"变化: ¥{ppo_change:+.2f} ({ppo_change_pct:+.1f}%)\n"
        else:
            result_text += "持平\n"
        
        self.comparison_result.setText(result_text)

    def open_spec_dialog_by_id(self, user_product_id):
        prod = self.db.safe_fetchall(
            "SELECT id, title FROM products WHERE name=? AND store_id=?",
            (user_product_id, self.store_id)
        )
        if prod:
            sys_id, prod_title = prod[0]
            self.open_spec_dialog(sys_id, user_product_id, prod_title)

    def open_spec_dialog(self, sys_id, prod_id, prod_title):
        """通过 main_app 打开规格对话框，避免 dialogs 依赖主模块中的 ProductSpecDialog"""
        return self.main_app.open_product_spec_dialog(self.db, sys_id, prod_id, prod_title, self)

    def import_orders(self):
        """导入订单功能"""
        file_path, _ = remembered_open_file(
            self, self.db, "选择订单文件",
            "Excel文件 (*.xlsx *.xls);;CSV文件 (*.csv);;所有支持的文件 (*.xlsx *.xls *.csv)"
        )
        if not file_path:
            return
        try:
            rows = self._read_import_rows(file_path)
            if rows is None:
                return
            headers = [str(value).strip() if value is not None else "" for value in rows[0]]
            if not headers or all(h == "" for h in headers):
                QMessageBox.warning(self, "错误", "文件没有找到有效的表头行")
                return
            col_mapping = self._auto_detect_columns(headers)
            
            # 直接弹出手动选择界面让用户确认
            col_mapping = self._show_column_mapping_dialog(headers, col_mapping)
            if col_mapping is None:
                return
            
            product_id_col = col_mapping["product_id"]
            spec_code_col = col_mapping["spec_code"]
            quantity_col = col_mapping["quantity"]
            date_col = col_mapping.get("order_date")
            status_col = col_mapping.get("order_status")
            actual_amount_col = col_mapping.get("actual_amount")
            products_in_store = self.db.safe_fetchall(
                "SELECT id, name FROM products WHERE store_id=? AND COALESCE(is_archived, 0)=0 ORDER BY sort_order", (self.store_id,)
            )
            if not products_in_store:
                QMessageBox.information(self, "提示", "当前店铺没有任何商品，请先添加商品")
                return
            product_code_to_id = {str(p[1]): p[0] for p in products_in_store}
            product_codes_in_store = set(str(p[1]) for p in products_in_store)
            all_store_specs = {}
            for prod_id, prod_code in products_in_store:
                specs = self.db.safe_fetchall(
                    "SELECT spec_code FROM product_specs WHERE product_id=?", (prod_id,)
                )
                all_store_specs[prod_id] = set(str(s[0]) for s in specs if s[0])
            
            order_data = {}
            excel_product_codes_found = set()
            total_row_count = 0
            matched_count = 0
            for row in rows[1:]:
                total_row_count += 1
                if total_row_count > 10000:
                    break
                try:
                    product_id_value = str(row[product_id_col]).strip() if product_id_col < len(row) else ""
                    spec_code_value = str(row[spec_code_col]).strip() if spec_code_col < len(row) else ""
                    quantity_value = row[quantity_col] if quantity_col is not None and quantity_col < len(row) else None
                    date_value = None
                    if date_col is not None and date_col < len(row):
                        date_value = row[date_col]
                    status_value = None
                    if status_col is not None and status_col < len(row):
                        status_value = str(row[status_col]).strip() if row[status_col] else ""
                    actual_amount_value = 0.0
                    if actual_amount_col is not None and actual_amount_col < len(row):
                        actual_amount_value = self._parse_amount_value(row[actual_amount_col])
                except:
                    continue
                if not product_id_value or product_id_value == "None":
                    continue
                if product_id_value not in product_codes_in_store:
                    continue
                excel_product_codes_found.add(product_id_value)
                prod_db_id = product_code_to_id.get(product_id_value)
                if prod_db_id is None:
                    continue
                quantity = 1
                if quantity_value is not None:
                    try:
                        quantity = max(1, int(quantity_value))
                    except (ValueError, TypeError):
                        quantity = 1
                order_date_str = self._format_import_order_date(date_value)
                spec_codes = all_store_specs.get(prod_db_id, set())
                spec_code_str = str(spec_code_value).strip() if spec_code_value else ""
                if spec_code_str and spec_code_str != "None" and spec_code_str in spec_codes:
                    if status_col is not None and status_value:
                        is_valid_order = ("已发货" in status_value) or ("已收货" in status_value)
                        if not is_valid_order:
                            continue
                    matched_count += 1
                    key = (product_id_value, spec_code_str)
                    if key not in order_data:
                        order_data[key] = {"count": 0, "refund_count": 0, "dates": [], "actual_amount": 0.0}
                    order_data[key]["count"] += quantity
                    order_data[key]["actual_amount"] += actual_amount_value
                    if order_date_str:
                        order_data[key]["dates"].append(order_date_str)
                    if status_value:
                        is_refund = "退款成功" in status_value
                        if is_refund:
                            order_data[key]["refund_count"] += quantity
            missing_product_codes = product_codes_in_store - excel_product_codes_found
            if missing_product_codes:
                msg = f"以下商品ID在表格中没有订单记录：\n{', '.join(missing_product_codes)}\n\n是否继续同步（未匹配的商品链接权重将设为0）？"
                reply = QMessageBox.question(self, "部分商品无订单", msg, QMessageBox.Yes | QMessageBox.No)
                if reply == QMessageBox.No:
                    return
            import_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # 先计算当前权重（用于保存到快照）
            temp_product_orders = {}
            for (product_code, _spec_code), data in order_data.items():
                temp_product_orders[product_code] = (
                    temp_product_orders.get(product_code, 0) + data["count"]
                )
            temp_total_orders = sum(temp_product_orders.values())
            temp_product_weights = {}
            for prod_id in self.product_weights:
                locked = self.product_weights[prod_id].get("locked", 0)
                weight = self.product_weights[prod_id].get("weight", 0)
                if locked:
                    temp_product_weights[prod_id] = weight
                elif prod_id in temp_product_orders and temp_total_orders > 0:
                    temp_product_weights[prod_id] = (temp_product_orders[prod_id] / temp_total_orders) * 100
                else:
                    temp_product_weights[prod_id] = 0

            # 保存历史快照
            total_products = len(set(prod_id for prod_id, spec_code in order_data.keys()))
            total_specs = len(order_data)
            total_orders = sum(data["count"] for data in order_data.values())
            total_amount = sum(data.get("actual_amount", 0) for data in order_data.values())
            
            # 保存快照数据到 import_history 表（包含权重信息）
            snapshot_data = json.dumps({
                "orders": {f"{prod_id}_{spec_code}": data for (prod_id, spec_code), data in order_data.items()},
                "weights": temp_product_weights
            })
            
            rows_to_insert = []
            for (product_id_val, spec_code), data in order_data.items():
                earliest_date = min(data["dates"]) if data["dates"] else None
                latest_date = max(data["dates"]) if data["dates"] else None
                date_range = f"{earliest_date}~{latest_date}" if earliest_date and latest_date else None
                rows_to_insert.append((
                    self.store_id, product_id_val, spec_code, data["count"], import_time,
                    date_range, data.get("actual_amount", 0), data.get("refund_count", 0),
                ))
            self.db.replace_imported_orders(
                self.store_id,
                (self.store_id, import_time, os.path.basename(file_path), total_products,
                 total_specs, total_orders, total_amount, snapshot_data),
                rows_to_insert,
            )
            self._update_waste_link_tasks_after_order_import(products_in_store, temp_product_orders)
            self._sync_product_refund_rates_from_imported_orders()
            self.calculate_weights_from_orders()
            self.load_products()
            self._flush_imported_data_to_archive()
            self.main_app.show_toast(f"✅ 已导入 {len(order_data)} 条订单数据")
            if hasattr(self.main_app, "update_daily_task_button_badge"):
                self.main_app.update_daily_task_button_badge()
            QTimer.singleShot(0, lambda: self.main_app.refresh_store_cards(self.store_id))
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入订单失败：\n{str(e)}")

    def _snapshot_product_order_totals(self, snapshot_data, product_codes=None):
        try:
            snapshot = json.loads(snapshot_data or "{}")
            orders = snapshot.get("orders") or {}
        except Exception:
            return {}
        sorted_codes = sorted((str(code or "") for code in (product_codes or []) if code), key=len, reverse=True)
        totals = {}
        for key, data in orders.items():
            key_text = str(key)
            product_code = next((code for code in sorted_codes if key_text.startswith(code + "_")), key_text.split("_", 1)[0])
            try:
                count = float((data or {}).get("count") or 0)
            except (TypeError, ValueError):
                count = 0
            totals[product_code] = totals.get(product_code, 0) + count
        return totals

    def _update_waste_link_tasks_after_order_import(self, products_in_store, current_product_orders):
        histories = self.db.safe_fetchall(
            "SELECT snapshot_data FROM import_history WHERE store_id=? ORDER BY id DESC LIMIT 2",
            (self.store_id,),
        )
        if len(histories) < 2:
            return
        product_codes = [str(product_code or "") for _product_db_id, product_code in products_in_store]
        previous_orders = self._snapshot_product_order_totals(histories[1][0], product_codes)
        today = datetime.now()
        created_time = today.strftime("%Y-%m-%d %H:%M:%S")
        recovered_ids = []
        created = 0
        violation_ids = {
            row[0] for row in self.db.safe_fetchall(
                "SELECT id FROM products WHERE store_id=? AND COALESCE(is_violation, 0)=1",
                (self.store_id,),
            )
        }
        for product_db_id, product_code in products_in_store:
            if product_db_id in violation_ids:
                continue
            product_code = str(product_code or "")
            if current_product_orders.get(product_code, 0) > 0:
                recovered_ids.append(product_db_id)
                continue
            if previous_orders.get(product_code, 0) > 0:
                continue
            exists = self.db.safe_fetchall(
                """SELECT id FROM daily_tasks
                   WHERE store_id=? AND product_id=? AND is_completed=0
                     AND task_content LIKE '【废物链接】%' LIMIT 1""",
                (self.store_id, product_db_id),
            )
            if exists:
                continue
            title_rows = self.db.safe_fetchall("SELECT title FROM products WHERE id=?", (product_db_id,))
            title = str(title_rows[0][0] or "") if title_rows else ""
            self.db.safe_execute(
                """INSERT INTO daily_tasks
                   (store_id, product_id, year, month, day, task_content, created_time)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (
                    self.store_id, product_db_id, today.year, today.month, today.day,
                    f"【废物链接】连续两周导入订单表无订单。商品ID：{product_code}；标题：{title}",
                    created_time,
                ),
            )
            created += 1
        if recovered_ids:
            placeholders = ",".join("?" for _ in recovered_ids)
            self.db.safe_execute(
                f"""UPDATE daily_tasks SET is_completed=1
                    WHERE store_id=? AND is_completed=0
                      AND task_content LIKE '【废物链接】%'
                      AND product_id IN ({placeholders})""",
                (self.store_id, *recovered_ids),
            )
        if created and getattr(self, "main_app", None) and hasattr(self.main_app, "show_toast"):
            self.main_app.show_toast(f"已新增 {created} 个废物链接任务")

    def _sync_product_refund_rates_from_imported_orders(self):
        """把订单规格毛利权重表里的商品退款率同步到 products.return_rate。"""
        try:
            rows = self.db.safe_fetchall(
                """
                SELECT
                    product_id,
                    COALESCE(SUM(COALESCE(order_count, 0)), 0) AS total_orders,
                    COALESCE(SUM(COALESCE(refund_count, 0)), 0) AS total_refunds
                FROM imported_orders
                WHERE store_id=?
                GROUP BY product_id
                """,
                (self.store_id,)
            )
            updated_rates = {}
            for product_code, total_orders, total_refunds in rows:
                product_code = str(product_code or "").strip()
                total_orders = float(total_orders or 0)
                total_refunds = float(total_refunds or 0)
                if not product_code or total_orders <= 0:
                    continue
                refund_rate = max(0.0, min(100.0, total_refunds / total_orders * 100))
                updated_rates[product_code] = refund_rate
                self.db.safe_execute(
                    "UPDATE products SET return_rate=? WHERE store_id=? AND name=?",
                    (refund_rate, self.store_id, product_code)
                )

            product_spec_dialog = getattr(getattr(self, "main_app", None), "product_spec_dialog", None)
            if product_spec_dialog is not None:
                product_code = str(getattr(product_spec_dialog, "product_code", "") or "").strip()
                if product_code in updated_rates and hasattr(product_spec_dialog, "return_rate_input"):
                    product_spec_dialog.return_rate_input.setText(f"{updated_rates[product_code]:.2f}")
                    if hasattr(product_spec_dialog, "calculate_roi_metrics"):
                        product_spec_dialog.calculate_roi_metrics()
        except Exception as e:
            print(f"同步商品退款率失败: {e}")

    def _auto_detect_columns(self, headers):
        """自动检测列映射"""
        mapping = {"product_id": None, "spec_code": None, "quantity": None, "order_date": None, "order_status": None, "actual_amount": None}
        product_id_keywords = ["商品id", "商品ID", "id", "产品id", "产品ID", "product_id"]
        spec_code_keywords = ["规格编码", "规格code", "spec_code", "规格code", "sku", "SKU"]
        quantity_keywords = ["数量", "订单数量", "quantity", "count", "num", "销售数量"]
        order_date_keywords = ["日期", "date", "时间", "time", "订单日期", "下单日期", "成交时间"]
        order_status_keywords = ["订单状态", "状态", "order_status", "order state"]
        for idx, header in enumerate(headers):
            header_lower = header.lower().strip()
            if mapping["spec_code"] is None and "规格维度" in header:
                mapping["spec_code"] = idx
            if mapping["actual_amount"] is None:
                for kw in ["实收金额", "实付金额", "支付金额", "成交金额", "订单实收", "actual_amount", "pay_amount", "payment"]:
                    if kw.lower() in header_lower:
                        mapping["actual_amount"] = idx
                        break
            if mapping["product_id"] is None:
                for kw in product_id_keywords:
                    if kw in header_lower:
                        mapping["product_id"] = idx
                        break
            if mapping["spec_code"] is None:
                for kw in spec_code_keywords:
                    if kw in header_lower:
                        mapping["spec_code"] = idx
                        break
            if mapping["quantity"] is None:
                for kw in quantity_keywords:
                    if kw in header_lower:
                        mapping["quantity"] = idx
                        break
            if mapping["order_date"] is None:
                for kw in order_date_keywords:
                    if kw in header_lower:
                        mapping["order_date"] = idx
                        break
            if mapping["order_status"] is None:
                for kw in order_status_keywords:
                    if kw in header:
                        mapping["order_status"] = idx
                        break
        return mapping

    def _show_column_mapping_dialog(self, headers, auto_mapping):
        """显示列映射对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("📋 列映射选择")
        dialog.resize(500, 400)
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel("请为每个字段选择对应的Excel列："))
        layout.addSpacing(10)
        
        combo_product_id = QComboBox()
        combo_product_id.addItems(["-- 不选择 --"] + headers)
        if auto_mapping.get("product_id") is not None:
            combo_product_id.setCurrentIndex(auto_mapping["product_id"] + 1)
        layout.addWidget(QLabel("商品ID列 *："))
        layout.addWidget(combo_product_id)
        
        combo_spec_code = QComboBox()
        combo_spec_code.addItems(["-- 不选择 --"] + headers)
        if auto_mapping.get("spec_code") is not None:
            combo_spec_code.setCurrentIndex(auto_mapping["spec_code"] + 1)
        layout.addWidget(QLabel("规格编码列 *："))
        layout.addWidget(combo_spec_code)
        
        combo_quantity = QComboBox()
        combo_quantity.addItems(["-- 不选择（默认为1） --"] + headers)
        if auto_mapping.get("quantity") is not None:
            combo_quantity.setCurrentIndex(auto_mapping["quantity"] + 1)
        layout.addWidget(QLabel("数量列："))
        layout.addWidget(combo_quantity)
        
        combo_order_date = QComboBox()
        combo_order_date.addItems(["-- 不选择 --"] + headers)
        if auto_mapping.get("order_date") is not None:
            combo_order_date.setCurrentIndex(auto_mapping["order_date"] + 1)
        layout.addWidget(QLabel("订单日期列："))
        layout.addWidget(combo_order_date)
        
        combo_order_status = QComboBox()
        combo_order_status.addItems(["-- 不选择 --"] + headers)
        if auto_mapping.get("order_status") is not None:
            combo_order_status.setCurrentIndex(auto_mapping["order_status"] + 1)
        layout.addWidget(QLabel("订单状态列（用于识别退款）："))
        layout.addWidget(combo_order_status)

        combo_actual_amount = QComboBox()
        combo_actual_amount.addItems(["-- 不选择 --"] + headers)
        if auto_mapping.get("actual_amount") is not None:
            combo_actual_amount.setCurrentIndex(auto_mapping["actual_amount"] + 1)
        layout.addWidget(QLabel("实收金额列："))
        layout.addWidget(combo_actual_amount)
        
        layout.addSpacing(20)
        btn_layout = QHBoxLayout()
        btn_ok = QPushButton("确认")
        btn_ok.setStyleSheet("QPushButton { background-color: #27ae60; color: white; padding: 8px 20px; border-radius: 4px; }")
        btn_cancel = QPushButton("取消")
        btn_layout.addWidget(btn_ok)
        btn_layout.addWidget(btn_cancel)
        layout.addLayout(btn_layout)
        
        result = {"product_id": None, "spec_code": None, "quantity": None, "order_date": None, "order_status": None, "actual_amount": None}
        
        def on_ok():
            result["product_id"] = combo_product_id.currentIndex() - 1 if combo_product_id.currentIndex() > 0 else None
            result["spec_code"] = combo_spec_code.currentIndex() - 1 if combo_spec_code.currentIndex() > 0 else None
            result["quantity"] = combo_quantity.currentIndex() - 1 if combo_quantity.currentIndex() > 0 else None
            result["order_date"] = combo_order_date.currentIndex() - 1 if combo_order_date.currentIndex() > 0 else None
            result["order_status"] = combo_order_status.currentIndex() - 1 if combo_order_status.currentIndex() > 0 else None
            result["actual_amount"] = combo_actual_amount.currentIndex() - 1 if combo_actual_amount.currentIndex() > 0 else None
            dialog.accept()
        
        btn_ok.clicked.connect(on_ok)
        btn_cancel.clicked.connect(dialog.reject)
        
        if dialog.exec_() == QDialog.Accepted:
            return result
        return None
    
    def update_orders_display(self):
        """更新单量列显示"""
        self._normalize_imported_order_store_ids()
        # 获取当前导入的数据
        current_data = self.db.safe_fetchall("""
            SELECT product_id, spec_code, order_count, refund_count
            FROM imported_orders
            WHERE store_id=?
        """, (self.store_id,))

        # 计算每个商品的总订单数和退款数（直接用 product_id 即商品ID字符串）
        prod_order_totals = {}
        prod_refund_data = {}
        total_refund_sum = 0
        for prod_id, spec_code, order_count, refund_count in current_data:
            if prod_id not in prod_order_totals:
                prod_order_totals[prod_id] = 0
                prod_refund_data[prod_id] = []
            prod_order_totals[prod_id] += order_count or 0
            prod_refund_data[prod_id].append((spec_code, order_count or 0, refund_count or 0))
            total_refund_sum += refund_count or 0

        # 遍历表格行
        for row in range(self.table.rowCount()):
            prod_id_item = self.table.item(row, 1)
            if not prod_id_item:
                continue
            user_product_id = prod_id_item.data(Qt.UserRole)
            if not user_product_id:
                continue

            order_label_widget = self.table.cellWidget(row, 8)
            if order_label_widget and user_product_id:
                order_label = order_label_widget.layout().itemAt(0).widget()
                if order_label:
                    if user_product_id in prod_order_totals:
                        order_label.setText(f"{prod_order_totals[user_product_id]}单")
                        order_label.setStyleSheet("color: black; font-size: 19px;")
                    else:
                        order_label.setText("0单")
                        order_label.setStyleSheet("color: #95a5a6; font-size: 19px;")

            refund_orders_label = self.table.cellWidget(row, 12)
            refund_ratio_label = self.table.cellWidget(row, 13)
            if user_product_id and user_product_id in prod_refund_data:
                spec_data = prod_refund_data[user_product_id]
                total_orders = sum(d[1] for d in spec_data)
                total_refund = sum(d[2] for d in spec_data)
                if total_orders > 0 and total_refund > 0:
                    refund_rate = total_refund / total_orders * 100
                    if refund_orders_label:
                        refund_orders_label.setText(f"{refund_rate:.2f}%")
                        refund_orders_label.setStyleSheet("color: #e74c3c; font-size: 19px; font-weight: bold;")
                    max_refund_spec = None
                    max_refund_rate_val = -1
                    valid_spec_codes = self._get_valid_spec_codes(user_product_id)
                    display_spec_data = [
                        item for item in spec_data
                        if not valid_spec_codes or (item[0] and str(item[0]).strip() in valid_spec_codes)
                    ]
                    for spec_code, oc, rc in display_spec_data:
                        if oc > 0 and rc > 0:
                            sr = rc / oc
                            if sr > max_refund_rate_val:
                                max_refund_rate_val = sr
                                max_refund_spec = spec_code
                    if refund_ratio_label:
                        refund_spec_mode = self._get_spec_display_mode("store_margin_refund_spec_display")
                        if max_refund_spec:
                            refund_spec_text = self._get_spec_display_text(user_product_id, max_refund_spec, refund_spec_mode)
                            self._set_spec_label_text(refund_ratio_label, refund_spec_text, refund_spec_mode, alert=True)
                        else:
                            refund_ratio_label.setText("无")
                            refund_ratio_label.setStyleSheet("color: #95a5a6; font-size: 19px;")
                            self._set_spec_label_text(refund_ratio_label, "无", refund_spec_mode)
                else:
                    if refund_orders_label:
                        refund_orders_label.setText("无")
                        refund_orders_label.setStyleSheet("color: #95a5a6; font-size: 19px;")
                    if refund_ratio_label:
                        refund_ratio_label.setText("无")
                        refund_ratio_label.setStyleSheet("color: #95a5a6; font-size: 19px;")
            else:
                if refund_orders_label:
                    refund_orders_label.setText("无")
                    refund_orders_label.setStyleSheet("color: #95a5a6; font-size: 19px;")
                if refund_ratio_label:
                    refund_ratio_label.setText("无")
                    refund_ratio_label.setStyleSheet("color: #95a5a6; font-size: 19px;")

        self.update_total_orders_label()
        self.update_order_range_label()

        # 强制刷新表格显示
        from PyQt5.QtWidgets import QApplication
        from PyQt5.QtCore import QTimer
        QApplication.processEvents()
    
    def update_current_history_label(self):
        """更新当前使用数据标签"""
        if not hasattr(self, 'lbl_current_history'):
            return

        latest_record = self.db.safe_fetchall("""
            SELECT start_date, end_date, actual_orders, actual_amount, gross_profit,
                   refund_amount, refund_orders, promotion_fee, deduction, other_service, other,
                   gross_margin_rate, refund_rate_by_amount, refund_rate_by_orders,
                   unit_price, promotion_ratio, tech_fee,
                   net_profit, net_margin_rate, profit_per_order
            FROM manual_margin_data WHERE store_id=? ORDER BY start_date DESC, end_date DESC LIMIT 1
        """, (self.store_id,))

        if latest_record:
            record = latest_record[0]
            start_date = record[0] if record[0] else ""
            end_date = record[1] if record[1] else ""

            if start_date and len(start_date) >= 10:
                start_display = start_date[5:10]
            else:
                start_display = start_date

            if end_date and len(end_date) >= 10:
                end_display = end_date[5:10]
            else:
                end_display = end_date

            if start_display == end_display:
                date_str = start_display
            else:
                date_str = f"{start_display}~{end_display}"

            net_profit = record[17] if record[17] else 0
            net_margin = record[18] if record[18] else 0

            self.lbl_current_history.setText(
                f"📍 最新: {date_str} | 净利润: ¥{net_profit:.2f} ({net_margin:.2f}%)"
            )
            self.lbl_current_history.setStyleSheet("""
                QLabel {
                    color: #27ae60;
                    font-size: 12px;
                    font-weight: bold;
                    padding: 6px 12px;
                    background-color: #e8f8f0;
                    border-radius: 4px;
                }
            """)
        else:
            self.lbl_current_history.setText("📍 当前: 暂无数据")
            self.lbl_current_history.setStyleSheet("""
                QLabel {
                    color: #7f8c8d;
                    font-size: 12px;
                    padding: 6px 12px;
                    background-color: #f5f5f5;
                    border-radius: 4px;
                }
            """)
    
    def update_total_orders_label(self):
        """更新总单量标签"""
        self._normalize_imported_order_store_ids()
        total_data = self.db.safe_fetchall("""
            SELECT SUM(io.order_count) FROM imported_orders io
            WHERE io.store_id=? AND EXISTS (
                SELECT 1 FROM products p
                JOIN product_specs ps ON ps.product_id=p.id AND ps.spec_code=io.spec_code
                  AND COALESCE(ps.is_temporarily_off_shelf, 0)=0
                WHERE p.store_id=io.store_id AND p.name=io.product_id
                  AND COALESCE(p.is_archived, 0)=0 AND COALESCE(p.is_violation, 0)=0
            )
        """, (self.store_id,))
        total = total_data[0][0] if total_data and total_data[0][0] else 0
        self.lbl_total_orders.setText(f"总单量: {total}")

    def update_order_range_label(self):
        """更新当前订单时间范围标签"""
        self._normalize_imported_order_store_ids()
        date_data = self.db.safe_fetchall("""
            SELECT order_date FROM imported_orders WHERE store_id=? AND order_date IS NOT NULL
        """, (self.store_id,))
        if not date_data:
            self.lbl_order_range.setText("当前订单时间范围: --")
            return
        all_dates = []
        for (date_range,) in date_data:
            if date_range and '~' in date_range:
                parts = date_range.split('~')
                all_dates.extend(parts)
            elif date_range:
                all_dates.append(date_range)
        if not all_dates:
            self.lbl_order_range.setText("当前订单时间范围: --")
            return
        try:
            parsed_dates = []
            for d in all_dates:
                if '/' in d:
                    m, day = d.split('/')
                    parsed_dates.append((int(m), int(day)))
            if parsed_dates:
                parsed_dates.sort()
                min_d = parsed_dates[0]
                max_d = parsed_dates[-1]
                if min_d != max_d:
                    range_str = f"{min_d[0]}/{min_d[1]}-{max_d[0]}/{max_d[1]}"
                else:
                    range_str = f"{min_d[0]}/{min_d[1]}"
                self.lbl_order_range.setText(f"当前订单时间范围: {range_str}")
            else:
                self.lbl_order_range.setText("当前订单时间范围: --")
        except:
            self.lbl_order_range.setText("当前订单时间范围: --")

    def update_compare_columns(self):
        """更新对比列数据 - 按订单时间范围与上一期对比"""
        self._normalize_imported_order_store_ids()
        # 获取当前导入的数据
        current_data = self.db.safe_fetchall("""
            SELECT product_id, spec_code, order_count, order_date
            FROM imported_orders
            WHERE store_id=?
        """, (self.store_id,))
        
        # 如果没有任何数据，显示 "-"
        if not current_data:
            for row in range(self.table.rowCount()):
                weight_compare_widget = self.table.cellWidget(row, 7)
                if weight_compare_widget:
                    weight_label = weight_compare_widget.layout().itemAt(0).widget()
                    weight_label.setText("-")
                    weight_label.setStyleSheet("color: #95a5a6; font-size: 19px;")
                    
                order_compare_widget = self.table.cellWidget(row, 9)
                if order_compare_widget:
                    order_label = order_compare_widget.layout().itemAt(0).widget()
                    order_label.setText("-")
                    order_label.setStyleSheet("color: #95a5a6; font-size: 19px;")
            return
        
        # 获取当前订单的日期范围（用于找上一期）
        current_date_range = None
        for _, _, _, order_date in current_data:
            if order_date and '~' in order_date:
                current_date_range = order_date
                break
        
        # 解析当前日期范围获取结束日期
        current_end_date = None
        if current_date_range:
            try:
                parts = current_date_range.split('~')
                if len(parts) == 2:
                    current_end_date = parts[1].strip()
            except:
                pass
        
        # 找上一期历史记录（按订单日期范围排序，找小于当前结束日期的最接近的那一期）
        last_history_data = None
        if current_end_date:
            # 获取所有历史记录
            all_history = self.db.safe_fetchall("""
                SELECT id, snapshot_data
                FROM import_history
                WHERE store_id=? AND snapshot_data IS NOT NULL AND snapshot_data != ''
                ORDER BY import_time DESC
            """, (self.store_id,))

            for hist_id, snapshot_data in all_history:
                try:
                    snapshot = json.loads(snapshot_data)
                    # 从订单数据中解析日期范围
                    orders = snapshot.get("orders", {})
                    all_dates = []
                    for key, data in orders.items():
                        if isinstance(data, dict) and "dates" in data:
                            for date_val in data.get("dates", []):
                                if date_val and '/' in date_val:
                                    try:
                                        if '~' in date_val:
                                            for p in date_val.split('~'):
                                                if '/' in p:
                                                    m, d = p.split('/')
                                                    all_dates.append((int(m), int(d)))
                                        else:
                                            m, d = date_val.split('/')
                                            all_dates.append((int(m), int(d)))
                                    except:
                                        pass
                    if all_dates:
                        all_dates.sort()
                        prev_end_date = f"{all_dates[-1][0]}/{all_dates[-1][1]}"
                        # 找小于当前结束日期的最接近的那一期
                        curr_parts = current_end_date.split('/')
                        curr_m, curr_d = int(curr_parts[0]), int(curr_parts[1])
                        prev_m, prev_d = int(all_dates[-1][0]), int(all_dates[-1][1])
                        if prev_m < curr_m or (prev_m == curr_m and prev_d < curr_d):
                            last_history_data = (snapshot_data, snapshot)
                            break
                except:
                    pass

        # 如果没找到按日期的对比，说明没有更早的历史，显示 "无"
        if not last_history_data:
            for row in range(self.table.rowCount()):
                weight_compare_widget = self.table.cellWidget(row, 7)
                if weight_compare_widget:
                    weight_label = weight_compare_widget.layout().itemAt(0).widget()
                    weight_label.setText("无")
                    weight_label.setStyleSheet("color: #95a5a6; font-size: 19px;")
                    
                order_compare_widget = self.table.cellWidget(row, 9)
                if order_compare_widget:
                    order_label = order_compare_widget.layout().itemAt(0).widget()
                    order_label.setText("无")
                    order_label.setStyleSheet("color: #95a5a6; font-size: 19px;")
            return
        
        # 解析上一期的快照数据
        try:
            last_snapshot = json.loads(last_history_data[0])
            last_orders = last_snapshot.get("orders", {})
        except:
            for row in range(self.table.rowCount()):
                weight_compare_widget = self.table.cellWidget(row, 7)
                if weight_compare_widget:
                    weight_label = weight_compare_widget.layout().itemAt(0).widget()
                    weight_label.setText("无")
                    weight_label.setStyleSheet("color: #95a5a6; font-size: 19px;")
                    
                order_compare_widget = self.table.cellWidget(row, 9)
                if order_compare_widget:
                    order_label = order_compare_widget.layout().itemAt(0).widget()
                    order_label.setText("无")
                    order_label.setStyleSheet("color: #95a5a6; font-size: 19px;")
            return
        
        # 计算每个商品的总订单数
        product_current = {}
        for prod_id, spec_code, order_count, _ in current_data:
            if prod_id not in product_current:
                product_current[prod_id] = {"orders": 0, "specs": {}}
            product_current[prod_id]["orders"] += order_count
            product_current[prod_id]["specs"][spec_code] = order_count

        # 解析上一期快照数据的 key（使用商品ID字符串）
        product_last = {}
        for key, data in last_orders.items():
            parts = key.split("_", 1)
            if len(parts) >= 2:
                user_product_id = parts[0]
                spec_code = parts[1]
                if user_product_id not in product_last:
                    product_last[user_product_id] = {"orders": 0, "specs": {}}
                product_last[user_product_id]["orders"] += data["count"]
                product_last[user_product_id]["specs"][spec_code] = data["count"]

        # 更新表格中的对比列
        for row in range(self.table.rowCount()):
            prod_id_item = self.table.item(row, 1)
            if not prod_id_item:
                continue
            user_product_id = prod_id_item.data(Qt.UserRole)
            if not user_product_id:
                continue

            # 权重对比（基于权重数值变化）
            weight_compare_widget = self.table.cellWidget(row, 7)
            if not weight_compare_widget:
                continue
            weight_label = weight_compare_widget.layout().itemAt(0).widget()
            
            current_weight = self.product_weights.get(user_product_id, {}).get("weight", 0)
            
            last_weights = last_snapshot.get("weights", {})
            last_weight = last_weights.get(user_product_id, None)
            
            if last_weight is None:
                if user_product_id in product_last:
                    last_total_orders = sum(pdata["orders"] for pdata in product_last.values())
                    if last_total_orders > 0:
                        last_weight = (product_last[user_product_id]["orders"] / last_total_orders) * 100
                    else:
                        last_weight = 0
                else:
                    last_weight = 0
            
            weight_change = current_weight - last_weight
            
            if abs(weight_change) < 0.001:
                weight_label.setText("⚪ 0.00%")
                weight_label.setStyleSheet("color: #7f8c8d; font-size: 19px;")
            elif weight_change > 0:
                weight_label.setText(f"↑{weight_change:.2f}%")
                weight_label.setStyleSheet("color: #27ae60; font-size: 19px; font-weight: bold;")
            else:
                weight_label.setText(f"↓{abs(weight_change):.2f}%")
                weight_label.setStyleSheet("color: #c0392b; font-size: 19px; font-weight: bold;")

            # 单量对比
            order_compare_widget = self.table.cellWidget(row, 9)
            if not order_compare_widget:
                continue
            order_label = order_compare_widget.layout().itemAt(0).widget()
            if user_product_id and user_product_id in product_current and user_product_id in product_last:
                current_orders = product_current[user_product_id]["orders"]
                last_orders_count = product_last[user_product_id]["orders"]
                order_change = current_orders - last_orders_count

                if order_change > 0:
                    order_label.setText(f"↑{order_change}")
                    order_label.setStyleSheet("color: #27ae60; font-size: 19px; font-weight: bold;")
                elif order_change < 0:
                    order_label.setText(f"↓{abs(order_change)}")
                    order_label.setStyleSheet("color: #c0392b; font-size: 19px; font-weight: bold;")
                else:
                    order_label.setText("⚪ 0")
                    order_label.setStyleSheet("color: #7f8c8d; font-size: 19px;")
            else:
                # 商品不在对比数据中，显示 "无"
                order_label.setText("无")
                order_label.setStyleSheet("color: #95a5a6; font-size: 19px;")

    def show_import_history(self):
        """显示导入历史记录对话框"""
        dialog = ImportHistoryDialog(self.store_id, self.store_name, self.db, self)
        dialog.exec_()
        self.update_total_orders_label()
        self.update_order_range_label()


# ==================== 导入历史记录对话框类 ====================

class ImportHistoryDialog(QDialog):
    """导入历史记录对话框"""
    def __init__(self, store_id, store_name, db, parent=None):
        super().__init__(parent)
        self.store_id = store_id
        self.store_name = store_name
        self.db = db
        self.parent_window = parent
        self.setWindowTitle(f"📜 {store_name} - 全部记录")
        self.resize(900, 600)
        self.setStyleSheet("background-color: #f5f5f5;")
        self.init_ui()
        self.load_history()
        self.check_old_data()
    
    def check_old_data(self):
        """检查是否有旧数据（imported_orders有数据但历史快照无效）"""
        # 检查 imported_orders 是否有数据
        current_orders = self.db.safe_fetchall(
            "SELECT product_id, spec_code, order_count, import_time, actual_amount FROM imported_orders WHERE store_id=?", (self.store_id,)
        )
        has_current_orders = current_orders and len(current_orders) > 0

        # 检查历史记录是否有有效快照（不仅要存在，还要 snapshot_data 不为空且能解析）
        history_records = self.db.safe_fetchall(
            "SELECT snapshot_data FROM import_history WHERE store_id=? ORDER BY import_time DESC LIMIT 1",
            (self.store_id,)
        )
        has_valid_history = False
        if history_records and history_records[0][0]:
            try:
                snapshot = json.loads(history_records[0][0])
                if snapshot and "orders" in snapshot:
                    has_valid_history = True
            except:
                has_valid_history = False

        # 如果有当前订单但没有有效历史，自动为旧数据创建快照
        if has_current_orders and not has_valid_history:
            # 根据旧数据创建快照
            from datetime import datetime
            orders_data = {}
            total_products = set()
            total_specs = 0
            total_orders = 0
            total_amount = 0
            
            for prod_id, spec_code, order_count, import_time, actual_amount in current_orders:
                key = f"{prod_id}_{spec_code}"
                orders_data[key] = {"count": order_count, "dates": [], "actual_amount": actual_amount or 0}
                total_products.add(prod_id)
                total_specs += 1
                total_orders += order_count
                total_amount += actual_amount or 0
            
            # 创建历史快照
            snapshot_data = json.dumps({"orders": orders_data})
            self.db.safe_execute("""
                INSERT INTO import_history (store_id, import_time, file_name, total_products, total_specs, total_orders, total_amount, snapshot_data)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (self.store_id, import_time if import_time else datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                  "旧数据导入", len(total_products), total_specs, total_orders, total_amount, snapshot_data))
            
            if self.parent_window:
                self.parent_window.main_app.show_toast("✅ 已为旧数据创建历史快照")
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # 标题
        title_label = QLabel(f"📊 {self.store_name} - 订单全部记录")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50; padding: 10px;")
        layout.addWidget(title_label)
        
        # 表格
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            "导入时间", "文件名", "订单时间范围", "商品数", "总单量", "操作"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        layout.addWidget(self.table)
        
        # 底部按钮
        btn_layout = QHBoxLayout()
        
        self.btn_close = QPushButton("关闭")
        self.btn_close.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                padding: 8px 16px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        self.btn_close.clicked.connect(self.accept)
        
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_close)
        
        layout.addLayout(btn_layout)
    
    def load_history(self):
        """加载历史记录 - 按订单日期范围排序（最新日期排最上面）"""
        # 调试：检查 snapshot_data 格式
        self.db._check_and_migrate_snapshot_data()

        records = self.db.safe_fetchall("""
            SELECT id, import_time, file_name, total_products, total_specs, total_orders, total_amount, snapshot_data
            FROM import_history
            WHERE store_id=?
        """, (self.store_id,))

        # 计算每个记录的订单日期范围，并按结束日期降序排序
        def get_order_end_date(record):
            _, _, _, _, _, _, _, snapshot_data = record
            if snapshot_data:
                try:
                    snapshot = json.loads(snapshot_data)
                    if snapshot and "orders" in snapshot:
                        all_dates = []
                        for key, data in snapshot["orders"].items():
                            if isinstance(data, dict) and "dates" in data:
                                for date_val in data.get("dates", []):
                                    if date_val and '/' in date_val:
                                        try:
                                            if '~' in date_val:
                                                for p in date_val.split('~'):
                                                    if '/' in p:
                                                        m, d = p.split('/')
                                                        all_dates.append((int(m), int(d)))
                                            else:
                                                m, d = date_val.split('/')
                                                all_dates.append((int(m), int(d)))
                                        except:
                                            pass
                        if all_dates:
                            all_dates.sort()
                            return all_dates[-1]  # 返回结束日期
                except:
                    pass
            return (0, 0)  # 默认最小的日期

        # 按订单结束日期降序排序
        records.sort(key=get_order_end_date, reverse=True)

        self.table.setRowCount(len(records))

        for row, record in enumerate(records):
            hist_id, import_time, file_name, total_products, total_specs, total_orders, total_amount, snapshot_data = record

            # 导入时间
            time_item = QTableWidgetItem(import_time)
            time_item.setFlags(Qt.ItemIsEnabled)
            self.table.setItem(row, 0, time_item)

            # 文件名
            file_item = QTableWidgetItem(file_name if file_name else "未知")
            file_item.setFlags(Qt.ItemIsEnabled)
            self.table.setItem(row, 1, file_item)

            # 订单时间范围（从snapshot_data中提取所有日期计算范围）
            order_range_str = "无日期"
            if snapshot_data:
                try:
                    snapshot = json.loads(snapshot_data)
                    if snapshot and "orders" in snapshot:
                        all_dates = []
                        for key, data in snapshot["orders"].items():
                            if isinstance(data, dict) and "dates" in data:
                                for date_val in data.get("dates", []):
                                    if date_val and '/' in date_val:
                                        try:
                                            m, d = date_val.split('/')
                                            all_dates.append((int(m), int(d)))
                                        except:
                                            pass
                        if all_dates:
                            all_dates.sort()
                            min_date = all_dates[0]
                            max_date = all_dates[-1]
                            if min_date != max_date:
                                order_range_str = f"{min_date[0]}/{min_date[1]}-{max_date[0]}/{max_date[1]}"
                            else:
                                order_range_str = f"{min_date[0]}/{min_date[1]}"
                except:
                    order_range_str = "解析失败"

            range_item = QTableWidgetItem(order_range_str)
            range_item.setFlags(Qt.ItemIsEnabled)
            range_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 2, range_item)

            # 商品数
            prod_item = QTableWidgetItem(str(total_products))
            prod_item.setFlags(Qt.ItemIsEnabled)
            prod_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 3, prod_item)

            # 总单量
            orders_item = QTableWidgetItem(str(total_orders))
            orders_item.setFlags(Qt.ItemIsEnabled)
            orders_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 4, orders_item)

            # 操作按钮（应用和删除）
            btn_widget = QWidget()
            btn_layout = QHBoxLayout(btn_widget)
            btn_layout.setContentsMargins(1, 1, 1, 1)
            btn_layout.setAlignment(Qt.AlignCenter)
            btn_layout.setSpacing(1)
            
            btn_apply = QPushButton("应用")
            btn_apply.setFixedSize(45, 28)
            btn_apply.setStyleSheet("""
                QPushButton {
                    background-color: #27ae60;
                    color: white;
                    border: none;
                    border-radius: 3px;
                    font-size: 11px;
                    font-weight: bold;
                    padding: 1px 2px;
                }
                QPushButton:hover {
                    background-color: #229954;
                }
            """)
            btn_apply.clicked.connect(lambda checked, hid=hist_id: self.apply_history(hid))
            btn_layout.addWidget(btn_apply)
            
            btn_delete = QPushButton("删除")
            btn_delete.setFixedSize(45, 28)
            btn_delete.setStyleSheet("""
                QPushButton {
                    background-color: #e74c3c;
                    color: white;
                    border: none;
                    border-radius: 3px;
                    font-size: 11px;
                    padding: 1px 2px;
                }
                QPushButton:hover {
                    background-color: #c0392b;
                }
            """)
            btn_delete.clicked.connect(lambda checked, hid=hist_id: self.delete_single_history(hid))
            btn_layout.addWidget(btn_delete)
            
            self.table.setCellWidget(row, 5, btn_widget)
    
    def delete_single_history(self, history_id):
        """删除单条历史记录"""
        # 检查是否删除的是最新的历史记录，如果是则同时清空 imported_orders
        latest_history = self.db.safe_fetchall("""
            SELECT id FROM import_history WHERE store_id=? ORDER BY import_time DESC LIMIT 1
        """, (self.store_id,))
        
        is_latest = latest_history and latest_history[0][0] == history_id
        
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("确认删除")
        msg_box.setText("确定要删除这条导入记录吗？")
        msg_box.setIcon(QMessageBox.Warning)
        
        yes_btn = msg_box.addButton("确定", QMessageBox.YesRole)
        no_btn = msg_box.addButton("取消", QMessageBox.NoRole)
        
        yes_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        no_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 20px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        
        msg_box.setDefaultButton(no_btn)
        msg_box.exec_()
        
        if msg_box.clickedButton() == yes_btn:
            self.db.safe_execute("DELETE FROM import_history WHERE id=?", (history_id,))
            
            # 检查是否还有历史记录
            remaining_history = self.db.safe_fetchall("""
                SELECT COUNT(*) FROM import_history WHERE store_id=?
            """, (self.store_id,))
            
            # 如果删除的是最新记录或者没有任何历史记录了，清空 imported_orders
            if is_latest or (remaining_history and remaining_history[0][0] == 0):
                self.db.safe_execute("DELETE FROM imported_orders WHERE store_id=?", (self.store_id,))

            self.load_history()

            self.parent_window.update_total_orders_label()
            self.parent_window.update_order_range_label()
            self.parent_window.main_app.show_toast("✅ 已删除")
    
    def delete_selected(self):
        """删除选中的历史记录"""
        selected_rows = []
        for row in range(self.table.rowCount()):
            check_item = self.table.item(row, 0)
            if check_item and check_item.checkState() == Qt.Checked:
                selected_rows.append(row)
        
        if not selected_rows:
            QMessageBox.warning(self, "提示", "请先选择要删除的记录")
            return
        
        reply = QMessageBox.question(
            self,
            "确认删除",
            f"确定要删除选中的 {len(selected_rows)} 条记录吗？",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            for row in reversed(selected_rows):
                hist_id = self.table.item(row, 0).data(Qt.UserRole)
                self.db.safe_execute("DELETE FROM import_history WHERE id=?", (hist_id,))
            self.load_history()
    
    def apply_history(self, history_id):
        """应用历史记录的订单数据"""
        # 获取历史记录
        history_records = self.db.safe_fetchall(
            "SELECT import_time, snapshot_data FROM import_history WHERE id=?",
            (history_id,)
        )

        if not history_records or not history_records[0][1]:
            return

        try:
            snapshot = json.loads(history_records[0][1])
            orders_data = snapshot.get("orders", {})
        except:
            return

        # 清空当前的 imported_orders
        self.db.safe_execute("DELETE FROM imported_orders WHERE store_id=?", (self.store_id,))

        # 恢复历史订单数据
        for key, data in orders_data.items():
            parts = key.split("_", 1)
            if len(parts) >= 2:
                user_product_id = parts[0]
                spec_code = parts[1]
                order_count = data.get("count", 0)
                refund_count = data.get("refund_count", 0)
                actual_amount = data.get("actual_amount", 0)
                dates = data.get("dates", [])
                earliest_date = min(dates) if dates else None
                latest_date = max(dates) if dates else None
                date_range = f"{earliest_date}~{latest_date}" if earliest_date and latest_date else None

                self.db.safe_execute("""
                    INSERT OR REPLACE INTO imported_orders
                    (store_id, product_id, spec_code, order_count, import_time, order_date, actual_amount, refund_count)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (self.store_id, user_product_id, spec_code, order_count,
                      datetime.now().strftime("%Y-%m-%d %H:%M:%S"), date_range, actual_amount, refund_count))

        # 关闭对话框
        self.accept()

        # 刷新界面显示
        if self.parent_window:
            self.parent_window.calculate_weights_from_orders()
            self.parent_window.load_products()
            self.parent_window.update_compare_columns()
            self.parent_window.update_product_avg_price()
            self.parent_window.calculate_total_margin()
            self.parent_window.update_total_orders_label()
            self.parent_window.update_order_range_label()
            main_app = self.parent_window.main_app
            if hasattr(main_app, "refresh_external_products"):
                product_ids = [
                    row[0] for row in self.parent_window.db.safe_fetchall(
                        "SELECT id FROM products WHERE store_id=? AND COALESCE(is_archived, 0)=0",
                        (self.parent_window.store_id,),
                    )
                ]
                main_app.refresh_external_products(product_ids)
            self.parent_window._flush_imported_data_to_archive()
            self.parent_window.main_app.show_toast("✅ 已应用")


class StoreMarginExcelExporter:
    """Headless store-margin Excel exporter used by batch export."""

    MARGIN_HEADERS = [
        "日期", "实发订单", "实发金额", "毛利润", "毛利率", "退款金额", "金额退款率",
        "退款订单", "订单退款率", "件单价", "推广费", "推广占比",
        "技术服务费", "扣款", "其他服务", "其他", "净利润",
        "净利率", "单笔利润", "日盈亏",
    ]
    ORDER_HEADERS = [
        "图片", "商品 ID", "商品标题", "综合成本", "客单价", "毛利", "权重 (%)",
        "权重对比 较上周", "单量", "单量对比 较上周", "销售额", "主卖规格",
        "退款率", "退款占比 最多规格", "链接备注",
    ]

    @staticmethod
    def select_detail_products(parent, db, main_app, stores):
        dialog = QDialog(None)
        dialog.setWindowTitle("选择详细展示链接")
        dialog.setWindowFlags(
            Qt.Window
            | Qt.WindowMinimizeButtonHint
            | Qt.WindowMaximizeButtonHint
            | Qt.WindowCloseButtonHint
        )
        dialog.setWindowModality(Qt.NonModal)
        dialog.setAttribute(Qt.WA_DeleteOnClose, False)
        dialog.resize(920, 680)
        apply_window_icon(dialog, "store")
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(8)

        tip = QLabel("单击左侧商品气泡加入展示箱；按住 Ctrl 连续单击也会逐个加入。导出后“商品规格售卖情况”只展示右侧展示箱里的链接。")
        tip.setWordWrap(True)
        layout.addWidget(tip)

        store_combo = QComboBox()
        for store_id, store_name in stores:
            store_combo.addItem(str(store_name or f"店铺{store_id}"), int(store_id))
        if len(stores) > 1:
            layout.addWidget(store_combo)

        search = QLineEdit()
        search.setPlaceholderText("搜索商品ID或标题")
        sort_combo = QComboBox()
        sort_combo.addItem("按单量", "order")
        sort_combo.addItem("按净利率", "net_margin")
        sort_combo.addItem("按净利润", "net_profit")
        sort_combo.addItem("按商品类型", "category")
        sort_combo.addItem("按毛利率", "gross_margin")
        sort_combo.addItem("按投产", "roi")
        sort_combo.addItem("按投产倍数", "roi_multiple")
        sort_index = sort_combo.findData(getattr(main_app, "product_sort_mode", "order"))
        sort_combo.setCurrentIndex(sort_index if sort_index >= 0 else 0)
        sort_combo.setFixedWidth(140)
        sort_combo.setToolTip("仅调整当前选择窗口的链接顺序，不改变主界面排序")
        search_row = QHBoxLayout()
        search_row.addWidget(search, 1)
        search_row.addWidget(QLabel("排序："))
        search_row.addWidget(sort_combo)
        layout.addLayout(search_row)

        lists = QHBoxLayout()
        left_box = QVBoxLayout()
        right_box = QVBoxLayout()
        left_title = QLabel("当前店铺链接")
        right_title = QLabel("展示箱（点击移除）")
        left_title.setStyleSheet("font-weight: bold; font-size: 14px;")
        right_title.setStyleSheet("font-weight: bold; font-size: 14px;")
        left_box.addWidget(left_title)
        right_box.addWidget(right_title)
        list_widget = QListWidget()
        selected_widget = QListWidget()
        for widget in (list_widget, selected_widget):
            widget.setSelectionMode(QAbstractItemView.SingleSelection)
            widget.setResizeMode(QListWidget.Adjust)
            widget.setUniformItemSizes(False)
            widget.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
            widget.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
            widget.setStyleSheet("""
                QListWidget { border: 1px solid #d1d5db; background: #f8fafc; }
                QListWidget::item { margin: 4px; padding: 0px; border: none; background: transparent; }
                QListWidget::item:selected { background: #dbeafe; color: #111827; }
            """)
        left_box.addWidget(list_widget, 1)
        right_box.addWidget(selected_widget, 1)
        lists.addLayout(left_box, 1)
        lists.addLayout(right_box, 1)
        layout.addLayout(lists, 1)

        selections = {}
        product_cache = {}
        left_items = {}
        bubble_width = 330
        bubble_height = 108
        highlight_timer = QTimer(dialog)
        highlight_timer.setSingleShot(True)
        highlighted_bubble = {"widget": None}

        def products_for_store(store_id):
            sort_mode = sort_combo.currentData() or "order"
            cache_key = (store_id, sort_mode)
            if cache_key not in product_cache:
                rows = db.safe_fetchall(
                    """SELECT id, name, title, image_data, sort_order, product_category_label
                       FROM products
                       WHERE store_id=? AND COALESCE(is_archived, 0)=0""",
                    (store_id,),
                )
                if hasattr(main_app, "_sort_products_for_display"):
                    if hasattr(main_app, "_prepare_product_card_caches"):
                        main_app._prepare_product_card_caches(store_id)
                    rows = main_app._sort_products_for_display(rows, sort_mode)
                else:
                    rows = sorted(rows, key=lambda p: (p[4] if p[4] is not None else p[0], p[0]))
                product_cache[cache_key] = rows
                selections.setdefault(store_id, set())
            return product_cache[cache_key]

        def current_store_id():
            return int(store_combo.currentData())

        def make_bubble(sys_id, product_code, title, image_data):
            bubble = QWidget()
            bubble.setObjectName("ExportProductBubble")
            bubble.setFixedSize(bubble_width, bubble_height)
            bubble.setProperty("normal_style", "#ExportProductBubble { background-color: #545e47; border: none; border-radius: 8px; }")
            bubble.setProperty("highlight_style", "#ExportProductBubble { background-color: #00e5ff; border: 3px solid #111111; border-radius: 8px; }")
            bubble.setStyleSheet(bubble.property("normal_style"))
            outer = QHBoxLayout(bubble)
            outer.setContentsMargins(4, 4, 4, 4)
            outer.setSpacing(6)

            img = QLabel("无图")
            img.setFixedSize(100, 100)
            img.setAlignment(Qt.AlignCenter)
            img.setStyleSheet("background: #eef2f7; color: #6b7280; border-radius: 4px; font-size: 12px;")
            if image_data:
                pixmap = QPixmap()
                if pixmap.loadFromData(image_data):
                    img.setPixmap(pixmap.scaled(100, 100, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation))
            outer.addWidget(img)

            info = QVBoxLayout()
            info.setContentsMargins(0, 0, 0, 0)
            info.setSpacing(2)
            code = QLabel(str(product_code))
            code.setStyleSheet("color: #fffdf5; font-weight: bold; font-size: 12px; background: transparent;")
            name = QLabel(str(title or ""))
            name.setWordWrap(False)
            name.setFixedHeight(18)
            name.setStyleSheet("color: #202020; font-weight: bold; font-size: 11px; background-color: rgba(255,255,255,115); border-radius: 5px; padding: 1px 4px;")
            name.setToolTip(str(title or ""))

            try:
                metrics = main_app.get_product_gross_margin_metrics(sys_id) if hasattr(main_app, "get_product_gross_margin_metrics") else {}
            except Exception:
                metrics = {}
            margin = metrics.get("gross_margin_pct")
            margin_text = f"毛利率:{float(margin):.1f}%" if margin is not None else "毛利率:-"
            try:
                order_count = main_app._get_product_order_count(str(product_code), current_store_id()) if hasattr(main_app, "_get_product_order_count") else 0
            except Exception:
                order_count = 0
            try:
                net_margin = main_app._calculate_product_net_margin(sys_id) if hasattr(main_app, "_calculate_product_net_margin") else None
            except Exception:
                net_margin = None
            net_text = f"净利率:{float(net_margin):.1f}%" if net_margin is not None else "净利率:-"
            rows = db.safe_fetchall("SELECT current_roi, is_natural_flow, is_sitewide_managed FROM products WHERE id=?", (sys_id,))
            roi, natural, sitewide = rows[0] if rows else (0, 0, 0)
            mode = "自然流" if natural else "全站" if sitewide else f"投产:{float(roi or 0):.2f}"
            metrics_label = QLabel(f"{margin_text}  单量:{int(order_count or 0)}单\n{net_text}  {mode}")
            metrics_label.setStyleSheet("color: #fffdf5; font-size: 12px; font-weight: bold; background: transparent;")
            metrics_label.setWordWrap(True)
            info.addWidget(code)
            info.addWidget(name)
            info.addWidget(metrics_label)
            info.addStretch()
            outer.addLayout(info, 1)
            return bubble

        def set_bubble_highlight(bubble, active):
            bubble.setStyleSheet(bubble.property("highlight_style") if active else bubble.property("normal_style"))

        def reload_products():
            store_id = current_store_id()
            highlight_timer.stop()
            clear_search_highlight()
            list_widget.clear()
            left_items.clear()
            for sys_id, product_code, title, image_data, *_rest in products_for_store(store_id):
                item = QListWidgetItem()
                item.setData(Qt.UserRole, str(product_code))
                item.setSizeHint(QSize(bubble_width + 12, bubble_height + 10))
                list_widget.addItem(item)
                bubble = make_bubble(sys_id, product_code, title, image_data)
                list_widget.setItemWidget(item, bubble)
                left_items[str(product_code)] = (item, bubble)
                if str(title or "").strip():
                    left_items.setdefault(str(title or "").strip().casefold(), (item, bubble))
            left_title.setText(f"当前店铺链接（{list_widget.count()}）")
            reload_showcase()

        def clear_search_highlight():
            bubble = highlighted_bubble.get("widget")
            if bubble is not None:
                try:
                    set_bubble_highlight(bubble, False)
                except Exception:
                    pass
            highlighted_bubble["widget"] = None

        def apply_search_highlight():
            query = search.text().strip()
            clear_search_highlight()
            if not query:
                return
            match = left_items.get(query) or left_items.get(query.casefold())
            if not match:
                return
            item, bubble = match
            list_widget.scrollToItem(item, QAbstractItemView.PositionAtCenter)
            list_widget.setCurrentItem(item)
            set_bubble_highlight(bubble, True)
            highlighted_bubble["widget"] = bubble
            highlight_timer.start(2000)

        def reload_showcase():
            store_id = current_store_id()
            selected_widget.clear()
            selected = selections.setdefault(store_id, set())
            products = {str(row[1]): row for row in products_for_store(store_id)}
            for product_code in [str(row[1]) for row in products_for_store(store_id) if str(row[1]) in selected]:
                sys_id, _code, title, image_data, *_rest = products[product_code]
                item = QListWidgetItem()
                item.setData(Qt.UserRole, product_code)
                item.setSizeHint(QSize(bubble_width + 12, bubble_height + 10))
                selected_widget.addItem(item)
                selected_widget.setItemWidget(item, make_bubble(sys_id, product_code, title, image_data))
            right_title.setText(f"展示箱（{selected_widget.count()}，点击移除）")

        def add_to_showcase(item):
            store_id = current_store_id()
            selected = selections.setdefault(store_id, set())
            selected.add(str(item.data(Qt.UserRole)))
            reload_showcase()

        def remove_from_showcase(item):
            selections.setdefault(current_store_id(), set()).discard(str(item.data(Qt.UserRole)))
            reload_showcase()

        actions = QHBoxLayout()
        btn_add_all = QPushButton("全部加入展示箱")
        btn_clear = QPushButton("清空展示箱")
        actions.addWidget(btn_add_all)
        actions.addWidget(btn_clear)
        actions.addStretch()
        layout.addLayout(actions)

        btn_add_all.clicked.connect(lambda: (selections.__setitem__(current_store_id(), {str(row[1]) for row in products_for_store(current_store_id())}), reload_showcase()))
        btn_clear.clicked.connect(lambda: (selections.__setitem__(current_store_id(), set()), reload_showcase()))
        store_combo.currentIndexChanged.connect(reload_products)
        sort_combo.currentIndexChanged.connect(reload_products)
        search.textChanged.connect(apply_search_highlight)
        highlight_timer.timeout.connect(clear_search_highlight)
        list_widget.itemClicked.connect(add_to_showcase)
        selected_widget.itemClicked.connect(remove_from_showcase)
        reload_products()

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.button(QDialogButtonBox.Ok).setText("确认导出")
        buttons.button(QDialogButtonBox.Cancel).setText("取消")
        layout.addWidget(buttons)

        result = {"accepted": False}
        loop = QEventLoop()

        def finish(accepted):
            result["accepted"] = accepted
            dialog.hide()
            loop.quit()

        buttons.accepted.connect(lambda: finish(True))
        buttons.rejected.connect(lambda: finish(False))
        dialog.rejected.connect(lambda: finish(False))
        holders = getattr(main_app, "_store_margin_export_selector_windows", [])
        holders.append(dialog)
        main_app._store_margin_export_selector_windows = holders
        dialog.show()
        dialog.raise_()
        dialog.activateWindow()
        loop.exec_()

        if not result["accepted"]:
            return None
        for store_id, _store_name in stores:
            products_for_store(int(store_id))
        return selections

    def __init__(self, store_id, store_name, main_app, image_quality="clear", detail_product_ids=None):
        self.store_id = store_id
        self.store_name = store_name
        self.main_app = main_app
        self.db = main_app.db
        self.export_image_quality = image_quality or "clear"
        self.export_detail_product_ids = None if detail_product_ids is None else {str(value) for value in detail_product_ids}
        self._excel_export_image_cache = {}
        rows = self.db.safe_fetchall(
            "SELECT id, name FROM products WHERE store_id=? AND COALESCE(is_archived, 0)=0",
            (self.store_id,),
        )
        self.sys_id_to_user_id = {row[0]: row[1] for row in rows}
        try:
            self._normalize_imported_order_store_ids()
        except Exception:
            pass

    def _write_historical_export_sheet(self, wb, sheet_name="过往数据分析", create_sheet=False):
        ws = wb.create_sheet() if create_sheet else wb.active
        ws.title = sheet_name
        ws.append(self.MARGIN_HEADERS)
        row_types = {}
        compare_directions = {}
        records = self.load_manual_data()
        for index, record in enumerate(records):
            ws.append(self._format_manual_record_for_export(record))
            data_row = ws.max_row
            row_types[data_row] = "data"
            ws.row_dimensions[data_row].height = 34
            if index > 0:
                compare_values, directions = self._manual_compare_export_row(record, records[index - 1])
                ws.append(compare_values)
                compare_row = ws.max_row
                row_types[compare_row] = "compare"
                ws.row_dimensions[compare_row].height = 24
                for col_idx, direction in enumerate(directions, start=1):
                    compare_directions[(compare_row, col_idx)] = direction
        self._style_historical_export_sheet(
            ws,
            row_types,
            compare_directions,
            {
                1: 20, 2: 12, 3: 14, 4: 14, 5: 12,
                6: 14, 7: 12, 8: 12, 9: 12, 10: 12,
                11: 12, 12: 12, 13: 14, 14: 12, 15: 14,
                16: 12, 17: 14, 18: 12, 19: 14, 20: 14,
            },
        )
        ws.freeze_panes = "A2"
        self._append_reading_mode_images_to_historical_sheet(ws)

    def _product_avg_price_and_amount_for_export(self, product_code, sys_id):
        imported_totals = self.db.safe_fetchall(
            """SELECT COALESCE(SUM(order_count), 0), COALESCE(SUM(actual_amount), 0)
               FROM imported_orders
               WHERE store_id=? AND product_id=?""",
            (self.store_id, product_code),
        )
        imported_orders = imported_totals[0][0] if imported_totals else 0
        imported_amount = imported_totals[0][1] if imported_totals else 0
        if imported_orders and imported_amount and imported_amount > 0:
            return imported_amount / imported_orders, imported_amount

        spec_sales = self.db.safe_fetchall(
            "SELECT ps.sale_price, io.order_count FROM product_specs ps "
            "LEFT JOIN imported_orders io ON io.product_id=? AND io.spec_code=ps.spec_code "
            "WHERE ps.product_id=? AND COALESCE(ps.is_temporarily_off_shelf, 0)=0",
            (product_code, sys_id),
        )
        total_amount = 0.0
        total_orders = 0
        for sale_price, order_count in spec_sales:
            if sale_price and order_count:
                total_amount += float(sale_price) * int(order_count)
                total_orders += int(order_count)
        if total_orders > 0:
            return total_amount / total_orders, total_amount
        return None, None

    def _order_refund_summary_for_export(self, product_code):
        spec_counts = self.db.safe_fetchall(
            """SELECT spec_code, order_count, refund_count
               FROM imported_orders
               WHERE store_id=? AND product_id=?""",
            (self.store_id, product_code),
        )
        if not spec_counts:
            return "无", "无"
        total_orders = sum(int(row[1] or 0) for row in spec_counts)
        total_refund = sum(int(row[2] or 0) for row in spec_counts)
        if total_orders <= 0 or total_refund <= 0:
            return "0.00%" if total_orders > 0 else "无", "无"

        refund_rate = total_refund / total_orders * 100
        valid_spec_codes = self._get_valid_spec_codes(product_code)
        display_spec_data = [
            item for item in spec_counts
            if not valid_spec_codes or (item[0] and str(item[0]).strip() in valid_spec_codes)
        ]
        max_refund_spec = None
        max_refund_rate = -1
        for spec_code, order_count, refund_count in display_spec_data:
            order_count = int(order_count or 0)
            refund_count = int(refund_count or 0)
            if order_count > 0 and refund_count > 0:
                current_rate = refund_count / order_count
                if current_rate > max_refund_rate:
                    max_refund_rate = current_rate
                    max_refund_spec = spec_code
        if max_refund_spec:
            mode = self._get_spec_display_mode("store_margin_refund_spec_display")
            spec_text = self._get_spec_display_text(product_code, max_refund_spec, mode)
        else:
            spec_text = "无"
        return f"{refund_rate:.2f}%", spec_text

    def _write_orders_export_sheet(self, wb):
        ws = wb.create_sheet("店铺商品权重")
        ws.append(self.ORDER_HEADERS)
        image_refs = []
        image_size = self._export_product_image_size()
        products = self.db.safe_fetchall(
            """SELECT id, name, title, image_data, sort_order, product_category_label,
                      store_weight, store_weight_locked, product_memo
               FROM products WHERE store_id=? AND COALESCE(is_archived, 0)=0
                 AND COALESCE(is_violation, 0)=0""",
            (self.store_id,),
        )
        if hasattr(self.main_app, "_sort_products_for_display"):
            extra_by_id = {product[0]: product[6:] for product in products}
            sorted_products = self.main_app._sort_products_for_display([
                (product[0], product[1], product[2], product[3], product[4], product[5])
                for product in products
            ])
            products = [
                (*product, *extra_by_id.get(product[0], (0, 0)))
                for product in sorted_products
            ]
        else:
            products = sorted(products, key=lambda p: (p[4] if p[4] is not None else p[0], p[0]))

        for product in products:
            sys_id, product_code, title, image_data = product[0], product[1], product[2], product[3]
            store_weight = product[6] if len(product) > 6 else 0
            cost, price, margin = self.get_product_margin(sys_id)
            avg_price, sales_amount = self._product_avg_price_and_amount_for_export(product_code, sys_id)
            if avg_price is not None:
                price = avg_price
            order_rows = self.db.safe_fetchall(
                """SELECT COALESCE(SUM(order_count), 0)
                   FROM imported_orders WHERE store_id=? AND product_id=?""",
                (self.store_id, product_code),
            )
            total_orders = int(order_rows[0][0] or 0) if order_rows else 0
            main_spec_code, spec_orders = self.get_main_spec(product_code)
            if spec_orders > 0 and main_spec_code:
                mode = self._get_spec_display_mode("store_margin_main_spec_display")
                main_spec = self._get_spec_display_text(product_code, main_spec_code, mode)
            else:
                main_spec = "无"
            refund_rate, refund_spec = self._order_refund_summary_for_export(product_code)
            ws.append([
                "",
                str(product_code or ""),
                str(title or ""),
                f"¥{cost:.2f}" if cost else "¥0.00",
                f"¥{price:.2f}" if price else "-",
                f"{margin:.2f}%" if margin else "0.00%",
                f"{float(store_weight or 0):.2f}%",
                "-",
                f"{total_orders}单",
                "-",
                f"¥{float(sales_amount or 0):.2f}" if sales_amount else "-",
                main_spec,
                refund_rate,
                refund_spec,
                str(product[8] or "") if len(product) > 8 else "",
            ])
            excel_row = ws.max_row
            self._set_square_image_cell(ws, excel_row, 1, image_size)
            self._add_export_image(ws, image_data, f"A{excel_row}", image_size, image_refs, "商品")
        ws._image_stream_refs = image_refs
        self._style_excel_sheet(ws, {
            1: self._excel_column_width_for_pixels(image_size),
            2: 14, 3: 24, 4: 11, 5: 11, 6: 10, 7: 10,
            8: 12, 9: 10, 10: 12, 11: 12, 12: 16, 13: 10, 14: 18, 15: 28,
        })
        ws.freeze_panes = "A2"

    _current_store_product_codes = StoreMarginDialog._current_store_product_codes
    _normalize_imported_order_store_ids = StoreMarginDialog._normalize_imported_order_store_ids
    get_user_id_by_sys_id = StoreMarginDialog.get_user_id_by_sys_id
    get_sys_id_by_user_id = StoreMarginDialog.get_sys_id_by_user_id
    _get_spec_display_mode = StoreMarginDialog._get_spec_display_mode
    _get_spec_display_text = StoreMarginDialog._get_spec_display_text
    _get_valid_spec_codes = StoreMarginDialog._get_valid_spec_codes
    get_main_spec = StoreMarginDialog.get_main_spec
    get_product_margin = StoreMarginDialog.get_product_margin
    load_manual_data = StoreMarginDialog.load_manual_data
    _format_manual_record_for_export = StoreMarginDialog._format_manual_record_for_export
    _export_period_day_text = StoreMarginDialog._export_period_day_text
    _export_day_text = StoreMarginDialog._export_day_text
    _manual_record_numeric_values_for_export = StoreMarginDialog._manual_record_numeric_values_for_export
    _manual_compare_export_row = StoreMarginDialog._manual_compare_export_row
    _style_excel_sheet = StoreMarginDialog._style_excel_sheet
    _style_historical_export_sheet = StoreMarginDialog._style_historical_export_sheet
    _export_month_text = StoreMarginDialog._export_month_text
    _reading_mode_images_for_export = StoreMarginDialog._reading_mode_images_for_export
    _add_reading_export_image = StoreMarginDialog._add_reading_export_image
    _high_res_export_image_stream = StoreMarginDialog._high_res_export_image_stream
    _thumbnail_export_image_stream = StoreMarginDialog._thumbnail_export_image_stream
    _append_reading_mode_images_to_historical_sheet = StoreMarginDialog._append_reading_mode_images_to_historical_sheet
    _safe_export_float = StoreMarginDialog._safe_export_float
    _fmt_export_money = StoreMarginDialog._fmt_export_money
    _fmt_export_pct = StoreMarginDialog._fmt_export_pct
    _fmt_export_number = StoreMarginDialog._fmt_export_number
    _split_spec_name_for_export = StoreMarginDialog._split_spec_name_for_export
    _promotion_summary_for_export = StoreMarginDialog._promotion_summary_for_export
    _roi_summary_for_export = StoreMarginDialog._roi_summary_for_export
    _previous_week_range_for_export = StoreMarginDialog._previous_week_range_for_export
    _weekday_cn_for_export = StoreMarginDialog._weekday_cn_for_export
    _operation_record_color_for_export = StoreMarginDialog._operation_record_color_for_export
    _style_operation_record_export_row = StoreMarginDialog._style_operation_record_export_row
    _first_number_for_export = StoreMarginDialog._first_number_for_export
    _brief_change_text_for_export = StoreMarginDialog._brief_change_text_for_export
    _record_briefs_for_export = StoreMarginDialog._record_briefs_for_export
    _products_for_specs_export = StoreMarginDialog._products_for_specs_export
    _product_order_map_for_export = StoreMarginDialog._product_order_map_for_export
    _product_export_context = StoreMarginDialog._product_export_context
    _set_range_fill = StoreMarginDialog._set_range_fill
    _style_product_specs_block = StoreMarginDialog._style_product_specs_block
    _style_product_specs_export_range = StoreMarginDialog._style_product_specs_export_range
    _excel_column_width_for_pixels = StoreMarginDialog._excel_column_width_for_pixels
    _excel_row_height_for_pixels = StoreMarginDialog._excel_row_height_for_pixels
    _set_square_image_cell = StoreMarginDialog._set_square_image_cell
    _export_image_quality_config = StoreMarginDialog._export_image_quality_config
    _export_product_image_size = StoreMarginDialog._export_product_image_size
    _export_image_embed_size = StoreMarginDialog._export_image_embed_size
    _square_image_stream = StoreMarginDialog._square_image_stream
    _add_export_image = StoreMarginDialog._add_export_image
    _operation_record_export_cells = StoreMarginDialog._operation_record_export_cells
    _write_product_specs_export_sheet = StoreMarginDialog._write_product_specs_export_sheet
    export_margin_excel_to_path = StoreMarginDialog.export_margin_excel_to_path
