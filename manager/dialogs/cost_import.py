# -*- coding: utf-8 -*-
"""成本表导入配置对话框"""
import os
import shutil
import tempfile
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    import pandas as pd  # type: ignore

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (
    QCheckBox,
    QComboBox,
    QDialog,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QVBoxLayout,
    QWidget,
)


def _copy_to_temp_file(file_path):
    suffix = os.path.splitext(file_path)[1]
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    temp_path = temp_file.name
    temp_file.close()
    try:
        shutil.copy2(file_path, temp_path)
    except Exception:
        try:
            os.remove(temp_path)
        except OSError:
            pass
        raise
    return temp_path


def _read_cost_file_direct(file_path, **kwargs):
    import pandas as pd  # type: ignore

    if file_path.lower().endswith(".csv"):
        try:
            return pd.read_csv(file_path, encoding="utf-8-sig", **kwargs)
        except UnicodeDecodeError:
            return pd.read_csv(file_path, encoding="gbk", **kwargs)
    return pd.read_excel(file_path, engine="openpyxl", **kwargs)


def read_cost_file(file_path, **kwargs):
    """Read a cost file, falling back to a temporary copy if the source is open."""
    try:
        return _read_cost_file_direct(file_path, **kwargs)
    except PermissionError as original_error:
        temp_path = None
        try:
            temp_path = _copy_to_temp_file(file_path)
            return _read_cost_file_direct(temp_path, **kwargs)
        except Exception as copy_error:
            raise PermissionError(
                "当前文件被独占锁定，无法读取已保存内容。请保存后另存副本，或关闭 Excel/WPS 后再导入。"
            ) from copy_error or original_error
        finally:
            if temp_path:
                try:
                    os.remove(temp_path)
                except OSError:
                    pass


def _extract_cell_fill_color(cell):
    fill = getattr(cell, "fill", None)
    if not fill or not fill.fill_type:
        return ""
    color = fill.fgColor
    if not color or color.type != "rgb" or not color.rgb:
        return ""
    rgb = str(color.rgb).upper()
    if len(rgb) == 8:
        rgb = rgb[2:]
    if len(rgb) != 6 or rgb in ("000000", "FFFFFF"):
        return ""
    return f"#{rgb}"


def _read_cost_row_colors_direct(file_path, name_col_idx=None, spec_col_idx=None):
    if not file_path.lower().endswith((".xlsx", ".xlsm")):
        return {}
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=False, data_only=True)
    try:
        sheet = workbook.active
        colors = {}
        priority_cols = []
        if name_col_idx is not None:
            priority_cols.append(int(name_col_idx) + 1)
        if spec_col_idx is not None:
            priority_cols.append(int(spec_col_idx) + 1)

        for data_idx, excel_row in enumerate(range(2, sheet.max_row + 1)):
            selected_color = ""
            for col in priority_cols:
                selected_color = _extract_cell_fill_color(sheet.cell(excel_row, col))
                if selected_color:
                    break
            if not selected_color:
                for cell in sheet[excel_row]:
                    if cell.value is None:
                        continue
                    selected_color = _extract_cell_fill_color(cell)
                    if selected_color:
                        break
            if selected_color:
                colors[data_idx] = selected_color
        return colors
    finally:
        workbook.close()


def read_cost_row_colors(file_path, name_col_idx=None, spec_col_idx=None):
    """Read row background colors from a saved Excel file. CSV and unsupported colors return empty."""
    try:
        return _read_cost_row_colors_direct(file_path, name_col_idx, spec_col_idx)
    except PermissionError as original_error:
        temp_path = None
        try:
            temp_path = _copy_to_temp_file(file_path)
            return _read_cost_row_colors_direct(temp_path, name_col_idx, spec_col_idx)
        except Exception:
            return {}
        finally:
            if temp_path:
                try:
                    os.remove(temp_path)
                except OSError:
                    pass
    except Exception:
        return {}


class CostImportDialog(QDialog):
    """成本表导入配置对话框。"""

    def __init__(self, file_path, parent=None, cost_mode="total"):
        super().__init__(parent)
        self.file_path = file_path
        self.column_names = []
        self.cost_mode = "detail" if str(cost_mode).lower() == "detail" else "total"

        self.setWindowTitle("导入成本表 - 选择列")
        self.resize(620, 650 if self.cost_mode == "detail" else 595)
        self.init_ui()
        self.load_columns()

    def init_ui(self):
        layout = QVBoxLayout(self)
        self._debug_cid_label = QLabel("【板块:成本导入对话框\n文件:cost_import.py】商品类型/商品名称/规格编码/数量/成本价/重量/导入执行")
        self._debug_cid_label.setStyleSheet("background-color: #DDA0DD; color: #000; font-weight: bold; padding: 1px;")
        self._debug_cid_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        layout.addWidget(self._debug_cid_label)

        mode_text = "详细成本模式：请选择产品成本和重量列。" if self.cost_mode == "detail" else "总成本模式：请选择总成本列。"
        lbl_info = QLabel(f"文件已加载：{os.path.basename(self.file_path)}\n{mode_text}\n商品类型、商品名称和数量可不导入。")
        lbl_info.setStyleSheet("font-weight: bold; padding: 10px;")
        layout.addWidget(lbl_info)

        layout_name = QHBoxLayout()
        layout_name.addWidget(QLabel("【商品名称】"))
        self.combo_name = QComboBox()
        self.combo_name.setMinimumWidth(320)
        layout_name.addWidget(self.combo_name)
        layout.addLayout(layout_name)

        layout_category = QHBoxLayout()
        layout_category.addWidget(QLabel("【商品类型】"))
        self.combo_category = QComboBox()
        self.combo_category.setMinimumWidth(320)
        layout_category.addWidget(self.combo_category)
        layout.addLayout(layout_category)

        layout_spec = QHBoxLayout()
        layout_spec.addWidget(QLabel("【规格编码/SKU】"))
        self.combo_spec = QComboBox()
        self.combo_spec.setMinimumWidth(320)
        layout_spec.addWidget(self.combo_spec)
        layout.addLayout(layout_spec)

        layout_quantity = QHBoxLayout()
        layout_quantity.addWidget(QLabel("【数量】"))
        self.combo_quantity = QComboBox()
        self.combo_quantity.setMinimumWidth(320)
        layout_quantity.addWidget(self.combo_quantity)
        layout.addLayout(layout_quantity)

        layout_price = QHBoxLayout()
        self.lbl_price = QLabel("【产品成本】" if self.cost_mode == "detail" else "【总成本】")
        layout_price.addWidget(self.lbl_price)
        self.combo_price = QComboBox()
        self.combo_price.setMinimumWidth(320)
        layout_price.addWidget(self.combo_price)
        layout.addLayout(layout_price)

        layout_weight = QHBoxLayout()
        self.lbl_weight = QLabel("【单个重量kg】")
        self.combo_weight = QComboBox()
        self.combo_weight.setMinimumWidth(320)
        layout_weight.addWidget(self.lbl_weight)
        layout_weight.addWidget(self.combo_weight)
        layout.addLayout(layout_weight)
        self.lbl_weight.setVisible(self.cost_mode == "detail")
        self.combo_weight.setVisible(self.cost_mode == "detail")

        self.chk_unit_by_quantity = QCheckBox("导入的成本/重量是整组数量，按数量折算为单个成本和单个重量")
        self.chk_unit_by_quantity.setChecked(True)
        self.chk_unit_by_quantity.setVisible(self.cost_mode == "detail")
        self.chk_unit_by_quantity.setStyleSheet("color: #555; padding: 4px;")
        layout.addWidget(self.chk_unit_by_quantity)

        self.chk_import_attribute = QCheckBox("导入产品属性")
        self.chk_import_attribute.toggled.connect(self._toggle_attribute_mapping)
        layout.addWidget(self.chk_import_attribute)

        self.attribute_widget = QWidget()
        attribute_layout = QVBoxLayout(self.attribute_widget)
        attribute_layout.setContentsMargins(18, 0, 0, 0)
        self.combo_attr_size = self._make_attribute_combo_row(attribute_layout, "尺寸")
        self.combo_attr_pages = self._make_attribute_combo_row(attribute_layout, "张数")
        self.combo_attr_print = self._make_attribute_combo_row(attribute_layout, "印刷工艺")
        layout.addWidget(self.attribute_widget)
        self._toggle_attribute_mapping(False)

        layout.addStretch()
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        self.btn_confirm = QPushButton("确认导入")
        self.btn_confirm.setStyleSheet("background-color: #28a745; color: white; font-weight: bold; padding: 10px 30px;")
        self.btn_confirm.clicked.connect(self.accept)
        self.btn_cancel = QPushButton("取消")
        self.btn_cancel.clicked.connect(self.reject)
        btn_layout.addWidget(self.btn_confirm)
        btn_layout.addWidget(self.btn_cancel)
        layout.addLayout(btn_layout)

    def _make_attribute_combo_row(self, parent_layout, label):
        row = QHBoxLayout()
        row.addWidget(QLabel(f"【{label}】"))
        combo = QComboBox()
        combo.setMinimumWidth(320)
        row.addWidget(combo)
        parent_layout.addLayout(row)
        return combo

    def _toggle_attribute_mapping(self, checked):
        self.attribute_widget.setVisible(bool(checked))

    def _read_columns(self):
        return read_cost_file(self.file_path, nrows=0, header=0).columns.tolist()

    def load_columns(self):
        try:
            import pandas as pd  # type: ignore

            if not os.path.exists(self.file_path):
                raise Exception("文件路径不存在")

            raw_columns = self._read_columns()
            valid_columns = [str(c).strip() for c in raw_columns if str(c).strip() != "" and str(c).lower() != "nan"]
            if len(valid_columns) == 0:
                df_backup = read_cost_file(self.file_path, nrows=1, header=None)
                backup_row = df_backup.iloc[0].tolist()
                raise Exception(f"读取到的第一行为空或无效。实际读取内容：{backup_row}")

            self.column_names = raw_columns
            self.combo_name.clear()
            self.combo_spec.clear()
            self.combo_quantity.clear()
            self.combo_price.clear()
            self.combo_category.clear()
            self.combo_weight.clear()
            for combo in (self.combo_attr_size, self.combo_attr_pages, self.combo_attr_print):
                combo.clear()
                combo.addItem("（不导入）", None)
            self.combo_name.addItem("（不导入）", None)
            self.combo_quantity.addItem("（不导入）", None)
            self.combo_category.addItem("（不导入）", None)
            if self.cost_mode == "detail":
                self.combo_price.addItem("（请选择产品成本列）", None)
                self.combo_weight.addItem("（请选择重量列）", None)

            for idx, name in enumerate(self.column_names):
                name_str = str(name).strip()
                if not name_str or name_str.lower() == "nan":
                    continue
                col_letter = ""
                num = idx + 1
                while num > 0:
                    num, rem = divmod(num - 1, 26)
                    col_letter = chr(65 + rem) + col_letter
                display_text = f"{name_str} ({col_letter}列)"
                self.combo_name.addItem(display_text, idx)
                self.combo_spec.addItem(display_text, idx)
                self.combo_quantity.addItem(display_text, idx)
                self.combo_price.addItem(display_text, idx)
                self.combo_category.addItem(display_text, idx)
                self.combo_weight.addItem(display_text, idx)
                for combo in (self.combo_attr_size, self.combo_attr_pages, self.combo_attr_print):
                    combo.addItem(display_text, idx)

            if self.combo_spec.count() == 0:
                raise Exception("所有列名均为空")
            self.auto_match_columns()
        except PermissionError as e:
            QMessageBox.critical(
                self,
                "文件无法读取",
                f"{str(e)}\n\n"
                "说明：软件只能读取表格已经保存到磁盘的内容，Excel/WPS 中尚未保存的修改不会被导入。",
            )
            self.reject()
        except Exception as e:
            import traceback

            print(traceback.format_exc())
            QMessageBox.critical(
                self,
                "读取失败",
                f"无法读取列名：\n\n{str(e)}\n\n"
                "排查建议：\n1. 检查第一行是否有表头。\n"
                "2. 如果文件正在打开，请先保存当前修改。\n"
                "3. 如果仍失败，尝试另存为新的 .xlsx 再导入。",
            )
            self.reject()

    def auto_match_columns(self):
        name_keywords = ["商品名称", "商品名", "产品名称", "产品名", "名称", "品名", "标题", "name", "title"]
        category_keywords = ["商品类型", "产品类型", "类型", "分类", "类别", "品类", "类目", "category", "type"]
        spec_keywords = ["规格", "编码", "SKU", "Code", "ID", "型号", "商品编号", "SPU", "No"]
        quantity_keywords = ["数量", "件数", "个数", "库存数量", "qty", "quantity", "count", "num"]
        weight_keywords = ["重量", "单重", "单个重量", "净重", "weight"]
        attribute_keywords = {
            "size": ["\u5c3a\u5bf8", "\u89c4\u683c\u5c3a\u5bf8", "\u5927\u5c0f", "\u5c3a\u7801", "size"],
            "pages": ["\u5f20\u6570", "\u9875\u6570", "\u6570\u91cf\u5f20", "pages", "sheets"],
            "print": ["\u5370\u5237\u5de5\u827a", "\u5370\u5237", "\u5de5\u827a", "printing", "print"],
        }
        preferred_price_keywords = ["产品成本", "单品成本", "单个成本", "产品单价"] if self.cost_mode == "detail" else ["总成本"]
        price_keywords = ["产品成本", "单品成本", "单个成本", "产品单价", "进价", "成本", "价格", "Price", "Cost", "单价", "Money"] if self.cost_mode == "detail" else ["成本", "价格", "Price", "Cost", "单价", "进价", "Money"]
        name_found = category_found = spec_found = quantity_found = price_found = weight_found = False
        attr_found = {"size": False, "pages": False, "print": False}

        for idx, name in enumerate(self.column_names):
            name_str = str(name).strip()
            if not name_str or name_str.lower() == "nan":
                continue
            name_lower = name_str.lower()
            for key, combo in (
                ("size", self.combo_attr_size),
                ("pages", self.combo_attr_pages),
                ("print", self.combo_attr_print),
            ):
                if not attr_found[key] and any(k.lower() in name_lower for k in attribute_keywords[key]):
                    i = combo.findData(idx)
                    if i >= 0:
                        combo.setCurrentIndex(i)
                        self.chk_import_attribute.setChecked(True)
                    attr_found[key] = True
            if not name_found and any(k.lower() in name_lower for k in name_keywords):
                i = self.combo_name.findData(idx)
                if i >= 0:
                    self.combo_name.setCurrentIndex(i)
                name_found = True
            if not category_found and any(k.lower() in name_lower for k in category_keywords):
                i = self.combo_category.findData(idx)
                if i >= 0:
                    self.combo_category.setCurrentIndex(i)
                category_found = True
            if not spec_found and any(k.lower() in name_lower for k in spec_keywords):
                i = self.combo_spec.findData(idx)
                if i >= 0:
                    self.combo_spec.setCurrentIndex(i)
                spec_found = True
            if not quantity_found and any(k.lower() in name_lower for k in quantity_keywords):
                i = self.combo_quantity.findData(idx)
                if i >= 0:
                    self.combo_quantity.setCurrentIndex(i)
                quantity_found = True
            if not price_found and any(k.lower() in name_lower for k in preferred_price_keywords):
                i = self.combo_price.findData(idx)
                if i >= 0:
                    self.combo_price.setCurrentIndex(i)
                price_found = True
            if self.cost_mode == "detail" and not weight_found and any(k.lower() in name_lower for k in weight_keywords):
                i = self.combo_weight.findData(idx)
                if i >= 0:
                    self.combo_weight.setCurrentIndex(i)
                weight_found = True
            if (
                name_found and category_found and spec_found and quantity_found and price_found
                and (self.cost_mode != "detail" or weight_found)
                and all(attr_found.values())
            ):
                break

        if not price_found:
            for idx, name in enumerate(self.column_names):
                name_str = str(name).strip()
                if not name_str or name_str.lower() == "nan":
                    continue
                name_lower = name_str.lower()
                if self.cost_mode == "detail" and "总成本" in name_str:
                    continue
                if any(k.lower() in name_lower for k in price_keywords):
                    i = self.combo_price.findData(idx)
                    if i >= 0:
                        self.combo_price.setCurrentIndex(i)
                    break

    def get_mapping(self):
        spec_idx = self.combo_spec.currentData()
        price_idx = self.combo_price.currentData()
        name_idx = self.combo_name.currentData()
        quantity_idx = self.combo_quantity.currentData()
        category_idx = self.combo_category.currentData()
        weight_idx = self.combo_weight.currentData() if self.cost_mode == "detail" else None
        attr_indices = None
        if self.chk_import_attribute.isChecked():
            attr_indices = {
                "size": self.combo_attr_size.currentData(),
                "pages": self.combo_attr_pages.currentData(),
                "print": self.combo_attr_print.currentData(),
            }
        if spec_idx is None or price_idx is None:
            return None, None, None, None, None, None, None
        if self.cost_mode == "detail" and weight_idx is None:
            return int(spec_idx), int(price_idx), (int(name_idx) if name_idx is not None else None), (int(quantity_idx) if quantity_idx is not None else None), (int(category_idx) if category_idx is not None else None), None, attr_indices
        return (
            int(spec_idx),
            int(price_idx),
            (int(name_idx) if name_idx is not None else None),
            (int(quantity_idx) if quantity_idx is not None else None),
            (int(category_idx) if category_idx is not None else None),
            (int(weight_idx) if weight_idx is not None else None),
            attr_indices,
        )

    def should_unit_by_quantity(self):
        return self.cost_mode == "detail" and self.chk_unit_by_quantity.isChecked()
