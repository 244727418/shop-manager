# ================= 版本信息 =================
VERSION = "5.19"
CURRENT_RELEASE_NOTES = (
    "1. 成本库新增局域网实时同步，可创建或加入成本组织，并自动同步规格、图片、商品类型和操作历史；修复组合产品派生成本导致旧值回写的问题。\n"
    "2. 素材库新增手机扫码上传，支持按商品类型和规格上传原图，并优化素材去重、引用及成本缩略图联动。\n"
    "3. 优化平台规格与活动数据处理，完善添加编码页面打开、未匹配项定位、营销活动同步和抓取操作记录。\n"
    "4. 优化订单退款率、主卖规格、销售额与链接盈亏展示，并改进商品卡片指标说明及毛利 Excel 导出。"
)

# ================= 系统标准库 =================
import sys
import os
import json
import calendar
import traceback
import re
import subprocess
import ctypes
import time
import hashlib
from ctypes import wintypes
import shutil
import socket
import threading
import tempfile
from copy import copy
from datetime import datetime, timedelta
from importlib import import_module

try:
    from manager.crash_report import append_event, append_exception, install_crash_reporting
except ImportError:
    from crash_report import append_event, append_exception, install_crash_reporting

if "--update-agent" not in sys.argv:
    install_crash_reporting(VERSION)

# Windows下隐藏控制台窗口的常量（防止黑框闪烁）
if sys.platform == 'win32':
    CREATE_NO_WINDOW = 0x08000000
else:
    CREATE_NO_WINDOW = 0

# ================= 第三方库 =================
_PANDAS_AVAILABLE = None
_OPENPYXL_AVAILABLE = None


def has_pandas():
    global _PANDAS_AVAILABLE
    if _PANDAS_AVAILABLE is None:
        try:
            import pandas  # type: ignore  # noqa: F401
            _PANDAS_AVAILABLE = True
        except ImportError as e:
            print(f"警告: 缺少 pandas - {e}")
            _PANDAS_AVAILABLE = False
    return _PANDAS_AVAILABLE


def has_openpyxl():
    global _OPENPYXL_AVAILABLE
    if _OPENPYXL_AVAILABLE is None:
        try:
            import openpyxl  # type: ignore  # noqa: F401
            _OPENPYXL_AVAILABLE = True
        except ImportError as e:
            print(f"警告: 缺少 openpyxl - {e}")
            _OPENPYXL_AVAILABLE = False
    return _OPENPYXL_AVAILABLE

# ================= PyQt5 核心库 =================
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QPushButton, QLabel, QTableWidget, QTableWidgetItem, QHeaderView,
    QInputDialog, QFileDialog, QMessageBox, QScrollArea, QTimeEdit,
    QTextEdit, QTextBrowser, QAbstractItemView, QFrame, QDialog, QComboBox,   
    QSpinBox, QTableView, QStyle, QStyledItemDelegate, QLineEdit,
    QCalendarWidget, QDateEdit, QStatusBar, QProgressBar, QProgressDialog, QSplitter,
    QGroupBox, QRadioButton, QCheckBox, QListWidget, QListWidgetItem,
    QTreeWidget, QTreeWidgetItem, QMenu, QAction, QToolBar, QSystemTrayIcon,
    QSizePolicy, QShortcut, QStyleOptionViewItem,
    QKeySequenceEdit, QDialogButtonBox,
    QToolTip, QLayout, QStackedWidget
)

from PyQt5.QtCore import (
    Qt, QTimer, QEvent, QRect, QMimeData, QThread, pyqtSignal,
    QModelIndex, QSize, QPoint, QUrl, QSettings, QTranslator, QLocale,
    QAbstractTableModel, QSortFilterProxyModel, QTime, QDate
)

from PyQt5.QtGui import (
    QPixmap, QColor, QIcon, QFont, QDrag, QStandardItemModel, QStandardItem,
    QFontMetrics, QDoubleValidator, QIntValidator, QRegExpValidator,
    QPainter, QPen, QBrush, QCursor, QKeySequence, QPalette, QImage, QFontDatabase
)
try:
    from PyQt5 import sip
except ImportError:
    import sip


def _qobject_alive(obj):
    if obj is None:
        return False
    try:
        return not sip.isdeleted(obj)
    except TypeError:
        return True


from PyQt5.QtSvg import QSvgRenderer
from PyQt5.QtNetwork import QLocalServer, QLocalSocket

import sqlite3

try:
    from manager.db import SafeDatabaseManager
except ImportError:
    from db import SafeDatabaseManager

try:
    from manager.data_root import DataRootManager
except ImportError:
    from data_root import DataRootManager

try:
    from manager.pinyin_search import all_terms_match, split_search_terms, text_matches
except ImportError:
    from pinyin_search import all_terms_match, split_search_terms, text_matches

try:
    from manager.pdd_browser_monitor import PddBrowserMonitor
except ImportError:
    from pdd_browser_monitor import PddBrowserMonitor

try:
    from manager.widgets import ProductWidget, StoreWidget, RecordRow, InPlaceEditor
except ImportError:
    from widgets import ProductWidget, StoreWidget, RecordRow, InPlaceEditor

def _import_local_attr(module_name, attr_name):
    try:
        module = import_module(f"manager.{module_name}")
    except ImportError:
        module = import_module(module_name)
    return getattr(module, attr_name)


class _LazyAttr:
    def __init__(self, module_name, attr_name):
        self.module_name = module_name
        self.attr_name = attr_name
        self._value = None

    def _load(self):
        if self._value is None:
            self._value = _import_local_attr(self.module_name, self.attr_name)
        return self._value

    def __call__(self, *args, **kwargs):
        return self._load()(*args, **kwargs)

    def __getattr__(self, name):
        return getattr(self._load(), name)

    def __repr__(self):
        return repr(self._load())

OperationRecordDialog = _LazyAttr("dialogs.records", "OperationRecordDialog")
DailyRecordDialog = _LazyAttr("dialogs.records", "DailyRecordDialog")
StoreMarginDialog = _LazyAttr("dialogs.store_margin", "StoreMarginDialog")
StoreMarginExcelExporter = _LazyAttr("dialogs.store_margin", "StoreMarginExcelExporter")
CostImportDialog = _LazyAttr("dialogs.cost_import", "CostImportDialog")
CostLibraryDialog = _LazyAttr("dialogs.cost_library", "CostLibraryDialog")
MaterialLibraryDialog = _LazyAttr("dialogs.material_library", "MaterialLibraryDialog")
MaterialMobileService = _LazyAttr("material_mobile_service", "MaterialMobileService")
CostSyncService = _LazyAttr("cost_sync_service", "CostSyncService")
ApiConfigDialog = _LazyAttr("dialogs.api_config", "ApiConfigDialog")
ProfitAnalysisDialog = _LazyAttr("dialogs.profit", "ProfitAnalysisDialog")
ProfitCalculatorDialog = _LazyAttr("dialogs.profit", "ProfitCalculatorDialog")
ProfitHistoryDialog = _LazyAttr("dialogs.profit", "ProfitHistoryDialog")
DailyTaskDialog = _LazyAttr("dialogs.daily_task", "DailyTaskDialog")
TaskReminderPopupDialog = _LazyAttr("dialogs.daily_task", "TaskReminderPopupDialog")
ProductSpecDialog = _LazyAttr("dialogs.product_spec", "ProductSpecDialog")


def read_cost_file(*args, **kwargs):
    return _import_local_attr("dialogs.cost_import", "read_cost_file")(*args, **kwargs)


def read_cost_row_colors(*args, **kwargs):
    return _import_local_attr("dialogs.cost_import", "read_cost_row_colors")(*args, **kwargs)

try:
    from manager.delegates import SpecNameDelegate, CenterAlignDelegate, WeightDelegate
except ImportError:
    from delegates import SpecNameDelegate, CenterAlignDelegate, WeightDelegate

try:
    from manager.prompts import get_default_prompt
except ImportError:
    from prompts import get_default_prompt

try:
    from manager.ui_utils import convert_markdown_to_html
except ImportError:
    from ui_utils import convert_markdown_to_html

try:
    from manager.window_icons import apply_window_icon, get_window_icon, icon_path, install_window_icon_filter
except ImportError:
    from window_icons import apply_window_icon, get_window_icon, icon_path, install_window_icon_filter

try:
    from manager.file_dialog_memory import remembered_existing_directory, remembered_open_file, remembered_save_file
except ImportError:
    from file_dialog_memory import remembered_existing_directory, remembered_open_file, remembered_save_file

try:
    from manager.update_manager import (
        UpdateAdminDialog,
        UpdateBroadcastListener,
        UpdateDownloadWorker,
        UpdatePublishService,
        app_dir,
        bind_trusted_update_source,
        fetch_manifest,
        find_newer_local_package,
        install_update_agent_task,
        is_newer_version,
        is_trusted_update_manifest,
        load_pending_update,
        load_global_update_settings,
        notify_update_agent,
        normalize_server_url,
        run_update_agent,
        save_global_update_setting,
        start_update_agent_task,
        uninstall_update_agent_task,
    )
except ImportError:
    from update_manager import (
        UpdateAdminDialog,
        UpdateBroadcastListener,
        UpdateDownloadWorker,
        UpdatePublishService,
        app_dir,
        bind_trusted_update_source,
        fetch_manifest,
        find_newer_local_package,
        install_update_agent_task,
        is_newer_version,
        is_trusted_update_manifest,
        load_pending_update,
        load_global_update_settings,
        notify_update_agent,
        normalize_server_url,
        run_update_agent,
        save_global_update_setting,
        start_update_agent_task,
        uninstall_update_agent_task,
    )


SINGLE_INSTANCE_KEY = "shop_manager_v3_7_single_instance"
SINGLE_INSTANCE_MUTEX_NAME = "shop_manager_v3_7_single_instance_mutex"
MAIN_WINDOW_TITLE = f"电商店铺操作记录管理工具 v{VERSION}"
DEFAULT_GLOBAL_HOTKEYS = {
    "quick_hotkey_main": "Ctrl+Shift+Z",
    "quick_hotkey_cost_library": "Ctrl+Shift+C",
    "quick_hotkey_material_library": "Ctrl+Shift+S",
}
AUTO_START_REGISTRY_PATH = r"Software\Microsoft\Windows\CurrentVersion\Run"
AUTO_START_VALUE_NAME = "ShopManager"
DEFAULT_UPDATE_ADMIN_PASSWORD_HASH = hashlib.sha256("244".encode("utf-8")).hexdigest()
ERROR_ALREADY_EXISTS = 183
SW_SHOW = 5
SW_RESTORE = 9

_USER32 = None
if sys.platform == "win32":
    _USER32 = ctypes.WinDLL("user32", use_last_error=True)
    _USER32.RegisterHotKey.argtypes = (wintypes.HWND, ctypes.c_int, wintypes.UINT, wintypes.UINT)
    _USER32.RegisterHotKey.restype = wintypes.BOOL
    _USER32.UnregisterHotKey.argtypes = (wintypes.HWND, ctypes.c_int)
    _USER32.UnregisterHotKey.restype = wintypes.BOOL


def startup_release_notes(settings, current_version):
    manifest = (settings or {}).get("last_update_manifest")
    if not isinstance(manifest, dict):
        return ""
    version = str(current_version or "").strip()
    if str(manifest.get("version") or "").strip() != version:
        return ""
    if str((settings or {}).get("release_notes_seen_version") or "").strip() == version:
        return ""
    return str(manifest.get("notes") or "").strip()


def verify_update_admin_password(password):
    settings = load_global_update_settings()
    password_hash = settings.get("update_admin_password_hash") or DEFAULT_UPDATE_ADMIN_PASSWORD_HASH
    return hashlib.sha256(str(password or "").encode("utf-8")).hexdigest() == password_hash


def auto_start_command():
    frozen = bool(getattr(sys, "frozen", False))
    executable = sys.executable
    if not frozen and sys.platform == "win32":
        pythonw = os.path.join(os.path.dirname(executable), "pythonw.exe")
        if os.path.isfile(pythonw):
            executable = pythonw
    args = [executable]
    if not frozen:
        args.append(os.path.abspath(sys.argv[0]))
    args.append("--autostart")
    return subprocess.list2cmdline(args)


def set_auto_start(enabled):
    if sys.platform != "win32":
        return False
    try:
        import winreg
        with winreg.CreateKeyEx(
            winreg.HKEY_CURRENT_USER,
            AUTO_START_REGISTRY_PATH,
            0,
            winreg.KEY_READ | winreg.KEY_SET_VALUE,
        ) as key:
            if enabled:
                command = auto_start_command()
                try:
                    current, _ = winreg.QueryValueEx(key, AUTO_START_VALUE_NAME)
                except FileNotFoundError:
                    current = ""
                if current != command:
                    winreg.SetValueEx(key, AUTO_START_VALUE_NAME, 0, winreg.REG_SZ, command)
            else:
                try:
                    winreg.DeleteValue(key, AUTO_START_VALUE_NAME)
                except FileNotFoundError:
                    pass
        return True
    except Exception as e:
        print(f"设置开机自启失败: {e}")
        return False


def notify_existing_instance():
    socket = QLocalSocket()
    socket.connectToServer(SINGLE_INSTANCE_KEY)
    if not socket.waitForConnected(300):
        socket.abort()
        return False

    socket.write(b"activate")
    socket.flush()
    socket.waitForBytesWritten(300)
    socket.disconnectFromServer()
    return True


def acquire_single_instance_mutex():
    if sys.platform != 'win32':
        return None, False

    kernel32 = ctypes.windll.kernel32
    handle = kernel32.CreateMutexW(None, False, SINGLE_INSTANCE_MUTEX_NAME)
    if not handle:
        return None, False
    return handle, kernel32.GetLastError() == ERROR_ALREADY_EXISTS


def activate_existing_window_by_title():
    if sys.platform != 'win32':
        return False

    user32 = ctypes.windll.user32
    hwnd = user32.FindWindowW(None, MAIN_WINDOW_TITLE)
    if not hwnd:
        return False

    user32.ShowWindow(hwnd, SW_SHOW)
    user32.ShowWindow(hwnd, SW_RESTORE)
    user32.SetForegroundWindow(hwnd)
    return True


def setup_single_instance(window_holder):
    if notify_existing_instance():
        return None, True

    server = QLocalServer()

    def activate_window():
        window = window_holder.get("window")
        if window is not None:
            window.show_window()

    def handle_new_connection():
        while server.hasPendingConnections():
            client = server.nextPendingConnection()
            client.readAll()
            client.disconnectFromServer()
        QTimer.singleShot(0, activate_window)

    if not server.listen(SINGLE_INSTANCE_KEY):
        if notify_existing_instance():
            return None, True
        QLocalServer.removeServer(SINGLE_INSTANCE_KEY)
        if not server.listen(SINGLE_INSTANCE_KEY):
            if notify_existing_instance():
                return None, True
            return None, False

    server.newConnection.connect(handle_new_connection)
    return server, False



class TodayColumnDelegate(QStyledItemDelegate):
    def __init__(self, parent=None, today_col=-1):
        super().__init__(parent)
        self._today_col = today_col
        self._today_bg = QColor("#dfeeda")
        self._today_text = QColor("#2f5d47")

    def set_today_col(self, col):
        self._today_col = col

    def paint(self, painter, option, index):
        option = QStyleOptionViewItem(option)
        option.state &= ~QStyle.State_HasFocus
        if index.column() == self._today_col and self._today_col >= 0:
            painter.fillRect(option.rect, self._today_bg)
            option.palette.setColor(QPalette.Text, self._today_text)
            option.palette.setColor(QPalette.HighlightedText, self._today_text)
        super().paint(painter, option, index)


class OperationRecordDelegate(TodayColumnDelegate):
    METRIC_STYLES = {
        "涨价": ("#fdecea", "#c0392b"),
        "降价": ("#e8f8ef", "#1e8449"),
        "改价": ("#fff3cd", "#b7791f"),
        "新增规格": ("#e8f8ef", "#1e8449"),
        "删除规格": ("#fdecea", "#c0392b"),
        "提投产": ("#fdecea", "#c0392b"),
        "降投产": ("#e8f8ef", "#1e8449"),
        "改投产": ("#e8f0fe", "#2f5fb3"),
        "改出价": ("#e0f7fa", "#00838f"),
        "改图": ("#e8f5e9", "#2e7d32"),
        "改优惠": ("#fff3cd", "#b7791f"),
        "改活动": ("#f3e8ff", "#7d3c98"),
        "改模式": ("#eceff1", "#455a64"),
        "改标题": ("#fff7e6", "#a35f00"),
        "新建链接": ("#e8f8ef", "#1e8449"),
        "下架": ("#eceff1", "#455a64"),
        "恢复": ("#e8f8ef", "#1e8449"),
        "删除": ("#fdecea", "#c0392b"),
        "同步": ("#e8f0fe", "#2f5fb3"),
        "记录": ("#edf2f7", "#34495e"),
        "规格售价": ("#e8f4ff", "#1f6fb2"),
        "规格新增": ("#e8f8ef", "#1e8449"),
        "规格删除": ("#fdecea", "#c0392b"),
        "规格名称": ("#f3e8ff", "#7d3c98"),
        "优惠券": ("#fff3cd", "#b7791f"),
        "新客立减": ("#fce4ec", "#ad1457"),
        "投产": ("#e8f0fe", "#2f5fb3"),
        "成交出价": ("#e0f7fa", "#00838f"),
        "退货率": ("#fff0e6", "#d35400"),
        "推广模式": ("#eceff1", "#455a64"),
        "投产/出价模式": ("#e8f0fe", "#2f5fb3"),
        "主轮播图": ("#e8f5e9", "#2e7d32"),
        "限时限量购": ("#fdecea", "#c0392b"),
        "营销活动": ("#f3e8ff", "#7d3c98"),
        "综合毛利": ("#e8f8ef", "#1e8449"),
        "商品标题": ("#fff7e6", "#a35f00"),
        "新建链接": ("#e8f8ef", "#1e8449"),
    }

    def __init__(self, parent=None, today_col=-1):
        super().__init__(parent, today_col)
        self._time_bg = QColor("#eef2f7")
        self._time_fg = QColor("#34495e")
        self._spec_bg = QColor("#e9f2ff")
        self._spec_fg = QColor("#245269")
        self._text_fg = QColor("#1f2d3d")
        self._selected_bg = QColor("#dfeeda")
        self._fallback_font = QFont("Microsoft YaHei", 11)
        self._fallback_font.setBold(True)

    def paint(self, painter, option, index):
        records = index.data(Qt.UserRole)
        if not self._has_structured_records(records):
            super().paint(painter, option, index)
            return

        painter.save()
        self._paint_background(painter, option, index)
        painter.setClipRect(option.rect.adjusted(1, 1, -1, -1))

        x = option.rect.left() + 3
        y = option.rect.top() + 3
        right = option.rect.right() - 3
        line_height = 20
        normal_font = QFont("Microsoft YaHei", 10)
        bold_font = QFont("Microsoft YaHei", 10)
        bold_font.setBold(True)

        for part in self._record_parts(records):
            if y + line_height > option.rect.bottom() - 2:
                break
            cx = x
            time_text = part.get("time", "")
            if time_text:
                cx = self._draw_pill_or_wrap(painter, x, cx, y, line_height, time_text, self._time_bg, self._time_fg, bold_font, right)
                if isinstance(cx, tuple):
                    cx, y = cx

            metric = part.get("metric", "记录")
            bg, fg = self._metric_colors(metric)
            cx = self._draw_pill_or_wrap(painter, x, cx, y, line_height, metric, bg, fg, bold_font, right)
            if isinstance(cx, tuple):
                cx, y = cx

            spec_text = part.get("spec", "")
            if spec_text:
                spec_width = QFontMetrics(bold_font).horizontalAdvance(spec_text) + 12
                if cx + spec_width > right:
                    y += line_height
                    if y + line_height > option.rect.bottom() - 2:
                        break
                    cx = x
                y = self._draw_wrapped_block(painter, cx, y + 2, spec_text, self._spec_bg, self._spec_fg, bold_font, right)
                cx = x

            text = part.get("text", "")
            if text:
                start_x = cx if cx > x else x
                if start_x >= right:
                    start_x = x
                    y += line_height
                if y + line_height > option.rect.bottom() - 2:
                    break
                y = self._draw_wrapped_text(painter, start_x, y, text, normal_font, right, option.rect.bottom() - 2)
            else:
                y += line_height

        painter.restore()

    def _paint_background(self, painter, option, index):
        if option.state & QStyle.State_Selected:
            painter.fillRect(option.rect, self._selected_bg)
            return
        bg = index.data(Qt.BackgroundRole)
        if isinstance(bg, QBrush) and bg.style() != Qt.NoBrush:
            painter.fillRect(option.rect, bg)
            return
        if index.column() == self._today_col and self._today_col >= 0:
            painter.fillRect(option.rect, self._today_bg)
            return
        painter.fillRect(option.rect, option.palette.base())

    def _has_structured_records(self, records):
        if not isinstance(records, list):
            return False
        return any(
            isinstance(r, dict) and (r.get("changes") or str(r.get("text", "") or "").strip())
            for r in records
        )

    def _record_parts(self, records):
        parts = []
        for record in sorted(records or [], key=lambda item: str(item.get("time", "")) if isinstance(item, dict) else "", reverse=True):
            if not isinstance(record, dict):
                continue
            time_text = str(record.get("time", "") or "")
            changes = record.get("changes") or []
            if changes:
                labels = []
                for change in changes:
                    if not isinstance(change, dict):
                        continue
                    text = str(change.get("text", "") or "")
                    metric = str(change.get("metric", "记录") or "记录")
                    label = self._summary_label(metric, text, change)
                    if label not in labels:
                        labels.append(label)
                if labels:
                    parts.append({"time": time_text, "metric": "、".join(labels), "spec": "", "text": ""})
            elif record.get("text"):
                text = str(record.get("text", ""))
                parts.append({"time": time_text, "metric": self._summary_label("记录", text, record), "spec": "", "text": ""})
        return parts

    def _summary_label(self, metric, text, change):
        old_value = change.get("old", "") if isinstance(change, dict) else ""
        new_value = change.get("new", "") if isinstance(change, dict) else ""
        combined = f"{metric} {text} {old_value} {new_value}"

        if "新建链接" in combined:
            return "新建链接"
        if "主轮播图" in combined or "图片" in combined or "改图" in combined:
            return "改图"
        if "商品标题" in combined or "标题" in combined:
            return "改标题"
        if "链接下架" in combined or "下架" in combined:
            return "下架"
        if "恢复" in combined:
            return "恢复"
        if "删除链接" in combined:
            return "删除"
        if "规格新增" in combined or "新增规格" in combined:
            return "新增规格"
        if "规格删除" in combined or "删除规格" in combined:
            return "删除规格"
        if "成交出价" in combined or "出价" in combined:
            return "改出价"
        if "投产" in combined or "ROI" in combined or "roi" in combined:
            trend = self._numeric_trend(old_value, new_value)
            if trend > 0:
                return "提投产"
            if trend < 0:
                return "降投产"
            return "改投产"
        if "售价" in combined or "价格" in combined or "毛利" in combined or "涨价" in combined or "降价" in combined:
            if "涨价" in combined:
                return "涨价"
            if "降价" in combined:
                return "降价"
            trend = self._numeric_trend(old_value, new_value)
            if trend > 0:
                return "涨价"
            if trend < 0:
                return "降价"
            return "改价"
        if "优惠券" in combined or "新客立减" in combined or "优惠" in combined:
            return "改优惠"
        if "限时" in combined or "营销" in combined or "活动" in combined:
            return "改活动"
        if "推广模式" in combined or "无推广" in combined or "全站" in combined or "模式" in combined:
            return "改模式"
        if "同步" in combined:
            return "同步"
        metric = str(metric or "").strip()
        return metric if 0 < len(metric) <= 6 else "记录"

    def _numeric_trend(self, old_value, new_value):
        old_number = self._first_number(old_value)
        new_number = self._first_number(new_value)
        if old_number is None or new_number is None:
            return 0
        if new_number > old_number:
            return 1
        if new_number < old_number:
            return -1
        return 0

    def _first_number(self, value):
        match = re.search(r"-?\d+(?:\.\d+)?", str(value or ""))
        if not match:
            return None
        try:
            return float(match.group(0))
        except ValueError:
            return None

    def _metric_colors(self, metric):
        for key, colors in self.METRIC_STYLES.items():
            if key in metric:
                return QColor(colors[0]), QColor(colors[1])
        return QColor("#edf2f7"), QColor("#34495e")

    def _extract_spec_text(self, metric, text, change):
        if "规格" not in metric and "规格" not in text:
            return "", text

        bracket_matches = re.findall(r"\[([^\]]{1,80})\]", text)
        if bracket_matches:
            spec = bracket_matches[0].strip()
            cleaned = re.sub(r"\s*\[[^\]]{1,80}\]\s*", " ", text, count=1).strip()
            return spec, cleaned

        old_value = str(change.get("old", "") or "").strip()
        new_value = str(change.get("new", "") or "").strip()
        if metric == "规格名称" and (old_value or new_value):
            return (new_value or old_value), "规格名称已修改"

        for pattern in (
            r"^(.{1,80}?)(?:售价设置为|设置售价到|从.+?(?:涨价到|降价到))",
            r"^(?:新增规格|删除规格)\s*[:：]?\s*(.{1,80})$",
        ):
            match = re.search(pattern, text)
            if match:
                spec = match.group(1).strip()
                cleaned = text.replace(spec, "", 1).strip()
                return spec, cleaned
        return "", text

    def _draw_pill(self, painter, x, y, text, bg_color, fg_color, font, right, max_width=None):
        painter.setFont(font)
        fm = QFontMetrics(font)
        width = fm.horizontalAdvance(text) + 8
        if max_width:
            width = min(width, max_width)
        if x + width > right:
            return x
        rect = QRect(x, y + 1, width, 17)
        painter.setPen(Qt.NoPen)
        painter.setBrush(bg_color)
        painter.drawRoundedRect(rect, 4, 4)
        painter.setPen(fg_color)
        elided = fm.elidedText(text, Qt.ElideRight, width - 6)
        painter.drawText(rect.adjusted(3, 0, -3, 0), Qt.AlignCenter, elided)
        return x + width + 3

    def _draw_pill_or_wrap(self, painter, line_x, x, y, line_height, text, bg_color, fg_color, font, right):
        next_x = self._draw_pill(painter, x, y, text, bg_color, fg_color, font, right)
        if next_x != x or x == line_x:
            return next_x
        y += line_height
        return self._draw_pill(painter, line_x, y, text, bg_color, fg_color, font, right), y

    def _draw_wrapped_block(self, painter, x, y, text, bg_color, fg_color, font, right):
        painter.setFont(font)
        fm = QFontMetrics(font)
        text_rect = fm.boundingRect(QRect(0, 0, max(20, right - x - 8), 1000), Qt.TextWordWrap, text)
        rect = QRect(x, y, max(20, right - x), text_rect.height() + 6)
        painter.setPen(Qt.NoPen)
        painter.setBrush(bg_color)
        painter.drawRoundedRect(rect, 4, 4)
        painter.setPen(fg_color)
        painter.drawText(rect.adjusted(4, 3, -4, -3), Qt.TextWordWrap | Qt.AlignLeft | Qt.AlignTop, text)
        return rect.bottom() + 3

    def _draw_wrapped_text(self, painter, x, y, text, font, right, bottom):
        painter.setFont(font)
        painter.setPen(self._text_fg)
        fm = QFontMetrics(font)
        rect = QRect(x, y + 2, max(20, right - x), max(18, bottom - y))
        bounds = fm.boundingRect(rect, Qt.TextWordWrap | Qt.AlignLeft | Qt.AlignTop, text)
        painter.drawText(rect, Qt.TextWordWrap | Qt.AlignLeft | Qt.AlignTop, text)
        return min(bottom + 1, y + max(20, bounds.height() + 4))


class ArchiveProgressDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("存档")
        self.setFixedSize(380, 140)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint & ~Qt.WindowCloseButtonHint)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 15, 20, 15)
        layout.setSpacing(12)

        self.status_label = QLabel("正在准备...")
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setStyleSheet("font-size: 13px; color: #333; font-weight: bold;")
        layout.addWidget(self.status_label)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFixedHeight(22)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 1px solid #ddd;
                border-radius: 10px;
                background-color: #f0f0f0;
                text-align: center;
                font-size: 11px;
                font-weight: bold;
            }
            QProgressBar::chunk {
                background-color: #3498db;
                border-radius: 10px;
            }
        """)
        layout.addWidget(self.progress_bar)

    def set_status(self, text, value=None):
        self.status_label.setText(text)
        if value is not None:
            self.progress_bar.setValue(value)

    def set_error(self, text):
        self.status_label.setText(text)
        self.status_label.setStyleSheet("font-size: 13px; color: #e74c3c; font-weight: bold;")
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 1px solid #ddd;
                border-radius: 10px;
                background-color: #f0f0f0;
                text-align: center;
                font-size: 11px;
                font-weight: bold;
            }
            QProgressBar::chunk {
                background-color: #e74c3c;
                border-radius: 10px;
            }
        """)


APP_FONT_PRESETS = {
    "默认（微软雅黑）": "Microsoft YaHei",
    "黄油体": "ZCOOL QingKe HuangYou",
    "卡通体": "ZCOOL KuaiLe",
    "书法体": "Ma Shan Zheng",
    "宋体": "SimSun",
    "黑体": "SimHei",
    "圆体": "Resource Han Rounded CN",
}
BUNDLED_APP_FONT_FILES = {
    "黄油体": "ZCOOLQingKeHuangYou-Regular.ttf",
    "卡通体": "ZCOOLKuaiLe-Regular.ttf",
    "书法体": "MaShanZheng-Regular.ttf",
    "圆体": "ResourceHanRoundedCN-Regular.ttf",
}


def load_bundled_app_fonts():
    manager_dir = os.path.join(sys._MEIPASS, "manager") if getattr(sys, "frozen", False) else os.path.dirname(__file__)
    loaded = set()
    for name, filename in BUNDLED_APP_FONT_FILES.items():
        font_id = QFontDatabase.addApplicationFont(os.path.join(manager_dir, "fonts", filename))
        families = QFontDatabase.applicationFontFamilies(font_id) if font_id >= 0 else []
        if families:
            APP_FONT_PRESETS[name] = families[0]
            loaded.add(name)
    return loaded


def _set_application_font_family(family):
    app = QApplication.instance()
    font = app.font()
    font.setFamily(family)
    app.setFont(font)
    for widget in app.allWidgets():
        widget_font = QFont(widget.font())
        widget_font.setFamily(family)
        widget.setFont(widget_font)
        widget.updateGeometry()
        widget.update()


class SettingsDialog(QDialog):
    FONT_PRESETS = APP_FONT_PRESETS
    DEFAULT_FONT = "默认（微软雅黑）"
    HOTKEY_FIELDS = [
        ("quick_hotkey_main", "主界面快速呼出", DEFAULT_GLOBAL_HOTKEYS["quick_hotkey_main"]),
        ("quick_hotkey_cost_library", "成本库快速呼出", DEFAULT_GLOBAL_HOTKEYS["quick_hotkey_cost_library"]),
        ("quick_hotkey_material_library", "素材库快速呼出", DEFAULT_GLOBAL_HOTKEYS["quick_hotkey_material_library"]),
    ]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("设置")
        apply_window_icon(self, "settings")
        self.resize(680, 700)
        self.setMinimumSize(640, 680)
        self.hotkey_inputs = {}
        self.init_ui()
        self.load_settings()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        title = QLabel("⚙️ 程序设置")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50;")
        layout.addWidget(title)

        layout.addWidget(QLabel("<hr>"))

        self.auto_start_checkbox = QCheckBox("开机自启")
        self.auto_start_checkbox.setStyleSheet("""
            QCheckBox {
                spacing: 10px;
                font-size: 14px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
            }
        """)
        self.auto_start_checkbox.setToolTip("勾选后，程序将在Windows启动时自动运行")
        layout.addWidget(self.auto_start_checkbox)

        self.update_admin_checkbox = QCheckBox("管理员更新模式")
        self.update_admin_checkbox.setStyleSheet("""
            QCheckBox {
                spacing: 10px;
                font-size: 14px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
            }
        """)
        self.update_admin_checkbox.setToolTip("仅主电脑勾选。勾选后可以开启局域网更新服务并推送更新。")
        layout.addWidget(self.update_admin_checkbox)

        font_layout = QHBoxLayout()
        font_layout.addWidget(QLabel("软件字体"))
        self.font_combo = QComboBox()
        for name, family in self.FONT_PRESETS.items():
            self.font_combo.addItem(name)
            self.font_combo.setItemData(self.font_combo.count() - 1, QFont(family, 11), Qt.FontRole)
        self.font_combo.currentTextChanged.connect(
            lambda name: self.font_combo.setFont(QFont(self.FONT_PRESETS.get(name, "Microsoft YaHei"), 11))
        )
        self.font_combo.setToolTip("内置字体无需安装，保存后立即应用")
        font_layout.addWidget(self.font_combo, 1)
        layout.addLayout(font_layout)

        layout.addSpacing(10)
        layout.addWidget(QLabel("<hr>"))

        self.hotkey_panel = QWidget()
        hotkey_panel_layout = QVBoxLayout(self.hotkey_panel)
        hotkey_panel_layout.setContentsMargins(0, 0, 0, 0)
        hotkey_panel_layout.setSpacing(6)
        hotkey_title = QLabel("🚀 快速呼出快捷键")
        hotkey_title.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50;")
        hotkey_panel_layout.addWidget(hotkey_title)

        hotkey_hint = QLabel("用于从系统托盘/后台快速呼出窗口；成本库和素材库未打开时会自动打开。")
        hotkey_hint.setStyleSheet("font-size: 12px; color: #666;")
        hotkey_hint.setWordWrap(True)
        hotkey_panel_layout.addWidget(hotkey_hint)

        hotkey_grid = QGridLayout()
        hotkey_grid.setHorizontalSpacing(8)
        hotkey_grid.setVerticalSpacing(6)
        hotkey_grid.setColumnStretch(1, 1)
        for row, (key, label, _default_value) in enumerate(self.HOTKEY_FIELDS):
            label_widget = QLabel(label)
            editor = QKeySequenceEdit()
            editor.setObjectName(key)
            if hasattr(editor, "setMaximumSequenceLength"):
                editor.setMaximumSequenceLength(1)
            editor.setMinimumWidth(260)
            editor.setFixedHeight(34)
            editor.setStyleSheet("""
                QKeySequenceEdit {
                    border: 1px solid #c8d4c4;
                    border-radius: 4px;
                    background-color: white;
                    font-size: 13px;
                }
                QKeySequenceEdit:focus {
                    border: 1px solid #5f8a62;
                }
                QKeySequenceEdit QLineEdit {
                    border: none;
                    padding: 4px 8px;
                    background-color: white;
                    font-size: 13px;
                }
            """)
            editor.setToolTip("点击后按下新的快捷键；按退格可清空。建议使用 Ctrl+Shift+字母，避免和系统快捷键冲突。")
            clear_btn = QPushButton("清空")
            clear_btn.setFixedWidth(52)
            clear_btn.setFixedHeight(34)
            clear_btn.setStyleSheet("QPushButton { padding: 1px; }")
            clear_btn.clicked.connect(lambda _checked=False, e=editor: e.clear())
            hotkey_grid.addWidget(label_widget, row, 0)
            hotkey_grid.addWidget(editor, row, 1)
            hotkey_grid.addWidget(clear_btn, row, 2)
            self.hotkey_inputs[key] = editor
        hotkey_panel_layout.addLayout(hotkey_grid)
        layout.addWidget(self.hotkey_panel)

        layout.addWidget(QLabel("<hr>"))

        shortcuts_title = QLabel("⌨️ 快捷键说明")
        shortcuts_title.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50; padding-top: 5px;")
        layout.addWidget(shortcuts_title)

        shortcuts_table = QTableWidget()
        shortcuts_table.setColumnCount(2)
        shortcuts_table.setHorizontalHeaderLabels(["快捷键", "功能说明"])
        shortcuts_table.setRowCount(5)
        shortcuts_table.verticalHeader().setVisible(False)
        shortcuts_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        shortcuts_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        shortcuts_table.setItem(0, 0, QTableWidgetItem("Ctrl+F"))
        shortcuts_table.setItem(0, 1, QTableWidgetItem("聚焦搜索框，快速搜索商品"))
        shortcuts_table.setItem(1, 0, QTableWidgetItem("Ctrl+S"))
        shortcuts_table.setItem(1, 1, QTableWidgetItem("快速保存当前账号到本地存档"))
        shortcuts_table.setItem(2, 0, QTableWidgetItem("可设置"))
        shortcuts_table.setItem(2, 1, QTableWidgetItem("快速呼出主界面并临时置顶"))
        shortcuts_table.setItem(3, 0, QTableWidgetItem("可设置"))
        shortcuts_table.setItem(3, 1, QTableWidgetItem("快速打开/呼出成本库并临时置顶"))
        shortcuts_table.setItem(4, 0, QTableWidgetItem("可设置"))
        shortcuts_table.setItem(4, 1, QTableWidgetItem("快速打开/呼出素材库并临时置顶"))
        
        shortcuts_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        shortcuts_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        shortcuts_table.setColumnWidth(0, 100)
        shortcuts_table.verticalHeader().setDefaultSectionSize(35)
        shortcuts_table.setFixedHeight(175)
        shortcuts_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ddd;
                border-radius: 4px;
                background-color: #f9f9f9;
                gridline-color: #e0e0e0;
            }
            QTableWidget::item {
                padding: 5px 10px;
                font-size: 13px;
            }
            QHeaderView::section {
                background-color: #3498db;
                color: white;
                font-weight: bold;
                padding: 6px;
                border: none;
            }
        """)
        
        layout.addWidget(shortcuts_table)

        layout.addStretch()

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        btn_save = QPushButton("保存")
        btn_save.setFixedSize(100, 35)
        btn_save.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #219a52;
            }
        """)
        btn_save.clicked.connect(self.save_settings)
        btn_layout.addWidget(btn_save)
        
        btn_cancel = QPushButton("取消")
        btn_cancel.setFixedSize(100, 35)
        btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        btn_cancel.clicked.connect(self.reject)
        btn_layout.addWidget(btn_cancel)
        
        layout.addLayout(btn_layout)

    def load_settings(self):
        parent = self.parent()
        db = getattr(parent, "db", None)
        is_enabled = str(db.get_setting("auto_start_enabled", "1") if db else "1") != "0"
        self.auto_start_checkbox.setChecked(is_enabled)
        settings = load_global_update_settings()
        self.update_admin_checkbox.setChecked(str(settings.get("update_admin_mode", "0")) == "1")
        font_name = db.get_setting("app_font", self.DEFAULT_FONT) if db else self.DEFAULT_FONT
        self.font_combo.setCurrentText(font_name if font_name in self.FONT_PRESETS else self.DEFAULT_FONT)
        for key, _label, default_value in self.HOTKEY_FIELDS:
            value = db.get_setting(key, default_value) if db else default_value
            self.hotkey_inputs[key].setKeySequence(QKeySequence(value or ""))

    def save_settings(self):
        auto_start_enabled = self.auto_start_checkbox.isChecked()
        font_family = None
        if self.parent() and hasattr(self.parent(), "db"):
            settings = load_global_update_settings()
            was_admin = str(settings.get("update_admin_mode", "0")) == "1"
            admin_verified = str(settings.get("update_admin_verified", "0")) == "1"
            if self.update_admin_checkbox.isChecked() and not was_admin and not admin_verified:
                password, ok = QInputDialog.getText(
                    self,
                    "管理员验证",
                    "请输入开发者密码：",
                    QLineEdit.Password,
                )
                if not ok:
                    return
                if not verify_update_admin_password(password):
                    QMessageBox.warning(self, "密码错误", "开发者密码不正确。")
                    self.update_admin_checkbox.setChecked(False)
                    return
            save_global_update_setting("update_admin_verified", "1")
            save_global_update_setting("update_admin_mode", "1" if self.update_admin_checkbox.isChecked() else "0")
            font_name = self.font_combo.currentText()
            self.parent().db.set_setting("app_font", font_name)
            font_family = self.FONT_PRESETS[font_name]
            for key, _label, _default_value in self.HOTKEY_FIELDS:
                sequence = self.hotkey_inputs[key].keySequence().toString(QKeySequence.NativeText)
                self.parent().db.set_setting(key, sequence)
        
        if set_auto_start(auto_start_enabled):
            if self.parent() and hasattr(self.parent(), "db"):
                self.parent().db.set_setting("auto_start_enabled", "1" if auto_start_enabled else "0")
            failed_hotkeys = []
            if self.parent() and hasattr(self.parent(), "apply_global_hotkeys"):
                failed_hotkeys = self.parent().apply_global_hotkeys(show_message=False)
            if failed_hotkeys:
                QMessageBox.warning(
                    self,
                    "已保存",
                    "设置已保存，但以下快捷键注册失败，可能被系统或其它软件占用：\n"
                    + "\n".join(failed_hotkeys)
                )
            self.accept()
            if font_family:
                _set_application_font_family(font_family)
                self.parent().repaint()
            if not failed_hotkeys and hasattr(self.parent(), "show_toast"):
                self.parent().show_toast("设置已保存", 500)
        else:
            QMessageBox.warning(self, "错误", "开机自启设置失败，请检查 Windows 启动项权限。")
            self.reject()


class MainTableState:
    """主界面当前月份表格行状态，作为后续 Model/View 化的统一数据入口。"""

    def __init__(self):
        self.rows = []
        self.row_by_index = {}
        self.product_row_by_id = {}
        self.store_row_by_id = {}

    def clear(self):
        self.rows.clear()
        self.row_by_index.clear()
        self.product_row_by_id.clear()
        self.store_row_by_id.clear()

    def add_store(self, row_index, store_id, store_name, row_height):
        row = {
            "type": "store",
            "row": row_index,
            "store_id": store_id,
            "store_name": store_name,
            "product_id": None,
            "product_code": "",
            "title": store_name,
            "row_height": row_height,
        }
        self.rows.append(row)
        self.row_by_index[row_index] = row
        self.store_row_by_id[store_id] = row_index
        return row

    def add_product(self, row_index, store_id, product_id, product_code, title, row_height):
        row = {
            "type": "product",
            "row": row_index,
            "store_id": store_id,
            "store_name": "",
            "product_id": product_id,
            "product_code": product_code,
            "title": title,
            "row_height": row_height,
        }
        self.rows.append(row)
        self.row_by_index[row_index] = row
        self.product_row_by_id[product_id] = row_index
        return row


class OrderedFlowLayout(QLayout):
    """按插入顺序从左到右换行的可变尺寸布局。"""

    def __init__(self, parent=None, spacing=10):
        super().__init__(parent)
        self._items = []
        self.setContentsMargins(0, 0, 0, 0)
        self.setSpacing(spacing)

    def addItem(self, item):
        self._items.append(item)

    def count(self):
        return len(self._items)

    def itemAt(self, index):
        return self._items[index] if 0 <= index < len(self._items) else None

    def takeAt(self, index):
        return self._items.pop(index) if 0 <= index < len(self._items) else None

    def expandingDirections(self):
        return Qt.Orientations(Qt.Orientation(0))

    def hasHeightForWidth(self):
        return True

    def heightForWidth(self, width):
        return self._do_layout(QRect(0, 0, width, 0), True)

    def setGeometry(self, rect):
        super().setGeometry(rect)
        self._do_layout(rect, False)

    def sizeHint(self):
        return self.minimumSize()

    def minimumSize(self):
        size = QSize()
        for item in self._items:
            if item.widget() is not None and item.widget().isHidden():
                continue
            size = size.expandedTo(item.minimumSize())
        return size

    def _do_layout(self, rect, test_only):
        x, y, line_height = rect.x(), rect.y(), 0
        for item in self._items:
            if item.widget() is not None and item.widget().isHidden():
                continue
            hint = item.sizeHint()
            if x > rect.x() and x + hint.width() > rect.right() + 1:
                x = rect.x()
                y += line_height + self.spacing()
                line_height = 0
            if not test_only:
                item.setGeometry(QRect(QPoint(x, y), hint))
            x += hint.width() + self.spacing()
            line_height = max(line_height, hint.height())
        return y + line_height - rect.y()


def _promotion_link_visible(data, show_ordered_data_only):
    return (
        not show_ordered_data_only
        or (data is not None and float(data.get("net_orders") or 0) > 0)
    )

class ShopManagerApp(QMainWindow):
    startup_update_checked = pyqtSignal(object, bool, object)

    PRODUCT_ROW_HEIGHT = 112
    STORE_ROW_HEIGHT = 120
    TABLE_HEADER_HEIGHT = 28
    FROZEN_PRODUCT_CONTENT_WIDTH = 560
    
    def __init__(self):
        
        super().__init__()
        append_event("startup:window_init:start")
        self.data_root_manager = DataRootManager()
        append_event("startup:data_root_ready")
        self.startup_archive_account, startup_db_path = self.data_root_manager.resolve_startup_account()
        self.db = SafeDatabaseManager(startup_db_path) if startup_db_path else SafeDatabaseManager()
        self.db.init_default_prompts()
        font_name = self.db.get_setting("app_font", SettingsDialog.DEFAULT_FONT)
        _set_application_font_family(
            SettingsDialog.FONT_PRESETS.get(font_name, SettingsDialog.FONT_PRESETS[SettingsDialog.DEFAULT_FONT])
        )
        self.started_by_auto_start = "--autostart" in sys.argv
        auto_start_enabled = str(self.db.get_setting("auto_start_enabled", "1")) != "0"
        if not set_auto_start(auto_start_enabled):
            append_event("startup:auto_start_sync_failed")
        append_event("startup:db_ready")
        
        if getattr(sys, 'frozen', False):
            script_dir = os.path.dirname(sys.executable)
        else:
            script_dir = os.path.dirname(os.path.abspath(__file__))
        
        self.current_date = datetime.now()
        self.year = self.current_date.year
        self.month = self.current_date.month
        self.row_store_map = {}
        self.row_data_map = {}
        self.product_store_map = {}
        self.visible_product_ids = set()
        self.main_table_state = MainTableState()
        self.product_sort_mode = self.db.get_setting("product_sort_mode", "created_at") or "created_at"
        self.product_sort_descending = self.db.get_setting("product_sort_descending", "1") != "0"
        self.main_view_mode = "data"
        self.bubble_product_widgets = {}
        self._data_mode_refresh_pending = False
        self._pdd_link_refresh_pending = False
        self._manual_refresh_pending = False
        self._main_view_change_token = None

        self.is_loading = False  # 防止重复加载
        self._is_quitting = False  # 防止退出时重复触发自动上传
        self._is_switching_local_account = False  # 防止本地账号切换重复触发
        self.tray_icon = None
        self._today_col = -1  # 今日列索引，-1表示不是当月
        self.current_category_filter = ""
        self.current_search_match_ids = None
        self._search_highlighted_rows = set()
        self.current_store_filter = set()  # 店铺筛选状态
        self.store_sheet_buttons = {}
        self.daily_task_dialog = None
        self.material_mobile_service = None
        self.cost_sync_service = None
        self.store_margin_dialogs = {}
        self.promotion_data_dialogs = {}
        self.record_dialogs = []
        self._retired_widget_parking = QWidget(self)
        self._retired_widget_parking.hide()
        self._account_switch_retained_widgets = []
        self.product_spec_dialog = None
        self._active_reminder_ids = set()
        self._task_reminder_popup_active = False
        self._record_tooltip_cell = None
        self._record_tooltip_text = ""
        self._record_tooltip_pos = QPoint()
        self._button_hint_widgets = {}
        self._button_hint_active = None
        self._button_hint_timer = QTimer(self)
        self._button_hint_timer.setSingleShot(True)
        self._button_hint_timer.timeout.connect(self._show_button_hint)
        self._button_hint_label = QLabel(self)
        self._button_hint_label.setWindowFlags(Qt.ToolTip | Qt.FramelessWindowHint)
        self._button_hint_label.setAttribute(Qt.WA_TransparentForMouseEvents, True)
        self._button_hint_label.setAttribute(Qt.WA_ShowWithoutActivating, True)
        self._button_hint_label.setFocusPolicy(Qt.NoFocus)
        self._button_hint_label.setMaximumWidth(480)
        self._button_hint_label.setWordWrap(True)
        self._button_hint_label.setStyleSheet(
            "QLabel { background: white; color: #111; border: 1px solid #b8b8b8; padding: 5px 7px; }"
        )
        self._button_hint_label.hide()
        self._global_hotkey_actions = {}
        self._registered_global_hotkeys = set()
        self._global_hotkey_hwnd = None
        self._global_hotkey_last_trigger = {}
        self._local_shortcuts = []
        self.current_version = VERSION
        self.update_publish_service = UpdatePublishService()
        self.update_listener = None
        self.update_download_worker = None
        self.update_download_in_progress = False
        self.update_download_requested_version = ""
        self.update_cache_worker = None
        self.update_progress_dialog = None
        self.pending_downloaded_update_path = ""
        self.pending_update_manifest = None
        self._local_update_prompt_active = False
        self._dismissed_local_update_version = ""
        self.local_update_scan_timer = None
        self.pending_test_message = ""
        self.update_notification_queue = []
        self.update_notification_active = False
        self.startup_update_checked.connect(self._finish_startup_update_check)
        # 初始化存档管理器
        self.archive_manager = None
        try:
            try:
                from manager.archive_manager import ArchiveManager
            except ImportError:
                from archive_manager import ArchiveManager
            self.archive_manager = ArchiveManager(self.db)
            if self.startup_archive_account:
                account_id = self.startup_archive_account.get("id")
                if account_id:
                    self.archive_manager.switch_account(account_id)
                    self.archive_manager.set_active_data_account(account_id)
            print(f"[DATA_ARCHIVE] db_path={getattr(self.db, 'db_path', '')}")
            append_event("startup:archive_manager:ok")
        except Exception as e:
            append_exception("startup:archive_manager:failed", error=e)
            print(f"存档管理器初始化失败: {e}")
        append_event("startup:archive_ready")

        append_event("startup:init_ui:start")
        self.init_ui()
        append_event("startup:init_ui:done")
        QTimer.singleShot(100, self.finish_startup)

    def finish_startup(self):
        """Load data and background features after the event loop starts."""
        append_event("startup:finish_startup:start")
        self.init_system_tray()
        self.load_data_safe()
        self.start_global_reminder_check()
        self.update_archive_account_label()
        if MaterialMobileService.is_enabled():
            try:
                self.ensure_material_mobile_service()
            except Exception as e:
                append_exception("material_mobile_service:start", error=e)
        if self.db.get_cost_sync_state():
            try:
                self.ensure_cost_sync_service()
            except Exception as e:
                append_exception("cost_sync_service:start", error=e)
        self.init_shortcuts()
        self.apply_global_hotkeys(show_message=False)
        self.start_update_features()
        QTimer.singleShot(800, self.show_release_notes_once)
        append_event("startup:finish_startup:done")

    def start_global_reminder_check(self):
        """启动主界面的全局待办提醒检查。"""
        if getattr(self, "task_reminder_timer", None) and self.task_reminder_timer.isActive():
            return
        self.task_reminder_timer = QTimer(self)
        self.task_reminder_timer.timeout.connect(self.check_due_task_reminders)
        self.task_reminder_timer.start(10000)
        QTimer.singleShot(500, self.check_due_task_reminders)

    def get_product_pending_task_lines(self, product_id):
        """返回链接当前未完成任务，提醒任务优先显示提醒时间。"""
        reminder_rows = self.db.safe_fetchall(
            """SELECT task_content, remind_time FROM task_reminders
               WHERE product_id=? AND is_reminded=0 ORDER BY remind_time""",
            (product_id,),
        )
        task_rows = self.db.safe_fetchall(
            """SELECT task_content, created_time FROM daily_tasks
               WHERE product_id=? AND is_completed=0 ORDER BY created_time DESC""",
            (product_id,),
        )
        lines = []
        reminder_contents = set()
        for content, remind_time in reminder_rows:
            content = str(content or "").strip()
            if content:
                lines.append(f"提醒时间：{remind_time}\n任务内容：{content}")
                reminder_contents.add(content)
        for content, created_time in task_rows:
            content = str(content or "").strip()
            if content and content not in reminder_contents:
                lines.append(f"创建时间：{created_time}\n任务内容：{content}")
        return lines

    def create_product_reminder(self, store_id, product_id, task_content, remind_time):
        task_content = str(task_content or "").strip()
        remind_time = str(remind_time or "").strip()
        if not store_id or not product_id or not task_content or not remind_time:
            raise ValueError("提醒缺少店铺、链接、内容或时间")
        created_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.db.safe_execute(
            """INSERT INTO task_reminders
               (store_id, product_id, task_content, remind_time, created_time)
               VALUES (?, ?, ?, ?, ?)""",
            (store_id, product_id, task_content, remind_time, created_time),
        )
        record = self.record_product_operation(
            product_id,
            f"创建任务：{task_content}；提醒时间：{remind_time}",
            metric="创建任务",
            new=remind_time,
            change_type="task_reminder_created",
        )
        self.force_refresh_product_widget(product_id)
        dialog = getattr(self, "daily_task_dialog", None)
        if dialog:
            dialog.load_reminders()
        spec_dialog = getattr(self, "product_spec_dialog", None)
        if getattr(spec_dialog, "product_id", None) == product_id and hasattr(spec_dialog, "refresh_metric_change_panel"):
            spec_dialog.refresh_metric_change_panel()
        self.show_toast(f"已设置提醒：{remind_time}")
        return record

    def check_due_task_reminders(self):
        """检查已到时间的待办提醒，并按顺序强制弹窗。"""
        if getattr(self, "_task_reminder_popup_active", False):
            return
        try:
            now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            rows = self.db.safe_fetchall(
                """SELECT id, store_id, product_id, task_content, remind_time
                   FROM task_reminders
                   WHERE is_reminded = 0 AND remind_time <= ?
                   ORDER BY remind_time ASC
                   LIMIT 1""",
                (now_text,)
            )
            if not rows:
                return
            rem_id, store_id, product_id, task_content, remind_time = rows[0]
            if rem_id in self._active_reminder_ids:
                return

            reminder = self._build_task_reminder_payload(rem_id, store_id, product_id, task_content, remind_time)
            self._active_reminder_ids.add(rem_id)
            self._task_reminder_popup_active = True
            try:
                self.show_window()
                dialog = TaskReminderPopupDialog(reminder, self)
                dialog.raise_()
                dialog.activateWindow()
                result = dialog.exec_()
                if result == QDialog.Accepted and getattr(dialog, "completed", False):
                    self.db.safe_execute("DELETE FROM task_reminders WHERE id = ?", (rem_id,))
                    self._refresh_after_task_reminder_completed(product_id)
            finally:
                self._active_reminder_ids.discard(rem_id)
                self._task_reminder_popup_active = False
                QTimer.singleShot(100, self.check_due_task_reminders)
        except Exception as e:
            self._task_reminder_popup_active = False
            print(f"全局待办提醒检查失败: {e}")

    def _build_task_reminder_payload(self, rem_id, store_id, product_id, task_content, remind_time):
        store_name = ""
        product_code = str(product_id)
        product_title = ""
        product_image_data = None
        link_type = "未分类"
        try:
            store_rows = self.db.safe_fetchall("SELECT name FROM stores WHERE id=?", (store_id,))
            if store_rows and store_rows[0][0]:
                store_name = store_rows[0][0]
        except Exception:
            pass
        try:
            product_rows = self.db.safe_fetchall(
                """SELECT name, title, image_data, COALESCE(link_type, ''),
                          COALESCE(product_category_label, '')
                   FROM products WHERE id=?""",
                (product_id,),
            )
            if product_rows and product_rows[0][0]:
                product_code = str(product_rows[0][0])
                product_title = product_rows[0][1] or ""
                product_image_data = product_rows[0][2]
                link_type = product_rows[0][3] or product_rows[0][4] or "未分类"
        except Exception:
            pass
        return {
            "id": rem_id,
            "store_id": store_id,
            "store_name": store_name,
            "product_id": product_id,
            "product_code": product_code,
            "product_title": product_title,
            "product_image_data": product_image_data,
            "link_type": link_type,
            "task_content": task_content,
            "remind_time": remind_time,
        }

    def _refresh_after_task_reminder_completed(self, product_id):
        try:
            self.force_refresh_product_widget(product_id)
        except Exception as e:
            print(f"刷新提醒商品标签失败: {e}")
        dialog = getattr(self, "daily_task_dialog", None)
        if dialog:
            try:
                dialog.load_reminders()
                dialog.load_tasks()
            except Exception as e:
                print(f"刷新每日任务窗口失败: {e}")

    def init_system_tray(self):
        """初始化系统托盘"""
        if self.tray_icon is not None:
            if not self.tray_icon.isVisible():
                self.tray_icon.show()
            return
        if not QSystemTrayIcon.isSystemTrayAvailable():
            append_event("startup:tray_unavailable")
            return
        self.tray_icon = QSystemTrayIcon(self)
        icon = get_window_icon("main")
        if icon.isNull():
            icon = self.create_star_icon()
        self.tray_icon.setIcon(icon)
        
        tray_menu = QMenu()
        self.show_action = QAction("⭐ 显示主窗口", self)
        self.show_action.triggered.connect(self.show_window)
        tray_menu.addAction(self.show_action)
        
        tray_menu.addSeparator()
        
        self.settings_action = QAction("⚙️ 设置", self)
        self.settings_action.triggered.connect(self.show_settings_dialog)
        tray_menu.addAction(self.settings_action)
        
        tray_menu.addSeparator()
        
        quit_action = QAction("❌ 退出", self)
        quit_action.triggered.connect(self.quit_application)
        tray_menu.addAction(quit_action)
        
        self.tray_icon.setContextMenu(tray_menu)
        self.tray_icon.activated.connect(self.on_tray_activated)
        self.tray_icon.show()
        
        if not self.started_by_auto_start:
            self.tray_icon.showMessage(
                "电商店铺操作记录管理工具",
                "程序已最小化到系统托盘，双击图标可显示窗口",
                QSystemTrayIcon.Information,
                3000
            )
    
    def create_star_icon(self):
        """创建星星图标"""
        import sys
        if getattr(sys, 'frozen', False):
            icons_dir = os.path.join(sys._MEIPASS, "manager", "icons")
        else:
            icons_dir = os.path.join(os.path.dirname(__file__), "icons")
        svg_path = os.path.join(icons_dir, "xingxing.svg")
        main_icon_path = icon_path("main")
        if os.path.exists(main_icon_path):
            return QIcon(main_icon_path)
        if os.path.exists(svg_path):
            renderer = QSvgRenderer(svg_path)
            pixmap = QPixmap(32, 32)
            pixmap.fill(Qt.transparent)
            painter = QPainter(pixmap)
            renderer.render(painter)
            painter.end()
            return QIcon(pixmap)
        else:
            pixmap = QPixmap(32, 32)
            pixmap.fill(Qt.transparent)
            painter = QPainter(pixmap)
            painter.setRenderHint(QPainter.Antialiasing)
            font = QFont()
            font.setPixelSize(24)
            painter.setFont(font)
            painter.drawText(pixmap.rect(), Qt.AlignCenter, "⭐")
            painter.end()
            return QIcon(pixmap)
    
    def on_tray_activated(self, reason):
        """托盘图标被激活时的处理"""
        if reason == QSystemTrayIcon.DoubleClick:
            self.show_window()
    
    def show_window(self):
        """显示窗口"""
        self.showNormal()
        self.raise_()
        self.activateWindow()
        QTimer.singleShot(100, self.show_pending_update_notification)

    def _get_pdd_browser_monitor(self):
        account = self.archive_manager.get_active_data_account() if self.archive_manager else None
        profile_base_dir = self.archive_manager._archive_browser_dir_for_account(account) if account else None
        if not profile_base_dir:
            raise RuntimeError("当前数据未绑定存档账号，无法确定浏览器数据目录")

        monitor = getattr(self, "pdd_browser_monitor", None)
        if monitor and os.path.abspath(os.path.dirname(monitor.legacy_profile_root)) != os.path.abspath(profile_base_dir):
            monitor.stop()
            monitor = None
        if monitor is None:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            monitor = PddBrowserMonitor(base_dir, profile_base_dir=profile_base_dir)
            self.pdd_browser_monitor = monitor
        return monitor

    def open_pinduoduo(self):
        store_id = self.store_combo.currentData() if hasattr(self, "store_combo") else None
        if not store_id:
            QMessageBox.information(self, "请选择店铺", "请先在主界面店铺筛选中选择具体店铺，再打开拼多多商家后台。")
            return
        try:
            state = self._get_pdd_browser_monitor().activate_store_browser(store_id, open_url=True, open_new_tab=False)
            self.statusBar().showMessage(f"已按店铺 {store_id} 打开拼多多商家后台，端口 {state.get('port')}", 3000)
        except Exception as e:
            QMessageBox.warning(self, "打开失败", f"打开拼多多商家后台失败：{e}")

    def _open_pdd_fetch_for_store(self, store_id, mode, product_id=""):
        if not store_id:
            QMessageBox.information(self, "请选择店铺", "请先选择要添加链接的店铺。")
            return
        titles = {
            "code": "抓取添加编码",
            "price": "抓取价格管理",
            "promotion_status": "抓取推广状态",
        }
        title = titles.get(mode, "拼多多链接抓取")
        product_id = str(product_id or "").strip()
        try:
            monitor = self._get_pdd_browser_monitor()
            monitor.activate_store_browser(store_id, open_url=True, open_new_tab=False)
            if mode == "code" and product_id:
                QApplication.clipboard().setText(product_id)
            try:
                from manager.dialogs.store_margin import PddProductMatchDialog, PddPromotionStatusDialog
            except ImportError:
                from dialogs.store_margin import PddProductMatchDialog, PddPromotionStatusDialog
            cache_name = f"pdd_{mode}_fetch_dialogs"
            if not hasattr(self, cache_name) or getattr(self, cache_name) is None:
                setattr(self, cache_name, {})
            dialogs = getattr(self, cache_name)
            dialog = dialogs.get(store_id)
            if dialog is None:
                if mode == "promotion_status":
                    dialog = PddPromotionStatusDialog(
                        self.db, monitor, lambda sid=store_id: sid, parent=None, owner=self
                    )
                else:
                    dialog = PddProductMatchDialog(
                        self.db, monitor, default_store_id=store_id, parent=None, mode=mode,
                        store_id_provider=lambda sid=store_id: sid, owner=self,
                    )
                dialog.destroyed.connect(lambda _=None, sid=store_id, name=cache_name: getattr(self, name, {}).pop(sid, None))
                dialogs[store_id] = dialog
            if dialog.isMinimized():
                dialog.showNormal()
            dialog.setWindowState(dialog.windowState() & ~Qt.WindowMinimized | Qt.WindowActive)
            dialog.show()
            dialog.raise_()
            dialog.activateWindow()
            if mode == "code" and product_id and hasattr(dialog, "lbl_summary"):
                dialog.lbl_summary.setText(f"已打开商家端并复制商品ID：{product_id}。请手动进入添加/编辑编码界面后开始抓取。")
            self.statusBar().showMessage(f"已打开店铺 {store_id} 的拼多多{title}窗口", 3000)
        except Exception as e:
            QMessageBox.warning(self, "拼多多链接抓取", f"打开{title}窗口失败：{e}")

    def open_pdd_code_fetch_for_store(self, store_id, product_id=""):
        self._open_pdd_fetch_for_store(store_id, "code", product_id)

    def open_pdd_price_fetch_for_store(self, store_id):
        self._open_pdd_fetch_for_store(store_id, "price")

    def open_pdd_promotion_status_fetch_for_store(self, store_id):
        self._open_pdd_fetch_for_store(store_id, "promotion_status")

    def show_settings_dialog(self):
        """打开设置对话框"""
        dialog = SettingsDialog(self)
        dialog.exec_()

    def show_tutorial_catalog(self):
        """打开可重复阅读的只读功能教程。"""
        controller = getattr(self, "tutorial_controller", None)
        if controller is None:
            try:
                from manager.tutorial import TutorialController
            except ImportError:
                from tutorial import TutorialController
            controller = TutorialController(self)
            self.tutorial_controller = controller
        controller.show_catalog()

    def get_current_release_notes(self):
        return CURRENT_RELEASE_NOTES

    def show_current_release_notes(self, notes=None):
        QMessageBox.information(
            self,
            f"v{self.current_version} 本次版本更新内容",
            str(notes or self.get_current_release_notes()).strip(),
        )

    def show_release_notes_once(self):
        settings = load_global_update_settings()
        notes = startup_release_notes(settings, self.current_version)
        if not notes:
            installed = load_pending_update(
                self.current_version,
                verify_hash=False,
                allow_current=True,
            )
            if installed and str(installed.get("version") or "").strip() == self.current_version:
                fallback_settings = dict(settings)
                fallback_settings["last_update_manifest"] = installed
                notes = startup_release_notes(fallback_settings, self.current_version)
        if not notes:
            return False
        save_global_update_setting("release_notes_seen_version", self.current_version)
        self.show_current_release_notes(notes)
        return True

    def resolve_tutorial_target(self, name):
        """返回当前数据卡片界面的真实教程目标。"""
        target = None
        if name == "first_store_bubble":
            for section in getattr(self, "_data_mode_store_sections", {}).values():
                if _qobject_alive(section) and not section.isHidden():
                    candidate = section.findChild(StoreWidget)
                    if _qobject_alive(candidate) and not candidate.isHidden():
                        target = candidate
                        break
        elif name == "first_product_bubble":
            for candidate in getattr(self, "bubble_product_widgets", {}).values():
                if _qobject_alive(candidate) and not candidate.isHidden():
                    target = candidate
                    break
        if target is not None and hasattr(self, "data_mode_scroll"):
            self.data_mode_scroll.ensureWidgetVisible(target, 20, 20)
            return target
        return getattr(self, "data_mode_container", self)

    def open_tutorial_screen(self, screen):
        """只打开教程要展示的界面，不触发保存、导入、同步或外部程序。"""
        self.show_window()
        if screen == "main":
            return self, False, ""

        if screen == "settings":
            dialog = SettingsDialog(self)
            apply_window_icon(dialog, "settings")
            dialog.setAttribute(Qt.WA_DeleteOnClose, True)
            dialog.show()
            return dialog, True, ""

        if screen == "margin_input":
            try:
                from manager.dialogs.input_data_dialog import InputDataDialog
            except ImportError:
                from dialogs.input_data_dialog import InputDataDialog
            dialog = InputDataDialog(self, tutorial_mode=True)
            apply_window_icon(dialog, "store")
            dialog.setAttribute(Qt.WA_DeleteOnClose, True)
            dialog.show()
            return dialog, True, ""

        if screen == "api":
            dialog = ApiConfigDialog(self.db, self)
            dialog.setAttribute(Qt.WA_DeleteOnClose, True)
            dialog.show()
            return dialog, True, ""

        if screen == "archive":
            try:
                from manager.archive_manager import ArchiveDialog
            except ImportError:
                from archive_manager import ArchiveDialog
            dialog = ArchiveDialog(self.db, self.archive_manager, self)
            apply_window_icon(dialog, "archive")
            dialog.setAttribute(Qt.WA_DeleteOnClose, True)
            dialog.show()
            return dialog, True, ""

        if screen == "cost":
            existing = getattr(self, "cost_library_dialog", None)
            created = not _qobject_alive(existing)
            self.show_cost_library()
            return getattr(self, "cost_library_dialog", None), created, ""

        if screen == "material":
            existing = getattr(self, "material_library_dialog", None)
            created = not _qobject_alive(existing)
            self.show_material_library()
            return getattr(self, "material_library_dialog", None), created, ""

        if screen == "daily":
            existing = getattr(self, "daily_task_dialog", None)
            if _qobject_alive(existing):
                existing.show()
                existing.raise_()
                return existing, False, ""
            dialog = DailyTaskDialog(self.db, self)
            dialog.setAttribute(Qt.WA_DeleteOnClose, True)
            dialog.destroyed.connect(lambda _obj=None: setattr(self, "daily_task_dialog", None))
            self.daily_task_dialog = dialog
            dialog.show()
            return dialog, True, ""

        if screen == "records":
            rows = self.db.safe_fetchall("SELECT id FROM stores ORDER BY sort_order, id LIMIT 1")
            if not rows:
                return self, False, "当前账号还没有店铺，先查看示例；创建店铺后可打开真实操作记录窗口。"
            dialog = self.open_store_record_window(rows[0][0])
            return dialog, True, ""

        if screen == "store_margin":
            rows = self.db.safe_fetchall("SELECT id, name FROM stores ORDER BY sort_order, id LIMIT 1")
            if not rows:
                return self, False, "当前账号还没有店铺，先查看示例；创建店铺后可进入店铺毛利管理。"
            store_id, store_name = rows[0]
            existing = self.store_margin_dialogs.get(store_id)
            created = not _qobject_alive(existing)
            self.open_store_margin_dialog(store_id, store_name or "示例店铺")
            return self.store_margin_dialogs.get(store_id), created, ""

        if screen == "product_spec":
            existing = getattr(self, "product_spec_dialog", None)
            if _qobject_alive(existing):
                existing.show()
                existing.raise_()
                return existing, False, ""
            rows = self.db.safe_fetchall(
                "SELECT id, name, title FROM products WHERE COALESCE(is_archived, 0)=0 ORDER BY id LIMIT 1"
            )
            if not rows:
                return self, False, "当前账号还没有链接，先查看示例；添加链接后可打开真实规格毛利窗口。"
            product_id, product_code, title = rows[0]
            dialog = self.open_product_spec_dialog(
                self.db, product_id, product_code or "示例商品", title or "示例商品", self
            )
            return dialog, True, ""

        if screen == "promotion":
            rows = self.db.safe_fetchall("SELECT id FROM stores ORDER BY sort_order, id LIMIT 1")
            if not rows:
                return self, False, "当前账号还没有店铺，先查看示例；导入推广数据后可进入分析窗口。"
            store_id = rows[0][0]
            existing = self.promotion_data_dialogs.get(store_id)
            created = not _qobject_alive(existing)
            self.open_promotion_data_for_store(store_id)
            return self.promotion_data_dialogs.get(store_id), created, ""

        return self, False, "该功能会打开外部程序或立即执行操作，教程仅展示入口，不会实际触发。"

    def start_update_features(self):
        """接收前台广播，并监测后台代理已下载的更新。"""
        if self.update_listener is None:
            self.update_listener = UpdateBroadcastListener(self)
            self.update_listener.updateReceived.connect(self.on_update_broadcast_received)
            self.update_listener.start()
        QTimer.singleShot(250, self.check_local_update_package_on_startup)
        QTimer.singleShot(500, self.ensure_update_agent_task)
        if self.local_update_scan_timer is None:
            self.local_update_scan_timer = QTimer(self)
            self.local_update_scan_timer.setInterval(3000)
            self.local_update_scan_timer.timeout.connect(self.check_local_update_package_on_startup)
            self.local_update_scan_timer.start()

    def ensure_update_agent_task(self):
        settings = load_global_update_settings()
        publisher = str(settings.get("update_publish_enabled", "0")) == "1"
        trusted = bool(
            str(settings.get("trusted_update_server_id") or "").strip()
            and str(settings.get("trusted_update_server_url") or "").strip()
        )
        if trusted and str(settings.get("auto_update_enabled", "0")) != "1":
            save_global_update_setting("auto_update_enabled", "1")

        def configure_task():
            if publisher or trusted:
                ok, error = install_update_agent_task(run_now=True)
                if not ok:
                    append_event(f"update:agent_task_failed error={error}")
            else:
                uninstall_update_agent_task()

        threading.Thread(target=configure_task, daemon=True).start()

    def _trust_update_source_if_needed(self, manifest):
        server_id = str(manifest.get("server_id") or "").strip()
        if not server_id:
            return False
        settings = load_global_update_settings()
        trusted_id = str(settings.get("trusted_update_server_id") or "").strip()
        if trusted_id:
            if not is_trusted_update_manifest(manifest, settings):
                QMessageBox.warning(self, "已拦截更新", "更新来源与首次绑定的主电脑不一致。")
                return None
            return True

        source = str(manifest.get("server_host") or manifest.get("_sender_ip") or "局域网主电脑")
        reply = QMessageBox.question(
            self,
            "信任更新主电脑",
            f"是否信任主电脑“{source}”并自动接收更新？\n\n"
            "信任后，即使主软件未打开，后台更新助手也会下载经过 SHA256 校验的 EXE。\n"
            "下载完成后再由你选择立即重启或稍后自行重启。",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )
        if reply != QMessageBox.Yes:
            return None
        try:
            bind_trusted_update_source(manifest)
            ok, error = install_update_agent_task(run_now=True)
            if not ok:
                raise RuntimeError(error or "后台更新助手启动失败")
            self.show_toast("已信任主电脑，正在后台下载更新", 1800)
            return True
        except Exception as e:
            QMessageBox.warning(self, "信任失败", f"更新来源保存失败：{e}")
            return None

    def check_local_update_package_on_startup(self, manual=False):
        if self._local_update_prompt_active:
            return True
        if not manual and (not self.isVisible() or self.isMinimized()):
            return False
        pending_candidate = load_pending_update(self.current_version, verify_hash=False)
        if pending_candidate and (
            not manual
            and str(pending_candidate.get("version") or "").strip() == self._dismissed_local_update_version
        ):
            return False
        candidate = pending_candidate
        if candidate:
            candidate["_local_path"] = candidate.get("_cached_path")
        if not candidate:
            candidate = find_newer_local_package(self.current_version, app_dir())
        if not candidate:
            return False
        remote_version = str(candidate.get("version") or "").strip()
        if not remote_version or (not manual and remote_version == self._dismissed_local_update_version):
            return False
        source_path = os.path.abspath(str(candidate.get("_local_path") or ""))
        if not os.path.isfile(source_path):
            return False
        self._local_update_prompt_active = True
        try:
            if pending_candidate:
                validated = load_pending_update(self.current_version)
                if not validated or str(validated.get("version") or "") != remote_version:
                    QMessageBox.warning(self, "更新校验失败", "更新文件校验失败，请重新检查更新。")
                    return True
                source_path = os.path.abspath(str(validated.get("_cached_path") or ""))
            target_path = source_path
            if os.path.dirname(source_path).lower() != os.path.abspath(app_dir()).lower():
                filename = os.path.basename(source_path)
                target_path = os.path.join(app_dir(), filename)
                if os.path.abspath(target_path).lower() == os.path.abspath(sys.executable).lower():
                    target_path = os.path.join(app_dir(), f"shop_manager_v{remote_version}.exe")
                temp_dir = os.path.join(app_dir(), ".update_tmp")
                os.makedirs(temp_dir, exist_ok=True)
                temp_path = os.path.join(temp_dir, f"{filename}.{os.getpid()}.ready")
                try:
                    shutil.copy2(source_path, temp_path)
                    os.replace(temp_path, target_path)
                finally:
                    if os.path.exists(temp_path):
                        os.remove(temp_path)
                    try:
                        os.rmdir(temp_dir)
                    except OSError:
                        pass
            if self._ask_update_restart(target_path, remote_version):
                self.quit_application()
            else:
                self._dismissed_local_update_version = remote_version
        except Exception as e:
            QMessageBox.warning(self, "更新准备失败", str(e))
        finally:
            self._local_update_prompt_active = False
        return True

    def _ask_update_restart(self, target_path, remote_version=""):
        self._update_exe_to_launch = target_path
        dialog = QMessageBox(self)
        dialog.setWindowTitle("更新下载完成")
        version_text = f" v{remote_version}" if remote_version else ""
        dialog.setText(
            f"新版本{version_text}已下载并校验完成：\n"
            f"{os.path.basename(target_path)}\n\n请选择重启时间。"
        )
        restart_button = dialog.addButton("立即重启", QMessageBox.AcceptRole)
        dialog.addButton("稍后自行重启", QMessageBox.RejectRole)
        dialog.setDefaultButton(restart_button)
        dialog.exec_()
        restart_now = dialog.clickedButton() is restart_button
        if not restart_now:
            self.show_toast("稍后退出软件时将自动打开新版本", 2200)
        return restart_now

    def verify_developer_password(self):
        settings = load_global_update_settings()
        if str(settings.get("update_admin_verified", "0")) == "1":
            return True
        password, ok = QInputDialog.getText(
            self,
            "开发者模式",
            "请输入开发者密码：",
            QLineEdit.Password,
        )
        if not ok:
            return False
        if not verify_update_admin_password(password):
            QMessageBox.warning(self, "密码错误", "开发者密码不正确。")
            return False
        save_global_update_setting("update_admin_verified", "1")
        return True

    def show_update_center(self):
        """打开开发者更新发布中心。"""
        if not self.verify_developer_password():
            return
        save_global_update_setting("update_admin_mode", "1")
        dialog = UpdateAdminDialog(self, self.update_publish_service, self)
        dialog.exec_()

    def check_update_on_startup(self, show_no_update=False):
        """启动时只检查一次保存的更新地址。"""
        settings = load_global_update_settings()
        server_url = normalize_server_url(settings.get("update_server_url", "") or "")
        if not server_url:
            if show_no_update:
                QMessageBox.information(self, "更新", "还没有设置更新地址。")
            return
        if not show_no_update:
            threading.Thread(
                target=self._check_update_on_startup_worker,
                args=(server_url,),
                daemon=True,
            ).start()
            return
        try:
            manifest = fetch_manifest(server_url, timeout=3)
        except Exception as e:
            if show_no_update:
                QMessageBox.warning(self, "检查失败", f"检查更新失败：{e}")
            return
        if self.is_local_update_broadcast(manifest):
            if show_no_update:
                QMessageBox.information(self, "更新", "这是本机发布的更新地址，已跳过本机提示。")
            return
        self.handle_update_manifest(manifest, show_no_update=show_no_update)

    def _check_update_on_startup_worker(self, server_url):
        try:
            self.startup_update_checked.emit(fetch_manifest(server_url, timeout=3), False, None)
        except Exception as e:
            self.startup_update_checked.emit(None, False, e)

    def _finish_startup_update_check(self, manifest, show_no_update=False, error=None):
        if error is not None:
            if show_no_update:
                QMessageBox.warning(self, "检查失败", f"检查更新失败：{error}")
            return
        if not manifest:
            return
        if self.is_local_update_broadcast(manifest):
            if show_no_update:
                QMessageBox.information(self, "更新", "这是本机发布的更新地址，已跳过本机提示。")
            return
        self.handle_update_manifest(manifest, show_no_update=show_no_update)

    def check_saved_or_server_update(self):
        """手动检查更新：优先使用最近一次广播收到的下载地址。"""
        if self.check_local_update_package_on_startup(manual=True):
            return
        settings = load_global_update_settings()
        saved_manifest = settings.get("last_update_manifest")
        if isinstance(saved_manifest, dict):
            saved_version = str(saved_manifest.get("version", "")).strip()
            saved_url = str(saved_manifest.get("url", "")).strip()
            if (
                saved_version
                and saved_url
                and not self.is_local_update_broadcast(saved_manifest)
                and is_newer_version(saved_version, self.current_version)
            ):
                self.handle_update_manifest(saved_manifest, show_no_update=True)
                return
        self.check_update_on_startup(show_no_update=True)

    def remember_update_manifest(self, manifest):
        """保存最近一次收到的更新包地址，便于稍后手动检查和下载。"""
        if not isinstance(manifest, dict):
            return
        update_url = str(manifest.get("url", "") or "").strip()
        if not update_url.startswith(("http://", "https://")):
            return
        stored_manifest = {k: v for k, v in manifest.items() if not str(k).startswith("_")}
        save_global_update_setting("last_update_manifest", stored_manifest)
        save_global_update_setting("last_update_url", update_url)
        save_global_update_setting("last_update_version", str(manifest.get("version", "") or "").strip())
        save_global_update_setting("last_update_filename", manifest.get("filename", "") or os.path.basename(update_url))
        server_url = "/".join(update_url.split("/")[:3])
        save_global_update_setting("update_server_url", normalize_server_url(server_url))

    def get_local_update_ips(self):
        ips = {"127.0.0.1", "::1", "localhost"}
        try:
            ips.add(socket.gethostbyname(socket.gethostname()))
        except Exception:
            pass
        try:
            for item in socket.getaddrinfo(socket.gethostname(), None):
                ip = item[4][0]
                if ip:
                    ips.add(ip)
        except Exception:
            pass
        probe = None
        try:
            probe = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            probe.connect(("8.8.8.8", 80))
            ips.add(probe.getsockname()[0])
        except Exception:
            pass
        finally:
            if probe:
                probe.close()
        return {ip for ip in ips if ip}

    def is_local_update_broadcast(self, manifest):
        local_ips = self.get_local_update_ips()
        sender_ip = str(manifest.get("_sender_ip", "") or "").strip()
        if sender_ip in local_ips:
            return True
        update_url = str(manifest.get("url", "") or "").strip()
        if "://" in update_url:
            try:
                host = update_url.split("://", 1)[1].split("/", 1)[0].split("@")[-1].split(":", 1)[0]
            except Exception:
                host = ""
            if host in local_ips:
                return True
        return False

    def enqueue_update_notification(self, item):
        if not isinstance(item, dict):
            return
        if self.update_notification_active or self.update_notification_queue:
            return
        self.update_notification_queue.append(item)
        self.schedule_update_notification()

    def schedule_update_notification(self):
        if self.isVisible() and not self.isMinimized():
            self.raise_()
            self.activateWindow()
            QTimer.singleShot(0, self.show_pending_update_notification)

    def on_update_broadcast_received(self, manifest):
        """在线时收到可信更新广播后直接下载。"""
        if self.is_local_update_broadcast(manifest):
            return
        if manifest.get("message_kind") == "test":
            self.enqueue_update_notification({
                "type": "test",
                "message": manifest.get("message", "局域网测试消息"),
            })
            return
        self.remember_update_manifest(manifest)
        remote_version = str(manifest.get("version", "")).strip()
        if is_newer_version(remote_version, self.current_version):
            if (
                self.update_download_in_progress
                or self.update_download_requested_version
                or self.pending_downloaded_update_path
            ):
                return
            self.update_download_requested_version = remote_version
            try:
                self.handle_update_manifest(manifest, show_no_update=False, auto_download=True)
            finally:
                if not self.update_download_in_progress:
                    self.update_download_requested_version = ""

    def show_pending_update_notification(self):
        if self.pending_downloaded_update_path:
            target_path = self.pending_downloaded_update_path
            self.pending_downloaded_update_path = ""
            if self._ask_update_restart(target_path):
                self.quit_application()
            return
        if self.update_notification_active:
            return
        if self.update_notification_queue:
            item = self.update_notification_queue.pop(0)
            self.update_notification_active = True
            try:
                if item.get("type") == "test":
                    QMessageBox.information(self, "收到局域网测试", item.get("message") or "局域网测试消息")
                elif item.get("type") == "update":
                    self.handle_update_manifest(item.get("manifest") or {}, show_no_update=False)
            finally:
                self.update_notification_active = False
                if self.update_notification_queue:
                    QTimer.singleShot(0, self.show_pending_update_notification)
            return
        if self.pending_test_message:
            message = self.pending_test_message
            self.pending_test_message = ""
            QMessageBox.information(self, "收到局域网测试", message)
            return
        if self.pending_update_manifest:
            manifest = self.pending_update_manifest
            self.pending_update_manifest = None
            self.handle_update_manifest(manifest, show_no_update=False)

    def handle_update_manifest(self, manifest, show_no_update=False, auto_download=False):
        self.remember_update_manifest(manifest)
        remote_version = str(manifest.get("version", "")).strip()
        if not remote_version:
            return
        if not is_newer_version(remote_version, self.current_version):
            if show_no_update:
                QMessageBox.information(self, "更新", f"当前已经是最新版本：v{self.current_version}")
            return
        trust_state = self._trust_update_source_if_needed(manifest)
        if trust_state is None:
            return
        if trust_state:
            start_update_agent_task()
            payload = dict(manifest)
            notify_update_agent(payload)
            QTimer.singleShot(1000, lambda: notify_update_agent(payload))
            return
        if auto_download:
            self.download_update_package(manifest)
            return
        filename = manifest.get("filename") or os.path.basename(manifest.get("url", ""))
        message = (
            f"发现新版本 v{remote_version}\n\n"
            f"文件：{filename}\n"
            f"下载位置：{app_dir()}\n\n"
            "下载完成后将自动关闭旧版本并打开新版本，请勿操作。"
        )
        reply = QMessageBox.question(
            self,
            "发现更新",
            message,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )
        if reply == QMessageBox.Yes:
            self.download_update_package(manifest)

    def download_update_package(self, manifest):
        if self.update_download_in_progress:
            return
        self.update_download_in_progress = True
        self.update_download_requested_version = str(manifest.get("version", "")).strip()
        try:
            self.update_progress_dialog = QProgressDialog("正在下载更新...", "", 0, 100, self)
            self.update_progress_dialog.setWindowTitle("下载更新")
            self.update_progress_dialog.setWindowModality(Qt.WindowModal)
            self.update_progress_dialog.setMinimumDuration(0)
            self.update_progress_dialog.setAutoClose(False)
            self.update_progress_dialog.setAutoReset(False)
            self.update_progress_dialog.setCancelButton(None)
            self.update_progress_dialog.setValue(0)
            self.update_progress_dialog.show()

            self.update_download_worker = UpdateDownloadWorker(manifest, app_dir(), self)
            self.update_download_worker.progressChanged.connect(
                lambda value: self._set_update_progress(
                    value,
                    f"正在后台下载新版本：{value}%",
                )
            )
            self.update_download_worker.finishedOk.connect(self.on_update_download_finished)
            self.update_download_worker.failed.connect(self.on_update_download_failed)
            self.update_download_worker.finished.connect(self.on_update_download_worker_stopped)
            self.update_download_worker.start()
        except Exception as e:
            self.update_download_in_progress = False
            self.update_download_requested_version = ""
            self.update_download_worker = None
            if self.update_progress_dialog:
                self.update_progress_dialog.close()
                self.update_progress_dialog = None
            QMessageBox.warning(self, "下载失败", f"无法启动更新下载：{e}")

    def _set_update_progress(self, value, text):
        dialog = self.update_progress_dialog
        if not _qobject_alive(dialog):
            return
        value = max(0, min(100, int(value)))
        dialog.setLabelText(f"{text}\n总进度：{value}%")
        dialog.setValue(value)
        QApplication.processEvents()

    def on_update_download_finished(self, target_path):
        if self.update_progress_dialog:
            self._set_update_progress(100, "新版本下载完成")
            self.update_progress_dialog.close()
            self.update_progress_dialog = None
        self.pending_downloaded_update_path = target_path

    def on_update_download_worker_stopped(self):
        worker = self.update_download_worker
        self.update_download_worker = None
        self.update_download_in_progress = False
        self.update_download_requested_version = ""
        if worker is not None:
            worker.deleteLater()
        if not self.pending_downloaded_update_path:
            return
        if self.isVisible() and not self.isMinimized():
            QTimer.singleShot(0, self.show_pending_update_notification)
        elif self.tray_icon:
            self.tray_icon.showMessage(
                "更新下载完成",
                "打开软件后可选择是否重启到新版本",
                QSystemTrayIcon.Information,
                3000,
            )

    def on_update_download_failed(self, error):
        if self.update_progress_dialog:
            self.update_progress_dialog.close()
            self.update_progress_dialog = None
        QMessageBox.warning(self, "下载失败", f"下载更新失败：{error}")

    def show_shortcuts_dialog(self):
        """打开快捷键说明对话框（复用设置对话框）"""
        self.show_settings_dialog()

    def init_shortcuts(self):
        """初始化快捷键（仅在主界面激活时生效）"""
        if self._local_shortcuts:
            return
        for sequence, callback in (
            ("Ctrl+F", self.focus_search),
            ("Ctrl+S", self.quick_save_archive),
        ):
            shortcut = QShortcut(QKeySequence(sequence), self)
            shortcut.activated.connect(callback)
            self._local_shortcuts.append(shortcut)

    def apply_global_hotkeys(self, show_message=False):
        """注册系统级快速呼出快捷键。"""
        self.unregister_global_hotkeys()
        if _USER32 is None:
            return ["当前系统暂不支持全局快捷键"]

        hotkeys = [
            (1, "主界面", "quick_hotkey_main", self.quick_show_main_window),
            (2, "成本库", "quick_hotkey_cost_library", self.quick_show_cost_library),
            (3, "素材库", "quick_hotkey_material_library", self.quick_show_material_library),
        ]
        failures = []
        hwnd = int(self.winId())
        for hotkey_id, label, setting_key, callback in hotkeys:
            sequence_text = self.db.get_setting(setting_key, DEFAULT_GLOBAL_HOTKEYS[setting_key])
            sequence_text = str(sequence_text or "").strip()
            if not sequence_text:
                continue
            parsed = self._parse_global_hotkey(sequence_text)
            if not parsed:
                failures.append(f"{label}: {sequence_text}")
                continue
            modifiers, virtual_key = parsed
            ctypes.set_last_error(0)
            if not _USER32.RegisterHotKey(hwnd, hotkey_id, modifiers, virtual_key):
                error_code = ctypes.get_last_error()
                failures.append(f"{label}: {sequence_text}（Windows错误 {error_code}）")
                continue
            self._registered_global_hotkeys.add(hotkey_id)
            self._global_hotkey_actions[hotkey_id] = callback

        self._global_hotkey_hwnd = hwnd if self._registered_global_hotkeys else None
        append_event(
            f"hotkeys:registered={len(self._registered_global_hotkeys)} "
            f"failed={len(failures)} details={' | '.join(failures) if failures else 'none'}"
        )
        if show_message and failures:
            QMessageBox.warning(self, "快捷键注册失败", "\n".join(failures))
        return failures

    def unregister_global_hotkeys(self):
        if _USER32 is None:
            self._registered_global_hotkeys.clear()
            self._global_hotkey_actions.clear()
            return
        try:
            hwnd = getattr(self, "_global_hotkey_hwnd", None)
            for hotkey_id in list(getattr(self, "_registered_global_hotkeys", set())):
                if hwnd:
                    _USER32.UnregisterHotKey(hwnd, hotkey_id)
        except Exception as e:
            print(f"注销全局快捷键失败: {e}")
        self._registered_global_hotkeys.clear()
        self._global_hotkey_actions.clear()
        self._global_hotkey_hwnd = None

    def _parse_global_hotkey(self, sequence_text):
        try:
            sequence = QKeySequence(sequence_text)
            if sequence.isEmpty():
                return None
            combined = int(sequence[0])
            modifier_mask = int(Qt.ShiftModifier) | int(Qt.ControlModifier) | int(Qt.AltModifier) | int(Qt.MetaModifier)
            key = combined & ~modifier_mask
            if key <= 0:
                return None
            modifiers = 0x4000  # MOD_NOREPEAT
            if combined & int(Qt.ControlModifier):
                modifiers |= 0x0002
            if combined & int(Qt.AltModifier):
                modifiers |= 0x0001
            if combined & int(Qt.ShiftModifier):
                modifiers |= 0x0004
            if combined & int(Qt.MetaModifier):
                modifiers |= 0x0008
            virtual_key = self._qt_key_to_vk(key)
            if not virtual_key:
                return None
            return modifiers, virtual_key
        except Exception as e:
            print(f"解析全局快捷键失败 {sequence_text}: {e}")
            return None

    def _qt_key_to_vk(self, key):
        if int(Qt.Key_A) <= key <= int(Qt.Key_Z):
            return key
        if int(Qt.Key_0) <= key <= int(Qt.Key_9):
            return key
        if int(Qt.Key_F1) <= key <= int(Qt.Key_F24):
            return 0x70 + (key - int(Qt.Key_F1))
        special_keys = {
            int(Qt.Key_Escape): 0x1B,
            int(Qt.Key_Tab): 0x09,
            int(Qt.Key_Backspace): 0x08,
            int(Qt.Key_Return): 0x0D,
            int(Qt.Key_Enter): 0x0D,
            int(Qt.Key_Insert): 0x2D,
            int(Qt.Key_Delete): 0x2E,
            int(Qt.Key_Home): 0x24,
            int(Qt.Key_End): 0x23,
            int(Qt.Key_PageUp): 0x21,
            int(Qt.Key_PageDown): 0x22,
            int(Qt.Key_Left): 0x25,
            int(Qt.Key_Up): 0x26,
            int(Qt.Key_Right): 0x27,
            int(Qt.Key_Down): 0x28,
            int(Qt.Key_Space): 0x20,
        }
        return special_keys.get(int(key))

    def nativeEvent(self, eventType, message):
        if sys.platform == "win32":
            try:
                msg = wintypes.MSG.from_address(int(message))
                if msg.message == 0x0312:  # WM_HOTKEY
                    hotkey_id = int(msg.wParam)
                    callback = self._global_hotkey_actions.get(hotkey_id)
                    if callback:
                        now = time.monotonic()
                        last = self._global_hotkey_last_trigger.get(hotkey_id, 0)
                        if now - last < 0.35:
                            return True, 0
                        self._global_hotkey_last_trigger[hotkey_id] = now
                        QTimer.singleShot(0, callback)
                        return True, 0
            except Exception as e:
                print(f"处理全局快捷键失败: {e}")
        return super().nativeEvent(eventType, message)

    def _raise_window_temporarily_on_top(self, window):
        if window is None:
            return
        try:
            if window.isMinimized():
                window.showNormal()
            else:
                window.show()
            window.raise_()
            window.activateWindow()
            if sys.platform == "win32":
                try:
                    hwnd = int(window.winId())
                    user32 = ctypes.windll.user32
                    user32.SetWindowPos(hwnd, -1, 0, 0, 0, 0, 0x0001 | 0x0002 | 0x0010 | 0x0040)
                    user32.SetWindowPos(hwnd, -2, 0, 0, 0, 0, 0x0001 | 0x0002 | 0x0010 | 0x0040)
                except Exception:
                    pass
        except Exception as e:
            print(f"临时置顶窗口失败: {e}")

    def _restore_window_flags(self, window, flags):
        try:
            if window is None:
                return
            window.setWindowFlags(flags)
            window.show()
            window.raise_()
        except RuntimeError:
            pass
        except Exception as e:
            print(f"恢复窗口置顶状态失败: {e}")

    def _is_toggle_hide_target(self, window):
        if window is None or not window.isVisible() or window.isMinimized():
            return False
        if sys.platform == "win32":
            try:
                return ctypes.windll.user32.GetForegroundWindow() == int(window.winId())
            except Exception:
                pass
        if window.isActiveWindow():
            return True
        active = QApplication.activeWindow()
        while active is not None:
            if active is window:
                return True
            active = active.parentWidget()
        return False

    def _toggle_window_visibility(self, window_getter, show_callback):
        try:
            window = window_getter()
            if self._is_toggle_hide_target(window):
                window.hide()
                return
            show_callback()
            window = window_getter()
            if window is not None:
                self._raise_window_temporarily_on_top(window)
        except RuntimeError:
            show_callback()
            window = window_getter()
            if window is not None:
                self._raise_window_temporarily_on_top(window)
        except Exception as e:
            self._log_runtime_exception("toggle_window_visibility", e)

    def quick_show_main_window(self):
        self._toggle_window_visibility(lambda: self, self.show_window)

    def quick_show_cost_library(self):
        self._toggle_window_visibility(lambda: getattr(self, "cost_library_dialog", None), self.show_cost_library)

    def quick_show_material_library(self):
        self._toggle_window_visibility(lambda: getattr(self, "material_library_dialog", None), self.show_material_library)

    def focus_search(self):
        """聚焦搜索框，全选文本"""
        self.search_input.setFocus()
        self.search_input.selectAll()

    def ensure_archive_account_allowed(self, target_account):
        """确保本地表格数据只能保存到其归属账号。"""
        if not self.archive_manager or not target_account:
            return False

        active_account = self.archive_manager.get_active_data_account()
        if active_account:
            if active_account.get('id') == target_account.get('id'):
                return True
            QMessageBox.warning(
                self,
                "账号不一致，已取消保存",
                f"当前表格数据属于账号：{active_account.get('name', '未知')}\n"
                f"你当前选择的存档账号：{target_account.get('name', '未知')}\n\n"
                f"为避免覆盖错误存档，本次保存已取消。\n"
                f"请先切换到正确账号，或读取对应账号的数据后再保存。"
            )
            return False

        reply = QMessageBox.question(
            self,
            "确认数据归属账号",
            f"当前本地表格数据还没有绑定存档账号。\n\n"
            f"是否确认这份本地数据属于账号：{target_account.get('name', '未知')}？\n\n"
            f"确认后以后只能保存到这个账号，避免误覆盖其他账号存档。",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return False
        self.archive_manager.set_active_data_account(target_account['id'])
        self.update_archive_account_label()
        return True

    def quick_save_archive(self):
        """快速保存到本地存档。"""
        if not self.archive_manager:
            self.show_toast("❌ 存档未初始化", 1000)
            return

        current = self.archive_manager.get_current_account()
        if not current:
            self.show_toast("⚠️ 请先选择存档账号", 1000)
            return
        if not self.archive_manager.get_data_root():
            self.show_toast("⚠️ 请先在存档窗口设置存档母文件夹", 1800)
            return
        if not self.ensure_archive_account_allowed(current):
            return

        progress_dialog = ArchiveProgressDialog(self)
        progress_dialog.setWindowTitle("保存存档")
        progress_dialog.set_status("💾 正在保存到存档...", 20)
        progress_dialog.show()
        QApplication.processEvents()

        try:
            success, result = self.archive_manager.save_local_profile(current['id'])
            if success:
                self.archive_manager.update_last_save_time(current['id'])
                progress_dialog.set_status("✅ 保存完成", 100)
                QApplication.processEvents()
                QTimer.singleShot(500, progress_dialog.close)
                QTimer.singleShot(1000, lambda p=result: self.show_toast(f"✅ 已保存到存档：{p}", 1800))
            else:
                progress_dialog.set_error(f"❌ 保存失败: {result}")
                progress_dialog.progress_bar.setValue(100)
                QTimer.singleShot(1500, progress_dialog.close)
                QTimer.singleShot(2000, lambda r=result: self.show_toast(f"❌ 保存失败: {r}", 1000))

        except Exception as e:
            print(f"快速保存存档失败: {e}")
            progress_dialog.set_error(f"❌ 保存异常")
            progress_dialog.progress_bar.setValue(100)
            QTimer.singleShot(1500, progress_dialog.close)
            QTimer.singleShot(2000, lambda: self.show_toast(f"❌ 保存异常", 1000))

    def autosave_current_archive(self):
        """Commit the current DB and keep it bound to the active archive account."""
        try:
            if hasattr(self.db, "conn"):
                self.db.conn.commit()
        except Exception as e:
            print(f"实时提交存档失败: {e}")

        if not self.archive_manager or not self.archive_manager.get_data_root():
            return True, getattr(self.db, "db_path", "")

        account = self.archive_manager.get_active_data_account() or self.archive_manager.get_current_account()
        if not account or not account.get("id"):
            return False, "当前数据没有绑定存档账号"

        if not self.archive_manager.get_active_data_account():
            self.archive_manager.set_active_data_account(account["id"])

        ok, result = self.archive_manager.save_local_profile(account["id"])
        if ok:
            return True, result
        return False, result

    def _close_account_scoped_windows(self):
        """关闭会持有旧数据库或旧店铺 ID 的窗口。"""
        try:
            append_event("account_switch:close_windows:start")
            for dialog in list(getattr(self, "store_margin_dialogs", {}).values()):
                self._dispose_account_window(dialog)
            self.store_margin_dialogs.clear()

            for dialog in list(getattr(self, "promotion_data_dialogs", {}).values()):
                self._dispose_account_window(dialog)
            self.promotion_data_dialogs.clear()

            for cache_name in (
                "pdd_code_fetch_dialogs",
                "pdd_price_fetch_dialogs",
                "pdd_promotion_status_fetch_dialogs",
            ):
                dialogs = getattr(self, cache_name, None)
                if isinstance(dialogs, dict):
                    for dialog in list(dialogs.values()):
                        self._dispose_account_window(dialog)
                    dialogs.clear()

            for dialog in list(getattr(self, "record_dialogs", [])):
                self._dispose_account_window(dialog)
            self.record_dialogs = []

            product_spec_dialog = getattr(self, "product_spec_dialog", None)
            if product_spec_dialog is not None:
                self._dispose_account_window(product_spec_dialog)
                self.product_spec_dialog = None

            daily_dialog = getattr(self, "daily_task_dialog", None)
            if daily_dialog is not None:
                self._dispose_account_window(daily_dialog)
                self.daily_task_dialog = None

            for attr_name in ("cost_library_dialog", "material_library_dialog"):
                dialog = getattr(self, attr_name, None)
                if dialog is not None:
                    self._dispose_account_window(dialog)
                    setattr(self, attr_name, None)
            app = QApplication.instance()
            if app is not None:
                app.processEvents()
            append_event("account_switch:close_windows:done")
        except Exception as e:
            print(f"关闭账号相关窗口失败: {e}")

    def _dispose_account_window(self, dialog):
        if dialog is None:
            return
        try:
            if sip.isdeleted(dialog):
                return
        except Exception:
            return
        try:
            if hasattr(dialog, "_prepare_for_account_switch"):
                dialog._prepare_for_account_switch()
            dialog.setAttribute(Qt.WA_DeleteOnClose, True)
            dialog.close()
            dialog.deleteLater()
            QApplication.removePostedEvents(dialog)
        except Exception:
            pass

    def _close_store_scoped_windows(self, store_id, product_ids=()):
        margin_dialog = self.store_margin_dialogs.pop(store_id, None)
        if margin_dialog is not None:
            self._dispose_account_window(getattr(margin_dialog, "promotion_data_dialog", None))
            self._dispose_account_window(margin_dialog)
        self._dispose_account_window(getattr(self, "promotion_data_dialogs", {}).pop(store_id, None))

        remaining_records = []
        for dialog in self.record_dialogs:
            if getattr(dialog, "store_id", None) == store_id:
                self._dispose_account_window(dialog)
            else:
                remaining_records.append(dialog)
        self.record_dialogs = remaining_records

        for cache_name in (
            "pdd_code_fetch_dialogs",
            "pdd_price_fetch_dialogs",
            "pdd_promotion_status_fetch_dialogs",
        ):
            dialogs = getattr(self, cache_name, None)
            if isinstance(dialogs, dict):
                self._dispose_account_window(dialogs.pop(store_id, None))

        product_spec_dialog = getattr(self, "product_spec_dialog", None)
        if getattr(product_spec_dialog, "product_id", None) in set(product_ids):
            self._dispose_account_window(product_spec_dialog)
            self.product_spec_dialog = None

    def _clear_account_scoped_ui_state(self):
        """清理换账号后不能跨库复用的行映射、筛选和搜索状态。"""
        try:
            for timer_name in ("_scroll_save_timer", "search_timer", "_record_tooltip_timer"):
                timer = getattr(self, timer_name, None)
                if timer is not None and timer.isActive():
                    timer.stop()

            if hasattr(self, "tag_filter_menu") and self.tag_filter_menu:
                self.tag_filter_menu.close()
            if hasattr(self, "store_filter_menu") and self.store_filter_menu:
                self.store_filter_menu.close()

            self.current_search_match_ids = None
            self._search_highlighted_rows = set()
            self.current_category_filter = ""
            self.current_store_filter = set()
            self.current_filter_tags = set()
            self._active_reminder_ids.clear()
            self._task_reminder_popup_active = False

            if hasattr(self, "search_input"):
                self.search_input.blockSignals(True)
                self.search_input.clear()
                self.search_input.blockSignals(False)
            if hasattr(self, "category_filter_input"):
                self.category_filter_input.blockSignals(True)
                self.category_filter_input.clear()
                self.category_filter_input.blockSignals(False)

            for btn_name in (
                "btn_filter_coupon", "btn_filter_new_customer", "btn_filter_limited_time",
                "btn_filter_marketing", "btn_filter_natural_flow", "btn_filter_sitewide",
                "btn_filter_garbage", "btn_filter_waste",
                "btn_filter_profit", "btn_filter_loss", "btn_filter_break_even",
                "btn_filter_missing_roi_bid",
            ):
                btn = getattr(self, btn_name, None)
                if btn is not None:
                    btn.blockSignals(True)
                    btn.setChecked(False)
                    btn.blockSignals(False)

            if hasattr(self, "btn_tag_filter"):
                self.btn_tag_filter.setText("🏷️ 筛选")
            if hasattr(self, "btn_store_filter"):
                self.btn_store_filter.setText("🏪 店铺")

            self.row_data_map.clear()
            self.row_store_map.clear()
            self.product_store_map.clear()
            self.visible_product_ids.clear()
            if hasattr(self, "table"):
                self.table.clearSelection()
                self.table.setRowCount(0)
            if hasattr(self, "frozen_table"):
                self.frozen_table.clearSelection()
                self.frozen_table.setRowCount(0)
            if hasattr(self, "data_mode_layout"):
                self._clear_layout_widgets(self.data_mode_layout)
                self.bubble_product_widgets = {}
        except Exception as e:
            print(f"清理账号界面状态失败: {e}")

    def _refresh_account_bound_controls_from_db(self):
        """从当前数据库恢复会随账号变化的控件状态。"""
        try:
            self.product_sort_mode = self.db.get_setting("product_sort_mode", "created_at") or "created_at"
            self.product_sort_descending = self.db.get_setting("product_sort_descending", "1") != "0"
            if hasattr(self, "product_sort_combo"):
                idx = self.product_sort_combo.findData(self.product_sort_mode)
                if idx >= 0:
                    self.product_sort_combo.blockSignals(True)
                    self.product_sort_combo.setCurrentIndex(idx)
                    self.product_sort_combo.blockSignals(False)
            if hasattr(self, "btn_product_sort_direction"):
                self._update_product_sort_direction_button()

            if hasattr(self, "btn_real_promotion_mode"):
                self.btn_real_promotion_mode.blockSignals(True)
                self.btn_real_promotion_mode.setChecked(self.db.get_setting("real_promotion_data_mode", "0") == "1")
                self.btn_real_promotion_mode.blockSignals(False)
                self._update_real_promotion_mode_button_style()

            self.main_view_mode = "data"
            self._apply_main_view_mode(refresh=False)
        except Exception as e:
            print(f"刷新账号控件状态失败: {e}")

    @staticmethod
    def _update_account_switch_progress(progress, value, text):
        if progress is None:
            return
        progress.setLabelText(text)
        progress.setValue(value)
        progress.repaint()

    def replace_database_from_local_profile(self, profile_path, account_id, progress=None):
        """用本地账号档案替换当前主库，并刷新界面。"""
        append_event(f"account_switch:start account_id={account_id}")
        if not os.path.exists(profile_path):
            return False, "本地账号数据文件不存在"

        profile_path = os.path.abspath(profile_path)
        db_path = os.path.abspath(self.db.db_path)
        temp_path = f"{db_path}.switching"

        if profile_path == db_path:
            return True, db_path

        try:
            self.stop_cost_sync_service()
            self._update_account_switch_progress(progress, 45, "正在关闭当前账号窗口...")
            self._close_account_scoped_windows()
            self._clear_account_scoped_ui_state()

            try:
                self.db.conn.commit()
            except Exception:
                pass
            try:
                self.db.conn.close()
            except Exception:
                pass

            if self.data_root_manager.get_data_root():
                self._update_account_switch_progress(progress, 55, "正在打开目标账号数据库...")
                self.db = SafeDatabaseManager(profile_path)
                self.db.init_default_prompts()
                if self.archive_manager:
                    self.archive_manager.db = self.db
                    self.archive_manager.switch_account(account_id)
                    self.archive_manager.set_active_data_account(account_id)
                self._refresh_account_bound_controls_from_db()
                append_event("account_switch:before_load_data")
                self._update_account_switch_progress(progress, 65, "正在读取店铺和链接...")
                self.load_data_safe(restore_position=False)
                self.update_archive_account_label()
                if self.main_view_mode == "data":
                    self._update_account_switch_progress(progress, 80, "正在生成主界面气泡...")
                    self._refresh_data_mode_view()
                self.restart_cost_sync_service()
                self._update_account_switch_progress(progress, 95, "正在恢复账号界面状态...")
                append_event("account_switch:done direct_profile")
                return True, profile_path

            self._update_account_switch_progress(progress, 55, "正在复制目标账号数据库...")
            if os.path.exists(temp_path):
                os.remove(temp_path)
            shutil.copy2(profile_path, temp_path)
            os.replace(temp_path, db_path)

            self.db = SafeDatabaseManager()
            self.db.init_default_prompts()
            if self.archive_manager:
                self.archive_manager.db = self.db
                self.archive_manager.switch_account(account_id)
            self._refresh_account_bound_controls_from_db()
            append_event("account_switch:before_load_data")
            self._update_account_switch_progress(progress, 65, "正在读取店铺和链接...")
            self.load_data_safe(restore_position=False)
            if self.archive_manager:
                self.archive_manager.set_active_data_account(account_id)
            self.update_archive_account_label()
            if self.main_view_mode == "data":
                self._update_account_switch_progress(progress, 80, "正在生成主界面气泡...")
                self._refresh_data_mode_view()
            self.restart_cost_sync_service()
            self._update_account_switch_progress(progress, 95, "正在恢复账号界面状态...")
            append_event("account_switch:done copied_profile")
            return True, db_path
        except Exception as e:
            append_exception("account_switch:failed", error=e)
            try:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            except Exception:
                pass
            try:
                self.db = SafeDatabaseManager()
                if self.archive_manager:
                    self.archive_manager.db = self.db
                self.restart_cost_sync_service()
            except Exception:
                pass
            return False, str(e)

    def _resolve_next_local_profile_account(self):
        """按当前应用账号自动推断下一个要切换的本地账号。"""
        if not self.archive_manager:
            return None, None, "存档管理器未初始化"

        available_profiles = self.archive_manager.get_accounts_with_local_profiles()
        if not available_profiles:
            return None, None, "没有找到任何本地账号数据。"

        active_account = self.archive_manager.get_active_data_account()
        active_id = active_account.get('id') if active_account else None
        candidates = [
            (acc, path)
            for acc, path in available_profiles
            if acc.get('id') != active_id
        ]
        if not candidates:
            return None, None, "没有其他可切换的本地账号数据。"

        if len(candidates) == 1:
            target_account, _ = candidates[0]
            return target_account, target_account.get('id'), None

        current = self.archive_manager.get_current_account()
        if current:
            for acc, _ in candidates:
                if acc.get('id') == current.get('id'):
                    return acc, acc.get('id'), None

        target_account, _ = candidates[0]
        return target_account, target_account.get('id'), None

    def show_local_account_switch_menu(self):
        """显示可切换的本地账号列表，点击账号后自动切换。"""
        if self._is_switching_local_account:
            return
        if not self.archive_manager:
            QMessageBox.warning(self, "提示", "存档管理器未初始化")
            return

        available_profiles = self.archive_manager.get_accounts_with_local_profiles()
        if not available_profiles:
            QMessageBox.warning(self, "提示", "没有找到任何本地账号数据。")
            return

        active_account = self.archive_manager.get_active_data_account()
        active_id = active_account.get('id') if active_account else None
        menu = QMenu(self)
        action_accounts = {}
        for account, _profile_path in available_profiles:
            account_id = account.get('id')
            account_name = account.get('name', '未知')
            action_text = f"✓ {account_name}" if account_id == active_id else account_name
            action = QAction(action_text, menu)
            action.setEnabled(account_id != active_id)
            action_accounts[action] = (account, account_id)
            menu.addAction(action)

        if menu.isEmpty():
            QMessageBox.warning(self, "提示", "没有可切换的本地账号数据。")
            return

        action = menu.exec_(self.btn_switch_local_account.mapToGlobal(self.btn_switch_local_account.rect().bottomLeft()))
        if action in action_accounts:
            account, account_id = action_accounts[action]
            self.switch_local_account(account, account_id)

    def switch_local_account(self, target_account=None, target_id=None):
        """切换到指定本地账号；未指定时兼容旧逻辑切到下一个账号。"""
        if self._is_switching_local_account:
            return

        progress = None
        self._is_switching_local_account = True
        if hasattr(self, 'btn_switch_local_account'):
            self.btn_switch_local_account.setEnabled(False)

        try:
            if not self.archive_manager:
                QMessageBox.warning(self, "提示", "存档管理器未初始化")
                return

            if target_account is None or target_id is None:
                target_account, target_id, error = self._resolve_next_local_profile_account()
                if error:
                    QMessageBox.warning(self, "提示", error)
                    return

            progress = QProgressDialog("正在检查账号数据...", None, 0, 100, self)
            progress.setWindowTitle("切换账号")
            progress.setWindowModality(Qt.WindowModal)
            progress.setMinimumDuration(0)
            progress.setAutoClose(False)
            progress.setAutoReset(False)
            progress.setValue(5)
            progress.show()
            progress.repaint()

            profile_ok, profile_path = self.archive_manager.load_local_profile(target_id)
            if not profile_ok:
                QMessageBox.warning(
                    self,
                    "提示",
                    f"账号「{target_account.get('name', '未知')}」暂无本地数据。\n请先保存该账号本地数据，或读取该账号存档。"
                )
                return

            self._update_account_switch_progress(progress, 15, "正在准备目标账号数据...")
            normalize_ok, normalized_path = self.archive_manager.ensure_local_profile_normalized(target_id, profile_path)
            if not normalize_ok:
                QMessageBox.critical(self, "错误", f"本地账号数据迁移失败：{normalized_path}")
                return
            profile_path = normalized_path

            active_account = self.archive_manager.get_active_data_account()
            if not active_account:
                current = self.archive_manager.get_current_account()
                if not current:
                    QMessageBox.warning(self, "请先绑定当前数据", "当前主表格数据还没有绑定本地账号，请先在存档中选择当前数据所属账号。")
                    return

                if not self.archive_manager.set_active_data_account(current['id']):
                    QMessageBox.critical(self, "错误", "绑定当前应用账号失败")
                    return
                active_account = current
                self.update_archive_account_label()

            if active_account.get('id') == target_id:
                QMessageBox.information(self, "提示", f"当前主表格已经是账号「{target_account.get('name', '未知')}」的数据。")
                return

            self._update_account_switch_progress(progress, 30, "正在保存当前账号数据...")
            save_ok, save_result = self.archive_manager.save_local_profile(active_account['id'])
            if not save_ok:
                QMessageBox.critical(self, "切换已取消", f"当前应用账号数据保存失败，已取消切换：\n{save_result}")
                return

            self._update_account_switch_progress(progress, 40, "正在切换账号...")
            ok, result = self.replace_database_from_local_profile(profile_path, target_id, progress)
            if ok:
                self._update_account_switch_progress(progress, 100, "切换完成")
                append_event("account_switch:return_to_handler")
            else:
                QMessageBox.critical(self, "错误", f"应用本地账号失败：{result}")
        finally:
            if progress is not None:
                progress.close()
            self._is_switching_local_account = False
            if hasattr(self, 'btn_switch_local_account'):
                self.btn_switch_local_account.setEnabled(True)
            append_event("account_switch:handler_done")

    def auto_save_archive_on_exit(self):
        """退出软件时自动保存当前数据到本地存档。"""
        if not self.archive_manager:
            return True

        current = self.archive_manager.get_active_data_account() or self.archive_manager.get_current_account()
        if not current:
            return True
        if not self.archive_manager.get_data_root():
            reply = QMessageBox.question(
                self,
                "未设置存档母文件夹",
                "当前还没有设置存档母文件夹，无法保存到统一本地存档。\n\n是否直接退出？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            return reply == QMessageBox.Yes

        updating = bool(getattr(self, "_update_exe_to_launch", None))
        progress_dialog = None
        if updating:
            self._set_update_progress(84, "正在保存当前数据，请勿操作...")
        else:
            progress_dialog = ArchiveProgressDialog(self)
            progress_dialog.setWindowTitle("退出前保存存档")
            progress_dialog.set_status("💾 正在保存到存档...", 20)
            progress_dialog.show()
            QApplication.processEvents()

        try:
            success, result = self.autosave_current_archive()
            if success:
                self.archive_manager.update_last_save_time(current['id'])
                if updating:
                    self._set_update_progress(87, "数据保存完成，正在关闭旧版本...")
                else:
                    progress_dialog.set_status("✅ 保存完成，正在退出...", 100)
                    QApplication.processEvents()
                    QTimer.singleShot(300, progress_dialog.close)
                return True

            if progress_dialog:
                progress_dialog.set_error(f"❌ 保存失败: {result}")
            QApplication.processEvents()
            QMessageBox.warning(self, "退出前保存失败", f"保存到存档失败：\n{result}")
            return False
        except Exception as e:
            print(f"退出自动保存失败: {e}")
            if progress_dialog:
                progress_dialog.set_error("❌ 保存异常")
            QApplication.processEvents()
            QMessageBox.warning(self, "退出前保存失败", f"保存异常：\n{str(e)}")
            return False
        finally:
            if progress_dialog:
                progress_dialog.close()

    def quit_application(self):
        """退出应用"""
        if self.update_download_in_progress:
            self.show_toast("更新正在下载，请等待下载完成", 1800)
            return
        if self._is_quitting:
            return
        self._is_quitting = True
        self._cancel_button_hint()
        self._cancel_record_tooltip()
        if not self.auto_save_archive_on_exit():
            self._is_quitting = False
            return
        update_exe = getattr(self, "_update_exe_to_launch", None)
        if update_exe:
            try:
                self._set_update_progress(88, "旧版本即将关闭，正在启动新版本...")
                launcher_env = os.environ.copy()
                launcher_env.update({
                    "SHOP_UPDATE_OLD_PID": str(os.getpid()),
                    "SHOP_UPDATE_EXE": update_exe,
                    "SHOP_UPDATE_CWD": os.path.dirname(update_exe),
                })
                launcher_script = os.path.join(
                    tempfile.gettempdir(), f"shop_manager_update_{os.getpid()}.ps1"
                )
                launcher_script_text = r"""
$ErrorActionPreference = 'Stop'
Add-Type -AssemblyName System.Windows.Forms
Add-Type -AssemblyName System.Drawing

$form = New-Object System.Windows.Forms.Form
$form.Text = '正在更新'
$form.Width = 440
$form.Height = 175
$form.StartPosition = 'CenterScreen'
$form.TopMost = $true
$form.ControlBox = $false

$label = New-Object System.Windows.Forms.Label
$label.Left = 28
$label.Top = 20
$label.Width = 375
$label.Height = 25

$bar = New-Object System.Windows.Forms.ProgressBar
$bar.Left = 28
$bar.Top = 52
$bar.Width = 375
$bar.Height = 22
$bar.Minimum = 0
$bar.Maximum = 100

$percent = New-Object System.Windows.Forms.Label
$percent.Left = 28
$percent.Top = 82
$percent.Width = 375
$percent.Height = 22
$percent.TextAlign = 'MiddleCenter'

$form.Controls.AddRange(@($label, $bar, $percent))

function Set-Stage([string]$text, [int]$value) {
    $value = [Math]::Max(0, [Math]::Min(100, $value))
    $label.Text = $text
    $bar.Value = $value
    $percent.Text = "总进度：$value%"
    [System.Windows.Forms.Application]::DoEvents()
}

try {
    Set-Stage '正在等待旧版本完全关闭，请勿操作...' 88
    $form.Show()
    [System.Windows.Forms.Application]::DoEvents()
    Wait-Process -Id $env:SHOP_UPDATE_OLD_PID -ErrorAction SilentlyContinue

    Set-Stage '旧版本已关闭，正在启动新版本...' 92
    Start-Process -FilePath $env:SHOP_UPDATE_EXE -WorkingDirectory $env:SHOP_UPDATE_CWD

    $started = Get-Date
    $limit = $started.AddSeconds(120)
    while ((Get-Date) -lt $limit) {
        $window = Get-Process -ErrorAction SilentlyContinue |
            Where-Object { $_.MainWindowTitle -like '电商店铺操作记录管理工具 v*' } |
            Select-Object -First 1
        if ($window) {
            Set-Stage '新版本已打开，更新完成。' 100
            Start-Sleep -Milliseconds 600
            break
        }
        $elapsed = ((Get-Date) - $started).TotalSeconds
        $value = [Math]::Min(99, 94 + [int]($elapsed / 20))
        Set-Stage '新版程序正在加载，请勿操作...' $value
        Start-Sleep -Milliseconds 200
    }
}
catch {
    [System.Windows.Forms.MessageBox]::Show("新版启动失败：`n$($_.Exception.Message)", '更新失败') | Out-Null
}
finally {
    $form.Close()
    Remove-Item -LiteralPath $PSCommandPath -Force -ErrorAction SilentlyContinue
}
"""
                with open(launcher_script, "w", encoding="utf-8-sig", newline="\r\n") as script_file:
                    script_file.write(launcher_script_text)
                subprocess.Popen(
                    [
                        "powershell.exe", "-NoProfile", "-Sta", "-ExecutionPolicy", "Bypass",
                        "-WindowStyle", "Hidden", "-File", launcher_script,
                    ],
                    creationflags=CREATE_NO_WINDOW,
                    env=launcher_env,
                )
            except Exception as e:
                self._is_quitting = False
                self._update_exe_to_launch = None
                QMessageBox.warning(self, "启动失败", f"无法打开新版本：\n{e}")
                return
        self.unregister_global_hotkeys()
        if self.update_listener is not None and hasattr(self.update_listener, "stop"):
            self.update_listener.stop()
            self.update_listener = None
        if getattr(self, "update_publish_service", None) is not None:
            self.update_publish_service.stop()
        if str(load_global_update_settings().get("update_publish_enabled", "0")) == "1":
            start_update_agent_task()
        self._cleanup_before_quit()
        if self.tray_icon:
            self.tray_icon.hide()
        QApplication.quit()

    def _cleanup_before_quit(self):
        self.stop_cost_sync_service()
        service = getattr(self, "material_mobile_service", None)
        if service is not None:
            service.stop()
            self.material_mobile_service = None
        monitor = getattr(self, "pdd_browser_monitor", None)
        if monitor is not None and hasattr(monitor, "stop"):
            monitor.stop()
            self.pdd_browser_monitor = None
        if getattr(self, "db", None) is not None and hasattr(self.db, "close"):
            self.db.close()
    
    def closeEvent(self, event):
        """关闭窗口时最小化到托盘；托盘菜单“退出”才真正结束进程。"""
        self._cancel_button_hint()
        self._cancel_record_tooltip()
        if getattr(self, "_is_quitting", False):
            self._cleanup_before_quit()
            event.accept()
            return
        if self.tray_icon and self.tray_icon.isVisible():
            self.hide()
            self.tray_icon.showMessage(
                "电商店铺操作记录管理工具",
                "程序已最小化到系统托盘，双击图标可显示窗口",
                QSystemTrayIcon.Information,
                2000
            )
            event.ignore()
        else:
            event.accept()
    
    def changeEvent(self, event):
        """窗口状态改变事件"""
        if event.type() == QEvent.WindowStateChange:
            if self.windowState() & Qt.WindowMinimized:
                self._cancel_button_hint()
                self._cancel_record_tooltip()
                # 最小化到任务栏（默认行为，不做额外处理）
                pass
        elif event.type() == QEvent.ActivationChange and self.isActiveWindow():
            QTimer.singleShot(100, self.check_local_update_package_on_startup)
            QTimer.singleShot(100, self.show_pending_update_notification)
        super().changeEvent(event)

    def center_on_screen(self):
        screen = QApplication.screenAt(QCursor.pos()) or QApplication.primaryScreen()
        if not screen:
            return
        screen_rect = screen.availableGeometry()
        frame = self.frameGeometry()
        frame.moveCenter(screen_rect.center())
        frame.moveTop(max(screen_rect.top(), frame.top() - 40))
        self.move(frame.topLeft())

    def init_ui(self):
        self.setWindowTitle(MAIN_WINDOW_TITLE)
        self.resize(1430, 1000)
        self.center_on_screen()

        self.resource_timer = None

        toolbar = QHBoxLayout()
        secondary_toolbar = QHBoxLayout()
        btn_prev = QPushButton("◀ 上个月")
        btn_prev.setObjectName("tutorial_prev_month")
        btn_prev.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 1px 12px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        btn_prev.clicked.connect(self.prev_month)
        self.lbl_month = QLabel(f"{self.year}年 {self.month}月")
        self.lbl_month.setFont(QFont("Arial", 14, QFont.Bold))
        btn_next = QPushButton("下个月 ▶")
        btn_next.setObjectName("tutorial_next_month")
        btn_next.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 1px 12px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        btn_next.clicked.connect(self.next_month)
        
        search_layout = QHBoxLayout()
        search_label = QLabel("搜索:")
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("输入商品ID、标题或备注...")
        self.search_timer = QTimer(self)
        self.search_timer.setSingleShot(True)
        self.search_timer.timeout.connect(self.apply_realtime_search)
        self.search_input.textChanged.connect(lambda: self.search_timer.start(200))
        
        self.btn_tag_filter = QPushButton("🏷️ 筛选")
        self.btn_tag_filter.setFixedWidth(80)
        self.btn_tag_filter.setStyleSheet("""
            QPushButton {
                border: 1px solid #3498db;
                background-color: transparent;
                color: #3498db;
                border-radius: 3px;
                padding: 4px 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #3498db;
                color: white;
            }
        """)
        self.btn_tag_filter.clicked.connect(self.show_tag_filter_menu)

        self.btn_store_filter = QPushButton("🏪 店铺")
        self.btn_store_filter.setFixedWidth(80)
        self.btn_store_filter.setStyleSheet("""
            QPushButton {
                border: 1px solid #27ae60;
                background-color: transparent;
                color: #27ae60;
                border-radius: 3px;
                padding: 4px 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #27ae60;
                color: white;
            }
        """)
        self.btn_store_filter.clicked.connect(self.show_store_filter_menu)

        self.tag_filter_menu = QFrame(self, Qt.Popup | Qt.FramelessWindowHint)
        self.tag_filter_menu.hide()
        self.tag_filter_menu.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 5px;
            }
            QCheckBox {
                padding: 5px;
                spacing: 8px;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
            }
        """)
        
        filter_layout = QVBoxLayout(self.tag_filter_menu)
        filter_layout.setContentsMargins(10, 10, 10, 10)
        filter_layout.setSpacing(5)
        
        filter_title = QLabel("🏷️ 选择筛选标签")
        filter_title.setStyleSheet("font-weight: bold; font-size: 14px; color: #2c3e50; padding-bottom: 5px;")
        filter_layout.addWidget(filter_title)

        category_title = QLabel("商品类型筛选")
        category_title.setStyleSheet("font-weight: bold; font-size: 13px; color: #2c3e50;")
        filter_layout.addWidget(category_title)

        self.category_filter_input = QLineEdit()
        self.category_filter_input.setPlaceholderText("输入商品类型关键字...")
        self.category_filter_input.setAttribute(Qt.WA_InputMethodEnabled, True)
        self.category_filter_input.setInputMethodHints(Qt.ImhNone)
        self.category_filter_input.textChanged.connect(self.refresh_category_filter_candidates)
        filter_layout.addWidget(self.category_filter_input)

        self.category_candidate_widget = QWidget()
        self.category_candidate_layout = QVBoxLayout(self.category_candidate_widget)
        self.category_candidate_layout.setContentsMargins(0, 0, 0, 0)
        self.category_candidate_layout.setSpacing(4)
        filter_layout.addWidget(self.category_candidate_widget)

        self.btn_clear_category_filter = QPushButton("清除类型筛选")
        self.btn_clear_category_filter.setStyleSheet("""
            QPushButton {
                border: 1px solid #7f8c8d;
                background-color: transparent;
                color: #7f8c8d;
                border-radius: 4px;
                padding: 5px 10px;
                text-align: left;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
                color: white;
            }
        """)
        self.btn_clear_category_filter.clicked.connect(self.clear_category_filter)
        filter_layout.addWidget(self.btn_clear_category_filter)
        
        icons_dir = os.path.join(os.path.dirname(__file__), "icons")
        
        filter_layout.addWidget(QLabel("<hr>"))
        
        self.btn_filter_coupon = self._create_filter_button("优惠券", "coupon.svg", "#d81e06")
        self.btn_filter_coupon.setCheckable(True)
        self.btn_filter_new_customer = self._create_filter_button("新客立减", "new_customer.svg", "#9b59b6")
        self.btn_filter_new_customer.setCheckable(True)
        self.btn_filter_limited_time = self._create_filter_button("限时限量购", "limited-time.svg", "#e74c3c")
        self.btn_filter_limited_time.setCheckable(True)
        self.btn_filter_marketing = self._create_filter_button("营销活动", "marketing.svg", "#9b59b6")
        self.btn_filter_marketing.setCheckable(True)
        self.btn_filter_natural_flow = self._create_filter_button("无推广", None, "#16a085")
        self.btn_filter_natural_flow.setCheckable(True)
        self.btn_filter_sitewide = self._create_filter_button("全站托管", None, "#8e44ad")
        self.btn_filter_sitewide.setCheckable(True)

        self.btn_filter_garbage = self._create_filter_button("垃圾链接", None, "#dc2626")
        self.btn_filter_garbage.setCheckable(True)
        self.btn_filter_waste = self._create_filter_button("废物链接", None, "#7c2d12")
        self.btn_filter_waste.setCheckable(True)
        
        profit_label = QLabel("利润标签")
        profit_label.setStyleSheet("font-weight: bold; font-size: 13px; color: #2c3e50;")
        
        self.btn_filter_profit = self._create_filter_button("赚钱 (≥1%)", None, "#27ae60")
        self.btn_filter_profit.setCheckable(True)

        self.btn_filter_loss = self._create_filter_button("亏钱 (<-2%)", None, "#e74c3c")
        self.btn_filter_loss.setCheckable(True)

        self.btn_filter_break_even = self._create_filter_button("保本 (-2%~1%)", None, "#f39c12")
        self.btn_filter_break_even.setCheckable(True)

        self.btn_filter_missing_roi_bid = self._create_filter_button("未填投产/出价", None, "#8e44ad")
        self.btn_filter_missing_roi_bid.setCheckable(True)
        
        filter_layout.addWidget(self.btn_filter_coupon)
        filter_layout.addWidget(self.btn_filter_new_customer)
        filter_layout.addWidget(self.btn_filter_limited_time)
        filter_layout.addWidget(self.btn_filter_marketing)
        filter_layout.addWidget(self.btn_filter_natural_flow)
        filter_layout.addWidget(self.btn_filter_sitewide)

        task_label = QLabel("链接标签")
        task_label.setStyleSheet("font-weight: bold; font-size: 13px; color: #2c3e50;")
        filter_layout.addWidget(task_label)
        task_filter_layout = QHBoxLayout()
        task_filter_layout.setSpacing(6)
        task_filter_layout.addWidget(self.btn_filter_garbage)
        task_filter_layout.addWidget(self.btn_filter_waste)
        filter_layout.addLayout(task_filter_layout)

        filter_layout.addWidget(QLabel("<hr>"))
        filter_layout.addWidget(profit_label)
        filter_layout.addWidget(self.btn_filter_profit)
        filter_layout.addWidget(self.btn_filter_loss)
        filter_layout.addWidget(self.btn_filter_break_even)
        filter_layout.addWidget(self.btn_filter_missing_roi_bid)
        
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        
        btn_save_filter = QPushButton("💾 保存筛选")
        btn_save_filter.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #219a52;
            }
        """)
        # 保存按钮连接筛选并关闭窗口
        btn_save_filter.clicked.connect(lambda: self.apply_tag_filter(close_menu=True))
        
        btn_clear_filter = QPushButton("清空")
        btn_clear_filter.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        btn_clear_filter.clicked.connect(self.clear_tag_filter)
        
        btn_layout.addWidget(btn_save_filter)
        btn_layout.addWidget(btn_clear_filter)
        filter_layout.addLayout(btn_layout)
        
        self.current_filter_tags = set()
        
        # 筛选按钮连接实时筛选（不关闭窗口）
        self.btn_filter_coupon.toggled.connect(lambda: self.apply_tag_filter(close_menu=False))
        self.btn_filter_new_customer.toggled.connect(lambda: self.apply_tag_filter(close_menu=False))
        self.btn_filter_limited_time.toggled.connect(lambda: self.apply_tag_filter(close_menu=False))
        self.btn_filter_marketing.toggled.connect(lambda: self.apply_tag_filter(close_menu=False))
        self.btn_filter_natural_flow.toggled.connect(lambda: self.apply_tag_filter(close_menu=False))
        self.btn_filter_sitewide.toggled.connect(lambda: self.apply_tag_filter(close_menu=False))
        self.btn_filter_garbage.toggled.connect(lambda: self.apply_tag_filter(close_menu=False))
        self.btn_filter_waste.toggled.connect(lambda: self.apply_tag_filter(close_menu=False))
        self.btn_filter_profit.toggled.connect(lambda: self.apply_tag_filter(close_menu=False))
        self.btn_filter_loss.toggled.connect(lambda: self.apply_tag_filter(close_menu=False))
        self.btn_filter_break_even.toggled.connect(lambda: self.apply_tag_filter(close_menu=False))
        self.btn_filter_missing_roi_bid.toggled.connect(lambda: self.apply_tag_filter(close_menu=False))
        
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(self.btn_tag_filter)
        toolbar.addLayout(search_layout)
        
        icons_dir = os.path.join(os.path.dirname(__file__), "icons")
        
        btn_add_store = QPushButton("添加店铺")
        btn_add_store.setObjectName("tutorial_add_store")
        btn_add_store.setFixedSize(82, 30)
        btn_add_store.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 1px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        btn_add_store.clicked.connect(self.add_store)
        secondary_toolbar.addWidget(btn_add_store)
        
        btn_daily_task = QPushButton("📋 每日任务")
        btn_daily_task.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 4px 12px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        btn_daily_task.clicked.connect(self.show_daily_task_dialog)
        secondary_toolbar.addWidget(btn_daily_task)
        self.btn_daily_task = btn_daily_task
        self.daily_task_badge = QLabel("!", btn_daily_task)
        self.daily_task_badge.setAlignment(Qt.AlignCenter)
        self.daily_task_badge.setFixedSize(16, 16)
        self.daily_task_badge.setStyleSheet("background:#ff2d2d; color:white; border-radius:8px; font-weight:bold; font-size:12px;")
        self.daily_task_badge.hide()
        QTimer.singleShot(0, self.update_daily_task_button_badge)

        btn_export = QPushButton("📊批量导出")
        btn_export.setObjectName("tutorial_batch_export")
        btn_export.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 4px 12px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        btn_export.clicked.connect(self.batch_export_store_margin_excel)
        secondary_toolbar.addWidget(btn_export)

        self.store_sheet_scroll = QScrollArea()
        self.store_sheet_scroll.setWidgetResizable(True)
        self.store_sheet_scroll.setFrameShape(QFrame.NoFrame)
        self.store_sheet_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.store_sheet_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.store_sheet_scroll.setFixedHeight(38)
        self.store_sheet_scroll.setStyleSheet("""
            QScrollArea {
                background: transparent;
                border: none;
            }
            QScrollBar:horizontal {
                height: 6px;
                background: #edf2ec;
                border-radius: 3px;
            }
            QScrollBar::handle:horizontal {
                background: #b8c7b4;
                border-radius: 3px;
            }
        """)
        self.store_sheet_container = QWidget()
        self.store_sheet_layout = QHBoxLayout(self.store_sheet_container)
        self.store_sheet_layout.setContentsMargins(8, 2, 8, 2)
        self.store_sheet_layout.setSpacing(4)
        self.store_sheet_scroll.setWidget(self.store_sheet_container)
        secondary_toolbar.addWidget(self.store_sheet_scroll, 1)
        secondary_toolbar.addStretch()

        toolbar.addWidget(btn_prev)
        toolbar.addWidget(self.lbl_month)
        toolbar.addWidget(btn_next)

        self.product_sort_combo = QComboBox()
        self.product_sort_combo.addItem("按创建时间", "created_at")
        self.product_sort_combo.addItem("按单量", "order")
        self.product_sort_combo.addItem("按净利率", "net_margin")
        self.product_sort_combo.addItem("按净利润", "net_profit")
        self.product_sort_combo.addItem("按商品类型", "category")
        self.product_sort_combo.addItem("按毛利率", "gross_margin")
        self.product_sort_combo.addItem("按投产", "roi")
        self.product_sort_combo.addItem("按投产倍数", "roi_multiple")
        self.product_sort_combo.setFixedWidth(118)
        sort_index = self.product_sort_combo.findData(self.product_sort_mode)
        self.product_sort_combo.setCurrentIndex(sort_index if sort_index >= 0 else 0)
        self.product_sort_combo.currentIndexChanged.connect(self.on_product_sort_changed)
        self.product_sort_combo.setToolTip("排序所有店铺内显示的链接")
        self.product_sort_combo.setStyleSheet("""
            QComboBox {
                border: 1px solid #6c757d;
                border-radius: 4px;
                padding: 2px 8px;
                font-weight: bold;
                font-size: 13px;
                background: white;
            }
        """)
        toolbar.addWidget(self.product_sort_combo)

        self.btn_product_sort_direction = QPushButton()
        self.btn_product_sort_direction.setFixedSize(58, 30)
        self.btn_product_sort_direction.clicked.connect(self.toggle_product_sort_direction)
        self._update_product_sort_direction_button()
        toolbar.addWidget(self.btn_product_sort_direction)

        toolbar.addStretch()

        self.btn_tutorial = QPushButton("📖 功能教程")
        self.btn_tutorial.setFixedSize(108, 30)
        self.btn_tutorial.clicked.connect(self.show_tutorial_catalog)
        toolbar.addWidget(self.btn_tutorial)

        # 状态栏左下角按钮区域
        bottom_left_widget = QWidget()
        bottom_left_layout = QHBoxLayout(bottom_left_widget)
        bottom_left_layout.setContentsMargins(5, 0, 0, 0)
        bottom_left_layout.setSpacing(5)

        self.btn_api_config = QPushButton("🔑 API配置")
        self.btn_api_config.setFixedSize(80, 26)
        self.btn_api_config.setStyleSheet("""
            QPushButton {
                background-color: #e67e22;
                color: white;
                font-weight: bold;
                border-radius: 4px;
                font-size: 13px;
                padding: 1px;
            }
            QPushButton:hover {
                background-color: #d35400;
            }
        """)
        self.btn_api_config.clicked.connect(self.show_api_config_dialog)
        bottom_left_layout.addWidget(self.btn_api_config)

        self.btn_view_cost = QPushButton("📦 成本库")
        self.btn_view_cost.setFixedSize(102, 26)
        self.btn_view_cost.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                font-weight: bold;
                border-radius: 4px;
                font-size: 13px;
                padding: 1px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        self.btn_view_cost.clicked.connect(self.show_cost_library)
        bottom_left_layout.addWidget(self.btn_view_cost)

        self.btn_material_library = QPushButton("素材库")
        self.btn_material_library.setFixedSize(78, 26)
        self.btn_material_library.setStyleSheet("""
            QPushButton {
                background-color: #20c997;
                color: white;
                font-weight: bold;
                border-radius: 4px;
                font-size: 13px;
                padding: 1px;
            }
            QPushButton:hover {
                background-color: #17a589;
            }
        """)
        self.btn_material_library.clicked.connect(self.show_material_library)
        bottom_left_layout.addWidget(self.btn_material_library)

        self.btn_real_promotion_mode = QPushButton("真实推广数据模式")
        self.btn_real_promotion_mode.setCheckable(True)
        self.btn_real_promotion_mode.setFixedSize(125, 26)
        self.btn_real_promotion_mode.setChecked(self.db.get_setting("real_promotion_data_mode", "0") == "1")
        self.btn_real_promotion_mode.clicked.connect(self.toggle_real_promotion_data_mode)
        bottom_left_layout.addWidget(self.btn_real_promotion_mode)
        self._update_real_promotion_mode_button_style()

        self.btn_update_center = QPushButton("开发者模式")
        self.btn_update_center.setFixedSize(92, 26)
        self.btn_update_center.setStyleSheet("""
            QPushButton {
                background-color: #34495e;
                color: white;
                font-weight: bold;
                border-radius: 4px;
                font-size: 13px;
                padding: 1px;
            }
            QPushButton:hover {
                background-color: #2c3e50;
            }
        """)
        self.btn_update_center.clicked.connect(self.show_update_center)
        self.btn_update_center.setToolTip("输入开发者密码后推送局域网更新")
        bottom_left_layout.addWidget(self.btn_update_center)

        self.btn_check_update = QPushButton("检查更新")
        self.btn_check_update.setFixedSize(82, 26)
        self.btn_check_update.setStyleSheet("""
            QPushButton {
                background-color: #2d8f6f;
                color: white;
                font-weight: bold;
                border-radius: 4px;
                font-size: 13px;
                padding: 1px;
            }
            QPushButton:hover {
                background-color: #23745a;
            }
        """)
        self.btn_check_update.clicked.connect(self.check_saved_or_server_update)
        self.btn_check_update.setToolTip("检查更新，优先使用最近一次收到的更新下载地址")
        bottom_left_layout.addWidget(self.btn_check_update)

        self.statusBar().addWidget(bottom_left_widget)

        # 状态栏右下角按钮区域
        status_widget = QWidget()
        status_layout = QHBoxLayout(status_widget)
        status_layout.setContentsMargins(0, 0, 5, 0)

        self.lbl_archive_account = QLabel("未登录")
        self.lbl_archive_account.setStyleSheet("color: #888; font-size: 12px; padding: 0 5px;")
        self.lbl_archive_account.setAlignment(Qt.AlignVCenter)
        status_layout.addWidget(self.lbl_archive_account)

        self.btn_switch_local_account = QPushButton("🔁 切换账号")
        self.btn_switch_local_account.setFixedSize(90, 26)
        self.btn_switch_local_account.setStyleSheet("""
            QPushButton {
                background-color: #f39c12;
                color: white;
                font-weight: bold;
                border-radius: 4px;
                font-size: 13px;
                padding: 1px;
            }
            QPushButton:hover {
                background-color: #d68910;
            }
        """)
        self.btn_switch_local_account.clicked.connect(self.show_local_account_switch_menu)
        self.btn_switch_local_account.setToolTip("选择要切换的本地账号数据")
        status_layout.addWidget(self.btn_switch_local_account)

        self.btn_archive = QPushButton("💾 存档")
        self.btn_archive.setFixedSize(80, 26)
        self.btn_archive.setStyleSheet("""
            QPushButton {
                background-color: #009688;
                color: white;
                font-weight: bold;
                border-radius: 4px;
                font-size: 13px;
                padding: 1px;
            }
            QPushButton:hover {
                background-color: #00796b;
            }
        """)
        self.btn_archive.clicked.connect(self.show_archive_dialog)
        self.btn_archive.setToolTip("本地存档账号管理")
        status_layout.addWidget(self.btn_archive)

        self.btn_pinduoduo = QPushButton("🛒 拼多多")
        self.btn_pinduoduo.setFixedSize(80, 26)
        self.btn_pinduoduo.clicked.connect(self.open_pinduoduo)
        self.btn_pinduoduo.setToolTip("打开拼多多商家后台")
        status_layout.addWidget(self.btn_pinduoduo)

        self._apply_main_button_visuals(
            top_buttons=[
                (btn_prev, "#5f86a8", "#4f7390"),
                (btn_next, "#5f86a8", "#4f7390"),
                (self.btn_tutorial, "#4f8f74", "#417a62"),
            ],
            secondary_buttons=[
                (btn_add_store, "#5f8a62", "#4e744f"),
                (btn_daily_task, "#b86a62", "#9f5952"),
                (btn_export, "#4f8f83", "#41786e"),
            ],
            bottom_buttons=[
                (self.btn_api_config, "#b88445", "#9d7139"),
                (self.btn_view_cost, "#687886", "#566673"),
                (self.btn_material_library, "#4f9b86", "#428573"),
                (self.btn_update_center, "#536b7c", "#455a69"),
                (self.btn_check_update, "#4f8f74", "#417a62"),
                (self.btn_switch_local_account, "#c7923e", "#aa7b33"),
                (self.btn_archive, "#4a938a", "#3e7c74"),
                (self.btn_pinduoduo, "#c94d54", "#aa3f46"),
            ],
        )
        self._update_real_promotion_mode_button_style()

        self._install_main_button_hints({
            btn_prev: "切换到上个月的数据。",
            btn_next: "切换到下个月的数据。",
            self.btn_tutorial: "查看软件全部功能，并进入不会保存数据的分步新手教程。",
            self.btn_tag_filter: "按商品类型、活动状态、推广方式、垃圾/废物标签和盈亏状态筛选链接。",
            self.btn_store_filter: "选择需要显示的店铺；店铺管理也可使用顶部店铺标签右键菜单。",
            self.btn_product_sort_direction: "切换当前链接排序的升序和降序。",
            self.btn_clear_category_filter: "清除当前商品类型关键字筛选。",
            self.btn_filter_coupon: "只显示设置了优惠券的链接。",
            self.btn_filter_new_customer: "只显示设置了新客立减的链接。",
            self.btn_filter_limited_time: "只显示参加限时限量购的链接。",
            self.btn_filter_marketing: "只显示参加营销活动的链接。",
            self.btn_filter_natural_flow: "只显示标记为无推广的链接。",
            self.btn_filter_sitewide: "只显示使用全站托管的链接。",
            self.btn_filter_garbage: "只显示带有垃圾标签的链接。",
            self.btn_filter_waste: "只显示带有废物标签的链接。",
            self.btn_filter_profit: "只显示净利率不低于 1% 的赚钱链接。",
            self.btn_filter_loss: "只显示净利率低于 -2% 的亏钱链接。",
            self.btn_filter_break_even: "只显示净利率处于 -2% 到 1% 的保本链接。",
            self.btn_filter_missing_roi_bid: "只显示尚未填写投产或出价的链接。",
            btn_save_filter: "应用当前筛选条件并关闭筛选面板。",
            btn_clear_filter: "清空所有标签和商品类型筛选条件。",
            btn_add_store: "创建一个新店铺；空店铺也会显示店铺行。",
            btn_daily_task: "打开每日任务，查看待办、废物链接和垃圾链接。",
            btn_export: "批量导出店铺毛利，可选择详细版或单文件多 Sheet 简化版。",
            self.btn_api_config: "配置 AI API 密钥和各类提示词。",
            self.btn_view_cost: "打开成本库，管理规格编码、成本和商品类型链接。",
            self.btn_material_library: "打开素材库，管理商品类型和产品素材。",
            self.btn_real_promotion_mode: "切换真实推广数据模式；每个店铺显示自己最近一次导入的数据。",
            self.btn_update_center: "输入开发者密码后，管理并推送局域网更新。",
            self.btn_check_update: "检查是否有可下载的软件新版本。",
            self.btn_switch_local_account: "切换当前使用的本地店铺账号数据。",
            self.btn_archive: "打开本地存档账号管理和存档操作。",
            self.btn_pinduoduo: "按当前店铺打开拼多多商家后台。",
        })

        self.statusBar().addPermanentWidget(status_widget)

        # 1. 创建表格
        self.table = QTableWidget()
        self.model = QStandardItemModel()
        self.frozen_table = QTableWidget()
        
        # 2. 初始化表格属性
        self.setup_tables()
        
        # 2.5 设置今日列高亮代理
        self.today_delegate = OperationRecordDelegate(self.table, -1)
        self.table.setItemDelegate(self.today_delegate)
        
        # 3. 【关键】连接选中同步信号
        self.table.selectionModel().selectionChanged.connect(self.sync_frozen_selection)
        self.frozen_table.selectionModel().selectionChanged.connect(self.sync_main_selection)
        
        # 4. 【关键】安装事件过滤器 (给两个表格的视口都安装，用于拦截滚轮)
        self.table.viewport().installEventFilter(self)
        self.frozen_table.viewport().installEventFilter(self)
        
        # 5. 绑定双击事件 (打开规格弹窗)
        self.frozen_table.cellDoubleClicked.connect(self.open_product_spec_dialog_from_table)
        
        # --- 布局代码 (保持你原有的不变) ---
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(2, 2, 2, 2)
        main_layout.setSpacing(2)

        main_layout.addLayout(toolbar)
        main_layout.addLayout(secondary_toolbar)

        self.main_splitter = QSplitter(Qt.Horizontal)
        self.main_splitter.setChildrenCollapsible(False)
        self.main_splitter.setHandleWidth(1)
        self.main_splitter.setStyleSheet("QSplitter::handle { background: #b7b7b7; margin: 0px; }")
        self.main_splitter.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.main_splitter.addWidget(self.frozen_table)
        self.main_splitter.addWidget(self.table)
        self.main_splitter.setStretchFactor(0, 0)
        self.main_splitter.setStretchFactor(1, 1)
        self.main_splitter.splitterMoved.connect(self._save_splitter_frozen_width)

        operation_page = QWidget()
        operation_layout = QVBoxLayout(operation_page)
        operation_layout.setContentsMargins(0, 0, 0, 0)
        operation_layout.addWidget(self.main_splitter)

        self.data_mode_scroll = QScrollArea()
        self.data_mode_scroll.setWidgetResizable(True)
        self.data_mode_scroll.setFrameShape(QFrame.NoFrame)
        self.data_mode_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.data_mode_container = QWidget()
        self.data_mode_container.setContextMenuPolicy(Qt.CustomContextMenu)
        self.data_mode_container.customContextMenuRequested.connect(
            lambda pos: self._show_main_blank_context_menu(self.data_mode_container.mapToGlobal(pos))
        )
        self.data_mode_layout = QVBoxLayout(self.data_mode_container)
        self.data_mode_layout.setContentsMargins(10, 10, 10, 10)
        self.data_mode_layout.setSpacing(10)
        self.data_mode_layout.setAlignment(Qt.AlignTop)
        self.data_mode_scroll.setWidget(self.data_mode_container)
        self.data_mode_sticky_header = QLabel(self.data_mode_scroll.viewport())
        self.data_mode_sticky_header.setContextMenuPolicy(Qt.CustomContextMenu)
        self.data_mode_sticky_header.installEventFilter(self)
        self.data_mode_sticky_header.customContextMenuRequested.connect(
            lambda pos: self._show_sticky_store_context_menu(
                self.data_mode_sticky_header.mapToGlobal(pos)
            )
        )
        self.data_mode_sticky_header.hide()
        self._sticky_store_widget = None
        self._sticky_store_cache_key = None
        self.data_mode_scroll.verticalScrollBar().valueChanged.connect(
            self._update_sticky_store_header
        )

        self.main_view_stack = QStackedWidget()
        self.main_view_stack.addWidget(operation_page)
        self.main_view_stack.addWidget(self.data_mode_scroll)
        main_layout.addWidget(self.main_view_stack, 1)
        self._apply_main_view_mode(refresh=False)
        
        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        central_widget.setStyleSheet("background-color: #f7f1e3;")
        self.setCentralWidget(central_widget)
        
        # --- Toast 提示代码 (保持你原有的不变) ---
        self.toast_label = QLabel(self)
        self.toast_label.setStyleSheet("""
            background-color: rgba(0, 0, 0, 0.7);
            color: white;
            padding: 10px 20px;
            border-radius: 5px;
            font-size: 14px;
        """)
        self.toast_label.setAlignment(Qt.AlignCenter)
        self.toast_label.hide()
        
        self.toast_timer = QTimer(self)
        self.toast_timer.setSingleShot(True)
        self.toast_timer.timeout.connect(self.hide_toast)

    def resizeEvent(self, event):
        """
        【中文功能说明】
        窗口大小改变事件：当用户拉伸窗口时触发。
        作用：强制更新冻结表的位置和大小，防止按钮被遮挡或左右错位。
        """
        super().resizeEvent(event)
        self.update_frozen_geometry()
        if hasattr(self, "data_mode_sticky_header"):
            self._update_sticky_store_header()

    def _solid_button_style(self, bg, hover=None, fg="#ffffff"):
        hover = hover or bg
        return f"""
            QPushButton {{
                background-color: {bg};
                color: {fg};
                border: 1px solid {hover};
                border-radius: 5px;
                padding: 0px;
                font-weight: bold;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {hover};
            }}
            QPushButton:pressed {{
                background-color: {hover};
                padding: 0px;
            }}
        """

    def _outline_button_style(self, accent, hover_bg=None):
        hover_bg = hover_bg or accent
        return f"""
            QPushButton {{
                background-color: #f8fbf6;
                color: {accent};
                border: 1px solid {accent};
                border-radius: 5px;
                padding: 0px;
                font-weight: bold;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {hover_bg};
                color: white;
            }}
            QPushButton:pressed {{
                background-color: {hover_bg};
                color: white;
                padding: 0px;
            }}
        """

    def _apply_main_button_visuals(self, top_buttons, secondary_buttons, bottom_buttons):
        for button, bg, hover in top_buttons:
            button.setFixedHeight(30)
            button.setStyleSheet(self._solid_button_style(bg, hover))
        for button, bg, hover in secondary_buttons:
            button.setFixedHeight(30)
            button.setStyleSheet(self._solid_button_style(bg, hover))
        for button, bg, hover in bottom_buttons:
            button.setFixedHeight(28)
            button.setStyleSheet(self._solid_button_style(bg, hover))
        self.btn_tag_filter.setFixedSize(86, 30)
        self.btn_tag_filter.setStyleSheet(self._outline_button_style("#4f7f74", "#4f7f74"))
        self.btn_store_filter.setFixedSize(86, 30)
        self.btn_store_filter.setStyleSheet(self._outline_button_style("#5f8a62", "#5f8a62"))
        self.search_input.setFixedHeight(30)
        self.search_input.setStyleSheet("""
            QLineEdit {
                background-color: #fbfcf8;
                border: 1px solid #c8d4c4;
                border-radius: 5px;
                padding: 2px 8px;
                color: #26362f;
                font-size: 13px;
            }
            QLineEdit:focus {
                border: 1px solid #5f8a62;
                background-color: white;
            }
        """)
        self.product_sort_combo.setFixedHeight(30)
        self.product_sort_combo.setStyleSheet("""
            QComboBox {
                border: 1px solid #9aac99;
                border-radius: 5px;
                padding: 2px 8px;
                font-weight: bold;
                font-size: 13px;
                background: #fbfcf8;
                color: #26362f;
            }
            QComboBox:hover {
                border-color: #5f8a62;
            }
        """)
        self.btn_product_sort_direction.setStyleSheet(self._outline_button_style("#6b7280", "#6b7280"))

    def _store_sheet_button_style(self, active=False):
        if active:
            return """
                QPushButton {
                    background-color: #4f8f6f;
                    color: white;
                    border: 1px solid #3d755a;
                    border-bottom: 3px solid #2f5d47;
                    border-radius: 6px;
                    padding: 0px;
                    font-weight: bold;
                    text-align: center;
                }
                QPushButton:hover { background-color: #477f64; }
            """
        return """
            QPushButton {
                background-color: #f4f7f2;
                color: #365044;
                border: 1px solid #c8d6c4;
                border-bottom: 2px solid #b0c4aa;
                border-radius: 6px;
                padding: 0px;
                font-weight: bold;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #e7efe4;
                border-color: #91ad89;
            }
        """

    def refresh_store_sheets(self):
        if not hasattr(self, "store_sheet_layout"):
            return
        try:
            while self.store_sheet_layout.count():
                item = self.store_sheet_layout.takeAt(0)
                widget = item.widget()
                if widget:
                    self._retire_widget(widget)
            self.store_sheet_buttons = {}

            stores = self.db.safe_fetchall("SELECT id, name FROM stores ORDER BY sort_order, id")
            count_rows = self.db.safe_fetchall(
                "SELECT store_id, COUNT(*) FROM products WHERE COALESCE(is_archived, 0)=0 GROUP BY store_id"
            )
            counts = {int(row[0]): int(row[1] or 0) for row in count_rows if row and row[0] is not None}
            total_count = sum(counts.values())

            all_btn = self._create_store_sheet_button(None, "全部", total_count)
            self.store_sheet_layout.addWidget(all_btn)
            self.store_sheet_buttons[None] = all_btn

            for store_id, store_name in stores:
                btn = self._create_store_sheet_button(store_id, str(store_name or "未命名店铺"), counts.get(int(store_id), 0))
                self.store_sheet_layout.addWidget(btn)
                self.store_sheet_buttons[int(store_id)] = btn

            self.store_sheet_layout.addStretch()
            self.update_store_sheet_selection()
        except Exception as e:
            print(f"刷新店铺sheet失败: {e}")

    def _create_store_sheet_button(self, store_id, name, count):
        text = f"{name}（{count}条）"
        display_text = text if len(text) <= 18 else f"{text[:17]}…"
        button = QPushButton(display_text)
        button.setFixedHeight(30)
        button.setMinimumWidth(90 if store_id is None else 128)
        button.setCursor(Qt.PointingHandCursor)
        button.setToolTip(f"{name}\n当前链接：{count} 条")
        button.clicked.connect(lambda _checked=False, sid=store_id: self.select_store_sheet(sid))
        if store_id is not None:
            button.setContextMenuPolicy(Qt.CustomContextMenu)
            button.customContextMenuRequested.connect(
                lambda pos, sid=store_id, widget=button: self.show_store_sheet_context_menu(sid, widget.mapToGlobal(pos))
            )
        return button

    def show_store_sheet_context_menu(self, store_id, global_pos):
        menu = QMenu(self)
        menu.setStyleSheet("QMenu { font-size: 14px; padding: 6px; } QMenu::item { padding: 8px 28px 8px 16px; min-width: 150px; }")
        add_action = menu.addAction("添加链接")
        code_action = menu.addAction("抓取添加编码")
        price_action = menu.addAction("抓取价格管理")
        promotion_action = menu.addAction("抓取推广状态")
        promotion_data_action = menu.addAction("推广数据分析")
        record_action = menu.addAction("店铺操作记录")
        delete_action = menu.addAction("删除店铺")
        selected = menu.exec_(global_pos)
        if selected == add_action:
            self.add_product(store_id)
        elif selected == code_action:
            self.open_pdd_code_fetch_for_store(store_id)
        elif selected == price_action:
            self.open_pdd_price_fetch_for_store(store_id)
        elif selected == promotion_action:
            self.open_pdd_promotion_status_fetch_for_store(store_id)
        elif selected == promotion_data_action:
            self.open_promotion_data_for_store(store_id)
        elif selected == record_action:
            self.open_store_record_window(store_id)
        elif selected == delete_action:
            rows = self.db.safe_fetchall("SELECT name FROM stores WHERE id=?", (store_id,))
            store_name = rows[0][0] if rows else "该店铺"
            self.delete_store_by_id(store_id, store_name)

    def delete_store_by_id(self, store_id, store_name=""):
        if QMessageBox.question(
            self,
            "确认",
            f"确定删除店铺 '{store_name or '该店铺'}' 及其所有商品和记录吗？\n此操作不可恢复！",
        ) != QMessageBox.Yes:
            return
        try:
            product_ids = [
                int(row[0]) for row in self.db.safe_fetchall(
                    "SELECT id FROM products WHERE store_id=?", (store_id,)
                )
            ]
            self._close_store_scoped_windows(store_id, product_ids)
            self.db.delete_store_cascade(store_id)
            QTimer.singleShot(
                0,
                lambda sid=int(store_id), pids=product_ids: self.refresh_after_store_deleted(sid, pids),
            )
        except Exception as e:
            append_exception("delete_store:failed", error=e)
            QMessageBox.warning(self, "错误", f"删除店铺失败：{e}")

    def refresh_after_store_renamed(self, store_id, store_name):
        def update_widget(widget):
            if not _qobject_alive(widget):
                return
            widget.store_name = store_name
            widget.label.setText(f" {store_name}")
            if getattr(widget, "display_mode", "table") == "bubble":
                widget._apply_bubble_store_label_styles()

        section = getattr(self, "_data_mode_store_sections", {}).get(store_id)
        if _qobject_alive(section):
            update_widget(section.findChild(StoreWidget))
        for row, row_store_id in self.row_store_map.items():
            if row_store_id == store_id:
                update_widget(self.frozen_table.cellWidget(row, 0))

        state_row = self.main_table_state.store_row_by_id.get(store_id)
        if state_row is not None and state_row in self.main_table_state.row_by_index:
            state = self.main_table_state.row_by_index[state_row]
            state["store_name"] = store_name
            state["title"] = store_name

        margin_dialog = self.store_margin_dialogs.get(store_id)
        if _qobject_alive(margin_dialog):
            margin_dialog.store_name = store_name
            margin_dialog.setWindowTitle(f"🏪 店铺毛利管理 - {store_name}")
        for dialog in self.record_dialogs:
            if getattr(dialog, "store_id", None) == store_id and _qobject_alive(dialog):
                dialog.store_name = store_name
                if hasattr(dialog, "_update_window_title"):
                    dialog._update_window_title()
        self.refresh_store_sheets()
        self._update_sticky_store_header(force=True)

    def update_store_sheet_selection(self):
        if not hasattr(self, "store_sheet_buttons"):
            return
        selected = set(getattr(self, "current_store_filter", set()) or set())
        active_store_id = next(iter(selected)) if len(selected) == 1 else None
        for store_id, button in self.store_sheet_buttons.items():
            is_active = (not selected and store_id is None) or (len(selected) == 1 and store_id == active_store_id)
            button.setStyleSheet(self._store_sheet_button_style(is_active))

    def select_store_sheet(self, store_id):
        append_event(f"ui:select_store_sheet:start store_id={store_id}")
        try:
            if store_id is None:
                self.current_store_filter = set()
                self.db.set_setting("store_filter_ids", "")
                self.btn_store_filter.setText("🏪 店铺")
            else:
                self.current_store_filter = {int(store_id)}
                self.db.set_setting("store_filter_ids", str(int(store_id)))
                self.btn_store_filter.setText("🏪 店铺 (1)")
            self.update_store_sheet_selection()
            if self.main_view_mode == "data" and not self._has_active_product_filter_inputs():
                if self.current_store_filter:
                    self._set_visible_product_ids(
                        pid for pid, sid in self.product_store_map.items()
                        if sid in self.current_store_filter
                    )
                else:
                    self._set_visible_product_ids(self.product_store_map.keys())
                if not self._apply_data_mode_store_visibility():
                    self._refresh_data_mode_view_if_active()
                append_event("ui:select_store_sheet:done")
                return
            self.apply_tag_filter(close_menu=False, show_message=False)
            append_event("ui:select_store_sheet:done")
        except Exception as e:
            print(f"切换店铺sheet失败: {e}")
            append_exception("ui:select_store_sheet:failed", error=e)

    def _visible_product_row_count(self):
        try:
            return len(self._visible_product_ids())
        except Exception:
            return 0

    def _is_qobject_alive(self, obj):
        return _qobject_alive(obj)

    def _install_main_button_hints(self, hints):
        for button, text in hints.items():
            if not self._is_qobject_alive(button):
                continue
            button.setToolTip("")
            self._button_hint_widgets[button] = str(text or "").strip()
            button.installEventFilter(self)

    def _cancel_button_hint(self):
        self._button_hint_timer.stop()
        self._button_hint_active = None
        self._button_hint_label.hide()

    def _show_button_hint(self):
        button = self._button_hint_active
        if not self._is_qobject_alive(button):
            self._cancel_button_hint()
            return
        try:
            if not button.isVisible() or not button.rect().contains(button.mapFromGlobal(QCursor.pos())):
                self._cancel_button_hint()
                return
            text = self._button_hint_widgets.get(button, "")
            if not text:
                self._cancel_button_hint()
                return
            self._button_hint_label.setText(text)
            self._button_hint_label.adjustSize()
            pos = button.mapToGlobal(QPoint(
                (button.width() - self._button_hint_label.width()) // 2,
                button.height() + 6,
            ))
            screen = QApplication.screenAt(button.mapToGlobal(button.rect().center()))
            if screen:
                area = screen.availableGeometry()
                pos.setX(max(area.left(), min(pos.x(), area.right() - self._button_hint_label.width() + 1)))
                if pos.y() + self._button_hint_label.height() > area.bottom():
                    pos.setY(button.mapToGlobal(QPoint(0, -self._button_hint_label.height() - 6)).y())
            self._button_hint_label.move(pos)
            self._button_hint_label.show()
            self._button_hint_label.raise_()
        except RuntimeError:
            self._cancel_button_hint()

    def eventFilter(self, obj, event):
        """
        事件过滤器：处理滚轮事件，实现丝滑滚动功能。
        冻结表格滚动时同步到主表格。

        配置说明：
        - speed_factor: 滚动速度因子，值越大滚动越慢/越精细
          推荐范围：60-500
          120: 每次滚动约1像素
          240: 每次滚动约0.5像素
          480: 每次滚动约0.25像素
        """
        if obj is getattr(self, "data_mode_sticky_header", None):
            if event.type() == QEvent.MouseButtonDblClick and event.button() == Qt.LeftButton:
                store_widget = getattr(self, "_sticky_store_widget", None)
                if _qobject_alive(store_widget):
                    store_widget.open_store_margin_dialog()
                return True

        if obj in getattr(self, "_button_hint_widgets", {}):
            if event.type() == QEvent.Enter:
                self._cancel_button_hint()
                self._button_hint_active = obj
                self._button_hint_timer.start(500)
            elif event.type() in (QEvent.Leave, QEvent.MouseButtonPress, QEvent.FocusOut, QEvent.Hide):
                if self._button_hint_active is obj:
                    self._cancel_button_hint()
            elif event.type() == QEvent.ToolTip:
                return True

        if hasattr(self, "table") and obj == self.table.viewport():
            if event.type() == QEvent.ToolTip:
                return True
            if event.type() == QEvent.Leave:
                self._cancel_record_tooltip()
            elif event.type() == QEvent.MouseMove:
                self._handle_record_tooltip_mouse_move(event)

        if event.type() == QEvent.Wheel:
            if not hasattr(self, "table") or not hasattr(self, "frozen_table"):
                return super().eventFilter(obj, event)
            v_scroll = self.table.verticalScrollBar()
            h_scroll = self.table.horizontalScrollBar()

            if not v_scroll:
                return super().eventFilter(obj, event)

            delta_y = event.angleDelta().y()
            delta_x = event.angleDelta().x()

            speed_factor = 120.0

            if obj == self.frozen_table.viewport():
                if delta_y != 0:
                    current_value = v_scroll.value()
                    if (current_value == v_scroll.minimum() and delta_y > 0) or \
                       (current_value == v_scroll.maximum() and delta_y < 0):
                        self._accumulated_v = 0
                    else:
                        step = delta_y / speed_factor
                        self._accumulated_v = getattr(self, '_accumulated_v', 0) + step
                        if abs(self._accumulated_v) >= 1.0:
                            scroll_step = int(self._accumulated_v)
                            self._accumulated_v -= scroll_step
                            new_value = current_value - scroll_step
                            v_scroll.setValue(max(v_scroll.minimum(), min(v_scroll.maximum(), new_value)))
                return True

            if obj == self.table.viewport():
                if delta_y != 0:
                    current_value = v_scroll.value()
                    if (current_value == v_scroll.minimum() and delta_y > 0) or \
                       (current_value == v_scroll.maximum() and delta_y < 0):
                        self._accumulated_v = 0
                    else:
                        step = delta_y / speed_factor
                        self._accumulated_v = getattr(self, '_accumulated_v', 0) + step
                        if abs(self._accumulated_v) >= 1.0:
                            scroll_step = int(self._accumulated_v)
                            self._accumulated_v -= scroll_step
                            new_value_y = current_value - scroll_step
                            v_scroll.setValue(max(v_scroll.minimum(), min(v_scroll.maximum(), new_value_y)))

                if h_scroll and delta_x != 0:
                    current_h_value = h_scroll.value()
                    if (current_h_value == h_scroll.minimum() and delta_x > 0) or \
                       (current_h_value == h_scroll.maximum() and delta_x < 0):
                        self._accumulated_h = 0
                    else:
                        h_step = delta_x / speed_factor
                        self._accumulated_h = getattr(self, '_accumulated_h', 0) + h_step
                        if abs(self._accumulated_h) >= 1.0:
                            h_scroll_step = int(self._accumulated_h)
                            self._accumulated_h -= h_scroll_step
                            new_value_x = current_h_value - h_scroll_step
                            h_scroll.setValue(max(h_scroll.minimum(), min(h_scroll.maximum(), new_value_x)))

                return True

        return super().eventFilter(obj, event)

    def _handle_record_tooltip_mouse_move(self, event):
        index = self.table.indexAt(event.pos())
        if not index.isValid() or index.column() < 0:
            self._cancel_record_tooltip()
            return

        item = self.table.item(index.row(), index.column())
        text = item.toolTip() if item else ""
        if not text:
            self._cancel_record_tooltip()
            return

        cell = (index.row(), index.column())
        global_pos = event.globalPos() if hasattr(event, "globalPos") else self.table.viewport().mapToGlobal(event.pos())
        if cell != self._record_tooltip_cell:
            QToolTip.hideText()
            if hasattr(self, "_record_tooltip_timer"):
                self._record_tooltip_timer.stop()
            self._record_tooltip_cell = cell
            self._record_tooltip_text = text
            self._record_tooltip_pos = global_pos
            self._record_tooltip_timer.start(900)
            return

        self._record_tooltip_pos = global_pos

    def _show_pending_record_tooltip(self):
        cell = getattr(self, "_record_tooltip_cell", None)
        text = getattr(self, "_record_tooltip_text", "")
        if not cell or not text:
            return
        row, col = cell
        index = self.table.model().index(row, col)
        if not index.isValid():
            return
        current_index = self.table.indexAt(self.table.viewport().mapFromGlobal(QCursor.pos()))
        if not current_index.isValid() or current_index.row() != row or current_index.column() != col:
            return
        rect = self.table.visualRect(index)
        QToolTip.showText(self._record_tooltip_pos, text, self.table.viewport(), rect)

    def _cancel_record_tooltip(self):
        if hasattr(self, "_record_tooltip_timer"):
            self._record_tooltip_timer.stop()
        self._record_tooltip_cell = None
        self._record_tooltip_text = ""
        QToolTip.hideText()

    def show_toast(self, message, duration=3000):
        """显示悬浮提示
        
        Args:
            message: 提示消息
            duration: 显示时长（毫秒），默认3000毫秒
        """
        self.toast_label.setText(message)
        self.toast_label.adjustSize() # 根据文字调整大小
        
        # 计算位置：让提示居中显示
        x = (self.width() - self.toast_label.width()) // 2
        y = self.height() - 100 # 距离底部 100 像素，或者改成 self.height()//2 居中
        
        self.toast_label.move(x, y)
        self.toast_label.show()
        
        # 指定时长后自动隐藏
        self.toast_timer.start(duration)

    def hide_toast(self):
        """隐藏提示"""
        self.toast_label.hide()

    def _on_scroll_changed(self):
        """滚动位置变化时自动保存（带防抖）"""
        self._scroll_save_timer.stop()
        self._scroll_save_timer.start(500)

    def _save_scroll_position_to_db(self):
        """实际保存滚动位置到数据库"""
        try:
            v_scroll = self.table.verticalScrollBar()
            h_scroll = self.table.horizontalScrollBar()
            if v_scroll:
                self.db.set_setting("scroll_vertical", v_scroll.value())
            if h_scroll:
                self.db.set_setting("scroll_horizontal", h_scroll.value())
        except Exception as e:
            print(f"保存滚动位置失败: {e}")

    def setup_tables(self):
        # --- 主表设置 ---
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setFixedHeight(self.TABLE_HEADER_HEIGHT)
        self.table.horizontalHeader().setMinimumHeight(self.TABLE_HEADER_HEIGHT)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.table.verticalHeader().hide()
        self.table.cellDoubleClicked.connect(self.open_editor)
        self.table.setWordWrap(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.table.setFocusPolicy(Qt.NoFocus)
        self.table.verticalHeader().setDefaultSectionSize(self.PRODUCT_ROW_HEIGHT)
        self.table.setMouseTracking(True)
        self.table.viewport().setMouseTracking(True)
        self.table.setAlternatingRowColors(True)
        
        # --- 冻结表设置 ---
        self.frozen_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.frozen_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.frozen_table.horizontalHeader().setFixedHeight(self.TABLE_HEADER_HEIGHT)
        self.frozen_table.horizontalHeader().setMinimumHeight(self.TABLE_HEADER_HEIGHT)
        self.frozen_table.verticalHeader().hide()
        self.frozen_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.frozen_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.frozen_table.setWordWrap(True)
        self.frozen_table.setMouseTracking(True)
        self.frozen_table.viewport().setMouseTracking(True)
        self.frozen_table.setAlternatingRowColors(True)
        
        # 【关键】确保冻结表也能整行选中和获取焦点
        self.frozen_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.frozen_table.setFocusPolicy(Qt.NoFocus)
        self.frozen_table.verticalHeader().setDefaultSectionSize(self.PRODUCT_ROW_HEIGHT)
        
        # 样式：右边框加粗
        self.frozen_table.setStyleSheet("QTableWidget { border-right: 2px solid #555; background-color: white; }")
        
        # --- 信号连接 ---
        # 1. 垂直滚动条同步
        self.table.verticalScrollBar().valueChanged.connect(self.frozen_table.verticalScrollBar().setValue)
        self.frozen_table.verticalScrollBar().valueChanged.connect(self.table.verticalScrollBar().setValue)
        
        # 2. 滚动位置自动保存（使用防抖，避免频繁写入）
        self._scroll_save_timer = QTimer(self)
        self._scroll_save_timer.setSingleShot(True)
        self._scroll_save_timer.timeout.connect(self._save_scroll_position_to_db)
        self.table.verticalScrollBar().valueChanged.connect(self._on_scroll_changed)
        self.table.horizontalScrollBar().valueChanged.connect(self._on_scroll_changed)
        self._record_tooltip_timer = QTimer(self)
        self._record_tooltip_timer.setSingleShot(True)
        self._record_tooltip_timer.timeout.connect(self._show_pending_record_tooltip)
        
        # 3. 行高同步
        self.table.verticalHeader().sectionResized.connect(self.sync_row_height)
        
        # 4. 保存左侧冻结列宽设置到数据库
        self._suppress_col0_width_save = False
        self.frozen_table.horizontalHeader().sectionResized.connect(self._save_frozen_col_width)
            # 主表样式
        self.table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #c8d4c4;
                border-left: none;
                gridline-color: #d7dfd3;
                background: #fbfcf8;
                alternate-background-color: #f4f8f2;
                color: #26362f;
            }

            /* 选中行的样式 */
            QTableWidget::item:selected {
                background-color: #dfeeda;
                color: #17251e;
                border: none;
                padding: 0px;
                outline: none;
            }

            QTableWidget::item:focus {
                border: none;
                outline: none;
            }

            /* 当窗口失去焦点时的选中行样式 */
            QTableWidget::item:selected:!active {
                background-color: #dfeeda;
                color: #17251e;
                padding: 0px;
                outline: none;
            }

            /* 单元格基础样式 */
            QTableWidget::item {
                padding: 0px;
                border: none;
            }
        """)
        
        # 冻结表样式（和主表保持一致，但确保文字清晰可读）
        self.frozen_table.setStyleSheet("""
            QTableWidget {
                border: none;
                border: 1px solid #c8d4c4;
                border-right: none;
                background-color: #fbfcf8;
                alternate-background-color: #f4f8f2;
                color: #26362f;
                font-weight: bold;
                gridline-color: #d7dfd3;
            }

            QTableWidget::item {
                color: #26362f;
                font-weight: bold;
                padding: 0px;
                border: none;
            }

            /* 选中行的样式 */
            QTableWidget::item:selected {
                background-color: #dfeeda;
                color: #17251e;
                padding: 0px;
                outline: none;
            }

            QTableWidget::item:focus {
                border: none;
                outline: none;
            }

            /* 失焦时的选中行样式 */
            QTableWidget::item:selected:!active {
                background-color: #dfeeda;
                color: #17251e;
                padding: 0px;
            }
        """)
    def _sync_frozen_from_main(self):
        """主表 -> 冻结表"""
        indexes = self.table.selectionModel().selectedIndexes()
        if not indexes:
            self.frozen_table.clearSelection()
            self._update_frozen_selection_highlight(None)
            self.frozen_table.viewport().update()
            return
        row = indexes[0].row()
        if 0 <= row < self.frozen_table.rowCount():
            self.frozen_table.blockSignals(True)
            self.frozen_table.selectRow(row)
            self._update_frozen_selection_highlight(row)
            self.frozen_table.viewport().update()
            self.frozen_table.update()
            self.frozen_table.blockSignals(False)

    def _sync_main_from_frozen(self):
        """冻结表 -> 主表"""
        indexes = self.frozen_table.selectionModel().selectedRows()
        if not indexes:
            self.table.clearSelection()
            self._update_frozen_selection_highlight(None)
            self.table.viewport().update()
            return
        row = indexes[0].row()
        if 0 <= row < self.table.rowCount():
            self.table.blockSignals(True)
            self._update_frozen_selection_highlight(row)
            self.table.viewport().update()
            self.table.update()
            self.table.blockSignals(False)    

    def sync_row_height(self, logicalIndex, oldSize, newSize):
        try:
            self.frozen_table.setRowHeight(logicalIndex, newSize)
        except Exception as e:
            print(f"同步行高失败：{e}")
        
    def sync_col_width(self, logicalIndex, oldSize, newSize):
        try:
            return
        except Exception as e:
            print(f"同步列宽失败：{e}")

    def sync_frozen_selection(self, selected, deselected):
        """主表选中变化时，同步冻结表选中状态"""
        indexes = selected.indexes()
        if not indexes:
            self.frozen_table.clearSelection()
            self._update_frozen_selection_highlight(None)
            return
        row = indexes[0].row()
        if 0 <= row < self.frozen_table.rowCount():
            self.frozen_table.blockSignals(True)
            self.frozen_table.selectRow(row)
            self._update_frozen_selection_highlight(row)
            self.frozen_table.viewport().update()
            self.frozen_table.update()
            self.frozen_table.blockSignals(False)

    def sync_main_selection(self, selected, deselected):
        """冻结表选中变化时，同步主表选中状态"""
        indexes = selected.indexes()
        if not indexes:
            self.table.clearSelection()
            self._update_frozen_selection_highlight(None)
            return
        row = indexes[0].row()
        if 0 <= row < self.table.rowCount():
            self.table.blockSignals(True)
            self._update_frozen_selection_highlight(row)
            self.table.viewport().update()
            self.table.update()
            self.table.blockSignals(False)

    def _update_frozen_selection_highlight(self, selected_row):
        selected_color = QColor("#ffe0b2")
        normal_color = QColor("#ffffff")
        for row in range(self.frozen_table.rowCount()):
            widget = self.frozen_table.cellWidget(row, 0)
            if not widget:
                continue
            for target in (widget, self._product_widget_at_row(row)):
                if not target:
                    continue
                palette = target.palette()
                palette.setColor(QPalette.Window, selected_color if row == selected_row else normal_color)
                target.setAutoFillBackground(True)
                target.setPalette(palette)

    def update_frozen_geometry(self):
        try:
            if hasattr(self, "main_splitter"):
                sizes = self.main_splitter.sizes()
                total = sum(sizes) if sizes else max(1, self.width())
                default_width = max(320, int(total * 0.4) - 200)
                saved_width = int(self.db.get_setting("col_0_width", default_width))
                if self.db.get_setting("col_0_width_minus_100_applied", "0") != "1":
                    saved_width = max(320, saved_width - 100)
                    self.db.set_setting("col_0_width", saved_width)
                    self.db.set_setting("col_0_width_minus_100_applied", "1")
                if self.db.get_setting("col_0_width_minus_200_applied", "0") != "1":
                    saved_width = max(320, saved_width - 200)
                    self.db.set_setting("col_0_width", saved_width)
                    self.db.set_setting("col_0_width_minus_200_applied", "1")
                user_adjusted = self.db.get_setting("col_0_width_user_adjusted", "0") == "1"
                left_width = max(320, saved_width if user_adjusted else default_width)
                right_width = max(400, total - left_width)
                self.main_splitter.setSizes([left_width, right_width])
                self._fit_frozen_column_to_viewport()
        except Exception as e:
            print(f"更新冻结表几何位置失败：{e}")

    def _fit_frozen_column_to_viewport(self):
        try:
            if not hasattr(self, "frozen_table") or self.frozen_table.columnCount() <= 0:
                return
            width = max(1, self.frozen_table.viewport().width())
            self._suppress_col0_width_save = True
            self.frozen_table.setColumnWidth(0, width)
            self._suppress_col0_width_save = False
        except Exception as e:
            self._suppress_col0_width_save = False
            print(f"同步冻结列宽失败: {e}")

    def _product_widget_at_row(self, row):
        widget = self.frozen_table.cellWidget(row, 0)
        if isinstance(widget, ProductWidget):
            return widget
        return widget.findChild(ProductWidget) if widget else None

    def _wrap_product_widget(self, product_widget):
        product_widget.setFixedWidth(self.FROZEN_PRODUCT_CONTENT_WIDTH)
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        layout.addWidget(product_widget, 0, Qt.AlignLeft | Qt.AlignTop)
        layout.addStretch(1)
        return container

    def force_refresh_frozen_table(self, store_id=None):
        """强制刷新 frozen_table 的显示，确保数据更新后能正确显示"""
        try:
            cache = getattr(self, "_latest_promotion_data_cache", None)
            if isinstance(cache, dict):
                if store_id is None:
                    cache.clear()
                else:
                    for key in list(cache):
                        if isinstance(key, tuple) and len(key) > 1 and key[1] == store_id:
                            cache.pop(key, None)
            if self.main_view_mode == "data" and store_id is not None:
                self.refresh_store_cards(store_id)
                self._apply_data_mode_store_visibility()
                return
            self.frozen_table.viewport().update()
            self.frozen_table.update()
            for row in range(self.frozen_table.rowCount()):
                widget = self._product_widget_at_row(row)
                if _qobject_alive(widget):
                    widget.update_margin_display()
                    widget.update_promo_badges()
                    widget.update_task_badge()
        except Exception as e:
            print(f"强制刷新frozen_table失败: {e}")

    def is_real_promotion_data_mode(self):
        return self.db.get_setting("real_promotion_data_mode", "0") == "1"

    def get_real_promotion_hidden_metrics(self):
        cached = getattr(self, "_real_promotion_hidden_metrics", None)
        if cached is None:
            raw = self.db.get_setting("real_promotion_hidden_metrics", "") or ""
            cached = {key for key in raw.split(",") if key}
            self._real_promotion_hidden_metrics = cached
        return cached

    def _show_main_blank_context_menu(self, global_pos):
        if self.main_view_mode != "data":
            return
        menu = QMenu(self)
        menu.setStyleSheet("QMenu { font-size: 14px; padding: 6px; } QMenu::item { padding: 8px 24px; min-width: 130px; }")
        refresh_action = menu.addAction("刷新主界面")
        settings_action = None
        if self.is_real_promotion_data_mode():
            settings_action = menu.addAction("真实推广数据显示设置")
        selected = menu.exec_(global_pos)
        if selected == refresh_action:
            self._request_manual_main_refresh()
        elif settings_action is not None and selected == settings_action:
            self._open_real_promotion_display_settings()

    def _current_main_view_change_token(self):
        try:
            data_version = self.db.conn.execute("PRAGMA data_version").fetchone()[0]
            return int(self.db.conn.total_changes), int(data_version)
        except Exception:
            return None

    def _request_manual_main_refresh(self):
        if self._manual_refresh_pending:
            return
        self._manual_refresh_pending = True
        QTimer.singleShot(0, self._run_manual_incremental_refresh)

    def _run_manual_incremental_refresh(self):
        self._manual_refresh_pending = False
        if self.is_loading:
            self.show_toast("主界面正在刷新，请稍后再试")
            return
        try:
            current_token = self._current_main_view_change_token()
            if self._main_view_change_token is None:
                self._main_view_change_token = current_token
            if current_token == self._main_view_change_token:
                self.show_toast("没有内容变更")
                return

            db_store_ids = {int(row[0]) for row in self.db.safe_fetchall("SELECT id FROM stores")}
            db_product_map = {
                int(product_id): int(store_id)
                for product_id, store_id in self.db.safe_fetchall(
                    "SELECT id, store_id FROM products WHERE COALESCE(is_archived, 0)=0"
                )
            }
            ui_store_ids = {int(store_id) for store_id in self.row_store_map.values()}
            ui_product_map = {int(pid): int(sid) for pid, sid in self.product_store_map.items()}
            if db_store_ids != ui_store_ids or db_product_map != ui_product_map:
                append_event("manual_refresh:structure_changed")
                self.load_data_safe(restore_position=True)
            else:
                append_event("manual_refresh:incremental")
                for product_id in list(self.bubble_product_widgets):
                    self.force_refresh_product_widget(product_id)
                self._refresh_store_margin_widgets()
                self.refresh_store_sheets()
                self._reorder_data_mode_bubbles()
                self._apply_data_mode_store_visibility()
                self.update_daily_task_button_badge()
            self._main_view_change_token = self._current_main_view_change_token()
            self.show_toast("主界面已刷新")
        except Exception as e:
            append_exception("manual_refresh:failed", error=e)
            self.show_toast(f"刷新失败：{e}")

    def _open_real_promotion_display_settings(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("真实推广数据显示设置")
        dialog.setMinimumWidth(420)
        layout = QVBoxLayout(dialog)
        show_ordered_only = QCheckBox("只显示有推广数据且单量大于 0 的链接")
        show_ordered_only.setChecked(self.db.get_setting("real_promotion_show_ordered_data_only", "1") == "1")
        layout.addWidget(show_ordered_only)
        layout.addWidget(QLabel("选择需要显示的数据："))
        fields_widget = QWidget()
        fields_layout = QGridLayout(fields_widget)
        fields_layout.setContentsMargins(0, 0, 0, 0)
        fields_layout.setSpacing(6)
        hidden = self.get_real_promotion_hidden_metrics()
        field_checks = {}
        for index, (key, label) in enumerate((
            ("avg_price", "客单价"),
            ("gross_margin_rate", "毛利率"),
            ("cost", "花费"), ("transaction_amount", "交易额"),
            ("net_orders", "净成交单量"), ("net_roi", "净投产比"),
            ("net_profit", "净利润"), ("net_margin_rate", "净利率"),
            ("profit_status", "盈亏状态"), ("roi_multiple", "投产倍数"),
            ("promotion_share", "曝光占比"), ("amount_per_order", "每笔成交"),
            ("cost_per_order", "每笔花费"), ("ctr", "点击率"),
            ("conversion_rate", "点击转化率"),
        )):
            check = QCheckBox(label)
            check.setChecked(key not in hidden)
            fields_layout.addWidget(check, index // 3, index % 3)
            field_checks[key] = check
        layout.addWidget(fields_widget)
        selection_row = QHBoxLayout()
        select_all = QPushButton("全选")
        select_none = QPushButton("全不选")
        invert = QPushButton("反选")
        select_all.clicked.connect(lambda: [check.setChecked(True) for check in field_checks.values()])
        select_none.clicked.connect(lambda: [check.setChecked(False) for check in field_checks.values()])
        invert.clicked.connect(lambda: [check.setChecked(not check.isChecked()) for check in field_checks.values()])
        selection_row.addWidget(select_all)
        selection_row.addWidget(select_none)
        selection_row.addWidget(invert)
        selection_row.addStretch()
        layout.addLayout(selection_row)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        dialog.adjustSize()
        dialog.move(self.frameGeometry().center() - dialog.rect().center())
        if dialog.exec_() != QDialog.Accepted:
            return
        hidden.clear()
        hidden.update(key for key, check in field_checks.items() if not check.isChecked())
        self.db.set_setting("real_promotion_hidden_metrics", ",".join(sorted(hidden)))
        self.db.set_setting("real_promotion_show_ordered_data_only", "1" if show_ordered_only.isChecked() else "0")
        self._apply_real_promotion_display_settings(show_ordered_only.isChecked())

    def _apply_real_promotion_display_settings(self, show_data_only):
        selected_stores = set(getattr(self, "current_store_filter", set()) or set())
        for product_id, bubble in getattr(self, "bubble_product_widgets", {}).items():
            store_id = self.product_store_map.get(product_id) or self.get_product_card_data(product_id).get("store_id")
            data = self.get_latest_promotion_data(store_id, bubble.prod_code) if show_data_only else None
            visible = _promotion_link_visible(data, show_data_only)
            bubble.setVisible(visible)
            if visible:
                bubble.update_margin_display()
                if not self.is_real_promotion_data_mode():
                    bubble.update_link_order_count()
        for store_id, section in getattr(self, "_data_mode_store_sections", {}).items():
            has_visible_product = any(not bubble.isHidden() for bubble in section.findChildren(ProductWidget))
            section.setVisible(has_visible_product and (not selected_stores or store_id in selected_stores))
            section.layout().invalidate()
        self.data_mode_layout.invalidate()
        self.data_mode_container.updateGeometry()

    def _update_real_promotion_mode_button_style(self):
        if not hasattr(self, "btn_real_promotion_mode"):
            return
        enabled = self.btn_real_promotion_mode.isChecked()
        self.btn_real_promotion_mode.setText("真实推广: 开" if enabled else "真实推广数据模式")
        bg = "#9a6a5f" if enabled else "#687886"
        hover = "#81574e" if enabled else "#566673"
        self.btn_real_promotion_mode.setFixedHeight(28)
        self.btn_real_promotion_mode.setStyleSheet(f"""
            QPushButton {{
                background-color: {bg};
                color: white;
                font-weight: bold;
                border: 1px solid {hover};
                border-radius: 5px;
                font-size: 12px;
                padding: 0px;
            }}
            QPushButton:hover {{
                background-color: {hover};
            }}
        """)
    def _apply_main_view_mode(self, refresh=True):
        if not hasattr(self, "main_view_stack"):
            return
        self.main_view_mode = "data"
        if self.db.get_setting("main_view_mode", "data") != "data":
            self.db.set_setting("main_view_mode", "data")
        self.main_view_stack.setCurrentIndex(1 if self.main_view_stack.count() > 1 else 0)
        if refresh:
            self._refresh_data_mode_view_if_active()

    def _refresh_data_mode_view_if_active(self):
        if self.main_view_mode != "data" or self.is_loading or self._data_mode_refresh_pending:
            return
        self._data_mode_refresh_pending = True
        QTimer.singleShot(0, self._run_pending_data_mode_refresh)

    def _run_pending_data_mode_refresh(self):
        try:
            if self.main_view_mode == "data" and not self.is_loading:
                self._refresh_data_mode_view()
        except Exception as e:
            print(f"刷新纯数据模式失败: {e}")
            traceback.print_exc()
            self.show_toast(f"纯数据模式刷新失败: {e}")
        finally:
            self._data_mode_refresh_pending = False

    def _visible_product_ids(self):
        if hasattr(self, "visible_product_ids"):
            return set(self.visible_product_ids or set())
        return {product_id for product_id in self.row_data_map.values() if product_id}

    def _set_visible_product_ids(self, product_ids):
        self.visible_product_ids = {int(pid) for pid in (product_ids or []) if pid}

    def _empty_store_visible(self, store_id):
        show_data_only = (
            self.is_real_promotion_data_mode()
            and self.db.get_setting("real_promotion_show_ordered_data_only", "1") == "1"
        )
        selected_stores = getattr(self, "current_store_filter", set())
        return (
            not show_data_only
            and getattr(self, "current_search_match_ids", None) is None
            and not self._has_active_product_filter_inputs()
            and (not selected_stores or store_id in selected_stores)
        )

    def _create_data_mode_store_section(self, store_id, store_name):
        section = QWidget()
        section_layout = QVBoxLayout(section)
        section_layout.setContentsMargins(0, 0, 0, 0)
        section_layout.setSpacing(10)
        store_widget = StoreWidget(store_id, store_name, self, display_mode="bubble")
        store_widget.setFixedHeight(StoreWidget.BUBBLE_HEIGHT)
        section_layout.addWidget(store_widget)

        flow_widget = QWidget()
        flow_widget.setContextMenuPolicy(Qt.CustomContextMenu)
        flow_widget.customContextMenuRequested.connect(
            lambda pos, widget=flow_widget: self._show_main_blank_context_menu(widget.mapToGlobal(pos))
        )
        flow_layout = OrderedFlowLayout(flow_widget, spacing=10)
        section_layout.addWidget(flow_widget)
        self.data_mode_layout.addWidget(section)
        self._data_mode_store_sections[store_id] = section
        return section, flow_layout

    def _show_sticky_store_context_menu(self, global_pos):
        store_widget = getattr(self, "_sticky_store_widget", None)
        if _qobject_alive(store_widget):
            store_widget.show_store_context_menu(global_pos)

    def _update_sticky_store_header(self, *_args, force=False):
        sticky = getattr(self, "data_mode_sticky_header", None)
        scroll = getattr(self, "data_mode_scroll", None)
        if not _qobject_alive(sticky) or not _qobject_alive(scroll):
            return

        viewport = scroll.viewport()
        headers = []
        for store_id, section in getattr(self, "_data_mode_store_sections", {}).items():
            if not _qobject_alive(section) or section.isHidden() or section.layout().count() == 0:
                continue
            store_widget = section.layout().itemAt(0).widget()
            if not _qobject_alive(store_widget):
                continue
            headers.append((
                store_id,
                store_widget,
                store_widget.mapTo(viewport, QPoint(0, 0)),
            ))

        if not headers:
            sticky.hide()
            self._sticky_store_widget = None
            self._sticky_store_cache_key = None
            return

        active_index = 0
        for index, (_store_id, _widget, position) in enumerate(headers):
            if position.y() <= 0:
                active_index = index
            else:
                break

        store_id, store_widget, position = headers[active_index]
        if position.y() >= 0:
            sticky.hide()
            self._sticky_store_widget = store_widget
            return

        cache_key = (store_id, id(store_widget), store_widget.width(), store_widget.height())
        if force or cache_key != self._sticky_store_cache_key:
            sticky.setPixmap(store_widget.grab())
            self._sticky_store_cache_key = cache_key

        sticky_y = 0
        if active_index + 1 < len(headers):
            next_y = headers[active_index + 1][2].y()
            sticky_y = min(0, next_y - store_widget.height())
        sticky.setGeometry(
            position.x(),
            sticky_y,
            store_widget.width(),
            store_widget.height(),
        )
        self._sticky_store_widget = store_widget
        sticky.show()
        sticky.raise_()

    def _refresh_data_mode_view(self):
        if not hasattr(self, "data_mode_layout"):
            return
        self._data_mode_render_token = getattr(self, "_data_mode_render_token", 0) + 1
        render_token = self._data_mode_render_token
        self._prepare_product_card_caches()
        saved_scroll = self.data_mode_scroll.verticalScrollBar().value()
        self.data_mode_sticky_header.hide()
        self._sticky_store_widget = None
        self._sticky_store_cache_key = None
        self._clear_layout_widgets(self.data_mode_layout)
        self.bubble_product_widgets = {}
        self._data_mode_store_sections = {}
        visible_ids = self._visible_product_ids()
        stores = self.db.safe_fetchall("SELECT id, name FROM stores ORDER BY sort_order")
        real_mode = self.is_real_promotion_data_mode()
        self._latest_promotion_data_cache = {} if real_mode else None
        render_jobs = []

        for store_id, store_name in stores:
            all_products = self._sort_products_for_display(self.db.safe_fetchall(
                "SELECT id, name, title, image_data, sort_order, product_category_label, created_at "
                "FROM products WHERE store_id=? AND COALESCE(is_archived, 0)=0",
                (store_id,),
            ))
            products = [product for product in all_products if product[0] in visible_ids]
            if not products and (all_products or not self._empty_store_visible(store_id)):
                continue

            section, flow_layout = self._create_data_mode_store_section(store_id, store_name)
            if products:
                section.hide()
            render_jobs.extend((store_id, section, flow_layout, product) for product in products)

        self._data_mode_render_jobs = render_jobs
        self._data_mode_render_index = 0
        self._data_mode_render_saved_scroll = saved_scroll
        QTimer.singleShot(0, lambda token=render_token: self._render_next_data_mode_link(token))

    def _render_next_data_mode_link(self, render_token):
        if render_token != getattr(self, "_data_mode_render_token", None):
            return
        index = self._data_mode_render_index
        jobs = self._data_mode_render_jobs
        if index >= len(jobs):
            self._apply_data_mode_store_visibility()
            saved_scroll = self._data_mode_render_saved_scroll
            QTimer.singleShot(0, lambda value=saved_scroll: self.data_mode_scroll.verticalScrollBar().setValue(value))
            QTimer.singleShot(0, self._update_sticky_store_header)
            self._product_margin_metrics_cache = None
            self._product_card_data_cache = None
            self._main_view_change_token = self._current_main_view_change_token()
            append_event(f"ui:data_mode_render:done count={len(jobs)}")
            return

        store_id, section, flow_layout, product = jobs[index]
        self._data_mode_render_index += 1
        product_id, product_code, title, image_data = product[:4]
        try:
            if (
                _qobject_alive(section)
                and self.product_store_map.get(product_id) == store_id
                and not _qobject_alive(self.bubble_product_widgets.get(product_id))
            ):
                bubble = ProductWidget(
                    product_id, product_code, title, image_data, self, display_mode="bubble"
                )
                flow_layout.addWidget(bubble)
                self.bubble_product_widgets[product_id] = bubble
                visible = product_id in self._visible_product_ids()
                show_data_only = (
                    self.is_real_promotion_data_mode()
                    and self.db.get_setting("real_promotion_show_ordered_data_only", "1") == "1"
                )
                if visible and show_data_only:
                    visible = _promotion_link_visible(
                        self.get_latest_promotion_data(store_id, product_code), show_data_only
                    )
                bubble.setVisible(visible)
                search_ids = getattr(self, "current_search_match_ids", None)
                bubble.set_search_highlight(search_ids is not None and product_id in search_ids)
                if visible:
                    section.show()
        except Exception as e:
            append_exception(f"ui:data_mode_render:product_failed product_id={product_id}", error=e)

        QTimer.singleShot(20, lambda token=render_token: self._render_next_data_mode_link(token))

    def refresh_after_product_added(self, product_id, store_id):
        append_event(f"ui:product_added_refresh:start product_id={product_id} store_id={store_id}")
        if self.main_view_mode != "data":
            self.load_data_safe()
            return
        try:
            rows = self.db.safe_fetchall(
                "SELECT id, name, title, image_data FROM products WHERE id=? AND COALESCE(is_archived, 0)=0",
                (product_id,),
            )
            if not rows:
                self._refresh_data_mode_view_if_active()
                return
            product_id, product_code, title, image_data = rows[0]
            section = getattr(self, "_data_mode_store_sections", {}).get(store_id)
            if section is None:
                store_rows = self.db.safe_fetchall("SELECT name FROM stores WHERE id=?", (store_id,))
                if not store_rows:
                    return
                section, flow_layout = self._create_data_mode_store_section(store_id, store_rows[0][0] or "")
            else:
                flow_widget = section.layout().itemAt(1).widget()
                flow_layout = flow_widget.layout() if flow_widget is not None else None
            if not isinstance(flow_layout, OrderedFlowLayout):
                self._refresh_data_mode_view_if_active()
                return

            self.product_store_map[product_id] = store_id
            occupied_rows = set(self.row_data_map) | set(self.row_store_map)
            next_row = max(occupied_rows) + 1 if occupied_rows else 0
            self.row_data_map[next_row] = product_id
            self._prepare_product_card_caches(store_id)

            bubble = ProductWidget(product_id, product_code, title, image_data, self, display_mode="bubble")
            flow_layout.addWidget(bubble)
            self.bubble_product_widgets[product_id] = bubble
            self.visible_product_ids.add(int(product_id))
            self.refresh_store_sheets()
            self.refresh_store_cards(store_id)
            self._reorder_data_mode_bubbles()
            self.apply_tag_filter(close_menu=False, show_message=False)
            append_event(f"ui:product_added_refresh:done product_id={product_id}")
        except Exception as e:
            append_exception("refresh_after_product_added:failed", error=e)
            self._refresh_data_mode_view_if_active()

    def refresh_after_product_deleted(self, product_id, store_id):
        append_event(f"ui:product_deleted_refresh:start product_id={product_id} store_id={store_id}")
        if self.main_view_mode != "data":
            self.load_data_safe()
            return
        try:
            product_id = int(product_id)
            bubble = getattr(self, "bubble_product_widgets", {}).pop(product_id, None)
            if bubble is not None:
                parent_layout = bubble.parentWidget().layout() if bubble.parentWidget() is not None else None
                if isinstance(parent_layout, OrderedFlowLayout):
                    for index in range(parent_layout.count() - 1, -1, -1):
                        item = parent_layout.itemAt(index)
                        if item is not None and item.widget() is bubble:
                            parent_layout.takeAt(index)
                            break
                    parent_layout.invalidate()
                self._prepare_widget_tree_for_delete(bubble)
                self._retire_widget(bubble)

            self.product_store_map.pop(product_id, None)
            self.visible_product_ids.discard(product_id)
            for row, row_product_id in list(self.row_data_map.items()):
                if row_product_id == product_id:
                    self.row_data_map.pop(row, None)
            cache = getattr(self, "_product_card_data_cache", None)
            if isinstance(cache, dict):
                cache.pop(product_id, None)
            metric_cache = getattr(self, "_product_margin_metrics_cache", None)
            if isinstance(metric_cache, dict):
                metric_cache.pop(product_id, None)

            self.refresh_store_sheets()
            self.refresh_store_cards(store_id)
            self._apply_data_mode_store_visibility()
            append_event(f"ui:product_deleted_refresh:done product_id={product_id}")
        except Exception as e:
            append_exception("refresh_after_product_deleted:failed", error=e)
            self.load_data_safe()

    def refresh_after_store_deleted(self, store_id, product_ids):
        append_event(f"ui:store_deleted_refresh:start store_id={store_id}")
        try:
            self._refresh_after_store_deleted(store_id, product_ids)
            append_event(f"ui:store_deleted_refresh:done store_id={store_id}")
        except Exception as e:
            append_exception("refresh_after_store_deleted:failed", error=e)
            self.show_toast("店铺已删除，但界面刷新失败，请重新打开软件")

    def _refresh_after_store_deleted(self, store_id, product_ids):
        if self.main_view_mode != "data":
            self.load_data_safe()
            return
        product_ids = {int(product_id) for product_id in product_ids}
        section = getattr(self, "_data_mode_store_sections", {}).pop(store_id, None)
        section_alive = _qobject_alive(section)
        if section_alive:
            self.data_mode_layout.removeWidget(section)
            self._prepare_widget_tree_for_delete(section)
            self._retire_widget(section)

        for product_id in product_ids:
            bubble = self.bubble_product_widgets.pop(product_id, None)
            if not section_alive and _qobject_alive(bubble):
                self._prepare_widget_tree_for_delete(bubble)
                self._retire_widget(bubble)
            self.product_store_map.pop(product_id, None)
            self.visible_product_ids.discard(product_id)
        for row, product_id in list(self.row_data_map.items()):
            if product_id in product_ids:
                self.row_data_map.pop(row, None)
        for row, row_store_id in list(self.row_store_map.items()):
            if row_store_id == store_id:
                self.row_store_map.pop(row, None)

        state = self.main_table_state
        state.rows = [row for row in state.rows if row.get("store_id") != store_id]
        state.row_by_index = {row["row"]: row for row in state.rows}
        state.product_row_by_id = {
            row["product_id"]: row["row"] for row in state.rows if row.get("product_id") is not None
        }
        state.store_row_by_id.pop(store_id, None)
        self.current_store_filter.discard(store_id)
        self.db.set_setting("store_filter_ids", ",".join(map(str, self.current_store_filter)))
        self.btn_store_filter.setText(
            f"🏪 店铺 ({self._visible_product_row_count()})"
            if self.current_store_filter else "🏪 店铺"
        )
        self._product_card_data_cache = None
        self._product_margin_metrics_cache = None
        self.refresh_store_sheets()
        self.update_store_sheet_selection()
        self._apply_data_mode_store_visibility()
        self.update_daily_task_button_badge()

    def _apply_data_mode_store_visibility(self):
        sections = getattr(self, "_data_mode_store_sections", {})
        if self.main_view_mode != "data" or not sections:
            return False
        visible_ids = self._visible_product_ids()
        show_data_only = (
            hasattr(self, "is_real_promotion_data_mode")
            and self.is_real_promotion_data_mode()
            and self.db.get_setting("real_promotion_show_ordered_data_only", "1") == "1"
        )
        for product_id, bubble in list(getattr(self, "bubble_product_widgets", {}).items()):
            if not _qobject_alive(bubble):
                self.bubble_product_widgets.pop(product_id, None)
                continue
            visible = product_id in visible_ids
            if visible and show_data_only:
                store_id = self.product_store_map.get(product_id) or self.get_product_card_data(product_id).get("store_id")
                visible = _promotion_link_visible(
                    self.get_latest_promotion_data(store_id, bubble.prod_code),
                    show_data_only,
                )
            bubble.setVisible(visible)
        for store_id, section in sections.items():
            if not _qobject_alive(section):
                continue
            store_bubbles = [
                bubble for product_id, bubble in self.bubble_product_widgets.items()
                if self.product_store_map.get(product_id) == store_id and _qobject_alive(bubble)
            ]
            section.setVisible(
                any(not bubble.isHidden() for bubble in store_bubbles)
                or (not store_bubbles and self._empty_store_visible(store_id))
            )
        self.data_mode_layout.invalidate()
        self.data_mode_container.updateGeometry()
        sticky = getattr(self, "data_mode_sticky_header", None)
        if _qobject_alive(sticky):
            sticky.hide()
            self._sticky_store_widget = None
            self._sticky_store_cache_key = None
            QTimer.singleShot(0, self._update_sticky_store_header)
        return True

    def toggle_real_promotion_data_mode(self, *_args):
        enabled = self.btn_real_promotion_mode.isChecked()
        self.db.set_setting("real_promotion_data_mode", "1" if enabled else "0")
        self._update_real_promotion_mode_button_style()
        if self.main_view_mode == "data":
            self._latest_promotion_data_cache = {} if enabled else None
            show_data_only = enabled and self.db.get_setting("real_promotion_show_ordered_data_only", "1") == "1"
            self._apply_real_promotion_display_settings(show_data_only)
            for section in getattr(self, "_data_mode_store_sections", {}).values():
                store_widget = section.findChild(StoreWidget)
                if store_widget is not None:
                    store_widget.refresh_bubble_metrics()
            self._update_sticky_store_header(force=True)
        else:
            self.load_data_safe()

    def _save_frozen_col_width(self, logicalIndex, oldSize, newSize):
        return

    def _save_splitter_frozen_width(self, *_args):
        try:
            if not hasattr(self, "main_splitter"):
                return
            sizes = self.main_splitter.sizes()
            if not sizes:
                return
            saved_width = int(sizes[0])
            self.db.set_setting("col_0_width", max(320, saved_width))
            self.db.set_setting("col_0_width_user_adjusted", "1")
            self._fit_frozen_column_to_viewport()
        except Exception as e:
            print(f"保存冻结区域宽度失败: {e}")

    def get_latest_promotion_data(self, store_id, product_code):
        cache = getattr(self, "_latest_promotion_data_cache", None)
        cache_key = ("product", store_id, str(product_code or ""))
        if isinstance(cache, dict) and cache_key in cache:
            return cache[cache_key]
        cutoff = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        date_key = ("store_date", store_id)
        date_cached = isinstance(cache, dict) and date_key in cache
        store_date = cache.get(date_key) if date_cached else None
        if not date_cached:
            date_rows = self.db.safe_fetchall(
                "SELECT MAX(record_date) FROM promotion_daily_data WHERE store_id=? AND record_date<=?",
                (store_id, cutoff),
            )
            store_date = date_rows[0][0] if date_rows and date_rows[0][0] else ""
            if isinstance(cache, dict):
                cache[date_key] = store_date
        if not store_date:
            if isinstance(cache, dict):
                cache[cache_key] = None
            return None
        rows = self.db.safe_fetchall("""
            SELECT record_date, cost, transaction_amount, net_transaction_amount, net_roi,
                   net_orders, promotion_impression_share, cost_per_net_order,
                   ctr, click_conversion_rate, net_profit, net_margin_rate
            FROM promotion_daily_data
            WHERE store_id=? AND product_id=? AND record_date=?
            LIMIT 1
        """, (store_id, product_code, store_date))
        if not rows:
            rows = self.db.safe_fetchall("""
                SELECT record_date, cost, transaction_amount, net_transaction_amount, net_roi,
                       net_orders, promotion_impression_share, cost_per_net_order,
                       ctr, click_conversion_rate, net_profit, net_margin_rate
                FROM promotion_daily_data
                WHERE store_id=? AND product_id=? AND record_date<=?
                ORDER BY record_date DESC
                LIMIT 1
            """, (store_id, product_code, cutoff))
            if not rows:
                if isinstance(cache, dict):
                    cache[cache_key] = None
                return None
        row = rows[0]
        result = {
            "record_date": row[0],
            "cost": float(row[1] or 0),
            "transaction_amount": float(row[2] or 0),
            "net_transaction_amount": float(row[3] or 0),
            "net_roi": float(row[4] or 0),
            "net_orders": float(row[5] or 0),
            "promotion_impression_share": float(row[6] or 0),
            "cost_per_net_order": float(row[7] or 0),
            "ctr": float(row[8] or 0),
            "click_conversion_rate": float(row[9] or 0),
            "net_profit": None if row[10] is None else float(row[10]),
            "net_margin_rate": None if row[11] is None else float(row[11]),
        }
        if isinstance(cache, dict):
            cache[cache_key] = result
        return result

    def force_refresh_product_widget(self, product_id):
        """根据 product_id 强制刷新对应的 ProductWidget"""
        try:
            self._product_margin_metrics_cache = None
            self._product_card_data_cache = None
            rows = self.db.safe_fetchall("SELECT name, title, image_data FROM products WHERE id=?", (product_id,))
            product_code, product_title, image_data = rows[0] if rows else ("", "", None)

            def refresh_identity(widget):
                if not _qobject_alive(widget):
                    return False
                widget.prod_code = product_code or widget.prod_code
                widget.prod_title = product_title or widget.prod_title
                if hasattr(widget, "code_label"):
                    widget.code_label.setText(str(widget.prod_code))
                if hasattr(widget, "title_label"):
                    widget.title_label.setText(widget.prod_title)
                    widget.title_label.setToolTip(widget.prod_title)
                if hasattr(widget, "set_image_from_data"):
                    widget.set_image_from_data(image_data)
                return True

            task_states = getattr(self, "_product_task_states", None)
            if isinstance(task_states, dict):
                task_states.pop(product_id, None)
            refreshed = False
            bubble = getattr(self, "bubble_product_widgets", {}).get(product_id)
            if _qobject_alive(bubble):
                refresh_identity(bubble)
                bubble.update_product_category_display()
                bubble.update_product_memo_display()
                bubble.update_margin_display()
                bubble.update_promo_badges()
                bubble.update_task_badge()
                bubble._update_bubble_width()
                bubble.update()
                refreshed = True
            for row, prod_id in self.row_data_map.items():
                if prod_id == product_id:
                    widget = self._product_widget_at_row(row)
                    if _qobject_alive(widget):
                        refresh_identity(widget)
                        widget.update_margin_display()
                        widget.update_promo_badges()
                        widget.update_task_badge()
                        widget.update()
                    self.frozen_table.viewport().update()
                    self.frozen_table.update()
                    return
            if refreshed:
                return
            store_id = self.product_store_map.get(product_id)
            if store_id is None:
                store_rows = self.db.safe_fetchall("SELECT store_id FROM products WHERE id=?", (product_id,))
                store_id = store_rows[0][0] if store_rows else None
            if store_id is not None:
                self.refresh_store_cards(store_id)
        except Exception as e:
            print(f"强制刷新ProductWidget失败: {e}")

    def refresh_store_cards(self, store_id=None):
        """Refresh existing product cards after store-scoped data changes."""
        try:
            append_event(f"refresh_store_cards:start store_id={store_id}")
            self._prepare_product_card_caches(store_id)
            refreshed = 0
            for product_id, bubble in list(getattr(self, "bubble_product_widgets", {}).items()):
                if store_id is not None and self.product_store_map.get(product_id) != store_id:
                    continue
                if not _qobject_alive(bubble):
                    self.bubble_product_widgets.pop(product_id, None)
                    continue
                bubble.db = self.db
                bubble.update_product_category_display()
                bubble.update_product_memo_display()
                bubble.update_margin_display()
                bubble.update_promo_badges()
                bubble.update_task_badge()
                bubble._update_bubble_width()
                bubble.update()
                refreshed += 1

            for row, product_id in list(self.row_data_map.items()):
                if store_id is not None and self.product_store_map.get(product_id) != store_id:
                    continue
                widget = self._product_widget_at_row(row)
                if not _qobject_alive(widget):
                    continue
                widget.db = self.db
                widget.update_product_category_display()
                widget.update_product_memo_display()
                widget.update_margin_display()
                widget.update_promo_badges()
                widget.update_task_badge()
                widget.update()
                self._set_product_display_row_height(row)
                refreshed += 1
            self.frozen_table.viewport().update()
            self.frozen_table.update()
            self._refresh_store_margin_widgets(store_id)
            self._update_sticky_store_header(force=True)
            self._main_view_change_token = self._current_main_view_change_token()
            append_event(f"refresh_store_cards:done count={refreshed}")
        except Exception as e:
            append_exception("refresh_store_cards:failed", error=e)
            self.load_data_safe()

    def schedule_pdd_link_refresh(self, store_id=None, product_id=None, created=False):
        if product_id:
            if created and hasattr(self, "refresh_after_product_added"):
                self.refresh_after_product_added(product_id, store_id)
                return
            self.force_refresh_product_widget(product_id)
            self._refresh_store_margin_widgets(store_id)
            return
        if self._pdd_link_refresh_pending:
            return
        self._pdd_link_refresh_pending = True
        QTimer.singleShot(1200, lambda sid=store_id: self._run_pdd_link_refresh(sid))

    def _run_pdd_link_refresh(self, store_id=None):
        if self.is_loading:
            QTimer.singleShot(1200, lambda sid=store_id: self._run_pdd_link_refresh(sid))
            return
        self._pdd_link_refresh_pending = False
        self.load_data_safe()
        self.refresh_store_cards(store_id)

    def _refresh_store_margin_widgets(self, store_id=None):
        def refresh(widget):
            if not isinstance(widget, StoreWidget):
                return
            if store_id is not None and widget.store_id != store_id:
                return
            if getattr(widget, "display_mode", "table") == "bubble":
                widget.refresh_bubble_metrics()
            else:
                widget.refresh_margin_display()

        for row in range(getattr(self.frozen_table, "rowCount", lambda: 0)()):
            refresh(self.frozen_table.cellWidget(row, 0))
        for section in getattr(self, "_data_mode_store_sections", {}).values():
            refresh(section.findChild(StoreWidget))




    def save_scroll_position(self):
        """保存当前滚动位置和选中的商品ID到数据库"""
        if self.main_view_mode == "data" and hasattr(self, "data_mode_scroll"):
            v_value = self.data_mode_scroll.verticalScrollBar().value()
            self.db.set_setting("scroll_vertical", v_value)
            return v_value, 0, None
        v_scroll = self.table.verticalScrollBar()
        h_scroll = self.table.horizontalScrollBar()
        
        v_value = v_scroll.value() if v_scroll else 0
        h_value = h_scroll.value() if h_scroll else 0
        
        self.db.set_setting("scroll_vertical", v_value)
        self.db.set_setting("scroll_horizontal", h_value)
        
        selected_rows = self.table.selectionModel().selectedIndexes()
        selected_product_id = None
        if selected_rows:
            row = selected_rows[0].row()
            if row in self.row_data_map:
                selected_product_id = self.row_data_map[row]
                self.db.set_setting("selected_product_id", selected_product_id)
        
        return v_value, h_value, selected_product_id
    
    def restore_scroll_position(self, scroll_value, selected_product_id, h_scroll_value=None):
        """恢复滚动位置和选中状态"""
        if self.main_view_mode == "data" and hasattr(self, "data_mode_scroll"):
            self.data_mode_scroll.verticalScrollBar().setValue(scroll_value or 0)
            return
        v_scroll = self.table.verticalScrollBar()
        h_scroll = self.table.horizontalScrollBar()
        
        if selected_product_id:
            for row, prod_id in self.row_data_map.items():
                if prod_id == selected_product_id:
                    self._update_frozen_selection_highlight(row)
                    if v_scroll:
                        v_scroll.setValue(scroll_value)
                    if h_scroll and h_scroll_value is not None:
                        h_scroll.setValue(h_scroll_value)
                    return
        
        if v_scroll:
            v_scroll.setValue(scroll_value)
        if h_scroll and h_scroll_value is not None:
            h_scroll.setValue(h_scroll_value)

    def _load_product_operation_records(self, product_id, year, month, day):
        try:
            rows = self.db.safe_fetchall(
                "SELECT records_json FROM records WHERE product_id=? AND year=? AND month=? AND day=?",
                (product_id, year, month, day),
            )
            return json.loads(rows[0][0]) if rows and rows[0][0] else []
        except Exception:
            return []

    def _track_record_dialog(self, dialog):
        self.record_dialogs.append(dialog)
        dialog.destroyed.connect(
            lambda _=None, d=dialog: self.record_dialogs.remove(d)
            if d in self.record_dialogs else None
        )

    def open_product_record_window(self, product_id, year=None, month=None, day=None):
        rows = self.db.safe_fetchall(
            "SELECT name, title, store_id, product_memo FROM products WHERE id=? AND COALESCE(is_archived, 0)=0",
            (product_id,),
        )
        if not rows:
            self.load_data_safe()
            QMessageBox.warning(self, "链接已失效", "当前链接不属于当前账号数据，已刷新主界面。")
            return
        prod_code, prod_title, store_id, product_memo = rows[0]
        store_name = ""
        store_rows = self.db.safe_fetchall("SELECT name FROM stores WHERE id=?", (store_id,))
        if store_rows:
            store_name = store_rows[0][0] or ""
        now = datetime.now()
        year, month, day = year or now.year, month or now.month, day or now.day

        def load_records(y, m, d):
            return self._load_product_operation_records(product_id, y, m, d)

        def save_records(new_data, y, m, d):
            old_records = self._load_product_operation_records(product_id, y, m, d)
            self._sync_main_image_history_from_record_save(product_id, old_records, new_data, y, m, d)
            if new_data:
                data = self._sort_records_by_time(new_data)
                self.db.safe_execute(
                    "INSERT OR REPLACE INTO records (product_id, year, month, day, records_json) VALUES (?, ?, ?, ?, ?)",
                    (product_id, y, m, d, json.dumps(data, ensure_ascii=False)),
                )
            else:
                self.db.safe_execute(
                    "DELETE FROM records WHERE product_id=? AND year=? AND month=? AND day=?",
                    (product_id, y, m, d),
                )
            self.refresh_store_cards(store_id)

        def save_memo(memo):
            self.db.safe_execute("UPDATE products SET product_memo=? WHERE id=?", (memo, product_id))
            self._product_card_data_cache = None

        dialog = OperationRecordDialog(
            load_records(year, month, day), product_id, prod_code, year, month, day,
            save_records, self, store_id=store_id, store_name=store_name,
            load_callback=load_records, save_with_date=True,
            title_prefix=f"链接操作记录 - {prod_code}",
            product_memo=product_memo, memo_save_callback=save_memo,
            quick_reminder_callback=lambda text, when: self.create_product_reminder(
                store_id, product_id, text, when
            ),
        )
        dialog.setAttribute(Qt.WA_DeleteOnClose, True)
        apply_window_icon(dialog, "record")
        self._track_record_dialog(dialog)
        dialog.show()
        dialog.raise_()
        dialog.activateWindow()
        return dialog

    def open_store_record_window(self, store_id, year=None, month=None, day=None):
        store_rows = self.db.safe_fetchall("SELECT name FROM stores WHERE id=?", (store_id,))
        if not store_rows:
            self.load_data_safe()
            QMessageBox.warning(self, "店铺已失效", "当前店铺不属于当前账号数据，已刷新主界面。")
            return
        store_name = store_rows[0][0] or ""
        now = datetime.now()
        year, month, day = year or now.year, month or now.month, day or now.day

        def load_records(y, m, d):
            return self.db.get_store_record(store_id, y, m, d)

        def save_records(new_data, y, m, d):
            if new_data:
                self.db.save_store_record(store_id, y, m, d, self._sort_records_by_time(new_data))
            else:
                self.db.safe_execute(
                    "DELETE FROM store_records WHERE store_id=? AND year=? AND month=? AND day=?",
                    (store_id, y, m, d),
                )
            self.refresh_store_cards(store_id)

        dialog = OperationRecordDialog(
            load_records(year, month, day), store_id, store_name, year, month, day,
            save_records, self, store_id=store_id, store_name=store_name,
            load_callback=load_records, save_with_date=True,
            title_prefix=f"店铺操作记录 - {store_name}",
        )
        dialog.setAttribute(Qt.WA_DeleteOnClose, True)
        apply_window_icon(dialog, "record")
        self._track_record_dialog(dialog)
        dialog.show()
        dialog.raise_()
        dialog.activateWindow()
        return dialog

    def open_store_margin_dialog(self, store_id, store_name):
        """打开店铺毛利管理对话框（供 StoreWidget 调用，避免 widgets 依赖本模块 Dialog）"""
        try:
            current_store = self.db.safe_fetchall("SELECT name FROM stores WHERE id=?", (store_id,))
        except Exception as e:
            QMessageBox.warning(self, "打开失败", f"读取当前账号店铺失败：{e}")
            return

        if not current_store:
            self.load_data_safe()
            QMessageBox.warning(self, "店铺已失效", "当前店铺不属于已切换后的账号数据，已重新刷新主界面。")
            return

        store_name = current_store[0][0] or store_name

        def on_margin_changed(sid, new_margin):
            self.refresh_store_cards(sid)

        existing = self.store_margin_dialogs.get(store_id)
        if existing is not None:
            if existing.isMinimized():
                existing.showNormal()
            else:
                existing.show()
            existing.raise_()
            existing.activateWindow()
            if hasattr(existing, "activate_keyboard_shortcuts"):
                existing.activate_keyboard_shortcuts()
            return
        try:
            dialog = StoreMarginDialog(store_id, store_name, self, None, on_margin_changed)
            apply_window_icon(dialog, "store")
            dialog.setModal(False)
            dialog.setWindowModality(Qt.NonModal)
            dialog.destroyed.connect(lambda _=None, sid=store_id: self.store_margin_dialogs.pop(sid, None))
            self.store_margin_dialogs[store_id] = dialog
            dialog.show()
            dialog.raise_()
            dialog.activateWindow()
        except Exception as e:
            self._log_runtime_exception("open_store_margin_dialog", e)
            QMessageBox.critical(
                self,
                "打开失败",
                f"打开店铺毛利管理失败：{e}\n\n已记录到存档根目录的“崩溃报告.log”"
            )

    def open_promotion_data_for_store(self, store_id):
        rows = self.db.safe_fetchall("SELECT name FROM stores WHERE id=?", (store_id,))
        if not rows:
            QMessageBox.warning(self, "店铺已失效", "当前店铺不存在，已取消打开推广数据分析。")
            return
        existing = self.promotion_data_dialogs.get(store_id)
        if _qobject_alive(existing):
            existing.showNormal() if existing.isMinimized() else existing.show()
            existing.raise_()
            existing.activateWindow()
            return
        try:
            PromotionDataDialog = _import_local_attr("dialogs.promotion_data", "PromotionDataDialog")
            dialog = PromotionDataDialog(store_id, rows[0][0] or f"店铺{store_id}", self.db, self, None)
            dialog.setWindowModality(Qt.NonModal)
            dialog.setAttribute(Qt.WA_DeleteOnClose, True)
            dialog.destroyed.connect(
                lambda _=None, sid=store_id: self.promotion_data_dialogs.pop(sid, None)
            )
            self.promotion_data_dialogs[store_id] = dialog
            dialog.show()
            dialog.raise_()
            dialog.activateWindow()
        except Exception as e:
            self._log_runtime_exception("open_promotion_data_for_store", e)
            QMessageBox.critical(self, "打开失败", f"打开推广数据分析失败：{e}")

    def _log_runtime_exception(self, context, error):
        append_exception(context, error=error)

    def _handle_uncaught_exception(self, exc_type, exc_value, exc_traceback):
        append_exception("uncaught", exc_info=(exc_type, exc_value, exc_traceback))
        traceback.print_exception(exc_type, exc_value, exc_traceback)

    def refresh_store_weight_sync_flag(self, store_id):
        """刷新店铺的权重已同步标签（供 StoreMarginDialog 调用）"""
        for row in range(self.frozen_table.rowCount()):
            widget = self.frozen_table.cellWidget(row, 0)
            if widget and hasattr(widget, 'store_id') and widget.store_id == store_id:
                if hasattr(widget, 'refresh_sync_flag'):
                    widget.refresh_sync_flag()
                break

    def open_product_spec_dialog(self, db, product_id, product_code, product_title, parent):
        """打开规格与毛利对话框（供 StoreMarginDialog 等调用，避免 dialogs 依赖本模块）"""
        existing = getattr(self, "product_spec_dialog", None)
        if existing is not None:
            try:
                if getattr(existing, "product_id", None) == product_id:
                    existing.load_specs()
                    existing._reset_undo_history()
                    if existing.isMinimized():
                        existing.showNormal()
                    else:
                        existing.show()
                    existing.raise_()
                    existing.activateWindow()
                    return existing
                existing.close()
            except RuntimeError:
                pass
            self.product_spec_dialog = None

        dialog = ProductSpecDialog(db, product_id, product_code, product_title, self)
        apply_window_icon(dialog, "spec")
        dialog.main_app = self
        dialog.destroyed.connect(lambda _=None, d=dialog: setattr(self, "product_spec_dialog", None) if getattr(self, "product_spec_dialog", None) is d else None)
        self.product_spec_dialog = dialog
        dialog.show()
        dialog.raise_()
        dialog.activateWindow()
        if hasattr(dialog, "activate_keyboard_shortcuts"):
            dialog.activate_keyboard_shortcuts()
        return dialog

    def open_profit_calculator_dialog(self, margin_rate, avg_price, store_id, store_name, scope, parent, db, return_rate=0.0, quick_mode=False):
        """打开利润计算器对话框（供 StoreMarginDialog 等调用）"""
        dialog = ProfitCalculatorDialog(
            margin_rate, avg_price, store_id, store_name, scope, parent, db,
            return_rate=return_rate, quick_mode=quick_mode,
        )
        apply_window_icon(dialog, "promotion")
        self.profit_calculator_dialog = dialog
        dialog.destroyed.connect(lambda _=None: setattr(self, "profit_calculator_dialog", None))
        dialog.show()

    def _get_product_order_count(self, prod_code, store_id=None):
        """获取商品的单量（从 imported_orders 表汇总）"""
        try:
            cache = getattr(self, "_product_order_count_cache", None)
            key = (store_id, str(prod_code or ""))
            if isinstance(cache, dict) and key in cache:
                return cache.get(key, 0)
            if store_id is not None:
                rows = self.db.safe_fetchall(
                    "SELECT COALESCE(SUM(order_count), 0) FROM imported_orders "
                    "WHERE store_id=? AND product_id=?",
                    (store_id, prod_code),
                )
            else:
                rows = self.db.safe_fetchall(
                    "SELECT COALESCE(SUM(order_count), 0) FROM imported_orders WHERE product_id=?",
                    (prod_code,),
                )
            total = rows[0][0] if rows else 0
            if isinstance(cache, dict):
                cache[key] = total
            return total
        except Exception:
            return 0

    def on_product_sort_changed(self):
        if not hasattr(self, "product_sort_combo"):
            return
        mode = self.product_sort_combo.currentData() or "created_at"
        if mode == self.product_sort_mode:
            return
        self.product_sort_mode = mode
        self.db.set_setting("product_sort_mode", mode)
        self._refresh_product_sort_display()

    def _update_product_sort_direction_button(self):
        descending = getattr(self, "product_sort_descending", True)
        self.btn_product_sort_direction.setText("↓ 降序" if descending else "↑ 升序")
        self.btn_product_sort_direction.setToolTip("当前最新/数值较大在前" if descending else "当前最早/数值较小在前")

    def toggle_product_sort_direction(self):
        self.product_sort_descending = not getattr(self, "product_sort_descending", True)
        self.db.set_setting("product_sort_descending", "1" if self.product_sort_descending else "0")
        self._update_product_sort_direction_button()
        self._refresh_product_sort_display()

    def _refresh_product_sort_display(self):
        if self.main_view_mode == "data":
            self._prepare_product_card_caches()
            try:
                self._reorder_data_mode_bubbles()
            finally:
                self._product_margin_metrics_cache = None
                self._product_card_data_cache = None
        else:
            self.load_data_safe()

    def _reorder_data_mode_bubbles(self):
        reordered_stores = 0
        for store_id, section in getattr(self, "_data_mode_store_sections", {}).items():
            if not _qobject_alive(section):
                continue
            section_layout = section.layout()
            if section_layout is None:
                continue
            flow_widget = None
            for index in range(section_layout.count()):
                item = section_layout.itemAt(index)
                widget = item.widget() if item is not None else None
                if _qobject_alive(widget) and isinstance(widget.layout(), OrderedFlowLayout):
                    flow_widget = widget
                    break
            flow_layout = flow_widget.layout() if flow_widget is not None else None
            if not isinstance(flow_layout, OrderedFlowLayout):
                continue
            flow_layout._items = [
                item for item in flow_layout._items
                if item.widget() is None or _qobject_alive(item.widget())
            ]
            products = self._sort_products_for_display(self.db.safe_fetchall(
                "SELECT id, name, title, image_data, sort_order, product_category_label, created_at "
                "FROM products WHERE store_id=? AND COALESCE(is_archived, 0)=0",
                (store_id,),
            ))
            order = {product[0]: index for index, product in enumerate(products)}
            before = [getattr(item.widget(), "prod_id", None) for item in flow_layout._items]
            flow_layout._items.sort(
                key=lambda item: order.get(getattr(item.widget(), "prod_id", None), len(order))
            )
            if before != [getattr(item.widget(), "prod_id", None) for item in flow_layout._items]:
                reordered_stores += 1
            flow_layout.invalidate()
            flow_layout.setGeometry(flow_widget.rect())
            flow_layout.activate()
            flow_widget.updateGeometry()
            flow_widget.update()
            section_layout.invalidate()
            section.updateGeometry()
        append_event(f"ui:product_sort:reorder mode={self.product_sort_mode} changed_stores={reordered_stores}")
        self.data_mode_layout.invalidate()
        self.data_mode_layout.activate()
        self.data_mode_container.updateGeometry()
        self.data_mode_container.update()

    def refresh_after_product_spec_save(self, product_id):
        if self.main_view_mode != "data":
            QTimer.singleShot(0, self.load_data_safe)
            return
        self.force_refresh_product_widget(product_id)
        store_id = getattr(self, "product_store_map", {}).get(product_id)
        section = getattr(self, "_data_mode_store_sections", {}).get(store_id)
        store_widget = section.findChild(StoreWidget) if _qobject_alive(section) else None
        if _qobject_alive(store_widget):
            store_widget.refresh_bubble_metrics()
        self._reorder_data_mode_bubbles()

    def refresh_external_products(self, product_ids):
        """Refresh products changed by independent dialogs without rebuilding the main view."""
        pending = getattr(self, "_pending_external_product_refreshes", set())
        pending.update(int(pid) for pid in (product_ids or []) if pid)
        self._pending_external_product_refreshes = pending
        if not pending or getattr(self, "_external_product_refresh_scheduled", False):
            return
        self._external_product_refresh_scheduled = True
        QTimer.singleShot(0, self._drain_external_product_refreshes)

    def _drain_external_product_refreshes(self):
        if self.is_loading:
            QTimer.singleShot(150, self._drain_external_product_refreshes)
            return
        pending = getattr(self, "_pending_external_product_refreshes", set())
        batch = list(pending)[:4]
        for product_id in batch:
            pending.discard(product_id)
            self.force_refresh_product_widget(product_id)
        if pending:
            QTimer.singleShot(16, self._drain_external_product_refreshes)
        else:
            self._external_product_refresh_scheduled = False

    def _calculate_product_net_margin(self, product_id):
        try:
            margin_metrics = self.get_product_gross_margin_metrics(product_id)
            gross_margin_pct = margin_metrics.get("gross_margin_pct")
            avg_price = float(margin_metrics.get("avg_final_price") or 0)
            if gross_margin_pct is None:
                return None
            card_data = self.get_product_card_data(product_id)
            if not card_data:
                return None
            current_roi = float(card_data.get("current_roi") or 0)
            return_rate = float(card_data.get("return_rate") or 0)
            is_natural_flow = card_data.get("is_natural_flow") or 0
            is_sitewide_managed = card_data.get("is_sitewide_managed") or 0
            store_id = card_data.get("store_id")
            product_code = card_data.get("product_code")
            roi_input_mode = card_data.get("roi_input_mode") or "roi"
            transaction_bid = card_data.get("transaction_bid") or 0
            sitewide_roi = card_data.get("sitewide_roi") or 0
            effective_roi = sitewide_roi if is_sitewide_managed and not is_natural_flow else current_roi
            if roi_input_mode == "bid" and not is_sitewide_managed and not is_natural_flow and float(transaction_bid or 0) > 0 and avg_price > 0:
                effective_roi = avg_price / float(transaction_bid)
            if not is_natural_flow and effective_roi <= 0:
                return None
            margin_rate_decimal = gross_margin_pct / 100
            if self.is_real_promotion_data_mode() and store_id and product_code:
                promo = self.get_latest_promotion_data(store_id, product_code)
                if not promo:
                    return None
                if promo.get("net_margin_rate") is not None:
                    return float(promo.get("net_margin_rate"))
                cost = float(promo.get("cost") or 0)
                net_amount = float(promo.get("net_transaction_amount") or 0)
                if net_amount > 0:
                    net_profit = net_amount * margin_rate_decimal - cost - net_amount * 0.006
                    return net_profit / net_amount * 100
                return -100.0 if cost > 0 else 0.0
            if is_natural_flow:
                return (margin_rate_decimal * (1 - return_rate / 100) - 0.006) * 100
            return (margin_rate_decimal * (1 - return_rate / 100) - 0.006 - (1 / effective_roi)) * 100
        except Exception as e:
            print(f"计算链接净利率失败: {e}")
            return None

    def _build_product_sort_info(self, product, mode=None):
        product_id, product_code, _title, _image_data, sort_order, category_label = product[:6]
        created_at = str(product[6] or "") if len(product) > 6 else ""
        mode = mode or self.product_sort_mode or "created_at"
        category_label = str(category_label or "").strip()
        if mode == "category":
            calculated_category = self.db.calculate_product_category_label(product_id)
            if calculated_category != category_label:
                category_label = self.db.update_product_category_label(product_id)
        store_id = self.get_product_card_data(product_id).get("store_id")
        order_count = self._get_product_order_count(product_code, store_id)
        net_margin = self._calculate_product_net_margin(product_id) if mode in ("order", "net_margin", "net_profit", "category") else None
        net_profit = self._calculate_product_net_profit(product_id, product_code, order_count, net_margin) if mode == "net_profit" else None
        sort_metrics = (
            self._calculate_product_sort_metrics(product_id)
            if mode in ("gross_margin", "roi", "roi_multiple")
            else {"gross_margin": None, "roi": None, "roi_multiple": None}
        )
        fallback_order = sort_order if sort_order is not None else product_id
        return {
            "product": product,
            "order_count": order_count,
            "net_margin": net_margin,
            "net_profit": net_profit,
            "gross_margin": sort_metrics.get("gross_margin"),
            "roi": sort_metrics.get("roi"),
            "roi_multiple": sort_metrics.get("roi_multiple"),
            "category_label": category_label,
            "created_at": created_at,
            "fallback_order": fallback_order,
            "product_id": product_id,
        }

    def _calculate_product_sort_metrics(self, product_id):
        metrics = {"gross_margin": None, "roi": None, "roi_multiple": None}
        try:
            card_data = self.get_product_card_data(product_id)
            if not card_data:
                return metrics
            current_roi = card_data.get("current_roi") or 0
            return_rate = card_data.get("return_rate") or 0
            is_natural_flow = card_data.get("is_natural_flow") or 0
            is_sitewide_managed = card_data.get("is_sitewide_managed") or 0
            sitewide_roi = card_data.get("sitewide_roi") or 0
            effective_roi = sitewide_roi if is_sitewide_managed and not is_natural_flow else current_roi
            if not is_natural_flow and effective_roi and effective_roi > 0:
                metrics["roi"] = float(effective_roi)
            gross_metrics = self.get_product_gross_margin_metrics(product_id)
            metrics["gross_margin"] = gross_metrics.get("gross_margin_pct")
            if metrics["gross_margin"] is not None:
                net_margin_formula = float(metrics["gross_margin"]) / 100 * (1 - float(return_rate) / 100) - 0.006
                if metrics["roi"] is not None and net_margin_formula > 0:
                    metrics["roi_multiple"] = float(effective_roi) * net_margin_formula
        except Exception as e:
            print(f"计算链接排序指标失败: {e}")
        return metrics

    def _calculate_product_net_profit(self, product_id, product_code, order_count, net_margin):
        if net_margin is None:
            return None
        try:
            if self.is_real_promotion_data_mode():
                store_id = self.product_store_map.get(product_id) or self.get_product_card_data(product_id).get("store_id")
                data = self.get_latest_promotion_data(store_id, product_code) if store_id else None
                if data:
                    if data.get("net_profit") is not None:
                        return float(data["net_profit"])
                    return float(data.get("net_transaction_amount") or 0) * float(net_margin) / 100
                return None
            metrics = self.get_product_gross_margin_metrics(product_id)
            avg_price = metrics.get("avg_final_price")
            if avg_price is None:
                return None
            return float(avg_price) * float(order_count or 0) * float(net_margin) / 100
        except Exception as e:
            print(f"计算链接净利润排序值失败: {e}")
            return None

    def _descending_metric_sort_key(self, info, metric_name):
        value = info.get(metric_name)
        return (
            1 if value is None else 0,
            -(value if value is not None else -10**9),
            -info["order_count"],
            1 if info["net_margin"] is None else 0,
            -(info["net_margin"] if info["net_margin"] is not None else -10**9),
            info["fallback_order"],
            info["product_id"],
        )

    def _product_metric_sort_key(self, info):
        net_margin = info["net_margin"]
        return (
            -info["order_count"],
            1 if net_margin is None else 0,
            -(net_margin if net_margin is not None else -10**9),
            info["fallback_order"],
            info["product_id"],
        )

    def _prepare_product_card_caches(self, store_id=None):
        if store_id is not None and not isinstance(getattr(self, "_product_card_data_cache", None), dict):
            store_id = None
        store_filter = " AND p.store_id=?" if store_id is not None else ""
        params = (store_id,) if store_id is not None else ()
        rows = self.db.safe_fetchall(
            """SELECT p.id, p.name, p.coupon_amount, p.new_customer_discount, p.current_roi,
                      p.return_rate, p.net_break_even_roi, p.is_limited_time,
                      p.is_marketing, COALESCE(p.marketing_activity, ''), p.is_natural_flow, p.is_sitewide_managed,
                      p.store_id, COALESCE(p.roi_input_mode, 'roi'),
                      COALESCE(p.transaction_bid, 0), p.product_category_label,
                      p.link_type, p.product_memo, COALESCE(s.sitewide_roi, 0),
                      COALESCE(p.is_violation, 0)
               FROM products p
               LEFT JOIN stores s ON s.id=p.store_id
               WHERE COALESCE(p.is_archived, 0)=0""" + store_filter,
            params,
        )
        fields = (
            "product_code", "coupon_amount", "new_customer_discount", "current_roi", "return_rate",
            "net_break_even_roi", "is_limited_time", "is_marketing", "marketing_activity", "is_natural_flow",
            "is_sitewide_managed", "store_id",
            "roi_input_mode", "transaction_bid", "product_category_label", "link_type",
            "product_memo", "sitewide_roi", "is_violation",
        )
        card_data = {
            row[0]: dict(zip(fields, row[1:])) for row in rows
        }
        product_ids = set(card_data)
        if store_id is None or not isinstance(getattr(self, "_product_card_data_cache", None), dict):
            self._product_card_data_cache = card_data
            self._product_margin_metrics_cache = self.db.calculate_products_gross_margin_metrics(product_ids)
            task_flags = {
                product_id: (bool(garbage), bool(waste), bool(pending), garbage_content or "")
                for product_id, garbage, waste, pending, garbage_content in self.db.safe_fetchall(
                    """SELECT product_id,
                              MAX(CASE WHEN task_content LIKE ? THEN 1 ELSE 0 END),
                              MAX(CASE WHEN task_content LIKE ? THEN 1 ELSE 0 END),
                              MAX(CASE WHEN task_content NOT LIKE ? AND task_content NOT LIKE ? THEN 1 ELSE 0 END),
                              MAX(CASE WHEN task_content LIKE ? THEN task_content ELSE '' END)
                       FROM daily_tasks WHERE is_completed=0 GROUP BY product_id""",
                    ("【垃圾链接】%", "【废物链接】%", "【垃圾链接】%", "【废物链接】%", "【垃圾链接】%"),
                )
            }
            reminder_ids = {
                row[0] for row in self.db.safe_fetchall(
                    "SELECT DISTINCT product_id FROM task_reminders WHERE is_reminded=0"
                )
            }
            self._product_task_states = {
                product_id: (
                    *task_flags.get(product_id, (False, False, False))[:2],
                    task_flags.get(product_id, (False, False, False))[2] or product_id in reminder_ids,
                    task_flags.get(product_id, (False, False, False, ""))[3],
                )
                for product_id in product_ids
            }
            self._product_order_count_cache = {
                (sid, str(product_code or "")): float(order_count or 0)
                for sid, product_code, order_count in self.db.safe_fetchall(
                    "SELECT store_id, product_id, SUM(order_count) FROM imported_orders "
                    "GROUP BY store_id, product_id"
                )
            }
            return

        cache = self._product_card_data_cache
        for product_id, data in list(cache.items()):
            if data.get("store_id") == store_id:
                cache.pop(product_id, None)
        cache.update(card_data)

        metric_cache = getattr(self, "_product_margin_metrics_cache", None)
        if not isinstance(metric_cache, dict):
            metric_cache = {}
            self._product_margin_metrics_cache = metric_cache
        for product_id in product_ids:
            metric_cache.pop(product_id, None)
        metric_cache.update(self.db.calculate_products_gross_margin_metrics(product_ids))

        if not isinstance(getattr(self, "_product_task_states", None), dict):
            self._product_task_states = {}
        if product_ids:
            placeholders = ",".join("?" for _ in product_ids)
            ids = tuple(product_ids)
            task_flags = {
                product_id: (bool(garbage), bool(waste), bool(pending), garbage_content or "")
                for product_id, garbage, waste, pending, garbage_content in self.db.safe_fetchall(
                    f"""SELECT product_id,
                               MAX(CASE WHEN task_content LIKE ? THEN 1 ELSE 0 END),
                               MAX(CASE WHEN task_content LIKE ? THEN 1 ELSE 0 END),
                               MAX(CASE WHEN task_content NOT LIKE ? AND task_content NOT LIKE ? THEN 1 ELSE 0 END),
                               MAX(CASE WHEN task_content LIKE ? THEN task_content ELSE '' END)
                        FROM daily_tasks WHERE is_completed=0
                          AND product_id IN ({placeholders}) GROUP BY product_id""",
                    ("【垃圾链接】%", "【废物链接】%", "【垃圾链接】%", "【废物链接】%", "【垃圾链接】%", *ids),
                )
            }
            reminder_ids = {
                row[0] for row in self.db.safe_fetchall(
                    f"SELECT DISTINCT product_id FROM task_reminders WHERE is_reminded=0 AND product_id IN ({placeholders})",
                    ids,
                )
            }
            for product_id in product_ids:
                self._product_task_states[product_id] = (
                    *task_flags.get(product_id, (False, False, False))[:2],
                    task_flags.get(product_id, (False, False, False))[2] or product_id in reminder_ids,
                    task_flags.get(product_id, (False, False, False, ""))[3],
                )

        if not isinstance(getattr(self, "_product_order_count_cache", None), dict):
            self._product_order_count_cache = {}
        self._product_order_count_cache = {
            key: value for key, value in self._product_order_count_cache.items() if key[0] != store_id
        }
        self._product_order_count_cache.update({
            (sid, str(product_code or "")): float(order_count or 0)
            for sid, product_code, order_count in self.db.safe_fetchall(
                "SELECT store_id, product_id, SUM(order_count) FROM imported_orders "
                "WHERE store_id=? GROUP BY store_id, product_id",
                (store_id,),
            )
        })

    def get_product_card_data(self, product_id):
        cache = getattr(self, "_product_card_data_cache", None)
        if isinstance(cache, dict) and product_id in cache:
            return cache.get(product_id, {})
        rows = self.db.safe_fetchall(
            """SELECT p.name, p.coupon_amount, p.new_customer_discount, p.current_roi,
                      p.return_rate, p.net_break_even_roi, p.is_limited_time,
                      p.is_marketing, COALESCE(p.marketing_activity, ''), p.is_natural_flow, p.is_sitewide_managed,
                      p.store_id, COALESCE(p.roi_input_mode, 'roi'),
                      COALESCE(p.transaction_bid, 0), p.product_category_label,
                      p.link_type, p.product_memo, COALESCE(s.sitewide_roi, 0),
                      COALESCE(p.is_violation, 0)
               FROM products p LEFT JOIN stores s ON s.id=p.store_id WHERE p.id=?""",
            (product_id,),
        )
        if not rows:
            return {}
        fields = (
            "product_code", "coupon_amount", "new_customer_discount", "current_roi", "return_rate",
            "net_break_even_roi", "is_limited_time", "is_marketing", "marketing_activity", "is_natural_flow",
            "is_sitewide_managed", "store_id",
            "roi_input_mode", "transaction_bid", "product_category_label", "link_type",
            "product_memo", "sitewide_roi", "is_violation",
        )
        data = dict(zip(fields, rows[0]))
        if isinstance(cache, dict):
            cache[product_id] = data
        return data

    def get_product_gross_margin_metrics(self, product_id, fresh=False):
        cache = getattr(self, "_product_margin_metrics_cache", None)
        if fresh:
            metrics = self.db.calculate_product_gross_margin_metrics(product_id)
            if isinstance(cache, dict):
                cache[product_id] = metrics
            return metrics
        if cache is None:
            return self.db.calculate_product_gross_margin_metrics(product_id)
        if product_id not in cache:
            cache[product_id] = self.db.calculate_product_gross_margin_metrics(product_id)
        return cache[product_id]

    def _sort_products_for_display(self, products_raw, mode=None):
        mode = mode or self.product_sort_mode or "created_at"
        infos = [self._build_product_sort_info(product, mode) for product in products_raw]
        product_ids = [info["product_id"] for info in infos]
        task_states = getattr(self, "_product_task_states", None)
        task_sort_groups = {}
        if isinstance(task_states, dict):
            task_sort_groups = {
                product_id: 2 if task_states.get(product_id, (False, False))[1]
                else 1 if task_states.get(product_id, (False, False))[0] else 0
                for product_id in product_ids
            }
        elif product_ids:
            placeholders = ",".join("?" for _ in product_ids)
            task_sort_groups = dict(self.db.safe_fetchall(
                f"""SELECT product_id,
                           MAX(CASE
                               WHEN task_content LIKE '【废物链接】%' THEN 2
                               WHEN task_content LIKE '【垃圾链接】%' THEN 1
                               ELSE 0
                           END)
                    FROM daily_tasks
                    WHERE is_completed=0
                      AND product_id IN ({placeholders})
                      AND (task_content LIKE '【垃圾链接】%' OR task_content LIKE '【废物链接】%')
                    GROUP BY product_id""",
                tuple(product_ids),
            ))
        for info in infos:
            info["task_sort_group"] = int(task_sort_groups.get(info["product_id"], 0) or 0)
            if info["task_sort_group"] == 2 and info["net_margin"] is None:
                info["net_margin"] = self._calculate_product_net_margin(info["product_id"])
        if mode == "created_at":
            infos.sort(key=lambda info: (info["created_at"], info["product_id"]), reverse=True)
        elif mode == "net_margin":
            infos.sort(
                key=lambda info: (
                    1 if info["net_margin"] is None else 0,
                    -(info["net_margin"] if info["net_margin"] is not None else -10**9),
                    -info["order_count"],
                    info["fallback_order"],
                    info["product_id"],
                )
            )
        elif mode == "net_profit":
            infos.sort(
                key=lambda info: (
                    1 if info["net_profit"] is None else 0,
                    -(info["net_profit"] if info["net_profit"] is not None else -10**9),
                    -info["order_count"],
                    info["fallback_order"],
                    info["product_id"],
                )
            )
        elif mode == "gross_margin":
            infos.sort(key=lambda info: self._descending_metric_sort_key(info, "gross_margin"))
        elif mode == "roi":
            infos.sort(key=lambda info: self._descending_metric_sort_key(info, "roi"))
        elif mode == "roi_multiple":
            infos.sort(key=lambda info: self._descending_metric_sort_key(info, "roi_multiple"))
        elif mode == "category":
            group_stats = {}
            for info in infos:
                label = info["category_label"]
                if not label:
                    continue
                stat = group_stats.setdefault(label, {"max_order": 0, "max_margin": None})
                stat["max_order"] = max(stat["max_order"], info["order_count"])
                margin = info["net_margin"]
                if margin is not None and (stat["max_margin"] is None or margin > stat["max_margin"]):
                    stat["max_margin"] = margin
            infos.sort(
                key=lambda info: (
                    1 if not info["category_label"] else 0,
                    -group_stats.get(info["category_label"], {}).get("max_order", 0),
                    1 if group_stats.get(info["category_label"], {}).get("max_margin") is None else 0,
                    -(
                        group_stats.get(info["category_label"], {}).get("max_margin")
                        if group_stats.get(info["category_label"], {}).get("max_margin") is not None
                        else -10**9
                    ),
                    info["category_label"],
                    *self._product_metric_sort_key(info),
                )
            )
        else:
            infos.sort(key=self._product_metric_sort_key)
        if not getattr(self, "product_sort_descending", True):
            infos.reverse()
        infos.sort(
            key=lambda info: (
                info["task_sort_group"],
                1 if info["task_sort_group"] == 2 and info["net_margin"] is None else 0,
                -(info["net_margin"] if info["task_sort_group"] == 2 and info["net_margin"] is not None else 0),
                info["fallback_order"] if info["task_sort_group"] == 2 else 0,
                info["product_id"] if info["task_sort_group"] == 2 else 0,
            )
        )
        return [info["product"] for info in infos]

    def _load_data_mode_safe(self, restore_position=True):
        if self.is_loading:
            return
        append_event(f"load_data_safe:data:start restore={restore_position}")
        account_switching = getattr(self, "_is_switching_local_account", False)
        saved_scroll = 0
        if restore_position and hasattr(self, "data_mode_scroll"):
            saved_scroll = self.data_mode_scroll.verticalScrollBar().value()
        self.is_loading = True
        try:
            self.lbl_month.setText(f"{self.year}年 {self.month}月")
            self.refresh_store_sheets()
            self.table.blockSignals(True)
            self.frozen_table.blockSignals(True)
            self.table.setRowCount(0)
            self.frozen_table.setRowCount(0)
            self.table.setColumnCount(0)
            self.frozen_table.setColumnCount(0)

            self.row_data_map.clear()
            self.row_store_map.clear()
            self.product_store_map.clear()
            self.main_table_state.clear()

            row_idx = 0
            stores = self.db.safe_fetchall("SELECT id, name FROM stores ORDER BY sort_order")
            for store_id, store_name in stores:
                self.row_store_map[row_idx] = store_id
                self.main_table_state.add_store(row_idx, store_id, store_name, self.STORE_ROW_HEIGHT)
                row_idx += 1
                products = self.db.safe_fetchall(
                    "SELECT id, name, title, image_data, sort_order, product_category_label, created_at "
                    "FROM products WHERE store_id=? AND COALESCE(is_archived, 0)=0",
                    (store_id,),
                )
                for product in products:
                    product_id, product_code, product_title = product[:3]
                    self.row_data_map[row_idx] = product_id
                    self.product_store_map[product_id] = store_id
                    self.main_table_state.add_product(
                        row_idx, store_id, product_id, product_code, product_title, self.PRODUCT_ROW_HEIGHT
                    )
                    row_idx += 1

            self._set_visible_product_ids(self.product_store_map.keys())
        except Exception as e:
            print(f"加载纯数据失败: {e}")
            QMessageBox.critical(self, "错误", f"加载纯数据失败: {e}")
        finally:
            self.table.blockSignals(False)
            self.frozen_table.blockSignals(False)
            append_event("load_data_safe:data:done")
            self.is_loading = False
            self._product_margin_metrics_cache = None
            self._product_card_data_cache = None
            self._main_view_change_token = self._current_main_view_change_token()

        if account_switching:
            append_event("load_data_safe:data:post_timers_skipped_for_account_switch")
            return
        QTimer.singleShot(30, self._reapply_search_and_filters_after_load)
        if restore_position and hasattr(self, "data_mode_scroll"):
            QTimer.singleShot(60, lambda value=saved_scroll: self.data_mode_scroll.verticalScrollBar().setValue(value))
    def load_data_safe(self, restore_position=True):
        """安全加载数据，防止闪退"""
        if self.is_loading:
            return  # 防止重复加载
        if self.main_view_mode != "operation":
            self._load_data_mode_safe(restore_position)
            return
        append_event(f"load_data_safe:start restore={restore_position}")
        account_switching = getattr(self, "_is_switching_local_account", False)
        
        # 保存当前滚动位置和选中状态
        v_scroll_value, h_scroll_value, selected_product_id = 0, 0, None
        if restore_position:
            v_scroll_value, h_scroll_value, selected_product_id = self.save_scroll_position()
        
        # 尝试从数据库读取上次保存的位置（用于切换月份时保持位置）
        saved_v = self.db.get_setting("scroll_vertical", 0)
        saved_h = self.db.get_setting("scroll_horizontal", 0)
        saved_product = self.db.get_setting("selected_product_id")
        try:
            if not restore_position:
                v_scroll_value = int(saved_v) if saved_v else 0
                h_scroll_value = int(saved_h) if saved_h else 0
            selected_product_id = saved_product
        except:
            pass
        
        self.is_loading = True
        self._prepare_product_card_caches()
        render_operation_view = self.main_view_mode == "operation"
        
        try:
            self.lbl_month.setText(f"{self.year}年 {self.month}月")
            days_in_month = calendar.monthrange(self.year, self.month)[1]
            
            stores = self.db.safe_fetchall("SELECT id, name FROM stores ORDER BY sort_order")
            self.refresh_store_sheets()
            
            # 临时禁用信号以提高性能
            self.table.blockSignals(True)
            self.frozen_table.blockSignals(True)
            
            # 清空表格
            for row in range(self.frozen_table.rowCount()):
                widget = self.frozen_table.cellWidget(row, 0)
                if widget:
                    self._prepare_widget_tree_for_delete(widget)
                    self.frozen_table.removeCellWidget(row, 0)
                    self._retire_widget(widget)
            while self.table.rowCount() > 0:
                self.table.removeRow(0)
            while self.frozen_table.rowCount() > 0:
                self.frozen_table.removeRow(0)
            
            total_cols = days_in_month
            self.table.setColumnCount(total_cols)
            self.frozen_table.setColumnCount(1)
            
            headers = []
            today = datetime.now()
            self._today_col = -1
            for i in range(1, days_in_month + 1):
                dt = datetime(self.year, self.month, i)
                weekday_idx = dt.weekday()
                weekday_name = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][weekday_idx]
                headers.append(f"{i}号 {weekday_name}")
                if self.year == today.year and self.month == today.month and i == today.day:
                    self._today_col = i - 1
            self.table.setHorizontalHeaderLabels(headers)
            self.frozen_table.setHorizontalHeaderLabels(["商品信息"])

            if hasattr(self, 'today_delegate'):
                self.today_delegate.set_today_col(self._today_col)

            # 恢复表头样式，并给今日列设置高亮
            header_style = "QHeaderView::section { background-color: #e9f0e6; color: #26362f; border: 1px solid #c8d4c4; padding: 3px; }"
            if self._today_col >= 0:
                header_style += f"QTableWidget::item:column({self._today_col}):selected, QHeaderView::section:column({self._today_col}) {{ background-color: #dfeeda; }}"
                header_style += f"QHeaderView::section:column({self._today_col}) {{ background-color: #cfe4c8; font-weight: bold; }}"
            self.table.horizontalHeader().setStyleSheet(header_style)
            self.frozen_table.horizontalHeader().setStyleSheet(header_style)
            self.table.horizontalHeader().setFixedHeight(self.TABLE_HEADER_HEIGHT)
            self.frozen_table.horizontalHeader().setFixedHeight(self.TABLE_HEADER_HEIGHT)
            
            self._suppress_col0_width_save = True
            self.frozen_table.setColumnWidth(0, max(1, self.frozen_table.viewport().width()))
            self._suppress_col0_width_save = False
            if not account_switching:
                QTimer.singleShot(0, self._fit_frozen_column_to_viewport)
            shrink_date_cols = self.db.get_setting("date_col_width_minus_20_applied", "0") != "1"
            shrink_date_cols_30 = self.db.get_setting("date_col_width_minus_30_applied", "0") != "1"
            for i in range(total_cols):
                setting_key = f"col_{i + 1}_width"
                width = int(self.db.get_setting(setting_key, 200))
                if shrink_date_cols:
                    width = max(120, width - 20)
                if shrink_date_cols_30:
                    width = max(120, width - 30)
                if shrink_date_cols or shrink_date_cols_30:
                    self.db.set_setting(setting_key, width)
                self.table.setColumnWidth(i, width)
            if shrink_date_cols:
                self.db.set_setting("date_col_width_minus_20_applied", "1")
            if shrink_date_cols_30:
                self.db.set_setting("date_col_width_minus_30_applied", "1")
                
            self.row_data_map.clear()
            self.row_store_map.clear()
            self.product_store_map.clear()
            self.main_table_state.clear()
            row_idx = 0
            
            for s_idx, store in enumerate(stores):
                store_id, store_name = store
                self.table.insertRow(row_idx)
                self.frozen_table.insertRow(row_idx)
                
                if render_operation_view:
                    store_widget = StoreWidget(store_id, store_name, self)
                    store_widget.setStyleSheet("""
                        #StoreWidget {
                            background-color: #e6f0e8;
                        }
                    """)
                    self.frozen_table.setCellWidget(row_idx, 0, store_widget)
                self.row_store_map[row_idx] = store_id
                self.main_table_state.add_store(row_idx, store_id, store_name, self.STORE_ROW_HEIGHT)
                
                if render_operation_view:
                    for day in range(1, days_in_month + 1):
                        item = QTableWidgetItem()
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        item.setBackground(QBrush(QColor("#e6f0e8")))
                        self.table.setItem(row_idx, day - 1, item)
                    self.render_store_records(row_idx, store_id, days_in_month)
                
                self.table.setRowHeight(row_idx, self.STORE_ROW_HEIGHT)
                self.frozen_table.setRowHeight(row_idx, self.STORE_ROW_HEIGHT)
                row_idx += 1
                
                products_raw = self.db.safe_fetchall(
                    "SELECT id, name, title, image_data, sort_order, product_category_label, created_at FROM products WHERE store_id=? AND COALESCE(is_archived, 0)=0",
                    (store_id,),
                )
                products = self._sort_products_for_display(products_raw) if render_operation_view else products_raw
                for prod in products:
                    p_id, p_code, p_title, p_img = prod[:4]  # 注意这里：p_code是商品ID，p_title是商品标题
                    self.table.insertRow(row_idx)
                    self.frozen_table.insertRow(row_idx)
                    
                    self.row_data_map[row_idx] = p_id
                    self.product_store_map[p_id] = store_id

                    if render_operation_view:
                        p_widget = ProductWidget(p_id, p_code, p_title, p_img, self)
                        self.frozen_table.setCellWidget(row_idx, 0, self._wrap_product_widget(p_widget))
                        self._set_product_display_row_height(row_idx)
                        row_height = self._get_product_display_row_height(row_idx)
                        self.render_records_for_product(row_idx, p_id, days_in_month)
                    else:
                        row_height = self.PRODUCT_ROW_HEIGHT
                        self.table.setRowHeight(row_idx, row_height)
                        self.frozen_table.setRowHeight(row_idx, row_height)
                    self.main_table_state.add_product(
                        row_idx,
                        store_id,
                        p_id,
                        p_code,
                        p_title,
                        row_height,
                    )
                    
                    row_idx += 1

            if render_operation_view:
                QApplication.processEvents()
                self.frozen_table.repaint()
            
        except Exception as e:
            print(f"加载数据失败: {e}")
            QMessageBox.critical(self, "错误", f"加载数据失败: {e}")
        finally:
            self.table.blockSignals(False)
            self.frozen_table.blockSignals(False)
            append_event("load_data_safe:done")
            self.is_loading = False
            self._product_margin_metrics_cache = None
            self._product_card_data_cache = None
            self._main_view_change_token = self._current_main_view_change_token()
            
        # 恢复滚动位置和选中状态
        if account_switching:
            append_event("load_data_safe:post_timers_skipped_for_account_switch")
            return
        QTimer.singleShot(10, lambda: self.restore_scroll_position(v_scroll_value, selected_product_id, h_scroll_value))
        
        QTimer.singleShot(10, self.update_frozen_geometry)
        QTimer.singleShot(30, self._reapply_search_and_filters_after_load)
    def _format_operation_record_for_cell(self, record):
        time_text = record.get("time", "")
        body = str(record.get("text", "") or "")
        body = re.sub(r"^自动记录[:：]\s*", "", body).strip()
        if not body and record.get("changes"):
            body = "；".join([c.get("text", "") for c in record.get("changes", []) if c.get("text")])
        return f"[{time_text}] {body}" if time_text else body

    def _format_operation_record_tooltip(self, records):
        lines = []
        for record in records or []:
            if not isinstance(record, dict):
                continue
            time_text = str(record.get("time", "") or "")
            changes = record.get("changes") or []
            if changes:
                for change in changes:
                    if not isinstance(change, dict):
                        continue
                    change_time = str(change.get("time", "") or time_text)
                    metric = str(change.get("metric", "") or "")
                    text = str(change.get("text", "") or "").strip()
                    prefix = f"[{change_time}]" if change_time else ""
                    if metric:
                        prefix = f"{prefix} {metric}".strip()
                    lines.append(f"{prefix}\n{text}" if prefix and text else (prefix or text))
            else:
                line = self._format_operation_record_for_cell(record)
                if line:
                    lines.append(line)
        return "\n\n".join([line for line in lines if line])

    def _apply_record_cell_style(self, item):
        font = QFont("Microsoft YaHei", 11)
        font.setBold(True)
        item.setFont(font)
        item.setForeground(QColor("#1f2d3d"))
        item.setTextAlignment(Qt.AlignTop | Qt.AlignLeft)

    def _get_product_display_row_height(self, row):
        widget = self._product_widget_at_row(row)
        if widget and hasattr(widget, "recommended_row_height"):
            try:
                return widget.recommended_row_height(self.PRODUCT_ROW_HEIGHT)
            except Exception as e:
                print(f"计算商品行高失败: {e}")
        return self.PRODUCT_ROW_HEIGHT

    def _set_product_display_row_height(self, row):
        row_height = self._get_product_display_row_height(row)
        self.table.setRowHeight(row, row_height)
        self.frozen_table.setRowHeight(row, row_height)
        state_row = getattr(self, "main_table_state", None)
        if state_row and row in state_row.row_by_index:
            state_row.row_by_index[row]["row_height"] = row_height

    def update_product_row_height(self, prod_id):
        for row, row_prod_id in self.row_data_map.items():
            if row_prod_id == prod_id:
                self._set_product_display_row_height(row)
                return

    def render_records_for_product(self, row, prod_id, days):
        try:
            # 1. 从数据库获取最新记录
            records = self.db.safe_fetchall(
                "SELECT day, records_json FROM records WHERE product_id=? AND year=? AND month=?", 
                (prod_id, self.year, self.month)
            )
            rec_dict = {}
            
            for r in records:
                try:
                    rec_dict[r[0]] = json.loads(r[1])
                except:
                    rec_dict[r[0]] = []
            
            # 定义基础参数
            for day in range(1, days + 1):
                cell_data = rec_dict.get(day, [])
                col = day - 1
                
                # 构建显示文本
                if cell_data:
                    display_text = "\n".join([self._format_operation_record_for_cell(item) for item in cell_data])
                else:
                    display_text = ""
                
                # 【关键修复 1】获取或创建单元格
                item = self.table.item(row, col)
                if not item:
                    item = QTableWidgetItem()
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    self.table.setItem(row, col, item)
                
                # ❌ 已删除 item.setWordWrap(True) 因为这方法不存在
                
                # 强制更新文本
                item.setText(display_text)
                item.setToolTip(self._format_operation_record_tooltip(cell_data) if cell_data else display_text)
                item.setData(Qt.UserRole, cell_data)
                
                # 确保文字靠上对齐，方便多行显示
                self._apply_record_cell_style(item)

            self._set_product_display_row_height(row)

        except Exception as e:
            print(f"渲染记录失败：{e}")
            import traceback
            traceback.print_exc()
    
    def render_store_records(self, row, store_id, days):
        try:
            rec_dict = self.db.get_store_record(store_id, self.year, self.month, 0)
            
            for day in range(1, days + 1):
                cell_data = rec_dict.get(day, [])
                col = day - 1
                
                if cell_data:
                    display_text = "\n".join([self._format_operation_record_for_cell(item) for item in cell_data])
                else:
                    display_text = ""
                
                item = self.table.item(row, col)
                if not item:
                    item = QTableWidgetItem()
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    self.table.setItem(row, col, item)
                
                item.setText(display_text)
                item.setToolTip(self._format_operation_record_tooltip(cell_data) if cell_data else display_text)
                item.setData(Qt.UserRole, cell_data)
                self._apply_record_cell_style(item)

        except Exception as e:
            print(f"渲染店铺记录失败：{e}")
            import traceback
            traceback.print_exc()
    
    def open_editor(self, row, col):
        if col < 0:
            return
        day = col + 1
        
        if row in self.row_store_map:
            self.open_store_record_editor(row, day)
            return
        
        if row not in self.row_data_map:
            return
            
        prod_id = self.row_data_map[row]
        try:
            res = self.db.safe_fetchall("SELECT records_json FROM records WHERE product_id=? AND year=? AND month=? AND day=?",
                                   (prod_id, self.year, self.month, day))
            records = json.loads(res[0][0]) if res else []
        except:
            records = []

        prod_code = str(prod_id)
        prod_store_id = self.product_store_map.get(prod_id)
        try:
            prod_res = self.db.safe_fetchall("SELECT name, product_memo FROM products WHERE id=?", (prod_id,))
            if prod_res and prod_res[0][0]:
                prod_code = prod_res[0][0]
            product_memo = prod_res[0][1] if prod_res else ""
        except:
            product_memo = ""

        store_name = ""
        if prod_store_id:
            try:
                store_res = self.db.safe_fetchall("SELECT name FROM stores WHERE id=?", (prod_store_id,))
                if store_res and store_res[0][0]:
                    store_name = store_res[0][0]
            except:
                pass

        def save_callback(new_data):
            try:
                self._sync_main_image_history_from_record_save(prod_id, records, new_data, self.year, self.month, day)
                if new_data:
                    new_data = self._sort_records_by_time(new_data)
                    self.db.safe_execute("INSERT OR REPLACE INTO records (product_id, year, month, day, records_json) VALUES (?, ?, ?, ?, ?)",
                                    (prod_id, self.year, self.month, day, json.dumps(new_data)))
                else:
                    self.db.safe_execute("DELETE FROM records WHERE product_id=? AND year=? AND month=? AND day=?",
                                    (prod_id, self.year, self.month, day))
                self.load_data_safe()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"保存记录失败：{e}")
                self.load_data_safe()

        def save_memo(memo):
            self.db.safe_execute("UPDATE products SET product_memo=? WHERE id=?", (memo, prod_id))
            self._product_card_data_cache = None

        dialog = OperationRecordDialog(
            records, prod_id, prod_code, self.year, self.month, day, save_callback,
            self, store_id=prod_store_id, store_name=store_name,
            product_memo=product_memo, memo_save_callback=save_memo,
            quick_reminder_callback=lambda text, when: self.create_product_reminder(
                prod_store_id, prod_id, text, when
            ),
        )
        apply_window_icon(dialog, "record")
        dialog.exec_()

    def _sort_records_by_time(self, records):
        def key(record):
            text = str(record.get("time", "") if isinstance(record, dict) else "")
            try:
                hour, minute = text.split(":", 1)
                return (int(hour), int(minute))
            except Exception:
                return (99, 99)
        return sorted(records or [], key=key)

    def _main_image_history_id_from_record(self, record):
        if not isinstance(record, dict):
            return None
        history_id = record.get("history_id")
        if history_id:
            return history_id
        for change in record.get("changes", []) or []:
            if isinstance(change, dict) and change.get("type") == "main_carousel_image" and change.get("history_id"):
                return change.get("history_id")
        return None

    def _main_image_history_ids_from_records(self, records):
        ids = set()
        for record in records or []:
            history_id = self._main_image_history_id_from_record(record)
            if history_id:
                ids.add(history_id)
        return ids

    def _sync_main_image_history_from_record_save(self, product_id, old_records, new_records, year, month, day):
        old_ids = self._main_image_history_ids_from_records(old_records)
        new_ids = self._main_image_history_ids_from_records(new_records)
        deleted_ids = old_ids - new_ids
        for history_id in deleted_ids:
            self.db.safe_execute(
                "DELETE FROM product_image_history WHERE id=? AND product_id=?",
                (history_id, product_id)
            )

        for record in new_records or []:
            history_id = self._main_image_history_id_from_record(record)
            if not history_id:
                continue
            time_text = record.get("time", "")
            if not time_text:
                continue
            changed_at = f"{year:04d}-{month:02d}-{day:02d} {time_text}:00"
            self.db.safe_execute(
                "UPDATE product_image_history SET changed_at=? WHERE id=? AND product_id=?",
                (changed_at, history_id, product_id)
            )

    def open_store_record_editor(self, row, day):
        store_id = self.row_store_map[row]
        
        records = self.db.get_store_record(store_id, self.year, self.month, day)
        
        store_name = ""
        try:
            store_res = self.db.safe_fetchall("SELECT name FROM stores WHERE id=?", (store_id,))
            if store_res and store_res[0][0]:
                store_name = store_res[0][0]
        except:
            pass
        
        def save_callback(new_data):
            try:
                if new_data:
                    self.db.save_store_record(store_id, self.year, self.month, day, new_data)
                else:
                    self.db.safe_execute("DELETE FROM store_records WHERE store_id=? AND year=? AND month=? AND day=?",
                                    (store_id, self.year, self.month, day))
                self.load_data_safe()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"保存记录失败：{e}")
                self.load_data_safe()

        dialog = OperationRecordDialog(records, store_id, store_name, self.year, self.month, day, save_callback, self, store_id=store_id, store_name=store_name)
        apply_window_icon(dialog, "record")
        dialog.exec_()
    
    def _activate_cell(self, row, col):
        """辅助方法：安全地选中并刷新指定单元格"""
        try:
            if 0 <= row < self.table.rowCount() and 0 <= col < self.table.columnCount():
                # 选中单元格
                self.table.setCurrentCell(row, col)
                # 强制视口重绘
                self.table.viewport().update()
                # 把焦点给表格，防止焦点还停留在空气里
                self.table.setFocus()
        except:
            pass

    def open_product_spec_dialog_from_table(self, row, col):
        """双击格子打开规格管理弹窗"""
        if row not in self.row_data_map:
            return
        
        product_id = self.row_data_map[row]  # 数据库自增ID
        
        # 从冻结列控件获取商品ID和标题
        widget = self._product_widget_at_row(row)
        prod_code = "未知ID"
        prod_title = "未知标题"
        
        if widget:
            # 从ProductWidget中获取
            prod_code = widget.prod_code      # 用户输入的ID
            prod_title = widget.prod_title    # 商品标题
        
        self.open_product_spec_dialog(self.db, product_id, prod_code, prod_title, self)

    def prev_month(self):
        try:
            if self.month == 1:
                self.month = 12
                self.year -= 1
            else:
                self.month -= 1
            self.load_data_safe()
        except Exception as e:
            print(f"切换上个月失败: {e}")

    def next_month(self):
        try:
            if self.month == 12:
                self.month = 1
                self.year += 1
            else:
                self.month += 1
            self.load_data_safe()
        except Exception as e:
            print(f"切换下个月失败: {e}")

    def add_store(self):
        try:
            name, ok = QInputDialog.getText(self, "添加店铺", "请输入店铺名称:")
            name = name.strip() if ok else ""
            if name:
                result = self.db.safe_fetchall("SELECT MAX(sort_order) FROM stores")
                max_order = result[0][0] if result and result[0][0] is not None else 0
                cursor = self.db.safe_execute("INSERT INTO stores (name, sort_order) VALUES (?, ?)", (name, max_order + 1))
                store_id = int(cursor.lastrowid)
                self.refresh_store_sheets()
                if self.main_view_mode == "data" and hasattr(self, "data_mode_layout"):
                    if not hasattr(self, "_data_mode_store_sections"):
                        self._data_mode_store_sections = {}
                    if store_id not in self._data_mode_store_sections:
                        self._create_data_mode_store_section(store_id, name)
                    occupied_rows = set(self.row_data_map) | set(self.row_store_map)
                    row = max(occupied_rows) + 1 if occupied_rows else 0
                    self.row_store_map[row] = store_id
                    self.main_table_state.add_store(row, store_id, name, self.STORE_ROW_HEIGHT)
                    self._apply_data_mode_store_visibility()
                self.show_toast(f"已添加店铺：{name}")
        except Exception as e:
            print(f"添加店铺失败: {e}")
            QMessageBox.warning(self, "错误", f"添加店铺失败: {e}")

    def record_product_operation(self, product_db_id, text, metric="操作记录", old="", new="", change_type="product_operation"):
        """给商品当天日期格追加一条结构化操作记录。"""
        try:
            if not product_db_id:
                return
            now = datetime.now()
            time_str = now.strftime("%H:%M")
            year, month, day = now.year, now.month, now.day
            text = str(text or "").strip()
            metric = str(metric or "操作记录").strip()
            if str(change_type).startswith("pdd_price"):
                metric, text = "改价", f"改价：{new or text.split('：', 1)[-1]}"
            elif str(change_type).startswith("pdd_promotion"):
                metric, text = "修改推广", f"修改推广：{old} → {new}"
            record = {
                "time": time_str,
                "text": text,
                "changes": [{
                    "time": time_str,
                    "metric": metric,
                    "old": "" if old is None else str(old),
                    "new": "" if new is None else str(new),
                    "text": text,
                    "type": change_type,
                }]
            }
            rows = self.db.safe_fetchall(
                "SELECT records_json FROM records WHERE product_id=? AND year=? AND month=? AND day=?",
                (product_db_id, year, month, day)
            )
            records = []
            if rows and rows[0][0]:
                try:
                    records = json.loads(rows[0][0])
                except Exception:
                    records = []
            if not isinstance(records, list):
                records = []
            if not str(change_type).startswith("task_"):
                cutoff = now - timedelta(minutes=10)
                records = [item for item in records if not (
                    isinstance(item, dict)
                    and any(str(change.get("metric", "")) == metric for change in item.get("changes", []) if isinstance(change, dict))
                    and self._record_time_on_date(item, now) >= cutoff
                )]
            records.append(record)
            records = self._sort_records_by_time(records)
            self.db.safe_execute(
                "INSERT OR REPLACE INTO records (product_id, year, month, day, records_json) VALUES (?, ?, ?, ?, ?)",
                (product_db_id, year, month, day, json.dumps(records, ensure_ascii=False))
            )
            return record
        except Exception as e:
            print(f"记录商品操作失败: {e}")

    def _record_time_on_date(self, record, now):
        try:
            hour, minute = str(record.get("time", "")).split(":", 1)
            return now.replace(hour=int(hour), minute=int(minute), second=0, microsecond=0)
        except (TypeError, ValueError):
            return datetime.min

    def add_product(self, store_id, copy_from_id=None):
        """添加商品 - 支持手动输入商品ID和标题，copy_from_id用于复制同款"""
        append_event(f"ui:add_product:dialog_open store_id={store_id} copy_from_id={copy_from_id}")
        try:
            # 如果是复制模式，获取原商品信息
            copy_data = {}
            if copy_from_id:
                rows = self.db.safe_fetchall(
                    "SELECT name, title, coupon_amount, new_customer_discount, image_path, product_memo FROM products WHERE id=?",
                    (copy_from_id,)
                )
                if rows and rows[0]:
                    copy_data = {
                        'name': rows[0][0],
                        'title': rows[0][1],
                        'coupon_amount': rows[0][2],
                        'new_customer_discount': rows[0][3],
                        'image_path': rows[0][4],
                        'product_memo': rows[0][5]
                    }
            
            # 创建一个自定义对话框
            dialog = QDialog(self)
            dialog.setWindowTitle("添加新商品 - 复制同款" if copy_from_id else "添加新商品")
            apply_window_icon(dialog, "main")
            dialog.setFixedSize(500, 350)
            
            layout = QVBoxLayout(dialog)
            
            # 商品ID输入
            id_layout = QHBoxLayout()
            id_layout.addWidget(QLabel("商品ID:"))
            id_input = QLineEdit()
            id_input.setPlaceholderText("请输入商品ID（用于搜索和绑定链接）")
            id_layout.addWidget(id_input)
            layout.addLayout(id_layout)
            
            # 商品标题输入
            title_layout = QHBoxLayout()
            title_layout.addWidget(QLabel("商品标题:"))
            title_input = QLineEdit()
            title_input.setPlaceholderText("请输入商品标题")
            if copy_data:
                title_input.setText(copy_data.get('title', ''))
            title_layout.addWidget(title_input)
            layout.addLayout(title_layout)
            
            # 优惠券金额
            coupon_layout = QHBoxLayout()
            coupon_layout.addWidget(QLabel("优惠券金额:"))
            coupon_input = QLineEdit()
            coupon_input.setPlaceholderText("请输入优惠券金额")
            if copy_data and copy_data.get('coupon_amount'):
                coupon_input.setText(str(copy_data.get('coupon_amount')))
            coupon_layout.addWidget(coupon_input)
            layout.addLayout(coupon_layout)
            
            # 新客立减
            newcust_layout = QHBoxLayout()
            newcust_layout.addWidget(QLabel("新客立减:"))
            newcust_input = QLineEdit()
            newcust_input.setPlaceholderText("请输入新客立减金额")
            if copy_data and copy_data.get('new_customer_discount'):
                newcust_input.setText(str(copy_data.get('new_customer_discount')))
            newcust_layout.addWidget(newcust_input)
            layout.addLayout(newcust_layout)
            
            # 提示标签
            tip_text = "提示：复制同款模式 - 除商品ID外，其他信息已从原商品复制，请修改ID后保存。"
            if not copy_from_id:
                tip_text = "提示：商品ID是您手动输入的链接ID，用于搜索；商品标题是商品名称。"
            tip_label = QLabel(tip_text)
            tip_label.setStyleSheet("color: #666; font-size: 10px;")
            tip_label.setWordWrap(True)
            layout.addWidget(tip_label)
            
            # 按钮
            btn_layout = QHBoxLayout()
            btn_pdd_code_fetch = QPushButton("抓取添加编码")
            btn_pdd_code_fetch.setToolTip("打开当前店铺的拼多多商家端网页，并打开抓取添加编码窗口")
            btn_pdd_code_fetch.clicked.connect(lambda _checked=False, sid=store_id: self.open_pdd_code_fetch_for_store(sid))
            btn_pdd_code_fetch.setStyleSheet(
                "QPushButton { background-color: #f39c12; color: white; padding: 6px 12px; border-radius: 4px; font-weight: bold; }"
                "QPushButton:hover { background-color: #d68910; }"
            )
            btn_ok = QPushButton("确定")
            btn_ok.clicked.connect(dialog.accept)
            btn_cancel = QPushButton("取消")
            btn_cancel.clicked.connect(dialog.reject)
            btn_layout.addWidget(btn_pdd_code_fetch)
            btn_layout.addStretch()
            btn_layout.addWidget(btn_ok)
            btn_layout.addWidget(btn_cancel)
            layout.addLayout(btn_layout)
            
            # 显示对话框
            if dialog.exec_() != QDialog.Accepted:
                append_event("ui:add_product:cancel")
                return
                
            product_id = id_input.text().strip()
            product_title = title_input.text().strip()
            coupon_amount = coupon_input.text().strip()
            new_customer_discount = newcust_input.text().strip()
            
            if not product_id:
                QMessageBox.warning(self, "提示", "商品ID不能为空！")
                return
            if not product_title:
                QMessageBox.warning(self, "提示", "商品标题不能为空！")
                return
            
            # 检查商品ID是否已存在
            existing = self.db.safe_fetchall("SELECT id FROM products WHERE name=?", (product_id,))
            if existing:
                QMessageBox.warning(self, "提示", f"商品ID '{product_id}' 已存在，请使用不同的ID！")
                return
            
            # 获取当前店铺的最大排序值
            result = self.db.safe_fetchall("SELECT MAX(sort_order) FROM products WHERE store_id=?", (store_id,))
            max_order = result[0][0] if result and result[0][0] is not None else 0
            
            # 插入数据库
            self.db.safe_execute(
                "INSERT INTO products (store_id, name, title, coupon_amount, new_customer_discount, image_path, sort_order, product_memo, is_natural_flow) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)",
                (store_id, product_id, product_title, 
                 float(coupon_amount) if coupon_amount else None,
                 float(new_customer_discount) if new_customer_discount else None,
                 copy_data.get('image_path') if copy_from_id else None,
                 max_order + 1,
                 copy_data.get('product_memo') if copy_from_id else None)
            )
            
            # 获取新插入商品的数据库自增ID（不是用户输入的商品ID）
            new_product_db_id = self.db.safe_fetchall("SELECT last_insert_rowid()")[0][0]
            
            # 如果是复制模式，复制规格信息（使用数据库自增ID）
            if copy_from_id:
                specs = self.db.safe_fetchall(
                    "SELECT spec_name, spec_code, sale_price, weight_percent, is_locked FROM product_specs WHERE product_id=? ORDER BY id",
                    (copy_from_id,)
                )
                if specs:
                    for spec in specs:
                        self.db.safe_execute(
                            "INSERT INTO product_specs (product_id, spec_name, spec_code, sale_price, weight_percent, is_locked) VALUES (?, ?, ?, ?, ?, ?)",
                            (new_product_db_id, spec[0], spec[1], spec[2], spec[3], spec[4])
                        )
            
            # 显示成功提示
            append_event(f"ui:add_product:inserted product_id={new_product_db_id} store_id={store_id}")
            self.show_toast(f"✅ 商品添加成功\nID: {product_id}\n标题: {product_title}")
            self.record_product_operation(
                new_product_db_id,
                f"新建链接：商品ID {product_id}，标题：{product_title}",
                metric="新建链接",
                old="",
                new=product_id,
                change_type="product_created",
            )
            self.refresh_after_product_added(new_product_db_id, store_id)
            
        except Exception as e:
            print(f"添加商品失败: {e}")
            append_exception("ui:add_product:failed", error=e)
            QMessageBox.warning(self, "错误", f"添加商品失败: {e}")
            
    def perform_search(self):
        self.apply_realtime_search()

    def _split_search_terms(self, text):
        return split_search_terms(text)

    def apply_realtime_search(self):
        try:
            query = self.search_input.text().strip().casefold()
            if not query:
                self.clear_search_filter()
                return

            had_search_filter = getattr(self, "current_search_match_ids", None) is not None
            self._stop_exact_search_flash()
            self.current_search_match_ids = None
            self.clear_search_highlight()

            product_ids = [pid for pid in self.row_data_map.values() if pid]
            if not product_ids:
                self.clear_search_filter()
                self.show_toast("没有找到对应的内容", 500)
                return

            placeholders = ",".join(["?"] * len(product_ids))
            rows = self.db.safe_fetchall(
                f"SELECT id, name, title, product_memo FROM products WHERE id IN ({placeholders})",
                tuple(product_ids)
            )
            exact_ids = {
                pid for pid, name, _title, _memo in rows
                if str(name or "").strip().casefold() == query
            }
            if len(exact_ids) == 1:
                if had_search_filter:
                    self.apply_tag_filter(close_menu=False, show_message=False)
                exact_id = next(iter(exact_ids))
                for row, pid in self.row_data_map.items():
                    if pid == exact_id:
                        self._start_exact_search_flash(row)
                        return

            terms = self._split_search_terms(query)
            matching_ids = {
                pid for pid, name, title, memo in rows
                if all_terms_match(terms, name, title, memo)
            }
            if not matching_ids:
                if had_search_filter:
                    self.apply_tag_filter(close_menu=False, show_message=False)
                self.show_toast("没有找到对应的内容", 500)
                return
            self.current_search_match_ids = matching_ids
            self.apply_tag_filter(close_menu=False, show_message=False)
            matching_rows = {
                row for row, pid in self.row_data_map.items()
                if pid in self.current_search_match_ids
            }
            if matching_rows:
                self._scroll_to_first_search_match(matching_rows)
        except Exception as e:
            print(f"实时搜索失败: {e}")

    def clear_search_filter(self):
        had_search_filter = getattr(self, "current_search_match_ids", None) is not None
        self._stop_exact_search_flash()
        if not had_search_filter and not getattr(self, "_search_highlighted_rows", set()):
            return
        self.current_search_match_ids = None
        self.clear_search_highlight()
        if had_search_filter:
            self.apply_tag_filter(close_menu=False, show_message=False)

    def clear_search_highlight(self):
        try:
            for row in list(getattr(self, "_search_highlighted_rows", set())):
                self._set_row_search_highlight(row, False)
            self._search_highlighted_rows = set()
        except Exception as e:
            print(f"清除搜索高亮失败: {e}")

    def _set_row_search_highlight(self, row, active):
        widget = self._product_widget_at_row(row)
        if widget and hasattr(widget, "set_search_highlight"):
            widget.set_search_highlight(active)

        bubble = self.bubble_product_widgets.get(self.row_data_map.get(row))
        if bubble and hasattr(bubble, "set_search_highlight"):
            bubble.set_search_highlight(active)

        color = QBrush(QColor("#fff8d8")) if active else QBrush()
        for col in range(self.table.columnCount()):
            item = self.table.item(row, col)
            if item:
                item.setBackground(color)

    def _start_exact_search_flash(self, row):
        self._stop_exact_search_flash()
        if row not in self.row_data_map:
            return
        if not hasattr(self, "_exact_search_flash_timer"):
            self._exact_search_flash_timer = QTimer(self)
            self._exact_search_flash_timer.setInterval(250)
            self._exact_search_flash_timer.timeout.connect(self._advance_exact_search_flash)
        self._exact_search_flash_row = row
        self._exact_search_flash_remaining = 8
        self._exact_search_flash_on = True
        self._search_highlighted_rows = {row}
        self._set_row_search_highlight(row, True)
        self._scroll_to_first_search_match({row})
        self._exact_search_flash_timer.start()

    def _advance_exact_search_flash(self):
        row = getattr(self, "_exact_search_flash_row", None)
        if row is None:
            return
        self._exact_search_flash_remaining -= 1
        if self._exact_search_flash_remaining <= 0:
            self._stop_exact_search_flash()
            return
        self._exact_search_flash_on = not self._exact_search_flash_on
        self._set_row_search_highlight(row, self._exact_search_flash_on)

    def _stop_exact_search_flash(self):
        timer = getattr(self, "_exact_search_flash_timer", None)
        if timer and timer.isActive():
            timer.stop()
        row = getattr(self, "_exact_search_flash_row", None)
        if row is not None:
            self._set_row_search_highlight(row, False)
        self._exact_search_flash_row = None
        self._exact_search_flash_remaining = 0
        self._exact_search_flash_on = False
        if row is not None:
            self._search_highlighted_rows.discard(row)

    def _scroll_to_first_search_match(self, rows, retries=6, defer=True):
        try:
            if not rows:
                return
            if self.main_view_mode == "data":
                if defer:
                    QTimer.singleShot(
                        80,
                        lambda target_rows=set(rows), remaining=retries: self._scroll_to_first_search_match(
                            target_rows, remaining, False
                        ),
                    )
                    return
                for row in sorted(rows):
                    bubble = self.bubble_product_widgets.get(self.row_data_map.get(row))
                    if bubble:
                        bar = self.data_mode_scroll.verticalScrollBar()
                        top = bubble.mapTo(self.data_mode_container, QPoint(0, 0)).y()
                        target = top + bubble.height() // 2 - self.data_mode_scroll.viewport().height() // 2
                        bar.setValue(max(bar.minimum(), min(bar.maximum(), target)))
                        return
                if retries > 0:
                    QTimer.singleShot(
                        60,
                        lambda target_rows=set(rows), remaining=retries - 1: self._scroll_to_first_search_match(
                            target_rows, remaining, False
                        ),
                    )
                return
            visible_rows = [
                row for row in sorted(rows)
                if not self.table.isRowHidden(row) and not self.frozen_table.isRowHidden(row)
            ]
            target_row = visible_rows[0] if visible_rows else sorted(rows)[0]
            h_scroll = self.table.horizontalScrollBar()
            saved_x = h_scroll.value() if h_scroll else None
            target_index = self.frozen_table.model().index(target_row, 0)
            self.frozen_table.scrollTo(target_index, QAbstractItemView.PositionAtCenter)
            if h_scroll and saved_x is not None:
                h_scroll.setValue(saved_x)
            self.update_frozen_geometry()
        except Exception as e:
            print(f"搜索跳转失败: {e}")

    def _reapply_search_and_filters_after_load(self):
        try:
            if self.search_input.text().strip():
                self.apply_realtime_search()
            elif getattr(self, "current_search_match_ids", None) is not None or getattr(self, "_search_highlighted_rows", set()):
                self.clear_search_filter()
            if (
                getattr(self, "current_category_filter", "")
                or getattr(self, "current_filter_tags", None)
                or getattr(self, "current_store_filter", None)
            ):
                self.apply_tag_filter(close_menu=False, show_message=False)
        except Exception as e:
            print(f"刷新搜索和筛选状态失败: {e}")
        finally:
            self._refresh_data_mode_view_if_active()
    
    def on_filter_toggle(self, state):
        if state == Qt.Unchecked:
            # 取消勾选时清除筛选
            self.clear_filter()
    
    def _create_filter_button(self, text, icon_name=None, color="#333"):
        """创建筛选按钮，带图标"""
        icons_dir = os.path.join(os.path.dirname(__file__), "icons")
        btn = QPushButton(text)
        
        if icon_name:
            icon_path = os.path.join(icons_dir, icon_name)
            if os.path.exists(icon_path):
                btn.setIcon(QIcon(icon_path))
        
        btn.setStyleSheet(f"""
            QPushButton {{
                border: 1px solid {color};
                background-color: transparent;
                color: {color};
                border-radius: 4px;
                padding: 6px 12px;
                text-align: left;
                icon-size: 20px;
            }}
            QPushButton:checked {{
                background-color: {color};
                color: white;
            }}
            QPushButton:hover {{
                background-color: {color};
                color: white;
            }}
        """)
        return btn
    
    def get_all_product_ids_with_current_store(self):
        """获取当前视图所有商品的ID（不受筛选影响）"""
        try:
            return [pid for pid in self.product_store_map.keys() if pid]
        except Exception as e:
            print(f"获取商品ID失败: {e}")
            return []

    def _prepare_widget_tree_for_delete(self, widget):
        cards = ([widget] if isinstance(widget, (ProductWidget, StoreWidget)) else [])
        cards += widget.findChildren(ProductWidget) + widget.findChildren(StoreWidget)
        for card in cards:
            card._disposing = True
            timer = getattr(card, "_code_click_timer", None)
            if timer is not None:
                timer.stop()
            for child in [card, *card.findChildren(QWidget)]:
                child.removeEventFilter(card)
                QApplication.removePostedEvents(child)

    def _clear_layout_widgets(self, layout):
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget:
                self._prepare_widget_tree_for_delete(widget)
                self._retire_widget(widget)

    def _retire_widget(self, widget):
        if getattr(self, "_is_switching_local_account", False) and self._is_account_card_widget(widget):
            # ponytail: park old account cards; PyQt crashes deleting these event-filtered widgets during refresh.
            widget.hide()
            widget.setParent(self._retired_widget_parking)
            QApplication.removePostedEvents(widget)
            self._account_switch_retained_widgets.append(widget)
            return
        widget.hide()
        widget.deleteLater()

    def _is_account_card_widget(self, widget):
        return isinstance(widget, (ProductWidget, StoreWidget)) or bool(widget.findChildren((ProductWidget, StoreWidget)))

    def _get_current_category_counts(self, keyword=""):
        product_ids = self.get_all_product_ids_with_current_store()
        if not product_ids:
            return []
        placeholders = ",".join(["?"] * len(product_ids))
        rows = self.db.safe_fetchall(
            f"""
            SELECT product_category_label, COUNT(*)
            FROM products
            WHERE id IN ({placeholders})
              AND COALESCE(product_category_label, '') <> ''
            GROUP BY product_category_label
            ORDER BY COUNT(*) DESC, product_category_label ASC
            """,
            tuple(product_ids)
        )
        keyword = str(keyword or "").strip().lower()
        result = []
        for label, count in rows:
            label_text = str(label or "").strip()
            if not label_text:
                continue
            if keyword and not text_matches(keyword, label_text):
                continue
            result.append((label_text, count))
        return result

    def refresh_category_filter_candidates(self):
        try:
            if not hasattr(self, "category_candidate_layout"):
                return
            self._clear_layout_widgets(self.category_candidate_layout)
            keyword = self.category_filter_input.text().strip() if hasattr(self, "category_filter_input") else ""
            if not keyword and not self.current_category_filter:
                hint = QLabel("输入关键字后显示类型候选")
                hint.setStyleSheet("color: #999; font-size: 12px; padding: 4px;")
                self.category_candidate_layout.addWidget(hint)
                return
            categories = self._get_current_category_counts(keyword)[:12]
            if not categories:
                hint = QLabel("没有匹配的商品类型")
                hint.setStyleSheet("color: #999; font-size: 12px; padding: 4px;")
                self.category_candidate_layout.addWidget(hint)
                return
            for label, count in categories:
                btn = QPushButton(f"{label} ({count})")
                btn.setCheckable(True)
                btn.setChecked(label == self.current_category_filter)
                btn.setStyleSheet("""
                    QPushButton {
                        border: 1px solid #3498db;
                        background-color: transparent;
                        color: #2c3e50;
                        border-radius: 4px;
                        padding: 5px 8px;
                        text-align: left;
                    }
                    QPushButton:checked {
                        background-color: #3498db;
                        color: white;
                    }
                    QPushButton:hover {
                        background-color: #3498db;
                        color: white;
                    }
                """)
                btn.clicked.connect(lambda checked=False, value=label: self.set_category_filter(value))
                self.category_candidate_layout.addWidget(btn)
        except Exception as e:
            print(f"刷新商品类型候选失败: {e}")

    def set_category_filter(self, category_label):
        self.current_category_filter = str(category_label or "").strip()
        if hasattr(self, "category_filter_input"):
            self.category_filter_input.blockSignals(True)
            self.category_filter_input.setText(self.current_category_filter)
            self.category_filter_input.blockSignals(False)
        self.refresh_category_filter_candidates()
        self.apply_tag_filter(close_menu=False)

    def clear_category_filter(self):
        self.current_category_filter = ""
        if hasattr(self, "category_filter_input"):
            self.category_filter_input.blockSignals(True)
            self.category_filter_input.clear()
            self.category_filter_input.blockSignals(False)
        self.refresh_category_filter_candidates()
        self.apply_tag_filter(close_menu=False)

    def _product_matches_category_filter(self, product_id):
        if not self.current_category_filter:
            return True
        rows = self.db.safe_fetchall(
            "SELECT product_category_label FROM products WHERE id=?",
            (product_id,)
        )
        label = str(rows[0][0] or "").strip() if rows else ""
        return label == self.current_category_filter
    
    def show_tag_filter_menu(self):
        """显示标签筛选下拉菜单"""
        append_event("ui:tag_filter_menu:toggle")
        if self.tag_filter_menu.isVisible():
            self.tag_filter_menu.hide()
            return
        self.refresh_category_filter_candidates()
        btn_rect = self.btn_tag_filter.rect()
        global_pos = self.btn_tag_filter.mapToGlobal(QPoint(0, btn_rect.bottom()))
        self.tag_filter_menu.adjustSize()
        window_left = self.mapToGlobal(QPoint(8, 0)).x()
        window_right = self.mapToGlobal(QPoint(self.width() - self.tag_filter_menu.width() - 8, 0)).x()
        self.tag_filter_menu.move(max(window_left, min(global_pos.x(), window_right)), global_pos.y())
        self.tag_filter_menu.show()
        self.tag_filter_menu.raise_()
        QTimer.singleShot(0, self.category_filter_input.setFocus)

    def show_store_filter_menu(self):
        """显示店铺筛选下拉菜单"""
        self.store_filter_menu = QDialog(self)
        self.store_filter_menu.setWindowFlags(Qt.FramelessWindowHint | Qt.Popup)
        self.store_filter_menu.setStyleSheet("""
            QDialog {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 5px;
            }
            QCheckBox {
                padding: 5px;
                spacing: 8px;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
            }
        """)

        filter_layout = QVBoxLayout(self.store_filter_menu)
        filter_layout.setContentsMargins(10, 10, 10, 10)
        filter_layout.setSpacing(5)

        filter_title = QLabel("🏪 选择筛选店铺")
        filter_title.setStyleSheet("font-weight: bold; font-size: 14px; color: #2c3e50; padding-bottom: 5px;")
        filter_layout.addWidget(filter_title)

        filter_layout.addWidget(QLabel("<hr>"))

        self.store_checkboxes = {}
        stores = self.db.safe_fetchall("SELECT id, name FROM stores ORDER BY sort_order")
        saved_filter_ids = self.db.get_setting("store_filter_ids", "")
        if saved_filter_ids:
            saved_filter_set = set(int(x) for x in saved_filter_ids.split(",") if x)
        else:
            saved_filter_set = set()
        for store_id, store_name in stores:
            cb = QCheckBox(store_name)
            cb.setCheckable(True)
            if store_id in saved_filter_set:
                cb.setChecked(True)
            cb.stateChanged.connect(lambda state, sid=store_id: self.apply_store_filter(sid))
            self.store_checkboxes[store_id] = cb
            filter_layout.addWidget(cb)

        filter_layout.addWidget(QLabel("<hr>"))

        btn_layout = QHBoxLayout()
        btn_save_filter = QPushButton("💾 保存筛选")
        btn_save_filter.setStyleSheet("""
            QPushButton { background-color: #27ae60; color: white; border: none; padding: 8px 16px; border-radius: 3px; font-weight: bold; }
            QPushButton:hover { background-color: #219a52; }
        """)
        btn_save_filter.clicked.connect(lambda: self.apply_store_filter(close_menu=True))
        btn_clear_filter = QPushButton("清空")
        btn_clear_filter.setStyleSheet("""
            QPushButton { background-color: #95a5a6; color: white; border: none; padding: 8px 16px; border-radius: 3px; font-weight: bold; }
            QPushButton:hover { background-color: #7f8c8d; }
        """)
        btn_clear_filter.clicked.connect(self.clear_store_filter_selection)
        btn_layout.addWidget(btn_save_filter)
        btn_layout.addWidget(btn_clear_filter)
        filter_layout.addLayout(btn_layout)

        self.current_store_filter = saved_filter_set.copy()
        self.store_filter_menu_selected_store = None

        btn_rect = self.btn_store_filter.rect()
        global_pos = self.btn_store_filter.mapToGlobal(QPoint(0, btn_rect.bottom()))
        self.store_filter_menu.move(global_pos)
        self.store_filter_menu.exec_()

    def apply_store_filter(self, store_id=None, close_menu=False, show_message=True):
        """应用店铺筛选

        Args:
            store_id: 如果指定，则只切换该店铺的选中状态
            close_menu: 是否关闭筛选菜单
        """
        append_event(f"ui:apply_store_filter:start store_id={store_id} close={close_menu}")
        handled_data_mode_visibility = False
        try:
            if store_id is not None:
                checkbox = self.store_checkboxes.get(store_id)
                if checkbox:
                    if checkbox.isChecked():
                        self.current_store_filter.add(store_id)
                    else:
                        self.current_store_filter.discard(store_id)

            if close_menu and self.store_filter_menu:
                self.store_filter_menu.close()

            if not self.current_store_filter:
                handled_data_mode_visibility = bool(self.clear_store_filter())
                return

            self.update_store_sheet_selection()
            if self._has_active_product_filter_inputs():
                self.db.set_setting("store_filter_ids", ",".join(map(str, self.current_store_filter)))
                self.apply_tag_filter(close_menu=False, show_message=False)
                handled_data_mode_visibility = self.main_view_mode == "data" and bool(getattr(self, "_data_mode_store_sections", {}))
                visible_count = self._visible_product_row_count()
                self.btn_store_filter.setText(f"🏪 店铺 ({visible_count})")
                if show_message:
                    self.show_toast(f"店铺筛选: 显示 {visible_count} 个商品")
                return

            self._set_visible_product_ids(
                pid for pid, sid in self.product_store_map.items()
                if sid in self.current_store_filter
            )

            filtered_count = self._visible_product_row_count()
            self.btn_store_filter.setText(f"🏪 店铺 ({filtered_count})")
            self.db.set_setting("store_filter_ids", ",".join(map(str, self.current_store_filter)) if self.current_store_filter else "")
            self.update_store_sheet_selection()
            if show_message:
                self.show_toast(f"店铺筛选: 显示 {filtered_count} 个商品")
            if self.main_view_mode == "data" and not self._has_active_product_filter_inputs():
                handled_data_mode_visibility = self._apply_data_mode_store_visibility()

        except Exception as e:
            print(f"应用店铺筛选失败: {e}")
            import traceback
            traceback.print_exc()
            append_exception("ui:apply_store_filter:failed", error=e)
            QMessageBox.warning(self, "筛选失败", f"店铺筛选出错: {e}")
        finally:
            if not handled_data_mode_visibility:
                self._refresh_data_mode_view_if_active()
            append_event("ui:apply_store_filter:done")

    def clear_store_filter_selection(self):
        """清空店铺筛选选择"""
        for cb in self.store_checkboxes.values():
            cb.setChecked(False)
        self.current_store_filter.clear()
        self.db.set_setting("store_filter_ids", "")
        self.update_store_sheet_selection()

    def clear_store_filter(self):
        """清除店铺筛选，显示所有商品"""
        self.clear_store_filter_selection()
        self.btn_store_filter.setText("🏪 店铺")

        if self._has_active_product_filter_inputs():
            self.apply_tag_filter(close_menu=False, show_message=False)
            return True

        self._set_visible_product_ids(self.product_store_map.keys())
        self.show_toast("已清除店铺筛选")
        self.current_store_filter = set()
        self.update_store_sheet_selection()
        if not self._apply_data_mode_store_visibility():
            self._refresh_data_mode_view_if_active()
        return True
    
    def calculate_profit_label(self, product_id):
        """根据净利润率计算利润标签: 赚钱>5%, 亏钱<5%, 保本=5%"""
        try:
            specs = self.db.safe_fetchall(
                """SELECT spec_code, sale_price, weight_percent FROM product_specs
                   WHERE product_id=? AND COALESCE(is_temporarily_off_shelf, 0)=0""",
                (product_id,)
            )
            
            if not specs:
                return 0
            
            prod_res = self.db.safe_fetchall(
                "SELECT coupon_amount, new_customer_discount FROM products WHERE id=?",
                (product_id,)
            )
            
            coupon = 0
            new_customer = 0
            if prod_res:
                coupon = prod_res[0][0] or 0
                new_customer = prod_res[0][1] or 0
            
            max_discount = max(coupon, new_customer)
            
            total_weight = 0
            total_profit = 0
            total_final_price = 0
            
            for spec_code, sale_price, weight in specs:
                if not sale_price or sale_price <= 0:
                    continue
                
                weight = weight or 0
                
                cost_res = self.db.safe_fetchall(
                    "SELECT cost_price FROM cost_library WHERE spec_code=?",
                    (spec_code,)
                )
                cost = cost_res[0][0] if cost_res and cost_res[0][0] else 0
                
                final_price = sale_price - max_discount
                
                if final_price > 0:
                    profit = final_price - cost
                    total_profit += profit * weight
                    total_final_price += final_price * weight
                    total_weight += weight
            
            if total_weight > 0:
                avg_profit = total_profit / total_weight
                avg_final_price = total_final_price / total_weight
                net_margin_rate = (avg_profit / avg_final_price) * 100 if avg_final_price > 0 else 0

                if net_margin_rate >= 1:
                    return 1
                elif net_margin_rate >= -2:
                    return 0
                else:
                    return -1
            
            return 0
        except Exception as e:
            print(f"计算利润标签失败: {e}")
            return 0
    
    def _calculate_profit_category(self, product_id):
        """根据净利率计算利润分类: 赚钱≥5%, 亏钱<5%, 保本-5%~5%"""
        try:
            rows = self.db.safe_fetchall(
                """SELECT spec_code, sale_price, weight_percent FROM product_specs
                   WHERE product_id=? AND COALESCE(is_temporarily_off_shelf, 0)=0""",
                (product_id,)
            )
            
            if not rows:
                return 'loss'
            
            product_rows = self.db.safe_fetchall(
                "SELECT coupon_amount, new_customer_discount, current_roi, return_rate, is_natural_flow, is_sitewide_managed, store_id FROM products WHERE id=?",
                (product_id,)
            )
            max_discount = 0
            current_roi = 0
            return_rate = 0
            is_natural_flow = 0
            is_sitewide_managed = 0
            sitewide_roi = 0
            
            if product_rows:
                coupon = product_rows[0][0] if product_rows[0][0] else 0
                new_customer = product_rows[0][1] if product_rows[0][1] else 0
                max_discount = max(coupon, new_customer)
                current_roi = product_rows[0][2] if product_rows[0][2] else 0
                return_rate = product_rows[0][3] if product_rows[0][3] else 0
                is_natural_flow = product_rows[0][4] if product_rows[0][4] else 0
                is_sitewide_managed = product_rows[0][5] if product_rows[0][5] else 0
                store_id = product_rows[0][6] if product_rows[0][6] else None
                if store_id:
                    store_rows = self.db.safe_fetchall("SELECT sitewide_roi FROM stores WHERE id=?", (store_id,))
                    sitewide_roi = store_rows[0][0] if store_rows and store_rows[0][0] else 0
            
            total_weighted_margin = 0.0
            total_weight = 0.0
            
            for r in rows:
                spec_code = r[0]
                sale_price = r[1]
                weight = r[2]
                
                if sale_price is None or weight is None:
                    continue
                
                cost_res = self.db.safe_fetchall(
                    "SELECT cost_price FROM cost_library WHERE spec_code=?", (spec_code,)
                )
                cost = cost_res[0][0] if cost_res else 0.0
                
                final_price = sale_price - max_discount
                if final_price > 0 and cost > 0:
                    margin = (final_price - cost) / final_price
                    total_weighted_margin += margin * weight
                    total_weight += weight
            
            if total_weight > 0:
                final_margin_pct = (total_weighted_margin / total_weight) * 100
                
                margin_rate_decimal = final_margin_pct / 100
                effective_roi = sitewide_roi if is_sitewide_managed and not is_natural_flow else current_roi
                final_net_margin_pct = -100
                if is_natural_flow:
                    final_net_margin_pct = (margin_rate_decimal * (1 - return_rate / 100) - 0.006) * 100
                elif effective_roi > 0 and return_rate >= 0:
                    final_net_margin_pct = (margin_rate_decimal * (1 - return_rate / 100) - 0.006 - (1 / effective_roi)) * 100
                
                if final_net_margin_pct >= 1:
                    return 'profit'
                elif final_net_margin_pct >= -2:
                    return 'break_even'
                else:
                    return 'loss'
            
            return 'loss'
        except Exception as e:
            print(f"计算利润分类失败: {e}")
            return 'loss'
    
    def apply_tag_filter(self, close_menu=False, show_message=True):
        """应用标签筛选
        
        Args:
            close_menu: 是否关闭筛选菜单，True=关闭，False=保持打开
        """
        append_event(f"ui:apply_tag_filter:start close={close_menu}")
        try:
            filters = {}
            
            if self.btn_filter_coupon.isChecked():
                filters['coupon'] = True
            if self.btn_filter_new_customer.isChecked():
                filters['new_customer'] = True
            if self.btn_filter_limited_time.isChecked():
                filters['limited_time'] = True
            if self.btn_filter_marketing.isChecked():
                filters['marketing'] = True
            if self.btn_filter_natural_flow.isChecked():
                filters['natural_flow'] = True
            if self.btn_filter_sitewide.isChecked():
                filters['sitewide'] = True
            if self.btn_filter_missing_roi_bid.isChecked():
                filters['missing_roi_bid'] = True

            task_filters = []
            if getattr(self, 'btn_filter_garbage', None) and self.btn_filter_garbage.isChecked():
                task_filters.append('garbage')
            if getattr(self, 'btn_filter_waste', None) and self.btn_filter_waste.isChecked():
                task_filters.append('waste')
            
            profit_filters = []
            if self.btn_filter_profit.isChecked():
                profit_filters.append('profit')
            if self.btn_filter_loss.isChecked():
                profit_filters.append('loss')
            if self.btn_filter_break_even.isChecked():
                profit_filters.append('break_even')
            
            # 只有明确要求关闭菜单时才关闭
            if close_menu:
                self.tag_filter_menu.close()
            
            has_category_filter = bool(getattr(self, "current_category_filter", ""))
            search_match_ids = getattr(self, "current_search_match_ids", None)
            has_search_filter = search_match_ids is not None
            has_store_filter = bool(getattr(self, "current_store_filter", set()))

            if not filters and not task_filters and not profit_filters and not has_category_filter and not has_search_filter:
                if has_store_filter:
                    self.apply_store_filter(close_menu=False, show_message=show_message)
                else:
                    self.btn_tag_filter.setText("🏷️ 筛选")
                    self._set_visible_product_ids(self.product_store_map.keys())
                    self.current_filter_tags = set()
                    if show_message:
                        self.show_toast("已显示所有商品")
                return
            
            all_product_ids = self.get_all_product_ids_with_current_store()
            if not all_product_ids:
                self.show_toast("当前视图无商品")
                return
            
            matching_ids = set()
            task_match_ids = None
            if task_filters:
                patterns = []
                if 'garbage' in task_filters:
                    patterns.append('【垃圾链接】%')
                if 'waste' in task_filters:
                    patterns.append('【废物链接】%')
                conditions = " OR ".join(["task_content LIKE ?"] * len(patterns))
                rows = self.db.safe_fetchall(
                    f"SELECT DISTINCT product_id FROM daily_tasks WHERE is_completed=0 AND ({conditions})",
                    tuple(patterns),
                )
                task_match_ids = {int(row[0]) for row in rows if row and row[0] is not None}
            
            for pid in all_product_ids:
                should_include = True
                
                if filters.get('coupon'):
                    coupon_res = self.db.safe_fetchall("SELECT coupon_amount FROM products WHERE id=?", (pid,))
                    if not coupon_res or not coupon_res[0][0]:
                        should_include = False
                
                if filters.get('new_customer') and should_include:
                    nc_res = self.db.safe_fetchall("SELECT new_customer_discount FROM products WHERE id=?", (pid,))
                    if not nc_res or not nc_res[0][0]:
                        should_include = False
                
                if filters.get('limited_time') and should_include:
                    lt_res = self.db.safe_fetchall("SELECT is_limited_time FROM products WHERE id=?", (pid,))
                    if not lt_res or not lt_res[0][0]:
                        should_include = False
                
                if filters.get('marketing') and should_include:
                    mk_res = self.db.safe_fetchall("SELECT is_marketing FROM products WHERE id=?", (pid,))
                    if not mk_res or not mk_res[0][0]:
                        should_include = False

                if filters.get('natural_flow') and should_include:
                    nf_res = self.db.safe_fetchall("SELECT is_natural_flow FROM products WHERE id=?", (pid,))
                    if not nf_res or not nf_res[0][0]:
                        should_include = False

                if filters.get('sitewide') and should_include:
                    sw_res = self.db.safe_fetchall("SELECT is_sitewide_managed, is_natural_flow FROM products WHERE id=?", (pid,))
                    if not sw_res or not sw_res[0][0] or sw_res[0][1]:
                        should_include = False

                if filters.get('missing_roi_bid') and should_include:
                    roi_res = self.db.safe_fetchall("SELECT current_roi, is_natural_flow, is_sitewide_managed FROM products WHERE id=?", (pid,))
                    current_roi = roi_res[0][0] if roi_res and roi_res[0][0] is not None else 0
                    is_natural_flow = roi_res[0][1] if roi_res else 0
                    is_sitewide_managed = roi_res[0][2] if roi_res else 0
                    try:
                        current_roi = float(current_roi)
                    except (TypeError, ValueError):
                        current_roi = 0
                    if current_roi > 0 or is_natural_flow or is_sitewide_managed:
                        should_include = False

                if task_match_ids is not None and should_include and pid not in task_match_ids:
                    should_include = False
                
                if profit_filters and should_include:
                    profit_category = self._calculate_profit_category(pid)
                    if profit_category not in profit_filters:
                        should_include = False

                if has_category_filter and should_include:
                    if not self._product_matches_category_filter(pid):
                        should_include = False

                if has_search_filter and should_include:
                    if pid not in search_match_ids:
                        should_include = False

                if has_store_filter and should_include:
                    if self.product_store_map.get(pid) not in self.current_store_filter:
                        should_include = False
                
                if should_include:
                    matching_ids.add(pid)
            
            self._set_visible_product_ids(matching_ids)
            
            filtered_count = len(matching_ids)
            self.current_filter_tags = filters.copy()
            if task_filters:
                self.current_filter_tags['tasks'] = task_filters
            if profit_filters:
                self.current_filter_tags['profit'] = profit_filters
            if has_category_filter:
                self.current_filter_tags['category'] = self.current_category_filter
            
            self.btn_tag_filter.setText(f"🏷️ 筛选 ({filtered_count})")
            if show_message:
                self.show_toast(f"筛选: 显示 {filtered_count} 个商品")
            
        except Exception as e:
            print(f"应用标签筛选失败: {e}")
            import traceback
            traceback.print_exc()
            append_exception("ui:apply_tag_filter:failed", error=e)
            QMessageBox.warning(self, "筛选失败", f"标签筛选出错: {e}")
        finally:
            if not self._apply_data_mode_store_visibility():
                self._refresh_data_mode_view_if_active()
            append_event("ui:apply_tag_filter:done")

    def _has_active_product_filter_inputs(self):
        try:
            if getattr(self, "current_category_filter", ""):
                return True
            buttons = [
                self.btn_filter_coupon,
                self.btn_filter_new_customer,
                self.btn_filter_limited_time,
                self.btn_filter_marketing,
                self.btn_filter_natural_flow,
                self.btn_filter_sitewide,
                self.btn_filter_garbage,
                self.btn_filter_waste,
                self.btn_filter_profit,
                self.btn_filter_loss,
                self.btn_filter_break_even,
                self.btn_filter_missing_roi_bid,
            ]
            return any(btn.isChecked() for btn in buttons)
        except Exception:
            return False
    
    def clear_tag_filter_selection(self):
        """清空标签筛选选择"""
        buttons = [
            self.btn_filter_coupon,
            self.btn_filter_new_customer,
            self.btn_filter_limited_time,
            self.btn_filter_marketing,
            self.btn_filter_natural_flow,
            self.btn_filter_sitewide,
            self.btn_filter_garbage,
            self.btn_filter_waste,
            self.btn_filter_profit,
            self.btn_filter_loss,
            self.btn_filter_break_even,
            self.btn_filter_missing_roi_bid,
        ]
        for btn in buttons:
            btn.blockSignals(True)
            btn.setChecked(False)
            btn.blockSignals(False)
        self.current_category_filter = ""
        if hasattr(self, "category_filter_input"):
            self.category_filter_input.blockSignals(True)
            self.category_filter_input.clear()
            self.category_filter_input.blockSignals(False)
        if hasattr(self, "category_candidate_layout"):
            self.refresh_category_filter_candidates()
    
    def clear_tag_filter(self, _checked=False, show_message=True):
        """清除标签筛选，显示所有商品"""
        self.clear_tag_filter_selection()
        self.btn_tag_filter.setText("🏷️ 筛选")
        if getattr(self, "current_store_filter", set()):
            self._set_visible_product_ids(
                pid for pid, sid in self.product_store_map.items()
                if sid in self.current_store_filter
            )
        else:
            self._set_visible_product_ids(self.product_store_map.keys())
        
        self.current_filter_tags = set()

        if getattr(self, "current_search_match_ids", None) is not None:
            self.apply_tag_filter(close_menu=False, show_message=False)
            return

        if show_message:
            self.show_toast("已清除标签筛选")
        if not self._apply_data_mode_store_visibility():
            self._refresh_data_mode_view_if_active()
    
    def clear_filter(self):
        self._set_visible_product_ids(self.product_store_map.keys())
        self.show_toast("已清除筛选，显示全部商品")
        self._refresh_data_mode_view_if_active()

    def _safe_excel_filename_part(self, value):
        text = str(value or "").strip() or "店铺"
        return "".join(ch for ch in text if ch not in r'\/:*?"<>|').strip() or "店铺"

    @staticmethod
    def _copy_excel_sheet_block(source, target, start_column):
        from openpyxl.formula.translate import Translator
        from openpyxl.utils import column_index_from_string, get_column_letter
        from openpyxl.utils.cell import coordinate_to_tuple

        column_offset = start_column - 1
        for row in source.iter_rows():
            for source_cell in row:
                target_cell = target.cell(source_cell.row, source_cell.column + column_offset)
                value = source_cell.value
                if source_cell.data_type == "f":
                    try:
                        value = Translator(value, origin=source_cell.coordinate).translate_formula(target_cell.coordinate)
                    except Exception:
                        pass
                target_cell.value = value
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.border = copy(source_cell.border)
                    target_cell.alignment = copy(source_cell.alignment)
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = copy(source_cell.protection)
                if source_cell.hyperlink:
                    target_cell._hyperlink = copy(source_cell.hyperlink)
                if source_cell.comment:
                    target_cell.comment = copy(source_cell.comment)

        for merged in source.merged_cells.ranges:
            target.merge_cells(
                start_row=merged.min_row,
                start_column=merged.min_col + column_offset,
                end_row=merged.max_row,
                end_column=merged.max_col + column_offset,
            )
        for column_name, dimension in source.column_dimensions.items():
            target.column_dimensions[get_column_letter(column_index_from_string(column_name) + column_offset)].width = dimension.width
        for row_index, dimension in source.row_dimensions.items():
            if dimension.height:
                target.row_dimensions[row_index].height = max(target.row_dimensions[row_index].height or 0, dimension.height)

        image_refs = getattr(target, "_image_stream_refs", [])
        image_refs.extend(getattr(source, "_image_stream_refs", []))
        for source_image in source._images:
            image = copy(source_image)
            if isinstance(source_image.anchor, str):
                row, column = coordinate_to_tuple(source_image.anchor)
                image.anchor = f"{get_column_letter(column + column_offset)}{row}"
            else:
                anchor = copy(source_image.anchor)
                anchor._from = copy(source_image.anchor._from)
                anchor._from.col += column_offset
                if hasattr(source_image.anchor, "to"):
                    anchor.to = copy(source_image.anchor.to)
                    anchor.to.col += column_offset
                image.anchor = anchor
            target.add_image(image)
        target._image_stream_refs = image_refs
        return start_column + source.max_column + 1

    def _select_stores_for_margin_batch_export(self):
        stores = self.db.safe_fetchall(
            "SELECT id, name FROM stores ORDER BY sort_order, id"
        )
        if not stores:
            QMessageBox.information(self, "提示", "当前没有可导出的店铺")
            return []

        dialog = QDialog(self)
        dialog.setWindowTitle("批量导出")
        dialog.resize(460, 570)
        apply_window_icon(dialog, "store")
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)

        tip = QLabel()
        tip.setWordWrap(True)
        tip.setStyleSheet("color: #374151; font-size: 13px;")
        layout.addWidget(tip)

        version_row = QHBoxLayout()
        version_row.addWidget(QLabel("导出版本:"))
        version_combo = QComboBox()
        version_combo.addItem("详细版本（每店独立完整文件）", "detailed")
        version_combo.addItem("单文件详细版（每店一个 Sheet）", "single_detailed")
        version_combo.addItem("简化版本（单文件、多店铺 Sheet）", "simple")
        saved_mode_index = version_combo.findData(self.db.get_setting("batch_margin_export_mode", "detailed"))
        version_combo.setCurrentIndex(max(0, saved_mode_index))
        version_row.addWidget(version_combo, 1)
        layout.addLayout(version_row)

        def update_tip():
            mode = version_combo.currentData()
            if mode == "simple":
                text = "选择需要导出的店铺。简化版本会生成一个 Excel 文件，每个店铺一个 Sheet，只包含过往数据分析和阅览模式图片。"
            elif mode == "single_detailed":
                text = "选择需要导出的店铺。单文件详细版每店一个 Sheet，过往数据、店铺权重和所选链接从左到右排列。"
            else:
                text = "选择需要导出的店铺。详细版本会为每个店铺生成独立的完整毛利 Excel 文件。"
            tip.setText(text)

        version_combo.currentIndexChanged.connect(update_tip)
        update_tip()

        quality_row = QHBoxLayout()
        quality_label = QLabel("图片清晰度:")
        quality_combo = QComboBox()
        quality_combo.addItem("清晰版（图片更清楚，文件较大）", "clear")
        quality_combo.addItem("均衡版（推荐，清晰度和体积平衡）", "balanced")
        quality_combo.addItem("轻量版（文件更小，图片够看）", "light")
        quality_combo.setCurrentIndex(1)
        quality_combo.setToolTip("只影响导出 Excel 里嵌入图片的分辨率，表格内显示尺寸不变。")
        quality_row.addWidget(quality_label)
        quality_row.addWidget(quality_combo, 1)
        layout.addLayout(quality_row)

        actions = QHBoxLayout()
        btn_all = QPushButton("全选")
        btn_none = QPushButton("全不选")
        btn_invert = QPushButton("反选")
        actions.addWidget(btn_all)
        actions.addWidget(btn_none)
        actions.addWidget(btn_invert)
        actions.addStretch()
        layout.addLayout(actions)

        list_widget = QListWidget()
        list_widget.setSelectionMode(QAbstractItemView.NoSelection)
        for store_id, store_name in stores:
            item = QListWidgetItem(str(store_name or f"店铺{store_id}"))
            item.setData(Qt.UserRole, (store_id, str(store_name or f"店铺{store_id}")))
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked)
            list_widget.addItem(item)
        layout.addWidget(list_widget, 1)

        btn_all.clicked.connect(lambda: [list_widget.item(i).setCheckState(Qt.Checked) for i in range(list_widget.count())])
        btn_none.clicked.connect(lambda: [list_widget.item(i).setCheckState(Qt.Unchecked) for i in range(list_widget.count())])
        btn_invert.clicked.connect(lambda: [
            list_widget.item(i).setCheckState(Qt.Unchecked if list_widget.item(i).checkState() == Qt.Checked else Qt.Checked)
            for i in range(list_widget.count())
        ])

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.button(QDialogButtonBox.Ok).setText("开始导出")
        buttons.button(QDialogButtonBox.Cancel).setText("取消")
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        if dialog.exec_() != QDialog.Accepted:
            return None

        self.db.set_setting("batch_margin_export_mode", version_combo.currentData() or "detailed")
        selected = []
        for i in range(list_widget.count()):
            item = list_widget.item(i)
            if item.checkState() == Qt.Checked:
                selected.append(item.data(Qt.UserRole))
        return {
            "stores": selected,
            "image_quality": quality_combo.currentData() or "balanced",
            "export_mode": version_combo.currentData() or "detailed",
        }

    def batch_export_store_margin_excel(self):
        if not has_openpyxl():
            QMessageBox.warning(self, "错误", "请先安装 openpyxl 库：pip install openpyxl")
            return

        selection = self._select_stores_for_margin_batch_export()
        if not selection:
            return
        selected_stores = selection.get("stores") or []
        image_quality = selection.get("image_quality") or "balanced"
        export_mode = selection.get("export_mode") or "detailed"
        if not selected_stores:
            QMessageBox.information(self, "提示", "请至少选择一个店铺")
            return
        if export_mode == "simple":
            self._batch_export_simplified_margin_excel(selected_stores, image_quality)
            return
        detail_map = StoreMarginExcelExporter.select_detail_products(
            self, self.db, self, selected_stores
        )
        if detail_map is None:
            return
        if export_mode == "single_detailed":
            self._batch_export_single_detailed_margin_excel(selected_stores, image_quality, detail_map)
            return

        folder = remembered_existing_directory(self, self.db, "选择批量导出保存文件夹")
        if not folder:
            return

        quality_names = {
            "clear": "清晰版",
            "balanced": "均衡版",
            "light": "轻量版",
        }
        quality_name = quality_names.get(image_quality, "均衡版")
        progress = QProgressDialog(f"正在逐个导出详细版本... 图片：{quality_name}", "取消", 0, len(selected_stores) * 100, self)
        progress.setWindowTitle("批量导出")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)

        exported_count = 0
        canceled_count = 0
        failed = []
        for index, (store_id, store_name) in enumerate(selected_stores, start=1):
            if progress.wasCanceled():
                break
            progress.setLabelText(f"正在导出：{store_name} ({index}/{len(selected_stores)})｜图片：{quality_name}")
            QApplication.processEvents()

            safe_name = self._safe_excel_filename_part(store_name)
            today = datetime.now().strftime("%Y%m%d")
            file_path = os.path.join(folder, f"店铺毛利_{safe_name}_{today}.xlsx")
            suffix = 2
            while os.path.exists(file_path):
                file_path = os.path.join(folder, f"店铺毛利_{safe_name}_{today}_{suffix}.xlsx")
                suffix += 1

            try:
                exporter = StoreMarginExcelExporter(
                    store_id,
                    store_name,
                    self,
                    image_quality=image_quality,
                    detail_product_ids=detail_map.get(int(store_id), set()),
                )
                exporter._batch_export_errors = []

                def update_store_progress(value, text, idx=index, total=len(selected_stores), name=store_name):
                    progress.setValue((idx - 1) * 100 + int(value))
                    progress.setLabelText(f"{name} ({idx}/{total})｜{int(value)}%｜图片：{quality_name}｜{text}")
                    QApplication.processEvents()
                    return not progress.wasCanceled()

                ok = exporter.export_margin_excel_to_path(file_path, update_store_progress)
                if ok:
                    exported_count += 1
                else:
                    errors = getattr(exporter, "_batch_export_errors", [])
                    if progress.wasCanceled():
                        canceled_count += 1
                    else:
                        failed.append((store_name, "\n".join(errors[-3:]) if errors else "导出返回失败，但没有提供具体错误"))
            except Exception as e:
                failed.append((store_name, traceback.format_exc()))
            finally:
                progress.setValue(index * 100)
                QApplication.processEvents()

        progress.close()

        msg = f"成功导出 {exported_count} 个店铺。"
        if exported_count:
            msg += f"\n保存文件夹：\n{folder}"
        if canceled_count:
            msg += f"\n取消或跳过 {canceled_count} 个店铺。"
        if failed:
            detail = "\n".join([f"- {name}: {err}" for name, err in failed[:8]])
            if len(failed) > 8:
                detail += f"\n... 还有 {len(failed) - 8} 个失败"
            QMessageBox.warning(self, "批量导出完成", f"{msg}\n\n失败 {len(failed)} 个：\n{detail}")
        else:
            QMessageBox.information(self, "批量导出完成", msg)

    def _batch_export_simplified_margin_excel(self, selected_stores, image_quality):
        from openpyxl import Workbook

        today = datetime.now().strftime("%Y%m%d")
        file_path, _ = remembered_save_file(
            self,
            self.db,
            "保存批量导出文件",
            f"批量导出_简化版_{today}.xlsx",
            "Excel 文件 (*.xlsx)",
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        total_progress = len(selected_stores) * 100 + 10
        progress = QProgressDialog("正在生成简化版批量导出...", "取消", 0, total_progress, self)
        progress.setWindowTitle("批量导出")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        wb = Workbook()
        default_sheet = wb.active
        used_names = set()
        exported_count = 0
        failed = []

        for index, (store_id, store_name) in enumerate(selected_stores, start=1):
            if progress.wasCanceled():
                break
            base_progress = (index - 1) * 100
            progress.setValue(base_progress + 5)
            progress.setLabelText(f"{store_name} ({index}/{len(selected_stores)})｜5%｜正在创建店铺 Sheet...")
            QApplication.processEvents()
            base_name = re.sub(r"[\\/*?:\[\]]", "_", str(store_name or f"店铺{store_id}")).strip() or f"店铺{store_id}"
            base_name = base_name[:31]
            sheet_name = base_name
            suffix = 2
            while sheet_name.casefold() in used_names:
                suffix_text = f"_{suffix}"
                sheet_name = f"{base_name[:31 - len(suffix_text)]}{suffix_text}"
                suffix += 1
            used_names.add(sheet_name.casefold())
            before_sheets = set(wb.sheetnames)
            try:
                exporter = StoreMarginExcelExporter(
                    store_id,
                    store_name,
                    self,
                    image_quality=image_quality,
                    detail_product_ids=set(),
                )
                progress.setValue(base_progress + 25)
                progress.setLabelText(f"{store_name} ({index}/{len(selected_stores)})｜25%｜正在读取过往数据...")
                QApplication.processEvents()
                exporter._write_historical_export_sheet(wb, sheet_name=sheet_name, create_sheet=True)
                progress.setValue(base_progress + 90)
                progress.setLabelText(f"{store_name} ({index}/{len(selected_stores)})｜90%｜正在嵌入本周图片并调整格式...")
                exported_count += 1
            except Exception:
                failed.append((store_name, traceback.format_exc()))
                for new_sheet in set(wb.sheetnames) - before_sheets:
                    wb.remove(wb[new_sheet])
            finally:
                progress.setValue(index * 100)
                QApplication.processEvents()

        canceled = progress.wasCanceled()
        if default_sheet in wb.worksheets:
            wb.remove(default_sheet)
        if exported_count:
            try:
                progress.setValue(len(selected_stores) * 100 + 5)
                progress.setLabelText("所有店铺已汇总，正在保存 Excel 文件...")
                QApplication.processEvents()
                wb.save(file_path)
                progress.setValue(total_progress)
            except Exception as e:
                progress.close()
                QMessageBox.critical(self, "批量导出失败", f"保存 Excel 文件失败：\n{e}")
                return
        progress.close()

        if not exported_count:
            QMessageBox.warning(self, "批量导出失败", "没有成功生成任何店铺 Sheet。")
            return
        message = f"成功导出 {exported_count} 个店铺到：\n{file_path}"
        if canceled:
            message += "\n已保存取消前完成的店铺。"
        if failed:
            detail = "\n".join(f"- {name}: {error.splitlines()[-1]}" for name, error in failed[:8])
            QMessageBox.warning(self, "批量导出完成", f"{message}\n\n失败 {len(failed)} 个：\n{detail}")
        else:
            QMessageBox.information(self, "批量导出完成", message)

    def _batch_export_single_detailed_margin_excel(self, selected_stores, image_quality, detail_map):
        from openpyxl import Workbook

        today = datetime.now().strftime("%Y%m%d")
        file_path, _ = remembered_save_file(
            self,
            self.db,
            "保存单文件详细版",
            f"批量导出_单文件详细版_{today}.xlsx",
            "Excel 文件 (*.xlsx)",
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        total_progress = len(selected_stores) * 100 + 10
        progress = QProgressDialog("正在生成单文件详细版...", "取消", 0, total_progress, self)
        progress.setWindowTitle("批量导出")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        wb = Workbook()
        default_sheet = wb.active
        used_names = set()
        exported_count = 0
        failed = []

        for index, (store_id, store_name) in enumerate(selected_stores, start=1):
            if progress.wasCanceled():
                break
            base_progress = (index - 1) * 100
            progress.setValue(base_progress + 3)
            progress.setLabelText(f"{store_name} ({index}/{len(selected_stores)})｜3%｜正在创建店铺 Sheet...")
            QApplication.processEvents()
            target = None
            try:
                base_name = re.sub(r"[\\/*?:\[\]]", "_", str(store_name or f"店铺{store_id}")).strip() or f"店铺{store_id}"
                base_name = base_name[:31]
                sheet_name = base_name
                suffix = 2
                while sheet_name.casefold() in used_names:
                    suffix_text = f"_{suffix}"
                    sheet_name = f"{base_name[:31 - len(suffix_text)]}{suffix_text}"
                    suffix += 1
                used_names.add(sheet_name.casefold())
                target = wb.create_sheet(sheet_name)
                exporter = StoreMarginExcelExporter(
                    store_id,
                    store_name,
                    self,
                    image_quality=image_quality,
                    detail_product_ids=detail_map.get(int(store_id), set()),
                )
                exporter._excel_export_image_cache = {}
                source_book = Workbook()
                progress.setValue(base_progress + 15)
                progress.setLabelText(f"{store_name} ({index}/{len(selected_stores)})｜15%｜正在读取过往数据和本周图片...")
                QApplication.processEvents()
                exporter._write_historical_export_sheet(source_book)
                progress.setValue(base_progress + 35)
                progress.setLabelText(f"{store_name} ({index}/{len(selected_stores)})｜35%｜正在按单量计算权重和对比数据...")
                QApplication.processEvents()
                exporter._write_orders_export_sheet(source_book)
                progress.setValue(base_progress + 52)
                progress.setLabelText(f"{store_name} ({index}/{len(selected_stores)})｜52%｜正在整理已选展示链接...")
                products = exporter._products_for_specs_export()
                original_detail_ids = exporter.export_detail_product_ids
                detail_products = products or [None]
                for product_index, product in enumerate(detail_products, start=1):
                    exporter.export_detail_product_ids = {str(product[1])} if product else set()
                    detail_progress = 52 + int(product_index / len(detail_products) * 28)
                    product_name = str(product[1]) if product else "无选中链接"
                    progress.setValue(base_progress + detail_progress)
                    progress.setLabelText(
                        f"{store_name} ({index}/{len(selected_stores)})｜{detail_progress}%｜"
                        f"正在生成链接 {product_name} ({product_index}/{len(detail_products)})..."
                    )
                    QApplication.processEvents()
                    exporter._write_product_specs_export_sheet(source_book)
                exporter.export_detail_product_ids = original_detail_ids

                next_column = 1
                source_sheets = source_book.worksheets
                for source_index, source in enumerate(source_sheets, start=1):
                    merge_progress = 80 + int(source_index / len(source_sheets) * 15)
                    progress.setValue(base_progress + merge_progress)
                    progress.setLabelText(
                        f"{store_name} ({index}/{len(selected_stores)})｜{merge_progress}%｜"
                        f"正在横向合并板块 {source_index}/{len(source_sheets)}：{source.title}..."
                    )
                    QApplication.processEvents()
                    next_column = self._copy_excel_sheet_block(source, target, next_column)
                target.freeze_panes = "A2"
                exported_count += 1
            except Exception:
                failed.append((store_name, traceback.format_exc()))
                if target is not None:
                    wb.remove(target)
            finally:
                progress.setValue(index * 100)
                QApplication.processEvents()

        canceled = progress.wasCanceled()
        wb.remove(default_sheet)
        if exported_count:
            try:
                progress.setValue(len(selected_stores) * 100 + 5)
                progress.setLabelText("所有店铺已汇总，正在写入图片、公式并保存 Excel...")
                QApplication.processEvents()
                wb.save(file_path)
                progress.setValue(total_progress)
            except Exception as e:
                progress.close()
                QMessageBox.critical(self, "批量导出失败", f"保存 Excel 文件失败：\n{e}")
                return
        progress.close()
        if not exported_count:
            QMessageBox.warning(self, "批量导出失败", "没有成功生成任何店铺 Sheet。")
            return
        message = f"成功导出 {exported_count} 个店铺到：\n{file_path}"
        if canceled:
            message += "\n已保存取消前完成的店铺。"
        if failed:
            detail = "\n".join(f"- {name}: {error.splitlines()[-1]}" for name, error in failed[:8])
            QMessageBox.warning(self, "批量导出完成", f"{message}\n\n失败 {len(failed)} 个：\n{detail}")
        else:
            QMessageBox.information(self, "批量导出完成", message)

    def _format_cost_quantity(self, value):
        if value is None:
            return ""
        text = str(value).strip()
        if not text or text.lower() == "nan":
            return ""
        numeric_text = text.replace(",", "")
        try:
            number = float(numeric_text)
        except ValueError:
            return text
        if number.is_integer():
            return str(int(number))
        return str(int(round(number)))

    def _format_cost_attribute_value(self, key, value):
        text = str(value or "").strip()
        if not text or text.lower() == "nan":
            return ""
        numeric_text = text.replace(",", "")
        try:
            number = float(numeric_text)
            if number.is_integer():
                text = str(int(number))
        except ValueError:
            pass
        if key == "size":
            return text if re.search(r"\bmm\b|毫米", text, flags=re.I) else f"{text}mm"
        if key == "pages":
            return text if "张" in text else f"{text}张"
        return text

    def import_cost_data(self):
        """导入成本表 - 最终版 (直接全量读取，无预览，只显示结果)"""
        # 1. 检查依赖
        try:
            import pandas as pd  # type: ignore
        except ImportError as e:
            import sys
            python_path = sys.executable
            QMessageBox.critical(self, "缺少依赖", 
                f"未检测到 pandas 库！\n\n"
                f"当前Python: {python_path}\n"
                f"错误: {str(e)}\n\n"
                f"请在终端运行:\n"
                f"pip install pandas openpyxl")
            return

        # 2. 选择文件
        file_path, _ = remembered_open_file(
            self,
            self.db,
            "选择成本表文件",
            "Excel 文件 (*.xlsx *.xlsm *.xls);;CSV 文件 (*.csv);;所有文件 (*)",
        )
        if not file_path:
            return

        # 4. 弹出配置对话框 (选择列)
        try:
            cost_mode = self.db.get_cost_library_mode() if hasattr(self.db, "get_cost_library_mode") else "total"
            dialog = CostImportDialog(file_path, self, cost_mode=cost_mode)
            apply_window_icon(dialog, "cost")
            if dialog.exec_() != QDialog.Accepted:
                return
            
            mapping = dialog.get_mapping()
            if len(mapping) == 6:
                spec_col_idx, price_col_idx, name_col_idx, quantity_col_idx, category_col_idx, weight_col_idx = mapping
                attribute_col_indices = None
            else:
                spec_col_idx, price_col_idx, name_col_idx, quantity_col_idx, category_col_idx, weight_col_idx, attribute_col_indices = mapping
            unit_by_quantity = dialog.should_unit_by_quantity() if hasattr(dialog, "should_unit_by_quantity") else False
            update_fields = dialog.get_update_fields() if hasattr(dialog, "get_update_fields") else {"category": True, "name": True, "quantity": True, "cost": True}
            
            if spec_col_idx is None or price_col_idx is None:
                QMessageBox.warning(self, "提示", "请先选择【规格编码】和【总成本/产品成本】所在的列！")
                return
            if cost_mode == "detail" and weight_col_idx is None:
                QMessageBox.warning(self, "提示", "详细成本模式下请先选择【单个重量kg】所在的列！")
                return
                
        except Exception as e:
            QMessageBox.critical(self, "配置错误", f"打开配置窗口失败:\n{str(e)}")
            return

        # 5. 开始全量读取和处理
        try:
            self.statusBar().showMessage("正在读取并处理数据...", 0)
            QApplication.processEvents()

            df = read_cost_file(file_path)

            if df.empty:
                QMessageBox.warning(self, "提示", "文件内容为空！")
                self.statusBar.showMessage("导入取消", 3000)
                return

            total_rows = len(df)
            print(f"文件总行数：{total_rows}")

            # 获取用户选中的列数据 (使用 iloc 按位置索引获取)
            # astype(str) 确保规格编码变成字符串，防止数字变科学计数法
            col_spec = df.iloc[:, spec_col_idx].astype(str)
            col_price = df.iloc[:, price_col_idx]
            col_name = df.iloc[:, name_col_idx] if name_col_idx is not None else None
            col_quantity = df.iloc[:, quantity_col_idx] if quantity_col_idx is not None else None
            col_category = df.iloc[:, category_col_idx] if category_col_idx is not None else None
            col_weight = df.iloc[:, weight_col_idx] if weight_col_idx is not None else None
            attribute_cols = {}
            import_product_attribute = attribute_col_indices is not None
            if attribute_col_indices:
                for key, col_idx in attribute_col_indices.items():
                    if col_idx is not None:
                        col_index = int(col_idx)
                        attribute_cols[key] = {
                            "series": df.iloc[:, col_index],
                            "header": str(df.columns[col_index]).strip(),
                        }
            row_colors = read_cost_row_colors(file_path, name_col_idx=name_col_idx, spec_col_idx=spec_col_idx)
            
            count_success = 0
            count_skip = 0
            count_error = 0
            count_history = 0
            count_changed = 0
            count_unit_converted = 0
            count_code_replaced = 0
            replaced_old_codes = set()
            changed_cost_spec_codes = set()
            combo_component_spec_codes = set()
            
            # 批量插入准备 (为了提高速度，可以每100条提交一次，这里为了简单逐条处理但加了事务优化)
            # 实际上 safe_execute 已经是逐条提交，对于几万行数据可能会慢，但最稳定
            
            if hasattr(self.db, "set_cost_history_source"):
                self.db.set_cost_history_source("import")
            self.db.conn.execute("BEGIN TRANSACTION") # 开启事务，极大提高写入速度

            for idx in range(total_rows):
                try:
                    # 获取规格
                    spec_val = col_spec.iloc[idx]
                    if not spec_val or spec_val.strip() == "" or spec_val.lower() == "nan":
                        count_skip += 1
                        continue
                    spec_code = spec_val.strip()

                    spec_name = ""
                    if col_name is not None:
                        name_val = col_name.iloc[idx]
                        if not pd.isna(name_val):
                            spec_name = str(name_val).strip()
                            if spec_name.lower() == "nan":
                                spec_name = ""

                    quantity = ""
                    if col_quantity is not None:
                        quantity_val = col_quantity.iloc[idx]
                        if not pd.isna(quantity_val):
                            quantity = self._format_cost_quantity(quantity_val)
                            if quantity.lower() == "nan":
                                quantity = ""

                    category_label = ""
                    if col_category is not None:
                        category_val = col_category.iloc[idx]
                        if not pd.isna(category_val):
                            category_label = str(category_val).strip()
                            if category_label.lower() == "nan":
                                category_label = ""
                    category_color = self.db.category_color_for_label(category_label)
                    source_bg_color = row_colors.get(idx, "")

                    product_attribute = ""
                    product_attribute_is_combo = 0
                    is_combo_name = self.db.is_cost_combo_name(spec_name) if hasattr(self.db, "is_cost_combo_name") else bool(re.search(r"\+|＋|﹢", spec_name or ""))
                    if attribute_cols and not is_combo_name:
                        attr_parts = []
                        for key in ("size", "pages", "print"):
                            attr_info = attribute_cols.get(key)
                            if not attr_info:
                                continue
                            col_attr = attr_info.get("series")
                            attr_val = col_attr.iloc[idx]
                            if pd.isna(attr_val):
                                continue
                            attr_text = str(attr_val).strip()
                            if not attr_text or attr_text.lower() == "nan":
                                continue
                            if key in ("size", "pages"):
                                attr_text = self._format_cost_attribute_value(key, attr_text)
                            header_text = str(attr_info.get("header") or "").strip()
                            attr_parts.append(f"{header_text}：{attr_text}" if header_text else attr_text)
                        product_attribute = "\n".join(attr_parts)
                    product_attribute_combo_disabled = 0
                    
                    # 获取价格
                    price_val = col_price.iloc[idx]
                    parsed_price = self.db.parse_cost_number(price_val, None) if hasattr(self.db, "parse_cost_number") else None
                    if parsed_price is None:
                        parsed_price = 0.0

                    product_cost = None
                    unit_weight = None
                    shipping_fee = None
                    misc_fee = None
                    cost_calc_mode = cost_mode
                    if cost_mode == "detail":
                        product_cost = float(parsed_price)
                        if col_weight is None:
                            count_error += 1
                            continue
                        unit_weight = self.db.parse_cost_number(col_weight.iloc[idx], None) if hasattr(self.db, "parse_cost_number") else None
                        if unit_weight is None or unit_weight <= 0:
                            count_error += 1
                            continue
                        if unit_by_quantity and hasattr(self.db, "parse_cost_quantity_factor"):
                            quantity_factor = self.db.parse_cost_quantity_factor(quantity)
                            if quantity_factor and quantity_factor > 1:
                                product_cost = product_cost / quantity_factor
                                unit_weight = unit_weight / quantity_factor
                                count_unit_converted += 1
                        cost_price, shipping_fee, misc_fee, _total_weight = self.db.calculate_detailed_cost(product_cost, quantity, unit_weight)
                    else:
                        cost_price = float(parsed_price)
                        cost_calc_mode = "total"

                    old_rows = self.db.cursor.execute(
                        """SELECT cost_price, COALESCE(product_attribute_is_combo, 0),
                                  product_cost, COALESCE(cost_calc_mode, 'total')
                           FROM cost_library WHERE spec_code=?""",
                        (spec_code,)
                    ).fetchall()
                    if not old_rows and spec_name:
                        match_rows = self.db.cursor.execute(
                            """SELECT spec_code
                               FROM cost_library
                               WHERE COALESCE(spec_name, '')=?
                                 AND (?='' OR COALESCE(category_label, '')=?)
                                 AND COALESCE(spec_code, '')<>?
                               ORDER BY CASE WHEN manual_sort_order IS NULL THEN 1 ELSE 0 END,
                                        manual_sort_order, sort_order, spec_code""",
                            (spec_name, category_label, category_label, spec_code),
                        ).fetchall()
                        old_code = next((str(row[0] or "").strip() for row in match_rows if str(row[0] or "").strip() not in replaced_old_codes), "")
                        if old_code:
                            self.db.rename_cost_spec_code(
                                old_code, spec_code, manage_transaction=False, mark_dirty=False
                            )
                            replaced_old_codes.add(old_code)
                            count_code_replaced += 1
                            old_rows = self.db.cursor.execute(
                                """SELECT cost_price, COALESCE(product_attribute_is_combo, 0),
                                          product_cost, COALESCE(cost_calc_mode, 'total')
                                   FROM cost_library WHERE spec_code=?""",
                                (spec_code,)
                            ).fetchall()
                    cost_row_exists = bool(old_rows)
                    old_cost = float(old_rows[0][0]) if old_rows and old_rows[0][0] is not None else None
                    existing_is_combo = bool(old_rows and old_rows[0][1])
                    old_product_cost = float(old_rows[0][2]) if old_rows and old_rows[0][2] is not None else None
                    update_cost = bool(update_fields.get("cost", True)) and not existing_is_combo
                    cost_changed = (not cost_row_exists) or (
                        update_cost and (old_cost is None or abs(cost_price - old_cost) > 0.001)
                    )
                    old_unit_cost = old_product_cost if cost_mode == "detail" else old_cost
                    new_unit_cost = product_cost if cost_mode == "detail" else cost_price
                    should_record_history = (
                        cost_mode == "detail" and update_cost
                        and old_unit_cost is not None and new_unit_cost is not None
                        and abs(float(new_unit_cost) - float(old_unit_cost)) > 0.001
                    )

                    self.db.cursor.execute(
                        """INSERT INTO cost_library
                           (spec_code, spec_name, quantity, category_label, category_color, cost_price, sort_order, source_bg_color,
                            product_cost, unit_weight, shipping_fee, misc_fee, cost_calc_mode, product_attribute, product_attribute_combo_disabled, product_attribute_is_combo)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                           ON CONFLICT(spec_code) DO UPDATE SET
                               spec_name=CASE WHEN ? THEN excluded.spec_name ELSE cost_library.spec_name END,
                               quantity=CASE WHEN ? THEN excluded.quantity ELSE cost_library.quantity END,
                               category_label=CASE WHEN ? THEN excluded.category_label ELSE cost_library.category_label END,
                               category_color=CASE WHEN ? THEN excluded.category_color ELSE cost_library.category_color END,
                               cost_price=CASE WHEN ? THEN excluded.cost_price ELSE cost_library.cost_price END,
                               sort_order=cost_library.sort_order,
                               source_bg_color=cost_library.source_bg_color,
                               product_cost=CASE WHEN ? THEN excluded.product_cost ELSE cost_library.product_cost END,
                               unit_weight=CASE WHEN ? THEN excluded.unit_weight ELSE cost_library.unit_weight END,
                               shipping_fee=CASE WHEN ? THEN excluded.shipping_fee ELSE cost_library.shipping_fee END,
                               misc_fee=CASE WHEN ? THEN excluded.misc_fee ELSE cost_library.misc_fee END,
                               cost_calc_mode=CASE WHEN ? THEN excluded.cost_calc_mode ELSE cost_library.cost_calc_mode END,
                               product_attribute=CASE
                                     WHEN ? AND COALESCE(excluded.product_attribute, '') <> '' THEN excluded.product_attribute
                                     ELSE cost_library.product_attribute
                                END,
                               product_attribute_combo_disabled=CASE
                                    WHEN ? AND COALESCE(excluded.product_attribute, '') <> '' THEN excluded.product_attribute_combo_disabled
                                    ELSE cost_library.product_attribute_combo_disabled
                                END,
                               product_attribute_is_combo=CASE
                                    WHEN ? AND COALESCE(excluded.product_attribute, '') <> '' THEN excluded.product_attribute_is_combo
                                    ELSE cost_library.product_attribute_is_combo
                                END""",
                        (spec_code, spec_name, quantity, category_label, category_color, cost_price, idx + 1, source_bg_color,
                         product_cost, unit_weight, shipping_fee, misc_fee, cost_calc_mode, product_attribute,
                         product_attribute_combo_disabled, product_attribute_is_combo,
                         int(update_fields.get("name", True)),
                         int(update_fields.get("quantity", True)),
                         int(update_fields.get("category", True)),
                         int(update_fields.get("category", True)),
                         int(update_cost),
                         int(update_cost),
                         int(update_cost),
                         int(update_cost),
                         int(update_cost),
                         int(update_cost),
                         int(import_product_attribute), int(import_product_attribute), int(import_product_attribute))
                    )

                    if should_record_history:
                        count_history += 1
                        count_changed += 1
                    if cost_changed:
                        changed_cost_spec_codes.add(spec_code)
                    if cost_mode == "detail" and update_cost:
                        combo_component_spec_codes.add(spec_code)
                    
                    count_success += 1
                    
                    # 每处理 1000 行提交一次事务，防止内存溢出并保持响应
                    if count_success % 1000 == 0:
                        self.db.conn.commit()
                        self.statusBar().showMessage(f"已处理 {count_success}/{total_rows} 条...", 0)
                        QApplication.processEvents()

                except Exception as row_err:
                    count_error += 1
                    # 单行错误不中断，继续下一条
                    # print(f"第 {idx+1} 行处理失败：{row_err}")
                    continue
            
            # 提交剩余事务
            self.db.conn.commit()
            if hasattr(self.db, "detect_cost_combo_candidates"):
                self.db.detect_cost_combo_candidates()
            if hasattr(self.db, "recalculate_cost_combinations_for_components"):
                changed_cost_spec_codes.update(
                    self.db.recalculate_cost_combinations_for_components(
                        combo_component_spec_codes, record_history=True, source="import"
                    )
                )
            self.db.set_setting("cost_sync_local_dirty", "1")
            self.db.normalize_cost_category_colors()
            self.db.update_all_product_category_labels()
            self.statusBar().showMessage("导入完成！", 3000)

            # 6. 显示结果
            msg = (f"✅ **导入完成！**\n\n"
                   f"📌 成本模式：{'详细成本模式' if cost_mode == 'detail' else '总成本模式'}\n"
                   f"📊 文件总行数：{total_rows}\n"
                   f"✅ 成功入库：{count_success} 条\n"
                   f"🔁 编码替换：{count_code_replaced} 条\n"
                   f"🕘 产品成本历史：{count_history} 条\n"
                   f"📈 产品成本变化：{count_changed} 条\n"
                   f"⏭️ 跳过空行：{count_skip} 条\n"
                   f"❌ 处理异常：{count_error} 条\n\n"
                   f"数据已更新至数据库 cost_library 表。")
            
            QMessageBox.information(self, "导入结果", msg)
            QTimer.singleShot(
                0,
                lambda codes=tuple(changed_cost_spec_codes): self.finish_cost_import_refresh(codes),
            )
            print(msg)

        except Exception as e:
            # 发生严重错误，回滚事务
            try:
                self.db.conn.rollback()
            except:
                pass
            
            import traceback
            error_detail = traceback.format_exc()
            print(error_detail)
            
            QMessageBox.critical(self, "严重错误", 
                                 f"❌ 导入过程中发生未知错误！\n\n"
                                 f"错误信息：{str(e)}\n\n"
                                 f"详细信息已打印到控制台。\n\n"
                                 f"建议：\n"
                                 f"1. 检查文件是否损坏。\n"
                                 f"2. 确保选中了正确的列。\n"
                                 f"3. 尝试将文件另存为新的 .xlsx 文件。")
            self.statusBar().showMessage("导入失败", 3000)
        finally:
            if hasattr(self.db, "set_cost_history_source"):
                self.db.set_cost_history_source("manual")

    def finish_cost_import_refresh(self, changed_spec_codes=()):
        self.sync_material_library_after_cost_import()
        codes = list(dict.fromkeys(str(code) for code in changed_spec_codes if code))
        dialog = getattr(self, "cost_library_dialog", None)
        if self._is_qobject_alive(dialog) and hasattr(dialog, "load_data"):
            if hasattr(dialog, "refresh_external_changes"):
                dialog.refresh_external_changes(codes, [], False)
            else:
                dialog.load_data()
        product_ids = set()
        for start in range(0, len(codes), 900):
            batch = codes[start:start + 900]
            placeholders = ",".join("?" for _ in batch)
            product_ids.update(
                row[0] for row in self.db.safe_fetchall(
                    f"""SELECT DISTINCT ps.product_id
                        FROM product_specs ps
                        JOIN products p ON p.id=ps.product_id
                        WHERE ps.spec_code IN ({placeholders})
                          AND COALESCE(p.is_archived, 0)=0""",
                    batch,
                )
            )
        self.refresh_external_products(product_ids)

    def sync_material_library_after_cost_import(self):
        """成本库导入后按最新商品类型/名称触发一次素材库文件夹同步。"""
        try:
            dialog = getattr(self, "material_library_dialog", None)
            owns_dialog = False
            if dialog is None:
                dialog = MaterialLibraryDialog(self.db, self)
                owns_dialog = True
            if not dialog.root_folder():
                if owns_dialog:
                    dialog.deleteLater()
                return False
            dialog.load_categories()
            dialog.sync_all_category_folders()
            if hasattr(dialog, "refresh_bubbles") and not owns_dialog:
                dialog.refresh_bubbles()
            if owns_dialog:
                dialog.deleteLater()
            return True
        except Exception as e:
            print(f"成本库导入后同步素材库失败: {e}")
            return False

    def show_cost_library(self):
        """打开成本库管理窗口"""
        existing = getattr(self, "cost_library_dialog", None)
        if existing is not None:
            if existing.isMinimized():
                existing.showNormal()
            else:
                existing.show()
            existing.raise_()
            existing.activateWindow()
            if hasattr(existing, "activate_keyboard_shortcuts"):
                existing.activate_keyboard_shortcuts()
            return
        dialog = CostLibraryDialog(self.db, main_window=self, parent=None)
        apply_window_icon(dialog, "cost")
        # Keep the populated table cached while this account is active.
        dialog.setAttribute(Qt.WA_DeleteOnClose, False)
        dialog.destroyed.connect(lambda _=None: setattr(self, "cost_library_dialog", None))
        self.cost_library_dialog = dialog
        dialog.show()
        dialog.raise_()
        dialog.activateWindow()
        if hasattr(dialog, "activate_keyboard_shortcuts"):
            dialog.activate_keyboard_shortcuts()

    def show_material_library(self):
        """打开素材库窗口"""
        existing = getattr(self, "material_library_dialog", None)
        if existing is not None:
            if existing.isMinimized():
                existing.showNormal()
            else:
                existing.show()
            existing.raise_()
            existing.activateWindow()
            return
        dialog = MaterialLibraryDialog(self.db, self)
        apply_window_icon(dialog, "material")
        dialog.destroyed.connect(lambda _=None: setattr(self, "material_library_dialog", None))
        self.material_library_dialog = dialog
        dialog.show()
        dialog.raise_()
        dialog.activateWindow()

    def open_link_material_library(self, product_db_id):
        """打开素材库并跳转到指定链接的素材目录"""
        self.show_material_library()
        dialog = getattr(self, "material_library_dialog", None)
        if dialog is None:
            return
        if hasattr(dialog, "open_link_material_for_product"):
            dialog.open_link_material_for_product(product_db_id)
        if dialog.isMinimized():
            dialog.showNormal()
        else:
            dialog.show()
        dialog.raise_()
        dialog.activateWindow()

    def _ensure_material_library_backend(self):
        dialog = getattr(self, "material_library_dialog", None)
        if _qobject_alive(dialog) and getattr(dialog, "db", None) is self.db:
            return dialog
        dialog = MaterialLibraryDialog(self.db, self)
        apply_window_icon(dialog, "material")
        dialog.destroyed.connect(lambda _=None: setattr(self, "material_library_dialog", None))
        self.material_library_dialog = dialog
        return dialog

    def _provide_material_mobile_context(self, action, payload):
        account = self.archive_manager.get_active_data_account() if self.archive_manager else None
        if not account:
            raise RuntimeError("电脑当前数据没有绑定软件账号")
        session = {
            "account_id": str(account.get("id") or ""),
            "account_name": str(account.get("name") or ""),
        }
        if action == "session":
            return session
        dialog = self._ensure_material_library_backend()
        if action == "catalog":
            return dialog.mobile_catalog_snapshot()
        if action == "target":
            return dialog.mobile_upload_target(payload.get("target") or {})
        if action == "uploaded":
            if dialog.isVisible():
                QTimer.singleShot(0, dialog.refresh_current_view)
            QTimer.singleShot(0, self.request_cost_thumbnail_scan)
            return {"ok": True}
        raise RuntimeError("未知的手机素材请求")

    def ensure_material_mobile_service(self):
        service = getattr(self, "material_mobile_service", None)
        if service is None:
            service = MaterialMobileService(self._provide_material_mobile_context, parent=self)
            self.material_mobile_service = service
        account = self.archive_manager.get_active_data_account() if self.archive_manager else None
        service.set_active_account(account)
        service.start()
        return service

    def _provide_cost_sync_context(self, action, payload):
        if action == "state":
            return self.db.get_cost_sync_state()
        if action == "local_snapshot":
            return {"snapshot": self.db.build_cost_sync_snapshot()}
        if action == "merge_snapshots":
            return {
                "snapshot": self.db.merge_cost_sync_snapshots(
                    payload.get("current") or {}, payload.get("incoming") or {}
                )
            }
        if action == "load_pending":
            try:
                pending = json.loads(self.db.get_setting("cost_sync_pending_json", "") or "{}")
            except (TypeError, ValueError, json.JSONDecodeError):
                pending = {}
            return {"snapshot": pending}
        if action == "save_pending":
            self.db.set_setting(
                "cost_sync_pending_json",
                json.dumps(payload.get("snapshot") or {}, ensure_ascii=False, separators=(",", ":")),
            )
            return {"ok": True}
        if action == "skip_initial_diff":
            return {"skip": self.db.get_setting("cost_sync_skip_initial_diff", "0") == "1"}
        if action == "clear_skip_initial_diff":
            self.db.set_setting("cost_sync_skip_initial_diff", "0")
            return {"ok": True}
        if action == "clear_local_dirty":
            self.db.set_setting("cost_sync_local_dirty", "0")
            return {"ok": True}
        if action == "remember_host":
            self.db.update_cost_sync_state(coordinator_host=payload.get("coordinator_host") or "")
            return {"ok": True}
        if action == "snapshot":
            state = self.db.get_cost_sync_state()
            try:
                snapshot = json.loads(state.get("snapshot_json") or "{}")
            except (TypeError, ValueError, json.JSONDecodeError):
                snapshot = {}
            return {
                "revision": int(state.get("revision") or 0),
                "snapshot": snapshot,
                "snapshot_hash": state.get("snapshot_hash") or "",
                "publisher_id": state.get("publisher_id") or "",
                "published_at": state.get("published_at") or "",
            }
        if action == "publish":
            publisher_id = payload.get("publisher_id") or ""
            result = self.db.publish_cost_sync_snapshot(
                payload.get("snapshot") or {}, publisher_id
            )
            service = getattr(self, "cost_sync_service", None)
            if not service or publisher_id != service.device_id:
                QTimer.singleShot(0, lambda data=dict(result): self._finish_cost_sync_changes(data))
            result = dict(result)
            result.pop("snapshot", None)
            return result
        if action == "apply_remote":
            result = self.db.apply_remote_cost_sync_snapshot(
                payload.get("snapshot") or {},
                payload.get("revision") or 0,
                payload.get("snapshot_hash") or "",
                payload.get("publisher_id") or "",
                payload.get("published_at") or "",
                bool(payload.get("replace_local")),
            )
            QTimer.singleShot(0, lambda data=dict(result): self._finish_cost_sync_changes(data))
            return result
        raise RuntimeError("未知的成本同步请求")

    def _finish_cost_sync_changes(self, result):
        changed_codes = list(dict.fromkeys(result.get("changed_codes") or []))
        image_changed_codes = list(dict.fromkeys(result.get("image_changed_codes") or []))
        if changed_codes or image_changed_codes:
            image_changed_codes = list(dict.fromkeys(
                image_changed_codes + self.db.inherit_single_multiplier_combo_thumbnails()
            ))
        categories_changed = bool(result.get("categories_changed"))
        history_changed_count = int(result.get("history_changed_count") or 0)
        if not changed_codes and not image_changed_codes and not categories_changed and not history_changed_count:
            return
        product_count = len(set(changed_codes) | set(image_changed_codes))
        if product_count:
            sync_message = f"已接收同步 {product_count} 个产品规格"
        elif categories_changed:
            sync_message = "已接收商品类型同步更新"
        else:
            sync_message = f"已接收同步 {history_changed_count} 条历史操作"
        hint_shown = False
        dialog = getattr(self, "cost_library_dialog", None)
        if self._is_qobject_alive(dialog) and hasattr(dialog, "load_data"):
            if hasattr(dialog, "refresh_external_changes"):
                dialog.refresh_external_changes(
                    changed_codes, image_changed_codes, categories_changed
                )
            else:
                dialog.load_data()
            if hasattr(dialog, "show_sync_hint"):
                dialog.show_sync_hint(sync_message)
                hint_shown = True
        if not hint_shown and hasattr(self, "show_toast"):
            self.show_toast(sync_message, 1500)
        if not changed_codes and not categories_changed:
            return
        try:
            self.db.update_all_product_category_labels()
        except Exception as exc:
            append_exception("cost_sync:update_product_categories", error=exc)
        self.sync_material_library_after_cost_import()
        product_ids = set()
        for start in range(0, len(changed_codes), 900):
            batch = changed_codes[start:start + 900]
            placeholders = ",".join("?" for _ in batch)
            product_ids.update(
                row[0] for row in self.db.safe_fetchall(
                    f"""SELECT DISTINCT ps.product_id
                        FROM product_specs ps
                        JOIN products p ON p.id=ps.product_id
                        WHERE ps.spec_code IN ({placeholders})
                          AND COALESCE(p.is_archived, 0)=0""",
                    batch,
                )
            )
        self.refresh_external_products(product_ids)

    def ensure_cost_sync_service(self):
        service = getattr(self, "cost_sync_service", None)
        if service is None:
            service = CostSyncService(self._provide_cost_sync_context, parent=self)
            self.cost_sync_service = service
        self.db.cost_sync_change_callback = service.notify_local_change
        service.start()
        return service

    def stop_cost_sync_service(self):
        service = getattr(self, "cost_sync_service", None)
        if service is not None:
            if getattr(self.db, "cost_sync_change_callback", None) == service.notify_local_change:
                self.db.cost_sync_change_callback = None
            service.stop()
            service.deleteLater()
            self.cost_sync_service = None

    def restart_cost_sync_service(self):
        self.stop_cost_sync_service()
        if self.db.get_cost_sync_state():
            try:
                self.ensure_cost_sync_service()
            except Exception as exc:
                append_exception("cost_sync_service:restart", error=exc)

    def open_product_material_library(self, spec_code):
        """打开素材库并跳转到指定成本库规格目录"""
        self.show_material_library()
        dialog = getattr(self, "material_library_dialog", None)
        if dialog is None:
            return
        if hasattr(dialog, "open_product_material_for_spec_code"):
            dialog.open_product_material_for_spec_code(spec_code)
        if dialog.isMinimized():
            dialog.showNormal()
        else:
            dialog.show()
        dialog.raise_()
        dialog.activateWindow()
        if hasattr(dialog, "activate_keyboard_shortcuts"):
            dialog.activate_keyboard_shortcuts()

    def material_images_for_cost_specs(self, spec_codes, white_only=False):
        dialog = self._ensure_material_library_backend()
        if white_only:
            return dialog.product_white_images_for_spec_codes(spec_codes)
        return dialog.product_material_images_for_spec_codes(spec_codes)

    def save_cost_thumbnail_to_material(self, spec_code, image_data):
        return self._ensure_material_library_backend().save_cost_thumbnail_to_product_material(
            spec_code, image_data
        )

    def request_cost_thumbnail_scan(self):
        dialog = getattr(self, "cost_library_dialog", None)
        if self._is_qobject_alive(dialog) and hasattr(dialog, "refresh_missing_thumbnails"):
            dialog.refresh_missing_thumbnails()
            return
        self.db.inherit_single_multiplier_combo_thumbnails()
        missing_codes = [
            str(row[0]) for row in self.db.safe_fetchall(
                """SELECT spec_code FROM cost_library
                   WHERE COALESCE(spec_code, '')<>''
                     AND COALESCE(product_attribute_is_combo, 0)=0
                     AND LENGTH(COALESCE(thumbnail_data, X''))=0"""
            )
        ]
        if not missing_codes:
            return
        for spec_code, path in self.material_images_for_cost_specs(
            missing_codes, white_only=True
        ).items():
            image_data = CostLibraryDialog._thumbnail_bytes_from_path(path)
            if image_data:
                self.db.set_cost_thumbnail(spec_code, image_data, only_if_empty=True)
        self.db.inherit_single_multiplier_combo_thumbnails()

    def open_product_material_library_for_link(self, product_db_id):
        """打开素材库，并按链接规格定位到产品素材或商品类型素材。"""
        self.show_material_library()
        dialog = getattr(self, "material_library_dialog", None)
        if dialog is None:
            return
        if hasattr(dialog, "open_product_material_for_link"):
            dialog.open_product_material_for_link(product_db_id)
        if dialog.isMinimized():
            dialog.showNormal()
        else:
            dialog.show()
        dialog.raise_()
        dialog.activateWindow()
        if hasattr(dialog, "activate_keyboard_shortcuts"):
            dialog.activate_keyboard_shortcuts()

    def show_archive_dialog(self):
        """打开存档窗口"""
        try:
            from manager.archive_manager import ArchiveDialog
            active_account = self.archive_manager.get_active_data_account() if self.archive_manager else None
            if active_account:
                self.archive_manager.switch_account(active_account['id'])
            dialog = ArchiveDialog(self.db, self.archive_manager, self)
            apply_window_icon(dialog, "archive")
            dialog.exec_()
            self.update_archive_account_label()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"打开存档窗口失败：\n{str(e)}")
            import traceback
            traceback.print_exc()

    def update_archive_account_label(self):
        """更新当前本地数据归属账号显示标签"""
        active_account = None
        try:
            if hasattr(self, 'archive_manager') and self.archive_manager:
                active_account = self.archive_manager.get_active_data_account()
                if active_account:
                    self.lbl_archive_account.setText(f"💾 当前应用: {active_account.get('name', '未知')}")
                    self.lbl_archive_account.setToolTip(f"当前数据库：{getattr(self.db, 'db_path', '')}")
                    self.lbl_archive_account.setStyleSheet("color: #27ae60; font-size: 11px; padding: 0 5px;")
                else:
                    self.lbl_archive_account.setText("当前应用: 未绑定")
                    self.lbl_archive_account.setToolTip(f"当前数据库：{getattr(self.db, 'db_path', '')}")
                    self.lbl_archive_account.setStyleSheet("color: #888; font-size: 11px; padding: 0 5px;")
            else:
                self.lbl_archive_account.setText("当前应用: 未绑定")
                self.lbl_archive_account.setToolTip("")
                self.lbl_archive_account.setStyleSheet("color: #888; font-size: 11px; padding: 0 5px;")
        except Exception as e:
            print(f"更新存档账号标签失败: {e}")
        service = getattr(self, "material_mobile_service", None)
        if service is not None:
            service.set_active_account(active_account)

    def show_api_config_dialog(self):
        """打开API配置窗口"""
        dialog = ApiConfigDialog(self.db, self)
        apply_window_icon(dialog, "api")
        dialog.show()
    
    def show_daily_task_dialog(self):
        """打开每日任务大盘窗口"""
        waste_rows = self.db.safe_fetchall(
            """SELECT COALESCE(MAX(id), 0) FROM daily_tasks
               WHERE is_completed=0 AND (task_content LIKE '【废物链接】%' OR task_content LIKE '【垃圾链接】%')"""
        )
        self.db.set_setting("daily_task_waste_seen_id", waste_rows[0][0] if waste_rows else 0)
        self.update_daily_task_button_badge()
        if self.daily_task_dialog:
            self.daily_task_dialog.show()
            self.daily_task_dialog.raise_()
            self.daily_task_dialog.activateWindow()
            return
        self.daily_task_dialog = DailyTaskDialog(self.db, self)
        apply_window_icon(self.daily_task_dialog, "daily")
        self.daily_task_dialog.setAttribute(Qt.WA_DeleteOnClose, True)
        self.daily_task_dialog.destroyed.connect(lambda _obj=None: setattr(self, "daily_task_dialog", None))
        self.daily_task_dialog.show()

    def update_daily_task_button_badge(self):
        try:
            latest_rows = self.db.safe_fetchall(
                """SELECT COALESCE(MAX(id), 0) FROM daily_tasks
                   WHERE is_completed=0 AND (task_content LIKE '【废物链接】%' OR task_content LIKE '【垃圾链接】%')"""
            )
            latest_id = latest_rows[0][0] if latest_rows else 0
            seen_id = int(self.db.get_setting("daily_task_waste_seen_id", "0") or 0)
            badge = getattr(self, "daily_task_badge", None)
            button = getattr(self, "btn_daily_task", None)
            if not badge or not button:
                return
            badge.move(max(0, button.width() - badge.width() + 2), -3)
            badge.setVisible(latest_id > seen_id)
            badge.setToolTip("有新的废物链接或垃圾链接待查看" if latest_id > seen_id else "")
        except Exception as e:
            print(f"更新每日任务红点失败: {e}")
    
    def update_resource_usage(self):
        """更新当前程序的资源使用情况"""
        try:
            cpu_percent = 0
            memory_info = "N/A"
            
            try:
                import psutil  # 系统资源监控
                # 获取当前进程
                current_process = psutil.Process()
                
                cpu_percent = current_process.cpu_percent(interval=None)
                
                # 获取当前进程的内存使用（单位 MB）
                memory_info_mb = current_process.memory_info().rss / 1024 / 1024
                memory_info = f"{memory_info_mb:.1f}MB"
                
            except Exception as e:
                memory_info = f"错误: {str(e)[:20]}"
            
            self.setWindowTitle(
                f"{MAIN_WINDOW_TITLE}    CPU:{cpu_percent}% | 内存:{memory_info}"
            )
            
        except Exception:
            self.setWindowTitle(MAIN_WINDOW_TITLE)

def frozen_runtime_self_test():
    import ssl
    import requests
    import openpyxl
    import xlrd
    import et_xmlfile
    from PyQt5.QtGui import QImageReader

    ssl.create_default_context()
    formats = {bytes(item).decode("ascii", "ignore").lower() for item in QImageReader.supportedImageFormats()}
    assert {"jpeg", "png", "webp"}.issubset(formats), formats
    return requests, openpyxl, xlrd, et_xmlfile


if __name__ == "__main__":
    if "--update-agent" in sys.argv:
        sys.exit(run_update_agent(VERSION))

    if "--runtime-self-test" in sys.argv:
        try:
            QApplication([])
            frozen_runtime_self_test()
        except Exception:
            with open(os.path.join(os.path.dirname(sys.executable), "runtime_self_test.log"), "w", encoding="utf-8") as f:
                f.write(traceback.format_exc())
            sys.exit(1)
        sys.exit(0)

    # 启用高分屏支持
    try:
        QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
        QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    except:
        pass
    
    app = QApplication(sys.argv)
    app.setQuitOnLastWindowClosed(False)
    load_bundled_app_fonts()
    install_window_icon_filter(app)
    tooltip_palette = app.palette()
    tooltip_palette.setColor(QPalette.Inactive, QPalette.ToolTipBase, QColor("#ffffff"))
    tooltip_palette.setColor(QPalette.Inactive, QPalette.ToolTipText, QColor("#111111"))
    app.setPalette(tooltip_palette)
    single_instance_mutex, already_running = acquire_single_instance_mutex()
    if already_running:
        if "--autostart" not in sys.argv:
            if not notify_existing_instance():
                activate_existing_window_by_title()
        sys.exit(0)

    window_holder = {}
    single_instance_server, should_exit = setup_single_instance(window_holder)
    if should_exit:
        sys.exit(0)

    font = QFont("微软雅黑", 10)
    app.setFont(font)
    main_icon = get_window_icon("main")
    if not main_icon.isNull():
        app.setWindowIcon(main_icon)
    
    style = """
    QPushButton {
        background-color: #f0f0f0;
        color: #333;
        border: 1px solid #ccc;
        padding: 8px 16px;
        border-radius: 6px;
        font-size: 13px;
    }
    QPushButton:hover {
        background-color: #e0e0e0;
    }
    QPushButton:pressed {
        background-color: #d0d0d0;
    }
    QPushButton:disabled {
        background-color: #f5f5f5;
        color: #999;
    }
    QToolTip {
        background-color: #ffffff;
        color: #111111;
        border: 1px solid #b8b8b8;
        padding: 4px 6px;
    }
    QLineEdit, QTextEdit, QSpinBox, QComboBox {
        border: 1px solid #dcdcdc;
        border-radius: 5px;
        padding: 6px 10px;
        background-color: white;
        font-size: 13px;
    }
    QLineEdit:focus, QTextEdit:focus, QSpinBox:focus, QComboBox:focus {
        border: 1px solid #3498db;
    }
    QComboBox::drop-down {
        border: none;
        width: 25px;
    }
    QComboBox::down-arrow {
        image: none;
        border-left: 5px solid transparent;
        border-right: 5px solid transparent;
        border-top: 6px solid #7f8c8d;
        margin-right: 8px;
    }
    QTableWidget {
        gridline-color: #000000;
        border: 2px solid #000000;
    }
    QTableWidget::item {
        padding: 5px;
    }
    QTableWidget::item:selected {
        background-color: #ffe0b2;
        color: black;
    }
    QTableWidget::item:focus {
        border: none;
        outline: none;
    }
    QHeaderView::section {
        background-color: #f8f9fa;
        color: #333;
        padding: 8px;
        border: none;
        border-bottom: 2px solid #3498db;
        font-weight: bold;
    }
    QScrollBar:vertical {
        border: none;
        background-color: #f0f0f0;
        width: 10px;
        margin: 0px;
    }
    QScrollBar::handle:vertical {
        background-color: #c0c0c0;
        border-radius: 5px;
        min-height: 20px;
    }
    QScrollBar::handle:vertical:hover {
        background-color: #a0a0a0;
    }
    QScrollBar:horizontal {
        border: none;
        background-color: #f0f0f0;
        height: 10px;
        margin: 0px;
    }
    QScrollBar::handle:horizontal {
        background-color: #c0c0c0;
        border-radius: 5px;
        min-width: 20px;
    }
    QScrollBar::handle:horizontal:hover {
        background-color: #a0a0a0;
    }
    QDialog {
        background-color: #fafafa;
    }
    QGroupBox {
        border: 1px solid #dcdcdc;
        border-radius: 8px;
        margin-top: 10px;
        padding-top: 10px;
        font-weight: bold;
    }
    QGroupBox::title {
        subcontrol-origin: margin;
        left: 10px;
        padding: 0 5px;
    }
    QMenu {
        background-color: white;
        border: 1px solid #dcdcdc;
        border-radius: 5px;
    }
    QMenu::item:selected {
        background-color: #3498db;
        color: white;
        border-radius: 3px;
    }
    QCheckBox, QRadioButton {
        spacing: 8px;
    }
    QCheckBox::indicator, QRadioButton::indicator {
        width: 18px;
        height: 18px;
        border-radius: 4px;
    }
    QCheckBox::indicator {
        border: 2px solid #dcdcdc;
        border-radius: 4px;
    }
    QRadioButton::indicator {
        border: 2px solid #dcdcdc;
        border-radius: 9px;
    }
    QCheckBox::indicator:checked, QRadioButton::indicator:checked {
        background-color: #3498db;
        border-color: #3498db;
    }
    QProgressBar {
        border: none;
        border-radius: 5px;
        background-color: #e0e0e0;
        text-align: center;
    }
    QProgressBar::chunk {
        background-color: #3498db;
        border-radius: 5px;
    }
    """
    app.setStyleSheet(style)
    
    append_event("startup:create_window")
    window = ShopManagerApp()
    append_event("startup:window_created")
    window_holder["window"] = window
    append_event("startup:before_show")
    if "--autostart" not in sys.argv:
        window.show()
    append_event("startup:after_show")
    append_event("startup:before_exec")
    sys.exit(app.exec_())
