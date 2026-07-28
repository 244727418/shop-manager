import os
import json
import tempfile
import time
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
import sys

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
sys.stdout.reconfigure(encoding="utf-8")

from PyQt5.QtCore import QEvent, QPoint, QRect, QTimer, Qt
from PyQt5.QtGui import QFont, QHelpEvent
from PyQt5.QtWidgets import QApplication, QComboBox, QDialog, QGridLayout, QLabel, QLineEdit, QMainWindow, QMenu, QPushButton, QTableWidget, QVBoxLayout, QWidget
from PyQt5 import sip

try:
    import manager.pdd_browser_monitor as pdd_browser_module
    import manager.shop_manager as shop_manager_module
    from manager.pdd_browser_monitor import PddBrowserMonitor
    from manager.archive_manager import ArchiveManager
    from manager.shop_manager import OrderedFlowLayout, ShopManagerApp, _promotion_link_visible, _set_application_font_family, auto_start_command
    from manager.db import SafeDatabaseManager
    from manager.widgets.product_store import ProductWidget, StoreWidget, _MetricState, _bubble_metric_foreground, _bubble_metric_typography, _expected_profit_for_100, _net_margin_background_color
    from manager.dialogs.promotion_data import PromotionDataDialog
    from manager.dialogs.records import OperationRecordDialog
    from manager.dialogs.store_margin import PddProductMatchDialog, StoreMarginDialog, StoreMarginExcelExporter
except ModuleNotFoundError:
    import pdd_browser_monitor as pdd_browser_module
    import shop_manager as shop_manager_module
    from pdd_browser_monitor import PddBrowserMonitor
    from archive_manager import ArchiveManager
    from shop_manager import OrderedFlowLayout, ShopManagerApp, _promotion_link_visible, _set_application_font_family, auto_start_command
    from db import SafeDatabaseManager
    from widgets.product_store import ProductWidget, StoreWidget, _MetricState, _bubble_metric_foreground, _bubble_metric_typography, _expected_profit_for_100, _net_margin_background_color
    from dialogs.promotion_data import PromotionDataDialog
    from dialogs.records import OperationRecordDialog
    from dialogs.store_margin import PddProductMatchDialog, StoreMarginDialog, StoreMarginExcelExporter


class SearchCheckWindow(QMainWindow):
    apply_realtime_search = ShopManagerApp.apply_realtime_search
    _split_search_terms = ShopManagerApp._split_search_terms
    clear_search_highlight = ShopManagerApp.clear_search_highlight
    _start_exact_search_flash = ShopManagerApp._start_exact_search_flash
    _advance_exact_search_flash = ShopManagerApp._advance_exact_search_flash
    _stop_exact_search_flash = ShopManagerApp._stop_exact_search_flash

    def __init__(self):
        super().__init__()
        self.search_input = QLineEdit("123456")
        self.row_data_map = {0: 1, 1: 2}
        self.product_store_map = {1: 10, 2: 10}
        self.current_store_filter = set()
        self.current_search_match_ids = {2}
        self._search_highlighted_rows = {1}
        self.main_view_mode = "operation"
        self.table = QTableWidget(2, 1)
        self.frozen_table = QTableWidget(2, 1)
        self.bubble_product_widgets = {}
        self.db = type("DB", (), {
            "safe_fetchall": lambda _self, *_args: [
                (1, "123456", "target", ""), (2, "999999", "other", "")
            ]
        })()
        self.highlight_events = []

    def _set_row_search_highlight(self, row, active):
        self.highlight_events.append((row, active))

    def apply_tag_filter(self, **_kwargs):
        for row in range(2):
            hidden = self.current_search_match_ids is not None and self.row_data_map[row] not in self.current_search_match_ids
            self.table.setRowHidden(row, hidden)
            self.frozen_table.setRowHidden(row, hidden)

    def _scroll_to_first_search_match(self, rows):
        self.scrolled_rows = set(rows)


def test_staggered_data_mode_render():
    app = QApplication.instance() or QApplication([])

    class StaggeredBubble(QLabel):
        def __init__(self, prod_id, prod_code, title, _image_data, _main_app, display_mode="bubble"):
            super().__init__(title)
            self.prod_id = prod_id
            self.prod_code = prod_code

        def set_search_highlight(self, active):
            self.search_highlighted = active

    class StaggeredHost:
        _render_next_data_mode_link = ShopManagerApp._render_next_data_mode_link

        def _visible_product_ids(self):
            return {1, 2}

        def is_real_promotion_data_mode(self):
            return False

        def _apply_data_mode_store_visibility(self):
            self.render_finished = True

        def _update_sticky_store_header(self):
            pass

        def _current_main_view_change_token(self):
            return (2, 1)

    section = QWidget()
    flow_layout = QVBoxLayout(section)
    staggered_host = StaggeredHost()
    staggered_host._data_mode_render_token = 2
    staggered_host._data_mode_render_index = 0
    staggered_host._data_mode_render_jobs = [
        (10, section, flow_layout, (1, "P1", "链接一", None)),
        (10, section, flow_layout, (2, "P2", "链接二", None)),
    ]
    staggered_host._data_mode_render_saved_scroll = 0
    staggered_host.product_store_map = {1: 10, 2: 10}
    staggered_host.bubble_product_widgets = {}
    staggered_host.current_search_match_ids = None
    staggered_host.db = type("DB", (), {"get_setting": lambda _self, *_args: "0"})()
    staggered_host.data_mode_scroll = type("Scroll", (), {
        "verticalScrollBar": lambda _self: type("Bar", (), {"setValue": lambda _self, _value: None})()
    })()

    original_product_widget = shop_manager_module.ProductWidget
    shop_manager_module.ProductWidget = StaggeredBubble
    try:
        staggered_host._render_next_data_mode_link(1)
        assert not staggered_host.bubble_product_widgets
        staggered_host._render_next_data_mode_link(2)
        assert list(staggered_host.bubble_product_widgets) == [1]
        deadline = time.time() + 1
        while not getattr(staggered_host, "render_finished", False) and time.time() < deadline:
            app.processEvents()
            time.sleep(0.01)
        assert list(staggered_host.bubble_product_widgets) == [1, 2]
        assert staggered_host.render_finished
    finally:
        shop_manager_module.ProductWidget = original_product_widget


def test_search_miss_keeps_all_links_visible():
    app = QApplication.instance() or QApplication([])
    search = SearchCheckWindow()
    search.search_input.setText("not-found")
    search.show_toast = lambda message, duration: setattr(search, "toast", (message, duration))

    search.apply_realtime_search()

    assert search.current_search_match_ids is None
    assert not search.table.isRowHidden(0) and not search.table.isRowHidden(1)
    assert search.toast == ("没有找到对应的内容", 500)


def test_store_task_ratio_uses_current_link_total():
    app = QApplication.instance() or QApplication([])

    class RatioDB:
        def safe_fetchall(self, query, params=()):
            assert params == (1,)
            if "COUNT(*) FROM products" in query:
                assert "is_archived" in query and "is_violation" not in query
                return [(42,)]
            return [(2, 3)]

    widget = StoreWidget.__new__(StoreWidget)
    QWidget.__init__(widget)
    widget.store_id = 1
    widget.display_mode = "bubble"
    widget.db = RatioDB()
    widget.task_ratio_widget = QWidget(widget)
    widget.garbage_ratio_badge = QLabel(widget)
    widget.garbage_ratio_label = QLabel(widget)
    widget.waste_ratio_badge = QLabel(widget)
    widget.waste_ratio_label = QLabel(widget)

    widget._refresh_garbage_ratio_label()

    assert "2/42" in widget.garbage_ratio_label.text()
    assert "3/42" in widget.waste_ratio_label.text()
    widget._store_foreground = "#fffdf5"
    widget.sync_flag_label = QLabel(widget)
    widget.label = QLabel(widget)
    widget.memo_label = QLabel(widget)
    widget.margin_label = QLabel(widget)
    widget.net_margin_label = QLabel(widget)
    widget.avg_price_label = QLabel(widget)
    widget._apply_bubble_store_label_styles()
    assert "color: #111111" in widget.garbage_ratio_label.styleSheet()
    assert "color: #111111" in widget.waste_ratio_label.styleSheet()


def test_font_family_switch_preserves_default_size():
    app = QApplication.instance() or QApplication([])
    original = QFont(app.font())
    enlarged = QFont(original)
    enlarged.setPointSize(13)
    app.setFont(enlarged)

    class RefreshProbe(QWidget):
        geometry_updates = 0
        paint_updates = 0

        def updateGeometry(self):
            self.geometry_updates += 1
            super().updateGeometry()

        def update(self):
            self.paint_updates += 1
            super().update()

    widget = RefreshProbe()
    try:
        _set_application_font_family("Microsoft YaHei")
        assert app.font().pointSize() == 13
        assert widget.font().family() == "Microsoft YaHei"
        assert widget.geometry_updates == 1
        assert widget.paint_updates == 1
    finally:
        widget.deleteLater()
        app.setFont(original)


def test_expected_profit_uses_natural_sales_or_promotion_spend_basis():
    assert round(_expected_profit_for_100(True, 0.30, 10, 4), 2) == 26.40
    assert round(_expected_profit_for_100(False, 0.30, 10, 4), 2) == 5.60
    assert _expected_profit_for_100(False, 0.30, 10, 0) is None


def test_created_time_sort_defaults_descending_and_direction_toggles():
    app = QApplication.instance() or QApplication([])
    with tempfile.TemporaryDirectory() as temp_dir:
        db = SafeDatabaseManager(str(Path(temp_dir) / "sort.db"))
        db.safe_execute("INSERT INTO stores (id, name, sort_order) VALUES (1, '店铺', 1)")
        db.safe_execute("INSERT INTO products (store_id, name) VALUES (1, '新链接')")
        assert db.safe_fetchall("SELECT created_at FROM products")[0][0]
        db.conn.close()

    sort_host = type("CreatedSortHost", (), {
        "_sort_products_for_display": ShopManagerApp._sort_products_for_display,
        "_build_product_sort_info": ShopManagerApp._build_product_sort_info,
        "get_product_card_data": lambda self, _product_id: {},
        "_get_product_order_count": lambda self, _code, _store_id: 0,
        "_calculate_product_net_margin": lambda self, _product_id: None,
    })()
    sort_host.product_sort_mode = "created_at"
    sort_host.product_sort_descending = True
    sort_host._product_task_states = {}
    products = [
        (1, "A", "", None, 1, "", "2026-07-01 10:00:00"),
        (2, "B", "", None, 2, "", "2026-07-03 10:00:00"),
        (3, "C", "", None, 3, "", "2026-07-02 10:00:00"),
    ]
    assert [row[0] for row in sort_host._sort_products_for_display(products)] == [2, 3, 1]
    sort_host.product_sort_descending = False
    assert [row[0] for row in sort_host._sort_products_for_display(products)] == [1, 3, 2]

    direction_host = type("DirectionHost", (), {
        "toggle": ShopManagerApp.toggle_product_sort_direction,
        "update": ShopManagerApp._update_product_sort_direction_button,
        "_update_product_sort_direction_button": ShopManagerApp._update_product_sort_direction_button,
        "_refresh_product_sort_display": lambda self: setattr(self, "refreshed", True),
    })()
    direction_host.product_sort_descending = True
    direction_host.btn_product_sort_direction = QPushButton()
    direction_host.db = type("DB", (), {
        "set_setting": lambda _self, key, value: setattr(direction_host, "saved", (key, value)),
    })()
    direction_host.toggle()
    assert not direction_host.product_sort_descending
    assert direction_host.btn_product_sort_direction.text() == "↑ 升序"
    assert direction_host.saved == ("product_sort_descending", "0") and direction_host.refreshed


def test_bubble_metrics_move_into_compact_chips_and_refresh_without_duplicates():
    app = QApplication.instance() or QApplication([])

    class BubbleHost:
        real_mode = False
        card_data = {}

        def is_real_promotion_data_mode(self):
            return self.real_mode

        def get_real_promotion_hidden_metrics(self):
            return set()

        def get_product_card_data(self, _prod_id):
            return self.card_data

    widget = ProductWidget.__new__(ProductWidget)
    QWidget.__init__(widget)
    widget.setFixedHeight(ProductWidget.BUBBLE_HEIGHT)
    widget.display_mode = "bubble"
    widget.prod_id = 1
    widget.main_app = BubbleHost()
    widget._search_highlight_active = False
    widget._bubble_foreground = "#171b18"
    widget._bubble_highlight_foreground = "#111111"
    widget._real_visible_metric_count = 99
    widget._bubble_avg_price = 25
    widget._bubble_expected_profit = None
    widget.margin_label = _MetricState("毛利率:20.00%")
    widget.link_order_label = _MetricState("单量:42单")
    widget.net_profit_label = _MetricState("净利率:3.00% 微盈利")
    widget.roi_label = _MetricState()
    widget.bubble_chips_widget = QWidget(widget)
    widget.bubble_chips_widget.setFixedHeight(37)
    widget.bubble_chip_labels = {
        key: QLabel(widget.bubble_chips_widget)
        for key in ("order", "promotion", "multiple", "status")
    }
    widget.bubble_metrics_widget = QWidget(widget)
    widget.bubble_metrics_layout = QGridLayout(widget.bubble_metrics_widget)
    widget.bubble_metric_chips = []
    widget.bubble_metrics_label = QLabel(widget)

    widget.main_app.card_data = {"current_roi": 4}
    widget.roi_label.setText("投产:4.00<br>投产倍数:1.20倍")
    widget._sync_bubble_metrics()
    assert widget.bubble_chip_labels["order"].text() == "单量42"
    assert widget.bubble_chip_labels["promotion"].text() == "推广中·稳定成本·投产4.00"
    assert widget.bubble_chip_labels["multiple"].text() == "倍数1.20倍"
    assert widget.bubble_chip_labels["status"].text() == "微盈利"
    assert all("color:#111" in chip.styleSheet() and "border-radius:6px" in chip.styleSheet()
               and "font-size" not in chip.styleSheet()
               and "padding:1px" in chip.styleSheet()
               for chip in widget.bubble_chip_labels.values())
    rendered = widget.bubble_metrics_label.text().replace("\u2060", "")
    assert "毛利率" in rendered and "净利率" in rendered and "客单价" in rendered
    assert rendered.count("<tr>") == 2 and rendered.count("<td ") == 3
    assert 'colspan="2"' in rendered
    assert "background-color:transparent" in rendered and "padding:1px" in rendered
    assert all(text not in rendered for text in ("单量", "投产:", "投产倍数", "微盈利"))
    assert [chip.text() for chip in widget.bubble_metric_chips] == [
        "毛利率:20.00%", "净利率:3.00%", "客单价:¥25.00",
    ]
    assert widget.bubble_metric_chips[0].toolTip() == "毛利率：商品毛利润占成交金额的比例。"
    assert widget.bubble_metric_chips[1].toolTip() == "净利率：净利润占成交金额的比例。"
    assert widget.bubble_metric_chips[2].toolTip() == "客单价：该链接平均每单的成交金额。"
    assert all(
        "border-radius:6px" in chip.styleSheet()
        and "padding:1px" in chip.styleSheet()
        and "background:transparent" not in chip.styleSheet()
        for chip in widget.bubble_metric_chips
    )

    widget._bubble_expected_profit = 12.3
    widget._sync_bubble_metrics()
    rendered = widget.bubble_metrics_label.text().replace("\u2060", "")
    assert rendered.count("<tr>") == 2 and rendered.count("<td ") == 4
    assert "客单价" in rendered and "预计盈亏" in rendered
    assert widget.bubble_metric_chips[3].toolTip().startswith("预计盈亏：按推广花费 100 元")
    widget._bubble_expected_profit = None

    widget.main_app.card_data = {"is_natural_flow": 1}
    widget.roi_label.setText("无推广")
    widget._bubble_expected_profit = 20
    widget._sync_bubble_metrics()
    assert widget.bubble_chip_labels["promotion"].text() == "无推广·自然流量"
    assert not widget.bubble_chip_labels["multiple"].text()
    assert "无推广" not in widget.bubble_metrics_label.text().replace("\u2060", "")
    assert widget.bubble_metric_chips[-1].toolTip().startswith("预计盈亏：按自然成交额 100 元")
    widget._bubble_expected_profit = None

    widget.main_app.card_data = {"is_sitewide_managed": 1, "sitewide_roi": 3.2}
    widget.roi_label.setText("全站:3.20<br>全站投产倍数:1.10倍")
    widget._sync_bubble_metrics()
    assert widget.bubble_chip_labels["promotion"].text() == "推广中·全站托管·投产3.20"
    assert widget.bubble_chip_labels["multiple"].text() == "倍数1.10倍"

    widget.main_app.card_data = {"roi_input_mode": "bid", "transaction_bid": 1.5}
    widget.roi_label.setText("出价:¥1.50<br>保本出价:¥2.00<br>出价倍数:0.75倍")
    widget._sync_bubble_metrics()
    assert widget.bubble_chip_labels["promotion"].text() == "推广中·成交出价¥1.50"
    assert widget.bubble_chip_labels["multiple"].text() == "出价倍数0.75倍"
    assert "保本出价" not in widget.bubble_metrics_label.text()

    widget.main_app.real_mode = True
    widget.main_app.card_data = {}
    widget.link_order_label.hide()
    widget.net_profit_label.setText(
        "净成交:12单 净投产比:2.10<br>净利润:¥30.00 净利率:6.00% 盈利"
    )
    widget.roi_label.setText(
        "投产倍数:1.30倍 曝光占比:25.00%<br>每笔成交:¥20.00 每笔花费:¥5.00"
    )
    widget._sync_bubble_metrics()
    assert widget.bubble_chip_labels["order"].text() == "净成交12"
    assert widget.bubble_chip_labels["promotion"].text() == "真实推广·净投产比2.10"
    assert widget.bubble_chip_labels["multiple"].text() == "倍数1.30倍"
    assert widget.bubble_chip_labels["status"].text() == "盈利"
    rendered = widget.bubble_metrics_label.text().replace("\u2060", "")
    assert all(text not in rendered for text in ("净成交", "净投产比", "投产倍数", "盈利"))
    assert "曝光占比" in rendered and "每笔花费" in rendered
    assert widget.height() == 104


def test_data_mode_layout_and_refresh():
    app = QApplication.instance() or QApplication([])

    if sys.platform == "win32" and not getattr(sys, "frozen", False):
        pythonw = Path(sys.executable).with_name("pythonw.exe")
        if pythonw.is_file():
            assert str(pythonw).lower() in auto_start_command().lower()

    class HintHost(ShopManagerApp):
        def __init__(self):
            QMainWindow.__init__(self)
            self._button_hint_widgets = {}
            self._button_hint_active = None
            self._button_hint_timer = QTimer(self)
            self._button_hint_timer.setSingleShot(True)
            self._button_hint_timer.timeout.connect(self._show_button_hint)
            self._button_hint_label = QLabel(self)
            self._button_hint_label.setWindowFlags(Qt.ToolTip | Qt.FramelessWindowHint)
            self._button_hint_label.hide()

    hint_host = HintHost()
    hint_button = QPushButton("提示", hint_host)
    next_hint_button = QPushButton("下一个", hint_host)
    hint_host._install_main_button_hints({hint_button: "新的按钮解释", next_hint_button: "下一个解释"})
    QApplication.sendEvent(hint_button, QEvent(QEvent.Enter))
    assert hint_host._button_hint_active is hint_button
    assert hint_host._button_hint_timer.isActive() and hint_host._button_hint_timer.interval() == 500
    QApplication.sendEvent(hint_button, QEvent(QEvent.Leave))
    assert hint_host._button_hint_active is None and not hint_host._button_hint_timer.isActive()
    QApplication.sendEvent(next_hint_button, QEvent(QEvent.Enter))
    assert hint_host._button_hint_active is next_hint_button
    assert hint_host._button_hint_timer.isActive() and hint_host._button_hint_timer.interval() == 500

    deferred_host = type("DeferredRefreshHost", (), {
        "request": ShopManagerApp._request_manual_main_refresh,
        "_run_manual_incremental_refresh": lambda self: setattr(self, "ran", True),
    })()
    deferred_host._manual_refresh_pending = False
    deferred_host.request()
    assert not hasattr(deferred_host, "ran")
    app.processEvents()
    assert deferred_host.ran

    unchanged_host = type("UnchangedRefreshHost", (), {
        "refresh": ShopManagerApp._run_manual_incremental_refresh,
        "_current_main_view_change_token": lambda self: (3, 1),
        "show_toast": lambda self, text: setattr(self, "toast", text),
    })()
    unchanged_host._manual_refresh_pending = True
    unchanged_host._main_view_change_token = (3, 1)
    unchanged_host.is_loading = False
    unchanged_host.refresh()
    assert unchanged_host.toast == "没有内容变更"

    class IncrementalRefreshDB:
        def safe_fetchall(self, query, _params=()):
            return [(1,)] if query == "SELECT id FROM stores" else [(10, 1)]

    incremental_host = type("IncrementalRefreshHost", (), {
        "refresh": ShopManagerApp._run_manual_incremental_refresh,
        "_current_main_view_change_token": lambda self: (4, 1),
        "force_refresh_product_widget": lambda self, product_id: setattr(self, "refreshed_product", product_id),
        "_refresh_store_margin_widgets": lambda self: setattr(self, "store_metrics_refreshed", True),
        "refresh_store_sheets": lambda self: None,
        "_reorder_data_mode_bubbles": lambda self: None,
        "_apply_data_mode_store_visibility": lambda self: True,
        "update_daily_task_button_badge": lambda self: None,
        "show_toast": lambda self, text: setattr(self, "toast", text),
        "load_data_safe": lambda self, **_kwargs: (_ for _ in ()).throw(AssertionError("must stay incremental")),
    })()
    incremental_host.db = IncrementalRefreshDB()
    incremental_host._manual_refresh_pending = True
    incremental_host._main_view_change_token = (3, 1)
    incremental_host.is_loading = False
    incremental_host.row_store_map = {0: 1}
    incremental_host.product_store_map = {10: 1}
    incremental_host.bubble_product_widgets = {10: object()}
    incremental_host.refresh()
    assert incremental_host.refreshed_product == 10
    assert incremental_host.store_metrics_refreshed and incremental_host.toast == "主界面已刷新"

    assert _bubble_metric_foreground(10, _net_margin_background_color(10)) == "#171b18"
    assert _bubble_metric_foreground(20, _net_margin_background_color(20)) == "#fffdf5"
    assert _bubble_metric_foreground(-5, _net_margin_background_color(-5)) == "#171b18"
    assert _bubble_metric_foreground(-10, _net_margin_background_color(-10)) == "#171b18"
    assert _bubble_metric_foreground(-15, _net_margin_background_color(-15)) == "#fffdf5"

    from openpyxl import Workbook

    exporter = StoreMarginExcelExporter.__new__(StoreMarginExcelExporter)
    exporter.load_manual_data = lambda: []
    exporter._style_historical_export_sheet = lambda *_args, **_kwargs: None
    exporter._append_reading_mode_images_to_historical_sheet = (
        lambda ws: ws.cell(3, 1, "reading-mode-image")
    )
    workbook = Workbook()
    exporter._write_historical_export_sheet(workbook, "店铺A", create_sheet=True)
    exporter._write_historical_export_sheet(workbook, "店铺B", create_sheet=True)
    workbook.remove(workbook["Sheet"])
    assert workbook.sheetnames == ["店铺A", "店铺B"]
    assert all(workbook[name]["A3"].value == "reading-mode-image" for name in workbook.sheetnames)

    test_single_file_detailed_excel_layout()

    host = QWidget()
    layout = OrderedFlowLayout(host, spacing=10)
    cards = []
    for width, height in ((320, 180), (400, 220), (350, 160)):
        card = QLabel("card")
        card.setFixedSize(width, height)
        cards.append(card)
        layout.addWidget(card)
        assert card.parentWidget() is host

    layout.setGeometry(QRect(0, 0, 750, 800))
    rects = [layout.itemAt(index).widget().geometry().getRect() for index in range(layout.count())]
    assert rects == [(0, 0, 320, 180), (330, 0, 400, 220), (0, 230, 350, 160)]
    assert layout.heightForWidth(750) == 390
    cards[1].hide()
    layout.setGeometry(QRect(0, 0, 750, 800))
    assert cards[2].geometry().getRect() == (330, 0, 350, 160)
    assert layout.heightForWidth(750) == 180

    search = SearchCheckWindow()
    search.apply_realtime_search()
    app.processEvents()
    assert search.current_search_match_ids is None
    assert not search.table.isRowHidden(0) and not search.table.isRowHidden(1)
    assert search._exact_search_flash_timer.interval() * search._exact_search_flash_remaining == 2000
    assert search.scrolled_rows == {0}
    for _ in range(8):
        search._advance_exact_search_flash()
    assert search._exact_search_flash_row is None

    class TaskFilterHost:
        apply_tag_filter = ShopManagerApp.apply_tag_filter
        get_all_product_ids_with_current_store = ShopManagerApp.get_all_product_ids_with_current_store

        def __init__(self):
            for name in (
                "coupon", "new_customer", "limited_time", "marketing", "natural_flow",
                "sitewide", "garbage", "waste", "profit", "loss", "break_even",
                "missing_roi_bid",
            ):
                button = QPushButton()
                button.setCheckable(True)
                setattr(self, f"btn_filter_{name}", button)
            self.btn_tag_filter = QPushButton()
            self.tag_filter_menu = QWidget()
            self.product_store_map = {1: 10, 2: 10, 3: 20, 4: 20}
            self.current_category_filter = ""
            self.current_search_match_ids = None
            self.current_store_filter = set()
            self.current_filter_tags = set()
            self.db = type("TaskFilterDB", (), {"safe_fetchall": self._fetch})()

        @staticmethod
        def _fetch(_db, query, params=()):
            if "FROM daily_tasks" not in query:
                return []
            ids = set()
            if "【垃圾链接】%" in params:
                ids.update((1, 3))
            if "【废物链接】%" in params:
                ids.update((2, 3))
            return [(product_id,) for product_id in sorted(ids)]

        def _set_visible_product_ids(self, product_ids):
            self.visible_ids = set(product_ids)

        def _apply_data_mode_store_visibility(self):
            return True

        def _refresh_data_mode_view_if_active(self):
            return None

        def show_toast(self, _text):
            return None

    task_filter = TaskFilterHost()
    task_filter.btn_filter_garbage.setChecked(True)
    task_filter.btn_filter_waste.setChecked(True)
    task_filter.apply_tag_filter(show_message=False)
    assert task_filter.visible_ids == {1, 2, 3}
    assert task_filter.current_filter_tags["tasks"] == ["garbage", "waste"]
    task_filter.btn_filter_waste.setChecked(False)
    task_filter.apply_tag_filter(show_message=False)
    assert task_filter.visible_ids == {1, 3}

    context_calls = []
    context_app = type("ContextApp", (), {
        "product_store_map": {8: 3},
        "open_pdd_code_fetch_for_store": lambda _self, store_id, product_id="": context_calls.append((store_id, product_id)),
    })()
    context_widget = ProductWidget.__new__(ProductWidget)
    QWidget.__init__(context_widget)
    context_widget.prod_id = 8
    context_widget.prod_code = "9513241661"
    context_widget.main_app = context_app
    context_widget.is_violation = False
    context_widget.refresh_violation_state = lambda *args, **kwargs: None
    original_menu_exec = QMenu.exec_
    QMenu.exec_ = lambda menu, *_args: next(action for action in menu.actions() if action.text() == "抓取添加编码")
    try:
        context_widget.show_product_context_menu(QPoint(0, 0))
    finally:
        QMenu.exec_ = original_menu_exec
    assert context_calls == [(3, "9513241661")]

    saved_memos = []
    saved_records = []
    memo_dialog = type("MemoRecordHost", (), {
        "new_text_edit": type("Text", (), {"text": lambda _self: ""})(),
        "memo_edit": type("Memo", (), {"toPlainText": lambda _self: "长期备注"})(),
        "records": [],
        "memo_save_callback": saved_memos.append,
        "save_with_date": False,
        "save_callback": lambda _self, records: saved_records.append(records),
        "accept": lambda _self: None,
    })()
    OperationRecordDialog.save(memo_dialog)
    assert saved_memos == ["长期备注"] and saved_records == [[]]

    quick_reminders = []
    quick_host = type("QuickReminderHost", (), {
        "new_text_edit": type("Text", (), {"text": lambda _self: "检查推广设置"})(),
        "quick_reminder_callback": lambda _self, text, when: quick_reminders.append((text, when)),
    })()
    before_quick_reminder = datetime.now()
    OperationRecordDialog.quick_set_reminder(quick_host)
    reminder_time = datetime.strptime(quick_reminders[0][1], "%Y-%m-%d %H:%M:%S")
    assert quick_reminders[0][0] == "检查推广设置"
    assert 23.99 * 3600 <= (reminder_time - before_quick_reminder).total_seconds() <= 24.01 * 3600

    class ReminderPayloadDB:
        def safe_fetchall(self, query, _params=()):
            if "FROM stores" in query:
                return [("测试店铺",)]
            if "FROM products" in query:
                return [("9513241661", "商品标题", b"image", "草稿本", "")]
            return []

    reminder_payload_host = type("ReminderPayloadHost", (), {"db": ReminderPayloadDB()})()
    reminder_payload = ShopManagerApp._build_task_reminder_payload(
        reminder_payload_host, 1, 3, 8, "检查推广设置", "2026-07-18 10:00:00"
    )
    assert reminder_payload["product_code"] == "9513241661"
    assert reminder_payload["product_image_data"] == b"image"
    assert reminder_payload["link_type"] == "草稿本"

    fast_monitor = PddBrowserMonitor.__new__(PddBrowserMonitor)
    browser_calls = []
    browser_events = []
    fast_monitor.activate_store_browser = lambda store_id, **_kwargs: browser_calls.append(store_id)
    fast_monitor.is_devtools_alive = lambda: True
    fast_monitor._get_pdd_target = lambda: {
        "url": "https://mms.pinduoduo.com/goods/goods_list",
        "webSocketDebuggerUrl": "ws://test",
    }
    fast_monitor.get_current_store_name = lambda expected_store_name="": (
        browser_events.append("store-check")
        or {"ok": True, "store_name": expected_store_name}
    )

    class FastDevToolsWebSocket:
        def __init__(self, *_args, **_kwargs):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            pass

        def call(self, method, _params):
            assert method == "Runtime.evaluate"
            browser_events.append("evaluate")
            return {"result": {"value": {"ok": True, "status": "fast"}}}

    original_devtools_websocket = pdd_browser_module.DevToolsWebSocket
    pdd_browser_module.DevToolsWebSocket = FastDevToolsWebSocket
    try:
        started = time.perf_counter()
        fast_result = fast_monitor.open_goods_list_and_search_product(
            "9513241661", expected_store_name="测试店铺", store_id=3
        )
        elapsed = time.perf_counter() - started
    finally:
        pdd_browser_module.DevToolsWebSocket = original_devtools_websocket
    assert fast_result["ok"] and browser_calls == [3]
    assert browser_events == ["evaluate", "store-check", "evaluate"]
    assert elapsed < 0.5

    tooltip_widget = ProductWidget.__new__(ProductWidget)
    QWidget.__init__(tooltip_widget)
    metric_chip = QLabel(tooltip_widget)
    metric_chip.setToolTip("预计盈亏：按推广花费 100 元估算，结合当前投产或成交出价计算。")
    tooltip_widget.bubble_metric_chips = [metric_chip]
    tooltip_widget.bubble_chip_labels = {}
    tooltip_widget.eventFilter(metric_chip, QHelpEvent(QEvent.ToolTip, QPoint(1, 1), QPoint(10, 10)))
    assert tooltip_widget._bubble_tooltip.text() == metric_chip.toolTip()
    assert tooltip_widget._bubble_tooltip.size() == tooltip_widget._bubble_tooltip.sizeHint()
    assert "background: #ffffff" in tooltip_widget._bubble_tooltip.styleSheet()
    assert "color: #111111" in tooltip_widget._bubble_tooltip.styleSheet()
    tooltip_widget._hide_bubble_tooltip()
    assert tooltip_widget._bubble_tooltip.isHidden()

    link_type_widget = ProductWidget.__new__(ProductWidget)
    QWidget.__init__(link_type_widget)
    link_type_widget.prod_id = 1
    link_type_widget.display_mode = "bubble"
    link_type_widget.category_label = QLabel()
    link_type_widget.main_app = type("LinkTypeApp", (), {
        "get_product_card_data": lambda _self, _product_id: {
            "product_category_label": "",
            "link_type": "这是一个很长的链接类型名称",
        },
    })()
    link_type_widget.update_product_category_display()
    assert link_type_widget.category_label.text() == "这是一个很长的链接类型名称"
    assert link_type_widget.category_label.width() <= 132
    assert 8 <= link_type_widget._category_font_size <= 12
    assert link_type_widget.category_label.height() == 18
    assert link_type_widget.category_label.contentsRect().left() >= 6
    assert link_type_widget.category_label.toolTip().endswith("双击打开原有编辑入口")

    task_badge_widget = ProductWidget.__new__(ProductWidget)
    QWidget.__init__(task_badge_widget)
    task_badge_widget.prod_id = 1
    task_badge_widget.display_mode = "bubble"
    task_badge_widget.reminder_badge = QLabel()
    task_badge_widget.garbage_badge = QLabel()
    task_badge_widget.garbage_streak_badge = QLabel(task_badge_widget.garbage_badge)
    task_badge_widget.waste_badge = QLabel()
    task_badge_widget.db = type("TaskBadgeDB", (), {
        "safe_fetchall": lambda _self, query, *_args: (
            [(1,)] if "FROM task_reminders" in query
            else [("【垃圾链接】连续3次测试",), ("【废物链接】测试",)]
        )
    })()
    task_badge_widget.update_task_badge()
    assert not task_badge_widget.reminder_badge.isHidden()
    assert not task_badge_widget.garbage_badge.isHidden()
    assert task_badge_widget.garbage_streak_badge.text() == "3"
    assert not task_badge_widget.garbage_streak_badge.isHidden()
    assert not task_badge_widget.waste_badge.isHidden()

    category_widget = ProductWidget.__new__(ProductWidget)
    QWidget.__init__(category_widget)
    category_widget.prod_id = 1
    category_widget.display_mode = "bubble"
    category_widget.category_label = QLabel()
    category_widget.main_app = type("CategoryHost", (), {
        "get_product_card_data": lambda _self, _product_id: {"link_type": "草稿本"}
    })()
    category_widget.update_product_category_display()
    assert category_widget.category_label.width() < 90

    class WasteTaskDB:
        def __init__(self):
            self.writes = []

        def safe_fetchall(self, query, _params=()):
            if "SELECT id, created_at FROM products" in query:
                now = datetime.now()
                return [
                    (1, (now - timedelta(days=2)).isoformat(sep=" ")),
                    (2, (now - timedelta(days=8)).isoformat(sep=" ")),
                    (3, (now - timedelta(days=8)).isoformat(sep=" ")),
                ]
            if "SELECT title FROM products" in query:
                return [("C",)]
            return []

        def safe_execute(self, query, params=()):
            self.writes.append((query, params))

    waste_host = type("WasteHost", (), {
        "store_id": 1,
        "main_app": None,
    })()
    waste_host.db = WasteTaskDB()
    StoreMarginDialog._update_waste_link_tasks_after_order_import(
        waste_host, [(1, "a"), (2, "b"), (3, "c")], {"b": 1}
    )
    waste_inserts = [params for query, params in waste_host.db.writes if "INSERT INTO daily_tasks" in query]
    assert len(waste_inserts) == 1 and waste_inserts[0][1] == 3

    class GarbageTaskDB:
        def __init__(self):
            self.queries = []
            self.created = []

        def safe_fetchall(self, query, params=()):
            self.queries.append((query, params))
            if "FROM promotion_daily_data" in query:
                return [
                    ("p", 0, 0, 0, 0, 10, 0, 0),
                    ("p", 0, 0, 0, 0, 20, 0, 0),
                ]
            if "FROM products" in query:
                return [(1, "p", "p", 0)]
            return []

        def safe_execute(self, query, params=()):
            self.created.append((query, params))

    garbage_db = GarbageTaskDB()
    created = SafeDatabaseManager.reconcile_garbage_link_tasks(
        garbage_db,
        1,
        "2026-07-11 12:00:00",
    )
    assert created == 1
    garbage_inserts = [params for query, params in garbage_db.created if "INSERT INTO daily_tasks" in query]
    assert len(garbage_inserts) == 1 and "连续2次" in garbage_inserts[0][5]

    ratio_widget = StoreWidget.__new__(StoreWidget)
    QWidget.__init__(ratio_widget)
    ratio_widget.store_id = 1
    ratio_widget.display_mode = "bubble"
    ratio_widget.task_ratio_widget = QWidget()
    ratio_widget.garbage_ratio_badge = QLabel()
    ratio_widget.garbage_ratio_label = QLabel()
    ratio_widget.waste_ratio_badge = QLabel()
    ratio_widget.waste_ratio_label = QLabel()
    ratio_widget.db = type("RatioDB", (), {"safe_fetchall": lambda _self, query, _params: [(19,)] if "COUNT(*)" in query else [(2, 3)]})()
    ratio_widget._refresh_garbage_ratio_label()
    assert "2/19" in ratio_widget.garbage_ratio_label.text()
    assert "3/19" in ratio_widget.waste_ratio_label.text()

    class StoreAddDB:
        def __init__(self):
            self.inserted = []

        def safe_fetchall(self, _query, _params=()):
            return [(3,)]

        def safe_execute(self, _query, params=()):
            self.inserted.append(params)
            return type("Cursor", (), {"lastrowid": 4})()

    store_add_host = type("StoreAddHost", (), {"add_store": ShopManagerApp.add_store})()
    store_add_host.db = StoreAddDB()
    store_add_host.main_view_mode = "operation"
    store_add_host.refresh_store_sheets = lambda: setattr(store_add_host, "sheets_refreshed", True)
    store_add_host.show_toast = lambda text: setattr(store_add_host, "toast", text)
    original_get_text = shop_manager_module.QInputDialog.getText
    shop_manager_module.QInputDialog.getText = staticmethod(lambda *_args: ("新店铺", True))
    try:
        store_add_host.add_store()
    finally:
        shop_manager_module.QInputDialog.getText = original_get_text
    assert store_add_host.db.inserted == [("新店铺", 4)]
    assert store_add_host.sheets_refreshed and store_add_host.toast == "已添加店铺：新店铺"

    miss = SearchCheckWindow()
    miss.search_input.setText("target")
    miss.apply_realtime_search()
    assert miss.current_search_match_ids == {1}
    assert miss.scrolled_rows == {0}
    assert not miss.table.isRowHidden(0) and miss.table.isRowHidden(1)

    class PromotionDB:
        def __init__(self):
            self.calls = []

        def safe_fetchall(self, query, params=()):
            self.calls.append((query, params))
            if "MAX(record_date)" in query:
                return [("2026-07-02",)]
            return [("2026-07-02", 10, 20, 18, 1.8, 2, 0.4, 5, 0.1, 0.2, 3, 16.7)]

    promotion_host = type("PromotionHost", (), {
        "get_latest_promotion_data": ShopManagerApp.get_latest_promotion_data,
    })()
    promotion_host.db = PromotionDB()
    promotion_host._latest_promotion_data_cache = {}
    first = promotion_host.get_latest_promotion_data(1, "123")
    second = promotion_host.get_latest_promotion_data(1, "123")
    assert first == second and len(promotion_host.db.calls) == 2
    promotion_host.get_latest_promotion_data(2, "123")
    assert promotion_host.db.calls[2][1][0] == 2

    class FallbackPromotionDB(PromotionDB):
        def safe_fetchall(self, query, params=()):
            self.calls.append((query, params))
            if "MAX(record_date)" in query:
                return [("2026-07-03",)]
            if "record_date=?" in query:
                return []
            return [("2026-07-02", 10, 20, 18, 1.8, 2, 0.4, 5, 0.1, 0.2, 3, 16.7)]

    fallback_host = type("FallbackPromotionHost", (), {
        "get_latest_promotion_data": ShopManagerApp.get_latest_promotion_data,
    })()
    fallback_host.db = FallbackPromotionDB()
    fallback_host._latest_promotion_data_cache = {}
    assert fallback_host.get_latest_promotion_data(1, "123")["record_date"] == "2026-07-02"

    class CacheDB:
        def __init__(self):
            self.calls = []

        def safe_fetchall(self, query, params=()):
            self.calls.append((query, params))
            product_a = (101, "A", 0, 0, 0, 0, 0, 0, 0, "", 0, 0, 1, "roi", 0, "", "", "", 0, 1)
            product_b = (202, "B", 0, 0, 0, 0, 0, 0, 0, "", 0, 0, 2, "roi", 0, "", "", "", 0, 0)
            if "FROM products p" in query and "WHERE p.id=?" in query:
                row = product_b if params == (202,) else product_a
                return [row[1:]]
            if "FROM products p" in query:
                return [product_a, product_b]
            if "FROM daily_tasks" in query:
                return [(101, 1, 0, 0, "【垃圾链接】连续2次测试")]
            if "FROM task_reminders" in query:
                return [(101,)]
            if "GROUP BY store_id, product_id" in query:
                return [(1, "A", 5), (2, "B", 7)]
            if "FROM imported_orders" in query:
                return [(7,)]
            return []

        def calculate_products_gross_margin_metrics(self, _product_ids):
            return {}

    cache_host = type("CacheHost", (), {
        "_prepare_product_card_caches": ShopManagerApp._prepare_product_card_caches,
        "get_product_card_data": ShopManagerApp.get_product_card_data,
        "_get_product_order_count": ShopManagerApp._get_product_order_count,
    })()
    cache_host.db = CacheDB()
    cache_host._product_card_data_cache = None
    cache_host._prepare_product_card_caches(1)
    assert cache_host.get_product_card_data(202)["store_id"] == 2
    assert cache_host.get_product_card_data(101)["is_violation"] == 1
    assert cache_host._product_task_states[101] == (True, False, True, "【垃圾链接】连续2次测试")

    cached_widget = ProductWidget.__new__(ProductWidget)
    QWidget.__init__(cached_widget)
    cached_widget.prod_id = 101
    cached_widget.main_app = cache_host
    cached_widget.db = type("NoQueryDB", (), {
        "safe_fetchall": lambda _self, *_args: (_ for _ in ()).throw(AssertionError("cached state queried the database")),
    })()
    cached_widget.display_mode = "bubble"
    cached_widget.reminder_badge = QLabel()
    cached_widget.garbage_badge = QLabel()
    cached_widget.waste_badge = QLabel()
    cached_widget.violation_overlay = QWidget()
    cached_widget.refresh_violation_state()
    cached_widget.update_task_badge()
    assert cached_widget.is_violation and not cached_widget.violation_overlay.isHidden()
    assert not cached_widget.reminder_badge.isHidden() and not cached_widget.garbage_badge.isHidden()
    assert cached_widget.waste_badge.isHidden()

    cache_host._product_order_count_cache = {(1, "A"): 5}
    assert cache_host._get_product_order_count("B", 2) == 7

    metric_state = _MetricState("净利率: -")
    metric_state.setText("净利率: 5%")
    metric_state.setToolTip("说明")
    metric_state.hide()
    assert metric_state.text() == "净利率: 5%" and metric_state.toolTip() == "说明"
    assert metric_state.isHidden()
    metric_state.show()
    assert not metric_state.isHidden()

    with tempfile.TemporaryDirectory() as profile_tmp:
        import shutil
        source = Path(profile_tmp) / "source.db"
        standard = Path(profile_tmp) / "account" / "local_current" / "backup.db"
        source.write_bytes(b"same-profile")
        standard.parent.mkdir(parents=True)
        shutil.copy2(source, standard)
        archive_host = ArchiveManager.__new__(ArchiveManager)
        archive_host.accounts = [{"id": "a", "folder": "account", "local_backup_path": str(standard.parent.parent)}]
        original_copy = shutil.copy2
        shutil.copy2 = lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("unchanged profile copied"))
        try:
            assert archive_host.ensure_local_profile_normalized("a", str(source)) == (True, str(standard))
        finally:
            shutil.copy2 = original_copy

    assert not _promotion_link_visible(None, True)
    assert not _promotion_link_visible({"net_orders": 0}, True)
    assert _promotion_link_visible({"net_orders": 1}, True)
    assert _promotion_link_visible(None, False)
    assert _bubble_metric_typography(True, 2) == (16, 22)
    assert _bubble_metric_typography(True, 5) == (12, 15)
    assert _bubble_metric_typography(True, 12) == (11, 12)

    class RealDisplayHost:
        def __init__(self, hidden=None):
            self.hidden = set(hidden or set())

        def get_real_promotion_hidden_metrics(self):
            return self.hidden

        def get_latest_promotion_data(self, *_args):
            return {
                "cost": 10,
                "transaction_amount": 120,
                "net_transaction_amount": 100,
                "net_roi": 10,
                "net_orders": 2,
                "promotion_impression_share": 0.3,
                "cost_per_net_order": 5,
                "ctr": 0.1,
                "click_conversion_rate": 0.2,
                "net_profit": 25,
                "net_margin_rate": 25,
            }

    def real_display_widget(hidden=None):
        widget = ProductWidget.__new__(ProductWidget)
        QWidget.__init__(widget)
        widget.display_mode = "bubble"
        widget.main_app = RealDisplayHost(hidden)
        widget.prod_code = "p"
        widget._search_highlight_active = False
        widget._bubble_background = "#545e47"
        widget._bubble_foreground = "#fffdf5"
        widget._bubble_highlight_color = "#00e5ff"
        widget._bubble_highlight_foreground = "#111111"
        widget.code_label = QLabel()
        widget.real_date_label = QLabel()
        widget.category_label = QLabel()
        widget.coupon_amount_label = QLabel()
        widget.new_customer_amount_label = QLabel()
        widget.metrics_panel = QWidget()
        widget.margin_label = QLabel()
        widget.net_profit_label = QLabel()
        widget.roi_label = QLabel()
        widget.link_order_label = QLabel()
        return widget

    real_widget = real_display_widget()
    real_widget._apply_real_promotion_display(1, 0.35, 3)
    assert "毛利率:" in real_widget.margin_label.text()
    assert "35.00%" in real_widget.margin_label.text()
    assert "毛利率:" in real_widget.margin_label.text().split("<br>")[0]

    hidden_real_widget = real_display_widget({"gross_margin_rate"})
    hidden_real_widget._apply_real_promotion_display(1, 0.35, 3)
    assert "毛利率:" not in hidden_real_widget.margin_label.text()
    assert "花费:" in hidden_real_widget.margin_label.text()

    class ToggleDB:
        def __init__(self):
            self.values = {"real_promotion_show_ordered_data_only": "1"}

        def set_setting(self, key, value):
            self.values[key] = value

        def get_setting(self, key, default=None):
            return self.values.get(key, default)

    toggle_host = type("ToggleHost", (), {
        "toggle": ShopManagerApp.toggle_real_promotion_data_mode,
        "_update_real_promotion_mode_button_style": lambda self: None,
        "_apply_real_promotion_display_settings": lambda self, value: setattr(self, "applied", value),
        "_update_sticky_store_header": lambda self, **kwargs: None,
        "load_data_safe": lambda self: setattr(self, "rebuilt", True),
    })()
    toggle_host.btn_real_promotion_mode = type("Button", (), {"isChecked": lambda self: True})()
    toggle_host.db = ToggleDB()
    toggle_host.main_view_mode = "data"
    toggle_host._data_mode_store_sections = {}
    toggle_host.toggle()
    assert toggle_host.applied is True and not hasattr(toggle_host, "rebuilt")

    class StoreFilterDB:
        def __init__(self):
            self.values = {}

        def set_setting(self, key, value):
            self.values[key] = value

    class TextButton:
        def setText(self, text):
            self.text = text

    store_filter_host = type("StoreFilterHost", (), {
        "select": ShopManagerApp.select_store_sheet,
        "_set_visible_product_ids": ShopManagerApp._set_visible_product_ids,
        "_visible_product_ids": ShopManagerApp._visible_product_ids,
        "_has_active_product_filter_inputs": lambda self: False,
        "_apply_data_mode_store_visibility": lambda self: setattr(self, "visibility_applied", True) or True,
        "_refresh_data_mode_view_if_active": lambda self: setattr(self, "rebuilt", True),
        "update_store_sheet_selection": lambda self: None,
        "apply_tag_filter": lambda self, **_kwargs: setattr(self, "tagged", True),
    })()
    store_filter_host.main_view_mode = "data"
    store_filter_host.current_store_filter = set()
    store_filter_host.product_store_map = {1: 10, 2: 20}
    store_filter_host.db = StoreFilterDB()
    store_filter_host.btn_store_filter = TextButton()
    store_filter_host.select(10)
    assert store_filter_host.current_store_filter == {10}
    assert store_filter_host.visible_product_ids == {1}
    assert store_filter_host.visibility_applied and not hasattr(store_filter_host, "rebuilt")

    class CheckButton:
        def __init__(self, checked=False):
            self.checked = checked

        def isChecked(self):
            return self.checked

    tag_store_host = type("TagStoreHost", (), {
        "apply": ShopManagerApp.apply_tag_filter,
        "_set_visible_product_ids": ShopManagerApp._set_visible_product_ids,
        "_apply_data_mode_store_visibility": lambda self: setattr(self, "visibility_applied", True) or True,
        "_refresh_data_mode_view_if_active": lambda self: setattr(self, "rebuilt", True),
        "get_all_product_ids_with_current_store": ShopManagerApp.get_all_product_ids_with_current_store,
        "show_toast": lambda self, *_args: None,
    })()
    tag_store_host.main_view_mode = "data"
    tag_store_host.current_store_filter = {10}
    tag_store_host.product_store_map = {1: 10, 2: 10, 3: 20}
    tag_store_host.btn_filter_coupon = CheckButton(True)
    for name in (
        "btn_filter_new_customer", "btn_filter_limited_time", "btn_filter_marketing",
        "btn_filter_natural_flow", "btn_filter_sitewide", "btn_filter_garbage",
        "btn_filter_waste", "btn_filter_profit",
        "btn_filter_loss", "btn_filter_break_even", "btn_filter_missing_roi_bid",
    ):
        setattr(tag_store_host, name, CheckButton(False))
    tag_store_host.btn_tag_filter = TextButton()
    tag_store_host.db = type("DB", (), {
        "safe_fetchall": lambda self, sql, params=(): [(5,)] if params and params[0] in (1, 3) else [(0,)]
    })()
    tag_store_host.apply(close_menu=False, show_message=False)
    assert tag_store_host.visible_product_ids == {1}
    assert tag_store_host.visibility_applied and not hasattr(tag_store_host, "rebuilt")

    visible_host = type("VisibleHost", (), {
        "apply": ShopManagerApp._apply_data_mode_store_visibility,
        "_visible_product_ids": ShopManagerApp._visible_product_ids,
        "_empty_store_visible": lambda self, store_id: store_id == 30,
        "_update_sticky_store_header": lambda self: setattr(self, "sticky_updated", True),
    })()
    visible_host.main_view_mode = "data"
    visible_host.product_store_map = {1: 10, 2: 20}
    visible_host.visible_product_ids = {1}
    visible_host.bubble_product_widgets = {}
    visible_host.data_mode_container = QWidget()
    visible_host.data_mode_sticky_header = QLabel()
    visible_host._sticky_store_widget = QLabel()
    visible_host._sticky_store_cache_key = (10,)
    visible_host.data_mode_layout = QVBoxLayout(visible_host.data_mode_container)
    first_section = QWidget()
    first_layout = QVBoxLayout(first_section)
    first_product = ProductWidget.__new__(ProductWidget)
    QWidget.__init__(first_product)
    first_layout.addWidget(first_product)
    second_section = QWidget()
    second_layout = QVBoxLayout(second_section)
    second_product = ProductWidget.__new__(ProductWidget)
    QWidget.__init__(second_product)
    second_layout.addWidget(second_product)
    empty_section = QWidget()
    QVBoxLayout(empty_section)
    visible_host.data_mode_layout.addWidget(first_section)
    visible_host.data_mode_layout.addWidget(second_section)
    visible_host.data_mode_layout.addWidget(empty_section)
    visible_host.bubble_product_widgets = {1: first_product, 2: second_product}
    visible_host._data_mode_store_sections = {10: first_section, 20: second_section, 30: empty_section}
    assert visible_host.apply()
    assert not first_section.isHidden() and second_section.isHidden()
    assert not empty_section.isHidden()
    assert not first_product.isHidden() and second_product.isHidden()
    app.processEvents()
    assert visible_host.sticky_updated

    class FakeBubble(QLabel):
        def __init__(self, prod_id, prod_code, title, image_data, main_app, display_mode="bubble"):
            super().__init__(str(prod_code))
            self.prod_id = prod_id
            self.prod_code = prod_code
            self.prod_title = title
            self.display_mode = display_mode

    original_product_widget = shop_manager_module.ProductWidget
    shop_manager_module.ProductWidget = FakeBubble
    try:
        add_section = QWidget()
        add_section_layout = QVBoxLayout(add_section)
        add_section_layout.addWidget(QLabel("store"))
        add_flow_widget = QWidget()
        add_flow_layout = OrderedFlowLayout(add_flow_widget, spacing=10)
        add_section_layout.addWidget(add_flow_widget)
        add_host = type("AddHost", (), {
            "add": ShopManagerApp.refresh_after_product_added,
            "_refresh_data_mode_view_if_active": lambda self: setattr(self, "rebuilt", True),
            "load_data_safe": lambda self: setattr(self, "loaded", True),
            "_prepare_product_card_caches": lambda self, store_id=None: setattr(self, "prepared_store", store_id),
            "refresh_store_sheets": lambda self: setattr(self, "sheets_refreshed", True),
            "refresh_store_cards": lambda self, store_id=None: setattr(self, "cards_refreshed", store_id),
            "_reorder_data_mode_bubbles": lambda self: setattr(self, "reordered", True),
            "apply_tag_filter": lambda self, **_kwargs: setattr(self, "filter_applied", True),
        })()
        add_host.main_view_mode = "data"
        add_host.db = type("DB", (), {
            "safe_fetchall": lambda self, *_args: [(9, "p9", "title9", None)]
        })()
        add_host._data_mode_store_sections = {10: add_section}
        add_host.product_store_map = {}
        add_host.row_data_map = {}
        add_host.row_store_map = {0: 10}
        add_host.visible_product_ids = set()
        add_host.bubble_product_widgets = {}
        add_host.add(9, 10)
        assert 9 in add_host.bubble_product_widgets
        assert add_host.row_data_map == {1: 9}
        assert add_host.product_store_map[9] == 10
        assert add_host.visible_product_ids == {9}
        assert add_flow_layout.count() == 1
        assert add_host.cards_refreshed == 10 and add_host.filter_applied
        assert not hasattr(add_host, "loaded") and not hasattr(add_host, "rebuilt")
    finally:
        shop_manager_module.ProductWidget = original_product_widget

    delete_section = QWidget()
    delete_section_layout = QVBoxLayout(delete_section)
    delete_section_layout.addWidget(QLabel("store"))
    delete_flow_widget = QWidget()
    delete_flow_layout = OrderedFlowLayout(delete_flow_widget, spacing=10)
    delete_bubble = QLabel("delete")
    delete_flow_layout.addWidget(delete_bubble)
    delete_section_layout.addWidget(delete_flow_widget)
    delete_host = type("DeleteHost", (), {
        "delete": ShopManagerApp.refresh_after_product_deleted,
        "load_data_safe": lambda self: setattr(self, "loaded", True),
        "_prepare_widget_tree_for_delete": ShopManagerApp._prepare_widget_tree_for_delete,
        "_retire_widget": lambda self, widget: widget.deleteLater(),
        "refresh_store_sheets": lambda self: setattr(self, "sheets_refreshed", True),
        "refresh_store_cards": lambda self, store_id=None: setattr(self, "cards_refreshed", store_id),
        "_apply_data_mode_store_visibility": lambda self: setattr(self, "visibility_applied", True) or True,
    })()
    delete_host.main_view_mode = "data"
    delete_host.bubble_product_widgets = {9: delete_bubble}
    delete_host.product_store_map = {9: 10}
    delete_host.visible_product_ids = {9}
    delete_host.row_data_map = {0: 9}
    delete_host._product_card_data_cache = {9: {}}
    delete_host._product_margin_metrics_cache = {9: {}}
    delete_host.delete(9, 10)
    assert delete_flow_layout.count() == 0
    assert delete_host.product_store_map == {}
    assert delete_host.visible_product_ids == set()
    assert delete_host.row_data_map == {}
    assert delete_host.cards_refreshed == 10 and delete_host.visibility_applied
    assert not hasattr(delete_host, "loaded")

    store_section = QWidget()
    store_section_layout = QVBoxLayout(store_section)
    store_bubble = QLabel("product")
    store_section_layout.addWidget(store_bubble)
    store_container = QWidget()
    store_layout = QVBoxLayout(store_container)
    store_layout.addWidget(store_section)
    store_state = type("State", (), {})()
    store_state.rows = [
        {"row": 0, "store_id": 10, "product_id": None},
        {"row": 1, "store_id": 10, "product_id": 9},
    ]
    store_state.row_by_index = {row["row"]: row for row in store_state.rows}
    store_state.product_row_by_id = {9: 1}
    store_state.store_row_by_id = {10: 0}
    store_delete_host = type("StoreDeleteHost", (), {
        "refresh": ShopManagerApp.refresh_after_store_deleted,
        "_refresh_after_store_deleted": ShopManagerApp._refresh_after_store_deleted,
        "_prepare_widget_tree_for_delete": lambda self, widget: widget.hide(),
        "_retire_widget": lambda self, widget: widget.hide(),
        "refresh_store_sheets": lambda self: setattr(self, "sheets_refreshed", True),
        "update_store_sheet_selection": lambda self: None,
        "_apply_data_mode_store_visibility": lambda self: True,
        "_visible_product_row_count": lambda self: len(self.visible_product_ids),
        "update_daily_task_button_badge": lambda self: None,
        "show_toast": lambda self, text: setattr(self, "error_toast", text),
    })()
    store_delete_host.main_view_mode = "data"
    store_delete_host.data_mode_layout = store_layout
    store_delete_host._data_mode_store_sections = {10: store_section}
    store_delete_host.bubble_product_widgets = {9: store_bubble}
    store_delete_host.product_store_map = {9: 10}
    store_delete_host.visible_product_ids = {9}
    store_delete_host.row_data_map = {1: 9}
    store_delete_host.row_store_map = {0: 10}
    store_delete_host.main_table_state = store_state
    store_delete_host.current_store_filter = {10}
    store_delete_host.btn_store_filter = TextButton()
    store_delete_host.db = type("DB", (), {"set_setting": lambda self, *_args: None})()
    store_delete_host.refresh(10, [9])
    assert store_layout.count() == 0
    assert not store_delete_host.bubble_product_widgets
    assert not store_delete_host.product_store_map and not store_delete_host.current_store_filter
    assert not store_delete_host.main_table_state.rows and store_delete_host.sheets_refreshed
    assert store_delete_host.btn_store_filter.text == "🏪 店铺"

    margin_dialog = type("MarginDialog", (), {"promotion_data_dialog": object()})()
    record_dialog = type("RecordDialog", (), {"store_id": 10})()
    other_record_dialog = type("RecordDialog", (), {"store_id": 20})()
    pdd_dialog = object()
    promotion_dialog = object()
    product_dialog = type("ProductDialog", (), {"product_id": 9})()
    disposed = []
    close_host = type("CloseHost", (), {
        "close": ShopManagerApp._close_store_scoped_windows,
        "_dispose_account_window": lambda self, dialog: disposed.append(dialog) if dialog is not None else None,
    })()
    close_host.store_margin_dialogs = {10: margin_dialog}
    close_host.promotion_data_dialogs = {10: promotion_dialog}
    close_host.record_dialogs = [record_dialog, other_record_dialog]
    close_host.pdd_code_fetch_dialogs = {10: pdd_dialog}
    close_host.pdd_price_fetch_dialogs = {}
    close_host.pdd_promotion_status_fetch_dialogs = {}
    close_host.product_spec_dialog = product_dialog
    close_host.close(10, [9])
    assert margin_dialog in disposed and margin_dialog.promotion_data_dialog in disposed
    assert record_dialog in disposed and pdd_dialog in disposed and product_dialog in disposed
    assert promotion_dialog in disposed
    assert close_host.record_dialogs == [other_record_dialog]
    assert close_host.store_margin_dialogs == {} and close_host.promotion_data_dialogs == {}
    assert close_host.pdd_code_fetch_dialogs == {}
    assert close_host.product_spec_dialog is None

    rename_state = type("State", (), {})()
    rename_state.store_row_by_id = {10: 0}
    rename_state.row_by_index = {0: {"store_name": "旧名", "title": "旧名"}}
    rename_margin = QLabel()
    rename_margin.store_name = "旧名"
    rename_record = QLabel()
    rename_record.store_id = 10
    rename_record.store_name = "旧名"
    rename_record._update_window_title = lambda: setattr(rename_record, "title_updated", True)
    rename_host = type("RenameHost", (), {
        "rename": ShopManagerApp.refresh_after_store_renamed,
        "refresh_store_sheets": lambda self: setattr(self, "sheets_refreshed", True),
        "_update_sticky_store_header": lambda self, **kwargs: None,
    })()
    rename_host._data_mode_store_sections = {}
    rename_host.row_store_map = {}
    rename_host.main_table_state = rename_state
    rename_host.store_margin_dialogs = {10: rename_margin}
    rename_host.record_dialogs = [rename_record]
    rename_host.rename(10, "新名")
    assert rename_state.row_by_index[0]["store_name"] == "新名"
    assert rename_margin.store_name == "新名" and "新名" in rename_margin.windowTitle()
    assert rename_record.store_name == "新名" and rename_record.title_updated
    assert rename_host.sheets_refreshed

    candidate_host = type("CandidateHost", (), {
        "refresh": ShopManagerApp.refresh_category_filter_candidates,
        "_clear_layout_widgets": ShopManagerApp._clear_layout_widgets,
        "_prepare_widget_tree_for_delete": ShopManagerApp._prepare_widget_tree_for_delete,
        "_retire_widget": lambda self, widget: widget.deleteLater(),
        "_get_current_category_counts": lambda self, keyword="": (_ for _ in ()).throw(AssertionError("should not query")),
    })()
    candidate_box = QWidget()
    candidate_host.category_candidate_layout = QVBoxLayout(candidate_box)
    candidate_host.category_filter_input = QLineEdit("")
    candidate_host.current_category_filter = ""
    candidate_host.refresh()
    assert candidate_host.category_candidate_layout.count() == 1

    def make_sort_section(product_ids):
        section = QWidget()
        section_layout = QVBoxLayout(section)
        section_layout.addWidget(QLabel("store"))
        flow_widget = QWidget()
        flow_layout = OrderedFlowLayout(flow_widget, spacing=10)
        for product_id in product_ids:
            bubble = QLabel(str(product_id))
            bubble.prod_id = product_id
            flow_layout.addWidget(bubble)
        section_layout.addWidget(flow_widget)
        return section, flow_layout

    section, flow_layout = make_sort_section((1, 2))
    section_two, flow_layout_two = make_sort_section((3, 4))
    data_container = QWidget()
    data_layout = QVBoxLayout(data_container)
    sort_host = type("SortHost", (), {
        "_reorder_data_mode_bubbles": ShopManagerApp._reorder_data_mode_bubbles,
        "_sort_products_for_display": lambda self, rows: list(reversed(rows)),
    })()
    sort_rows = {10: [(1,), (2,)], 20: [(3,), (4,)]}
    sort_host.db = type("DB", (), {
        "safe_fetchall": lambda self, _query, params=(): sort_rows[params[0]]
    })()
    sort_host.product_sort_mode = "order"
    sort_host._data_mode_store_sections = {10: section, 20: section_two}
    sort_host.data_mode_layout = data_layout
    sort_host.data_mode_container = data_container
    sort_host._reorder_data_mode_bubbles()
    assert [flow_layout.itemAt(i).widget().prod_id for i in range(2)] == [2, 1]
    assert [flow_layout_two.itemAt(i).widget().prod_id for i in range(2)] == [4, 3]

    sort_metrics = {
        1: {"order_count": 1, "net_margin": 5.0},
        2: {"order_count": 100, "net_margin": 90.0},
        3: {"order_count": 200, "net_margin": 10.0},
        4: {"order_count": 300, "net_margin": 20.0},
    }
    independent_sort_host = type("IndependentSortHost", (), {
        "_sort_products_for_display": ShopManagerApp._sort_products_for_display,
        "_product_metric_sort_key": ShopManagerApp._product_metric_sort_key,
        "_descending_metric_sort_key": ShopManagerApp._descending_metric_sort_key,
        "_build_product_sort_info": lambda self, product, mode=None: {
            "product": product,
            "product_id": product,
            "fallback_order": product,
            "order_count": sort_metrics[product]["order_count"],
            "net_margin": sort_metrics[product]["net_margin"],
            "net_profit": None,
            "gross_margin": None,
            "roi": None,
            "roi_multiple": None,
            "category_label": "",
        },
    })()
    independent_sort_host.db = type("TagDB", (), {
        "safe_fetchall": lambda _self, _query, params=(): [
            (2, 1),
            (3, 2),
            (4, 2),
        ] if len(params) == 4 else [],
    })()
    independent_sort_host._calculate_product_net_margin = lambda product_id: sort_metrics[product_id]["net_margin"]
    independent_sort_host.product_sort_mode = "order"
    assert independent_sort_host._sort_products_for_display([1, 2]) == [2, 1]
    assert independent_sort_host._sort_products_for_display([1, 2], "net_margin") == [2, 1]
    assert independent_sort_host._sort_products_for_display([1, 2, 3, 4]) == [1, 2, 4, 3]
    assert independent_sort_host._sort_products_for_display([1, 2, 3, 4], "net_margin") == [1, 2, 4, 3]
    assert independent_sort_host.product_sort_mode == "order"

    deleted_product = ProductWidget.__new__(ProductWidget)
    QWidget.__init__(deleted_product)
    sip.delete(deleted_product)
    assert ProductWidget.eventFilter(deleted_product, host, QEvent(QEvent.None_)) is False
    assert ProductWidget.eventFilter(QWidget(), host, QEvent(QEvent.None_)) is False

    disposable = ProductWidget.__new__(ProductWidget)
    QWidget.__init__(disposable)
    disposable._code_click_timer = QTimer(disposable)
    disposable._code_click_timer.start(1000)
    disposable_child = QLabel("child", disposable)
    disposable_child.installEventFilter(disposable)
    dispose_host = type("DisposeHost", (), {
        "dispose": ShopManagerApp._prepare_widget_tree_for_delete,
    })()
    dispose_host.dispose(disposable)
    assert disposable._disposing and not disposable._code_click_timer.isActive()

    cleanup_container = QWidget()
    cleanup_layout = QVBoxLayout(cleanup_container)
    cleanup_section = QWidget()
    cleanup_layout.addWidget(cleanup_section)
    cleanup_host = type("CleanupHost", (), {
        "_clear_layout_widgets": ShopManagerApp._clear_layout_widgets,
        "_prepare_widget_tree_for_delete": ShopManagerApp._prepare_widget_tree_for_delete,
        "_retire_widget": lambda self, widget: widget.deleteLater(),
    })()
    cleanup_host._clear_layout_widgets(cleanup_layout)
    QApplication.sendPostedEvents(None, QEvent.DeferredDelete)
    assert cleanup_layout.count() == 0 and sip.isdeleted(cleanup_section)

    class BubbleCheck:
        def __init__(self):
            self.calls = []

        def update_promo_badges(self):
            self.calls.append("promos")

        def update_margin_display(self):
            self.calls.append("margin")

        def update_link_order_count(self):
            self.calls.append("orders")

    bubble = BubbleCheck()
    refresh_host = type("RefreshHost", (), {
        "refresh": ShopManagerApp.refresh_after_product_spec_save,
        "force_refresh_product_widget": lambda self, product_id: (
            self.bubble_product_widgets[product_id].update_promo_badges(),
            self.bubble_product_widgets[product_id].update_margin_display(),
            self.bubble_product_widgets[product_id].update_link_order_count(),
        ),
        "_reorder_data_mode_bubbles": lambda self: setattr(self, "reordered", True),
        "load_data_safe": lambda self: setattr(self, "rebuilt", True),
    })()
    refresh_host.main_view_mode = "data"
    refresh_host.bubble_product_widgets = {7: bubble}
    refresh_host.product_store_map = {7: 3}
    refresh_host._data_mode_store_sections = {}
    refresh_host.refresh(7)
    assert bubble.calls == ["promos", "margin", "orders"]
    assert refresh_host.reordered and not hasattr(refresh_host, "rebuilt")

    metrics_db = type("MetricsDB", (), {
        "calls": 0,
        "calculate_product_gross_margin_metrics": lambda self, product_id: (
            setattr(self, "calls", self.calls + 1) or {"product_id": product_id}
        ),
    })()
    metrics_host = type("MetricsHost", (), {
        "get_metrics": ShopManagerApp.get_product_gross_margin_metrics,
    })()
    metrics_host.db = metrics_db
    metrics_host._product_margin_metrics_cache = {}
    assert metrics_host.get_metrics(7) == metrics_host.get_metrics(7)
    assert metrics_db.calls == 1

    refresh_cache_host = type("RefreshCacheHost", (), {
        "refresh": ShopManagerApp.refresh_after_product_spec_save,
        "load_data_safe": lambda self: setattr(self, "loaded", True),
    })()
    refresh_cache_host.main_view_mode = "operation"
    refresh_cache_host._product_margin_metrics_cache = {7: {"gross_margin_pct": 1}}
    refresh_cache_host.refresh(7)
    app.processEvents()
    assert refresh_cache_host.loaded

    with tempfile.TemporaryDirectory() as tmp:
        db = SafeDatabaseManager(str(Path(tmp) / "test.db"))
        db.safe_execute("INSERT INTO stores (id, name) VALUES (?, ?)", (1, "s"))
        db.safe_execute(
            "INSERT INTO products (id, store_id, name, title, use_manual_spec_weight, product_memo) VALUES (?, ?, ?, ?, ?, ?)",
            (1, 1, "p", "p", 1, "长期备注"),
        )
        db.safe_execute("INSERT INTO cost_library (spec_code, spec_name, cost_price) VALUES (?, ?, ?)", ("A", "A", 1))
        db.safe_execute("INSERT INTO cost_library (spec_code, spec_name, cost_price) VALUES (?, ?, ?)", ("B", "B", 9))
        db.safe_execute(
            "INSERT INTO product_specs (product_id, spec_name, spec_code, sale_price, weight_percent) VALUES (?, ?, ?, ?, ?)",
            (1, "A", "A", 10, 100),
        )
        db.safe_execute(
            "INSERT INTO product_specs (product_id, spec_name, spec_code, sale_price, weight_percent) VALUES (?, ?, ?, ?, ?)",
            (1, "B", "B", 10, 0),
        )
        db.safe_execute(
            "INSERT INTO imported_orders (store_id, product_id, spec_code, order_count, import_time) VALUES (?, ?, ?, ?, ?)",
            (1, "p", "B", 10, "2026-07-07 00:00:00"),
        )
        assert round(db.calculate_product_gross_margin_metrics(1)["gross_margin_pct"], 2) == 90.00
        assert round(db.calculate_products_gross_margin_metrics({1})[1]["gross_margin_pct"], 2) == 90.00
        export_app = type("ExportApp", (), {"db": db})()
        memo_exporter = StoreMarginExcelExporter(1, "s", export_app)
        export_context = memo_exporter._product_export_context((1, "p", "p", None, 0, ""))
        assert export_context["product_memo"] == "长期备注"
        db.safe_execute("UPDATE products SET is_violation=1 WHERE id=?", (1,))
        violation_context = memo_exporter._product_export_context((1, "p", "p", None, 0, ""))
        assert violation_context["product_memo"] == "长期备注；违规｜推广受限"
        assert memo_exporter._products_for_specs_export()[0][1] == "p"
        assert StoreMarginExcelExporter.ORDER_HEADERS[-1] == "链接备注"
        db.conn.close()
    print("data mode layout OK")


def test_sticky_store_header_selection():
    app = QApplication.instance() or QApplication([])

    class FakeScroll(QWidget):
        def __init__(self):
            super().__init__()
            self._viewport = QWidget(self)

        def viewport(self):
            return self._viewport

    class Header(QLabel):
        def __init__(self, name, top):
            super().__init__(name)
            self.top = top
            self.setFixedSize(260, 38)

        def mapTo(self, _parent, _point):
            return QPoint(10, self.top)

    first = Header("first", 10)
    second = Header("second", 200)
    sections = {}
    for store_id, header in ((1, first), (2, second)):
        section = QWidget()
        QVBoxLayout(section).addWidget(header)
        section.show()
        sections[store_id] = section

    host = type("StickyHost", (), {
        "update_sticky": ShopManagerApp._update_sticky_store_header,
    })()
    host.data_mode_scroll = FakeScroll()
    host.data_mode_sticky_header = QLabel(host.data_mode_scroll.viewport())
    host._data_mode_store_sections = sections
    host._sticky_store_widget = None
    host._sticky_store_cache_key = None

    host.update_sticky()
    assert host.data_mode_sticky_header.isHidden()
    first.top, second.top = -20, 100
    host.update_sticky()
    assert not host.data_mode_sticky_header.isHidden()
    assert host._sticky_store_widget is first
    first.top, second.top = -300, -1
    host.update_sticky()
    assert host._sticky_store_widget is second
    app.processEvents()


def test_single_file_detailed_excel_layout():
    from openpyxl import Workbook
    from PIL import Image as PilImage

    styled_sheet = Workbook().active
    styled_sheet.append(["实发金额", "一段需要保持单行完整显示的较长内容"])
    StoreMarginExcelExporter._style_historical_export_sheet(
        StoreMarginExcelExporter.__new__(StoreMarginExcelExporter), styled_sheet, {}, {}, {1: 12, 2: 12}
    )
    assert all(cell.font.sz == 20 for cell in styled_sheet[1])
    assert all(cell.alignment.shrink_to_fit and not cell.alignment.wrap_text for cell in styled_sheet[1])
    assert styled_sheet.row_dimensions[1].height is None
    compare_values, _directions = StoreMarginExcelExporter._manual_compare_export_row(
        StoreMarginExcelExporter.__new__(StoreMarginExcelExporter),
        ("2026-07-08", "2026-07-14", 20, *([0] * 17)),
        ("2026-07-01", "2026-07-07", 10, *([0] * 17)),
    )
    assert compare_values[1] == "↑100.0%"

    source = BytesIO()
    PilImage.new("RGB", (1600, 900), "red").save(source, format="PNG")
    image_exporter = StoreMarginExcelExporter.__new__(StoreMarginExcelExporter)
    image_exporter.store_id = 1
    image_exporter.export_image_quality = "light"
    image_exporter.db = type("ImageDB", (), {
        "safe_fetchall": lambda _self, *_args: [
            (0, source.getvalue(), "2026-07-14"),
            (1, source.getvalue(), "2026-07-14"),
        ],
    })()
    image_sheet = Workbook().active
    image_sheet.append(["header"] * 20)
    image_exporter._append_reading_mode_images_to_historical_sheet(image_sheet)
    assert image_sheet["A3"].value == "本周附带图片"
    assert not image_sheet.merged_cells.ranges
    assert [image.anchor for image in image_sheet._images] == ["A4", "B4"]
    assert image_sheet.max_row == 4 and all(image_sheet.cell(4, column).value is None for column in (1, 2))
    copied_image_book = Workbook()
    ShopManagerApp._copy_excel_sheet_block(image_sheet, copied_image_book.active, 4)
    assert [image.anchor for image in copied_image_book.active._images] == ["D4", "E4"]
    copied_image_book.save(BytesIO())

    source_sheet = Workbook().active
    source_sheet["A1"] = "表头"
    source_sheet["A2"] = 1
    source_sheet["B2"] = "=A2+1"
    source_sheet.merge_cells("A3:B3")
    combined_sheet = Workbook().active
    assert ShopManagerApp._copy_excel_sheet_block(source_sheet, combined_sheet, 4) == 7
    assert combined_sheet["D1"].value == "表头" and combined_sheet["E2"].value == "=D2+1"
    assert "D3:E3" in combined_sheet.merged_cells

    compact_exporter = StoreMarginExcelExporter.__new__(StoreMarginExcelExporter)
    compact_exporter.export_image_quality = "light"
    compact_exporter._products_for_specs_export = lambda: [(1, "P1", "长商品标题", None, 0, "错题本")]
    compact_exporter._product_export_context = lambda _product: {
        "sys_id": 1,
        "product_id": "P1",
        "title": "长商品标题",
        "product_memo": "重点链接",
        "category_label": "错题本",
        "image_data": None,
        "specs": [
            {"spec_image_data": None, "spec_name": f"规格{i}", "spec_code": f"S{i}", "cost": i, "final_price": i + 5, "weight": 25, "order_count": i * 10, "refund_count": i % 2}
            for i in range(1, 5)
        ],
        "promotion_mode": "稳定成本推广",
        "roi_input_mode": "roi",
        "current_roi": 3.2,
        "return_rate": 5,
        "is_natural_flow": False,
    }
    compact_exporter._promotion_summary_for_export = lambda _ctx: "店铺满减"
    compact_exporter._record_briefs_for_export = lambda _sys_id: ("7/1-7/7", [])
    compact_exporter._add_export_image = lambda *_args: None
    compact_book = Workbook()
    compact_exporter._write_product_specs_export_sheet(compact_book)
    compact_sheet = compact_book["商品规格售卖情况"]
    assert [compact_sheet.cell(1, col).value for col in range(2, 13)] == ["规格图", "规格名称", "规格编码", "成本", "实际价格", "毛利率", "毛利润", "权重", "单量", "退款订单", "退款占比"]
    assert compact_sheet["A2"].value == "P1" and compact_sheet["A2"].fill.fgColor.rgb[-6:] == "FFF2CC"
    assert "A3:A4" in compact_sheet.merged_cells and compact_sheet["A3"].fill.fgColor.rgb[-6:] == "DCEBFF"
    assert compact_sheet["A3"].font.sz == 13
    assert "A5:A6" in compact_sheet.merged_cells and compact_sheet["A5"].fill.fgColor.rgb[-6:] == "F4CCCC"
    assert compact_sheet["B7"].value == "错题本" and compact_sheet["D7"].value == "综合毛利率"
    assert compact_sheet["E7"].fill.fgColor.rgb[-6:] == "D9EAF7"
    assert [compact_sheet[cell].value for cell in ("A8", "C8", "E8")] == ["推广方式", "当前投产", "综合毛利率"]
    assert compact_sheet["D8"].fill.fgColor.rgb[-6:] == "E8F7EF" and compact_sheet["F9"].fill.fgColor.rgb[-6:] == "E8F7EF"
    assert compact_sheet.max_row == 9 and compact_sheet["I9"].value is None
    promotion_ctx = {"coupon": 0, "new_customer": 0, "store_discount_text": "满20减2", "is_limited_time": True, "is_marketing": False}
    promotion_ctx["uses_store_discount"] = False
    assert compact_exporter.__class__._promotion_summary_for_export(compact_exporter, promotion_ctx) == "限时限量购"
    promotion_ctx["uses_store_discount"] = True
    assert "店铺满减：满20减2" in compact_exporter.__class__._promotion_summary_for_export(compact_exporter, promotion_ctx)
    removed_titles = {"商品ID", "商品标题", "链接备注", "规格售卖情况", "投产比分析", "指标", "数值", "说明", "可编辑"}
    assert removed_titles.isdisjoint({cell.value for row in compact_sheet.iter_rows() for cell in row})
    assert compact_sheet.freeze_panes == "B2"
    compact_book.save(BytesIO())
    shifted_book = Workbook()
    ShopManagerApp._copy_excel_sheet_block(compact_sheet, shifted_book.active, 4)
    assert shifted_book.active["H7"].value.startswith("=IF(SUM(L2:L5)")
    shifted_book.save(BytesIO())

    class OrderDB:
        def safe_fetchall(self, query, _params=()):
            if "store_weight" in query:
                return [
                    (1, "LOW", "low", None, 1, "", 20, 0, "", 0),
                    (2, "HIGH", "high", None, 2, "", 80, 0, "重点链接", 1),
                ]
            if "GROUP BY product_id" in query:
                return [("LOW", 2), ("HIGH", 20)]
            if "FROM import_history" in query:
                previous = {"weights": {"LOW": 90, "HIGH": 10}, "orders": {"LOW_S": {"count": 2}, "HIGH_S": {"count": 5}}}
                return [("{}",), (json.dumps(previous),)]
            raise AssertionError(query)

    order_exporter = StoreMarginExcelExporter.__new__(StoreMarginExcelExporter)
    order_exporter.db = OrderDB()
    order_exporter.store_id = 1
    order_exporter.main_app = object()
    order_exporter._product_avg_price_and_amount_for_export = lambda *_args: (None, None)
    order_exporter.get_product_margin = lambda _sys_id: (0, 0, 0)
    order_exporter.get_main_spec = lambda _code: (None, 0)
    order_exporter._order_refund_summary_for_export = lambda _code: ("0.00%", "无")
    order_exporter._link_profit_breakdown = lambda _code: {"available": False}
    order_exporter._set_square_image_cell = lambda *_args: None
    order_exporter._add_export_image = lambda *_args: None
    order_exporter._style_excel_sheet = lambda *_args: None
    order_exporter._export_product_image_size = lambda: 96
    order_book = Workbook()
    order_exporter._write_orders_export_sheet(order_book)
    assert [order_book["店铺商品权重"][f"B{row}"].value for row in (2, 3)] == ["HIGH", "LOW"]
    assert order_book["店铺商品权重"]["H2"].value == "↑70.00%"
    assert order_book["店铺商品权重"]["J2"].value == "↑15"
    assert order_book["店铺商品权重"]["L2"].value == "无"
    assert order_book["店铺商品权重"]["P2"].value == "--"
    assert order_book["店铺商品权重"]["Q2"].value == "重点链接；违规｜推广受限"
    assert len(StoreMarginExcelExporter.ORDER_HEADERS) == 17

    record_exporter = StoreMarginExcelExporter.__new__(StoreMarginExcelExporter)
    record_exporter._previous_week_range_for_export = lambda: (
        datetime(2026, 7, 1).date(), datetime(2026, 7, 7).date()
    )
    record_exporter.db = type("RecordDB", (), {
        "safe_fetchall": lambda _self, *_args: [
            (2026, 7, 2, json.dumps([{"time": "09:00", "text": "较早"}])),
            (2026, 7, 6, json.dumps([{"time": "18:00", "text": "较晚"}])),
        ]
    })()
    _range_text, records = record_exporter._record_briefs_for_export(1)
    assert [record["text"].endswith(label) for record, label in zip(records, ("较晚", "较早"))] == [True, True]

    phase_exporter = StoreMarginExcelExporter.__new__(StoreMarginExcelExporter)
    phase_exporter._write_historical_export_sheet = lambda _wb: None
    phase_exporter._write_orders_export_sheet = lambda _wb: None
    phase_exporter._write_product_specs_export_sheet = lambda _wb: None
    phase_exporter._batch_export_errors = []
    phase_updates = []
    assert phase_exporter.export_margin_excel_to_path(
        BytesIO(), lambda value, text: phase_updates.append((value, text)) or True
    )
    assert [value for value, _text in phase_updates] == [5, 15, 32, 42, 60, 70, 90, 96, 100]
    assert any("权重、对比数据" in text for _value, text in phase_updates)


def test_batch_export_detail_store_sequence():
    app = QApplication.instance() or QApplication([])

    class FakeDB:
        def safe_fetchall(self, query, params=()):
            if "SELECT id, name, title" in query:
                store_id = params[0]
                return [(store_id, f"P{store_id}", f"商品{store_id}", None, 0, "")]
            return [(0, 0, 0)]

    observed = {}
    main_app = type("ExportSelectorHost", (), {
        "product_sort_mode": "net_profit",
        "_prepare_product_card_caches": lambda self, _store_id: None,
        "_sort_products_for_display": lambda self, rows, mode: observed.setdefault("sort_modes", []).append(mode) or rows,
        "get_product_gross_margin_metrics": lambda self, _product_id: {},
        "_get_product_order_count": lambda self, _product_code, _store_id: 0,
        "_calculate_product_net_margin": lambda self, _product_id: None,
    })()

    def drive_dialog():
        dialog = next(widget for widget in app.topLevelWidgets() if widget.windowTitle() == "选择详细展示链接")
        observed["visible_combos"] = len([combo for combo in dialog.findChildren(QComboBox) if combo.isVisible()])
        observed["first_store"] = any(label.text() == "当前店铺：甲（1/2）" for label in dialog.findChildren(QLabel))
        button = next(button for button in dialog.findChildren(QPushButton) if button.text() == "下一个店铺")
        button.click()
        observed["second_store"] = any(label.text() == "当前店铺：乙（2/2）" for label in dialog.findChildren(QLabel))
        observed["last_button"] = button.text()
        button.click()

    QTimer.singleShot(0, drive_dialog)
    selections = StoreMarginExcelExporter.select_detail_products(
        None, FakeDB(), main_app, [(1, "甲"), (2, "乙")]
    )
    assert selections == {1: set(), 2: set()}
    assert observed == {
        "sort_modes": ["order", "order"],
        "visible_combos": 1,
        "first_store": True,
        "second_store": True,
        "last_button": "确认导出",
    }


def test_batch_export_mode_persistence():
    app = QApplication.instance() or QApplication([])

    class FakeDB:
        mode = "single_detailed"

        def safe_fetchall(self, *_args):
            return [(1, "甲")]

        def get_setting(self, _key, _default=None):
            return self.mode

        def set_setting(self, _key, value):
            self.mode = value

    host = QWidget()
    host.db = FakeDB()
    host._select_stores_for_margin_batch_export = ShopManagerApp._select_stores_for_margin_batch_export.__get__(host)
    observed = []

    def accept_with(mode):
        dialog = next(widget for widget in app.topLevelWidgets() if widget.windowTitle() == "批量导出")
        combo = next(combo for combo in dialog.findChildren(QComboBox) if combo.findData("single_detailed") >= 0)
        observed.append(combo.currentData())
        combo.setCurrentIndex(combo.findData(mode))
        next(button for button in dialog.findChildren(QPushButton) if button.text() == "开始导出").click()

    QTimer.singleShot(0, lambda: accept_with("simple"))
    assert host._select_stores_for_margin_batch_export()["export_mode"] == "simple"
    QTimer.singleShot(0, lambda: accept_with("simple"))
    assert host._select_stores_for_margin_batch_export()["export_mode"] == "simple"
    assert observed == ["single_detailed", "simple"]


if __name__ == "__main__":
    test_data_mode_layout_and_refresh()
    test_sticky_store_header_selection()
    test_single_file_detailed_excel_layout()
    test_batch_export_detail_store_sequence()
    test_batch_export_mode_persistence()
