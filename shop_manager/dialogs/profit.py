# -*- coding: utf-8 -*-
"""利润分析、计算器、历史记录对话框"""
import re
from datetime import datetime

from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QComboBox,
    QTextBrowser, QMessageBox, QApplication, QGroupBox, QGridLayout,
    QLineEdit, QScrollArea, QFrame, QTableWidget, QTableWidgetItem,
    QFileDialog, QAbstractItemView, QWidget,
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor

try:
    from openpyxl.utils import get_column_letter
except Exception:
    get_column_letter = lambda c: chr(64 + c)


class ProfitAnalysisDialog(QDialog):
    """利润分析建议对话框"""
    def __init__(self, result_data, db_manager=None, parent=None):
        super().__init__(parent)
        self.result_data = result_data
        self.db = db_manager
        self.setWindowTitle("📊 利润分析建议")
        self.resize(550, 500)
        self.api_key = ""
        self.custom_prompt = ""
        self.prompt_is_system = 0
        self.init_ui()
        self.load_settings()
        self.generate_analysis()

    def init_ui(self):
        layout = QVBoxLayout(self)

        self._debug_pad_label = QLabel("【板块:利润分析对话框\n文件:profit.py】利润分析/AI建议/推广优化")
        self._debug_pad_label.setStyleSheet("background-color: #FFA07A; color: #000; font-weight: bold; padding: 1px;")
        self._debug_pad_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        layout.addWidget(self._debug_pad_label)

        header = QLabel("💡 利润分析建议")
        header.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50; padding: 10px;")
        layout.addWidget(header)

        self.info_label = QLabel("")
        self.info_label.setStyleSheet("""
            QLabel {
                background-color: #e8f4fc;
                border: 1px solid #3498db;
                border-radius: 4px;
                padding: 8px 12px;
                font-size: 13px;
                color: #2c3e50;
            }
        """)
        layout.addWidget(self.info_label)

        template_layout = QHBoxLayout()
        template_layout.addWidget(QLabel("📝 选择模板："))
        self.template_combo = QComboBox()
        self.template_combo.setMinimumWidth(200)
        self.template_combo.setStyleSheet("""
            QComboBox {
                padding: 5px;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
            }
        """)
        self.template_combo.currentIndexChanged.connect(self.on_template_changed)
        template_layout.addWidget(self.template_combo)
        template_layout.addStretch()
        layout.addLayout(template_layout)

        toolbar_layout = QHBoxLayout()
        toolbar_layout.addWidget(QLabel("📊 分析结果："))
        toolbar_layout.addStretch()

        self.btn_copy = QPushButton("📋 复制分析结果")
        self.btn_copy.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                padding: 5px 15px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        self.btn_copy.clicked.connect(self.copy_analysis)
        toolbar_layout.addWidget(self.btn_copy)
        layout.addLayout(toolbar_layout)

        self.analysis_text = QTextBrowser()
        self.analysis_text.setReadOnly(True)
        self.analysis_text.setMinimumHeight(280)
        self.analysis_text.setOpenExternalLinks(True)
        self.analysis_text.setStyleSheet("""
            QTextBrowser {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 12px;
                font-size: 14px;
                line-height: 1.6;
            }
        """)
        layout.addWidget(self.analysis_text)

        btn_layout = QHBoxLayout()

        self.btn_generate_ai = QPushButton("🤖 生成AI分析")
        self.btn_generate_ai.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                color: white;
                padding: 8px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
        """)
        self.btn_generate_ai.clicked.connect(self.generate_ai_analysis)
        btn_layout.addWidget(self.btn_generate_ai)

        btn_layout.addStretch()

        btn_refresh = QPushButton("🔄 刷新分析")
        btn_refresh.clicked.connect(self.generate_analysis)
        btn_layout.addWidget(btn_refresh)

        btn_close = QPushButton("关闭")
        btn_close.clicked.connect(self.accept)
        btn_layout.addWidget(btn_close)

        layout.addLayout(btn_layout)

    def copy_analysis(self):
        text = self.analysis_text.toPlainText().strip()
        if not text:
            QMessageBox.warning(self, "提示", "没有可复制的内容！")
            return

        clipboard = QApplication.clipboard()
        clipboard.setText(text)

        self.btn_copy.setText("✅ 已复制")
        QMessageBox.information(self, "✅ 复制成功", "分析结果已复制到剪贴板！")
        self.btn_copy.setText("📋 复制分析结果")

    def load_settings(self):
        if self.db:
            self.api_key = self.db.get_setting("ai_api_key", "") or ""
            result = self.db.get_active_prompt()
            if result:
                self.custom_prompt, self.prompt_is_system = result
            else:
                self.custom_prompt = ""
                self.prompt_is_system = 0
        else:
            self.custom_prompt = ""
            self.prompt_is_system = 0

        self.load_templates()
        self.update_info_label()

    def load_templates(self):
        self.template_combo.blockSignals(True)
        self.template_combo.clear()

        active_id = None

        if self.db:
            prompts = self.db.get_all_prompts()
            for p in prompts:
                prompt_id, name, content, is_active, is_system = p
                self.template_combo.addItem(name, prompt_id)
                if is_active:
                    active_id = prompt_id

        current_idx = self.template_combo.findData(active_id) if active_id is not None else 0
        if current_idx >= 0:
            self.template_combo.setCurrentIndex(current_idx)

        self.template_combo.blockSignals(False)

    def on_template_changed(self, index):
        prompt_id = self.template_combo.itemData(index)
        if prompt_id and self.db:
            self.db.set_active_prompt(prompt_id)
            result = self.db.get_active_prompt()
            if result:
                self.custom_prompt, self.prompt_is_system = result

    def update_info_label(self):
        r = self.result_data
        target_type = r.get("data_type", "")
        target_id = r.get("target_id")

        info_text = ""

        if target_type == "product" and target_id and self.db:
            try:
                rows = self.db.safe_fetchall(
                    "SELECT title FROM products WHERE id = ?",
                    (target_id,)
                )
                if rows and rows[0][0]:
                    info_text = f"📦 商品：{rows[0][0]}"

                    spec_rows = self.db.safe_fetchall(
                        "SELECT spec_name FROM product_specs WHERE product_id = ? ORDER BY spec_name",
                        (target_id,)
                    )
                    if spec_rows:
                        spec_names = [s[0] for s in spec_rows if s[0]]
                        if spec_names:
                            info_text += f"\n📋 规格：{', '.join(spec_names[:3])}"
                            if len(spec_names) > 3:
                                info_text += f" 等{len(spec_names)}个"
            except:
                pass
        elif target_type == "store" and target_id and self.db:
            try:
                rows = self.db.safe_fetchall(
                    "SELECT store_name FROM stores WHERE id = ?",
                    (target_id,)
                )
                if rows and rows[0][0]:
                    info_text = f"🏪 店铺：{rows[0][0]}"
            except:
                pass

        if not info_text:
            target_name = r.get("target_name", "")
            if target_name:
                info_text = f"📌 {target_name}"
            else:
                info_text = "📌 未获取到商品/店铺信息"

        self.info_label.setText(info_text)

    def get_context_info(self):
        r = self.result_data

        target_type = r.get("data_type", "")
        target_id = r.get("target_id")
        target_name = r.get("target_name", "")

        context = f"类型：{target_type}，名称：{target_name}"

        if target_type == "product" and target_id and self.db:
            try:
                rows = self.db.safe_fetchall(
                    "SELECT title FROM products WHERE id = ?",
                    (target_id,)
                )
                if rows:
                    title = rows[0][0] or ""
                    if title:
                        context = f"商品标题：{title}"
            except:
                pass
        elif target_type == "store" and target_id and self.db:
            try:
                rows = self.db.safe_fetchall(
                    "SELECT store_name FROM stores WHERE id = ?",
                    (target_id,)
                )
                if rows and rows[0][0]:
                    context = f"店铺名称：{rows[0][0]}"

                product_rows = self.db.safe_fetchall(
                    "SELECT DISTINCT p.title FROM products p WHERE p.store_id = ? LIMIT 5",
                    (target_id,)
                )
                if product_rows:
                    titles = [row[0] for row in product_rows if row[0]]
                    if titles:
                        context += f"，关联商品：{', '.join(titles[:3])}"
            except:
                pass

        if not context or context == f"类型：{target_type}，名称：{target_name}":
            context = f"类型：{target_type}，名称：{target_name}（未获取到详细信息）"

        return context

    def get_default_prompt(self):
        return """你是一位资深拼多多电商运营专家，拥有多年类目运营经验。请根据以下完整的推广数据，给出专业、深入、可操作的分析建议。

【分析对象】
{分析对象信息}

【今日战绩】
推广费：{推广费}元
投产比：{投产比}
退货率：{退货率}%
毛利率：{毛利率}%
客单价：{客单价}元

【自动计算出的数据】
成交金额：{成交金额}元
退款金额：{退款金额}元
实际成交：{实际成交}元
产品成本：{产品成本}元
毛利润：{毛利润}元
技术服务费：{技术服务费}元
净利润：{净利润}元
净利率：{净利率}%
推广占比：{推广占比}%
成交单量：{成交单量}单
每笔成交花费：{每笔成交花费}元/单
单笔利润：{单笔利润}元/单

【保本情况】
毛保本投产：{毛保本投产}
净保本投产：{净保本投产}
净保本1.25倍：{净保本1.25倍}
最佳投产：{最佳投产}
当前投产倍数：{当前投产倍数}

请按以下格式输出，要求内容详实、数据支撑、实用可执行：

📊 【盈利状况诊断】
（分析当前是否盈利，亏损原因，盈利/亏损幅度，与行业平均对比）

⚠️ 【问题点深度剖析】
（列出2-4个核心问题，每个问题要说明原因、影响程度、改进优先级）

🎯 【实战优化方案】
（列出3-5条具体可执行的优化建议，每条要包含：具体动作+预期效果+操作难度）

📈 【市场趋势与竞争分析】
（分析该类目当前市场趋势、竞争格局、消费者偏好变化、季节性因素等，提供前瞻性建议）

💎 【核心干货总结】
（用最精炼的语言总结2-3个最关键的决策点）

要求：
1. 数据必须完全准确，每项数据都要在分析中体现
2. 建议要具体可执行，避免空洞废话
3. 分析要深入本质，给出真正的干货
4. 适当引用行业经验和数据支撑
5. 总字数不少于500字，内容要充实详细"""

    def convert_markdown_to_html(self, text):
        if not text:
            return ""

        text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

        lines = text.split('\n')
        result = []
        in_list = False
        i = 0
        length = len(lines)

        while i < length:
            line = lines[i]

            if not line.strip():
                if in_list:
                    result.append("</ul>")
                    in_list = False
                i += 1
                continue

            h_match = re.match(r'^(#{1,6})\s+(.+)$', line)
            if h_match:
                if in_list:
                    result.append("</ul>")
                    in_list = False
                level = len(h_match.group(1))
                content = h_match.group(2)
                result.append(f"<h{level}>{content}</h{level}>")
                i += 1
                continue

            list_match = re.match(r'^(\d+)\.\s+(.+)$', line)
            if list_match:
                if not in_list:
                    result.append("<ul>")
                    in_list = True
                content = list_match.group(2)
                content = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', content)
                content = re.sub(r'\*(.+?)\*', r'<i>\1</i>', content)
                content = re.sub(r'`(.+?)`', r'<code>\1</code>', content)
                result.append(f"<li>{content}</li>")
                i += 1
                continue

            bullet_match = re.match(r'^[-*]\s+(.+)$', line)
            if bullet_match:
                if not in_list:
                    result.append("<ul>")
                    in_list = True
                content = bullet_match.group(1)
                content = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', content)
                content = re.sub(r'\*(.+?)\*', r'<i>\1</i>', content)
                content = re.sub(r'`(.+?)`', r'<code>\1</code>', content)
                result.append(f"<li>{content}</li>")
                i += 1
                continue

            if in_list:
                result.append("</ul>")
                in_list = False

            line = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', line)
            line = re.sub(r'\*(.+?)\*', r'<i>\1</i>', line)
            line = re.sub(r'`(.+?)`', r'<code>\1</code>', line)
            result.append(f"<p>{line}</p>")
            i += 1

        if in_list:
            result.append("</ul>")

        html = "".join(result)

        html = re.sub(r'(<br>\s*)+', '<br>', html)

        return html

    def generate_analysis(self):
        r = self.result_data
        suggestions = []

        suggestions.append("=" * 30)
        suggestions.append("📈 核心指标分析")
        suggestions.append("=" * 30)

        if r["net_profit"] < 0:
            suggestions.append("❌ 净利润为负：当前亏损 ¥{:.2f}".format(r["net_profit"]))
            suggestions.append("   建议：优化投产比或降低推广费用")
        else:
            suggestions.append("✅ 净利润为正：盈利 ¥{:.2f}".format(r["net_profit"]))

        suggestions.append("")

        if r["roi"] < r.get("break_even_roi", 0):
            suggestions.append("⚠️ 投产比低于毛保本 ({:.2f})".format(r.get("break_even_roi", 0)))
            suggestions.append("   需要提升投产比至 {:.2f} 以上".format(r.get("break_even_roi", 0)))
        else:
            suggestions.append("✅ 投产比高于毛保本，安全边际充足")

        suggestions.append("")

        if r["return_rate"] > 10:
            suggestions.append("⚠️ 退货率较高 ({:.1f}%)".format(r["return_rate"]))
            suggestions.append("   建议：关注产品质量和客户预期管理")
        elif r["return_rate"] > 5:
            suggestions.append("⚠️ 退货率中等 ({:.1f}%)".format(r["return_rate"]))
            suggestions.append("   建议：持续监控退货原因")
        else:
            suggestions.append("✅ 退货率正常 ({:.1f}%)".format(r["return_rate"]))

        suggestions.append("")
        suggestions.append("=" * 30)
        suggestions.append("🎯 优化建议")
        suggestions.append("=" * 30)

        current_roi = r.get("best_roi_from_net", 0)
        if current_roi > 0:
            suggestions.append("💰 建议目标投产比：{:.2f}（净保本1.4倍）".format(current_roi))

        if r.get("cost_per_transaction", 0) > 0:
            suggestions.append("📊 每笔成交推广成本：¥{:.2f}".format(r["cost_per_transaction"]))

        if r.get("profit_per_transaction", 0) > 0:
            suggestions.append("💵 每笔成交净利润：¥{:.2f}".format(r["profit_per_transaction"]))

        self.analysis_text.setText("\n".join(suggestions))

    def save_record_with_analysis(self, ai_analysis):
        try:
            record_date = datetime.now().strftime("%Y-%m-%d")

            if self.db:
                self.db.safe_execute(
                    """INSERT INTO profit_records
                    (data_type, target_id, target_name, record_date, promotion_amount, roi, return_rate,
                    margin_rate, avg_price, transaction_amount, refund_amount, actual_transaction_amount,
                    product_cost, gross_profit, tech_fee, net_profit, net_profit_rate, promotion_ratio,
                    break_even_roi, transaction_count, cost_per_transaction, profit_per_transaction,
                    best_roi, net_break_even_roi, net_break_even_125, net_break_even_value,
                    net_break_even_125_from_net, best_roi_from_net, current_roi_multiple, ai_analysis, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        self.result_data.get("data_type", "product"),
                        self.result_data.get("target_id"),
                        self.result_data.get("target_name", ""),
                        record_date,
                        self.result_data.get("promotion_amount", 0),
                        self.result_data.get("roi", 0),
                        self.result_data.get("return_rate", 0),
                        self.result_data.get("margin_rate", 0),
                        self.result_data.get("avg_price", 0),
                        self.result_data.get("transaction_amount", 0),
                        self.result_data.get("refund_amount", 0),
                        self.result_data.get("actual_transaction_amount", 0),
                        self.result_data.get("product_cost", 0),
                        self.result_data.get("gross_profit", 0),
                        self.result_data.get("tech_fee", 0),
                        self.result_data.get("net_profit", 0),
                        self.result_data.get("net_profit_rate", 0),
                        self.result_data.get("promotion_ratio", 0),
                        self.result_data.get("break_even_roi", 0),
                        self.result_data.get("transaction_count", 0),
                        self.result_data.get("cost_per_transaction", 0),
                        self.result_data.get("profit_per_transaction", 0),
                        self.result_data.get("best_roi", 0),
                        self.result_data.get("net_break_even_roi", 0),
                        self.result_data.get("net_break_even_125", 0),
                        self.result_data.get("net_break_even_value", 0),
                        self.result_data.get("net_break_even_125_from_net", 0),
                        self.result_data.get("best_roi_from_net", 0),
                        self.result_data.get("current_roi_multiple", 0),
                        ai_analysis,
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    )
                )
                print(f"[DEBUG] AI分析已自动保存到 {record_date}")
        except Exception as e:
            print(f"[DEBUG] 自动保存失败: {e}")

    def generate_ai_analysis(self):
        if not self.api_key:
            QMessageBox.warning(self, "⚠️ 提示", "请先在主界面点击「🔑 API配置」按钮导入API Key！")
            return

        r = self.result_data

        context_info = self.get_context_info()

        if self.custom_prompt:
            prompt = self.custom_prompt
            prompt_is_system = getattr(self, 'prompt_is_system', 0)
        else:
            prompt = self.get_default_prompt()
            prompt_is_system = 1

        if prompt_is_system:
            prompt = "【系统】" + prompt
        else:
            prompt = "【用户】" + prompt

        prompt = prompt.replace("{净保本1.25倍}", "{净保本1_25倍}")

        common_prompts = self.db.get_active_common_prompts() if self.db else []

        try:
            prompt = prompt.format(
                分析对象信息=context_info,
                推广费=r.get("promotion_amount", 0),
                投产比=r.get("roi", 0),
                退货率=r.get("return_rate", 0),
                毛利率=r.get("margin_rate", 0),
                客单价=r.get("avg_price", 0),
                成交金额=r.get("transaction_amount", 0),
                退款金额=r.get("refund_amount", 0),
                实际成交=r.get("actual_transaction_amount", 0),
                产品成本=r.get("product_cost", 0),
                毛利润=r.get("gross_profit", 0),
                技术服务费=r.get("tech_fee", 0),
                净利润=r.get("net_profit", 0),
                净利率=r.get("net_profit_rate", 0),
                推广占比=r.get("promotion_ratio", 0),
                成交单量=r.get("transaction_count", 0),
                每笔成交花费=r.get("cost_per_transaction", 0),
                单笔利润=r.get("profit_per_transaction", 0),
                毛保本投产=r.get("break_even_roi", 0),
                净保本投产=r.get("net_break_even_roi", 0),
                净保本1_25倍=r.get("net_break_even_roi", 0) * 1.25,
                最佳投产=r.get("net_break_even_roi", 0) * 1.4,
                当前投产倍数=(r.get("roi", 0) / r.get("net_break_even_roi", 1) if r.get("net_break_even_roi", 0) > 0 else 0),
            )

            if common_prompts:
                common_section = "\n\n【用户-重要运营常识提醒】\n" + "\n".join([f"- {p}" for p in common_prompts])
                prompt = prompt + common_section

            saved_titles = self.db.get_setting("selected_knowledge_titles", "")
            if saved_titles:
                title_list = [t.strip() for t in saved_titles.split(",") if t.strip()]
                if title_list:
                    knowledge_items = self.db.get_knowledge_items_by_titles(title_list)
                    if knowledge_items:
                        knowledge_section = "\n\n【本地知识库参考】\n"
                        for item in knowledge_items:
                            knowledge_section += f"\n【{item['title']}】\n{item['content']}\n"
                        prompt = prompt + knowledge_section
            else:
                use_rag = self.db.get_setting("use_rag_retrieval", "1") == "1"
                if use_rag:
                    rag_results = self.db.rag_retrieve(prompt[:500], top_k=3)
                    if rag_results:
                        knowledge_section = "\n\n【本地知识库参考（RAG检索）】\n"
                        for item in rag_results:
                            knowledge_section += f"\n【{item['title']}】(相似度:{item['similarity']:.2f})\n{item['content']}\n"
                        prompt = prompt + knowledge_section

            store_memo = ""
            if self.result_data.get("target_id"):
                try:
                    if self.result_data.get("data_type") == "store":
                        store_id = self.result_data.get("target_id")
                    else:
                        store_rows = self.db.safe_fetchall("SELECT store_id FROM products WHERE id=?", (self.result_data.get("target_id"),))
                        store_id = store_rows[0][0] if store_rows and store_rows[0] else None

                    if store_id:
                        memo_rows = self.db.safe_fetchall("SELECT memo FROM stores WHERE id=?", (store_id,))
                        store_memo = memo_rows[0][0] if memo_rows and memo_rows[0][0] else ""
                except Exception as e:
                    print(f"获取店铺备注失败: {e}")

            if store_memo:
                priority_section = f"\n\n【店铺运营指导大纲 - 最高优先级】\n{store_memo}\n\n请严格遵循上述店铺运营指导大纲进行所有分析和建议。"
                prompt = prompt + priority_section

        except Exception as e:
            QMessageBox.warning(self, "提示", f"提示词格式错误: {str(e)}")
            return

        self.analysis_text.setText("🤖 正在调用AI分析，请稍候...")
        self.btn_generate_ai.setEnabled(False)

        try:
            import requests

            api_key = self.api_key.strip()

            if not api_key:
                self.analysis_text.setText("❌ 请先配置API Key！")
                self.btn_generate_ai.setEnabled(True)
                return

            is_deepseek = (
                "sk-" in api_key and len(api_key) > 40
            ) or (
                api_key.startswith("deepseek-")
            )

            if is_deepseek or "sk-" in api_key:
                headers = {
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json"
                }
                api_url = "https://api.deepseek.com/v1/chat/completions"
                model = "deepseek-chat"
            else:
                headers = {
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json"
                }
                api_url = "https://api.openai.com/v1/chat/completions"
                model = "gpt-3.5-turbo"

            data = {
                "model": model,
                "messages": [{"role": "user", "content": prompt}],
                "max_tokens": 2000,
                "temperature": 0.7
            }

            response = requests.post(
                api_url,
                headers=headers,
                json=data,
                timeout=120
            )

            if response.status_code == 200:
                result = response.json()
                ai_response = result["choices"][0]["message"]["content"]

                html = self.convert_markdown_to_html(ai_response)
                self.analysis_text.setHtml(html)

                self.result_data["ai_analysis"] = ai_response

                self.save_record_with_analysis(ai_response)
            elif response.status_code == 401:
                self.analysis_text.setText("❌ API Key无效，请检查是否正确")
            elif response.status_code == 403:
                self.analysis_text.setText("❌ 访问被拒绝，请检查API Key权限")
            elif response.status_code == 429:
                self.analysis_text.setText("❌ 请求过于频繁，请稍后再试")
            else:
                self.analysis_text.setText(f"❌ API调用失败: {response.status_code}\n{response.text[:500]}")

        except requests.exceptions.Timeout:
            self.analysis_text.setText("❌ 超时：国内访问DeepSeek可能需要代理")
        except requests.exceptions.ConnectionError as ce:
            error_msg = f"连接失败: {str(ce)}"
            print(f"API Connection Error: {error_msg}")
            self.analysis_text.setText(f"❌ 连接失败，请检查API配置\n{error_msg}")
        except Exception as e:
            error_msg = f"调用失败: {str(e)}"
            print(f"API Error: {error_msg}")
            self.analysis_text.setText(f"❌ {error_msg}")

        finally:
            self.btn_generate_ai.setEnabled(True)


class ProfitCalculatorDialog(QDialog):
    """利润计算器对话框"""
    def __init__(self, margin_rate, avg_price, target_id=None, target_name="", data_type="product", parent=None, db_manager=None):
        super().__init__(parent)
        self.margin_rate = margin_rate
        self.avg_price = avg_price
        self.target_id = target_id
        self.target_name = target_name
        self.data_type = data_type
        self.last_result = None
        self.db = db_manager
        self.setWindowTitle("🧮 利润计算器")
        self.resize(700, 650)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)

        self._debug_pc_label = QLabel("【板块:利润计算器对话框\n文件:profit.py】推广费/投产比/退货率/毛利率/计算结果")
        self._debug_pc_label.setStyleSheet("background-color: #FFA07A; color: #000; font-weight: bold; padding: 1px;")
        self._debug_pc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        layout.addWidget(self._debug_pc_label)

        header = QLabel("💰 利润计算器")
        header.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50; padding: 10px;")
        layout.addWidget(header)

        input_group = QGroupBox("📝 输入参数")
        input_layout = QGridLayout()

        input_layout.addWidget(QLabel("推广费 (¥):"), 0, 0)
        self.promotion_input = QLineEdit()
        self.promotion_input.setPlaceholderText("请输入推广费...")
        self.promotion_input.setFixedWidth(150)
        input_layout.addWidget(self.promotion_input, 0, 1)

        input_layout.addWidget(QLabel("投产比 (ROI):"), 1, 0)
        self.roi_input = QLineEdit()
        self.roi_input.setPlaceholderText("请输入投产比...")
        self.roi_input.setFixedWidth(150)
        input_layout.addWidget(self.roi_input, 1, 1)

        input_layout.addWidget(QLabel("退货率 (%):"), 2, 0)
        self.return_rate_input = QLineEdit()
        self.return_rate_input.setPlaceholderText("请输入退货率...")
        self.return_rate_input.setFixedWidth(150)
        input_layout.addWidget(self.return_rate_input, 2, 1)

        input_layout.addWidget(QLabel("毛利率 (%):"), 3, 0)
        self.margin_rate_input = QLineEdit()
        self.margin_rate_input.setText(f"{self.margin_rate:.2f}")
        self.margin_rate_input.setReadOnly(True)
        self.margin_rate_input.setStyleSheet("background-color: #e8f5e9; color: #2e7d32;")
        self.margin_rate_input.setFixedWidth(150)
        input_layout.addWidget(self.margin_rate_input, 3, 1)

        input_layout.addWidget(QLabel("客单价 (¥):"), 4, 0)
        self.avg_price_input = QLineEdit()
        self.avg_price_input.setText(f"{self.avg_price:.2f}")
        self.avg_price_input.setReadOnly(True)
        self.avg_price_input.setStyleSheet("background-color: #e8f5e9; color: #2e7d32;")
        self.avg_price_input.setFixedWidth(150)
        input_layout.addWidget(self.avg_price_input, 4, 1)

        input_group.setLayout(input_layout)
        layout.addWidget(input_group)

        save_group = QGroupBox("💾 保存数据")
        save_layout = QHBoxLayout()

        save_layout.addWidget(QLabel("日期:"))
        self.date_combo = QComboBox()
        self.date_combo.addItems(["当天", "昨天"])
        self.date_combo.setFixedWidth(100)
        save_layout.addWidget(self.date_combo)

        self.btn_save = QPushButton("📥 保存记录")
        self.btn_save.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                font-weight: bold;
                padding: 5px 15px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        self.btn_save.clicked.connect(self.save_record)
        self.btn_save.setEnabled(False)
        save_layout.addWidget(self.btn_save)

        self.btn_view_history = QPushButton("📊 查看历史")
        self.btn_view_history.setStyleSheet("""
            QPushButton {
                background-color: #e67e22;
                color: white;
                font-weight: bold;
                padding: 5px 15px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #d35400;
            }
        """)
        self.btn_view_history.clicked.connect(self.view_history)
        save_layout.addWidget(self.btn_view_history)

        save_layout.addStretch()
        save_group.setLayout(save_layout)
        layout.addWidget(save_group)

        btn_calc_layout = QHBoxLayout()
        self.btn_calculate = QPushButton("🔢 计算利润")
        self.btn_calculate.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                font-weight: bold;
                padding: 10px 30px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        self.btn_calculate.clicked.connect(self.calculate)
        btn_calc_layout.addStretch()
        btn_calc_layout.addWidget(self.btn_calculate)
        btn_calc_layout.addStretch()
        layout.addLayout(btn_calc_layout)

        self.result_scroll = QScrollArea()
        self.result_scroll.setWidgetResizable(True)
        self.result_scroll.setMinimumHeight(280)

        self.result_container = QWidget()
        self.result_layout = QGridLayout(self.result_container)
        self.result_layout.setSpacing(8)
        self.result_layout.setContentsMargins(5, 5, 5, 5)

        self.result_scroll.setWidget(self.result_container)
        layout.addWidget(self.result_scroll)

        self.result_container.setStyleSheet("""
            QWidget {
                background-color: #fafafa;
            }
        """)

        btn_analysis_layout = QHBoxLayout()
        btn_analysis_layout.addStretch()

        self.btn_analysis = QPushButton("📊 分析建议")
        self.btn_analysis.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                padding: 5px 15px;
                border-radius: 3px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #138496;
            }
        """)
        self.btn_analysis.clicked.connect(self.show_analysis_dialog)
        btn_analysis_layout.addWidget(self.btn_analysis)

        btn_close_layout = QHBoxLayout()
        btn_close_layout.addStretch()
        btn_close = QPushButton("关闭")
        btn_close.clicked.connect(self.accept)
        btn_close_layout.addWidget(btn_close)

        layout.addLayout(btn_analysis_layout)
        layout.addLayout(btn_close_layout)

    def show_analysis_dialog(self):
        if not self.last_result:
            QMessageBox.warning(self, "提示", "请先计算利润后再查看分析！")
            return

        db = None
        if hasattr(self, 'db'):
            db = self.db
        dialog = ProfitAnalysisDialog(self.last_result, db, self)
        dialog.show()

    def calculate(self):
        try:
            promotion = float(self.promotion_input.text()) if self.promotion_input.text() else 0
            roi = float(self.roi_input.text()) if self.roi_input.text() else 0
            return_rate = float(self.return_rate_input.text()) if self.return_rate_input.text() else 0
            margin_rate = self.margin_rate / 100
            avg_price = self.avg_price
        except ValueError:
            QMessageBox.warning(self, "输入错误", "请输入有效的数字！")
            return

        if roi <= 0:
            QMessageBox.warning(self, "输入错误", "投产比必须大于0！")
            return

        transaction_amount = promotion * roi
        refund_amount = transaction_amount * (return_rate / 100)
        actual_transaction_amount = transaction_amount - refund_amount
        product_cost = transaction_amount * (1 - margin_rate)
        gross_profit = transaction_amount - product_cost
        tech_fee = transaction_amount * 0.006
        net_profit = actual_transaction_amount - product_cost - promotion - tech_fee
        net_profit_rate = (net_profit / transaction_amount * 100) if transaction_amount > 0 else 0
        promotion_ratio = (promotion / transaction_amount * 100) if transaction_amount > 0 else 0
        break_even_roi = 1 / margin_rate if margin_rate > 0 else 0
        transaction_count = (transaction_amount / avg_price) if avg_price > 0 else 0
        cost_per_transaction = (promotion / transaction_count) if transaction_count > 0 else 0

        profit_per_transaction = net_profit / transaction_count if transaction_count > 0 else 0
        best_roi = break_even_roi * 1.4 if break_even_roi > 0 else 0

        net_margin_formula = margin_rate * (1 - return_rate / 100) - 0.0006
        net_break_even_roi = 1 / net_margin_formula if net_margin_formula > 0 else 0

        net_break_even_value = transaction_amount / net_break_even_roi if net_break_even_roi > 0 else 0
        net_break_even_125 = net_break_even_value * 1.25 if net_break_even_value > 0 else 0
        net_break_even_125_from_net = net_break_even_roi * 1.25 if net_break_even_roi > 0 else 0
        best_roi_from_net = net_break_even_roi * 1.4 if net_break_even_roi > 0 else 0
        current_roi_multiple = roi / break_even_roi if break_even_roi > 0 else 0

        self.last_result = {
            "data_type": self.data_type,
            "target_id": self.target_id,
            "target_name": self.target_name,
            "promotion_amount": promotion,
            "roi": roi,
            "return_rate": return_rate,
            "margin_rate": self.margin_rate,
            "avg_price": avg_price,
            "transaction_amount": transaction_amount,
            "refund_amount": refund_amount,
            "actual_transaction_amount": actual_transaction_amount,
            "product_cost": product_cost,
            "gross_profit": gross_profit,
            "tech_fee": tech_fee,
            "net_profit": net_profit,
            "net_profit_rate": net_profit_rate,
            "promotion_ratio": promotion_ratio,
            "break_even_roi": break_even_roi,
            "transaction_count": transaction_count,
            "cost_per_transaction": cost_per_transaction,
            "profit_per_transaction": profit_per_transaction,
            "best_roi": best_roi,
            "net_break_even_roi": net_break_even_roi,
            "net_break_even_125": net_break_even_125,
            "net_break_even_value": net_break_even_value,
            "net_break_even_125_from_net": net_break_even_125_from_net,
            "best_roi_from_net": best_roi_from_net,
            "current_roi_multiple": current_roi_multiple,
        }

        results = [
            ("成交金额", f"¥{transaction_amount:,.2f}"),
            ("退款金额", f"¥{refund_amount:,.2f}"),
            ("实际成交", f"¥{actual_transaction_amount:,.2f}"),
            ("产品成本", f"¥{product_cost:,.2f}"),
            ("毛利润", f"¥{gross_profit:,.2f}"),
            ("技术服务费", f"¥{tech_fee:,.2f}"),
            ("净利润", f"¥{net_profit:,.2f}"),
            ("净利率", f"{net_profit_rate:.2f}%"),
            ("推广占比", f"{promotion_ratio:.2f}%"),
            ("成交单量", f"{transaction_count:.2f} 笔"),
            ("每笔成交花费", f"¥{cost_per_transaction:.2f}"),
            ("单笔利润", f"¥{profit_per_transaction:.2f}"),
            ("毛保本投产", f"{break_even_roi:.2f}"),
            ("净保本投产", f"{net_break_even_roi:.2f}"),
            ("净保本1.25倍", f"{net_break_even_roi * 1.25:.2f}"),
            ("最佳投产", f"{net_break_even_roi * 1.4:.2f}"),
            ("当前投产倍数", f"{roi / net_break_even_roi:.2f}倍" if net_break_even_roi > 0 else "0.00倍"),
        ]

        self.display_results_grid(results)

        self.btn_save.setEnabled(True)

    def display_results_grid(self, results):
        while self.result_layout.count():
            item = self.result_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        def calculate_cols():
            width = self.result_scroll.viewport().width() - 30
            if width <= 0:
                return 3
            unit_width = 160
            cols = max(1, width // unit_width)
            return min(cols, 6)

        cols = calculate_cols()

        for i, (label, value) in enumerate(results):
            row = i // cols
            col = i % cols

            card = QFrame()
            card.setFrameStyle(QFrame.StyledPanel | QFrame.Raised)
            card.setFixedWidth(150)
            card.setStyleSheet("""
                QFrame {
                    background-color: white;
                    border-radius: 5px;
                    border: 1px solid #e0e0e0;
                    padding: 4px;
                }
            """)

            card_layout = QVBoxLayout(card)
            card_layout.setContentsMargins(6, 4, 6, 4)
            card_layout.setSpacing(2)
            card_layout.setAlignment(Qt.AlignCenter)

            label_widget = QLabel(label)
            label_widget.setAlignment(Qt.AlignCenter)
            label_widget.setStyleSheet("""
                QLabel {
                    font-size: 12px;
                    font-weight: bold;
                    color: #555555;
                    padding: 0px;
                }
            """)

            value_widget = QLabel(value)
            value_widget.setAlignment(Qt.AlignCenter)
            value_widget.setStyleSheet("""
                QLabel {
                    font-size: 14px;
                    font-weight: bold;
                    padding: 2px;
                }
            """)

            if "净利润" in label or "单笔利润" in label:
                try:
                    val_num = float(value.replace("¥", "").replace(",", "").replace(" 笔", "").replace("倍", ""))
                    if val_num >= 0:
                        value_widget.setStyleSheet("""
                            QLabel {
                                font-size: 14px;
                                font-weight: bold;
                                color: #27ae60;
                                padding: 2px;
                            }
                        """)
                    else:
                        value_widget.setStyleSheet("""
                            QLabel {
                                font-size: 14px;
                                font-weight: bold;
                                color: #e74c3c;
                                padding: 2px;
                            }
                        """)
                except:
                    pass

            if "最佳" in label or "倍数" in label:
                value_widget.setStyleSheet("""
                    QLabel {
                        font-size: 14px;
                        font-weight: bold;
                        color: #9b59b6;
                        padding: 2px;
                    }
                """)

            card_layout.addWidget(label_widget)
            card_layout.addWidget(value_widget)

            self.result_layout.addWidget(card, row, col)

        for i in range(cols):
            self.result_layout.setColumnStretch(i, 1)

        self.result_scroll.resizeEvent = lambda event: self.on_resize_event()
        self._last_cols = cols

    def on_resize_event(self):
        if not hasattr(self, '_last_cols') or not self.last_result:
            return

        current_width = self.result_scroll.viewport().width()
        new_cols = max(1, min(6, (current_width - 30) // 160))

        if new_cols != self._last_cols and hasattr(self, 'last_result') and self.last_result:
            results = []
            r = self.last_result
            results.append(("成交金额", f"¥{r['transaction_amount']:,.2f}"))
            results.append(("退款金额", f"¥{r['refund_amount']:,.2f}"))
            results.append(("实际成交", f"¥{r['actual_transaction_amount']:,.2f}"))
            results.append(("产品成本", f"¥{r['product_cost']:,.2f}"))
            results.append(("毛利润", f"¥{r['gross_profit']:,.2f}"))
            results.append(("技术服务费", f"¥{r['tech_fee']:,.2f}"))
            results.append(("净利润", f"¥{r['net_profit']:,.2f}"))
            results.append(("净利率", f"{r['net_profit_rate']:.2f}%"))
            results.append(("推广占比", f"{r['promotion_ratio']:.2f}%"))
            results.append(("成交单量", f"{r['transaction_count']:.2f} 笔"))
            results.append(("每笔成交花费", f"¥{r['cost_per_transaction']:,.2f}"))
            results.append(("单笔利润", f"¥{r['profit_per_transaction']:.2f}"))
            results.append(("毛保本投产", f"{r['break_even_roi']:.2f}"))
            results.append(("净保本投产", f"{r['net_break_even_roi']:.2f}"))
            results.append(("净保本1.25倍", f"{r['net_break_even_roi'] * 1.25:.2f}"))
            results.append(("最佳投产", f"{r['net_break_even_roi'] * 1.4:.2f}"))
            results.append(("当前投产倍数", f"{r['roi'] / r['net_break_even_roi']:.2f}倍" if r['net_break_even_roi'] > 0 else "0.00倍"))

            self.display_results_grid(results)

    def get_record_date(self):
        from datetime import timedelta
        today = datetime.now()
        if self.date_combo.currentText() == "昨天":
            yesterday = today - timedelta(days=1)
            return yesterday.strftime("%Y-%m-%d")
        return today.strftime("%Y-%m-%d")

    def save_record(self):
        if not self.last_result:
            QMessageBox.warning(self, "提示", "请先计算利润后再保存！")
            return

        db = self.get_database()
        if not db:
            QMessageBox.warning(self, "错误", "无法获取数据库连接！")
            return

        record_date = self.get_record_date()

        db.safe_execute(
            """INSERT INTO profit_records
            (data_type, target_id, target_name, record_date, promotion_amount, roi, return_rate,
            margin_rate, avg_price, transaction_amount, refund_amount, actual_transaction_amount,
            product_cost, gross_profit, tech_fee, net_profit, net_profit_rate, promotion_ratio,
            break_even_roi, transaction_count, cost_per_transaction, profit_per_transaction,
            best_roi, net_break_even_roi, net_break_even_125, net_break_even_value,
            net_break_even_125_from_net, best_roi_from_net, current_roi_multiple, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                self.data_type, self.target_id, self.target_name, record_date,
                self.last_result["promotion_amount"], self.last_result["roi"], self.last_result["return_rate"],
                self.last_result["margin_rate"], self.last_result["avg_price"], self.last_result["transaction_amount"],
                self.last_result["refund_amount"], self.last_result["actual_transaction_amount"],
                self.last_result["product_cost"], self.last_result["gross_profit"], self.last_result["tech_fee"],
                self.last_result["net_profit"], self.last_result["net_profit_rate"], self.last_result["promotion_ratio"],
                self.last_result["break_even_roi"], self.last_result["transaction_count"],
                self.last_result["cost_per_transaction"], self.last_result["profit_per_transaction"],
                self.last_result["best_roi"], self.last_result["net_break_even_roi"], self.last_result["net_break_even_125"],
                self.last_result["net_break_even_value"], self.last_result["net_break_even_125_from_net"],
                self.last_result["best_roi_from_net"], self.last_result["current_roi_multiple"],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
        )

        QMessageBox.information(self, "保存成功", f"利润记录已保存到 {record_date}！")

    def get_database(self):
        if hasattr(self, 'db'):
            return self.db
        if self.parent() and hasattr(self.parent(), 'db'):
            return self.parent().db
        if self.parent() and hasattr(self.parent(), 'main_app') and hasattr(self.parent().main_app, 'db'):
            return self.parent().main_app.db
        return None

    def view_history(self):
        dialog = ProfitHistoryDialog(self.data_type, self.target_id, self.target_name, self.get_database(), self)
        dialog.show()


class ProfitHistoryDialog(QDialog):
    """利润历史记录查看窗口"""
    def __init__(self, data_type, target_id, target_name, db, parent=None):
        super().__init__(parent)
        self.data_type = data_type
        self.target_id = target_id
        self.target_name = target_name
        self.db = db
        self.setWindowTitle(f"📊 利润历史记录 - {target_name}")
        self.resize(1000, 600)
        self.init_ui()
        self.load_data()

    def init_ui(self):
        layout = QVBoxLayout(self)

        self._debug_phd_label = QLabel("【板块:利润历史对话框\n文件:profit.py】利润记录表格/导出/日期筛选/分析查看")
        self._debug_phd_label.setStyleSheet("background-color: #FFB6C1; color: #000; font-weight: bold; padding: 1px;")
        self._debug_phd_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        layout.addWidget(self._debug_phd_label)

        header = QLabel(f"💰 利润历史记录 - {self.target_name}")
        header.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50; padding: 10px;")
        layout.addWidget(header)

        self.table = QTableWidget()
        self.table.setColumnCount(17)
        self.table.setHorizontalHeaderLabels([
            "日期", "推广费", "投产比", "退货率", "毛利率", "客单价",
            "成交金额", "退款金额", "实际成交", "产品成本", "毛利润",
            "技术服务费", "净利润", "净利率", "推广占比", "保本投产", "分析建议"
        ])
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table.cellDoubleClicked.connect(self.show_full_analysis)
        layout.addWidget(self.table)

        btn_layout = QHBoxLayout()

        self.btn_export = QPushButton("📤 导出选中")
        self.btn_export.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                font-weight: bold;
                padding: 8px 20px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        self.btn_export.clicked.connect(self.export_selected)

        self.btn_delete = QPushButton("🗑️ 删除选中")
        self.btn_delete.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                font-weight: bold;
                padding: 8px 20px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        self.btn_delete.clicked.connect(self.delete_selected)

        self.btn_clear = QPushButton("🧹 清空全部")
        self.btn_clear.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                font-weight: bold;
                padding: 8px 20px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        self.btn_clear.clicked.connect(self.clear_all)

        btn_layout.addWidget(self.btn_export)
        btn_layout.addWidget(self.btn_delete)
        btn_layout.addWidget(self.btn_clear)
        btn_layout.addStretch()

        self.btn_close = QPushButton("关闭")
        self.btn_close.clicked.connect(self.accept)
        btn_layout.addWidget(self.btn_close)

        layout.addLayout(btn_layout)

    def load_data(self):
        if self.target_id:
            rows = self.db.safe_fetchall(
                """SELECT id, record_date, promotion_amount, roi, return_rate, margin_rate, avg_price,
                transaction_amount, refund_amount, actual_transaction_amount, product_cost, gross_profit,
                tech_fee, net_profit, net_profit_rate, promotion_ratio, break_even_roi, ai_analysis
                FROM profit_records WHERE data_type=? AND target_id=? ORDER BY record_date DESC""",
                (self.data_type, self.target_id)
            )
        else:
            rows = self.db.safe_fetchall(
                """SELECT id, record_date, promotion_amount, roi, return_rate, margin_rate, avg_price,
                transaction_amount, refund_amount, actual_transaction_amount, product_cost, gross_profit,
                tech_fee, net_profit, net_profit_rate, promotion_ratio, break_even_roi, ai_analysis
                FROM profit_records WHERE data_type=? ORDER BY record_date DESC""",
                (self.data_type,)
            )

        self.table.setRowCount(len(rows))
        self.record_ids = []
        self.ai_analysis_list = []

        for row_idx, row in enumerate(rows):
            self.record_ids.append(row[0])

            ai_analysis = row[17] if len(row) > 17 else ""
            self.ai_analysis_list.append(ai_analysis)

            self.table.setItem(row_idx, 0, QTableWidgetItem(row[1]))
            self.table.setItem(row_idx, 1, QTableWidgetItem(f"¥{row[2]:,.2f}" if row[2] else "¥0.00"))
            self.table.setItem(row_idx, 2, QTableWidgetItem(f"{row[3]:.2f}" if row[3] else "0.00"))
            self.table.setItem(row_idx, 3, QTableWidgetItem(f"{row[4]:.2f}%" if row[4] else "0.00%"))
            self.table.setItem(row_idx, 4, QTableWidgetItem(f"{row[5]:.2f}%" if row[5] else "0.00%"))
            self.table.setItem(row_idx, 5, QTableWidgetItem(f"¥{row[6]:,.2f}" if row[6] else "¥0.00"))
            self.table.setItem(row_idx, 6, QTableWidgetItem(f"¥{row[7]:,.2f}" if row[7] else "¥0.00"))
            self.table.setItem(row_idx, 7, QTableWidgetItem(f"¥{row[8]:,.2f}" if row[8] else "¥0.00"))
            self.table.setItem(row_idx, 8, QTableWidgetItem(f"¥{row[9]:,.2f}" if row[9] else "¥0.00"))
            self.table.setItem(row_idx, 9, QTableWidgetItem(f"¥{row[10]:,.2f}" if row[10] else "¥0.00"))
            self.table.setItem(row_idx, 10, QTableWidgetItem(f"¥{row[11]:,.2f}" if row[11] else "¥0.00"))
            self.table.setItem(row_idx, 11, QTableWidgetItem(f"¥{row[12]:,.2f}" if row[12] else "¥0.00"))

            net_profit_item = QTableWidgetItem(f"¥{row[13]:,.2f}" if row[13] else "¥0.00")
            if row[13] and row[13] < 0:
                net_profit_item.setForeground(QColor("#e74c3c"))
            else:
                net_profit_item.setForeground(QColor("#27ae60"))
            self.table.setItem(row_idx, 12, net_profit_item)

            self.table.setItem(row_idx, 13, QTableWidgetItem(f"{row[14]:.2f}%" if row[14] else "0.00%"))
            self.table.setItem(row_idx, 14, QTableWidgetItem(f"{row[15]:.2f}%" if row[15] else "0.00%"))
            self.table.setItem(row_idx, 15, QTableWidgetItem(f"{row[16]:.2f}" if row[16] else "0.00"))

            if ai_analysis and len(ai_analysis) > 10:
                display_text = ai_analysis[:12] + "..."
            else:
                display_text = ai_analysis if ai_analysis else "-"

            analysis_item = QTableWidgetItem(display_text)
            analysis_item.setForeground(QColor("#9b59b6"))
            self.table.setItem(row_idx, 16, analysis_item)

            for col in range(17):
                item = self.table.item(row_idx, col)
                if item:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)

    def show_full_analysis(self, row, col):
        if col == 16 and row < len(self.ai_analysis_list):
            ai_analysis = self.ai_analysis_list[row]
            if ai_analysis:
                dialog = QDialog(self)
                dialog.setWindowTitle("📝 完整分析建议")
                dialog.resize(600, 450)
                layout = QVBoxLayout(dialog)

                text_browser = QTextBrowser()
                text_browser.setReadOnly(True)
                text_browser.setStyleSheet("""
                    QTextBrowser {
                        background-color: #f8f9fa;
                        border: 1px solid #dee2e6;
                        padding: 12px;
                        font-size: 14px;
                        line-height: 1.6;
                    }
                """)

                html = self.convert_markdown_to_html(ai_analysis)
                text_browser.setHtml(html)
                layout.addWidget(text_browser)

                btn_layout = QHBoxLayout()
                btn_layout.addStretch()
                btn_copy = QPushButton("📋 复制")
                btn_copy.clicked.connect(lambda: self.copy_to_clipboard(ai_analysis, dialog))
                btn_layout.addWidget(btn_copy)
                btn_close = QPushButton("关闭")
                btn_close.clicked.connect(dialog.accept)
                btn_layout.addWidget(btn_close)
                layout.addLayout(btn_layout)

                dialog.exec_()

    def copy_to_clipboard(self, text, dialog):
        clipboard = QApplication.clipboard()
        clipboard.setText(text)
        QMessageBox.information(dialog, "✅ 复制成功", "内容已复制到剪贴板！")

    def convert_markdown_to_html(self, text):
        if not text:
            return ""

        text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

        lines = text.split('\n')
        result = []
        in_list = False
        i = 0
        length = len(lines)

        while i < length:
            line = lines[i]

            if not line.strip():
                if in_list:
                    result.append("</ul>")
                    in_list = False
                i += 1
                continue

            h_match = re.match(r'^(#{1,6})\s+(.+)$', line)
            if h_match:
                if in_list:
                    result.append("</ul>")
                    in_list = False
                level = len(h_match.group(1))
                content = h_match.group(2)
                result.append(f"<h{level}>{content}</h{level}>")
                i += 1
                continue

            list_match = re.match(r'^(\d+)\.\s+(.+)$', line)
            if list_match:
                if not in_list:
                    result.append("<ul>")
                    in_list = True
                content = list_match.group(2)
                content = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', content)
                content = re.sub(r'\*(.+?)\*', r'<i>\1</i>', content)
                content = re.sub(r'`(.+?)`', r'<code>\1</code>', content)
                result.append(f"<li>{content}</li>")
                i += 1
                continue

            bullet_match = re.match(r'^[-*]\s+(.+)$', line)
            if bullet_match:
                if not in_list:
                    result.append("<ul>")
                    in_list = True
                content = bullet_match.group(1)
                content = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', content)
                content = re.sub(r'\*(.+?)\*', r'<i>\1</i>', content)
                content = re.sub(r'`(.+?)`', r'<code>\1</code>', content)
                result.append(f"<li>{content}</li>")
                i += 1
                continue

            if in_list:
                result.append("</ul>")
                in_list = False

            line = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', line)
            line = re.sub(r'\*(.+?)\*', r'<i>\1</i>', line)
            line = re.sub(r'`(.+?)`', r'<code>\1</code>', line)
            result.append(f"<p>{line}</p>")
            i += 1

        if in_list:
            result.append("</ul>")

        html = "".join(result)

        html = re.sub(r'(<br>\s*)+', '<br>', html)

        return html

    def get_selected_ids(self):
        selected_ids = []
        for index in self.table.selectionModel().selectedRows():
            row = index.row()
            if row < len(self.record_ids):
                selected_ids.append(self.record_ids[row])
        return selected_ids

    def delete_selected(self):
        selected_ids = self.get_selected_ids()
        if not selected_ids:
            QMessageBox.warning(self, "提示", "请先选择要删除的记录！")
            return

        reply = QMessageBox.question(
            self, "确认删除",
            f"确定要删除选中的 {len(selected_ids)} 条记录吗？",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            for record_id in selected_ids:
                self.db.safe_execute("DELETE FROM profit_records WHERE id=?", (record_id,))
            QMessageBox.information(self, "删除成功", f"已删除 {len(selected_ids)} 条记录！")
            self.load_data()

    def clear_all(self):
        if not self.record_ids:
            QMessageBox.warning(self, "提示", "没有可清空的数据！")
            return

        reply = QMessageBox.question(
            self, "确认清空",
            f"确定要清空所有 {len(self.record_ids)} 条记录吗？此操作不可恢复！",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            if self.target_id:
                self.db.safe_execute(
                    "DELETE FROM profit_records WHERE data_type=? AND target_id=?",
                    (self.data_type, self.target_id)
                )
            else:
                self.db.safe_execute(
                    "DELETE FROM profit_records WHERE data_type=?",
                    (self.data_type,)
                )
            QMessageBox.information(self, "清空成功", "所有记录已清空！")
            self.load_data()

    def export_selected(self):
        selected_ids = self.get_selected_ids()
        if not selected_ids:
            QMessageBox.warning(self, "提示", "请先选择要导出的记录！")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出数据",
            f"利润记录_{self.target_name}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "利润记录"

            headers = ["日期", "推广费", "投产比", "退货率", "毛利率", "客单价",
                      "成交金额", "退款金额", "实际成交", "产品成本", "毛利润",
                      "技术服务费", "净利润", "净利率", "推广占比", "保本投产", "分析建议"]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(1, col, header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            placeholders = ",".join(["?"] * len(selected_ids))
            rows = self.db.safe_fetchall(
                f"""SELECT record_date, promotion_amount, roi, return_rate, margin_rate, avg_price,
                transaction_amount, refund_amount, actual_transaction_amount, product_cost, gross_profit,
                tech_fee, net_profit, net_profit_rate, promotion_ratio, break_even_roi, ai_analysis
                FROM profit_records WHERE id IN ({placeholders})""",
                selected_ids
            )

            for row_idx, row in enumerate(rows, 2):
                for col_idx, value in enumerate(row, 1):
                    if col_idx == 17:
                        ws.cell(row_idx, col_idx, value if value else "")
                    elif col_idx in [2, 7, 8, 9, 10, 11, 12, 13]:
                        ws.cell(row_idx, col_idx, f"¥{value:,.2f}" if value else "¥0.00")
                    elif col_idx in [4, 5, 14, 15]:
                        ws.cell(row_idx, col_idx, f"{value:.2f}%" if value else "0.00%")
                    elif col_idx == 16:
                        ws.cell(row_idx, col_idx, f"{value:.2f}" if value else "0.00")
                    else:
                        ws.cell(row_idx, col_idx, value)

            for col in range(1, 18):
                if col == 17:
                    ws.column_dimensions[get_column_letter(col)].width = 50
                else:
                    ws.column_dimensions[get_column_letter(col)].width = 14

            wb.save(file_path)
            QMessageBox.information(self, "导出成功", f"数据已导出到：\n{file_path}")

        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出时发生错误：\n{str(e)}")
